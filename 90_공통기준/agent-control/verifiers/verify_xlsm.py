#!/usr/bin/env python3
"""
verify_xlsm.py — xlsm 파일 2단계 검증 스크립트
1차: openpyxl 구조 검증 (시트/테이블/헤더/수식 패턴)
2차: COM 값 검증 (win32com — 실제 Excel로 열어 셀 값 확인)

GPT 합의 2026-04-01: 1단계 구조적 가드레일
사용 예:
  python verify_xlsm.py --task-dir 90_공통기준/작업중/20260401_plan_output \
      --artifact 04_생산계획/SP3M3_생산지시서_메크로자동화_v2.xlsm \
      --sheet PLAN_OUTPUT --table tblPlanOutput --table-ref A1:AK11
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
from pathlib import Path
from typing import Any

# ============================================================
# 공통 유틸
# ============================================================

DEFAULT_HEADERS = [
    "NO", "KIND", "SUB_PART", "TIME_TEXT", "ORDER_QTY", "CARRY_FLAG", "PLAN_KEY",
    "INFO_ROW", "CAR_TYPE", "PART_NO", "MODULE_PART", "SENSE_SPEC", "SENSE_LOC",
    "HOUSING", "HOLDER", "TORSION", "TORSION_LOC", "FRAME_LOC", "NOTE", "V_GEAR",
    "MODEL_1", "MODEL_2", "SUPPLY_LINE", "FRAME_PART", "OLD_MES_PART", "BOX_QTY",
    "P_TRAY_TYPE", "ANGLE_TYPE", "TEXT_ALL", "MODEL_GROUP", "PI_TYPE", "PLL_FLAG",
    "TORSION_NUM", "CAR_FAMILY", "DIR_KEY", "PTRAY_FLAG", "ANGLE_KEY",
]

DEFAULT_KEEP_SHEETS = [
    "SP3 LINE 기준 정보",
    "생산계획",
    "PLAN_RESULT_FIXED",
    "PLAN_BACKUP_FIXED",
    "PLAN_MOVE_SNAPSHOT",
]


def mkcheck(name: str, ok: bool, details: str = "", extra: dict[str, Any] | None = None) -> dict[str, Any]:
    item = {"name": name, "status": "pass" if ok else "fail", "details": details}
    if extra:
        item.update(extra)
    return item


def contains_all(text, patterns: list[str]) -> bool:
    if not text:
        return False
    upper = str(text).upper()
    return all(p.upper() in upper for p in patterns)


def compact_text(v) -> str:
    if v is None:
        return ""
    return re.sub(r'\s+', '', str(v)).upper()


def write_verify(path: Path, result: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


# ============================================================
# 1차: openpyxl 구조 검증
# ============================================================

def run_structure_checks(args: argparse.Namespace) -> tuple[list[dict[str, Any]], bool]:
    checks: list[dict[str, Any]] = []
    artifact = Path(args.artifact)

    if not artifact.exists():
        checks.append(mkcheck("artifact_exists", False, f"파일 없음: {artifact}"))
        return checks, False

    try:
        from openpyxl import load_workbook
        wb = load_workbook(artifact, keep_vba=True, data_only=False)
    except Exception as e:
        checks.append(mkcheck("workbook_open", False, f"openpyxl load 실패: {e}"))
        return checks, False

    # VBA archive
    vba_ok = getattr(wb, "vba_archive", None) is not None
    checks.append(mkcheck("vba_preserved", vba_ok, "vba_archive 존재 여부"))

    # sheet existence
    sheet_names = wb.sheetnames
    checks.append(mkcheck("all_sheets_present", len(sheet_names) >= 2, "시트 수 확인 전부", {"sheetnames": sheet_names}))

    for name in DEFAULT_KEEP_SHEETS + (args.must_keep_sheet or []):
        checks.append(mkcheck(f"keep_sheet::{name}", name in sheet_names, f"{name} 존재 여부"))

    # info sheet / table
    info_sheet_ok = args.info_sheet in sheet_names
    if info_sheet_ok:
        ws_info = wb[args.info_sheet]
        info_tables = {}
        for tbl in ws_info.tables.values():
            info_tables[tbl.name] = tbl.ref
        checks.append(mkcheck("info_table_exists", args.info_table in info_tables,
                              f"{args.info_table} 존재 여부", {"tables": info_tables}))
        if args.info_table in info_tables:
            checks.append(mkcheck("info_table_ref", info_tables[args.info_table] == args.info_table_ref,
                                  f"expected={args.info_table_ref}, actual={info_tables[args.info_table]}"))
    else:
        checks.append(mkcheck("info_sheet_exists", False, f"{args.info_sheet} 시트 없음"))

    # target sheet / table / headers / formulas
    if args.sheet in sheet_names:
        ws = wb[args.sheet]

        checks.append(mkcheck(
            "target_sheet_shape",
            ws.max_row >= args.min_rows and ws.max_column >= args.min_cols,
            f"rows={ws.max_row}, cols={ws.max_column}",
            {"rows": ws.max_row, "cols": ws.max_column}
        ))

        target_tables = {}
        for tbl in ws.tables.values():
            target_tables[tbl.name] = tbl.ref
        checks.append(mkcheck("target_table_exists", args.table in target_tables,
                              f"{args.table} 존재 여부", {"tables": target_tables}))
        if args.table in target_tables:
            checks.append(mkcheck("target_table_ref", target_tables[args.table] == args.table_ref,
                                  f"expected={args.table_ref}, actual={target_tables[args.table]}"))

        actual_headers = [ws.cell(row=1, column=i).value for i in range(1, len(DEFAULT_HEADERS) + 1)]
        checks.append(mkcheck("headers_match", actual_headers == DEFAULT_HEADERS,
                              "헤더 37개 일치 여부", {"actual_headers": actual_headers}))

        # 수식 패턴 검사
        formula_h2 = ws["H2"].value
        formula_i2 = ws["I2"].value
        formula_j2 = ws["J2"].value
        formula_ad2 = ws["AD2"].value if ws.max_column >= 30 else None
        formula_ai2 = ws["AI2"].value if ws.max_column >= 35 else None

        checks.append(mkcheck("formula_info_row_match", contains_all(formula_h2, ["MATCH("]),
                              str(formula_h2)))
        checks.append(mkcheck("formula_lookup_index", contains_all(formula_i2, ["INDEX("]) and contains_all(formula_j2, ["INDEX("]),
                              f"I2={formula_i2} / J2={formula_j2}"))
        checks.append(mkcheck("formula_model_group", contains_all(formula_ad2, ["LET(", "TEXTBEFORE("]),
                              str(formula_ad2)))
        checks.append(mkcheck("formula_dir_key", contains_all(formula_ai2, ["SEARCH("]),
                              str(formula_ai2)))
    else:
        checks.append(mkcheck("target_sheet_exists", False, f"{args.sheet} 시트 없음"))

    wb.close()
    ok = all(c["status"] == "pass" for c in checks)
    return checks, ok


# ============================================================
# 2차: COM 값 검증
# ============================================================

def run_com_checks(args: argparse.Namespace) -> tuple[list[dict[str, Any]], bool]:
    checks: list[dict[str, Any]] = []

    try:
        import pythoncom
        import win32com.client as win32
    except Exception as e:
        checks.append(mkcheck("com_import", False, f"pywin32 import 실패: {e}"))
        return checks, False

    # 원본 보호: temp 복사본으로 검증
    artifact = Path(args.artifact).resolve()
    temp_dir = tempfile.mkdtemp(prefix="verify_xlsm_")
    temp_artifact = Path(temp_dir) / artifact.name
    shutil.copy2(artifact, temp_artifact)

    app = None
    wb = None

    try:
        pythoncom.CoInitialize()
        app = win32.DispatchEx("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        try:
            app.EnableEvents = False
        except Exception:
            pass
        try:
            app.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        except Exception:
            pass

        wb = app.Workbooks.Open(str(temp_artifact), UpdateLinks=0, ReadOnly=False)
        ws = wb.Worksheets(args.sheet)

        # 샘플 입력 → 계산 → 값 확인
        ws.Range(args.sample_input_cell).Value = args.sample_sub_part
        app.CalculateFullRebuild()

        actual_car_type = ws.Range(args.expect_car_type_cell).Value
        actual_part_no = ws.Range(args.expect_part_no_cell).Value
        actual_module = ws.Range(args.expect_module_part_cell).Value
        actual_model_group = ws.Range(args.expect_model_group_cell).Value

        checks.append(mkcheck(
            "com_value_car_type",
            compact_text(actual_car_type) == compact_text(args.expect_car_type),
            f"expected={args.expect_car_type}, actual={actual_car_type}",
            {"actual": actual_car_type}
        ))
        checks.append(mkcheck(
            "com_value_part_no",
            compact_text(actual_part_no) == compact_text(args.expect_part_no),
            f"expected={args.expect_part_no}, actual={actual_part_no}",
            {"actual": actual_part_no}
        ))
        checks.append(mkcheck(
            "com_value_module_part",
            compact_text(actual_module) == compact_text(args.expect_module_part),
            f"expected={args.expect_module_part}, actual={actual_module}",
            {"actual": actual_module}
        ))
        checks.append(mkcheck(
            "com_value_model_group",
            compact_text(actual_model_group) == compact_text(args.expect_model_group),
            f"expected={args.expect_model_group}, actual={actual_model_group}",
            {"actual": actual_model_group}
        ))

        ok = all(c["status"] == "pass" for c in checks)
        return checks, ok

    except Exception as e:
        checks.append(mkcheck("com_runtime", False, f"COM 검증 실패: {e}",
                              {"traceback": traceback.format_exc()}))
        return checks, False

    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if app is not None:
                app.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        # temp 정리
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass


# ============================================================
# main
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="xlsm 2단계 검증: openpyxl 구조 + COM 값")
    parser.add_argument("--task-dir", required=True, help="작업 디렉토리 (verify.json 저장 위치)")
    parser.add_argument("--artifact", required=True, help="검증 대상 xlsm 파일 경로")

    # 구조 검증 옵션
    parser.add_argument("--sheet", default="PLAN_OUTPUT", help="대상 시트명")
    parser.add_argument("--table", default="tblPlanOutput", help="대상 테이블명")
    parser.add_argument("--table-ref", default="A1:AK11", help="테이블 참조 범위")
    parser.add_argument("--min-rows", type=int, default=2, help="최소 행 수")
    parser.add_argument("--min-cols", type=int, default=37, help="최소 열 수")
    parser.add_argument("--info-sheet", default="SP3 LINE 기준 정보")
    parser.add_argument("--info-table", default="tblInfo")
    parser.add_argument("--info-table-ref", default="A1:X1294")
    parser.add_argument("--must-keep-sheet", action="append", default=[])

    # COM 값 검증 옵션
    parser.add_argument("--sample-input-cell", default="C2", help="샘플 입력 셀")
    parser.add_argument("--sample-sub-part", default="147", help="샘플 SUB_PART 값")
    parser.add_argument("--expect-car-type-cell", default="I2", help="CAR_TYPE 확인 셀")
    parser.add_argument("--expect-car-type", default="SG2(180도)(니로)", help="기대 CAR_TYPE")
    parser.add_argument("--expect-part-no-cell", default="J2", help="PART_NO 확인 셀")
    parser.add_argument("--expect-part-no", default="", help="기대 PART_NO")
    parser.add_argument("--expect-module-part-cell", default="K2", help="MODULE_PART 확인 셀")
    parser.add_argument("--expect-module-part", default="", help="기대 MODULE_PART")
    parser.add_argument("--expect-model-group-cell", default="AD2", help="MODEL_GROUP 확인 셀")
    parser.add_argument("--expect-model-group", default="", help="기대 MODEL_GROUP")

    # 구조 검증 전용: skip COM
    parser.add_argument("--structure-only", action="store_true", help="1차 구조 검증만 실행")

    args = parser.parse_args()

    task_dir = Path(args.task_dir)
    verify_path = task_dir / "verify.json"
    task_dir.mkdir(parents=True, exist_ok=True)

    # 1차: 구조 검증
    structure_checks, structure_ok = run_structure_checks(args)

    # 2차: COM 값 검증
    if args.structure_only:
        com_checks, com_ok = [], True
    else:
        com_checks, com_ok = run_com_checks(args)

    result = {
        "artifact_type": "xlsm",
        "artifact_path": str(Path(args.artifact)),
        "status": "pass" if (structure_ok and com_ok) else "fail",
        "structure": {
            "status": "pass" if structure_ok else "fail",
            "checks": structure_checks,
        },
        "com": {
            "status": "pass" if com_ok else "fail",
            "checks": com_checks,
            "sample_input": {
                "cell": args.sample_input_cell,
                "value": args.sample_sub_part,
            },
            "expected": {
                "car_type": args.expect_car_type,
                "part_no": args.expect_part_no,
                "module_part": args.expect_module_part,
                "model_group": args.expect_model_group,
            },
        },
    }

    write_verify(verify_path, result)
    print(json.dumps(result, ensure_ascii=False, indent=2))

    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
