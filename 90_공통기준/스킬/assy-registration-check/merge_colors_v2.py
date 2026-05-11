"""
조립비 현황(8 모품번) + 라인배치(19 모품번) 컬러 cache 통합 → 양식·마스터 재작성.
- 27 모품번 컬러 발견 (조립비 8 + 라인배치 19, 중복 없음)
- 5 모품번 ERP 어디에도 0건 → 모품번 단위 그대로

라인배치 결과의 PROD_NO에 X/Z 접두사 또는 비표준 길이 있을 수 있음. 그대로 포함.
"""
import json
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
# split_unreg_colors 적용 직전 상태로 복원
MASTER_BAK = sorted((MASTER.parent).glob("기준정보_라인별정리_최종_V2_20260506_pre_unregsplit_*.xlsx"))
MASTER_BAK = MASTER_BAK[0] if MASTER_BAK else None
CHECK = REPO / "05_생산실적/조립비정산/05월/등록누락점검_SP3M3_20260509/등록누락점검_SP3M3_20260509.xlsx"
CACHE_ASSY = REPO / "05_생산실적/조립비정산/05월/_cache/erp_unreg_recheck.json"
CACHE_LINEBATCH = REPO / "05_생산실적/조립비정산/05월/_cache/linebatch_24pns.json"
CACHE_XLS1 = REPO / "05_생산실적/조립비정산/05월/_cache/xls1_colors.json"
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"


def load_unregistered_meta():
    """★미등록 모품번 + 면제 + 차종."""
    wb = openpyxl.load_workbook(CHECK, data_only=True, read_only=True)
    out = {}
    for sname, is_exempt in [("★미등록(일반)", False), ("★미등록(면제)", True)]:
        if sname not in wb.sheetnames: continue
        first = True
        for row in wb[sname].iter_rows(values_only=True):
            if first: first = False; continue
            if row and row[0]:
                pn = str(row[0]).strip()
                if not pn.startswith("W"):
                    out[pn] = {"car": str(row[1] or "").strip(), "exempt": is_exempt}
    return out


def merge_colors():
    """두 cache 통합 → {모품번: [컬러 list]}."""
    out = {}
    # 조립비 현황 cache (8 모품번 컬러)
    c1 = json.loads(CACHE_ASSY.read_text(encoding="utf-8"))
    for pn, e in c1.items():
        colors = [c for c in (e.get("colors") or []) if len(c) == len(pn) + 3]
        if colors:
            out[pn] = colors
    # 라인배치 cache (19 모품번 컬러)
    c2 = json.loads(CACHE_LINEBATCH.read_text(encoding="utf-8"))
    for pn, rows in c2.items():
        if not rows: continue
        prods = sorted({(r.get("PROD_NO") if isinstance(r, dict) else "") for r in rows if isinstance(r, dict) and r.get("PROD_NO")})
        if prods:
            out.setdefault(pn, [])
            for p in prods:
                if p not in out[pn]:
                    out[pn].append(p)
    # 1.XLS cache (생산계획 — MO 컬러 보강)
    if CACHE_XLS1.exists():
        c3 = json.loads(CACHE_XLS1.read_text(encoding="utf-8"))
        for pn, rows in c3.items():
            if not rows: continue
            prods = sorted({(r.get("PROD_NO") if isinstance(r, dict) else "") for r in rows if isinstance(r, dict) and r.get("PROD_NO")})
            # MO 도메인 룰 — MO 품번은 컬러 없음. PROD_NO를 모품번 단위로 정규화
            if pn.startswith("MO"):
                prods = [pn]  # 모품번 그대로
            if prods:
                out.setdefault(pn, [])
                for p in prods:
                    if p not in out[pn]:
                        out[pn].append(p)
    return out


def collect_part_set(ws):
    s = set()
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and len(r) >= 4 and r[3]:
            s.add(str(r[3]).strip())
    return s


def parse_max_seq(used, line):
    pats = [re.compile(r"^SP3M(\d+)$")] if line == "SP3M3" else [
        re.compile(r"^HCAMS02-(\d+)$"), re.compile(r"^HCAMS02_(\d+)$")]
    maxv = 0
    for x in used:
        for p in pats:
            m = p.match(x)
            if m and int(m.group(1)) > maxv: maxv = int(m.group(1))
    return maxv


def gen_new(used, line, n):
    while True:
        cand = f"SP3M{n}" if line == "SP3M3" else f"HCAMS02-{n:05d}"
        if cand not in used: return cand, n
        n += 1


def main():
    meta = load_unregistered_meta()
    print(f"[1/5] 미등록 모품번: {len(meta)}")

    colors_by_pn = merge_colors()
    print(f"[2/5] 컬러 통합: {len(colors_by_pn)} 모품번 컬러 발견")
    total_colors = sum(len(v) for v in colors_by_pn.values())
    print(f"  총 컬러 수: {total_colors}")
    pns_no_color = [pn for pn in meta if pn not in colors_by_pn]
    print(f"  컬러 0건 모품번: {len(pns_no_color)} → {pns_no_color}")

    # 마스터 백업 복원
    if not MASTER_BAK or not MASTER_BAK.exists():
        print(f"[ERR] 백업 없음")
        sys.exit(2)
    shutil.copy(MASTER_BAK, MASTER)
    bak2 = MASTER.with_name(MASTER.stem + f"_pre_mergecolors_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(MASTER, bak2)
    print(f"[3/5] 마스터 백업 복원 + pre 백업: {bak2.name}")

    wb = openpyxl.load_workbook(MASTER)
    sp_ws = wb[PRIMARY]; hc_ws = wb[PAIRED]
    sp_used = collect_part_set(sp_ws); hc_used = collect_part_set(hc_ws)
    sp_max = parse_max_seq(sp_used, PRIMARY); hc_max = parse_max_seq(hc_used, PAIRED)
    sp_next = sp_max + 1; hc_next = hc_max + 1

    def has_row(ws, pn):
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r and r[0] and str(r[0]).strip() == pn: return True
        return False

    # 행 build
    rows_to_add = []
    for root_pn, info in meta.items():
        car = info["car"]; exempt = info["exempt"]
        colors = colors_by_pn.get(root_pn) or [root_pn]  # 컬러 없으면 모품번 단위
        for pn in colors:
            rows_to_add.append({"pn": pn, "car": car, "exempt": exempt})

    err_rows = []
    sp_a = hc_a = sp_s = hc_s = 0
    for r in rows_to_add:
        pn = r["pn"]; car = r["car"]
        # SP3M3
        if has_row(sp_ws, pn):
            sp_s += 1
        else:
            new_part, sp_next = gen_new(sp_used, PRIMARY, sp_next)
            sp_used.add(new_part); sp_next += 1
            sp_ws.append([pn, TARGET_CMPY, PRIMARY, new_part, 1, "정단가", None, car])
            sp_a += 1
        err_rows.append({
            "품번": pn, "업체코드": TARGET_CMPY, "라인코드": PRIMARY,
            "조립품번": None, "Usage": 1, "단가구분": "정단가",
            "단가": None, "차종": car, "오류유형": "품번 신규 등록",
            "비고": "신규등록 (ELR 면제)" if r["exempt"] else "신규등록",
        })
        if not r["exempt"]:
            if has_row(hc_ws, pn):
                hc_s += 1
            else:
                new_part, hc_next = gen_new(hc_used, PAIRED, hc_next)
                hc_used.add(new_part); hc_next += 1
                hc_ws.append([pn, TARGET_CMPY, PAIRED, new_part, 1, "정단가", None, car])
                hc_a += 1
            err_rows.append({
                "품번": pn, "업체코드": TARGET_CMPY, "라인코드": PAIRED,
                "조립품번": None, "Usage": 1, "단가구분": "정단가",
                "단가": None, "차종": car, "오류유형": "품번 신규 등록", "비고": "신규등록",
            })
    wb.save(MASTER)
    print(f"[4/5] 마스터 append: SP3M3 +{sp_a}(skip {sp_s}) / HCAMS02 +{hc_a}(skip {hc_s})")

    # 양식 update
    ebak = ERR.with_name(ERR.stem + f"_bak_mergecolors_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR, ebak)
    wbe = openpyxl.load_workbook(ERR)
    wse = wbe["오류리스트"]
    to_del = []
    for r in range(4, wse.max_row + 1):
        if wse.cell(row=r, column=9).value == "품번 신규 등록":
            to_del.append(r)
    for r in sorted(to_del, reverse=True):
        wse.delete_rows(r, 1)
    print(f"[5/5] 양식: 기존 \"품번 신규 등록\" {len(to_del)}행 삭제 + 신규 {len(err_rows)}행 추가")

    template_cells = [wse.cell(row=4, column=c) for c in range(1, 11)]
    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]
    start = wse.max_row + 1
    for i, r in enumerate(err_rows, start=start):
        for c, k in enumerate(cols, start=1):
            cell = wse.cell(row=i, column=c, value=r.get(k))
            tc = template_cells[c - 1]
            cell.font = copy(tc.font); cell.alignment = copy(tc.alignment)
            cell.border = copy(tc.border); cell.number_format = tc.number_format
    wbe.save(ERR)
    print(f"\n=== 완료 ===")
    print(f"양식 신규: r{start}~r{wse.max_row}")


if __name__ == "__main__":
    main()
