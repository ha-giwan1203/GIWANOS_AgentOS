"""
W접두사 26행 cost=0 무효 행 정리.
- ERP raw에 같은 컬러+라인 행 2개 (옛 PART_NO cost=0 + 새 PART_NO cost>0)
- cost>0 행 우선 선택 룰로 양식 W접두사 행 재추출
- cost=0뿐이면 그 컬러는 미등록 = 양식 제외
"""
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
CACHE = REPO / "05_생산실적/조립비정산/05월/_cache/erp_w_strip.json"
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"

W_TO_CAR = {
    "89870BS000": "SP3(180도) (셀토스)", "89880BS000": "SP3(270도) (셀토스)",
    "89870BS200": "SP3(180도) (셀토스)", "89880BS200": "SP3(270도) (셀토스)",
    "89870BS500": "SP3 (180도) (셀토스)", "89880BS500": "SP3 (270도) (셀토스)",
    "89870DE200": "SP3(180도) (셀토스)", "89880DE200": "SP3(270도) (셀토스)",
}


def main():
    cache = json.loads(CACHE.read_text(encoding="utf-8"))

    err_rows = []
    for root, e in cache.items():
        car = W_TO_CAR.get(root, "")
        for color, rows in e.get("lines_by_color", {}).items():
            for line in [PRIMARY, PAIRED]:
                matches = [r for r in rows if r.get("ASSY_LINE_CD") == line
                           and r.get("ASSY_CMPY_CD") == TARGET_CMPY]
                # cost > 0 우선 (deactivate된 cost=0 행 제외)
                active = [r for r in matches if (r.get("ASSY_COST") or 0) > 0]
                if not active:
                    continue
                # 최고 cost 행 (보통 가장 최근 등록)
                m = sorted(active, key=lambda r: -(r.get("ASSY_COST") or 0))[0]
                err_rows.append({
                    "품번": color, "업체코드": TARGET_CMPY, "라인코드": line,
                    "조립품번": m.get("PART_NO"), "Usage": 1,
                    "단가구분": m.get("UNQCST_DIV_NM") or "정단가",
                    "단가": m.get("ASSY_COST"), "차종": car,
                    "오류유형": "라인코드 누락", "비고": "신규등록 (W접두사 sync)",
                })

    print(f"cost>0 행 추출: {len(err_rows)}")

    # 양식 백업 + W접두사 행 재작성
    bak = ERR.with_name(ERR.stem + f"_bak_wprefix_fix_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR, bak)
    print(f"양식 백업: {bak.name}")

    wb = openpyxl.load_workbook(ERR)
    ws = wb["오류리스트"]

    # 기존 W접두사 sync 행 모두 삭제 (비고에 "W접두사 sync" 포함)
    to_del = []
    for r in range(4, ws.max_row + 1):
        note = ws.cell(row=r, column=10).value
        if note and "W접두사 sync" in str(note):
            to_del.append(r)
    for r in sorted(to_del, reverse=True):
        ws.delete_rows(r, 1)
    print(f"기존 W접두사 행 삭제: {len(to_del)}")

    # 신규 행 append (정렬: 라인 → 품번)
    err_rows.sort(key=lambda x: (x["라인코드"], x["품번"]))
    template_cells = [ws.cell(row=4, column=c) for c in range(1, 11)]
    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]
    start = ws.max_row + 1
    for i, r in enumerate(err_rows, start=start):
        for c, k in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=c, value=r.get(k))
            tc = template_cells[c - 1]
            cell.font = copy(tc.font); cell.alignment = copy(tc.alignment)
            cell.border = copy(tc.border); cell.number_format = tc.number_format
    wb.save(ERR)
    print(f"신규 추가: {len(err_rows)}행 (r{start}~r{ws.max_row})")


if __name__ == "__main__":
    main()
