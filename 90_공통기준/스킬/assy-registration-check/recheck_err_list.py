"""
오류리스트(`오류리스트_04월_추가.xlsx`)의 43행 PROD_NO를 ERP 재조회해서 등록 여부·일치 확인.

검증 항목:
- ERP에 SP3M3·HCAMS02 라인 등록되어 있는가
- PART_NO 일치
- 단가 일치
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import lookup_lines
from _xlsx_style import format_workbook


ERR_XLSX = Path(r"\\210.216.217.180\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")
# 사실 ERR_XLSX의 정확한 UNC path
ERR_XLSX = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

OUT_DIR = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\조립비정산\05월\등록누락점검_SP3M3_20260509")
CACHE_OUT = OUT_DIR.parent / "_cache" / f"erp_recheck_{datetime.now():%Y%m%d_%H%M%S}.json"

TARGET_CMPY = "0109"


def load_err_rows():
    wb = openpyxl.load_workbook(ERR_XLSX, data_only=True)
    ws = wb["오류리스트"]
    rows = []
    for r in range(4, ws.max_row + 1):
        pn = ws.cell(row=r, column=1).value
        if not pn:
            continue
        rows.append({
            "row": r,
            "pn": str(pn).strip(),
            "line": ws.cell(row=r, column=3).value,
            "part": ws.cell(row=r, column=4).value,
            "cost": ws.cell(row=r, column=7).value,
            "car": ws.cell(row=r, column=8).value,
        })
    return rows


def main():
    rows = load_err_rows()
    print(f"[1/4] 오류리스트 load: {len(rows)}행")

    roots = sorted({r["pn"][:10] for r in rows})
    print(f"[2/4] unique 모품번: {len(roots)}건")
    print(f"      {roots}")

    # ERP 재조회
    print(f"[3/4] ERP 재조회 (appl_da=2026-05-11)")
    result = lookup_lines.lookup_root_pns_with_colors(roots, appl_da="2026-05-11", cmpy_cd=TARGET_CMPY)
    CACHE_OUT.parent.mkdir(parents=True, exist_ok=True)
    CACHE_OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"      raw cache → {CACHE_OUT.name}")

    # 검증
    n_ok = 0
    n_unreg = 0
    n_mismatch = 0
    for r in rows:
        pn = r["pn"]
        line = r["line"]
        root = pn[:10]
        entry = result.get(root, {})
        erp_rows = entry.get("lines_by_color", {}).get(pn, [])
        matched = [
            x for x in erp_rows
            if (x.get("ASSY_LINE_CD") == line and x.get("ASSY_CMPY_CD") == TARGET_CMPY)
        ]
        if not matched:
            r["erp_등록"] = "X"
            r["erp_part"] = None
            r["erp_cost"] = None
            r["판정"] = "ERP 미등록"
            n_unreg += 1
        else:
            m = matched[0]
            r["erp_등록"] = "O"
            r["erp_part"] = m.get("PART_NO")
            r["erp_cost"] = m.get("ASSY_COST")
            part_ok = str(r["part"] or "").strip() == str(r["erp_part"] or "").strip()
            cost_ok = r["cost"] == r["erp_cost"]
            if part_ok and cost_ok:
                r["판정"] = "일치"
                n_ok += 1
            else:
                r["판정"] = f"불일치(part={part_ok} cost={cost_ok})"
                n_mismatch += 1

    print(f"[4/4] 검증 결과")
    print(f"  일치       : {n_ok}")
    print(f"  ERP 미등록 : {n_unreg}")
    print(f"  불일치     : {n_mismatch}")

    # xlsx 저장
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERP재조회결과"
    ws.append(["오류행", "PROD_NO", "라인", "차종",
               "양식_PART_NO", "양식_단가",
               "ERP_등록", "ERP_PART_NO", "ERP_단가",
               "판정"])
    for r in rows:
        ws.append([r["row"], r["pn"], r["line"], r["car"],
                   r["part"], r["cost"],
                   r["erp_등록"], r["erp_part"], r["erp_cost"],
                   r["판정"]])

    out_xlsx = OUT_DIR / f"ERP재조회결과_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    format_workbook(wb, header_row=1)
    wb.save(out_xlsx)
    print(f"\n산출: {out_xlsx}")


if __name__ == "__main__":
    main()
