"""
xlsx 산출물 공용 서식 helper.
- 폰트: 맑은 고딕 11pt
- 정렬: 가운데
- 테두리: thin 검정
- 헤더: 굵은 흰색 + 청남색 배경

사용:
    from _xlsx_style import format_sheet, auto_column_widths
    ...
    format_sheet(ws, header_row=1, data_start=2)
    auto_column_widths(ws)
    wb.save(...)
"""
from copy import copy

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_FONT = Font(name="맑은 고딕", size=11)
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
DEFAULT_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
THIN = Side(style="thin", color="000000")
DEFAULT_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def format_sheet(ws, header_row=1, data_start=None, max_col=None, left_cols=None):
    """시트 전체 서식 일관 적용.

    Args:
        ws: openpyxl worksheet
        header_row: 헤더 행 번호 (1-base)
        data_start: 데이터 시작 행 번호 (None이면 header_row+1)
        max_col: 마지막 컬럼 (None이면 ws.max_column)
        left_cols: 좌측 정렬 컬럼 인덱스 set (예: {10}은 비고 컬럼만 왼쪽 정렬)
    """
    if max_col is None:
        max_col = ws.max_column
    if data_start is None:
        data_start = header_row + 1
    if left_cols is None:
        left_cols = set()

    # 헤더 서식
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = copy(HEADER_FONT)
        cell.alignment = copy(DEFAULT_ALIGN)
        cell.fill = copy(HEADER_FILL)
        cell.border = copy(DEFAULT_BORDER)

    # 데이터 서식 (값 있는 행만)
    for r in range(data_start, ws.max_row + 1):
        # 행 전체 빈 행은 skip
        if all(ws.cell(row=r, column=cc).value is None for cc in range(1, max_col + 1)):
            continue
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = copy(DEFAULT_FONT)
            cell.alignment = copy(LEFT_ALIGN if c in left_cols else DEFAULT_ALIGN)
            cell.border = copy(DEFAULT_BORDER)


def auto_column_widths(ws, max_col=None, sample_rows=200, min_width=8, max_width=40):
    """컬럼 너비 자동 조정 — 값 길이 기반."""
    if max_col is None:
        max_col = ws.max_column
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for r in range(1, min(ws.max_row + 1, sample_rows + 1)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                # 한글은 2배 폭 보정
                s = str(v)
                width = sum(2 if ord(ch) > 127 else 1 for ch in s) + 2
                if width > max_len:
                    max_len = min(max_width, width)
        ws.column_dimensions[col_letter].width = max_len


def format_workbook(wb, header_row=1, data_start=None, left_cols=None):
    """전 시트 일괄 서식."""
    for sname in wb.sheetnames:
        ws = wb[sname]
        format_sheet(ws, header_row=header_row, data_start=data_start, left_cols=left_cols)
        auto_column_widths(ws)
