"""
라인 기준정보 ERP 조립비 등록 누락 점검 — 진입점.

사용:
    python run_check.py SP3M3
    python run_check.py SP3M3 --applied 2026-05-09
    python run_check.py SP3M3 --master "<UNC 또는 로컬 경로>"
    python run_check.py SP3M3 --use-cache    # _cache JSON 재사용 (ERP 조회 skip)
"""
import argparse
import json
import os
import re
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import classify as classify_mod
import report as report_mod
import lookup_lines


REPO_ROOT = Path(r"C:\Users\User\Desktop\업무리스트")


def load_lines_config():
    cfg_path = SCRIPT_DIR / "lines.json"
    return json.loads(cfg_path.read_text(encoding="utf-8"))


def resolve_master_path(template: str, applied: datetime) -> str:
    """템플릿 placeholder 치환."""
    yyyy = applied.strftime("%Y")
    yy = applied.strftime("%y")
    mm = applied.strftime("%m")
    dd = applied.strftime("%d")
    return (template
            .replace("{YYYY}", yyyy)
            .replace("{YY}", yy)
            .replace("{MM}", mm)
            .replace("{DD}", dd))


def safe_copy_master(src: str) -> Path:
    """본체 락 회피 — 임시 복사."""
    if not os.path.exists(src):
        raise FileNotFoundError(f"마스터 파일 없음: {src}")
    td = tempfile.mkdtemp(prefix="assy_reg_")
    dst = Path(td) / Path(src).name
    shutil.copy(src, dst)
    return dst


def find_header_columns(ws, header_row, pn_header, exempt_header):
    """헤더 행에서 품번·예외 컬럼 인덱스 자동 감지."""
    pn_col = None
    exempt_col = None
    car_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if pn_col is None and (pn_header in s or s == pn_header):
            pn_col = c
        if exempt_header and exempt_col is None and (exempt_header in s or s == exempt_header):
            exempt_col = c
        if car_col is None and ("차종" == s or "차종" in s):
            car_col = c
    return pn_col, exempt_col, car_col


def normalize_pn(v):
    """엑셀 셀이 숫자형이어도 문자열로 정규화. 모품번은 보통 10자리 숫자/문자 mix."""
    if v is None:
        return None
    if isinstance(v, float):
        # 7460137000.0 같은 케이스
        if v.is_integer():
            v = int(v)
    s = str(v).strip()
    s = s.replace(" ", "")
    if not s:
        return None
    return s


def extract_master_rows(master_path: Path, sheet_name: str, header_row: int,
                       pn_header: str, exempt_header: str, exempt_value: str):
    wb = openpyxl.load_workbook(master_path, data_only=True, keep_vba=False, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"시트 '{sheet_name}' 없음. 시트 목록: {wb.sheetnames}")
    ws = wb[sheet_name]

    # 헤더 행 캐싱 (cell column, value) 리스트
    header_cells = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            header_cells.append((cell.column, str(cell.value).strip()))
        break

    # 1차 — 정확 매칭
    pn_col = exempt_col = car_col = None
    for c, s in header_cells:
        if pn_col is None and s == pn_header:
            pn_col = c
        if exempt_header and exempt_col is None and s == exempt_header:
            exempt_col = c
        if car_col is None and s == "차종":
            car_col = c
    # 2차 — 부분 매칭 (정확 매칭 실패한 것만)
    for c, s in header_cells:
        if pn_col is None and pn_header in s:
            pn_col = c
        if exempt_header and exempt_col is None and exempt_header in s:
            exempt_col = c
        if car_col is None and "차종" in s:
            car_col = c

    if pn_col is None:
        raise RuntimeError(f"품번 컬럼 헤더('{pn_header}') 미발견 — 헤더 후보: {header_cells}")
    print(f"  pn_col={pn_col} exempt_col={exempt_col} car_col={car_col}")

    rows = []
    pn_to_exempt_conflict = []
    seen = {}  # pn -> {"exempt": bool, "차종": str}

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if pn_col - 1 >= len(row):
            continue
        pn = normalize_pn(row[pn_col - 1])
        if not pn:
            continue
        # 헤더 잔여 / 영문 텍스트 등 — 품번 패턴 필터 (영숫자 5자 이상)
        if not re.match(r"^[A-Za-z0-9\-_/]{5,}$", pn):
            continue

        exempt_val_raw = row[exempt_col - 1] if exempt_col and exempt_col - 1 < len(row) else None
        exempt_str = "" if exempt_val_raw is None else str(exempt_val_raw).strip()
        is_exempt = bool(exempt_value) and (exempt_value in exempt_str)

        car = row[car_col - 1] if car_col and car_col - 1 < len(row) else None
        car_str = "" if car is None else str(car).replace("\n", " ").strip()

        if pn in seen:
            prev = seen[pn]
            if prev["exempt"] != is_exempt:
                pn_to_exempt_conflict.append(pn)
            continue
        seen[pn] = {"exempt": is_exempt, "차종": car_str}
        rows.append({"pn": pn, "exempt": is_exempt, "차종": car_str})

    if pn_to_exempt_conflict:
        print(f"  [WARN] 동일 품번 면제/비면제 충돌 {len(pn_to_exempt_conflict)}건 — 첫 행 적용: {pn_to_exempt_conflict[:5]}")

    exempt_n = sum(1 for r in rows if r["exempt"])
    print(f"  unique 품번: {len(rows)} (면제 {exempt_n} / 일반 {len(rows) - exempt_n})")
    return rows, exempt_n


def main():
    ap = argparse.ArgumentParser(description="라인 기준정보 ERP 조립비 등록 누락 점검")
    ap.add_argument("LINE", help="라인 키 (예: SP3M3) — lines.json에 정의된 키")
    ap.add_argument("--applied", default=None, help="적용일 YYYY-MM-DD (default: 오늘)")
    ap.add_argument("--master", default=None, help="마스터 파일 경로 직접 지정 (lines.json 템플릿 무시)")
    ap.add_argument("--use-cache", action="store_true", help="기존 _cache JSON 사용 (ERP 조회 skip)")
    ap.add_argument("--limit", type=int, default=0, help="테스트용 — 모품번 N개로 제한 (0=제한없음)")
    args = ap.parse_args()

    cfg = load_lines_config()
    if args.LINE not in cfg:
        print(f"[ERROR] '{args.LINE}' lines.json 미정의. 정의된 키: {list(cfg.keys())}")
        sys.exit(2)
    line_cfg = cfg[args.LINE]

    applied = datetime.fromisoformat(args.applied) if args.applied else datetime.today()
    applied_da = applied.strftime("%Y-%m-%d")
    print(f"=== {args.LINE} 등록 누락 점검 / 적용일 {applied_da} ===")

    primary_cd = line_cfg["primary_line_cd"]
    paired_cd = line_cfg.get("paired_line_cd") or ""

    # 마스터 경로 결정
    if args.master:
        master_src = args.master
    else:
        master_src = resolve_master_path(line_cfg["master_file"], applied)
    print(f"[1/4] 마스터: {master_src}")

    # 임시 복사 (UNC 락 회피)
    master_local = safe_copy_master(master_src)
    print(f"  → 로컬 사본: {master_local}")

    # 추출
    rows, exempt_n = extract_master_rows(
        master_local,
        line_cfg["sheet_name"],
        line_cfg["header_row"],
        line_cfg["pn_column_header"],
        line_cfg.get("exempt_column_header") or "",
        line_cfg.get("exempt_value") or "",
    )
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]
        print(f"  --limit 적용 → {len(rows)}건만 점검")

    # 산출 위치
    next_mm = (applied.month % 12) + 1
    next_year = applied.year + (1 if applied.month == 12 else 0)
    out_dir_base = REPO_ROOT / "05_생산실적" / "조립비정산" / f"{next_mm:02d}월"
    if not out_dir_base.exists():
        # fallback — 현재 월
        out_dir_base = REPO_ROOT / "05_생산실적" / "조립비정산" / f"{applied.month:02d}월"
    out_subdir = out_dir_base / f"등록누락점검_{args.LINE}_{applied.strftime('%Y%m%d')}"
    cache_dir = out_dir_base / "_cache"
    cache_path = cache_dir / f"erp_lookup_{args.LINE}_{applied.strftime('%Y%m%d')}.json"

    # ERP 조회
    if args.use_cache and cache_path.exists():
        print(f"[2/4] 캐시 사용: {cache_path}")
        erp_result = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        print(f"[2/4] ERP 조회 (모품번 {len(rows)}건)")
        pns = [r["pn"] for r in rows]
        erp_result = lookup_lines.lookup_root_pns_with_colors(pns, appl_da=applied_da, cmpy_cd="0109")
        report_mod.write_raw_cache(erp_result, cache_path)

    # 분류
    print(f"[3/4] 분류 (primary={primary_cd} paired={paired_cd or '단독'})")
    classified = classify_mod.classify(rows, erp_result, primary_cd, paired_cd)
    summary_n = classify_mod.summarize(classified)
    main_total = summary_n.pop("__total_main__")
    if main_total != len(rows):
        print(f"  [WARN] 메인 분류 합 {main_total} ≠ unique 모품번 {len(rows)} — 검증 결함 가능")
    print(f"  분류: {summary_n}")

    # 산출
    print(f"[4/4] 산출 → {out_subdir}")
    xlsx_path = out_subdir / f"등록누락점검_{args.LINE}_{applied.strftime('%Y%m%d')}.xlsx"
    md_path = out_subdir / "summary.md"
    report_mod.write_xlsx(classified, xlsx_path, primary_cd, paired_cd, applied_da)
    report_mod.write_summary_md(
        classified, md_path, primary_cd, paired_cd, applied_da,
        master_path=master_src, total_unique=len(rows), exempt_n=exempt_n,
    )

    # 1줄 요약 보고
    fail_keys = ["primary누락_일반", "paired누락_일반", "양쪽누락_일반", "미등록_일반",
                 "primary누락_면제", "미등록_면제", "이상_면제paired등록"]
    fail_n = sum(len(classified[k]) for k in fail_keys)
    print(f"\n=== 결과 ===")
    print(f"입력 {len(rows)} | 정상(일반) {len(classified['정상_일반'])} + 정상(면제) {len(classified['정상_면제'])} | 누락 {fail_n} | 컬러편차 {len(classified['컬러편차'])}")
    print(f"xlsx: {xlsx_path}")
    print(f"summary: {md_path}")


if __name__ == "__main__":
    main()
