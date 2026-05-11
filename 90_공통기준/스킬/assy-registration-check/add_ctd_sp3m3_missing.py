"""
CTD 컬러 SP3M3 라인 누락 행 추가.
- ERP raw에서 CTD 컬러 SP3M3 cost=0 = 옛 PART_NO만 (실질 미등록)
- 같은 모품번 OVS 컬러의 SP3M3 cost>0 단가 차용
- 양식 라인누락 행 추가
"""
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
CACHE = REPO / "05_생산실적/조립비정산/05월/_cache/erp_w_strip.json"
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
TARGET_CMPY = "0109"
W_TO_CAR = {
    "89870BS000": "SP3(180도) (셀토스)", "89880BS000": "SP3(270도) (셀토스)",
    "89870BS200": "SP3(180도) (셀토스)", "89880BS200": "SP3(270도) (셀토스)",
    "89870BS500": "SP3 (180도) (셀토스)", "89880BS500": "SP3 (270도) (셀토스)",
    "89870DE200": "SP3(180도) (셀토스)", "89880DE200": "SP3(270도) (셀토스)",
}


def main():
    cache = json.loads(CACHE.read_text(encoding="utf-8"))

    new_rows = []
    for root, e in cache.items():
        car = W_TO_CAR.get(root, "")
        lines_by_color = e.get("lines_by_color", {})
        # 같은 모품번 OVS 컬러 SP3M3 cost>0 찾기 (단가 차용 ref)
        ref_cost = None
        ref_part = None
        for color, rows in lines_by_color.items():
            sp_active = [r for r in rows if r.get("ASSY_LINE_CD") == PRIMARY
                         and r.get("ASSY_CMPY_CD") == TARGET_CMPY
                         and (r.get("ASSY_COST") or 0) > 0]
            if sp_active:
                ref_cost = sp_active[0].get("ASSY_COST")
                ref_part = sp_active[0].get("PART_NO")
                break

        # CTD 컬러 + SP3M3 cost=0뿐 → 라인누락 행 생성
        for color, rows in lines_by_color.items():
            sp_active = [r for r in rows if r.get("ASSY_LINE_CD") == PRIMARY
                         and r.get("ASSY_CMPY_CD") == TARGET_CMPY
                         and (r.get("ASSY_COST") or 0) > 0]
            if sp_active:
                continue  # 이미 cost>0 SP3M3 등록 = fix_wprefix에 포함됨
            # SP3M3 cost=0뿐 또는 SP3M3 행 없음 = 누락
            if ref_cost is None:
                continue  # 같은 모품번에 SP3M3 cost>0 ref 없음
            new_rows.append({
                "품번": color, "업체코드": TARGET_CMPY, "라인코드": PRIMARY,
                "조립품번": None,  # 신규 시퀀스 필요
                "Usage": 1, "단가구분": "정단가", "단가": ref_cost,
                "차종": car, "오류유형": "라인코드 누락",
                "비고": f"신규등록 (W접두사 CTD-SP3M3누락, 단가는 OVS 차용)",
            })

    print(f"CTD SP3M3 누락 행: {len(new_rows)}")
    for r in new_rows:
        print(f"  {r['품번']} / {r['라인코드']} / cost={r['단가']}")

    if not new_rows:
        return

    # 양식 append
    bak = ERR.with_name(ERR.stem + f"_bak_ctdmiss_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR, bak)
    wb = openpyxl.load_workbook(ERR)
    ws = wb["오류리스트"]
    template_cells = [ws.cell(row=4, column=c) for c in range(1, 11)]
    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]
    start = ws.max_row + 1
    for i, r in enumerate(new_rows, start=start):
        for c, k in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=c, value=r.get(k))
            tc = template_cells[c - 1]
            cell.font = copy(tc.font); cell.alignment = copy(tc.alignment)
            cell.border = copy(tc.border); cell.number_format = tc.number_format
    wb.save(ERR)
    print(f"\n양식 추가: {len(new_rows)}행 (r{start}~r{ws.max_row})")


if __name__ == "__main__":
    main()
