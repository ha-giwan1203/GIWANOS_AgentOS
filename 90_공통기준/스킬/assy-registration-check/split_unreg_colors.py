"""
미등록 32 모품번 컬러별 분리 — ERP 재조회 (cmpy_cd 빈칸으로 다른 업체 컬러 확보).

1. 미등록 32 모품번 ERP 재조회 (cmpy_cd="")
2. 컬러 정보 있는 모품번: 컬러 list 추출
3. 컬러 정보 없는 모품번: 모품번 그대로 (양산 자체 X)
4. 마스터 백업 복원 + 컬러별 행 추가
5. 양식 미등록 61행 제거 + 컬러별 행 추가
"""
import json
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import lookup_lines

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
MASTER_BAK = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506_pre_addunreg_20260511_135912.xlsx"
CHECK = REPO / "05_생산실적/조립비정산/05월/등록누락점검_SP3M3_20260509/등록누락점검_SP3M3_20260509.xlsx"
CACHE = REPO / "05_생산실적/조립비정산/05월/_cache/erp_unreg_recheck.json"
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"


def load_unregistered():
    wb = openpyxl.load_workbook(CHECK, data_only=True, read_only=True)
    out = []
    for sname, is_exempt in [("★미등록(일반)", False), ("★미등록(면제)", True)]:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first: first = False; continue
            if row and row[0] and not str(row[0]).strip().startswith("W"):
                out.append({"pn": str(row[0]).strip(),
                            "car": str(row[1] or "").strip(),
                            "exempt": is_exempt})
    return out


def collect_part_set(ws):
    s = set()
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and len(r) >= 4 and r[3]:
            s.add(str(r[3]).strip())
    return s


def parse_max_seq(used_set, line):
    if line == "SP3M3":
        pat = re.compile(r"^SP3M(\d+)$"); maxv = 0
        for x in used_set:
            m = pat.match(x)
            if m and int(m.group(1)) > maxv: maxv = int(m.group(1))
        return maxv
    else:
        maxv = 0
        for pat in [re.compile(r"^HCAMS02-(\d+)$"), re.compile(r"^HCAMS02_(\d+)$")]:
            for x in used_set:
                m = pat.match(x)
                if m and int(m.group(1)) > maxv: maxv = int(m.group(1))
        return maxv


def gen_new(used, line, n):
    while True:
        cand = f"SP3M{n}" if line == "SP3M3" else f"HCAMS02-{n:05d}"
        if cand not in used: return cand, n
        n += 1


def main():
    items = load_unregistered()
    print(f"[1/6] 미등록 모품번: {len(items)}")

    roots = [r["pn"] for r in items]
    if CACHE.exists():
        print(f"[2/6] cache 재사용: {CACHE.name}")
        result = json.loads(CACHE.read_text(encoding="utf-8"))
    else:
        print(f"[2/6] ERP 재조회 (cmpy_cd 빈칸, 32 모품번)")
        result = lookup_lines.lookup_root_pns_with_colors(roots, appl_da="2026-05-11", cmpy_cd="")
        CACHE.parent.mkdir(parents=True, exist_ok=True)
        CACHE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    # 컬러 정보 집계
    items_with_colors = []
    items_no_color = []
    for r in items:
        entry = result.get(r["pn"], {})
        colors = entry.get("colors") or []
        # 컬러 4자리 미사용 필터
        valid_colors = [c for c in colors if len(c) == len(r["pn"]) + 3]
        if valid_colors:
            for c in valid_colors:
                items_with_colors.append({"pn": c, "root": r["pn"], "car": r["car"], "exempt": r["exempt"]})
        else:
            items_no_color.append(r)
    print(f"[3/6] 컬러 있음: {len(items_with_colors)}개 컬러 / 컬러 없음: {len(items_no_color)} 모품번 그대로")

    # 4. 마스터 백업 복원 + 컬러별 추가
    print(f"[4/6] 마스터 백업 복원 후 컬러별 재추가")
    if not MASTER_BAK.exists():
        print(f"  [ERR] 백업 없음: {MASTER_BAK}")
        sys.exit(2)
    shutil.copy(MASTER_BAK, MASTER)
    bak2 = MASTER.with_name(MASTER.stem + f"_pre_unregsplit_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(MASTER, bak2)

    wb = openpyxl.load_workbook(MASTER)
    sp_ws = wb[PRIMARY]; hc_ws = wb[PAIRED]
    sp_used = collect_part_set(sp_ws); hc_used = collect_part_set(hc_ws)
    sp_max = parse_max_seq(sp_used, PRIMARY); hc_max = parse_max_seq(hc_used, PAIRED)
    sp_next = sp_max + 1; hc_next = hc_max + 1
    print(f"  사용중 max — SP3M3={sp_max} / HCAMS02={hc_max}")

    def has_row(ws, pn):
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r and r[0] and str(r[0]).strip() == pn: return True
        return False

    sp_a = hc_a = sp_s = hc_s = 0
    rows_to_add = items_with_colors + [{"pn": r["pn"], "root": r["pn"], "car": r["car"], "exempt": r["exempt"]} for r in items_no_color]

    err_rows = []
    for r in rows_to_add:
        pn = r["pn"]; car = r["car"]
        # SP3M3
        if has_row(sp_ws, pn):
            sp_s += 1
        else:
            new_part, sp_next = gen_new(sp_used, PRIMARY, sp_next)
            sp_used.add(new_part); sp_next += 1
            sp_ws.append([pn, TARGET_CMPY, PRIMARY, new_part, 1, "정단가", None, car])
            sp_a += 1
        err_rows.append({
            "품번": pn, "업체코드": TARGET_CMPY, "라인코드": PRIMARY,
            "조립품번": None,  # 양식엔 빈칸 — 사용자 ERP 신규 등록 시 부여
            "Usage": 1, "단가구분": "정단가", "단가": None, "차종": car,
            "오류유형": "품번 신규 등록",
            "비고": "신규등록 (ELR 면제)" if r["exempt"] else "신규등록",
        })
        if not r["exempt"]:
            if has_row(hc_ws, pn):
                hc_s += 1
            else:
                new_part, hc_next = gen_new(hc_used, PAIRED, hc_next)
                hc_used.add(new_part); hc_next += 1
                hc_ws.append([pn, TARGET_CMPY, PAIRED, new_part, 1, "정단가", None, car])
                hc_a += 1
            err_rows.append({
                "품번": pn, "업체코드": TARGET_CMPY, "라인코드": PAIRED,
                "조립품번": None, "Usage": 1, "단가구분": "정단가",
                "단가": None, "차종": car, "오류유형": "품번 신규 등록", "비고": "신규등록",
            })
    wb.save(MASTER)
    print(f"  마스터 append: SP3M3 +{sp_a}(skip {sp_s}) / HCAMS02 +{hc_a}(skip {hc_s})")

    # 5. 양식 미등록 61행 제거 + 컬러별 새로 추가
    print(f"[5/6] 양식 update")
    ebak = ERR.with_name(ERR.stem + f"_bak_unregsplit_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR, ebak)
    wbe = openpyxl.load_workbook(ERR)
    wse = wbe["오류리스트"]

    # 기존 "품번 신규 등록" 분류 행 모두 삭제 (행 번호 내림차순)
    to_del = []
    for r in range(4, wse.max_row + 1):
        if wse.cell(row=r, column=9).value == "품번 신규 등록":
            to_del.append(r)
    print(f"  기존 \"품번 신규 등록\" 삭제 대상: {len(to_del)}행")
    for r in sorted(to_del, reverse=True):
        wse.delete_rows(r, 1)

    # 컬러별 + no_color 새 행 추가 (서식 template = r4)
    template_cells = [wse.cell(row=4, column=c) for c in range(1, 11)]
    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]
    start = wse.max_row + 1
    for i, r in enumerate(err_rows, start=start):
        for c, k in enumerate(cols, start=1):
            cell = wse.cell(row=i, column=c, value=r.get(k))
            tc = template_cells[c - 1]
            cell.font = copy(tc.font); cell.alignment = copy(tc.alignment)
            cell.border = copy(tc.border); cell.number_format = tc.number_format
    wbe.save(ERR)
    print(f"  양식 신규 추가: {len(err_rows)}행 (r{start}~r{wse.max_row})")
    print(f"[6/6] 완료. 양식 총 행수: {wse.max_row - 3}")


if __name__ == "__main__":
    main()
