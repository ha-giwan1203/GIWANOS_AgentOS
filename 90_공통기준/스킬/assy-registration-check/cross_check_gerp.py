"""
미등록 품번 vs GERP 월별 raw 실적 교차 검증.

용도:
    assy-registration-check 산출물의 ★미등록 시트에 들어간 품번이
    GERP raw 실적에 등장하는지 확인 → 등장하면 "실적은 있는데 단가 미등록" 시급도 높은 정정 대상.

호출:
    python cross_check_gerp.py <검사_xlsx> <gerp_raw_xlsx> [--out <path>]
"""
import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _xlsx_style import format_workbook

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


# 소스별 컬럼 매핑 (1-base column index)
SOURCE_SCHEMA = {
    "gerp": {
        "header_row": 1,
        "data_start": 2,
        "cols": {
            "구분1": 1, "구분2": 2, "조립라인": 3, "조립업체": 4,
            "제품번호": 7, "주야": 14, "생산량": 15, "단가": 16, "조립금액": 17,
        },
        "cmpy_target_includes": "대원테크",  # GERP는 한글 업체명
    },
    "olderp": {
        "header_row": 3,
        "data_start": 4,
        "cols": {
            "기준일": 2, "조립업체": 3, "문서번호": 4, "제품번호": 5, "품명": 6,
            "조립라인": 7, "차종": 8, "기종": 9, "LOTNO": 10,
            "생산량": 11, "단가": 12, "조립금액": 13, "구분": 14,
        },
        # 구ERP 업체 표기는 "A00031[화인텍]" 형식. 대원테크 코드는 도메인 룰로 사용자 확인 필요.
        # 우선 raw 그대로 표시. 0109 한정 분류는 GERP 케이스만 적용.
        "cmpy_target_includes": "대원테크",
    },
}


def load_unregistered_pns(xlsx_path: Path):
    """등록누락점검_*.xlsx 의 ★미등록(일반)·★미등록(면제) 시트에서 모품번 추출."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("★미등록"):
            ws = wb[sheet_name]
            exempt = "면제" in sheet_name
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if not row or not row[0]:
                    continue
                pn = str(row[0]).strip()
                car = str(row[1]).strip() if row[1] else ""
                out.append({"pn": pn, "exempt": exempt, "차종": car})
    return out


def _normalize_variants(pn: str):
    """매칭 시도 prefix 변형 — W접두사 등 GERP 표기 차이 흡수."""
    out = [pn]
    if pn.startswith("W") and len(pn) > 1:
        out.append(pn[1:])  # W 제거 변형
    if pn.startswith("MO") and len(pn) > 2:
        out.append(pn[2:])  # MO 제거 (모듈품번 → 기본품번)
    return out


def scan_raw(path: Path, pn_set, source: str):
    """소스 raw scan — 제품번호가 pn_set의 어느 prefix(또는 변형)로 시작하면 매칭."""
    schema = SOURCE_SCHEMA[source]
    cols = schema["cols"]
    data_start = schema["data_start"]
    pn_col = cols["제품번호"]

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    prefix_to_root = {}
    for pn in pn_set:
        for v in _normalize_variants(pn):
            prefix_to_root.setdefault(v, pn)
    prefixes = sorted(prefix_to_root.keys(), key=len, reverse=True)

    matches_by_pn = defaultdict(list)
    seen_rows = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue
        seen_rows += 1
        if pn_col - 1 >= len(row):
            continue
        prod = row[pn_col - 1]
        if prod is None:
            continue
        prod_str = str(prod).strip()
        # 정수형 prod (예: 13585757) → str 변환 후 .0 제거
        if prod_str.endswith(".0"):
            prod_str = prod_str[:-2]
        if not prod_str:
            continue

        matched_variant = None
        for p in prefixes:
            if prod_str == p or prod_str.startswith(p):
                matched_variant = p
                break
        if matched_variant is None:
            continue
        matched = prefix_to_root[matched_variant]

        # 모든 매핑 컬럼 추출
        item = {"_matched_variant": matched_variant, "제품번호": prod_str}
        for label, c in cols.items():
            if label == "제품번호":
                continue
            if c - 1 < len(row):
                item[label] = row[c - 1]
            else:
                item[label] = None
        matches_by_pn[matched].append(item)
    return matches_by_pn, seen_rows


# 호환성 — 기존 호출자 유지
def scan_gerp(gerp_path: Path, pn_set):
    return scan_raw(gerp_path, pn_set, "gerp")


def write_report(items, matches_by_pn, gerp_rows, out_xlsx: Path, out_md: Path, gerp_path: Path, source="gerp"):
    cols = SOURCE_SCHEMA[source]["cols"]
    cmpy_target = SOURCE_SCHEMA[source]["cmpy_target_includes"]
    src_label = "GERP" if source == "gerp" else "구ERP"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 시트1: 매칭있음
    ws1 = wb.create_sheet("실적있음(시급)")
    ws1.append(["모품번", "차종", "면제", f"{src_label}행수", "총생산량", "조립라인_분포", "대원테크여부", "다른업체"])
    sw1_rows = []
    for it in items:
        pn = it["pn"]
        ms = matches_by_pn.get(pn, [])
        if not ms:
            continue
        total_qty = sum((m.get("생산량") or 0) for m in ms)
        line_set = defaultdict(int)
        for m in ms:
            line_set[(m.get("조립라인") or "", m.get("조립업체") or "")] += m.get("생산량") or 0
        line_dist = ", ".join(f"{l}/{c}={q}" for (l, c), q in sorted(line_set.items(), key=lambda x: -x[1]))
        cmpy_strs = [str(m.get("조립업체") or "") for m in ms]
        has_target = any(cmpy_target in c for c in cmpy_strs)
        other_cmpy = sorted({c for c in cmpy_strs if c and cmpy_target not in c})
        sw1_rows.append([pn, it["차종"], "ELR" if it["exempt"] else "", len(ms), total_qty, line_dist,
                         "Y" if has_target else "N", ", ".join(other_cmpy)])

    sw1_rows.sort(key=lambda r: -r[4])
    for r in sw1_rows:
        ws1.append(r)

    # 시트2: 매칭없음
    ws2 = wb.create_sheet("실적없음")
    ws2.append(["모품번", "차종", "면제"])
    for it in items:
        if it["pn"] not in matches_by_pn:
            ws2.append([it["pn"], it["차종"], "ELR" if it["exempt"] else ""])

    # 시트3: 매칭 raw — 모든 cols 노출
    ws3 = wb.create_sheet("매칭raw")
    raw_cols = ["모품번"] + list(cols.keys())
    ws3.append(raw_cols)
    for it in items:
        for m in matches_by_pn.get(it["pn"], []):
            ws3.append([it["pn"]] + [m.get(c) for c in cols.keys()])

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    format_workbook(wb, header_row=1)
    wb.save(out_xlsx)

    # md 요약
    parts = []
    parts.append(f"# 미등록 품번 vs {src_label} 실적 교차 검증\n")
    parts.append(f"- 검사 대상: 미등록 {len(items)}건")
    parts.append(f"- {src_label} raw: `{gerp_path}` ({gerp_rows}행)")
    parts.append(f"- 매칭 (실적있음): **{len(sw1_rows)}건**")
    parts.append(f"- 미매칭 (실적없음): {len(items) - len(sw1_rows)}건")
    parts.append("")

    if sw1_rows:
        parts.append(f"## ★ 실적있음 — 정정 시급 (양산했는데 단가 미등록)\n")
        parts.append(f"| 모품번 | 차종 | 면제 | {src_label}행수 | 총생산량 | 대원테크 | 다른업체 |")
        parts.append("|--------|------|------|----------|----------|----------|----------|")
        for r in sw1_rows:
            parts.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3]} | {r[4]:,} | {r[6]} | {r[7]} |")
        parts.append("")
        parts.append("→ 라인/업체 분포는 xlsx `실적있음(시급)` 시트 참조")
        parts.append("")

    parts.append("## 실적없음 (4월 미양산)\n")
    no_match = [it for it in items if it["pn"] not in matches_by_pn]
    if no_match:
        parts.append("| 모품번 | 차종 | 면제 |")
        parts.append("|--------|------|------|")
        for it in no_match:
            parts.append(f"| {it['pn']} | {it['차종']} | {'ELR' if it['exempt'] else ''} |")
    else:
        parts.append("(없음)")
    parts.append("")

    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_md.write_text("\n".join(parts), encoding="utf-8")

    return len(sw1_rows), len(items) - len(sw1_rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("check_xlsx", help="등록누락점검_*.xlsx 경로")
    ap.add_argument("raw_xlsx", help="실적 raw xlsx 경로")
    ap.add_argument("--source", choices=["gerp", "olderp"], default="gerp", help="raw 종류 (default: gerp)")
    ap.add_argument("--out-dir", default=None, help="산출 디렉터리 (default: check_xlsx 같은 폴더)")
    ap.add_argument("--out-suffix", default=None, help="산출 파일명 suffix (default: source 값)")
    args = ap.parse_args()

    check_path = Path(args.check_xlsx)
    raw_path = Path(args.raw_xlsx)
    if not check_path.exists():
        print(f"[ERR] 검사 xlsx 없음: {check_path}")
        sys.exit(2)
    if not raw_path.exists():
        print(f"[ERR] raw xlsx 없음: {raw_path}")
        sys.exit(2)

    suffix = args.out_suffix or args.source.upper()
    out_dir = Path(args.out_dir) if args.out_dir else check_path.parent
    out_xlsx = out_dir / f"실적교차_미등록_{suffix}.xlsx"
    out_md = out_dir / f"실적교차_미등록_{suffix}.md"

    items = load_unregistered_pns(check_path)
    print(f"[1/3] ★미등록 품번 추출: {len(items)}건")
    pn_set = {it["pn"] for it in items}

    src_label = "GERP" if args.source == "gerp" else "구ERP"
    print(f"[2/3] {src_label} raw scan: {raw_path.name}")
    matches, raw_rows = scan_raw(raw_path, pn_set, args.source)
    print(f"  매칭: {len(matches)}품번 / {src_label} {raw_rows}행")

    n_match, n_no = write_report(items, matches, raw_rows, out_xlsx, out_md, raw_path, source=args.source)
    print(f"[3/3] 산출 완료")
    print(f"  실적있음(시급): {n_match}건")
    print(f"  실적없음:       {n_no}건")
    print(f"  xlsx: {out_xlsx}")
    print(f"  md:   {out_md}")


if __name__ == "__main__":
    main()
