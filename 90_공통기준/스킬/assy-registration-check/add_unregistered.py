"""
★미등록(일반)/★미등록(면제) 분류 → 오류리스트에 "품번 신규 등록"으로 추가.
- W접두사 false positive 제외 (마스터 W vs ERP 실표기 차이로 분류된 케이스)
- 면제(ELR) → SP3M3만 / 일반 → SP3M3+HCAMS02
- 단가·조립품번 빈칸 (ERP·마스터 양쪽 0건이라 차용 불가, 사용자 ERP 등록 시 결정)
- 컬러 정보는 마스터 모품번 단위라 양식도 모품번 단위 행
"""
import sys
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
from _xlsx_style import format_sheet, auto_column_widths

CHECK = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\조립비정산\05월\등록누락점검_SP3M3_20260509\등록누락점검_SP3M3_20260509.xlsx")
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"


def load_unregistered():
    """★미등록(일반)/★미등록(면제) 시트에서 모품번 + 면제 + 차종 추출."""
    wb = openpyxl.load_workbook(CHECK, data_only=True, read_only=True)
    out = []
    for sname, is_exempt in [("★미등록(일반)", False), ("★미등록(면제)", True)]:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if row and row[0]:
                pn = str(row[0]).strip()
                car = str(row[1] or "").strip()
                out.append({"pn": pn, "car": car, "exempt": is_exempt})
    return out


def main():
    items = load_unregistered()
    print(f"★미등록 모품번: {len(items)}")

    # W접두사 false positive 제외 (마스터 W vs ERP 실표기 차이)
    filtered = [r for r in items if not r["pn"].startswith("W")]
    excluded = [r for r in items if r["pn"].startswith("W")]
    print(f"  W접두사 제외 {len(excluded)}건: {[r['pn'] for r in excluded]}")
    print(f"  진행 대상: {len(filtered)}")

    # 행 생성
    new_rows = []
    for r in filtered:
        pn = r["pn"]
        if r["exempt"]:
            new_rows.append({
                "품번": pn, "업체코드": TARGET_CMPY, "라인코드": PRIMARY,
                "조립품번": None, "Usage": 1, "단가구분": "정단가",
                "단가": None, "차종": r["car"],
                "오류유형": "품번 신규 등록", "비고": "신규등록 (ELR 면제)",
            })
        else:
            for line in [PRIMARY, PAIRED]:
                new_rows.append({
                    "품번": pn, "업체코드": TARGET_CMPY, "라인코드": line,
                    "조립품번": None, "Usage": 1, "단가구분": "정단가",
                    "단가": None, "차종": r["car"],
                    "오류유형": "품번 신규 등록", "비고": "신규등록",
                })

    print(f"  생성 행: {len(new_rows)}")

    # 백업
    bak = ERR.with_name(ERR.stem + f"_bak_addmissing_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR, bak)
    print(f"백업: {bak.name}")

    # 양식 append
    wb = openpyxl.load_workbook(ERR)
    ws = wb["오류리스트"]

    # template 서식 — 기존 r4 셀 서식 추출 (이미 reformat된 상태)
    template_cells = [ws.cell(row=4, column=c) for c in range(1, 11)]

    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]

    start = ws.max_row + 1
    for i, r in enumerate(new_rows, start=start):
        for c, k in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=c, value=r.get(k))
            tc = template_cells[c - 1]
            cell.font = copy(tc.font)
            cell.alignment = copy(tc.alignment)
            cell.border = copy(tc.border)
            cell.number_format = tc.number_format

    wb.save(ERR)
    print(f"\n추가 완료: {len(new_rows)}행 (r{start}~r{ws.max_row})")
    print(f"양식 총 행수: {ws.max_row - 3}")


if __name__ == "__main__":
    main()
