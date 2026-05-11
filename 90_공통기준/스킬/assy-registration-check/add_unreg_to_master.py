"""
미등록 32 모품번을 V2 마스터 SP3M3·HCAMS02 시트에 모품번 단위로 추가.
- 단가는 빈칸 (ERP·GERP 양쪽 0건이라 차용 불가)
- 조립품번은 신규 시퀀스 자동 부여 (마스터 + ERP cache 충돌 회피)
- 차종 컬럼은 마스터 .xlsm 차종 그대로
- 다음달 정산 빌더가 GERP raw 매칭 시 마스터 참조 — 마스터에 행 있어야 매칭 가능
"""
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
CHECK = REPO / "05_생산실적/조립비정산/05월/등록누락점검_SP3M3_20260509/등록누락점검_SP3M3_20260509.xlsx"
CACHE_SP3M3 = REPO / "05_생산실적/조립비정산/05월/_cache/erp_lookup_SP3M3_20260509.json"
CACHE_WSTRIP = REPO / "05_생산실적/조립비정산/05월/_cache/erp_w_strip.json"
CACHE_RECHECK = REPO / "05_생산실적/조립비정산/05월/_cache"

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"


def load_unregistered():
    wb = openpyxl.load_workbook(CHECK, data_only=True, read_only=True)
    out = []
    for sname, is_exempt in [("★미등록(일반)", False), ("★미등록(면제)", True)]:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False; continue
            if row and row[0]:
                out.append({"pn": str(row[0]).strip(),
                            "car": str(row[1] or "").strip(),
                            "exempt": is_exempt})
    return out


def collect_part_set(ws):
    s = set()
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r and len(r) >= 4 and r[3]:
            s.add(str(r[3]).strip())
    return s


def collect_erp_parts(cache, line):
    s = set()
    for root, entry in cache.items():
        for color, rows in entry.get("lines_by_color", {}).items():
            for r in rows:
                if r.get("ASSY_LINE_CD") == line and r.get("PART_NO"):
                    s.add(str(r["PART_NO"]).strip())
    return s


def parse_max_seq(used_set, line):
    if line == "SP3M3":
        pat = re.compile(r"^SP3M(\d+)$")
        maxv = 0
        for x in used_set:
            m = pat.match(x)
            if m:
                v = int(m.group(1))
                if v > maxv: maxv = v
        return maxv
    else:  # HCAMS02
        maxv = 0
        for pat in [re.compile(r"^HCAMS02-(\d+)$"), re.compile(r"^HCAMS02_(\d+)$")]:
            for x in used_set:
                m = pat.match(x)
                if m:
                    v = int(m.group(1))
                    if v > maxv: maxv = v
        return maxv


def gen_new(used_set, line, n):
    while True:
        if line == "SP3M3":
            cand = f"SP3M{n}"
        else:
            cand = f"HCAMS02-{n:05d}"
        if cand not in used_set:
            return cand, n
        n += 1


def main():
    items = load_unregistered()
    items = [r for r in items if not r["pn"].startswith("W")]
    print(f"[1/5] 미등록 모품번 (W접두사 제외): {len(items)}")

    # 백업
    bak = MASTER.with_name(MASTER.stem + f"_pre_addunreg_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(MASTER, bak)
    print(f"[2/5] 마스터 백업: {bak.name}")

    wb = openpyxl.load_workbook(MASTER)
    sp_ws = wb[PRIMARY]
    hc_ws = wb[PAIRED]

    # 충돌 set
    sp_used = collect_part_set(sp_ws)
    hc_used = collect_part_set(hc_ws)
    # cache PART_NO도 포함
    for cache_path in [CACHE_SP3M3, CACHE_WSTRIP]:
        if cache_path.exists():
            c = json.loads(cache_path.read_text(encoding="utf-8"))
            sp_used |= collect_erp_parts(c, PRIMARY)
            hc_used |= collect_erp_parts(c, PAIRED)
    # recheck cache
    for p in sorted(CACHE_RECHECK.glob("erp_recheck_*.json")):
        c = json.loads(p.read_text(encoding="utf-8"))
        sp_used |= collect_erp_parts(c, PRIMARY)
        hc_used |= collect_erp_parts(c, PAIRED)

    sp_max = parse_max_seq(sp_used, PRIMARY)
    hc_max = parse_max_seq(hc_used, PAIRED)
    print(f"[3/5] 사용중 max — SP3M3={sp_max} / HCAMS02={hc_max}")

    # 중복 검사: 마스터에 이미 (PROD_NO, 라인) 있으면 skip
    def has_row(ws, pn):
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r and r[0] and str(r[0]).strip() == pn:
                return True
        return False

    sp_next = sp_max + 1
    hc_next = hc_max + 1
    sp_added = hc_added = sp_skip = hc_skip = 0

    for r in items:
        pn = r["pn"]
        car = r["car"]
        # SP3M3
        if has_row(sp_ws, pn):
            sp_skip += 1
        else:
            new_part, sp_next = gen_new(sp_used, PRIMARY, sp_next)
            sp_used.add(new_part)
            sp_next += 1
            sp_ws.append([pn, TARGET_CMPY, PRIMARY, new_part, 1, "정단가", None, car])
            sp_added += 1
        # HCAMS02 (ELR 면제 제외)
        if not r["exempt"]:
            if has_row(hc_ws, pn):
                hc_skip += 1
            else:
                new_part, hc_next = gen_new(hc_used, PAIRED, hc_next)
                hc_used.add(new_part)
                hc_next += 1
                hc_ws.append([pn, TARGET_CMPY, PAIRED, new_part, 1, "정단가", None, car])
                hc_added += 1

    wb.save(MASTER)
    print(f"[4/5] 마스터 append: SP3M3 +{sp_added} (skip {sp_skip}) / HCAMS02 +{hc_added} (skip {hc_skip})")
    print(f"[5/5] 마스터 save: {MASTER.name}")
    print(f"\n=== 완료 ===")
    print(f"신규 시퀀스: SP3M3 {sp_max+1}~{sp_next-1} ({sp_added}개)")
    print(f"             HCAMS02 {hc_max+1}~{hc_next-1} ({hc_added}개)")


if __name__ == "__main__":
    main()
