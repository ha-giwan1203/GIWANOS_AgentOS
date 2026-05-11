"""
W접두사 8개 모품번 마스터 V2 sync.
- 마스터(.xlsm 생산지시서)엔 W접두사 표기 (W89870BS000 등)
- ERP·GERP 실표기는 W 없음 (89870BS000)
- V2 정산 빌더 마스터에는 W strip 형식으로 추가 (빌더가 GERP raw와 매칭하므로)

흐름:
1. W strip 8개 모품번 ERP 재조회 (Hdr+Dtl, 0109 SP3M3·HCAMS02 라인만)
2. ERP 컬러별 행에서 (PROD_NO, PART_NO, ASSY_COST) 추출
3. V2 마스터 SP3M3·HCAMS02 시트에 행 추가
4. 양식에도 라인누락 행으로 추가 (오류유형 "라인코드 누락")
"""
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import lookup_lines

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
CACHE_OUT = REPO / "05_생산실적/조립비정산/05월/_cache/erp_w_strip.json"
ERR = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

W_TARGETS = [
    "W89870BS000", "W89880BS000",
    "W89870BS200", "W89880BS200",
    "W89870BS500", "W89880BS500",
    "W89870DE200", "W89880DE200",
]

# W strip + 차종 매핑 (마스터 .xlsm 등재 정보)
W_TO_NORMAL_CAR = {
    "W89870BS000": ("89870BS000", "SP3(180도) (셀토스)"),
    "W89880BS000": ("89880BS000", "SP3(270도) (셀토스)"),
    "W89870BS200": ("89870BS200", "SP3(180도) (셀토스)"),
    "W89880BS200": ("89880BS200", "SP3(270도) (셀토스)"),
    "W89870BS500": ("89870BS500", "SP3 (180도) (셀토스)"),
    "W89880BS500": ("89880BS500", "SP3 (270도) (셀토스)"),
    "W89870DE200": ("89870DE200", "SP3(180도) (셀토스)"),
    "W89880DE200": ("89880DE200", "SP3(270도) (셀토스)"),
}

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"


def main():
    normal_pns = [W_TO_NORMAL_CAR[w][0] for w in W_TARGETS]
    print(f"[1/5] W strip 모품번: {normal_pns}")

    print(f"[2/5] ERP 재조회 (appl_da=2026-05-11)")
    result = lookup_lines.lookup_root_pns_with_colors(normal_pns, appl_da="2026-05-11", cmpy_cd=TARGET_CMPY)
    CACHE_OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  raw → {CACHE_OUT.name}")

    # 0109 + SP3M3/HCAMS02 라인 행 추출
    sp_rows = []  # 마스터 SP3M3 추가 행
    hc_rows = []
    err_rows = []  # 양식 추가 행

    for w_pn in W_TARGETS:
        normal, car = W_TO_NORMAL_CAR[w_pn]
        entry = result.get(normal, {})
        for color, rows in entry.get("lines_by_color", {}).items():
            sp_match = [r for r in rows if r.get("ASSY_LINE_CD") == PRIMARY and r.get("ASSY_CMPY_CD") == TARGET_CMPY]
            hc_match = [r for r in rows if r.get("ASSY_LINE_CD") == PAIRED and r.get("ASSY_CMPY_CD") == TARGET_CMPY]
            if sp_match:
                m = sp_match[0]
                sp_rows.append([color, TARGET_CMPY, PRIMARY, m.get("PART_NO"),
                                m.get("COMP_QTY", 1), m.get("UNQCST_DIV_NM") or "정단가",
                                m.get("ASSY_COST"), car])
                err_rows.append({
                    "품번": color, "업체코드": TARGET_CMPY, "라인코드": PRIMARY,
                    "조립품번": m.get("PART_NO"), "Usage": 1, "단가구분": "정단가",
                    "단가": m.get("ASSY_COST"), "차종": car,
                    "오류유형": "라인코드 누락", "비고": "신규등록 (W접두사 sync)",
                })
            if hc_match:
                m = hc_match[0]
                hc_rows.append([color, TARGET_CMPY, PAIRED, m.get("PART_NO"),
                                m.get("COMP_QTY", 1), m.get("UNQCST_DIV_NM") or "정단가",
                                m.get("ASSY_COST"), car])
                err_rows.append({
                    "품번": color, "업체코드": TARGET_CMPY, "라인코드": PAIRED,
                    "조립품번": m.get("PART_NO"), "Usage": 1, "단가구분": "정단가",
                    "단가": m.get("ASSY_COST"), "차종": car,
                    "오류유형": "라인코드 누락", "비고": "신규등록 (W접두사 sync)",
                })

    print(f"[3/5] 추출: SP3M3 {len(sp_rows)} / HCAMS02 {len(hc_rows)} / 양식 {len(err_rows)}")

    # 4. 마스터 V2 sync
    bak = MASTER.with_name(MASTER.stem + f"_pre_wprefix_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(MASTER, bak)
    print(f"[4/5] 마스터 백업: {bak.name}")

    wb = openpyxl.load_workbook(MASTER)
    # 중복 검사: 같은 PROD_NO + 라인 있으면 skip
    def get_existing_pns(ws):
        s = set()
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r and r[0]:
                s.add(str(r[0]).strip())
        return s

    sp_ws = wb[PRIMARY]
    hc_ws = wb[PAIRED]
    sp_existing = get_existing_pns(sp_ws)
    hc_existing = get_existing_pns(hc_ws)

    sp_added = hc_added = 0
    for r in sp_rows:
        if r[0] not in sp_existing:
            sp_ws.append(r)
            sp_existing.add(r[0])
            sp_added += 1
    for r in hc_rows:
        if r[0] not in hc_existing:
            hc_ws.append(r)
            hc_existing.add(r[0])
            hc_added += 1
    wb.save(MASTER)
    print(f"  마스터 append: SP3M3 +{sp_added} / HCAMS02 +{hc_added}")

    # 5. 양식 append (PermissionError 시 cache만 저장)
    err_json = REPO / "05_생산실적/조립비정산/05월/_cache/err_w_prefix_rows.json"
    err_json.write_text(json.dumps(err_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[5/5] 양식 추가 시도")
    try:
        ebak = ERR.with_name(ERR.stem + f"_bak_wprefix_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        shutil.copy(ERR, ebak)
        wbe = openpyxl.load_workbook(ERR)
        wse = wbe["오류리스트"]
        template_cells = [wse.cell(row=4, column=c) for c in range(1, 11)]
        cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
                "단가구분", "단가", "차종", "오류유형", "비고"]
        start = wse.max_row + 1
        for i, r in enumerate(err_rows, start=start):
            for c, k in enumerate(cols, start=1):
                cell = wse.cell(row=i, column=c, value=r.get(k))
                tc = template_cells[c - 1]
                cell.font = copy(tc.font)
                cell.alignment = copy(tc.alignment)
                cell.border = copy(tc.border)
                cell.number_format = tc.number_format
        wbe.save(ERR)
        print(f"  양식 append: {len(err_rows)}행 (r{start}~r{wse.max_row})")
    except PermissionError:
        print(f"  양식 PermissionError — cache 저장: {err_json}")
        print(f"  사용자 양식 닫은 후 별도 스크립트로 append 가능")


if __name__ == "__main__":
    main()
