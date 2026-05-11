"""
분류 결과 → xlsx (시트별) + summary.md.
"""
import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parent))
from _xlsx_style import format_workbook


SHEET_ORDER = [
    # 최상단 — 품번 자체 ERP 미등록 (가장 시급한 신규 등록 대상)
    ("미등록_일반",            "★미등록(일반)"),
    ("미등록_면제",            "★미등록(면제)"),
    # 라인 누락 — 품번은 있으나 라인 일부 미등록
    ("양쪽누락_일반",          "양쪽누락(일반)"),
    ("primary누락_일반",       "primary 누락(일반)"),
    ("paired누락_일반",        "paired 누락(일반)"),
    ("primary누락_면제",       "primary 누락(면제)"),
    # 도메인 룰 위배 후보
    ("이상_면제paired등록",     "이상(면제+paired)"),
    ("컬러편차",               "컬러편차"),
    # 정상
    ("정상_일반",              "정상(일반)"),
    ("정상_면제",              "정상(면제)"),
]


COLUMN_ORDER = ["pn", "차종", "exempt", "all_lines", "colors_n", "_diag"]
COLUMN_LABEL = {
    "pn": "모품번",
    "차종": "차종",
    "exempt": "면제(ELR등)",
    "all_lines": "ERP 등록 라인 목록",
    "colors_n": "컬러 수",
    "_diag": "진단",
}


def write_xlsx(classified, out_path: Path, primary_cd: str, paired_cd: str, applied_da: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for key, label in SHEET_ORDER:
        rows = classified.get(key, [])
        ws = wb.create_sheet(label[:31])  # sheet name 31자 제한
        # extra cols: row의 원본 헤더 키 union
        extra_keys = set()
        for r in rows:
            for k in r.keys():
                if k not in COLUMN_ORDER and not k.startswith("_"):
                    extra_keys.add(k)
        extra = sorted(extra_keys - {"colors"})
        cols = COLUMN_ORDER + extra

        labels = [COLUMN_LABEL.get(c, c) for c in cols]
        ws.append(labels)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            ws.append([r.get(c, "") for c in cols])

        # auto width 간단
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(
                10, min(40, len(str(labels[i-1])) * 2 + 2)
            )

    # raw 요약 시트
    ws = wb.create_sheet("raw요약")
    ws.append(["분류", "건수"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for key, label in SHEET_ORDER:
        ws.append([label, len(classified.get(key, []))])
    ws.append([])
    ws.append(["기준일", applied_da])
    ws.append(["primary 라인", primary_cd])
    ws.append(["paired 라인", paired_cd or "(없음 — 단독 라인)"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    format_workbook(wb, header_row=1, left_cols={cols.index("_diag")+1} if "_diag" in cols else None)
    wb.save(out_path)
    print(f"[report] xlsx → {out_path}")


def write_summary_md(classified, out_path: Path, primary_cd, paired_cd, applied_da, master_path, total_unique, exempt_n):
    parts = []
    parts.append(f"# 등록 누락 점검 — {primary_cd}\n")
    parts.append(f"- 기준일: {applied_da}")
    parts.append(f"- primary 라인: **{primary_cd}**")
    parts.append(f"- paired 라인: **{paired_cd or '(없음 — 단독 라인)'}**")
    parts.append(f"- 마스터 파일: `{master_path}`")
    parts.append(f"- 입력 모품번 unique: {total_unique}건 (면제 {exempt_n}건 / 일반 {total_unique - exempt_n}건)")
    parts.append("")

    # 최상단 — 품번 자체 미등록 강조 섹션
    n_unreg_general = len(classified.get("미등록_일반", []))
    n_unreg_exempt = len(classified.get("미등록_면제", []))
    n_unreg_total = n_unreg_general + n_unreg_exempt
    parts.append("## ★ 품번 자체 ERP 미등록 (신규 등록 시급)\n")
    parts.append(f"**총 {n_unreg_total}건** (일반 {n_unreg_general} / 면제 {n_unreg_exempt})")
    if n_unreg_total > 0:
        parts.append("")
        parts.append("ERP 조립비 현황관리(New)에서 0109 업체 기준 행 자체가 0건. 단가·라인을 신규 등록해야 정산 진입 가능.")
        parts.append("xlsx의 `★미등록(일반)` / `★미등록(면제)` 시트 참조.")
    else:
        parts.append("")
        parts.append("0건 — 모든 모품번이 ERP에 1행 이상 등록됨.")
    parts.append("")

    parts.append("## 분류별 건수\n")
    parts.append("| 분류 | 건수 | 액션 |")
    parts.append("|------|------|------|")
    action_map = {
        "primary누락_일반": f"{primary_cd} 라인 추가 등록 필요",
        "paired누락_일반": f"{paired_cd} 라인 추가 등록 필요",
        "양쪽누락_일반": f"{primary_cd}+{paired_cd} 두 라인 모두 추가 등록",
        "미등록_일반": "ERP 신규 단가 등록 필요",
        "primary누락_면제": f"{primary_cd} 라인 추가 등록 필요 (paired 면제)",
        "미등록_면제": "ERP 신규 단가 등록 필요 (paired 면제)",
        "이상_면제paired등록": f"면제 표기인데 {paired_cd} 등록됨 — 도메인 룰 확인",
        "컬러편차": "같은 모품번 컬러별 등록 편차 — 사용자 확인",
        "정상_일반": "정상",
        "정상_면제": "정상 (paired 면제)",
    }
    for key, label in SHEET_ORDER:
        n = len(classified.get(key, []))
        parts.append(f"| {label} | {n} | {action_map.get(key, '')} |")
    parts.append("")

    # 차종 분포 (누락 합산)
    fail_keys = ["primary누락_일반", "paired누락_일반", "양쪽누락_일반", "미등록_일반",
                 "primary누락_면제", "미등록_면제", "이상_면제paired등록"]
    car_counter = Counter()
    for k in fail_keys:
        for r in classified.get(k, []):
            car_counter[r.get("차종", "(미상)")] += 1
    if car_counter:
        parts.append("## 차종 분포 (누락 합산)\n")
        parts.append("| 차종 | 건수 |")
        parts.append("|------|------|")
        for car, n in car_counter.most_common():
            parts.append(f"| {car} | {n} |")
        parts.append("")

    # 컬러편차 모품번 list
    diff = classified.get("컬러편차", [])
    if diff:
        parts.append("## 컬러편차 모품번\n")
        for r in diff[:50]:
            parts.append(f"- `{r['pn']}` ({r.get('차종','')}) — {r.get('_diag','')}")
        if len(diff) > 50:
            parts.append(f"- ... 외 {len(diff)-50}건")
        parts.append("")

    parts.append("## 다음 행동\n")
    parts.append(f"1. xlsx 시트별 누락 행 확인")
    parts.append(f"2. ERP 단가관리 > 조립비관리 > 조립비 현황관리(New)에서 라인 추가/신규 등록")
    parts.append(f"3. 정정 후 본 스킬 재실행으로 0건 정합 확인")
    parts.append("")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(parts), encoding="utf-8")
    print(f"[report] summary → {out_path}")


def write_raw_cache(erp_result, out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(erp_result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[report] raw cache → {out_path}")
