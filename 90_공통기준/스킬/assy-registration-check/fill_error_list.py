"""
등록누락점검 결과 + ERP cache → 오류리스트 양식에 컬러별 행 추가.

원칙:
- 컬러편차 (같은 모품번 N개 컬러 중 일부만 SP3M3·HCAMS02 누락):
    누락 컬러 × 누락 라인(SP3M3·HCAMS02) 각 1행. 같은 모품번 등록 컬러의 같은 라인 행에서 PART_NO·ASSY_COST·UNQCST_DIV_NM 가져옴 (같은 기본품번 컬러별 단가 동일 도메인 룰).
- 라인 누락 (같은 모품번 SP3M3·HCAMS02 등록 0건):
    각 컬러 × 누락 라인 1행. cache에 같은 라인 행 없으므로 조립품번·단가 빈칸.
    같은 모품번 다른 라인 등록 정보를 비고에 적어 사용자 ERP 등록 시 참고.

양식 (10컬럼):
    품번 | 업체코드 | 라인코드 | 조립품번 | Usage | 단가구분 | 단가 | 차종 | 오류유형 | 비고
"""
import json
import sys
from pathlib import Path

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


CACHE_PATH = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\조립비정산\05월\_cache\erp_lookup_SP3M3_20260509.json")
CHECK_XLSX = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\조립비정산\05월\등록누락점검_SP3M3_20260509\등록누락점검_SP3M3_20260509.xlsx")
ERR_XLSX = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")

PRIMARY = "SP3M3"
PAIRED = "HCAMS02"
TARGET_CMPY = "0109"

# 도메인 룰
ROOT_PN_LEN = 10           # 모품번 자리수
VALID_COLOR_LEN = 3        # 정상 컬러 3자리 (모품번10 + 컬러3 = 13)
# 컬러 4자리(품번 14자리) = 미사용 품번 → 제외

ERROR_TYPE_LINE_MISSING = "라인코드 누락"
NOTE_NEW_REG = "신규등록"


def is_valid_color(pn: str, root_pn: str) -> bool:
    """컬러 부분 길이 검사. root_pn 길이 기준 + 3자리 컬러만 사용. 4자리는 미사용 품번."""
    # root_pn은 cache 키 (모품번). pn은 13자리 정상 또는 14자리 미사용.
    expected_len = len(root_pn) + VALID_COLOR_LEN
    return len(pn) == expected_len


def load_classify():
    """등록누락점검 xlsx에서 분류 행 추출 — 라인누락 4시트 + 컬러편차."""
    wb = openpyxl.load_workbook(CHECK_XLSX, data_only=True, read_only=True)
    out = {}
    for sname in ["양쪽누락(일반)", "primary 누락(일반)", "paired 누락(일반)",
                  "primary 누락(면제)", "컬러편차"]:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        first = True
        rows = []
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not row[0]:
                continue
            rows.append({
                "pn": str(row[0]).strip(),
                "차종": str(row[1] or "").strip().replace("\n", " "),
                "exempt": bool(row[2]) and str(row[2]).strip().upper() == "ELR",
                "all_lines": (row[3] or "").split(",") if row[3] else [],
                "colors_n": row[4],
                "_diag": row[5] if len(row) > 5 else "",
            })
        out[sname] = rows
    return out


def build_rows(cache, classified):
    """행 생성 — 라인누락(컬러별 분해) + 컬러편차(누락 컬러별)."""
    rows = []

    # === 라인 누락 4시트 — 컬러별 분해, cache에 같은 라인 0건이라 cost·조립품번 빈칸 ===
    for sheet, items in classified.items():
        if sheet == "컬러편차":
            continue
        is_exempt_sheet = "면제" in sheet
        for it in items:
            pn = it["pn"]
            entry = cache.get(pn, {})
            colors = entry.get("colors") or []
            lines_by_color = entry.get("lines_by_color", {})

            # 누락 라인 결정 (시트별)
            if sheet == "양쪽누락(일반)":
                missing_lines = [PRIMARY, PAIRED]
            elif sheet == "primary 누락(일반)":
                missing_lines = [PRIMARY]
            elif sheet == "paired 누락(일반)":
                missing_lines = [PAIRED]
            elif sheet == "primary 누락(면제)":
                missing_lines = [PRIMARY]
            else:
                continue

            # 같은 모품번 다른 라인 정보 (비고용)
            other_info = []
            for c in colors:
                for r in lines_by_color.get(c, []):
                    if r.get("ASSY_CMPY_CD") != TARGET_CMPY:
                        continue
                    line = r.get("ASSY_LINE_CD")
                    if line in missing_lines:
                        continue
                    other_info.append((c, line, r.get("ASSY_COST"), r.get("PART_NO")))
            # 라인별 cost 평균/대표값 1줄 요약
            line_summary = {}
            for c, line, cost, partno in other_info:
                line_summary.setdefault(line, []).append(cost)
            other_note = "; ".join(f"{l}={costs[0]}원" for l, costs in line_summary.items()) if line_summary else "다른 라인 등록 없음"

            for color in colors:
                # 컬러 4자리(미사용 품번) 필터
                if not is_valid_color(color, pn):
                    continue
                for miss in missing_lines:
                    rows.append({
                        "품번": color,
                        "업체코드": TARGET_CMPY,
                        "라인코드": miss,
                        "조립품번": None,
                        "Usage": 1,
                        "단가구분": "정단가",
                        "단가": None,
                        "차종": it["차종"],
                        "오류유형": ERROR_TYPE_LINE_MISSING,
                        "비고": NOTE_NEW_REG,
                    })

    # === 컬러편차 — 같은 모품번 다른 컬러의 같은 라인 정보 차용 ===
    diff_items = classified.get("컬러편차", [])
    # 컬러편차 시트는 모품번 N행 (라인별 진단). 같은 pn이 primary/paired 두 번 등장 가능. 통합.
    pn_to_lines = {}
    for it in diff_items:
        pn = it["pn"]
        diag = it["_diag"] or ""
        # diag 형식: "primary(SP3M3) 컬러편차: A=O, B=X, ..."
        if "primary" in diag:
            line = PRIMARY
        elif "paired" in diag:
            line = PAIRED
        else:
            continue
        pn_to_lines.setdefault(pn, {"차종": it["차종"], "lines": {}})["lines"][line] = diag

    for pn, info in pn_to_lines.items():
        entry = cache.get(pn, {})
        colors = entry.get("colors") or []
        lines_by_color = entry.get("lines_by_color", {})

        for line, diag in info["lines"].items():
            # 등록 컬러 (line 등록 있는 컬러)
            registered = []
            missing_colors = []
            for c in colors:
                rows_c = lines_by_color.get(c, [])
                has = any((r.get("ASSY_LINE_CD") == line and r.get("ASSY_CMPY_CD") == TARGET_CMPY) for r in rows_c)
                if has:
                    registered.append(c)
                else:
                    missing_colors.append(c)

            # 등록 컬러 행에서 cost·PART_NO 대표값 (컬러별 단가 동일 도메인 룰)
            ref_cost = None
            ref_partno = None
            ref_unq = "정단가"
            for c in registered:
                for r in lines_by_color.get(c, []):
                    if r.get("ASSY_LINE_CD") == line and r.get("ASSY_CMPY_CD") == TARGET_CMPY:
                        ref_cost = ref_cost or r.get("ASSY_COST")
                        ref_partno = ref_partno or r.get("PART_NO")
                        ref_unq = r.get("UNQCST_DIV_NM") or ref_unq
                        break
                if ref_cost is not None:
                    break

            for mc in missing_colors:
                # 컬러 4자리(미사용 품번) 필터
                if not is_valid_color(mc, pn):
                    continue
                rows.append({
                    "품번": mc,
                    "업체코드": TARGET_CMPY,
                    "라인코드": line,
                    "조립품번": None,    # 신규 시퀀스는 register_to_master가 부여 (충돌 방지)
                    "Usage": 1,
                    "단가구분": ref_unq,
                    "단가": ref_cost,    # 같은 모품번 등록 컬러 단가 차용 (도메인 룰: 컬러별 단가 동일)
                    "차종": info["차종"],
                    "오류유형": ERROR_TYPE_LINE_MISSING,
                    "비고": NOTE_NEW_REG,
                })

    return rows


def write_to_err_xlsx(rows):
    from copy import copy
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.load_workbook(ERR_XLSX)
    ws = wb["오류리스트"]

    # 데이터 영역 r4부터 끝까지 행 삭제
    max_r = ws.max_row
    if max_r >= 4:
        # 첫 데이터 행(r4) 서식 추출 (template)
        template_styles = []
        for c in range(1, 11):
            tc = ws.cell(row=4, column=c)
            template_styles.append({
                "font": copy(tc.font) if tc.font else Font(name="맑은 고딕", size=11),
                "alignment": copy(tc.alignment) if tc.alignment else Alignment(horizontal="center", vertical="center"),
                "border": copy(tc.border) if tc.border else None,
                "fill": copy(tc.fill) if tc.fill else None,
                "number_format": tc.number_format or "General",
            })
        ws.delete_rows(4, max_r - 3)
        print(f"  데이터 영역 클리어 (r4~r{max_r}, {max_r-3}행 삭제)")
    else:
        template_styles = None

    # default 서식 (template 없을 때)
    default_font = Font(name="맑은 고딕", size=11)
    default_align_center = Alignment(horizontal="center", vertical="center")
    default_align_left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    default_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 정렬: 라인코드 → 품번
    rows.sort(key=lambda x: (x["라인코드"], x["품번"]))

    cols = ["품번", "업체코드", "라인코드", "조립품번", "Usage",
            "단가구분", "단가", "차종", "오류유형", "비고"]
    # 컬럼별 정렬 룰 — 품번/조립품번/차종/오류유형/비고는 가운데, 단가/Usage는 가운데(숫자), 업체코드/라인코드 가운데
    align_per_col = {c: default_align_center for c in cols}

    for i, row in enumerate(rows, start=4):
        for c, k in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=c, value=row.get(k))
            if template_styles:
                ts = template_styles[c - 1]
                cell.font = copy(ts["font"])
                cell.alignment = copy(ts["alignment"])
                if ts["border"]:
                    cell.border = copy(ts["border"])
                if ts["fill"]:
                    cell.fill = copy(ts["fill"])
                cell.number_format = ts["number_format"]
            else:
                cell.font = copy(default_font)
                cell.alignment = copy(align_per_col[k])
                cell.border = copy(default_border)

    wb.save(ERR_XLSX)
    return len(rows)


def main():
    print(f"[1/4] cache load: {CACHE_PATH.name}")
    cache = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
    print(f"[2/4] 분류 시트 load: {CHECK_XLSX.name}")
    classified = load_classify()
    for k, v in classified.items():
        print(f"  {k}: {len(v)}건")
    print(f"[3/4] 행 생성")
    rows = build_rows(cache, classified)
    by_type = {}
    for r in rows:
        by_type[r["오류유형"]] = by_type.get(r["오류유형"], 0) + 1
    print(f"  생성: 총 {len(rows)}행")
    for t, n in by_type.items():
        print(f"    {t}: {n}")
    print(f"[4/4] 양식 저장: {ERR_XLSX.name}")
    n = write_to_err_xlsx(rows)
    print(f"  저장 완료 {n}행")


if __name__ == "__main__":
    main()
