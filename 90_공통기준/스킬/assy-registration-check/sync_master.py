"""
ERP cache → V2 마스터 sync 보고서.

비교 단위: (라인, PROD_NO)
- 마스터 있음 + ERP 있음 + 단가 일치 → 정상
- 마스터 있음 + ERP 있음 + 단가 차이 → 단가차이 (사용자 결정)
- 마스터 있음 + ERP 없음 → 마스터잔존 (양산 정지 가능성)
- 마스터 없음 + ERP 있음 → ERP누락 (마스터 추가 필요)

1차 보고서만. 사용자 검토 후 별도 단계로 마스터 update.
"""
import json
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _xlsx_style import format_workbook

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
CACHE = REPO / "05_생산실적/조립비정산/05월/_cache/erp_lookup_SP3M3_20260509.json"
OUT_DIR = REPO / "05_생산실적/조립비정산/05월/등록누락점검_SP3M3_20260509"

TARGET_CMPY = "0109"
LINES = ["SP3M3", "HCAMS02"]


def load_master_rows(line):
    """마스터 시트 데이터 → {PROD_NO: row}."""
    wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)
    ws = wb[line]
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0]:
            continue
        pn = str(r[0]).strip()
        if pn in out:
            # 중복 — 마지막 행 사용
            pass
        out[pn] = {
            "PROD_NO": pn,
            "ASSY_CMPY_CD": str(r[1] or "").strip(),
            "ASSY_LINE_CD": str(r[2] or "").strip(),
            "PART_NO": str(r[3] or "").strip() if r[3] else "",
            "COMP_QTY": r[4],
            "UNQCST_DIV_NM": str(r[5] or "").strip(),
            "ASSY_COST": r[6],
            "CHA_JONG": str(r[7] or "").strip() if r[7] else "",
        }
    return out


def load_erp_rows(cache, line):
    """ERP cache → {PROD_NO: row}. 0109 + line만."""
    out = {}
    for root, entry in cache.items():
        for color, rows in entry.get("lines_by_color", {}).items():
            for r in rows:
                if r.get("ASSY_LINE_CD") != line:
                    continue
                if r.get("ASSY_CMPY_CD") != TARGET_CMPY:
                    continue
                pn = (r.get("PROD_NO") or color or "").strip()
                if not pn:
                    continue
                out[pn] = r
    return out


def compare(master_rows, erp_rows):
    """비교 결과 4분류."""
    master_only = []
    erp_only = []
    cost_diff = []
    ok = []
    for pn in sorted(set(master_rows.keys()) | set(erp_rows.keys())):
        m = master_rows.get(pn)
        e = erp_rows.get(pn)
        if m and not e:
            master_only.append(m)
        elif e and not m:
            erp_only.append(e)
        elif m and e:
            m_cost = m.get("ASSY_COST")
            e_cost = e.get("ASSY_COST")
            if m_cost != e_cost:
                cost_diff.append({"master": m, "erp": e})
            else:
                ok.append(m)
    return master_only, erp_only, cost_diff, ok


def write_report(per_line):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUT_DIR / "마스터_ERP_sync_보고.xlsx"
    md_path = OUT_DIR / "마스터_ERP_sync_보고.md"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    md = ["# 마스터 vs ERP cache sync 보고\n",
          "비교 범위: SP3M3 점검 cache 보유 모품번 (279건)",
          "ERP 권위 가정. 차이 발견 시 사용자 검토 후 update.\n"]

    for line, (master_only, erp_only, cost_diff, ok) in per_line.items():
        md.append(f"\n## {line}")
        md.append(f"- 정상 (마스터=ERP): {len(ok)}")
        md.append(f"- 단가차이 (둘 다 있는데 단가 다름): **{len(cost_diff)}**")
        md.append(f"- ERP누락 (ERP만 있고 마스터 없음): **{len(erp_only)}**")
        md.append(f"- 마스터잔존 (마스터만 있고 ERP cache 범위 내 없음): {len(master_only)}")
        md.append("")

        # ERP누락 시트
        ws = wb.create_sheet(f"{line}_ERP누락"[:31])
        ws.append(["PROD_NO", "PART_NO", "단가", "Usage", "단가구분", "차종"])
        for r in erp_only:
            ws.append([r.get("PROD_NO"), r.get("PART_NO"), r.get("ASSY_COST"),
                       r.get("COMP_QTY"), r.get("UNQCST_DIV_NM"), r.get("CHA_JONG", "")])

        # 단가차이 시트
        ws = wb.create_sheet(f"{line}_단가차이"[:31])
        ws.append(["PROD_NO", "마스터 단가", "ERP 단가", "차이", "마스터 PART_NO", "ERP PART_NO"])
        for d in cost_diff:
            m, e = d["master"], d["erp"]
            mc, ec = m.get("ASSY_COST"), e.get("ASSY_COST")
            try:
                diff = (ec or 0) - (mc or 0)
            except TypeError:
                diff = "N/A"
            ws.append([m.get("PROD_NO"), mc, ec, diff, m.get("PART_NO"), e.get("PART_NO")])

        # 마스터잔존 시트 — 정보용
        ws = wb.create_sheet(f"{line}_마스터잔존"[:31])
        ws.append(["PROD_NO", "PART_NO", "단가"])
        for r in master_only[:300]:
            ws.append([r.get("PROD_NO"), r.get("PART_NO"), r.get("ASSY_COST")])

    format_workbook(wb, header_row=1)
    wb.save(xlsx_path)
    md_path.write_text("\n".join(md), encoding="utf-8")
    return xlsx_path, md_path


def main():
    cache = json.loads(CACHE.read_text(encoding="utf-8"))
    print(f"[1/3] cache load: {len(cache)} 모품번")

    per_line = {}
    for line in LINES:
        master = load_master_rows(line)
        erp = load_erp_rows(cache, line)
        # 마스터잔존은 cache 범위 내 모품번에 한정 — cache 모품번 10자리 prefix로 매칭
        cache_pn_prefixes = set(cache.keys())  # 10자리 모품번
        def in_cache_range(pn):
            for prefix in cache_pn_prefixes:
                if pn.startswith(prefix):
                    return True
            return False
        master_in_range = {k: v for k, v in master.items() if in_cache_range(k)}
        result = compare(master_in_range, erp)
        per_line[line] = result
        print(f"  [{line}] 마스터(범위내) {len(master_in_range)} / ERP cache {len(erp)}")
        print(f"    정상={len(result[3])} 단가차이={len(result[2])} ERP누락={len(result[1])} 마스터잔존={len(result[0])}")

    print(f"[2/3] 보고서 작성")
    xlsx, md = write_report(per_line)
    print(f"[3/3] 완료")
    print(f"  xlsx: {xlsx}")
    print(f"  md: {md}")


if __name__ == "__main__":
    main()
