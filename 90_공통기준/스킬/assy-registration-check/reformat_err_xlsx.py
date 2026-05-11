"""
오류리스트 양식의 값을 보존한 채로 서식만 일관 적용.
- 폰트: 맑은 고딕 11pt
- 정렬: 가운데 정렬 (수직·수평)
- 테두리: thin 검정
- 데이터 행만 (r4~), 헤더 행(r1~r3)은 그대로 보존

사용자가 추가 편집한 값(예: ERP 등록 4행 삭제, 단가 직접 입력)은 그대로 유지.
"""
import os
import sys
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

ERR_XLSX = Path(r"\\210.216.217.180\zz-group\★ 신규 시스템\1. GERP 조립비\완료\정합성 검증\★ 2026년 오류리스트\4월 오류리스트\대원테크\오류리스트_04월_추가.xlsx")


def main():
    bak = ERR_XLSX.with_name(ERR_XLSX.stem + f"_bak_reformat_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(ERR_XLSX, bak)
    print(f"백업: {bak.name}")

    wb = openpyxl.load_workbook(ERR_XLSX)
    ws = wb["오류리스트"]

    font = Font(name="맑은 고딕", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 데이터 행 r4~max_row, col A~J(1~10)
    applied = 0
    for r in range(4, ws.max_row + 1):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            # 값이 있는 행만 — 빈 행은 skip
            if cell.value is None and all(ws.cell(row=r, column=cc).value is None for cc in range(1, 11)):
                continue
            cell.font = copy(font)
            cell.alignment = copy(align)
            cell.border = copy(border)
            applied += 1

    # 컬럼 너비 자동 조정 (값 길이 기반 대략치)
    col_widths = {1: 18, 2: 10, 3: 12, 4: 22, 5: 8, 6: 10, 7: 10, 8: 18, 9: 18, 10: 14}
    for c, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    wb.save(ERR_XLSX)
    print(f"서식 적용: {applied}셀")
    print(f"컬럼 너비 조정: 10개 컬럼")


if __name__ == "__main__":
    main()
