"""
sync_master 보고서 기반 마스터 일괄 update.
- 단가차이 → ERP 단가·PART_NO로 마스터 update
- ERP누락 → 마스터에 ERP 행 추가
"""
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


REPO = Path(r"C:\Users\User\Desktop\업무리스트")
MASTER = REPO / "05_생산실적/조립비정산/01_기준정보/기준정보_라인별정리_최종_V2_20260506.xlsx"
CACHE = REPO / "05_생산실적/조립비정산/05월/_cache/erp_lookup_SP3M3_20260509.json"
LOG = REPO / "05_생산실적/조립비정산/05월/등록누락점검_SP3M3_20260509/apply_sync_log.txt"

TARGET_CMPY = "0109"
LINES = ["SP3M3", "HCAMS02"]


def load_erp(cache, line):
    out = {}
    for root, entry in cache.items():
        for color, rows in entry.get("lines_by_color", {}).items():
            for r in rows:
                if r.get("ASSY_LINE_CD") != line:
                    continue
                if r.get("ASSY_CMPY_CD") != TARGET_CMPY:
                    continue
                pn = (r.get("PROD_NO") or color or "").strip()
                if not pn:
                    continue
                out[pn] = r
    return out


def main():
    # 백업
    bak = MASTER.with_name(MASTER.stem + f"_pre_applysync_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy(MASTER, bak)
    print(f"백업: {bak.name}")

    cache = json.loads(CACHE.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(MASTER)

    log_lines = []
    log_lines.append(f"# apply_sync 로그 {datetime.now():%Y-%m-%d %H:%M:%S}\n")

    for line in LINES:
        ws = wb[line]
        erp_rows = load_erp(cache, line)

        # 마스터 PROD_NO → row_index 매핑
        pn_to_row = {}
        for r_idx in range(4, ws.max_row + 1):
            pn = ws.cell(row=r_idx, column=1).value
            if pn:
                pn = str(pn).strip()
                pn_to_row.setdefault(pn, r_idx)  # 첫 행 유지

        updated = 0
        appended = 0

        # 1. ERP에 있는 행 처리
        for pn, e in erp_rows.items():
            new_part = (e.get("PART_NO") or "").strip()
            new_cost = e.get("ASSY_COST")
            new_qty = e.get("COMP_QTY", 1)
            new_unq = (e.get("UNQCST_DIV_NM") or "정단가").strip()
            new_car = (e.get("CHA_JONG") or "").strip()

            if pn in pn_to_row:
                # 마스터에 있음 — 단가·PART_NO 차이 시 update
                r_idx = pn_to_row[pn]
                cur_cost = ws.cell(row=r_idx, column=7).value
                cur_part = (str(ws.cell(row=r_idx, column=4).value or "").strip())
                if cur_cost != new_cost or cur_part != new_part:
                    log_lines.append(f"[{line}] UPDATE r{r_idx} {pn}: cost {cur_cost}→{new_cost} / PART_NO {cur_part}→{new_part}")
                    ws.cell(row=r_idx, column=4, value=new_part)
                    ws.cell(row=r_idx, column=7, value=new_cost)
                    updated += 1
            else:
                # 마스터에 없음 — append
                ws.append([pn, TARGET_CMPY, line, new_part, new_qty, new_unq, new_cost, new_car])
                log_lines.append(f"[{line}] APPEND {pn} PART_NO={new_part} cost={new_cost}")
                appended += 1

        print(f"[{line}] update {updated} / append {appended}")
        log_lines.append(f"\n[{line}] 합계: update {updated} / append {appended}\n")

    wb.save(MASTER)
    print(f"마스터 save: {MASTER.name}")

    LOG.parent.mkdir(parents=True, exist_ok=True)
    LOG.write_text("\n".join(log_lines), encoding="utf-8")
    print(f"로그: {LOG}")


if __name__ == "__main__":
    main()
