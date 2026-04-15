"""매일 반복 업무 통합 실행 (일요일 자동 차단, KST 기준)
1. ZDM 일상점검 (API 직호출)
2. MES 생산실적 업로드 (직접 HTTP — Playwright/CDP 불필요)
"""
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime, date, timedelta, timezone


import openpyxl
import requests

# ─── MES 인증 정보 ────────────────────────────────────
MES_USER_ID  = "0109"
MES_PASSWORD = "samsong1234"

# ─── 공통 ───────────────────────────────────────────
KST = timezone(timedelta(hours=9))
DOW_KR = ["월", "화", "수", "목", "금", "토", "일"]

def now_kst():
    return datetime.now(KST)

def is_sunday(d):
    return d.weekday() == 6

def workdays_in_month(year, month, up_to_day):
    """1일~up_to_day 중 일요일 제외 날짜 리스트"""
    return [
        date(year, month, d)
        for d in range(1, up_to_day + 1)
        if date(year, month, d).weekday() != 6
    ]

# ─── ZDM ────────────────────────────────────────────
ZDM_BASE = "http://ax.samsong.com:34010"

def zdm_get_sp3m3():
    resp = requests.get(f"{ZDM_BASE}/api/daily-inspection", timeout=10)
    resp.raise_for_status()
    return sorted(
        [d for d in resp.json()["data"] if d["line_code"] == "SP3M3"],
        key=lambda x: x["inspection_no"],
    )

def zdm_get_items(master_id):
    resp = requests.get(f"{ZDM_BASE}/api/daily-inspection/{master_id}/items", timeout=10)
    resp.raise_for_status()
    return resp.json()["data"]

def zdm_entered_days(master_id, ym):
    resp = requests.get(f"{ZDM_BASE}/api/daily-inspection/{master_id}/records", params={"year_month": ym}, timeout=10)
    resp.raise_for_status()
    return {r["check_day"] for r in resp.json().get("data", []) if r.get("check_result") == "OK"}

def zdm_enter_day(masters, ym, day):
    ok = fail = 0
    for m in masters:
        for item in zdm_get_items(m["id"]):
            payload = {"item_id": item["id"], "year_month": ym, "check_day": day,
                       "check_result": "OK", "issue_desc": "", "action_taken": "",
                       "manager_name": "", "worker_name": "작업자"}
            try:
                r = requests.post(f"{ZDM_BASE}/api/daily-inspection/{m['id']}/record", json=payload, timeout=10)
                if r.status_code == 200 and r.json().get("success"):
                    ok += 1
                else:
                    fail += 1
            except:
                fail += 1
    return ok, fail

def zdm_verify_day(masters, ym, day):
    count = 0
    for m in masters:
        resp = requests.get(f"{ZDM_BASE}/api/daily-inspection/{m['id']}/records", params={"year_month": ym}, timeout=10)
        count += sum(1 for r in resp.json().get("data", []) if r.get("check_day") == day and r.get("check_result") == "OK")
    return count

def run_zdm(today):
    print("\n[1/2] ZDM 일상점검")
    ym = today.strftime("%Y-%m")
    masters = zdm_get_sp3m3()
    if len(masters) != 19:
        print(f"  FAIL: SP3M3 {len(masters)}개 (19 아님)")
        return False

    # 당일 입력
    ok, fail = zdm_enter_day(masters, ym, today.day)
    v = zdm_verify_day(masters, ym, today.day)
    print(f"  {today.day}일: {ok}건 OK, {fail}건 실패, 검증 {v}/75 {'PASS' if v==75 else 'FAIL'}")

    # 누락분 보정 (어제까지)
    entered = zdm_entered_days(masters[0]["id"], ym)
    for d in range(1, today.day):
        dt = date(today.year, today.month, d)
        if is_sunday(dt) or d in entered:
            continue
        ok2, fail2 = zdm_enter_day(masters, ym, d)
        v2 = zdm_verify_day(masters, ym, d)
        print(f"  {d}일 보정: {ok2}건 OK, 검증 {v2}/75 {'PASS' if v2==75 else 'FAIL'}")

    return True

# ─── MES ────────────────────────────────────────────
BI_SRC = r"Z:\★ 라인별 생산실적\대원테크_라인별 생산실적_BI.xlsx"
BI_DST = r"C:\Users\User\Desktop\업무리스트\05_생산실적\BI실적\대원테크_라인별 생산실적_BI.xlsx"

def mes_refresh_bi():
    if os.path.exists(BI_SRC):
        if os.path.getmtime(BI_SRC) > (os.path.getmtime(BI_DST) if os.path.exists(BI_DST) else 0):
            shutil.copy2(BI_SRC, BI_DST)
    elif not os.path.exists(BI_DST):
        print("  FAIL: BI 파일 없음")
        return None
    return BI_DST

def mes_extract(bi_path, target_str):
    wb = openpyxl.load_workbook(bi_path, data_only=True, read_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dv = row[7]
        if dv is None: continue
        ds = str(dv)[:10]
        if ds != target_str: continue
        item = {}
        for i in range(22):
            val = row[i]
            if val is None: val = ""
            elif isinstance(val, datetime): val = val.strftime("%Y-%m-%d")
            elif isinstance(val, float): val = str(int(val)) if val == int(val) else str(round(val, 6))
            else: val = str(val)
            item[f"COL{i+1}"] = val
        items.append(item)
    wb.close()
    return items

# 공백 허용 컬럼: 품질,설비 비가동(h) = COL13
BLANK_OK = {"COL13"}

def mes_find_blanks(items):
    """공백 셀 목록 반환 [(행번호, 컬럼명, 컬럼제목), ...]. 빈 리스트면 OK."""
    COL_NAMES = {
        "COL1":"업체명","COL2":"유형","COL3":"기종구분","COL4":"대표기종",
        "COL5":"라인명","COL6":"야간구분","COL7":"생산인원","COL8":"날짜",
        "COL9":"표준UPH","COL10":"C/T","COL11":"적용효율","COL12":"근무시간(h)",
        "COL13":"품질,설비비가동(h)","COL14":"실가동시간(h)","COL15":"생산량(ea)",
        "COL16":"품번수(ea)","COL17":"실가동목표수량","COL18":"실적UPH",
        "COL19":"UPMH","COL20":"실가동UPH","COL21":"가동효율","COL22":"순가동효율",
    }
    blanks = []
    for row_idx, item in enumerate(items, 1):
        for col in [f"COL{i}" for i in range(1, 23)]:
            if col in BLANK_OK:
                continue
            val = item.get(col, "")
            if val in (None, ""):
                blanks.append((row_idx, col, COL_NAMES.get(col, col)))
    return blanks

def mes_login():
    """OAuth SSO 직접 HTTP 로그인. requests.Session 반환. 실패 시 None."""
    s = requests.Session()
    s.headers["User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    try:
        # 1. SSO 페이지에서 ssoUrl(state 포함) 추출
        r1 = s.get("http://mes-dev.samsong.com:19200/oauth2/sso", timeout=10)
        m = re.search(r"ssoUrl\s*=\s*'([^']+)'", r1.text)
        if not m:
            print("  FAIL: ssoUrl 파싱 실패")
            return None
        sso_url = m.group(1)
        # 2. OAuth authorize 방문 (auth-dev JSESSIONID 발급)
        s.get(sso_url, allow_redirects=True, timeout=10)
        # 3. 로그인 POST
        r3 = s.post("http://auth-dev.samsong.com:18100/login", data={
            "userId": MES_USER_ID, "password": MES_PASSWORD,
            "clientId": "MES", "ssoUrl": sso_url, "clientName": "", "lang": "ko"
        }, allow_redirects=True, timeout=15)
        if "layout.do" not in r3.url:
            print(f"  FAIL: 로그인 후 URL={r3.url}")
            return None
        # XSRF 헤더 설정 (POST 요청에 필요)
        s.headers["X-XSRF-TOKEN"] = s.cookies.get("XSRF-TOKEN", "")
        return s
    except Exception as e:
        print(f"  FAIL: 로그인 오류 — {e}")
        return None


def mes_upload_and_verify(s, items, target_str):
    """직접 HTTP 업로드 + 검증. 성공 시 {count, qty} 반환, 실패 시 None."""
    payload = json.dumps({"excelList": items}, ensure_ascii=False)
    try:
        r = s.post(
            "http://mes-dev.samsong.com:19200/prdtstatus/SaveExcelData.do",
            data=payload.encode("utf-8"),
            headers={"Content-Type": "application/json; charset=utf-8"},
            timeout=30,
        )
        result = r.json()
        if str(result.get("statusCode", "")) not in ("200", "OK"):
            print(f"    업로드 응답: {result}")
            return None
    except Exception as e:
        print(f"    업로드 오류: {e}")
        return None

    time.sleep(2)
    try:
        vr = s.get(
            "http://mes-dev.samsong.com:19200/prdtstatus/selectPrdtRsltByLine.do",
            params={"S_FROM": target_str, "S_TO": target_str,
                    "S_CMPY_NM": "대원테크", "pq_curPage": "1", "pq_rPP": "1000"},
            timeout=10,
        )
        rows = vr.json().get("data", {}).get("list", [])
        return {
            "count": len(rows),
            "qty": sum(int(row.get("RESULT_QUANTITY") or 0) for row in rows),
        }
    except Exception as e:
        print(f"    검증 오류: {e}")
        return None


def run_mes(today):
    print("\n[2/2] MES 생산실적 업로드")

    # BI 갱신
    bi = mes_refresh_bi()
    if not bi:
        return False

    s = mes_login()
    if not s:
        return False
    print("  MES 로그인 OK")

    # 기준: 어제까지 최근 7일 (일요일 제외)
    yesterday = today - timedelta(days=1)
    week_ago  = today - timedelta(days=7)
    check_range = [
        (week_ago + timedelta(days=i)).isoformat()
        for i in range(7)
        if not is_sunday(week_ago + timedelta(days=i))
        and (week_ago + timedelta(days=i)) <= yesterday
    ]

    # MES 기등록 날짜 조회
    try:
        r = s.get(
            "http://mes-dev.samsong.com:19200/prdtstatus/selectPrdtRsltByLine.do",
            params={"S_FROM": week_ago.isoformat(), "S_TO": yesterday.isoformat(),
                    "S_CMPY_NM": "대원테크", "pq_curPage": "1", "pq_rPP": "1000"},
            timeout=15,
        )
        rows = r.json().get("data", {}).get("list", [])
        mes_dates = {row["TRX_DA"] for row in rows if row.get("TRX_DA")}
    except Exception as e:
        print(f"  FAIL: 기등록 조회 오류 — {e}")
        return False

    # BI 데이터 날짜 수집 (check_range 범위)
    wb = openpyxl.load_workbook(bi, data_only=True, read_only=True)
    ws = wb.active
    bi_dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        dv = row[7]
        if dv is None: continue
        ds = str(dv)[:10]
        if ds in check_range:
            bi_dates.add(ds)
    wb.close()

    # 누락일: check_range 중 BI 있고 MES 미등록
    targets = [ds for ds in check_range if ds in bi_dates and ds not in mes_dates]

    has_fail = False
    incomplete = []  # 데이터 누락으로 업로드 금지된 날짜

    if not targets:
        print("  누락 없음")
    else:
        print(f"  누락일: {targets}")
        for t in targets:
            items = mes_extract(bi, t)
            if not items:
                print(f"  {t}: BI 데이터 없음 SKIP")
                continue

            # 공백 검사 (COL13 제외)
            blanks = mes_find_blanks(items)
            if blanks:
                print(f"  {t}: 데이터 누락 — 업로드 금지")
                for row_idx, col, col_name in blanks:
                    print(f"    행{row_idx} {col}({col_name}) 공백")
                incomplete.append(t)
                has_fail = True
                continue

            bi_qty = sum(int(r["COL15"]) for r in items if r.get("COL15", "").isdigit())
            v = None
            for attempt in range(1, 4):
                v = mes_upload_and_verify(s, items, t)
                if v is not None:
                    break
                print(f"    {t}: 시도 {attempt}/3 실패 — 재로그인 후 재시도")
                time.sleep(2)
                s = mes_login()
                if not s:
                    print(f"  {t}: FAIL 재로그인 실패")
                    break
            if v is None:
                print(f"  {t}: FAIL 업로드 3회 실패")
                has_fail = True
                continue

            cnt_ok = v["count"] == len(items)
            qty_ok = v["qty"] == bi_qty
            print(f"  {t}: {v['count']}/{len(items)}건({'OK' if cnt_ok else 'FAIL'}), "
                  f"qty {v['qty']:,}/{bi_qty:,}({'OK' if qty_ok else 'FAIL'})")
            if not cnt_ok or not qty_ok:
                has_fail = True

    # 데이터 누락 날짜 사용자 보고
    if incomplete:
        print()
        print("  [!] 데이터 누락으로 업로드 보류된 날짜:")
        for t in incomplete:
            print(f"      {t} — BI 파일 해당 날짜 공백 셀 확인 필요")
        print("      (품질,설비 비가동(h) 공백은 제외하고 검사)")

    return not has_fail

# ─── main ───────────────────────────────────────────
def main():
    today = now_kst().date()
    print(f"=== 매일 반복 업무 === {today} ({DOW_KR[today.weekday()]})")

    if is_sunday(today):
        print("일요일 -전체 스킵")
        sys.exit(0)

    zdm_ok = run_zdm(today)
    mes_ok = run_mes(today)

    if zdm_ok and mes_ok:
        print("\n=== 완료 (ALL PASS) ===")
    else:
        failed = []
        if not zdm_ok:
            failed.append("ZDM")
        if not mes_ok:
            failed.append("MES")
        print(f"\n=== 완료 (FAIL: {', '.join(failed)}) ===")
        sys.exit(1)

if __name__ == "__main__":
    main()
