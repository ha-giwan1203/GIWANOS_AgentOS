"""94_라인지원 시트 빌더.

폴더 `{MM}월 지원/` 존재 시 화인텍·리노텍 등 다양한 양식 자동 파싱.
폴더 부재 시 "당월 지원 없음" + B=0.
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import NamedTuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

META_KEYS = {
    "날짜", "차종", "완성품품번", "서브품번", "수량",
    "품번", "지원라인", "합계금액", "합계", "비고",
    "지원구분", "대상라인", "단가", "시간",
}

NEG_TOKENS = ["지원받은", "이관", "받은"]
POS_TOKENS = ["지원해줌", "대원테크 →", "→ "]


class SupportRow(NamedTuple):
    file: str
    sign: int           # +1 가산 / -1 차감
    line: str           # 우리 라인 코드
    amount: float


def _detect_sign(filename: str, sheetname: str, title: str) -> int:
    """파일명·시트명·제목 토큰으로 부호 자동 판정."""
    text = f"{filename} {sheetname} {title}"
    for tok in NEG_TOKENS:
        if tok in text:
            return -1
    for tok in POS_TOKENS:
        if tok in text:
            return +1
    return 0  # 명확 안 됨


def _find_header_row(ws, max_scan: int = 6) -> int | None:
    """첫 max_scan행에서 헤더 행 자동 탐색. "날짜"·"수량"·"합계" 등 키워드 매칭."""
    needed = {"날짜", "수량"}
    for r in range(1, max_scan + 1):
        cells = {str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)
                 if ws.cell(r, c).value}
        if needed.issubset(cells):
            return r
    return None


def parse_support_file(path: Path, target_month: str) -> list[SupportRow]:
    """파일 한 개 파싱. target_month: "04" 형식."""
    rows: list[SupportRow] = []
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in wb.sheetnames:
        if sheet in ("요약", "summary", "Summary"):
            continue
        ws = wb[sheet]
        title = str(ws.cell(1, 1).value or "")
        sign = _detect_sign(path.name, sheet, title)
        if sign == 0:
            print(f"  [WARN] {path.name}/{sheet}: 부호 토큰 불명 — skip")
            continue

        hdr_row = _find_header_row(ws)
        if hdr_row is None:
            print(f"  [WARN] {path.name}/{sheet}: 헤더 행 미발견 — skip")
            continue

        headers = {c: ws.cell(hdr_row, c).value
                   for c in range(1, ws.max_column + 1)}
        # 라인 컬럼 = META_KEYS 제외 + 라인명 패턴 (대문자+숫자)
        line_cols: list[tuple[int, str]] = []
        qty_col = None
        date_col = 1
        for c, h in headers.items():
            if not h:
                continue
            hs = str(h).strip()
            if hs == "수량":
                qty_col = c
            elif hs == "날짜":
                date_col = c
            elif hs not in META_KEYS and re.match(r"^[A-Z][A-Z0-9]{2,}$", hs):
                line_cols.append((c, hs))

        if not line_cols or qty_col is None:
            print(f"  [WARN] {path.name}/{sheet}: 라인/수량 컬럼 미발견 — skip")
            continue

        agg: dict[str, float] = defaultdict(float)
        for r in range(hdr_row + 1, ws.max_row + 1):
            dv = ws.cell(r, date_col).value
            if dv is None:
                continue
            m = re.search(r"(?:^|[^0-9])(0[1-9]|1[0-2])(?:/|\b)",
                          str(dv).replace("\\", "/"))
            if not m or m.group(1) != target_month:
                continue
            qty = ws.cell(r, qty_col).value
            if not isinstance(qty, (int, float)) or qty == 0:
                continue
            for cc, lname in line_cols:
                unit = ws.cell(r, cc).value
                if isinstance(unit, (int, float)) and unit > 0:
                    agg[lname] += unit * qty

        for lname, amt in agg.items():
            rows.append(SupportRow(path.name, sign, lname, amt))
    return rows


def build_sheet_94(wb: openpyxl.Workbook, month_mm: str, base_dir: Path) -> dict:
    """본체 wb에 94_라인지원 시트 추가. 반환: 라인별 순액 dict."""
    folder = base_dir / f"{int(month_mm):02d}월 지원"
    if not folder.exists():
        folder = base_dir / f"{int(month_mm)}월 지원"
    sheet_name = "94_라인지원"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_font = Font(name="맑은 고딕", size=14, bold=True)
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    neg_fill = PatternFill("solid", fgColor="FCE4D6")
    pos_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws.cell(1, 1, f"{month_mm}월 라인지원 정산").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    line_total: dict[str, float] = defaultdict(float)
    pos_total = 0.0
    neg_total = 0.0

    if not folder.exists():
        ws.cell(3, 1, "당월 지원 없음 — 폴더 미존재").font = Font(italic=True, color="808080")
        ws.cell(5, 1, "지원 순액 (B)").font = Font(bold=True)
        ws.cell(5, 2, 0).alignment = right
        return {"line_total": dict(line_total), "B_sum": 0.0,
                "pos_total": 0, "neg_total": 0}

    files = sorted(folder.glob("*.xlsx"))
    if not files:
        ws.cell(3, 1, "당월 지원 없음 — 폴더 내 파일 0").font = Font(italic=True, color="808080")
        ws.cell(5, 1, "지원 순액 (B)").font = Font(bold=True)
        ws.cell(5, 2, 0).alignment = right
        return {"line_total": dict(line_total), "B_sum": 0.0,
                "pos_total": 0, "neg_total": 0}

    # 헤더
    headers = ["파일", "부호", "라인", "금액(원)", "비고"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row = 4
    for f in files:
        rows = parse_support_file(f, month_mm)
        for sr in rows:
            ws.cell(row, 1, sr.file)
            ws.cell(row, 2, "+" if sr.sign > 0 else "−").alignment = center
            ws.cell(row, 3, sr.line).alignment = center
            ws.cell(row, 4, sr.amount * sr.sign).alignment = right
            ws.cell(row, 4).number_format = "#,##0"
            ws.cell(row, 5, "받은지원금(차감)" if sr.sign < 0 else "준지원금(가산)")
            if sr.sign < 0:
                ws.cell(row, 4).fill = neg_fill
                neg_total += sr.amount
            else:
                ws.cell(row, 4).fill = pos_fill
                pos_total += sr.amount
            line_total[sr.line] += sr.amount * sr.sign
            row += 1

    # 합계 행
    row += 1
    ws.cell(row, 1, "지원 순액 (B = 준지원금 − 받은지원금)").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    b_sum = pos_total - neg_total
    ws.cell(row, 4, b_sum).font = Font(bold=True)
    ws.cell(row, 4).number_format = "#,##0"
    ws.cell(row, 4).alignment = right

    row += 1
    ws.cell(row, 1, "  준지원금(+)").alignment = right
    ws.cell(row, 4, pos_total).number_format = "#,##0"
    ws.cell(row, 4).alignment = right

    row += 1
    ws.cell(row, 1, "  받은지원금(−)").alignment = right
    ws.cell(row, 4, -neg_total).number_format = "#,##0"
    ws.cell(row, 4).alignment = right

    # 라인별 순액 표 (오른쪽)
    ws.cell(3, 7, "라인별 순액").font = header_font
    ws.cell(3, 7).fill = header_fill
    ws.cell(3, 7).alignment = center
    ws.cell(3, 8, "금액").font = header_font
    ws.cell(3, 8).fill = header_fill
    ws.cell(3, 8).alignment = center

    r2 = 4
    for ln, v in sorted(line_total.items(), key=lambda x: -abs(x[1])):
        ws.cell(r2, 7, ln).alignment = center
        ws.cell(r2, 8, v).number_format = "#,##0"
        ws.cell(r2, 8).alignment = right
        if v < 0:
            ws.cell(r2, 8).fill = neg_fill
        else:
            ws.cell(r2, 8).fill = pos_fill
        r2 += 1

    # 열 너비
    widths = {1: 38, 2: 6, 3: 12, 4: 16, 5: 24, 6: 2, 7: 12, 8: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    return {
        "line_total": dict(line_total),
        "B_sum": b_sum,
        "pos_total": pos_total,
        "neg_total": neg_total,
    }
