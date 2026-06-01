"""monthly-pnl-rollup — 본체 정산_수식버전_MM월.xlsx에 90/91 시트 추가.

설계:
- 디자인 토큰 + 헬퍼 함수 4종으로 모든 표를 동일 양식으로 그림
- 컬럼 폭 / 헤더 색 / 합계 강조 / 음수 빨강 / 0=하이픈 일관

호출: python run.py --month 04
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent))
from builders.sheet_94_support import parse_support_file  # noqa: E402

REPO = Path(r"C:\Users\User\Desktop\업무리스트")
SETTLE = REPO / "05_생산실적" / "조립비정산"
BI_PATH = REPO / "05_생산실적" / "BI실적" / "대원테크_라인별 생산실적_BI.xlsx"

LINES = ["SD9A01", "ANAAS04", "DRAAS11", "SP3M3", "HASMS02",
         "HCAMS02", "WAMAS01", "WABAS01", "WASAS01", "ISAMS03"]

# ============================================================
# 디자인 토큰 — 파스텔 + 세련된 톤
# ============================================================

# 색상 (파스텔 팔레트)
C_HEADER   = "BDD7EE"  # 표 헤더 연한 파랑
C_HEAD_TXT = "1F3864"  # 헤더 텍스트 진청 (배경이 연하므로 글자는 진하게)
C_TOTAL    = "FFF2CC"  # 합계 연한 노랑
C_POS      = "E2EFDA"  # 양수 연두
C_NEG      = "FCE4D6"  # 음수 연주황
C_SUB      = "DEEBF7"  # 서브 연파랑
C_SECTION  = "1F4E78"  # 섹션 텍스트 진청
C_MUTED    = "8C8C8C"  # 본문 회색
C_GRID     = "D0D7E2"  # 테두리 회색

# KPI 카드 5색 (파스텔)
C_CARD_A   = "D9E2F3"  # A 기본조립비 연파랑
C_CARD_B   = "FCE4D6"  # B 지원순액 연주황 (음수 가능)
C_CARD_C   = "E2EFDA"  # C 비용보전 연초록
C_CARD_D   = "F4CCCC"  # D 차감비용 연빨강
C_CARD_F   = "FFE699"  # 최종 카드 강조 노랑

# 라인별 박스 10색 (파스텔, 라인 시각 구분)
LINE_COLORS = {
    "SD9A01":  "E1D5F2",  # 보라 연한
    "ANAAS04": "FFF2CC",  # 노랑 연한
    "DRAAS11": "C9E5DA",  # 청록 연한
    "SP3M3":   "D6E5F4",  # 파랑 연한
    "HASMS02": "FCE4D6",  # 주황 연한
    "HCAMS02": "F4D6E5",  # 분홍 연한
    "WAMAS01": "DDDDDD",  # 회색
    "WABAS01": "CDDEEE",  # 하늘 연한
    "WASAS01": "D9EAD3",  # 연두 연한
    "ISAMS03": "FCE0BF",  # 살구 연한
}

# 신호등 색상
C_OK       = "C6E0B4"  # PASS 연두 (조금 진함)
C_WARN     = "FFE699"  # WARN/INFO 연노랑
C_FAIL     = "F4A6A6"  # FAIL 연빨강

# 폰트 (위계) — 시인성 위해 조금 축소
F_TITLE     = Font(name="맑은 고딕", size=14, bold=True, color="1F3864")
F_SECTION   = Font(name="맑은 고딕", size=11, bold=True, color=C_SECTION)
F_CARD_LBL  = Font(name="맑은 고딕", size=9, color=C_MUTED)
F_CARD_VAL  = Font(name="맑은 고딕", size=16, bold=True, color="1F3864")
F_CARD_NOTE = Font(name="맑은 고딕", size=8, color=C_MUTED)
F_FINAL_LBL = Font(name="맑은 고딕", size=11, bold=True, color="1F3864")
F_FINAL_VAL = Font(name="맑은 고딕", size=20, bold=True, color="1F3864")
F_HEAD      = Font(name="맑은 고딕", size=10, bold=True, color=C_HEAD_TXT)
F_DATA      = Font(name="맑은 고딕", size=10)
F_TOTAL     = Font(name="맑은 고딕", size=10, bold=True)
F_NOTE      = Font(name="맑은 고딕", size=9, italic=True, color=C_MUTED)
F_SIGNAL    = Font(name="맑은 고딕", size=12, bold=True)

# Fill
FILL_HEAD  = PatternFill("solid", fgColor=C_HEADER)
FILL_TOTAL = PatternFill("solid", fgColor=C_TOTAL)
FILL_POS   = PatternFill("solid", fgColor=C_POS)
FILL_NEG   = PatternFill("solid", fgColor=C_NEG)
FILL_SUB   = PatternFill("solid", fgColor=C_SUB)
FILL_OK    = PatternFill("solid", fgColor=C_OK)
FILL_WARN  = PatternFill("solid", fgColor=C_WARN)
FILL_FAIL  = PatternFill("solid", fgColor=C_FAIL)

# 정렬
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_LW = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 테두리
_SIDE = Side(style="thin", color=C_GRID)
_SIDE_TOP_THICK = Side(style="medium", color="000000")
BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)
BORDER_TOTAL = Border(left=_SIDE, right=_SIDE, top=_SIDE_TOP_THICK, bottom=_SIDE)

# 숫자 형식
FMT_NUM = "#,##0;[Red]-#,##0;-"
FMT_PCT = "0.0%;[Red]-0.0%;-"

# 표준 컬럼 폭 (A~G)
COL_WIDTHS = {1: 28, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14, 7: 42}


# ============================================================
# 헬퍼 함수 (Design System API)
# ============================================================

def apply_col_widths(ws, widths=None):
    """표준 컬럼 폭 적용."""
    w = widths or COL_WIDTHS
    for c, val in w.items():
        ws.column_dimensions[get_column_letter(c)].width = val


def title(ws, row, text, span=7):
    """시트 타이틀 (14pt bold)."""
    ws.cell(row, 1).value = text
    ws.cell(row, 1).font = F_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def section(ws, row, text, span=7):
    """섹션 헤더 (■ 12pt 진청)."""
    ws.cell(row, 1).value = f"■ {text}"
    ws.cell(row, 1).font = F_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def note(ws, row, text, span=7):
    """비고 (※ 9pt italic 회색)."""
    ws.cell(row, 1).value = f"※ {text}"
    ws.cell(row, 1).font = F_NOTE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def table_header(ws, row, headers):
    """표 헤더 — 흰글자 + 진청 fill + 가운데 + 테두리."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = ALIGN_C
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def data_row(ws, row, values, types):
    """데이터 행 — types 별 정렬·포맷·강조 자동 적용.

    types 코드:
      "L"  = 라벨 (좌측)
      "C"  = 코드 (가운데)
      "N"  = 숫자/금액 (우측, #,##0)
      "Np" = 숫자 — 양수 연두 / 음수 연주황
      "P"  = 퍼센트 (우측, 0.0%)
      "T"  = 텍스트 비고 (좌측, wrap)
      "S"  = 상태 (가운데, PASS=연두/INFO=노랑/WARN=연주황/FAIL=연주황)
    """
    for c, (val, t) in enumerate(zip(values, types), 1):
        cell = ws.cell(row, c)
        cell.value = val
        cell.font = F_DATA
        cell.border = BORDER
        if t == "L":
            cell.alignment = ALIGN_L
        elif t == "C":
            cell.alignment = ALIGN_C
        elif t == "N":
            cell.alignment = ALIGN_R
            cell.number_format = FMT_NUM
        elif t == "Np":
            cell.alignment = ALIGN_R
            cell.number_format = FMT_NUM
            if isinstance(val, (int, float)):
                if val > 0:
                    cell.fill = FILL_POS
                elif val < 0:
                    cell.fill = FILL_NEG
        elif t == "P":
            cell.alignment = ALIGN_R
            cell.number_format = FMT_PCT
        elif t == "T":
            cell.alignment = ALIGN_LW
        elif t == "S":
            cell.alignment = ALIGN_C
            s = str(val or "").upper()
            if s == "PASS":
                cell.fill = FILL_POS
            elif s in ("INFO", "WARN"):
                cell.fill = FILL_TOTAL
            elif s == "FAIL":
                cell.fill = FILL_NEG


def total_row(ws, row, values, types, label_text="합계"):
    """합계 행 — 노랑 fill + 굵게 + 상단 medium 테두리."""
    # 1열은 라벨
    ws.cell(row, 1).value = label_text
    for c in range(1, len(values) + 1):
        cell = ws.cell(row, c)
        cell.font = F_TOTAL
        cell.fill = FILL_TOTAL
        cell.border = BORDER_TOTAL
    for c, (val, t) in enumerate(zip(values, types), 1):
        cell = ws.cell(row, c)
        if c == 1:
            cell.alignment = ALIGN_C
            continue
        cell.value = val
        if t == "N":
            cell.alignment = ALIGN_R
            cell.number_format = FMT_NUM
        elif t == "C":
            cell.alignment = ALIGN_C
        elif t == "P":
            cell.alignment = ALIGN_R
            cell.number_format = FMT_PCT
        else:
            cell.alignment = ALIGN_L


# ============================================================
# 데이터 로더
# ============================================================

def extract_bi_night_total(line, year, month):
    if not BI_PATH.exists():
        return 0, 0
    wb = openpyxl.load_workbook(BI_PATH, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    night_val = next((r[5] for r in rows[1:]
                      if r[4] == line and r[5] not in (None, "주간")), None)
    if night_val is None:
        return 0, 0
    total = days = 0
    for r in rows[1:]:
        if r[4] != line or r[5] != night_val:
            continue
        d = r[7]
        if not isinstance(d, datetime) or d.year != year or d.month != month:
            continue
        if r[14]:
            total += r[14]
            days += 1
    return total, days


def load_settlement_totals(book_path, work_folder):
    """라인별 GERP 합계 + 야간수량 계산.

    우선순위:
    1. 본체 정산집계 cached value (기준단가 적용 K+L — 정산 권위값)
    2. cached 손실 시 GERP 원본 raw 합산 (fallback, 다중단가 미반영)
    """
    line_names = {
        "SD9A01": "아우터", "ANAAS04": "앵커", "DRAAS11": "디링",
        "SP3M3": "SP3", "HASMS02": "HASMS", "HCAMS02": "HCAMS",
        "WAMAS01": "웨빙", "WABAS01": "웨빙버클", "WASAS01": "WASAS",
        "ISAMS03": "이너센스",
    }
    agg = {c: {"day_amt": 0, "night_amt": 0,
               "day_qty": 0, "night_qty": 0} for c in LINES}

    # 1차: 본체 정산집계 cached value
    cached_ok = False
    wb = openpyxl.load_workbook(book_path, data_only=True, read_only=True)
    if "정산집계" in wb.sheetnames:
        ws = wb["정산집계"]
        sum_check = 0
        for r in range(2, ws.max_row + 1):
            code = ws.cell(r, 1).value
            if not code or code == "합계":
                continue
            day_amt = ws.cell(r, 5).value or 0
            night_amt = ws.cell(r, 6).value or 0
            night_qty = ws.cell(r, 4).value or 0
            if code in agg:
                agg[code]["day_amt"] = day_amt
                agg[code]["night_amt"] = night_amt
                agg[code]["night_qty"] = night_qty
                sum_check += day_amt + night_amt
        cached_ok = sum_check > 0
    wb.close()

    # 2차: cached 손실 시 GERP raw fallback
    if not cached_ok:
        print("  [WARN] 본체 정산집계 cached value 손실 — GERP raw fallback")
        print("        (사용자 엑셀로 본체 한 번 열고 저장 시 235M 정합값으로 복귀)")
        agg = {c: {"day_amt": 0, "night_amt": 0,
                   "day_qty": 0, "night_qty": 0} for c in LINES}
        src_dir = work_folder / "실적데이터"
        gerp_files = []
        if src_dir.exists():
            gerp_files = (list(src_dir.glob("G-ERP*.xlsx"))
                          + list(src_dir.glob("GERP*.xlsx")))
        if gerp_files:
            gerp_src = gerp_files[0]
            print(f"  GERP 원본 직접 합산: {gerp_src.name}")
            wb_g = openpyxl.load_workbook(gerp_src, data_only=True, read_only=True)
            ws_g = wb_g[wb_g.sheetnames[0]]
            col_line, col_shift, col_qty, col_amt, col_vendor = 3, 14, 15, 17, 21
            for row in ws_g.iter_rows(min_row=2, values_only=True):
                if len(row) < col_vendor:
                    continue
                vendor = row[col_vendor - 1]
                if vendor and str(vendor).strip() != "0109":
                    continue
                line = row[col_line - 1]
                if line not in agg:
                    continue
                shift = row[col_shift - 1]
                qty = row[col_qty - 1] or 0
                amt = row[col_amt - 1] or 0
                if not isinstance(amt, (int, float)):
                    continue
                if not isinstance(qty, (int, float)):
                    qty = 0
                if shift == "추가":
                    agg[line]["night_amt"] += amt
                    agg[line]["night_qty"] += qty
                else:
                    agg[line]["day_amt"] += amt
                    agg[line]["day_qty"] += qty
            wb_g.close()
    else:
        print("  본체 정산집계 cached value 사용 (기준단가 적용 권위값)")

    lines = {}
    grand = 0
    for code in LINES:
        a = agg[code]
        total = a["day_amt"] + a["night_amt"]
        lines[code] = {
            "name": line_names.get(code, code),
            "night_qty": a["night_qty"],
            "night_amt": a["night_amt"],
            "day_amt": a["day_amt"],
            "gerp_total": total,
        }
        grand += total
    return {"lines": lines, "grand_total": grand}


def load_line_stoppage(path):
    if not path.exists():
        return {"exists": False, "rows": [], "total": 0, "blame": []}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows, total = [], 0

    def num(v):
        if isinstance(v, (int, float)):
            return v
        if isinstance(v, str):
            try:
                return float(v.replace(",", "").strip())
            except ValueError:
                return 0
        return 0

    if "집계" in wb.sheetnames:
        ws_sum = wb["집계"]
        for row in ws_sum.iter_rows(values_only=True):
            if row and row[0] == "라인정지":
                amt = round(num(row[2] if len(row) > 2 else 0))
                cnt = row[1] if len(row) > 1 and row[1] else 0
                rows.append({"system": "G-ERP 라인보상",
                             "claim_type": "라인정지",
                             "count": cnt, "amount": amt})
                total += amt
                break

    blame = []
    if "통합집계" in wb.sheetnames:
        ws = wb["통합집계"]
        main_end = ws.max_row
        for r in range(4, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v == "합계":
                main_end = r
                break
            if not v:
                continue
            amt = round(num(ws.cell(r, 4).value))
            rows.append({"system": v, "claim_type": ws.cell(r, 2).value,
                         "count": ws.cell(r, 3).value or 0, "amount": amt})
            total += amt
        in_blame = False
        for r in range(main_end + 1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v == "귀책 부서":
                in_blame = True
                continue
            if not in_blame or not v:
                continue
            if v == "합계":
                break
            blame.append({"blame": v, "count": ws.cell(r, 2).value or 0,
                          "amount": ws.cell(r, 3).value or 0})
    wb.close()
    return {"exists": True, "rows": rows, "total": total, "blame": blame}


def load_error_types(book_path):
    """본체 유형별요약 시트 5행을 91 ④섹션으로 가져옴."""
    wb = openpyxl.load_workbook(book_path, data_only=True, read_only=True)
    if "유형별요약" not in wb.sheetnames:
        wb.close()
        return {"exists": False, "rows": [], "total_cnt": 0, "total_diff": 0}
    ws = wb["유형별요약"]
    rows = []
    total_cnt = total_diff = 0
    for r in range(4, ws.max_row + 1):
        kind = ws.cell(r, 1).value
        if not kind or kind == "합계":
            continue
        cnt = ws.cell(r, 2).value or 0
        gerp = ws.cell(r, 3).value or 0
        old = ws.cell(r, 4).value or 0
        diff = ws.cell(r, 5).value or 0
        rows.append({"kind": kind, "cnt": cnt, "gerp": gerp,
                     "old": old, "diff": diff})
        if isinstance(cnt, (int, float)):
            total_cnt += cnt
        if isinstance(diff, (int, float)):
            total_diff += diff
    wb.close()
    return {"exists": True, "rows": rows,
            "total_cnt": total_cnt, "total_diff": total_diff}


def load_support_data(folder, month_mm):
    result = {"exists": False, "details": [], "line_total": {},
              "pos_total": 0, "neg_total": 0, "B_sum": 0}
    if not folder.exists():
        return result
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        return result
    result["exists"] = True
    line_total = {}
    pos = neg = 0
    for f in files:
        rows = parse_support_file(f, month_mm)
        for sr in rows:
            result["details"].append(sr)
            line_total[sr.line] = line_total.get(sr.line, 0) + sr.amount * sr.sign
            if sr.sign > 0:
                pos += sr.amount
            else:
                neg += sr.amount
    result["line_total"] = line_total
    result["pos_total"] = pos
    result["neg_total"] = neg
    result["B_sum"] = pos - neg
    return result


def resolve_support_dir(work, mm):
    candidates = [
        work / f"{int(mm):02d}월 지원",
        work / f"{int(mm)}월 지원",
    ]
    for folder in candidates:
        if folder.exists():
            return folder
    return candidates[0]


def compute_night_info(lines_data, year, month):
    night_lines = [c for c in LINES
                   if lines_data["lines"].get(c, {}).get("night_qty", 0) > 0]
    bi_summary = {}
    for code in night_lines:
        qty, days = extract_bi_night_total(code, year, month)
        bi_summary[code] = {"qty": qty, "days": days}
    return {"night_lines": night_lines, "bi_summary": bi_summary}


def compute_pnl(lines_data, support, etc_data, night_info):
    """A·B·C·D·E 산출.

    E (BI 차이 청구금액) 산식:
    - SP3M3 한정 적용 (LOT B 분류 불가 라인)
    - 단가 170원 고정 (도메인 룰)
    - 차이 EA가 양수(BI > GERP)일 때만 청구 — 회사가 더 생산했는데 GERP 누락
    - 차이 EA가 음수면 E = 0 (BI가 적은 건 청구 안 함)
    """
    A = lines_data["grand_total"]
    B = support["B_sum"]
    # 라인정지·기타생산비용은 사내부서/납품업체 귀책 → 우리(대원테크)가 받을 금액
    # 사용자 룰 2026-06-01: D 차감 아닌 C 비용보전(가산)으로 처리
    C = etc_data["total"] if etc_data["exists"] else 0
    D = 0
    # E 산출 — SP3M3 한정 170원 고정
    E = 0
    e_detail = []
    NIGHT_UNIT_PRICE = 170  # SP3M3 BI 차이 청구 단가 (도메인 룰)
    for code in ("SP3M3",):  # 현재 SP3M3만 적용. 향후 라인 확장 가능
        if code not in night_info.get("night_lines", []):
            continue
        d = lines_data["lines"].get(code, {})
        bi = night_info.get("bi_summary", {}).get(code, {})
        gn = d.get("night_qty", 0)
        bn = bi.get("qty", 0)
        diff_ea = bn - gn  # BI > GERP면 누락분
        if diff_ea > 0:
            amt = diff_ea * NIGHT_UNIT_PRICE
            E += amt
            e_detail.append({"line": code, "diff_ea": diff_ea,
                             "unit_price": NIGHT_UNIT_PRICE, "amount": amt})
    return {"A": A, "B": B, "C": C, "D": D, "E": E,
            "e_detail": e_detail, "final": A + B + C - D + E}


# ============================================================
# 시트 빌더 (디자인 시스템 적용)
# ============================================================

def _line_card(ws, r0, c0, line_code, line_name, final_value,
               gerp_ref, support_amt, fill_color):
    """라인별 손익 박스 — 3행 구조 (라벨/값/비고). 외곽 테두리만."""
    fill = PatternFill("solid", fgColor=fill_color)
    s = Side(style="thin", color=C_GRID)
    for rr in range(r0, r0 + 3):
        for cc in range(c0, c0 + 3):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = Border(
                left=s if cc == c0 else None,
                right=s if cc == c0 + 2 else None,
                top=s if rr == r0 else None,
                bottom=s if rr == r0 + 2 else None,
            )
    # 라인 코드 + 이름 (r0)
    ws.merge_cells(start_row=r0, start_column=c0,
                   end_row=r0, end_column=c0 + 2)
    ws.cell(r0, c0).value = f"{line_code} ({line_name})"
    ws.cell(r0, c0).font = F_CARD_LBL
    ws.cell(r0, c0).alignment = Alignment(horizontal="left",
                                          vertical="center", indent=1)
    # 최종금액 (r0+1, 큰 숫자)
    ws.merge_cells(start_row=r0 + 1, start_column=c0,
                   end_row=r0 + 1, end_column=c0 + 2)
    ws.cell(r0 + 1, c0).value = final_value
    ws.cell(r0 + 1, c0).font = F_CARD_VAL
    ws.cell(r0 + 1, c0).alignment = ALIGN_C
    ws.cell(r0 + 1, c0).number_format = FMT_NUM
    # 비고 (r0+2)
    ws.merge_cells(start_row=r0 + 2, start_column=c0,
                   end_row=r0 + 2, end_column=c0 + 2)
    sup_str = f" / 지원 {support_amt:+,}" if support_amt else ""
    note_text = f"GERP {sup_str}".strip()
    ws.cell(r0 + 2, c0).value = note_text
    ws.cell(r0 + 2, c0).font = F_CARD_NOTE
    ws.cell(r0 + 2, c0).alignment = Alignment(horizontal="left",
                                              vertical="center", indent=1)
    ws.row_dimensions[r0].height = 15
    ws.row_dimensions[r0 + 1].height = 22
    ws.row_dimensions[r0 + 2].height = 13


def _kpi_card(ws, r0, c0, label, value, note, fill_color):
    """KPI 카드 1장 — 3행 × 3열 (라벨 / 값 / 비고). 외곽 테두리만."""
    fill = PatternFill("solid", fgColor=fill_color)
    s = Side(style="thin", color=C_GRID)
    for rr in range(r0, r0 + 3):
        for cc in range(c0, c0 + 3):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = Border(
                left=s if cc == c0 else None,
                right=s if cc == c0 + 2 else None,
                top=s if rr == r0 else None,
                bottom=s if rr == r0 + 2 else None,
            )
    # 라벨 (r0)
    ws.merge_cells(start_row=r0, start_column=c0,
                   end_row=r0, end_column=c0 + 2)
    ws.cell(r0, c0).value = label
    ws.cell(r0, c0).font = F_CARD_LBL
    ws.cell(r0, c0).alignment = Alignment(horizontal="left",
                                          vertical="center", indent=1)
    # 값 (r0+1, 큰 숫자)
    ws.merge_cells(start_row=r0 + 1, start_column=c0,
                   end_row=r0 + 1, end_column=c0 + 2)
    ws.cell(r0 + 1, c0).value = value
    ws.cell(r0 + 1, c0).font = F_CARD_VAL
    ws.cell(r0 + 1, c0).alignment = ALIGN_C
    ws.cell(r0 + 1, c0).number_format = FMT_NUM
    # 비고 (r0+2)
    ws.merge_cells(start_row=r0 + 2, start_column=c0,
                   end_row=r0 + 2, end_column=c0 + 2)
    ws.cell(r0 + 2, c0).value = note
    ws.cell(r0 + 2, c0).font = F_CARD_NOTE
    ws.cell(r0 + 2, c0).alignment = Alignment(horizontal="left",
                                              vertical="center", indent=1)
    ws.row_dimensions[r0].height = 15
    ws.row_dimensions[r0 + 1].height = 25
    ws.row_dimensions[r0 + 2].height = 13


def _final_card(ws, r0, c0, span, value):
    """최종 정산금액 강조 카드 (큰 노랑)."""
    fill = PatternFill("solid", fgColor=C_CARD_F)
    border_side = Side(style="medium", color="BF9000")
    border = Border(left=border_side, right=border_side,
                    top=border_side, bottom=border_side)
    for rr in range(r0, r0 + 4):
        for cc in range(c0, c0 + span):
            ws.cell(rr, cc).fill = fill
            ws.cell(rr, cc).border = border
    # 라벨
    ws.merge_cells(start_row=r0, start_column=c0,
                   end_row=r0, end_column=c0 + span - 1)
    ws.cell(r0, c0).value = "★ 최종 정산금액"
    ws.cell(r0, c0).font = F_FINAL_LBL
    ws.cell(r0, c0).alignment = ALIGN_C
    # 값 (큰 숫자, 2행 병합)
    ws.merge_cells(start_row=r0 + 1, start_column=c0,
                   end_row=r0 + 2, end_column=c0 + span - 1)
    ws.cell(r0 + 1, c0).value = value
    ws.cell(r0 + 1, c0).font = F_FINAL_VAL
    ws.cell(r0 + 1, c0).alignment = ALIGN_C
    ws.cell(r0 + 1, c0).number_format = FMT_NUM
    # 산식
    ws.merge_cells(start_row=r0 + 3, start_column=c0,
                   end_row=r0 + 3, end_column=c0 + span - 1)
    ws.cell(r0 + 3, c0).value = "A. 기본조립비 + B. 지원순액 + C. 비용보전 − D. 차감비용 + E. BI 차이 청구"
    ws.cell(r0 + 3, c0).font = F_CARD_NOTE
    ws.cell(r0 + 3, c0).alignment = ALIGN_C
    ws.row_dimensions[r0].height = 22
    ws.row_dimensions[r0 + 1].height = 32
    ws.row_dimensions[r0 + 2].height = 14
    ws.row_dimensions[r0 + 3].height = 18


def _signal_row(ws, r, item, result, note):
    """검증 신호등 행: 라벨 / 결과(큰 셀+이모지) / 수치."""
    icon_map = {"PASS": "✅", "INFO": "ℹ️", "WARN": "⚠️", "FAIL": "❌"}
    fill_map = {"PASS": FILL_OK, "INFO": FILL_WARN,
                "WARN": FILL_WARN, "FAIL": FILL_FAIL}
    result = (result or "").upper()
    icon = icon_map.get(result, "·")
    fill = fill_map.get(result, FILL_SUB)

    # 셀 A: 항목 라벨
    ws.cell(r, 1).value = item
    ws.cell(r, 1).font = Font(name="맑은 고딕", size=11, bold=True,
                              color="1F3864")
    ws.cell(r, 1).alignment = Alignment(horizontal="left",
                                        vertical="center", indent=1)
    ws.cell(r, 1).fill = FILL_SUB
    ws.cell(r, 1).border = BORDER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    # 셀 C: 결과 아이콘 + 라벨
    ws.cell(r, 3).value = f"{icon} {result}"
    ws.cell(r, 3).font = F_SIGNAL
    ws.cell(r, 3).alignment = ALIGN_C
    ws.cell(r, 3).fill = fill
    ws.cell(r, 3).border = BORDER

    # 셀 D~끝(c19): 수치·비고 — wrap_text + 충분한 폭
    ws.cell(r, 4).value = note
    ws.cell(r, 4).font = F_DATA
    ws.cell(r, 4).alignment = Alignment(horizontal="left",
                                        vertical="center",
                                        indent=1, wrap_text=True)
    ws.cell(r, 4).border = BORDER
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=19)
    ws.row_dimensions[r].height = 26


def build_sheet_summary(wb, mm, lines_data, support, etc_data,
                       night_info, pnl, err_types):
    """90_월정산_요약 — 대시보드 (KPI 카드 + 차트 + 신호등 + 보고)."""
    name = "90_월정산_요약"
    # 기존 시트 위치 보존 — 사용자가 순서 바꿔도 그 자리 유지
    if name in wb.sheetnames:
        existing_idx = wb.index(wb[name])
        del wb[name]
        ws = wb.create_sheet(name, existing_idx)
    else:
        ws = wb.create_sheet(name, 0)  # 신규 생성 시 맨 앞
    ws.sheet_view.showGridLines = False  # 눈금선 숨김 (대시보드 깔끔)

    # 컬럼 폭 — 카드 3열씩 5장 + gap 4 = 19열. A~S (전체 폭 축소)
    widths = {1: 10, 2: 10, 3: 10,        # 카드1 (A)
              4: 2,                        # gap
              5: 10, 6: 10, 7: 10,        # 카드2 (B)
              8: 2,                        # gap
              9: 10, 10: 10, 11: 10,      # 카드3 (C)
              12: 2,                       # gap
              13: 10, 14: 10, 15: 10,     # 카드4 (D)
              16: 2,                       # gap
              17: 10, 18: 10, 19: 10}     # 카드5 (E)
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ───── 타이틀 ─────
    ws.cell(1, 1).value = f"{mm}월 대원테크 조립비 월정산 대시보드"
    ws.cell(1, 1).font = F_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    ws.row_dimensions[1].height = 24
    ws.cell(2, 1).value = f"생성: {datetime.now():%Y-%m-%d %H:%M}  (수식: 본체 [정산집계] 참조)"
    ws.cell(2, 1).font = F_NOTE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    ws.row_dimensions[2].height = 13

    # ───── 라인별 행 매핑 ─────
    LINE_ROWS = {code: 2 + i for i, code in enumerate(LINES)}

    # ───── 최종 정산금액 박스 (r4~r6, 가로 전체, 3행) ─────
    final_fill = PatternFill("solid", fgColor=C_CARD_F)
    final_border_side = Side(style="medium", color="BF9000")
    final_border = Border(left=final_border_side, right=final_border_side,
                          top=final_border_side, bottom=final_border_side)
    for rr in range(4, 7):
        for cc in range(1, 20):
            ws.cell(rr, cc).fill = final_fill
            ws.cell(rr, cc).border = final_border

    # 라벨 (r4)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=19)
    ws.cell(4, 1).value = f"★ {mm}월 최종 정산금액"
    ws.cell(4, 1).font = F_FINAL_LBL
    ws.cell(4, 1).alignment = ALIGN_C
    ws.row_dimensions[4].height = 18

    # 최종금액 (r5, 큰 숫자) — KPI 카드 값 셀 참조
    # KPI 값 셀: A=A9, B=E9, C=I9, D=M9, E=Q9 (r0=8 기준 r0+1=9)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=19)
    ws.cell(5, 1).value = "=A9+E9+I9-M9+Q9"
    ws.cell(5, 1).font = F_FINAL_VAL
    ws.cell(5, 1).alignment = ALIGN_C
    ws.cell(5, 1).number_format = FMT_NUM
    ws.row_dimensions[5].height = 30

    # 산식 (r6)
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=19)
    ws.cell(6, 1).value = "산식: A 기본조립비 + B 지원순액 + C 비용보전 − D 차감비용 + E BI 차이 청구"
    ws.cell(6, 1).font = F_CARD_NOTE
    ws.cell(6, 1).alignment = ALIGN_C
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[7].height = 6  # gap

    # ───── KPI 카드 5장 (r8~r10) — 수식 참조 ─────
    _kpi_card(ws, 8, 1, "A. 기본조립비 (GERP)",
              "=정산집계!G12", "본체 [정산집계] G12", C_CARD_A)
    _kpi_card(ws, 8, 5, "B. 지원순액",
              pnl["B"], "준지원금 − 받은지원금", C_CARD_B)
    _kpi_card(ws, 8, 9, "C. 비용보전",
              pnl["C"], "별도 비용보전", C_CARD_C)
    _kpi_card(ws, 8, 13, "D. 차감비용",
              pnl["D"], "G-ERP 라인정지 + QIS", C_CARD_D)
    _kpi_card(ws, 8, 17, "E. BI 차이 청구",
              pnl["E"], "(BI−GERP) × 170원", C_CARD_F)

    # ───── 라인별 손익 박스 10장 (2행 × 5열, 수식 참조) ─────
    # KPI 카드 r8~r10 (3행). 라인 박스 r12~r14 (1행) / r16~r18 (2행)
    section(ws, 11, "■ 라인별 손익 (정산집계 셀 참조)")
    line_box_start = 12
    col_groups = [1, 5, 9, 13, 17]
    for idx, code in enumerate(LINES):
        row_offset = idx // 5  # 0 or 1
        col_idx = idx % 5
        r0 = line_box_start + row_offset * 4  # 3행 박스 + 1행 gap
        c0 = col_groups[col_idx]
        d = lines_data["lines"].get(code, {})
        b = support["line_total"].get(code, 0)
        gjg_row = LINE_ROWS[code]
        if code == "SP3M3" and pnl["E"]:
            final_formula = f"=정산집계!G{gjg_row}+({b})+({pnl['E']})"
        else:
            final_formula = f"=정산집계!G{gjg_row}+({b})"
        fill = LINE_COLORS.get(code, "F2F2F2")
        _line_card(ws, r0, c0, code, d.get("name", ""),
                   final_formula, f"=정산집계!G{gjg_row}", b, fill)

    # ───── 검증 신호등 (라인 박스 아래) ─────
    r = line_box_start + 8  # 2행 박스(r16~r18) 후 여유
    section(ws, r, "■ 검증·이슈 신호등")
    r += 1
    # 기준 안내 (한 줄)
    ws.cell(r, 1).value = (
        "※ PASS = 정상 / INFO = 확인·청구 필요 / WARN = 큰 차이 / FAIL = 데이터 누락")
    ws.cell(r, 1).font = F_NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19)
    r += 2

    # 야간 정합 라벨
    night_detail = []
    for code in night_info["night_lines"]:
        d = lines_data["lines"][code]
        bi = night_info["bi_summary"].get(code, {})
        gn = d["night_qty"]
        bn = bi.get("qty", 0)
        if gn > 0 and bn > 0:
            pct = (gn - bn) / gn * 100
            night_detail.append((code, gn, bn, pct))
    # 정합 기준: 모든 야간 라인 ±5% 이내 + E (BI 누락 청구) = 0 → PASS
    # E > 0 (청구 필요) 또는 ±5% 초과 → INFO
    # 10% 초과 + 청구 불가 → FAIL
    NIGHT_UNIT = 170
    night_label = "PASS"
    night_text_parts = []
    has_claim = False
    has_big_diff = False
    for code, gn, bn, pct in night_detail:
        diff_ea = bn - gn  # 양수면 BI > GERP (회사 누락분)
        if code == "SP3M3" and diff_ea > 0:
            claim_amt = diff_ea * NIGHT_UNIT
            night_text_parts.append(
                f"{code} 누락 {diff_ea:+,} EA → +{claim_amt:,}원 (E 가산)")
            has_claim = True
        elif abs(pct) <= 5:
            night_text_parts.append(f"{code} 정합 ({pct:+.1f}%)")
        else:
            night_text_parts.append(
                f"{code} 차이 ({pct:+.1f}%)")
            has_big_diff = True
    night_text = " / ".join(night_text_parts) or "야간 라인 없음"
    if has_claim or has_big_diff:
        night_label = "INFO"

    if err_types["exists"]:
        diff_text = f"차이 {err_types['total_cnt']}건 / {err_types['total_diff']:+,}원 — 91 ④번 (ERP 정정용)"
        if err_types["total_cnt"] == 0:
            erp_label = "PASS"
        elif abs(err_types["total_diff"]) < 100000:
            erp_label = "INFO"
        else:
            erp_label = "WARN"
    else:
        diff_text = "본체 [유형별요약] 시트 미존재"
        erp_label = "INFO"

    items = [
        ("GERP 전체금액", "PASS", f"A = {pnl['A']:,}원"),
        ("지원 자료",
         "PASS" if support["exists"] else "INFO",
         f"B = {pnl['B']:+,}원" + (f" ({len(support['details'])}건)" if support["exists"] else " — 당월 없음")),
        ("차감비용 자료",
         "PASS" if etc_data["exists"] else "FAIL",
          f"D = {pnl['D']:,}원 ({sum(x['count'] for x in etc_data['rows'])}건)" if etc_data["exists"] else "raw 미존재"),
        ("야간 BI 정합", night_label, night_text),
        ("GERP vs 구ERP", erp_label, diff_text),
        ("최종 손익", "PASS", f"{pnl['final']:,}원"),
    ]
    for item, result, n in items:
        _signal_row(ws, r, item, result, n)
        r += 1
    r += 2

    # ───── 보고 문장 ─────
    section(ws, r, "보고 문장 (복붙용)")
    r += 1
    night_str = ", ".join(night_info["night_lines"]) or "없음"
    template = [
        f"○ 2026년 {mm}월 대원테크 조립비 정산 결과 공유드립니다.",
        f"1. GERP 기준 기본 조립비는 {pnl['A']:,.0f}원입니다.",
        f"2. 야간 가동라인 {len(night_info['night_lines'])}개({night_str}) BI 비교 결과 차이 청구금액은 {pnl['E']:+,.0f}원입니다 (170원 × 누락 EA, 91 ①섹션).",
        f"3. GERP-구ERP 차이는 {err_types.get('total_cnt', 0)}건, {err_types.get('total_diff', 0):+,.0f}원 (91 ④섹션, ERP 정정용).",
        f"4. 타라인 지원 정산금액은 {pnl['B']:+,.0f}원, 라인정지·기타생산비 차감액은 {pnl['D']:,.0f}원입니다.",
        f"5. 최종 정산 반영금액은 {pnl['final']:,.0f}원입니다 (A+B+C−D+E).",
    ]
    for line in template:
        ws.cell(r, 1).value = line
        ws.cell(r, 1).font = F_DATA
        ws.cell(r, 1).alignment = ALIGN_L
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        r += 1

    # 차트 데이터 영역은 차트가 덮어서 시각적으로 가려짐 (hidden 제거 — 차트 렌더링 위해)


def build_sheet_detail(wb, mm, lines_data, support, etc_data,
                      night_info, err_types, pnl):
    """91_정산상세 — 4섹션 ①②③④."""
    name = "91_정산상세"
    # 기존 시트 위치 보존
    if name in wb.sheetnames:
        existing_idx = wb.index(wb[name])
        del wb[name]
        ws = wb.create_sheet(name, existing_idx)
    else:
        ws = wb.create_sheet(name)  # 신규 시 끝에
    ws.sheet_view.showGridLines = False  # 눈금선 숨김
    apply_col_widths(ws)

    title(ws, 1, f"{mm}월 정산 상세 (raw 근거)")

    # ───── ① 야간 가동라인 BI 정합 ─────
    r = 3
    section(ws, r, "① 야간 가동라인 BI 정합 (GERP 추가행 vs BI 원본)")
    r += 1
    table_header(ws, r, ["라인", "라인명", "GERP 야간", "BI 야간",
                         "BI 누락(EA)", "단가", "청구금액 E"])
    r += 1
    if not night_info["night_lines"]:
        ws.cell(r, 1).value = "당월 야간 가동라인 없음"
        ws.cell(r, 1).font = F_NOTE
        r += 1
    else:
        NIGHT_UNIT_PRICE = 170  # SP3M3 BI 차이 청구 단가 (도메인 룰)
        for code in night_info["night_lines"]:
            d = lines_data["lines"][code]
            bi = night_info["bi_summary"][code]
            gn = d["night_qty"]
            bn = bi["qty"]
            diff_ea = bn - gn  # 양수면 BI > GERP (누락분)
            data_row(ws, r,
                     [code, d["name"], gn, bn],
                     ["C", "C", "N", "N"])
            # 누락(EA) = BI - GERP. 양수만 청구 대상
            ws.cell(r, 5).value = f"=D{r}-C{r}"
            ws.cell(r, 5).font = F_DATA
            ws.cell(r, 5).number_format = FMT_NUM
            ws.cell(r, 5).alignment = ALIGN_R
            ws.cell(r, 5).border = BORDER
            # 단가
            unit_price = NIGHT_UNIT_PRICE if code == "SP3M3" else 0
            ws.cell(r, 6).value = unit_price if unit_price > 0 else None
            ws.cell(r, 6).font = F_DATA
            ws.cell(r, 6).number_format = FMT_NUM
            ws.cell(r, 6).alignment = ALIGN_R
            ws.cell(r, 6).border = BORDER
            # 청구금액 = MAX(0, 누락EA) × 단가
            if code == "SP3M3" and diff_ea > 0:
                claim_amt = diff_ea * NIGHT_UNIT_PRICE
            else:
                claim_amt = 0
            ws.cell(r, 7).value = claim_amt
            ws.cell(r, 7).font = F_TOTAL  # 청구금액 강조
            ws.cell(r, 7).number_format = FMT_NUM
            ws.cell(r, 7).alignment = ALIGN_R
            ws.cell(r, 7).border = BORDER
            if claim_amt > 0:
                ws.cell(r, 7).fill = FILL_POS
            # 누락 EA 양수면 강조
            if diff_ea > 0:
                ws.cell(r, 5).fill = FILL_POS
            r += 1
        # 합계 — E 항목 (소명 청구 합계)
        e_total = sum(
            max(0, night_info["bi_summary"].get(code, {}).get("qty", 0) - lines_data["lines"][code]["night_qty"]) * NIGHT_UNIT_PRICE
            for code in night_info["night_lines"] if code == "SP3M3"
        )
        total_row(ws, r,
                  ["E 항목 합계", "", "", "", "", "", e_total],
                  ["L", "C", "N", "N", "N", "N", "N"])
        r += 1
        note(ws, r, f"E 산식: BI 누락 EA × 170원 (SP3M3 한정). A는 GERP raw 지정단가 그대로 (35.84M 권위값). E는 GERP가 RSP 모듈품번만 잡고 누락한 일반 품번 야간 생산분을 소명·청구하기 위한 금액 — 손익에 +E 가산.")
        r += 1
    r += 2

    # ───── ② 차감비용 ─────
    section(ws, r, "② 차감비용 (손익 산식 D 차감비용)")
    r += 1
    if not etc_data["exists"]:
        ws.cell(r, 1).value = "라인정지_MM월_raw.xlsx 미존재"
        ws.cell(r, 1).font = F_NOTE
        r += 1
    else:
        table_header(ws, r, ["시스템", "청구유형", "건수", "금액(원)",
                             "비고", "", ""])
        r += 1
        for row_data in etc_data["rows"]:
            data_row(ws, r,
                     [row_data["system"], row_data["claim_type"],
                      row_data["count"], row_data["amount"], "", "", ""],
                     ["L", "C", "N", "N", "L", "L", "L"])
            r += 1
        total_row(ws, r,
                  ["합계", "",
                   sum(x["count"] for x in etc_data["rows"]),
                   etc_data["total"], "", "", ""],
                  ["L", "C", "N", "N", "L", "L", "L"])
        r += 2
        # 귀책 서브표
        if etc_data["blame"]:
            ws.cell(r, 1).value = "└ QIS 라인교체 — 귀책 부서별"
            ws.cell(r, 1).font = F_TOTAL
            r += 1
            table_header(ws, r, ["귀책 부서", "건수", "금액(원)", "", "", "", ""])
            r += 1
            for b in etc_data["blame"]:
                data_row(ws, r,
                         [b["blame"], b["count"], b["amount"], "", "", "", ""],
                         ["L", "N", "N", "L", "L", "L", "L"])
                r += 1
    r += 2

    # ───── ③ 라인지원 ─────
    section(ws, r, "③ 라인지원 (손익 산식 B 지원순액)")
    r += 1
    if not support["exists"]:
        ws.cell(r, 1).value = "당월 지원 없음"
        ws.cell(r, 1).font = F_NOTE
        r += 1
    else:
        table_header(ws, r, ["파일", "방향", "라인", "금액(원)",
                             "분류 (회사 입장)", "", ""])
        r += 1
        for sr in support["details"]:
            # 부호 정정: 회사 입장 — 화인텍 = 지출(우리가 줘야), 리노텍 = 수입(우리가 받아야)
            if sr.sign > 0:
                arrow = "← 받음"
                cls = "수입 (우리가 받을 돈)"
            else:
                arrow = "→ 지급"
                cls = "지출 (우리가 줘야 할 돈)"
            data_row(ws, r,
                     [sr.file, arrow, sr.line, sr.amount * sr.sign, cls, "", ""],
                     ["L", "C", "C", "Np", "L", "L", "L"])
            r += 1
        total_row(ws, r,
                  ["지원 순액 (B = 수입 − 지출)", "", "", support["B_sum"], "", "", ""],
                  ["L", "C", "C", "N", "L", "L", "L"])
        r += 1
        # 수입/지출 분해
        ws.cell(r, 3).value = "└ 수입(+)"
        ws.cell(r, 3).font = F_DATA
        ws.cell(r, 3).alignment = ALIGN_R
        ws.cell(r, 4).value = support["pos_total"]
        ws.cell(r, 4).font = F_DATA
        ws.cell(r, 4).number_format = FMT_NUM
        ws.cell(r, 4).alignment = ALIGN_R
        ws.cell(r, 5).value = "리노텍 등 우리가 받아야 할 금액"
        ws.cell(r, 5).font = F_NOTE
        r += 1
        ws.cell(r, 3).value = "└ 지출(−)"
        ws.cell(r, 3).font = F_DATA
        ws.cell(r, 3).alignment = ALIGN_R
        ws.cell(r, 4).value = -support["neg_total"]
        ws.cell(r, 4).font = F_DATA
        ws.cell(r, 4).number_format = FMT_NUM
        ws.cell(r, 4).alignment = ALIGN_R
        ws.cell(r, 5).value = "화인텍 등 우리가 줘야 할 금액"
        ws.cell(r, 5).font = F_NOTE
        r += 1
    r += 2

    # ───── ④ 차이 통합 — 소명·청구 작성용 (GERP-구ERP + 야간 BI) ─────
    section(ws, r, "④ 차이 통합 — 소명·청구 작성용 (받을 금액 합산)")
    r += 1
    if not err_types["exists"]:
        ws.cell(r, 1).value = "본체 [유형별요약] 시트 미존재"
        ws.cell(r, 1).font = F_NOTE
        r += 1
    else:
        # 받을 금액 산정 룰:
        # 차이금액 = GERP − 구ERP
        # 음수 → GERP 부족 → 청구 후 받을 금액 (절댓값)
        # 양수 → GERP 과도청구 → 구ERP 정정 대상 (받을 금액 아님)
        # 받을 금액 산정 룰 (project_mo_part_rules.md):
        # - 음수 차이 (GERP < 구ERP) → GERP 단가 미등록 → GERP 정정 후 청구
        # - 양수 차이 (GERP > 구ERP) → 구ERP 실적 누락 확정 → 구ERP 정정 후 매출 인정
        # - 양쪽 다 회사가 받아야 할 금액 → |차이| 절댓값
        # - 정상제외(이관)만 제외
        ACTION_GUIDE = {
            "구실적누락": "GERP 청구분이 구ERP 입고 누락 → 구ERP 정정 + 매출 인정 (받을 금액)",
            "GERP누락": "구ERP 입고분이 GERP 단가 미등록 → GERP 등록 + 청구 (받을 금액)",
            "수량차이": "양쪽 수량 불일치 → 정확한 쪽 확인 후 정정 (받을 금액)",
            "정산차이": "단가 차이 또는 기준단가 누락 → 기준정보 정정 (받을 금액)",
            "정상제외(이관)": "이관 처리 품번 → 정산 제외 (받을 금액 X)",
        }
        table_header(ws, r, ["구분", "건수", "차이금액", "받을 금액",
                             "처리 액션", "", ""])
        r += 1
        claim_total = 0
        for er in err_types["rows"]:
            kind = er["kind"]
            cnt = er.get("cnt", 0)
            diff = er.get("diff", 0)
            # 받을 금액 = |차이| 절댓값 (정상제외 제외)
            if kind == "정상제외(이관)":
                claim_amt = 0
            else:
                claim_amt = abs(diff)
            claim_total += claim_amt
            action = ACTION_GUIDE.get(kind, "")
            data_row(ws, r,
                     [kind, cnt, diff, claim_amt, action, "", ""],
                     ["L", "N", "Np", "N", "L", "L", "L"])
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 22
            r += 1
        # 야간 BI 차이 — E 항목으로 손익 가산 (재청구 아님, 별도 표시만)
        e_total = pnl.get("E", 0)
        e_cnt = len(pnl.get("e_detail", []))
        if e_total > 0:
            data_row(ws, r,
                     ["야간 BI 차이", e_cnt, 0, e_total,
                      "BI 야간 누락 × 170원 — E 항목으로 손익 가산 완료 (소명만 필요)",
                      "", ""],
                     ["L", "N", "Np", "N", "L", "L", "L"])
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 22
            r += 1
        total_row(ws, r,
                  ["합계", err_types["total_cnt"] + (e_cnt if e_total else 0),
                   err_types["total_diff"], claim_total + e_total,
                   "받을 금액 합계 = |차이금액| 절댓값 + 야간 BI (정상제외 제외)",
                   "", ""],
                  ["L", "N", "N", "N", "L", "L", "L"])
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        r += 1
        note(ws, r,
             "받을 금액 = 음수 차이(GERP 부족) 절댓값 + 야간 BI 누락. "
             "양수 차이는 GERP 과도청구로 구ERP 정정 대상이라 받을 금액 아님. "
             "야간 BI는 E 항목으로 손익에 이미 가산되어 있음 (이중 가산 X). "
             "상세 품번은 본체 [오류리스트] 시트.")

    ws.row_dimensions[1].height = 22


def force_excel_recalc(book_path):
    """PowerShell + Excel COM 자동 재계산·저장 → cached value 박힘."""
    import subprocess
    ps_cmd = f"""
$ErrorActionPreference = 'Stop'
Get-Process -Name EXCEL -ErrorAction SilentlyContinue | Stop-Process -Force -ErrorAction SilentlyContinue
Start-Sleep -Seconds 1
$Excel = New-Object -ComObject Excel.Application
$Excel.Visible = $false
$Excel.DisplayAlerts = $false
try {{
    $wb = $Excel.Workbooks.Open('{book_path}', 0, $false)
    Start-Sleep -Seconds 2
    $wb.Application.CalculateFullRebuild()
    Start-Sleep -Seconds 2
    $wb.Save()
    $wb.Close($true)
}} finally {{
    $Excel.Quit()
}}
"""
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=120
        )
        return result.returncode == 0
    except Exception as e:
        print(f"  [WARN] force_recalc 실패: {e}")
        return False


def verify_workbook_formulas(book_path):
    """OOXML <f> 수식 raw 검증."""
    import zipfile
    import re
    issues = []
    with zipfile.ZipFile(book_path) as zf:
        for fname in zf.namelist():
            if not (fname.startswith("xl/worksheets/sheet") and fname.endswith(".xml")):
                continue
            content = zf.read(fname).decode("utf-8")
            for m in re.finditer(r"<f[^>]*>([^<]+)</f>", content):
                f = m.group(1)
                if not f.strip():
                    issues.append((fname, "빈 수식", f))
                elif f.startswith(" "):
                    issues.append((fname, "공백 시작", f[:60]))
                elif f.count("'") % 2 != 0:
                    issues.append((fname, "홑따옴표 짝", f[:60]))
                elif f.count('"') % 2 != 0:
                    issues.append((fname, "쌍따옴표 짝", f[:60]))
                elif f.endswith(("(", ",", ".")):
                    issues.append((fname, "잘림", f[:60]))
    return issues


# ============================================================
# main
# ============================================================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True)
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    mm = args.month.zfill(2)
    work = SETTLE / f"{int(mm) + 1:02d}월"
    book = work / f"정산_수식버전_{mm}월.xlsx"
    stoppage = work / f"라인정지_{mm}월_raw.xlsx"
    support_dir = resolve_support_dir(work, mm)

    if not book.exists():
        print(f"[FAIL] 본체 미존재: {book}")
        sys.exit(1)

    if not args.no_backup:
        bak = book.with_suffix(f".{datetime.now():%Y%m%d_%H%M%S}.bak.xlsx")
        shutil.copy2(book, bak)
        print(f"[BACKUP] {bak.name}")

    print("[LOAD] GERP 원본 직접 합산 …")
    lines_data = load_settlement_totals(book, work)
    print(f"  라인 {len(lines_data['lines'])}개, GERP 합계 {lines_data['grand_total']:,}원")

    print("[LOAD] 본체 유형별요약 …")
    err_types = load_error_types(book)
    if err_types["exists"]:
        print(f"  차이 {err_types['total_cnt']}건 / {err_types['total_diff']:+,}원")
    else:
        print("  [WARN] 유형별요약 시트 미존재")

    print("[LOAD] 라인정지 통합집계 …")
    etc_data = load_line_stoppage(stoppage)
    if etc_data["exists"]:
        print(f"  {len(etc_data['rows'])}건, 합계 {etc_data['total']:,}원")
    else:
        print(f"  [WARN] {stoppage.name} 미존재")

    print(f"[LOAD] 지원자료 ({support_dir.name}) …")
    support = load_support_data(support_dir, mm)
    if support["exists"]:
        print(f"  지원 {len(support['details'])}행, B = {support['B_sum']:+,}원")
    else:
        print("  당월 지원 없음")

    print("[LOAD] BI 야간 추출 …")
    night_info = compute_night_info(lines_data, args.year, int(mm))
    print(f"  야간 라인: {night_info['night_lines']}")

    pnl = compute_pnl(lines_data, support, etc_data, night_info)
    print(f"\n[PNL] A={pnl['A']:,}  B={pnl['B']:+,}  C={pnl['C']:,}  D={pnl['D']:,}  E={pnl['E']:+,}")
    for d in pnl.get("e_detail", []):
        print(f"      E 상세 — {d['line']}: 누락 {d['diff_ea']:,} EA × {d['unit_price']}원 = {d['amount']:,}원")
    print(f"      최종 = {pnl['final']:,}원")

    print(f"\n[OPEN] {book.name}")
    wb = openpyxl.load_workbook(book, data_only=False)

    # 옛 90~95 6시트 잔존 시 삭제 (이 스킬 이전 버전 호환)
    # 90_월정산_요약 / 91_정산상세는 빌더 안에서 위치 보존하며 갱신하므로 여기서 안 지움
    for old in ["91_손익통합", "92_야간정합", "93_기타비용",
                "94_라인지원", "95_검증_이슈요약"]:
        if old in wb.sheetnames:
            del wb[old]

    print("[BUILD] 91_정산상세 …")
    build_sheet_detail(wb, mm, lines_data, support, etc_data,
                       night_info, err_types, pnl)
    print("[BUILD] 90_월정산_요약 …")
    build_sheet_summary(wb, mm, lines_data, support, etc_data,
                       night_info, pnl, err_types)

    # 본체 유형별요약 시트 hidden 처리 — 91 ④섹션에 흡수, 데이터는 보존
    # (재실행 시 데이터 재사용 위해 삭제 X)
    if "유형별요약" in wb.sheetnames:
        wb["유형별요약"].sheet_state = "hidden"
        print("[HIDDEN] 본체 유형별요약 시트 숨김 (데이터 보존, 91 ④섹션에 흡수)")

    print(f"[SAVE] {book}")
    wb.save(book)
    print(f"\n[BUILD OK] 본체 시트 수: {len(wb.sheetnames)}")

    print("\n[VERIFY] OOXML 수식 무결성 …")
    issues = verify_workbook_formulas(book)
    if issues:
        print(f"  ❌ 깨진 수식 {len(issues)}건:")
        for fname, kind, sample in issues[:20]:
            print(f"    [{fname}] {kind}: {sample!r}")
        sys.exit(2)
    print("  ✅ 모든 수식 정상")

    # cached value 박기 (다음 실행 시 정합값 사용)
    print("\n[RECALC] Excel COM 자동 재계산 …")
    if force_excel_recalc(str(book)):
        print("  ✅ cached value 박힘 (재실행 시 본체 정산집계 기준 사용)")
    else:
        print("  ⚠️ 자동 재계산 실패 — 엑셀에서 한 번 열고 저장하면 갱신됨")


if __name__ == "__main__":
    main()
