"""night-scan-compare — BI 야간 산출량 vs MES 스캔실적 일자별 비교.

사용:
  python run.py --line SP3M3 --month 4
  python run.py --line SP3M3 --month 4 --year 2026
  python run.py --line SP3M3 --month 4 --target "<커스텀 경로>"

기본 출력 경로:
  05_생산실적/조립비정산/{MM}월/야간_생산_{라인}.xlsx

세션143 v1.2 — 임시 스크립트(/c/temp) 3종 통합:
  Phase1 BI 추출 + Phase2 MES API + Phase3 엑셀 데이터 입력
  Phase4 테이블 ref + 셀 서식 + 조건부 서식
  Phase5 수식 복원 (KPI/SUMIFS/IF/판정/차트 셀참조)
"""
import sys, json, time, shutil, argparse
from copy import copy
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# d0-production-plan 함수 재사용 (OAuth 자동 로그인 + Chrome 9223 기동)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "d0-production-plan"))
from run import ensure_chrome_cdp, ensure_erp_login, CDP_URL  # noqa

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

# ============================================================
# 설정
# ============================================================
SHIFT_NIGHT = "02"
BI_PATH = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\BI실적\대원테크_라인별 생산실적_BI.xlsx")
DEFAULT_TARGET_DIR = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\조립비정산")
MES_LAYOUT = "http://mes-dev.samsong.com:19200/layout/layout.do"
MES_API = "http://mes-dev.samsong.com:19200/prdtstatus/selectPrdtRsltLine.do"

# 조건부 서식 색상 (매뉴얼 명세)
FILL_DIFF   = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
FILL_MATCH  = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
FILL_NOSCAN = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")


# ============================================================
# Phase 1: BI 추출
# ============================================================
def extract_bi(line: str, year: int, month: int) -> list[dict]:
    print(f"[phase1] BI 추출 — {line} {year}-{month:02d} 야간")
    if not BI_PATH.exists():
        raise FileNotFoundError(f"BI 원본 없음: {BI_PATH}")
    wb = openpyxl.load_workbook(BI_PATH, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    # 야간 식별: '주간'이 아닌 shift 값
    night_val = None
    for r in rows[1:]:
        if r[4] == line and r[5] not in (None, "주간"):
            night_val = r[5]
            break
    if night_val is None:
        raise RuntimeError(f"BI에서 {line} 야간 데이터 식별 실패")

    out = []
    for r in rows[1:]:
        if r[4] != line or r[5] != night_val:
            continue
        d = r[7]
        if not isinstance(d, datetime) or d.year != year or d.month != month:
            continue
        out.append({
            "date": d, "line": r[4], "shift": r[5], "people": r[6],
            "uph_std": r[8], "ct": r[9], "eff": r[10],
            "work_h": r[11], "down_h": r[12], "real_h": r[13],
            "qty": r[14],
        })
    out.sort(key=lambda x: x["date"])
    print(f"  추출: {len(out)}일 / 산출량 합계 {sum(x['qty'] or 0 for x in out)}")
    return out


# ============================================================
# Phase 2: MES 스캔실적 조회
# ============================================================
def fetch_mes_scans(page, line: str, dates: list[datetime]) -> tuple[dict, list]:
    print("[phase2] MES API 호출")
    scans, misses = {}, []
    for d in dates:
        ymd = d.strftime("%Y%m%d")
        url = f"{MES_API}?prdtDa={ymd}"
        try:
            res = page.evaluate("""
                async (url) => {
                    const r = await fetch(url, {
                        method: 'GET',
                        credentials: 'include',
                        headers: {'X-Requested-With': 'XMLHttpRequest', 'Accept': 'application/json, text/plain, */*'}
                    });
                    return {status: r.status, text: (await r.text()).slice(0, 30000)};
                }
            """, url)
            if res["status"] != 200:
                print(f"  {ymd}: HTTP {res['status']}")
                misses.append(d); continue
            data = json.loads(res["text"])
            lst = data.get("data", {}).get("list") or data.get("list") or []
            qty = None
            for item in lst:
                if item.get("lineCd") == line and item.get("shiftsCd") == SHIFT_NIGHT:
                    raw = item.get("prdtQty")
                    try:
                        qty = int(raw) if raw not in (None, "") else None
                    except Exception:
                        qty = None
                    break
            scans[ymd] = qty
            print(f"  {ymd}: scan = {qty}")
        except Exception as e:
            print(f"  {ymd}: fetch 실패 {e}")
            misses.append(d)
    return scans, misses


# ============================================================
# Phase 3+4+5: 엑셀 입력 + 양식 + 수식
# ============================================================
def _copy_row_format(ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d.font = copy(s.font)
            d.alignment = copy(s.alignment)
            d.border = copy(s.border)
            d.fill = copy(s.fill)
            d.number_format = s.number_format
            d.protection = copy(s.protection)


def update_excel(target_path: Path, bi_data: list[dict], scans: dict):
    print(f"[phase3-5] 엑셀 갱신 — {target_path}")
    bak = target_path.with_name(
        target_path.stem + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}" + target_path.suffix
    )
    shutil.copy2(target_path, bak)
    print(f"  백업: {bak.name}")

    wb = openpyxl.load_workbook(target_path)
    n_prod, n_scan, n_cmp, n_kpi = wb.sheetnames[:4]
    sh_prod, sh_scan, sh_cmp, sh_kpi = wb[n_prod], wb[n_scan], wb[n_cmp], wb[n_kpi]

    N = len(bi_data)
    PROD_LAST = 4 + N   # 생산기준DATA / 스캔실적DATA: 헤더 4 + N
    CMP_LAST  = 5 + N   # 일자별비교: 헤더 5 + N
    KPI_LAST  = 5 + N   # 월간요약 차트 데이터

    # ---- 생산기준DATA: 데이터 + 서식 + 테이블 ref + number_format
    for row in sh_prod.iter_rows(min_row=5, max_row=sh_prod.max_row):
        for c in row: c.value = None
    for i, item in enumerate(bi_data):
        r = 5 + i
        sh_prod.cell(r, 1, item["line"])
        sh_prod.cell(r, 2, item["shift"])
        sh_prod.cell(r, 3, item["people"])
        sh_prod.cell(r, 4, item["date"])
        sh_prod.cell(r, 5, item["uph_std"])
        sh_prod.cell(r, 6, item["ct"])
        sh_prod.cell(r, 7, item["eff"])
        sh_prod.cell(r, 8, item["work_h"])
        sh_prod.cell(r, 9, item["down_h"])
        sh_prod.cell(r, 10, item["real_h"])
        sh_prod.cell(r, 11, item["qty"])
    sh_prod.tables["tbl_prod"].ref = f"A4:K{PROD_LAST}"
    for r in range(6, PROD_LAST + 1):
        _copy_row_format(sh_prod, 5, r, 11)
    for r in range(5, PROD_LAST + 1):
        sh_prod.cell(r, 4).number_format = "yyyy-mm-dd"
        sh_prod.cell(r, 5).number_format = "#,##0"
        sh_prod.cell(r, 7).number_format = "0%"
        for c in (8, 9, 10):
            sh_prod.cell(r, c).number_format = "0.00"
        sh_prod.cell(r, 11).number_format = "#,##0"

    # ---- 스캔실적DATA: 값 입력 + 서식 + 테이블 ref
    for row in sh_scan.iter_rows(min_row=5, max_row=sh_scan.max_row):
        for c in row: c.value = None
    for i, item in enumerate(bi_data):
        r = 5 + i
        ymd = item["date"].strftime("%Y%m%d")
        scan_qty = scans.get(ymd)
        sh_scan.cell(r, 1, item["date"])
        sh_scan.cell(r, 2, item["line"])
        sh_scan.cell(r, 3, scan_qty)  # None → 빈 셀 (=MES없음)
        sh_scan.cell(r, 4, "MES데이터없음" if scan_qty is None else None)
    sh_scan.tables["tbl_scan"].ref = f"A4:D{PROD_LAST}"
    for r in range(6, PROD_LAST + 1):
        _copy_row_format(sh_scan, 5, r, 4)
    for r in range(5, PROD_LAST + 1):
        sh_scan.cell(r, 1).number_format = "yyyy-mm-dd"
        sh_scan.cell(r, 3).number_format = "#,##0"

    # ---- 일자별비교: 수식만 입력 (값 박기 금지 — feedback_excel_preserve_formulas.md)
    for row in sh_cmp.iter_rows(min_row=6, max_row=sh_cmp.max_row):
        for c in row: c.value = None
    # KPI 행
    sh_cmp.cell(3, 1).value = f"=SUM(C6:C{CMP_LAST})"
    sh_cmp.cell(3, 3).value = f"=SUM(D6:D{CMP_LAST})"
    sh_cmp.cell(3, 5).value = f"=SUM(E6:E{CMP_LAST})"
    sh_cmp.cell(3, 7).value = f'=COUNTIF(G6:G{CMP_LAST},"불일치")'
    for c in (1, 3, 5, 7):
        sh_cmp.cell(3, c).number_format = "#,##0"
    # 데이터 행
    for i, item in enumerate(bi_data):
        r = 6 + i
        sh_cmp.cell(r, 1, item["date"])
        sh_cmp.cell(r, 2, item["line"])
        sh_cmp.cell(r, 3).value = (
            f"=SUMIFS('{n_prod}'!$K$5:$K${PROD_LAST},"
            f"'{n_prod}'!$D$5:$D${PROD_LAST},$A{r},"
            f"'{n_prod}'!$A$5:$A${PROD_LAST},$B{r})"
        )
        sh_cmp.cell(r, 4).value = (
            f"=IF(COUNTIFS('{n_scan}'!$A$5:$A${PROD_LAST},$A{r},"
            f"'{n_scan}'!$B$5:$B${PROD_LAST},$B{r},"
            f"'{n_scan}'!$C$5:$C${PROD_LAST},\"<>\")=0,\"\","
            f"SUMIFS('{n_scan}'!$C$5:$C${PROD_LAST},"
            f"'{n_scan}'!$A$5:$A${PROD_LAST},$A{r},"
            f"'{n_scan}'!$B$5:$B${PROD_LAST},$B{r}))"
        )
        sh_cmp.cell(r, 5).value = f'=IF(D{r}="","",C{r}-D{r})'
        sh_cmp.cell(r, 6).value = f'=IF(D{r}="","",IFERROR(E{r}/C{r},0))'
        sh_cmp.cell(r, 7).value = f'=IF(D{r}="", "스캔미입력", IF(E{r}=0,"일치","불일치"))'
    sh_cmp.tables["tbl_compare"].ref = f"A5:H{CMP_LAST}"
    for r in range(7, CMP_LAST + 1):
        _copy_row_format(sh_cmp, 6, r, 8)
    for r in range(6, CMP_LAST + 1):
        sh_cmp.cell(r, 1).number_format = "yyyy-mm-dd"
        sh_cmp.cell(r, 3).number_format = "#,##0"
        sh_cmp.cell(r, 4).number_format = "#,##0"
        sh_cmp.cell(r, 5).number_format = "#,##0"
        sh_cmp.cell(r, 6).number_format = "0.0%"
    # 조건부 서식
    sh_cmp.conditional_formatting._cf_rules.clear()
    rng = f"A6:H{CMP_LAST}"
    sh_cmp.conditional_formatting.add(rng, FormulaRule(formula=['$G6="스캔미입력"'], fill=FILL_NOSCAN, stopIfTrue=True))
    sh_cmp.conditional_formatting.add(rng, FormulaRule(formula=['$G6="일치"'],     fill=FILL_MATCH))
    sh_cmp.conditional_formatting.add(rng, FormulaRule(formula=['$G6="불일치"'],   fill=FILL_DIFF))

    # ---- 월간요약: KPI 셀 + 차트 데이터 모두 수식 참조
    sh_kpi.cell(5, 2).value = f"='{n_cmp}'!A3"
    sh_kpi.cell(6, 2).value = f"='{n_cmp}'!C3"
    sh_kpi.cell(7, 2).value = f"='{n_cmp}'!E3"
    sh_kpi.cell(8, 2).value = f"='{n_cmp}'!G3"
    for r in range(5, 9):
        sh_kpi.cell(r, 2).number_format = "#,##0"
    for row in sh_kpi.iter_rows(min_row=6, max_row=sh_kpi.max_row, min_col=6, max_col=9):
        for c in row: c.value = None
    for i, item in enumerate(bi_data):
        r = 6 + i
        sh_kpi.cell(r, 6, item["date"])
        sh_kpi.cell(r, 7).value = f"='{n_cmp}'!C{r}"
        sh_kpi.cell(r, 8).value = f"='{n_cmp}'!D{r}"
        sh_kpi.cell(r, 9).value = f'=IF(H{r}="","",G{r}-H{r})'
    for r in range(7, KPI_LAST + 1):
        _copy_row_format(sh_kpi, 6, r, 9)
    for r in range(6, KPI_LAST + 1):
        sh_kpi.cell(r, 6).number_format = "yyyy-mm-dd"
        for c in (7, 8, 9):
            sh_kpi.cell(r, c).number_format = "#,##0"

    wb.save(target_path)
    print(f"  저장 완료")


# ============================================================
# main
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--line", required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--year", type=int, default=datetime.now().year)
    ap.add_argument("--target", default=None,
                    help="커스텀 출력 경로. 기본=05_생산실적/조립비정산/{MM}월/야간_생산_{LINE}.xlsx")
    args = ap.parse_args()

    if args.target:
        target_path = Path(args.target)
    else:
        target_path = DEFAULT_TARGET_DIR / f"{args.month:02d}월" / f"야간_생산_{args.line}.xlsx"
    if not target_path.exists():
        print(f"[ERR] 비교 엑셀 없음: {target_path}"); return 2

    bi_data = extract_bi(args.line, args.year, args.month)
    if not bi_data:
        print("[ERR] BI 해당 월 야간 데이터 0건"); return 2
    dates = [x["date"] for x in bi_data]

    print("[phase0] CDP 9223 attach + OAuth")
    ensure_chrome_cdp()
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(CDP_URL)
        ctx = browser.contexts[0]
        page = next((pg for pg in ctx.pages if pg.url.startswith("http")), None)
        if page is None:
            print("[ERR] 사용 가능한 page 없음"); return 2
        page.bring_to_front()
        ensure_erp_login(page)
        # MES 본 페이지 도달
        try:
            page.goto(MES_LAYOUT, wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)
        except Exception:
            pass
        deadline = time.time() + 30
        while time.time() < deadline:
            u = page.url
            if "mes-dev.samsong.com" in u and "/layout/" in u and "oauth2" not in u:
                break
            time.sleep(0.5)
        print(f"[phase0] MES 도달: {page.url}")

        scans, misses = fetch_mes_scans(page, args.line, dates)

    update_excel(target_path, bi_data, scans)

    total_qty  = sum(x["qty"] or 0 for x in bi_data)
    total_scan = sum(v for v in scans.values() if isinstance(v, int))
    print("\n=== 요약 ===")
    print(f"파일      : {target_path}")
    print(f"BI 산출량 : {total_qty:,}")
    print(f"MES 스캔  : {total_scan:,}")
    print(f"차이      : {total_qty - total_scan:,}")
    print(f"MES 미수집: {len(misses)}일")
    return 0


if __name__ == "__main__":
    sys.exit(main())
