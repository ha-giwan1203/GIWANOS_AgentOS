"""QIS 4탭(라인정지·재작업·선별작업·기타생산비용) 월별 자동 조회.

- Playwright 내장 Chromium을 별도 CDP 포트(9224)로 detach 기동 (회사 Chrome HTTP 차단 우회)
- 로그인 → 비용청구작성관리 진입 → 4탭 순회 → jqGrid getRowData
- 산출: 05_생산실적/조립비정산/{MM+1}월/QIS청구_{MM}월_raw.json + 요약 md

⛔ 작성·등록·저장·제출 버튼 절대 클릭 X. 조회 only.

사용:
  python qis_extract.py --month 2026-04
  python qis_extract.py --month 2026-04 --keep-open    # 작업 후 브라우저 유지
"""
import sys, os, json, time, re, subprocess, argparse, calendar
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from playwright.sync_api import sync_playwright

REPO = Path(__file__).resolve().parents[3]
SKILL_DIR = Path(__file__).parent
CRED_PATH = SKILL_DIR / ".oauth_qis.json"

QIS_HOME = "http://qis.samsong.co.kr/"
ACCOUNT_URL = "http://qis.samsong.co.kr/qis/claim/account/accountView.jsp?system_nm=QIS&menu_id=60012"
CDP_PORT = 9224
# Chromium user-data-dir — 홈 디렉토리에 (d0의 .flow-chrome-debug와 동일 패턴)
PROFILE = str(Path.home() / ".qis-chromium-profile")

TABS = [
    ("라인정지", "ui-id-1", "ui-id-2"),
    ("재작업",   "ui-id-3", "ui-id-4"),
    ("선별작업", "ui-id-5", "ui-id-6"),
    ("기타생산비용", "ui-id-7", "ui-id-8"),
]


def month_range(yyyymm: str):
    y, m = map(int, yyyymm.split("-"))
    last = calendar.monthrange(y, m)[1]
    return f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}"


def output_dir(yyyymm: str) -> Path:
    """도메인 관행: 매월 작업폴더 = 05_생산실적/조립비정산/{MM+1:02d}월/ (zero-padded)."""
    y, m = map(int, yyyymm.split("-"))
    nm = m + 1 if m < 12 else 1
    p = REPO / "05_생산실적" / "조립비정산" / f"{nm:02d}월"
    p.mkdir(parents=True, exist_ok=True)
    return p


def is_cdp_alive(port=CDP_PORT) -> bool:
    import urllib.request
    try:
        urllib.request.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=2)
        return True
    except Exception:
        return False


def ensure_chromium_cdp():
    """Playwright 내장 Chromium을 detach 모드로 CDP 9224 기동. 이미 살아있으면 skip."""
    if is_cdp_alive():
        return None
    with sync_playwright() as pw:
        exe = pw.chromium.executable_path
    flags = subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP if os.name == "nt" else 0
    proc = subprocess.Popen([
        exe,
        f"--remote-debugging-port={CDP_PORT}",
        f"--user-data-dir={PROFILE}",
        "--no-first-run", "--no-default-browser-check",
        "--unsafely-treat-insecure-origin-as-secure=http://qis.samsong.co.kr",
        "--disable-features=HttpsUpgrades,HttpsFirstBalancedMode,HttpsFirstModeIncognito,HttpsFirstModeForTypicallySecureUsers,HttpsFirstModeForAdvancedProtectionUsers",
        "--allow-running-insecure-content",
        "about:blank",
    ], creationflags=flags)
    # CDP up 대기
    for _ in range(15):
        time.sleep(1)
        if is_cdp_alive():
            return proc
    raise RuntimeError("CDP 9224 미기동")


def login_and_enter(ctx, user_id, pw_val):
    """로그인 + 비용청구작성관리 진입. 이미 로그인 상태면 진입만."""
    page = ctx.pages[0] if ctx.pages else ctx.new_page()
    page.goto(QIS_HOME, timeout=60000, wait_until="domcontentloaded")
    time.sleep(2)
    print(f"  [debug] after goto: url={page.url} title={page.title()!r}")
    has_login = page.locator('input[name="userid"]').count() > 0
    print(f"  [debug] login form: {has_login}")
    if has_login:
        page.fill('input[name="userid"]', user_id)
        page.fill('input[name="userpw"]', pw_val)
        page.evaluate("() => { const r=document.querySelector('input[name=company][value=BP]'); if(r){r.checked=true; r.click();} }")
        time.sleep(0.5)
        page.click("button.login-btn")
        time.sleep(5)
        print(f"  [debug] after login: url={page.url} title={page.title()!r}")
        print(f"  [debug] frames: {[(f.name, f.url[:60]) for f in page.frames]}")
    # detail frame 확보 + 비용청구작성관리 진입
    for _ in range(20):
        detail = next((f for f in page.frames if f.name == "detail"), None)
        if detail is not None: break
        time.sleep(0.5)
    if detail is None:
        print(f"  [debug] final frames: {[(f.name, f.url[:80]) for f in page.frames]}")
        raise RuntimeError("detail frame 미생성 (로그인 실패 의심)")
    detail.goto(ACCOUNT_URL, timeout=30000, wait_until="domcontentloaded")
    # 페이지 로드 완료 — 검색일자 input 등장까지 대기
    for _ in range(20):
        try:
            detail = next((f for f in page.frames if f.name == "detail"), None)
            if detail.locator('input[id^="STD_DT_"]').count() > 0:
                break
        except Exception: pass
        time.sleep(0.5)
    time.sleep(1)
    return page


def extract_tab(detail, label, tab_id, panel_id, dt_f, dt_t):
    """탭 활성화 + 검색일자 입력 + 검색 → jqGrid 데이터 + 합계."""
    detail.evaluate(f"() => document.getElementById('{tab_id}').click()")
    time.sleep(1.5)
    detail.evaluate(f"""
        (() => {{
            const panel = document.getElementById('{panel_id}');
            if(!panel) return;
            panel.querySelectorAll('input[id^="STD_DT_"]').forEach(i => i.value = '{dt_f}');
            panel.querySelectorAll('input[id^="END_DT_"]').forEach(i => i.value = '{dt_t}');
            // 검색 버튼 클릭
            for (const b of panel.querySelectorAll('button, input[type=button], a')) {{
                const t = (b.innerText || b.value || '').trim();
                if (t === '검색' || t === '조회' || t.includes('검색')) {{ b.click(); return; }}
            }}
        }})()
    """)
    time.sleep(4)
    return detail.evaluate(f"""
        (() => {{
            const panel = document.getElementById('{panel_id}');
            if(!panel) return {{err:'no panel'}};
            const tbl = panel.querySelector('table.ui-jqgrid-btable, table[role=grid][id]');
            if(!tbl) return {{err:'no jqgrid', html_len: panel.innerHTML.length}};
            const gid = tbl.id;
            let rows = [];
            try {{ rows = window.jQuery('#'+gid).jqGrid('getRowData'); }}
            catch(e) {{ return {{err: 'getRowData fail: '+e.message, gid}}; }}
            let sum_text = '';
            panel.querySelectorAll('*').forEach(el => {{
                const t = (el.innerText || '').trim();
                if (t.includes('청구비합계') && t.length < 80) sum_text = t;
            }});
            return {{gid, total: rows.length, rows, sum_text}};
        }})()
    """)


def save_outputs(results: dict, yyyymm: str):
    """QIS 4탭 raw를 기존 라인정지_MM월_raw.xlsx에 시트 추가. 별도 json·md 만들지 않음."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    out = output_dir(yyyymm)
    mm = yyyymm[-2:]
    xlsx = out / f"라인정지_{mm}월_raw.xlsx"
    if not xlsx.exists():
        # run.py 미실행 — 자체 xlsx 생성
        wb = openpyxl.Workbook(); wb.remove(wb.active)
    else:
        wb = openpyxl.load_workbook(xlsx)

    HF = Font(bold=True, color="FFFFFF")
    HB = PatternFill("solid", fgColor="305496")
    HEAD = Alignment(horizontal="center", vertical="center")

    for label, _, _ in TABS:
        sheet_name = f"QIS_{label}"
        if sheet_name in wb.sheetnames: del wb[sheet_name]
        rows = results.get(label, {}).get("rows", [])
        if not rows:
            continue  # 0건 탭은 시트 안 만듦 (_qis_meta 시트에 표시)
        ws = wb.create_sheet(sheet_name)
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for c in ws[1]:
            c.font = HF; c.fill = HB; c.alignment = HEAD
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    # QIS 4탭 메타
    if "_qis_meta" in wb.sheetnames: del wb["_qis_meta"]
    ms = wb.create_sheet("_qis_meta")
    ms.append(["tab", "count", "sum_text"])
    for c in ms[1]:
        c.font = HF; c.fill = HB; c.alignment = HEAD
    for label, _, _ in TABS:
        r = results.get(label, {})
        ms.append([label, r.get("total", 0), r.get("sum_text", "")])

    wb.save(xlsx)
    return xlsx


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--month", required=True, help="YYYY-MM")
    p.add_argument("--headless", action="store_true", help="창 안 띄우고 백그라운드")
    args = p.parse_args()

    cred = json.loads(CRED_PATH.read_text(encoding="utf-8"))
    user_id, pw_val = cred["id"], cred["pw"]
    df, dt = month_range(args.month)
    print(f"[start] QIS {args.month} ({df}~{dt})")

    results = {}
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=args.headless, args=[
            "--no-first-run", "--no-default-browser-check",
            "--unsafely-treat-insecure-origin-as-secure=http://qis.samsong.co.kr",
            "--disable-features=HttpsUpgrades,HttpsFirstBalancedMode,HttpsFirstModeIncognito,HttpsFirstModeForTypicallySecureUsers,HttpsFirstModeForAdvancedProtectionUsers",
            "--allow-running-insecure-content",
        ])
        ctx = browser.new_context()
        page = login_and_enter(ctx, user_id, pw_val)
        detail = next((f for f in page.frames if f.name == "detail"), None)
        if detail is None:
            raise RuntimeError("detail frame 없음")
        print(f"[detail] {detail.url}")

        for label, tab_id, panel_id in TABS:
            r = extract_tab(detail, label, tab_id, panel_id, df, dt)
            results[label] = r
            print(f"  {label}: total={r.get('total', 0)} err={r.get('err','')}")

        browser.close()

    xlsx = save_outputs(results, args.month)
    print(f"\n[OK] {xlsx}")
    for label, _, _ in TABS:
        r = results.get(label, {})
        print(f"  {label}: 건수={r.get('total', 0)}")


if __name__ == "__main__":
    main()
