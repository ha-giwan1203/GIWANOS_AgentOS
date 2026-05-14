"""월별 라인정지비용 통합 — G-ERP + QIS 단일 파일로.

선행 산출물:
  05_생산실적/조립비정산/{MM+1}월/라인정지_{MM}월_raw.xlsx  (run.py 출력)
  05_생산실적/조립비정산/{MM+1}월/라인정지_{MM}월_요약.md   (run.py 출력)
  05_생산실적/조립비정산/{MM+1}월/QIS청구_{MM}월_raw.json   (qis_extract.py 출력)

결과:
  - xlsx에 시트 추가: 통합집계 / QIS_기타생산비용 (등)
  - 요약 md를 G-ERP + QIS 통합 본문으로 재작성

사용:
  python merge_monthly.py --month 2026-04
"""
import sys, json, argparse, re, calendar
from pathlib import Path
from collections import defaultdict
try: sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception: pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

REPO = Path(__file__).resolve().parents[3]


def to_int(v):
    try: return int(str(v).replace(",", "").replace(" ", "") or 0)
    except: return 0


def month_dirs(yyyymm: str):
    y, m = map(int, yyyymm.split("-"))
    nm = m + 1 if m < 12 else 1
    out = REPO / "05_생산실적" / "조립비정산" / f"{nm}월"
    last = calendar.monthrange(y, m)[1]
    return out, f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}", f"{m:02d}"


def parse_gerp_summary(md_path: Path, meta_path: Path, xlsx_path: Path):
    """G-ERP 건수·합계 — 우선순위: meta json → md → xlsx 직접 계산.

    md는 merge_monthly 재실행으로 덮어써질 수 있어 신뢰 못함.
    meta json이 가장 신뢰. 없으면 raw xlsx에서 직접 계산.
    """
    # 1) meta json
    if meta_path.exists():
        try:
            m = json.loads(meta_path.read_text(encoding="utf-8"))
            return int(m.get("count", 0)), int(m.get("total_amount", 0))
        except Exception: pass

    # 2) raw xlsx 첫 시트(라인보상_*)에서 직접 계산
    if xlsx_path.exists():
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            gerp_sheet = next((n for n in wb.sheetnames if n.startswith("라인보상_")), None)
            if gerp_sheet:
                ws = wb[gerp_sheet]
                cnt = ws.max_row - 1
                # 합계 컬럼 후보 (run.py 매핑 fail로 청구비용 비어있을 수 있음)
                # → 도메인 공식: (주간 분당정지비 × 주간 정지시간) + (야간 분당정지비 × 야간 정지시간)
                # 그 컬럼이 raw에 있는 위치 찾기
                headers = [(c.value or "") for c in ws[1]]
                amt = 0
                # 1차: 청구비용·청구비합계 컬럼
                for kw in ("청구비합계", "청구비용", "청구비", "정지비"):
                    idxs = [i for i, h in enumerate(headers) if kw in str(h)]
                    if idxs:
                        i0 = idxs[0]
                        for r in range(2, ws.max_row + 1):
                            v = ws.cell(r, i0 + 1).value
                            try: amt += int(float(v or 0))
                            except: pass
                        if amt > 0:
                            wb.close()
                            return cnt, amt
                wb.close()
                return cnt, amt
        except Exception as e:
            print(f"  [warn] xlsx 합계 계산 실패: {e}")

    # 3) md fallback (덮어쓰기 전 원본만 신뢰)
    if md_path.exists():
        txt = md_path.read_text(encoding="utf-8")
        m_cnt = re.search(r"건수.*?\*\*(\d+)건\*\*", txt)
        m_amt = re.search(r"합계.*?\*\*([\d,]+)원\*\*", txt)
        cnt = int(m_cnt.group(1)) if m_cnt else 0
        amt = int(m_amt.group(1).replace(",", "")) if m_amt else 0
        return cnt, amt
    return 0, 0


HF = Font(bold=True, color="FFFFFF")
HB = PatternFill("solid", fgColor="305496")
HEAD = Alignment(horizontal="center", vertical="center")
TITLE = Font(bold=True, size=14, color="305496")
NOTE = Font(italic=True, color="C00000")


def merge(yyyymm: str):
    out, df, dt, mm = month_dirs(yyyymm)
    xlsx = out / f"라인정지_{mm}월_raw.xlsx"
    qis_json = out / f"QIS청구_{mm}월_raw.json"
    md = out / f"라인정지_{mm}월_요약.md"

    if not xlsx.exists():
        raise FileNotFoundError(f"G-ERP raw 없음: {xlsx} — run.py 먼저 실행")
    if not qis_json.exists():
        raise FileNotFoundError(f"QIS raw 없음: {qis_json} — qis_extract.py 먼저 실행")

    qis = json.loads(qis_json.read_text(encoding="utf-8"))

    # 1. xlsx 시트 추가
    wb = openpyxl.load_workbook(xlsx)
    for name in ("통합집계", "QIS_기타생산비용"):
        if name in wb.sheetnames: del wb[name]

    # QIS 데이터 시트
    rows = qis.get("기타생산비용", {}).get("rows", [])
    if rows:
        ws = wb.create_sheet("QIS_기타생산비용")
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        for c in ws[1]:
            c.font = HF; c.fill = HB; c.alignment = HEAD
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    # 통합집계
    ws = wb.create_sheet("통합집계", 0)
    ws.append([f"■ {yyyymm} 라인정지·라인교체 통합 — 대원테크(0109)"])
    ws["A1"].font = TITLE
    ws.append([])
    ws.append(["시스템", "청구유형", "건수", "금액(원)", "비고"])
    for c in ws[3]:
        c.font = HF; c.fill = HB; c.alignment = HEAD

    meta_path = out / f"라인정지_{mm}월_meta.json"
    gerp_cnt, gerp_amt = parse_gerp_summary(md, meta_path, xlsx)
    # 통합집계 결과를 다시 meta로 저장 (재실행 안정성)
    meta_path.write_text(json.dumps({"count": gerp_cnt, "total_amount": gerp_amt}, ensure_ascii=False), encoding="utf-8")
    ws.append(["G-ERP 라인보상상세현황", "라인정지", gerp_cnt, gerp_amt, "본사 등록"])

    for label in ("라인정지", "재작업", "선별작업", "기타생산비용"):
        rs = qis.get(label, {}).get("rows", [])
        cnt = len(rs)
        amt = sum(to_int(r.get("TOTAL_REQUIRE") or r.get("REWARD_LINE") or 0) for r in rs)
        type_name = "라인교체" if label == "기타생산비용" else label
        note = "협력사 작성" if cnt > 0 else "협력사 작성 0"
        ws.append([f"QIS Claim ({label} 탭)", type_name, cnt, amt, note])

    ws.append([])
    ws.append(["※ 청구유형이 다르므로 단순 합산 금지. 정산 본체에 별도 라인으로 들어감."])
    ws.cell(ws.max_row, 1).font = NOTE

    # QIS 라인교체 귀책 세분
    if rows:
        ws.append([])
        ws.append(["■ QIS 라인교체 귀책 세분"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="305496")
        ws.append(["귀책 부서", "건수", "금액(원)"])
        for c in ws[ws.max_row]:
            c.font = HF; c.fill = HB; c.alignment = HEAD
        by_dept = defaultdict(lambda: {"cnt": 0, "amt": 0})
        for r in rows:
            d = r.get("NAME_CHARGE") or r.get("DEPT_ING") or "?"
            by_dept[d]["cnt"] += 1
            by_dept[d]["amt"] += to_int(r.get("TOTAL_REQUIRE") or 0)
        for d, v in sorted(by_dept.items(), key=lambda x: -x[1]["amt"]):
            ws.append([d, v["cnt"], v["amt"]])

    for col, w in zip("ABCDE", (25, 18, 8, 14, 18)):
        ws.column_dimensions[col].width = w

    wb.save(xlsx)
    print(f"[OK] xlsx 시트 추가: {wb.sheetnames}")

    # 2. md 통합 재작성
    qis_etc_rows = qis.get("기타생산비용", {}).get("rows", [])
    qis_etc_amt = sum(to_int(r.get("TOTAL_REQUIRE") or 0) for r in qis_etc_rows)
    by_dept2 = defaultdict(lambda: {"cnt": 0, "amt": 0})
    for r in qis_etc_rows:
        d = r.get("NAME_CHARGE") or r.get("DEPT_ING") or "?"
        by_dept2[d]["cnt"] += 1
        by_dept2[d]["amt"] += to_int(r.get("TOTAL_REQUIRE") or 0)

    L = [f"# {yyyymm} 라인정지·라인교체 통합 요약 — 대원테크", "",
         f"- 기간: {df} ~ {dt}",
         "- 데이터 원천: G-ERP 라인보상상세현황 + QIS Claim관리 비용청구작성관리",
         "",
         "## 시스템별·청구유형별 (합산 금지)",
         "",
         "| 시스템 | 청구유형 | 건수 | 금액(원) | 비고 |",
         "|--------|---------|-----:|---------:|------|",
         f"| G-ERP 라인보상상세현황 | 라인정지 | {gerp_cnt} | {gerp_amt:,} | 본사 등록 |",
         f"| QIS Claim (라인정지 탭) | 라인정지 | 0 | 0 | 협력사 작성 0 |",
         f"| QIS Claim (재작업 탭) | 재작업 | 0 | 0 | 협력사 작성 0 |",
         f"| QIS Claim (선별작업 탭) | 선별작업 | 0 | 0 | 협력사 작성 0 |",
         f"| QIS Claim (기타생산비용 탭) | 라인교체 | {len(qis_etc_rows)} | {qis_etc_amt:,} | 협력사 작성 |",
         "",
         "**※ 청구유형 다름. 단순 합산 금지. 정산 본체에 별도 라인으로 들어감.**",
         ""]
    if qis_etc_rows:
        L.append("## QIS 라인교체 귀책 세분")
        L.append("")
        L.append("| 귀책 부서 | 건수 | 금액(원) |")
        L.append("|---------|-----:|---------:|")
        for d, v in sorted(by_dept2.items(), key=lambda x: -x[1]["amt"]):
            L.append(f"| {d} | {v['cnt']} | {v['amt']:,} |")
        L.append("")
    L += [
        "## G-ERP 라인정지 상세",
        "",
        f"xlsx 시트 `라인보상_{mm}월` / `집계` 참조.",
        "",
        "## 매월 회귀 확인 항목",
        "- QIS 협력사 작성 라인교체가 다음 달 G-ERP에 sync 등장하는지",
        "- QIS 라인정지/재작업/선별작업 탭 신규 등장 여부",
        "- G-ERP 미승인(`APPROVAL_YN ≠ Y`) / 미접수 잔존 건",
        "",
        "## 산출물",
        f"- `라인정지_{mm}월_raw.xlsx` — G-ERP + QIS_기타생산비용 + 통합집계 시트",
        f"- `QIS청구_{mm}월_raw.json` — QIS 4탭 raw",
    ]
    md.write_text("\n".join(L), encoding="utf-8")
    print(f"[OK] md: {md}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--month", required=True, help="YYYY-MM")
    args = p.parse_args()
    merge(args.month)


if __name__ == "__main__":
    main()
