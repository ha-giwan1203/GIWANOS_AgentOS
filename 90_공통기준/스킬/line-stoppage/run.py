"""월별 라인정지비(라인보상상세현황) G-ERP 자동 조회.

사용:
  python run.py --month 2026-04                  # 대원테크 4월
  python run.py --month 2026-04 --cmpy 0110      # 다른 업체
  python run.py --month 2026-04 --line SP3M3     # 라인 필터
  python run.py --month 2026-04 --no-browser     # API 1회 시도(현재 -9999, 미동작)

산출:
  05_생산실적/조립비정산/{MM+1}월/라인정지_{MM}월_raw.xlsx
  05_생산실적/조립비정산/{MM+1}월/라인정지_{MM}월_요약.md
"""
import sys, os, json, time, argparse, calendar
from datetime import datetime
from pathlib import Path
from collections import defaultdict

# UTF-8 출력 (Windows cp949 콘솔 회피)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

REPO_ROOT = Path(__file__).resolve().parents[3]
D0_SKILL = REPO_ROOT / "90_공통기준" / "스킬" / "d0-production-plan"
sys.path.insert(0, str(D0_SKILL))

from run import (  # type: ignore
    ensure_chrome_cdp, erp_login_via_http,
    _inject_cookies_to_playwright, ensure_erp_login,
    CDP_URL,
)
from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ERP = "http://erp-dev.samsong.com:19100"
PAGE_URL = f"{ERP}/costCharge/viewListCostBillDetail.do"


def gerp_blackout_now() -> bool:
    """매시 x0:10~13/20~23/30~33/40~43/50~53 G-ERP 차단."""
    m = datetime.now().minute
    return (10 <= m <= 13) or (20 <= m <= 23) or (30 <= m <= 33) or (40 <= m <= 43) or (50 <= m <= 53)


def month_range(yyyymm: str):
    y, m = map(int, yyyymm.split("-"))
    last = calendar.monthrange(y, m)[1]
    return (f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}")


def output_dir(yyyymm: str) -> Path:
    y, m = map(int, yyyymm.split("-"))
    next_m = m + 1 if m < 12 else 1
    next_y = y if m < 12 else y + 1
    p = REPO_ROOT / "05_생산실적" / "조립비정산" / f"{next_m}월"
    p.mkdir(parents=True, exist_ok=True)
    return p


def fetch_via_browser(yyyymm: str, cmpy: str, line: str) -> dict:
    """Playwright + CDP 9223. 라인보상상세현황 grid 직접 추출."""
    df, dt = month_range(yyyymm)
    ensure_chrome_cdp()
    time.sleep(2)

    sess = erp_login_via_http()
    if sess is None:
        raise RuntimeError("ERP OAuth 실패 — .oauth.json 확인")

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp(CDP_URL)
        ctx = browser.contexts[0]
        try:
            _inject_cookies_to_playwright(ctx, sess)
        except Exception as e:
            print(f"[warn] cookie 주입 실패: {e}")

        page = ctx.new_page()
        page.goto(PAGE_URL, timeout=30000)
        time.sleep(1.5)
        if "auth-dev" in page.url or "/login" in page.url:
            print("[oauth] 브라우저 로그인 fallback")
            ensure_erp_login(page)
            page.goto(PAGE_URL, timeout=30000)
            time.sleep(1.5)

        page.wait_for_selector("#searchOcrnDaF", timeout=15000)
        page.wait_for_selector("#btnSearch", timeout=5000)

        # 날짜
        page.evaluate(
            f"() => {{ $('#searchOcrnDaF').val('{df}'); $('#searchOcrnDaT').val('{dt}'); }}"
        )

        # 조립업체
        cmpy_options = page.evaluate(
            "() => Array.from(document.querySelectorAll('#searchCmpy option')).map(o => ({v: o.value, t: o.text}))"
        )
        cmpy_val = None
        for o in cmpy_options:
            if o["v"] == cmpy:
                cmpy_val = cmpy
                break
        if cmpy_val:
            page.evaluate(f"() => $('#searchCmpy').val('{cmpy_val}').trigger('change')")

        # 라인 필터 (선택)
        if line:
            line_options = page.evaluate(
                "() => Array.from(document.querySelectorAll('#searchLineCd option')).map(o => ({v: o.value, t: o.text}))"
            )
            for o in line_options:
                if line in (o.get("t") or "") or line == o.get("v"):
                    page.evaluate(f"() => $('#searchLineCd').val('{o['v']}').trigger('change')")
                    break

        page.click("#btnSearch")
        time.sleep(3)

        grid = page.evaluate(
            """() => {
                const $g = $('.pq-grid').first();
                if (!$g.length) return {err: 'no pq-grid', total: 0, cols: [], data: []};
                const g = $g.pqGrid('instance');
                if (!g) return {err: 'no instance', total: 0, cols: [], data: []};
                const data = g.option('dataModel.data') || [];
                const cm = g.option('colModel') || [];
                return {
                    total: data.length,
                    cols: cm.map(c => ({di: c.dataIndx, title: c.title})),
                    data: data,
                };
            }"""
        )
        return grid


def save_raw_xlsx(grid: dict, out_path: Path, yyyymm: str, cmpy: str):
    cols = grid["cols"]
    rows = grid["data"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"라인보상_{yyyymm[-2:]}월"
    headers = [{"di": c["di"], "title": (c.get("title") or c["di"]).replace("<BR>", " ")} for c in cols]
    ws.append([h["title"] for h in headers])
    for r in rows:
        ws.append([r.get(h["di"], "") for h in headers])
    # 스타일
    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="305496")
    for c in ws[1]:
        c.font = hf; c.fill = hb; c.alignment = Alignment(horizontal="center", vertical="center")
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(h["title"]) + 2, 10), 30)

    # 집계 시트
    sw = wb.create_sheet("집계")
    sw.append([f"■ {yyyymm} 라인보상상세현황 — 업체코드 {cmpy} ({len(rows)}건)"])
    sw.append([])
    _append_group(sw, "[청구유형별]", rows, lambda r: r.get("COST_BILL_DOC_TYPE_NM") or "?")
    _append_group(sw, "[라인별]", rows, lambda r: (r.get("LINE_NM") or "?").replace("<BR>", " ").strip())
    _append_group(sw, "[귀책업체/부서별]", rows, lambda r: (r.get("REASON_CMPY_NM") or r.get("RESP_DEPT_NM") or "(공란)").strip() or "(공란)")
    _append_group(sw, "[발생유형별]", rows, lambda r: r.get("STOP_REASON_DIV_NM") or "?")
    sw["A1"].font = Font(bold=True, size=14, color="305496")
    wb.save(out_path)


def _append_group(ws, title, rows, key_fn):
    g = defaultdict(lambda: {"cnt": 0, "amt": 0.0})
    for r in rows:
        k = key_fn(r)
        g[k]["cnt"] += 1
        g[k]["amt"] += float(r.get("COST_BILL_TOT") or 0)
    ws.append([title])
    ws.append(["항목", "건수", "금액(원)"])
    for k, v in sorted(g.items(), key=lambda x: -x[1]["amt"]):
        ws.append([k, v["cnt"], v["amt"]])
    ws.append([])


def save_summary_md(grid: dict, out_path: Path, yyyymm: str, cmpy: str):
    rows = grid["data"]
    total_amt = sum(float(r.get("COST_BILL_TOT") or 0) for r in rows)
    appr_y = sum(1 for r in rows if r.get("APPROVAL_YN") == "Y")
    acc_n = sum(1 for r in rows if r.get("ACCEPT_YN_NM") != "접수")
    blank_cmpy = sum(1 for r in rows if not (r.get("REASON_CMPY_NM") or "").strip())
    blank_car = sum(1 for r in rows if not (r.get("CARTYPE_NM") or "").strip())

    L = []
    L.append(f"# {yyyymm} 라인정지비 요약 — 업체 {cmpy}")
    L.append("")
    df, dt = month_range(yyyymm)
    L.append(f"- 기간: {df} ~ {dt}")
    L.append(f"- 건수: **{len(rows)}건**")
    L.append(f"- 합계: **{total_amt:,.0f}원**")
    L.append("")
    for title, key in [
        ("청구유형별", lambda r: r.get("COST_BILL_DOC_TYPE_NM") or "?"),
        ("라인별", lambda r: (r.get("LINE_NM") or "?").replace("<BR>", " ").strip()),
        ("귀책업체/부서별", lambda r: (r.get("REASON_CMPY_NM") or r.get("RESP_DEPT_NM") or "(공란)").strip() or "(공란)"),
        ("발생유형별", lambda r: r.get("STOP_REASON_DIV_NM") or "?"),
    ]:
        g = defaultdict(lambda: {"cnt": 0, "amt": 0.0})
        for r in rows:
            k = key(r); g[k]["cnt"] += 1; g[k]["amt"] += float(r.get("COST_BILL_TOT") or 0)
        L.append(f"## {title}")
        L.append("| 항목 | 건수 | 금액(원) |")
        L.append("|------|-----:|---------:|")
        for k, v in sorted(g.items(), key=lambda x: -x[1]["amt"]):
            L.append(f"| {k} | {v['cnt']} | {v['amt']:,.0f} |")
        L.append("")
    L.append("## 상태 점검")
    L.append(f"- 승인 Y: {appr_y}/{len(rows)}")
    L.append(f"- 미접수: {acc_n}건")
    L.append(f"- 귀책 공란: {blank_cmpy}건")
    L.append(f"- 차종 공란: {blank_car}건")
    L.append("")
    L.append("## 데이터 원천")
    L.append("- G-ERP 클레임관리 > 라인보상관리 > 라인보상상세현황")
    L.append(f"- 검색조건: 발생일자 {df}~{dt}, 조립업체 {cmpy}")
    out_path.write_text("\n".join(L), encoding="utf-8")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--month", required=True, help="YYYY-MM (예: 2026-04)")
    p.add_argument("--cmpy", default="0109", help="조립업체코드 (default 0109 대원테크)")
    p.add_argument("--line", default="", help="라인 필터 (예: SP3M3, 빈값=전체)")
    args = p.parse_args()

    if gerp_blackout_now():
        print("[wait] GERP 동기화 차단 시간대 — 60초 대기")
        time.sleep(60)

    yyyymm = args.month
    print(f"[start] {yyyymm} / 업체 {args.cmpy} / 라인 {args.line or '전체'}")

    grid = fetch_via_browser(yyyymm, args.cmpy, args.line)
    if grid.get("err"):
        print(f"[FAIL] grid 추출 실패: {grid['err']}")
        sys.exit(2)

    out = output_dir(yyyymm)
    mm = yyyymm[-2:]
    raw = out / f"라인정지_{mm}월_raw.xlsx"
    md = out / f"라인정지_{mm}월_요약.md"

    save_raw_xlsx(grid, raw, yyyymm, args.cmpy)
    save_summary_md(grid, md, yyyymm, args.cmpy)

    total = sum(float(r.get("COST_BILL_TOT") or 0) for r in grid["data"])
    # meta.json — merge_monthly가 재실행 안정성 위해 사용
    meta = out / f"라인정지_{mm}월_meta.json"
    meta.write_text(json.dumps({"count": len(grid["data"]), "total_amount": int(total)}, ensure_ascii=False), encoding="utf-8")
    print(f"[OK] {len(grid['data'])}건 / {total:,.0f}원")
    print(f"     {raw}")
    print(f"     {md}")
    print(f"     {meta}")


if __name__ == "__main__":
    main()
