"""교육 기록부 작성 — 대원테크 표준 양식 신규 생성.

사용:
  python run.py --line SD9A01 --topic "라인내8대유의사항" --body-file body.txt
  python run.py --line SP3M3 --topic "MGG역조립" --body-file body.txt --date 2026-05-18 --instructor "하지완 과장"

기준 양식: 06_생산관리/품질/교육기록부/대원테크_부적합교육_MX5_바코드이종_20260428.xlsx
"""
from __future__ import annotations

import argparse
import shutil
import sys
import io
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

REPO = Path(__file__).resolve().parents[3]
EDU_DIR = REPO / "06_생산관리" / "품질" / "교육기록부"
TEMPLATE = EDU_DIR / "대원테크_부적합교육_MX5_바코드이종_20260428.xlsx"


def build_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--line", required=True, help="라인명 (SD9A01·SP3M3·MX5·QX1 등)")
    p.add_argument("--topic", required=True, help="주제 슬러그 (파일명용, 공백 없이)")
    p.add_argument("--title", help="교육 제목 (G9). 미지정 시 --topic 사용")
    p.add_argument("--body-file", required=True, help="본문 텍스트 파일 경로 (utf-8)")
    p.add_argument("--date", default=datetime.now().strftime("%Y-%m-%d"))
    p.add_argument("--instructor", default="하지완 과장")
    p.add_argument("--duration", default="30분")
    p.add_argument("--internal", default="사내", choices=["사내", "사외"])
    p.add_argument("--audience", help="교육대상자. 미지정 시 '{LINE} 라인 작업자'")
    p.add_argument("--overwrite", action="store_true", help="동일 파일명 있을 때 덮어쓰기")
    return p.parse_args()


def target_path(line: str, topic: str, date: str) -> Path:
    ymd = date.replace("-", "")
    return EDU_DIR / f"대원테크_부적합교육_{line}_{topic}_{ymd}.xlsx"


def fill(target: Path, args: argparse.Namespace, body: str) -> None:
    wb = openpyxl.load_workbook(target, data_only=False)
    if "교육기록부" not in wb.sheetnames:
        raise SystemExit(f"FAIL: 시트 '교육기록부' 부재 — 양식 변형 의심: {target}")
    ws = wb["교육기록부"]

    ws["G6"] = datetime.strptime(args.date, "%Y-%m-%d")
    ws["V6"] = args.internal
    ws["G7"] = args.instructor
    ws["V7"] = args.duration
    ws["G8"] = args.audience or f"{args.line} 라인 작업자"
    ws["V8"] = "명(참석자 명단 별첨)"
    ws["G9"] = args.title or args.topic

    g10 = ws["G10"]
    g10.value = body
    g10.alignment = Alignment(
        horizontal=g10.alignment.horizontal or "left",
        vertical="top",
        wrap_text=True,
    )

    ws["W3"] = None
    ws["AC3"] = None
    ws["E44"] = args.title or args.topic
    ws["T45"] = args.instructor
    ws["B38"] = None

    for r in range(49, 66):
        for col in (4, 10, 13, 19, 25, 28):  # D J M S Y AB
            ws.cell(r, col).value = None

    wb.save(target)


def verify(target: Path, args: argparse.Namespace) -> None:
    wb = openpyxl.load_workbook(target, data_only=False)
    ws = wb["교육기록부"]
    merged = len(list(ws.merged_cells.ranges))
    size = target.stat().st_size
    assert size > 15000, f"size {size} < 15KB"
    assert 185 <= merged <= 189, f"merged cells {merged} (expected ~187)"
    assert ws["G6"].value is not None
    assert (args.line in str(ws["G8"].value)) , f"G8 라인 불일치: {ws['G8'].value}"
    print(f"OK size={size} merged={merged} G6={ws['G6'].value} G8={ws['G8'].value}")


def main() -> None:
    args = build_args()
    if not TEMPLATE.exists():
        raise SystemExit(f"FAIL: 기준 양식 부재 — {TEMPLATE}")

    target = target_path(args.line, args.topic, args.date)
    if target.exists() and not args.overwrite:
        # 자동 _v2/_v3 부여
        for n in range(2, 10):
            alt = target.with_name(target.stem + f"_v{n}.xlsx")
            if not alt.exists():
                target = alt
                break
        else:
            raise SystemExit(f"FAIL: 동명 파일 9개 초과 — {target}")

    body = Path(args.body_file).read_text(encoding="utf-8")
    shutil.copy(TEMPLATE, target)
    fill(target, args, body)
    verify(target, args)
    print(f"SAVED: {target}")


if __name__ == "__main__":
    main()
