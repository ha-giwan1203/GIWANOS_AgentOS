"""
오류리스트 GERP누락 행을 GERP에서 일괄 조회 → 등록 현황 수집 → 양식대로 작성.

흐름:
1. 정산_수식버전_{MM}월_VALUES.xlsx 오류리스트 시트에서 GERP누락 행 추출
2. erp_lookup.lookup_pns_lines로 라인 등록 현황 일괄 수집
3. 결과 분석 → 비고 컬럼에 ERP 정정 가이드 박기
4. 오류리스트 양식 그대로 + GERP확인 시트 추가하여 저장

비고 분류:
- "본라인 미등록 / GERP 다른라인 등록: {라인}({업체})"
- "GERP 미등록 — 신규 등록 필요"
- "본라인 등록 OK (확정)" — 이상 케이스, 빌더 점검

사용:
    python collect_gerp_status.py {MM} [--values {VALUES.xlsx}] [--out {산출폴더}]
"""
import argparse
import json
import shutil
import sys
import time
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from erp_lookup import lookup_pns_lines


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("month", help="정산 대상 월 MM")
    ap.add_argument("--values", help="VALUES.xlsx 경로 (기본: 05_생산실적/조립비정산/{MM+1}월/오류리스트_재검증_*/)")
    ap.add_argument("--out", help="산출 폴더")
    ap.add_argument("--appl-da", help="ERP 적용일 (YYYY-MM-DD, 기본 정산월 마지막 날)")
    args = ap.parse_args()

    mm = f'{int(args.month):02d}'
    repo_root = Path(__file__).resolve().parents[3]

    # 1. VALUES.xlsx 자동 탐색 — 정산 도메인 작업폴더(05_생산실적/조립비정산/{MM+1}월/)에서
    if args.values:
        values_path = Path(args.values)
    else:
        settle_month_dir = repo_root / "05_생산실적" / "조립비정산" / f"{int(mm)+1:02d}월"
        cand_dirs = sorted(settle_month_dir.glob("오류리스트_재검증_*"), reverse=True)
        if not cand_dirs:
            print(f"[ERROR] {settle_month_dir}/오류리스트_재검증_* 폴더 없음. run_check.py 먼저 실행")
            sys.exit(1)
        values_path = cand_dirs[0] / f"정산_수식버전_{mm}월_VALUES.xlsx"
    if not values_path.exists():
        print(f"[ERROR] VALUES 파일 없음: {values_path}")
        sys.exit(1)
    print(f"[1/5] VALUES 로드: {values_path}")

    out_dir = Path(args.out) if args.out else values_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # ERP 적용일 — 정산월 마지막 날 (4월 → 2026-04-30)
    appl_da = args.appl_da or f'2026-{mm}-30'

    # 2. 오류리스트 시트 GERP누락 추출
    wb = openpyxl.load_workbook(values_path, data_only=True)
    err_ws = wb['오류리스트']
    gerp_missing = []
    for r in range(5, err_ws.max_row + 1):
        if err_ws.cell(r, 18).value == 'GERP누락':
            gerp_missing.append({
                'r': r,
                'pn': str(err_ws.cell(r, 1).value).strip(),
                'line': str(err_ws.cell(r, 3).value or '').strip(),
                'q_diff': err_ws.cell(r, 17).value,
            })
    pns = [g['pn'] for g in gerp_missing]
    print(f"[2/5] GERP누락 행 {len(gerp_missing)}건 추출 (적용일 {appl_da})")

    # 3. ERP 일괄 조회
    print(f"[3/5] ERP 자동 진입 + 일괄 조회 시작 (예상 {len(pns) * 0.5 / 60:.1f}분)")
    raw = lookup_pns_lines(pns, appl_da=appl_da, cmpy_cd="0109")

    # 4. 결과 분석 → 비고 생성
    print(f"[4/5] 결과 분석 + 비고 생성")
    enriched = []
    summary = defaultdict(int)
    for g in gerp_missing:
        pn = g['pn']
        line = g['line']
        rows = raw.get(pn, [])
        # ERP 등록 라인 분류
        same_line_active = []
        other_line_active = []
        unconfirmed = []
        for r in rows:
            decide = r.get('DECIDE_NM', '')
            line_cd = r.get('ASSY_LINE_CD', '')
            cmpy_cd = r.get('ASSY_CMPY_CD', '')
            cmpy_nm = r.get('ASSY_CMPY_NM', '')
            cost = r.get('ASSY_COST', 0)
            entry = f"{line_cd}({cmpy_cd}/{cmpy_nm}) 단가={cost} {decide}"
            if decide == '확정':
                if line_cd == line:
                    same_line_active.append(entry)
                else:
                    other_line_active.append(entry)
            else:
                unconfirmed.append(entry)

        if same_line_active:
            category = '본라인등록확인필요'
            memo = f"본라인 등록 OK: {' | '.join(same_line_active)} → 빌더 점검"
        elif other_line_active:
            category = '라인매칭'
            memo = f"본라인 미등록 / 다른라인 등록: {' | '.join(other_line_active)}"
        elif unconfirmed:
            category = '검토중'
            memo = f"검토중: {' | '.join(unconfirmed)}"
        else:
            category = '신규등록'
            memo = "GERP 미등록 — 신규 등록 필요"
        summary[category] += 1
        enriched.append({**g, 'rows': rows, 'category': category, 'memo': memo})

    print(f"   분포: {dict(summary)}")

    # 5. 양식대로 산출 — VALUES 사본(본체 안 오류리스트 시트)에서 오류리스트만 추출
    # (별도 오류리스트_MM월.xlsx 파일은 9컬럼 옛 형식, VALUES 안 오류리스트 시트가 20컬럼 사용자 양식)
    out_xlsx = out_dir / f"오류리스트_{mm}월_GERP확인완료.xlsx"
    src_wb = openpyxl.load_workbook(values_path, data_only=True)
    src_err = src_wb['오류리스트']
    new_wb = openpyxl.Workbook()
    new_err = new_wb.active
    new_err.title = '오류리스트'
    # 오류리스트 시트 전체 복사 (헤더+데이터)
    for row in src_err.iter_rows(values_only=False):
        for cell in row:
            new_cell = new_err.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    # 컬럼 너비 복사
    for col_letter, dim in src_err.column_dimensions.items():
        new_err.column_dimensions[col_letter].width = dim.width
    new_wb.save(out_xlsx)
    print(f"[5/5] 양식 추출: {out_xlsx.name}")

    out_wb = openpyxl.load_workbook(out_xlsx)
    out_err = out_wb['오류리스트']
    # 비고 컬럼(col 20) GERP누락 행만 갱신
    pn_to_memo = {(g['pn'], g['line']): g for g in enriched}
    updated = 0
    for r in range(5, out_err.max_row + 1):
        if out_err.cell(r, 18).value != 'GERP누락':
            continue
        pn = str(out_err.cell(r, 1).value).strip()
        line = str(out_err.cell(r, 3).value or '').strip()
        g = pn_to_memo.get((pn, line))
        if g:
            out_err.cell(r, 20).value = f"[{g['category']}] {g['memo']}"
            updated += 1

    # GERP확인 시트 신설 (요약 + 분류별 카운트)
    if 'GERP확인' in out_wb.sheetnames:
        del out_wb['GERP확인']
    chk_ws = out_wb.create_sheet('GERP확인')
    chk_ws.append([f'{int(mm)}월 GERP누락 ERP 자동 조회 결과'])
    chk_ws['A1'].font = Font(bold=True, size=14)
    chk_ws.append([f'적용일: {appl_da} / 조회 {len(pns)}건 / 갱신 {updated}건'])
    chk_ws.append([])
    chk_ws.append(['분류', '건수', '의미', 'ERP 정정 방향'])
    for c in chk_ws[4]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D9E1F2')
    chk_ws.append(['신규등록', summary.get('신규등록', 0), 'GERP 어디에도 등록 안 됨', '본라인에 단가 신규 등록'])
    chk_ws.append(['라인매칭', summary.get('라인매칭', 0), '다른 라인엔 확정 등록', '본라인 추가 등록'])
    chk_ws.append(['본라인등록확인필요', summary.get('본라인등록확인필요', 0), '본라인 등록인데 GERP 실적 0', '빌더 SUMIFS 점검'])
    chk_ws.append(['검토중', summary.get('검토중', 0), '미확정 상태', '확정 처리 후 재조회'])
    chk_ws.append([])
    chk_ws.append(['상세'])
    chk_ws.append(['품번', '본라인', '차이qty', '분류', 'ERP 등록 현황'])
    for c in chk_ws[chk_ws.max_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D9E1F2')
    for g in enriched:
        chk_ws.append([g['pn'], g['line'], g['q_diff'], g['category'], g['memo']])
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        chk_ws.column_dimensions[col_letter].width = 18 if col_letter != 'E' else 80

    out_wb.save(out_xlsx)

    # JSON 로그 저장 (재실행 캐시용)
    log_path = out_dir / f"erp_gerp_status_raw_{mm}월.json"
    with open(log_path, 'w', encoding='utf-8') as f:
        json.dump({
            'extract_date': time.strftime('%Y-%m-%d %H:%M:%S'),
            'appl_da': appl_da,
            'gerp_missing_count': len(pns),
            'summary': dict(summary),
            'enriched': [{k: v for k, v in g.items() if k != 'rows'} for g in enriched],
            'raw': raw,
        }, f, ensure_ascii=False, indent=2)

    print(f"\n[DONE] 산출물:")
    print(f"  - {out_xlsx}")
    print(f"  - {log_path.name}")
    print(f"\n=== 분류 요약 ===")
    for k, v in summary.items():
        print(f"  {k}: {v}건")


if __name__ == '__main__':
    main()
