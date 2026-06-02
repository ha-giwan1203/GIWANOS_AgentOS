"""SP3M3 야간 보충 등록 모드.

정규 evening 경로의 `출력용` 시트가 stale일 때, 사용자가 명시한 시트에서
야간 품번을 다시 추출해 Phase 3~5를 좁게 수행한다.
"""
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import run as d0


LINE = "SP3M3"


def extract_night_items_from_sheet(ws):
    """B열 주간계획 경계 전까지 I=신MES 품번, K=지시 수량을 읽는다."""
    night_end = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and "주간계획" in str(v):
            night_end = r - 1
            break
    if night_end is None:
        raise RuntimeError(f"{ws.title} 시트 주간계획 경계 미발견")

    items = []
    skipped_korean = []
    for r in range(3, night_end + 1):
        prod_no = ws.cell(r, 9).value
        qty = ws.cell(r, 11).value
        if not (prod_no and qty):
            continue
        try:
            qty_int = int(qty)
        except (ValueError, TypeError):
            continue
        pno = str(prod_no).strip()
        if any("가" <= ch <= "힣" for ch in pno):
            skipped_korean.append({"row": r, "PROD_NO": pno, "QTY": qty_int})
            continue
        items.append({"PROD_NO": pno, "QTY": qty_int})
    return items, skipped_korean


def load_items(plan_path: Path, source_sheet: str):
    wb = load_workbook(plan_path, data_only=True)
    try:
        if source_sheet not in wb.sheetnames:
            raise RuntimeError(f"{plan_path.name}에 {source_sheet!r} 시트 없음")
        if source_sheet == "출력용":
            items = d0.extract_sp3m3_night(wb, allow_stale_output=True)
            skipped = []
        else:
            items, skipped = extract_night_items_from_sheet(wb[source_sheet])
        return items, skipped
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser(description="SP3M3 야간 보충 등록 모드")
    ap.add_argument("--target-date", required=True, help="YYYY-MM-DD 생산지시서 파일명 날짜이자 보충 등록 기준 생산일")
    ap.add_argument("--source-sheet", choices=["출력용", "생산계획"], default="생산계획")
    ap.add_argument("--dry-run", action="store_true", help="추출만 검증하고 xlsx 생성/ERP/MES 반영은 하지 않음")
    args = ap.parse_args()

    target_date = datetime.strptime(args.target_date, "%Y-%m-%d")
    plan_path = d0.find_plan_file(target_date)
    items, skipped = load_items(plan_path, args.source_sheet)
    if not items:
        raise RuntimeError(f"{args.source_sheet} 시트에서 SP3M3 야간 등록 대상 0건")

    out_xlsx = d0.UPLOAD_DIR / f"d0_{LINE}_evening_supplement_{target_date.strftime('%Y%m%d')}.xlsx"
    print(f"[supplement] plan={plan_path}")
    print(f"[supplement] source_sheet={args.source_sheet} target_date={target_date.strftime('%Y-%m-%d')}")
    print(f"[supplement] 추출 {len(items)}건, 한글 PROD_NO skip {len(skipped)}건")
    for row in skipped:
        print(f"  [skip] r={row['row']} {row['PROD_NO']} QTY={row['QTY']}")
    for item in items:
        print(f"  - {item['PROD_NO']} QTY={item['QTY']}")

    if args.dry_run:
        print(f"[supplement] dry-run: xlsx 생성/ERP D0 저장/MES 전송 미실행 (예정 xlsx={out_xlsx})")
        return

    print(
        f"[비가역 통보] SP3M3 야간 보충 등록 {len(items)}건 "
        f"prod_date={target_date.strftime('%Y-%m-%d')} ERP D0 저장/서열 반영 및 MES 전송 진입"
    )
    d0.make_upload_xlsx(items, target_date, out_xlsx)
    sess = d0.erp_login_via_http()
    if sess is None:
        raise RuntimeError("ERP OAuth 실패")
    d0.d0_upload_via_http(sess, out_xlsx, parse_only=False)

    cfg = d0.LINE_CONFIG[LINE]
    batch = d0.api_rank_batch_via_http(
        sess, items, LINE, cfg["save_url"], target_date, day_opt="1", strict_regdt=False
    )
    print(f"[supplement] api_rank_batch strict_regdt=False: {batch}")
    if batch["failed"] > 0 or batch["missing"] > 0:
        raise RuntimeError(f"보충 Phase4 실패/누락: {batch}")

    res = d0.final_save_via_http_with_pbom_guard(
        sess, cfg["save_url"], batch["last_m_row"], batch["last_parent_prod_id"],
        LINE, day_opt="1", processed_order=batch.get("processed_order"),
    )
    print(f"[supplement] final_save result: {res}")
    print("=== evening_supplement 완료 ===")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[supplement] FAIL: {exc}", file=sys.stderr)
        raise
