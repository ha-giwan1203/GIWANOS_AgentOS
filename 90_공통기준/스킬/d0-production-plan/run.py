"""/d0-plan — ERP D0추가생산지시 자동화 (SP3M3 MAIN + SD9A01 OUTER).

세션별 업로드:
  저녁 (evening): SP3M3 야간 + SD9A01 OUTER D+1 (생산일 = 내일)
  아침 (morning): SP3M3 주간 (생산일 = 내일, 누적 ≥ 3600 컷)

실행:
  python run.py --session evening
  python run.py --session morning
  python run.py --session auto --dry-run
  python run.py --session evening --line SP3M3
"""
import sys, io, os, re, json, time, base64, argparse, subprocess
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
from playwright.sync_api import sync_playwright
import pyautogui

# ============================================================
# 설정
# ============================================================
CDP_URL = "http://localhost:9223"
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\User\.flow-chrome-debug"
ERP_LAYOUT = "http://erp-dev.samsong.com:19100/layout/layout.do"
D0_URL = "http://erp-dev.samsong.com:19100/prdtPlanMng/viewListDoAddnPrdtPlanInstrMngNew.do"
OAUTH_LOGIN = "http://auth-dev.samsong.com:18100/login"
PLAN_ROOT = r"Z:\15. SP3 메인 CAPA점검\SP3M3\생산지시서"
REPO_ROOT = Path(__file__).parent.parent.parent.parent
SKILL_DIR = Path(__file__).resolve().parent
STATE_DIR = SKILL_DIR / "state"
OUTER_LOCK_FILE = STATE_DIR / "sd9a01_outer.lock"
UPLOAD_DIR = REPO_ROOT / "06_생산관리" / "D0_업로드"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
SMARTMES_TOKEN = "bfee3f3d-caf9-434d-abbb-2cb015ec2469"
SMARTMES_BASE = "http://lmes-dev.samsong.com:19220"
DAY_CUT_THRESHOLD = 3600

# 세션151: phase6 SmartMES 검증 실패 누적. main 종료 시 0이 아니면 exit 코드 2 (verify_run에서 자동복구 진입).
PHASE6_FAILED = []  # [(line_cd, prod_date_str), ...]

# 세션153 A안 2단계: phase3 D0 업로드를 requests로 처리할 때 사용할 session.
# main에서 --http-upload 옵션 활성 + erp_login_via_http() 성공 시 세팅.
# d0_upload(page, xlsx) 호출 시 이 변수 있으면 d0_upload_via_http로 분기 (호출자 시그니처 무변경).
_HTTP_UPLOAD_SESS = None

# 세션153 A안 3단계: 완전 브라우저-less. --http-only 옵션 활성 시 세팅.
# phase0~6 전부 requests path. ensure_chrome_cdp + playwright 호출 0.
_HTTP_ONLY_SESS = None

LINE_CONFIG = {
    "SP3M3": {
        "line_div_cd": "02",
        "save_url": "/prdtPlanMng/multiListMainSubPrdtPlanRankDecideMng.do",
    },
    "SD9A01": {
        "line_div_cd": "01",
        "save_url": "/prdtPlanMng/multiListOuterPrdtPlanRankDecideMng.do",
    },
}

PBOM_GUARD_MAX_ITER = 20
PBOM_MISSING_RE = re.compile(
    r"P-?BOM\s*등록\s*안\s*됨\.\s*\[\s*([^,\]]+)\s*,\s*([A-Z0-9_./-]+)\s*\]"
)
_PBOM_GUARD_PHASE1_LOGGED = set()


def _uniq_prod_nos(items):
    seen = set()
    prod_nos = []
    for it in items or []:
        pno = str(it.get("PROD_NO") or "").strip()
        if pno and pno not in seen:
            seen.add(pno)
            prod_nos.append(pno)
    return prod_nos


def parse_prod_no_csv(value):
    return {p.strip() for p in (value or "").split(",") if p.strip()}


def apply_manual_exclude(items, exclude_csv, label="exclude"):
    exc = parse_prod_no_csv(exclude_csv)
    if not exc:
        return items
    before = len(items)
    kept = [it for it in items if it.get("PROD_NO") not in exc]
    removed = before - len(kept)
    print(f"[exclude] {label}: {sorted(exc)} 제외 {before}->{len(kept)}건 (removed={removed})")
    return kept


def note_pbom_guard_phase1(items, line, no_pbom_guard=False):
    prod_nos = _uniq_prod_nos(items)
    if no_pbom_guard:
        print(f"[pbom-guard] disabled (--no-pbom-guard): line={line} prod_no={len(prod_nos)}건")
        return items
    key = (line, tuple(prod_nos))
    if key not in _PBOM_GUARD_PHASE1_LOGGED:
        _PBOM_GUARD_PHASE1_LOGGED.add(key)
        print(
            f"[pbom-guard] line={line} PROD_NO {len(prod_nos)}건 수집. "
            "전용 P-BOM 조회 endpoint 미확정 -> Phase5 sendMesFlag=N preflight로 MES 전송 전 차단"
        )
    return items


def filter_items_by_prod_nos(items, prod_nos):
    blocked = set(prod_nos or [])
    if not blocked:
        return items
    kept = [it for it in items if it.get("PROD_NO") not in blocked]
    print(f"[pbom-guard] items exclude {sorted(blocked)}: {len(items)}->{len(kept)}건")
    return kept


def _flatten_mes_text(value):
    if value is None:
        return []
    if isinstance(value, str):
        texts = [value]
        s = value.strip()
        if s.startswith("{") or s.startswith("["):
            try:
                texts.extend(_flatten_mes_text(json.loads(s)))
            except Exception:
                pass
        return texts
    if isinstance(value, dict):
        texts = []
        for v in value.values():
            texts.extend(_flatten_mes_text(v))
        return texts
    if isinstance(value, list):
        texts = []
        for v in value:
            texts.extend(_flatten_mes_text(v))
        return texts
    return [str(value)]


def parse_pbom_missing_from_mes_msg(mes_msg, line_cd=None):
    missing = []
    seen = set()
    for text in _flatten_mes_text(mes_msg):
        for m in PBOM_MISSING_RE.finditer(text):
            line = (m.group(1) or "").strip()
            pno = (m.group(2) or "").strip()
            if line_cd and line and line != line_cd:
                continue
            if pno and pno not in seen:
                seen.add(pno)
                missing.append(pno)
    return missing


def parse_mes_status_code(mes_msg):
    for text in _flatten_mes_text(mes_msg):
        s = text.strip()
        if not (s.startswith("{") or s.startswith("[")):
            continue
        try:
            obj = json.loads(s)
        except Exception:
            continue
        stack = [obj]
        while stack:
            cur = stack.pop()
            if isinstance(cur, dict):
                if "statusCode" in cur:
                    return cur.get("statusCode")
                stack.extend(cur.values())
            elif isinstance(cur, list):
                stack.extend(cur)
    return None


def assert_mes_msg_ok(res, context):
    status = (res or {}).get("statusCode")
    if status not in (None, 200, "200"):
        raise RuntimeError(f"{context}: statusCode={status} body={str(res)[:300]}")
    mes_msg = (res or {}).get("mesMsg") or ""
    if not mes_msg:
        return
    mes_status = parse_mes_status_code(mes_msg)
    if mes_status in (200, "200"):
        return
    raise RuntimeError(f"{context}: MES statusCode={mes_status} mesMsg={str(mes_msg)[:300]}")


# ============================================================
# Phase 0: 환경 준비
# ============================================================
def _suppress_chrome_crash_restore(profile_dir: str):
    """Chrome 비정상 종료 후 "페이지 복원" 알림 차단.

    세션110 보강: taskkill로 강제 종료된 Chrome을 다시 launch하면 Preferences의
    exit_type이 "Crashed"로 남아 자동 복원 다이얼로그 표시 → 무인 트리거에서 화면 거슬림.
    Preferences 파일의 exit_type / exited_cleanly 정리로 복원 알림 차단.
    """
    import json as _json
    from pathlib import Path as _Path
    prefs_path = _Path(profile_dir) / "Default" / "Preferences"
    if not prefs_path.exists():
        return
    try:
        data = _json.loads(prefs_path.read_text(encoding="utf-8"))
        prof = data.setdefault("profile", {})
        prof["exit_type"] = "Normal"
        prof["exited_cleanly"] = True
        prefs_path.write_text(_json.dumps(data, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"[phase0] Preferences 정리 실패 (무시): {e}")


def _kill_zombie_chrome(profile_dir: str):
    """자동화 프로필 Chrome 좀비만 선별 종료 + 잠금파일 정리.

    세션151 보강: 5/9(토) morning 실패 분석 결과.
    같은 user-data-dir로 Chrome 좀비가 살아있으면 새 launch가 single-instance 규칙으로
    좀비에 URL만 위임 → 9223 listen은 부활 안 함 → 화면엔 ERP 로그인 페이지만 뜨고 Python timeout.

    프로필 디렉토리 매칭(`--user-data-dir=...`)으로 자동화 Chrome만 식별하여
    사용자 일상 Chrome(다른 프로필)은 건드리지 않는다.
    """
    try:
        # CommandLine에 user-data-dir 매칭. profile_dir의 백슬래시는 PowerShell -like 패턴에서
        # 그대로 사용 가능 (와일드카드 *로 둘러싸므로 escape 불필요).
        ps_cmd = (
            "Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
            f"Where-Object {{ $_.CommandLine -like '*--user-data-dir={profile_dir}*' }} | "
            "Select-Object -ExpandProperty ProcessId"
        )
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True, timeout=10
        )
        pids = [p.strip() for p in result.stdout.splitlines() if p.strip().isdigit()]
        if pids:
            print(f"[phase0] zombie chrome 발견 (프로필 매칭): pids={pids} — 종료")
            for pid in pids:
                subprocess.run(["taskkill", "/F", "/PID", pid], capture_output=True, timeout=5)
            time.sleep(1.5)  # 종료 안정화 + 파일 핸들 release
        else:
            print("[phase0] zombie chrome 없음")
        # 잠금파일 잔재 정리 (Chrome 비정상 종료 후 SingletonLock 잔존 시 신규 인스턴스 차단됨)
        from pathlib import Path as _Path
        for lock_name in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
            lock_path = _Path(profile_dir) / lock_name
            try:
                if lock_path.exists() or lock_path.is_symlink():
                    lock_path.unlink()
                    print(f"[phase0] {lock_name} 정리")
            except Exception:
                pass
    except Exception as e:
        print(f"[phase0] zombie chrome 정리 실패 (무시): {e}")


def ensure_chrome_cdp():
    """CDP 9223 기동 확인."""
    import urllib.request
    try:
        urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=3)
        print("[phase0] CDP 9223 alive")
        return True
    except Exception:
        print("[phase0] CDP dead — launching Chrome")

    # 세션151: single-instance 위임 함정 차단 — 같은 프로필 좀비 정리 후 fresh launch
    _kill_zombie_chrome(CHROME_PROFILE)
    # Chrome 비정상 종료 잔재 정리 (세션110 — 복원 알림 차단)
    _suppress_chrome_crash_restore(CHROME_PROFILE)

    subprocess.Popen([
        CHROME_PATH,
        f"--remote-debugging-port=9223",
        f"--remote-debugging-address=127.0.0.1",  # 세션105 — IPv6 기본 바인딩 회피
        f"--user-data-dir={CHROME_PROFILE}",
        "--no-first-run", "--no-default-browser-check",
        "--disable-session-crashed-bubble",  # 세션110 — 복원 다이얼로그 차단
        "--hide-crash-restore-bubble",
        "--no-default-browser-check",
        ERP_LAYOUT,
    ])
    for i in range(10):
        time.sleep(1.5)
        try:
            urllib.request.urlopen(f"{CDP_URL}/json/version", timeout=2)
            print(f"[phase0] CDP up (try={i+1})")
            return True
        except Exception:
            continue
    raise RuntimeError("CDP 9223 기동 실패")


def _force_chrome_foreground():
    """Chrome window 강제 OS-foreground (pyautogui 입력 가로채기 방지).

    세션151 보강 (사용자 5/11 재지적 — 세션141 본질 진단 미반영분 흡수):
    page.bring_to_front()는 Chrome 내부 탭 활성화만, Windows OS 창 순서는 별개.
    자동 실행 시점에 다른 앱 창이 Chrome 위에 있으면 pyautogui.click/press가
    그 위 창에 떨어져 빈값 submit → `/login?error`.

    매칭 우선순위 (좁은 매칭 우선 — 카카오톡/슬랙 등 Chromium 기반 앱 오매칭 차단):
    1. CHROME_PROFILE(`--user-data-dir`)로 launch된 chrome.exe PID 매칭 (정확)
    2. PID 수집 실패 시 erp-dev/auth-dev/samsong URL 키워드 매칭 (보조)
    3. 둘 다 실패 → False (fallback 없음 — 카톡/슬랙에 끌려가지 않도록)

    Windows 한정. Linux/Mac에선 ctypes.windll ImportError → 무시.
    """
    try:
        import ctypes
        from ctypes import wintypes, byref
        user32 = ctypes.windll.user32

        # 1. EnumWindows로 Chrome_WidgetWin 클래스 + window의 PID 수집
        candidates = []  # [(hwnd, title, pid), ...]
        def cb(hwnd, _lparam):
            if not user32.IsWindowVisible(hwnd):
                return True
            cls = ctypes.create_unicode_buffer(256)
            user32.GetClassNameW(hwnd, cls, 256)
            if 'Chrome_WidgetWin' not in cls.value:
                return True
            title = ctypes.create_unicode_buffer(512)
            user32.GetWindowTextW(hwnd, title, 512)
            if not title.value:
                return True
            pid = wintypes.DWORD()
            user32.GetWindowThreadProcessId(hwnd, byref(pid))
            candidates.append((hwnd, title.value, pid.value))
            return True
        EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
        user32.EnumWindows(EnumWindowsProc(cb), 0)

        if not candidates:
            print("[phase0] chrome window 없음 — foreground 강제 안 함")
            return False

        # 2. 자동화 프로필 chrome.exe PID set 수집
        automation_pids = set()
        try:
            ps_cmd = (
                "Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
                f"Where-Object {{ $_.CommandLine -like '*--user-data-dir={CHROME_PROFILE}*' }} | "
                "Select-Object -ExpandProperty ProcessId"
            )
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", ps_cmd],
                capture_output=True, text=True, timeout=10
            )
            automation_pids = {int(p.strip()) for p in result.stdout.splitlines() if p.strip().isdigit()}
        except Exception as e:
            print(f"[phase0] 자동화 PID 수집 실패: {e}")

        # 3. PID 매칭 우선 (정확)
        if automation_pids:
            for hwnd, t, pid in candidates:
                if pid in automation_pids:
                    user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                    user32.SetForegroundWindow(hwnd)
                    print(f"[phase0] chrome OS-foreground (자동화 PID={pid}): {t[:60]}")
                    return True
            print(f"[phase0] 자동화 chrome PID 매칭 윈도우 없음 (pids={automation_pids})")

        # 4. URL 키워드 매칭 (보조 — PID 수집 실패 시만)
        for hwnd, t, _ in candidates:
            tl = t.lower()
            if 'erp-dev' in tl or 'auth-dev' in tl or 'samsong' in tl:
                user32.ShowWindow(hwnd, 9)
                user32.SetForegroundWindow(hwnd)
                print(f"[phase0] chrome OS-foreground (URL 키워드): {t[:60]}")
                return True

        # 5. 둘 다 실패 — 카톡/슬랙 등 오매칭 차단 위해 fallback 제거
        sample_titles = [t[:30] for _, t, _ in candidates[:3]]
        print(f"[phase0] 자동화 chrome 윈도우 미발견 — foreground 강제 안 함 (titles: {sample_titles})")
        return False
    except Exception as e:
        print(f"[phase0] foreground 강제 예외 (무시): {e}")
        return False


def _load_oauth_cred():
    """OAuth ID/PW 로컬 credential 파일 로드.

    세션153: pyautogui typewrite + Chrome 자동완성 의존 폐기. ID/PW를 .oauth.json에
    저장하고 page.fill()로 DOM 직접 입력 — OS focus 가로채기 위험 0%.
    """
    cred_path = Path(__file__).parent / ".oauth.json"
    if not cred_path.exists():
        raise RuntimeError(f".oauth.json 미존재: {cred_path} — ID/PW 파일 저장 후 재실행")
    cfg = json.loads(cred_path.read_text(encoding="utf-8"))
    if not cfg.get("id") or not cfg.get("pw"):
        raise RuntimeError(f".oauth.json id/pw 누락: {cred_path}")
    return cfg["id"], cfg["pw"]


def erp_login_via_http():
    """ERP OAuth SSO를 requests로 직접 처리 — 브라우저·CDP·pyautogui 의존 0.

    세션153 (2026-05-13) — A안 점진 전환 1단계.
    daily-routine mes_login() 패턴의 ERP 버전. 실측 캡처 PASS:
      r1 = GET erp-dev/oauth2/sso → ssoUrl 변수 추출
      r2 = GET ssoUrl (auth-dev/oauth/authorize, JSESSIONID 발급)
      r3 = POST auth-dev/login (userId/password/clientId=ERP/ssoUrl)
           → redirect layout.do?statusCode=200+OK
      r4 = GET erp-dev/layout/layout.do → XSRF-TOKEN 발급

    반환: requests.Session (cookies + X-XSRF-TOKEN header 세팅 완료). 실패 시 None.
    가로채기 시나리오 자체 0%. 가장 견고.
    """
    import requests as _req
    user_id, password = _load_oauth_cred()
    s = _req.Session()
    s.headers["User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    try:
        r1 = s.get("http://erp-dev.samsong.com:19100/oauth2/sso", timeout=10)
        m = re.search(r"ssoUrl\s*=\s*'([^']+)'", r1.text)
        if not m:
            print("[phase0:http] FAIL ssoUrl 파싱 실패")
            return None
        sso_url = m.group(1)
        s.get(sso_url, allow_redirects=True, timeout=10)
        r3 = s.post("http://auth-dev.samsong.com:18100/login", data={
            "userId": user_id, "password": password,
            "clientId": "ERP", "ssoUrl": sso_url, "clientName": "", "lang": "ko"
        }, allow_redirects=True, timeout=15)
        if "layout.do" not in r3.url:
            print(f"[phase0:http] FAIL 로그인 후 URL={r3.url}")
            return None
        s.get("http://erp-dev.samsong.com:19100/layout/layout.do", timeout=10)
        s.headers["X-XSRF-TOKEN"] = s.cookies.get("XSRF-TOKEN", "")
        print(f"[phase0:http] OAuth PASS (cookies: {list(s.cookies.keys())})")
        return s
    except Exception as e:
        print(f"[phase0:http] FAIL {e}")
        return None


def _inject_cookies_to_playwright(context, sess):
    """requests.Session cookie를 playwright context에 주입.

    HTTP OAuth로 받은 SESSION/XSRF-TOKEN/JSESSIONID cookie를 playwright context에
    추가해서, page.goto(D0_URL) 시 OAuth 페이지 안 거치고 즉시 통과되게 함.
    """
    cookies_for_pw = []
    for c in sess.cookies:
        domain = c.domain or "erp-dev.samsong.com"
        if domain.startswith("."):
            domain = domain
        cookies_for_pw.append({
            "name": c.name,
            "value": c.value,
            "domain": domain,
            "path": c.path or "/",
        })
    if cookies_for_pw:
        context.add_cookies(cookies_for_pw)
        print(f"[phase0:http] playwright cookie 주입 {len(cookies_for_pw)}건 ({[c['name'] for c in cookies_for_pw]})")


def ensure_erp_login(page):
    """auth-dev 로그인 페이지면 자동 로그인 — DOM 직접 입력 (page.fill).

    세션153 (2026-05-13): pyautogui 키보드 입력 + Chrome 비번관리자 자동완성 의존 폐기.
    실측 사고 — 다른 창이 Chrome 위에 떠 있으면 pyautogui 키가 가로채여 빈 form submit
    → /login?error 반복 (5/13 07:11~07:42 morning 자동화 51min 한계 도달 실패).
    page.fill()은 DOM API라 OS focus 무관 → 가로채기 불가능.

    ID/PW 원본: 같은 폴더 .oauth.json (gitignore 처리).
    """
    if "auth-dev.samsong.com" in page.url and "/login" in page.url:
        print("[phase0] OAuth 로그인 수행 (DOM 직접 입력)")
        user_id, password = _load_oauth_cred()
        page.bring_to_front()
        page.wait_for_selector('#userId', timeout=10000)
        page.fill('#userId', user_id)
        page.fill('#password', password)
        page.click('#loginBtn')
        time.sleep(3)
        if "/login?error" in page.url:
            print(f"[phase0] ⚠ submit 직후 /login?error — .oauth.json id/pw 검증 필요. URL: {page.url}")
            # raise 안 하고 caller(navigate_to_d0)의 재시도 분기에서 1회 재시도 흐름 유지


def _wait_oauth_complete(page, timeout_sec: float = 10.0):
    """OAuth 완료 대기 — auth-dev 떠나고 erp-dev 본 페이지(oauth2/sso 콜백 제외) 도달까지.

    세션110 보강: 기존 `"erp-dev.samsong.com" in url` 조건은 OAuth 콜백 중간 단계
    `oauth2/sso` URL도 매칭되어 잘못 break 발생 → 미완료 상태에서 D0_URL 시도 →
    auth-dev/login 페이지 redirect → btnExcelUpload timeout 실패 시나리오.

    세션131 [E] 보강: default 30→60s. 5일 중 4일 morning 자동화 OAuth timeout 실패
    실측 (4/27 4/29 4/30). 7시 ERP 부하 + cold start fresh launch에서 OAuth 콜백 redirect
    chain이 30s를 넘어가는 케이스 흡수.

    세션152: 60s → 10s. ID typewrite 명시 입력으로 자동완성 실패 케이스 차단된 후
    OAuth 콜백 정체 시 빠른 fallback 진입 우선. 사용자 명시 변경 (체감 대기 단축).
    """
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        url = page.url
        if ("erp-dev.samsong.com" in url
                and "auth-dev" not in url
                and "oauth2/sso" not in url
                and "/login" not in url):
            return True
        time.sleep(0.5)
    return False


def navigate_to_d0(browser, sess=None):
    """D0추가생산지시 화면 탭 확보.

    세션153: sess 인자 추가 — HTTP OAuth로 받은 requests.Session의 cookie를
    playwright context에 주입해서 ensure_erp_login(브라우저 OAuth) 자체를 우회.
    sess=None이거나 cookie 주입 후에도 login 페이지면 기존 page.fill fallback.
    """
    ctx = browser.contexts[0]

    # HTTP OAuth 성공 시 cookie 주입 → 브라우저 OAuth 스킵
    if sess is not None:
        try:
            _inject_cookies_to_playwright(ctx, sess)
        except Exception as e:
            print(f"[phase0:http] cookie 주입 실패 (page.fill fallback 진입): {e}")

    page = None
    for pg in ctx.pages:
        if "viewListDoAddnPrdtPlanInstrMngNew" in pg.url:
            page = pg; break
    if page is None:
        for pg in ctx.pages:
            if pg.url.startswith("http"):
                page = pg; break
    if page is None:
        raise RuntimeError("사용 가능한 탭 없음")

    page.bring_to_front()
    ensure_erp_login(page)

    # OAuth 완료 명시 대기 (세션110 보강 — oauth2/sso 콜백 부분 매칭 버그 수정 / 세션152 60s→10s)
    if not _wait_oauth_complete(page, timeout_sec=10.0):
        print(f"[phase0] OAuth 완료 대기 10s 실패 — 현재 URL: {page.url}")
        # login 페이지 다시 도달 시 1회 재시도
        if "auth-dev.samsong.com" in page.url and "/login" in page.url:
            print("[phase0] login 페이지 재도달 — 재로그인 1회 시도")
            ensure_erp_login(page)
            if not _wait_oauth_complete(page, timeout_sec=10.0):
                raise RuntimeError(f"OAuth 완료 2회 실패: {page.url}")
        elif "auth-dev.samsong.com" in page.url:
            # 세션124 [3way] 보강: 비login auth-dev 정착(클라이언트 선택 화면 등) 시 D0_URL 직접 이동 1회 시도
            print("[phase0] auth-dev 비login 정착 — D0_URL 직접 이동 1회 시도")
            _safe_goto(page, D0_URL)
            if not _wait_oauth_complete(page, timeout_sec=10.0):
                raise RuntimeError(f"OAuth 완료 실패: auth-dev 클라이언트 선택 화면에서 D0_URL 직접 이동 1회 시도 후에도 erp-dev 미도달 — 현재 URL: {page.url}")
        else:
            # 세션131 [E]: OAuth 콜백 URL(oauth2/sso 등) 정체 — D0_URL 직접 navigate fallback
            # 사용자 실측 관찰(2026-04-30): 로그인 성공 후 생산계획 탭 redirect 못 받고 멍때리다 실패.
            # ERP 서버가 OAuth 콜백 후 redirect 누락 의심. 4/29 auth-dev 비login 분기와 동일 패턴 —
            # 클라이언트가 능동 navigate로 우회. 4/27·4/29·4/30 3건 동일 증상 흡수.
            print(f"[phase0] OAuth 콜백 정체 ({page.url}) — D0_URL 직접 이동 1회 시도")
            _safe_goto(page, D0_URL)
            if not _wait_oauth_complete(page, timeout_sec=10.0):
                raise RuntimeError(f"OAuth 완료 실패 (D0_URL 직접 이동 후에도 미완료): {page.url}")

    try: page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception: pass
    time.sleep(2)

    if "viewListDoAddnPrdtPlanInstrMngNew" not in page.url:
        _safe_goto(page, D0_URL)
        try: page.wait_for_load_state("domcontentloaded", timeout=20000)
        except Exception: pass
        time.sleep(3)
    print(f"[phase0] D0 화면 진입: {page.url}")
    return page


def _safe_goto(page, url, max_retries=3):
    """page.goto 실패 시 단계적 재시도 + ERR_BLOCKED_BY_CLIENT 우회.

    1차~N차: page.goto (timeout 30s, 간격 증가)
    최종 fallback: JavaScript window.location.href 설정 (CDP 레벨 차단 우회)
    """
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            page.goto(url, timeout=30000)
            return
        except Exception as e:
            last_err = e
            msg = str(e)
            print(f"[phase0] goto 시도 {attempt}/{max_retries} 실패: {msg[:120]}")
            if "ERR_BLOCKED_BY_CLIENT" in msg and attempt == max_retries:
                # 마지막: JS navigation 우회
                print("[phase0] ERR_BLOCKED_BY_CLIENT 감지 — JS window.location 우회 시도")
                try:
                    page.evaluate(f"() => {{ window.location.href = {url!r}; }}")
                    time.sleep(5)
                    page.wait_for_load_state("domcontentloaded", timeout=20000)
                    if "viewListDoAddnPrdtPlanInstrMngNew" in page.url:
                        print(f"[phase0] JS 우회 성공: {page.url}")
                        return
                except Exception as e2:
                    print(f"[phase0] JS 우회도 실패: {e2}")
            time.sleep(3 * attempt)
    raise RuntimeError(f"goto {url} 실패 (전체 {max_retries}회 + JS 우회): {last_err}")


# ============================================================
# Phase 1: 파일 해석
# ============================================================
def find_plan_file(target_date: datetime) -> Path:
    """target_date에 해당하는 SP3M3 생산지시서 xlsm 탐색.

    월 boundary 케이스(예: 5/1 파일이 4월 폴더에 저장됨) 대응:
    target 폴더 → 이전월 폴더 → 다음월 폴더 순으로 같은 date_tag 검색.
    target 폴더가 없으면 자동 생성(빈 폴더로 두지 않고, fallback에서 발견 시 복사).
    """
    date_tag = f"{target_date.year%100:02d}.{target_date.month:02d}.{target_date.day:02d}"
    pattern = f"SP3M3_생산지시서_({date_tag})*.xlsm"
    year_root = Path(PLAN_ROOT) / f"{target_date.year}년 생산지시"
    target_folder = year_root / f"{target_date.month:02d}월"

    # 1순위: target 폴더
    if target_folder.exists():
        candidates = list(target_folder.glob(pattern))
        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            chosen = candidates[0]
            print(f"[phase1] 파일 선택: {chosen.name}")
            return chosen

    # 2순위: 인접 월 fallback (이전월 → 다음월)
    fallback_months = []
    if target_date.month > 1:
        fallback_months.append(target_date.month - 1)
    else:
        fallback_months.append(12)  # 1월이면 전년 12월은 별도 처리 필요 — 일단 같은 year_root 한정
    if target_date.month < 12:
        fallback_months.append(target_date.month + 1)

    for m in fallback_months:
        fb_folder = year_root / f"{m:02d}월"
        if not fb_folder.exists():
            continue
        candidates = list(fb_folder.glob(pattern))
        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            src = candidates[0]
            # target 폴더 자동 생성 + 파일 복사 (원본 보존)
            target_folder.mkdir(parents=True, exist_ok=True)
            dst = target_folder / src.name
            if not dst.exists():
                import shutil
                shutil.copy2(src, dst)
                print(f"[phase1] fallback: {fb_folder.name}에서 {src.name} 발견 → {target_folder.name}/ 복사")
            else:
                print(f"[phase1] fallback: {fb_folder.name}에서 {src.name} 발견 (target에 이미 존재)")
            print(f"[phase1] 파일 선택: {dst.name}")
            return dst

    # 3순위: 모두 실패
    if not target_folder.exists():
        raise FileNotFoundError(f"폴더 없음: {target_folder} (인접 월 폴더에도 {pattern} 미발견)")
    raise FileNotFoundError(f"파일 없음: {target_folder}/{pattern}")


def _find_section_end(ws, start_row, next_header_keyword=None):
    """start_row부터 시작해서 다음 헤더 또는 빈 행 만날 때까지의 끝 행."""
    r = start_row
    last = start_row
    while r <= ws.max_row:
        b = ws.cell(row=r, column=2).value
        i = ws.cell(row=r, column=9).value
        k = ws.cell(row=r, column=11).value
        # 다음 섹션 헤더 만나면 중단
        if b and next_header_keyword and next_header_keyword in str(b):
            break
        if i and k:
            last = r
        r += 1
    return last


def extract_sp3m3_night(wb):
    """출력용 시트 야간 섹션.

    구조:
      R1: `◀ D+1 야간계획`
      R2: 데이터 헤더 (순서/신MES 품번/지시)
      R3~: 실제 데이터
      R?: `◀ D+2 주간계획` (종료 경계)
    """
    ws = wb["출력용"]
    start = 3
    night_end = 34  # 보수적 default
    for r in range(2, ws.max_row+1):
        v = ws.cell(row=r, column=2).value
        if v and "주간계획" in str(v):
            night_end = r - 1
            break
    items = []
    skipped = []
    for r in range(start, night_end+1):
        part = ws.cell(row=r, column=9).value
        qty = ws.cell(row=r, column=11).value
        if not (part and qty):
            continue
        try:
            qty_int = int(qty)
        except (ValueError, TypeError):
            continue  # 헤더 행 등 숫자 아닌 값 skip
        pno = str(part).strip()
        # 한글 포함 PROD_NO는 자리표시 문자열 ("구형바코드사용" 등) — ERP 라인배치 미등록 확정, skip
        if any("가" <= ch <= "힣" for ch in pno):
            skipped.append({"row": r, "PROD_NO": pno, "QTY": qty_int})
            print(f"[phase1:skip] r={r} PROD_NO={pno!r} QTY={qty_int} — 한글 포함, ERP 등록 불가")
        else:
            items.append({"PROD_NO": pno, "QTY": qty_int})
    if skipped:
        print(f"[phase1] SP3M3 야간: 원본 {len(items) + len(skipped)}건, 등록 {len(items)}건, 제외 {len(skipped)}건")
    else:
        print(f"[phase1] SP3M3 야간: {len(items)}건")
    return items


def extract_sp3m3_day(wb, cut_threshold=DAY_CUT_THRESHOLD):
    """출력용 시트 주간 섹션 — 누적 ≥ cut_threshold 첫 행까지 포함.

    구조 (2026-04-25 실증):
      R32: `◀ D+2 주간계획` 섹션 헤더
      R33: 데이터 헤더 (순서/신MES 품번/지시)
      R34~: 실제 데이터

    day_start = 주간계획 헤더 r + 2 (데이터 헤더 한 줄 skip).
    방어: 헤더 행이 추가로 있어도 try/except로 숫자 아닌 K값은 skip.
    """
    ws = wb["출력용"]
    day_start = None
    for r in range(2, ws.max_row+1):
        v = ws.cell(row=r, column=2).value
        if v and "주간계획" in str(v):
            day_start = r + 2  # 섹션 헤더 + 데이터 헤더 둘 다 skip
            break
    if day_start is None:
        raise ValueError("출력용 시트 주간 헤더 미발견")
    items = []
    skipped = []
    cumsum = 0
    for r in range(day_start, ws.max_row+1):
        part = ws.cell(row=r, column=9).value
        qty = ws.cell(row=r, column=11).value
        if not (part and qty):
            continue
        try:
            qty_int = int(qty)
        except (ValueError, TypeError):
            continue
        pno = str(part).strip()
        cumsum += qty_int
        # 한글 포함 PROD_NO는 자리표시 문자열 ("구형바코드사용" 등) — ERP 라인배치 미등록 확정, skip
        if any('가' <= ch <= '힯' for ch in pno):
            skipped.append({"row": r, "PROD_NO": pno, "QTY": qty_int})
            print(f"[phase1:skip] r={r} PROD_NO={pno!r} QTY={qty_int} — 한글 포함, ERP 등록 불가")
        else:
            items.append({"PROD_NO": pno, "QTY": qty_int})
        if cumsum >= cut_threshold:
            msg = f"[phase1] SP3M3 주간 컷: 누적 {cumsum} (등록 {len(items)}건"
            if skipped:
                msg += f", 제외 {len(skipped)}건"
            msg += ")"
            print(msg)
            break
    else:
        msg = f"[phase1] SP3M3 주간: 누적 {cumsum} (컷 미도달, 등록 {len(items)}건"
        if skipped:
            msg += f", 제외 {len(skipped)}건"
        msg += ")"
        print(msg)
    return items


def extract_outer_d1(wb, target_line="SD9M01"):
    """OUTER 생산계획 시트 D+1 블록에서 B열 == target_line 필터."""
    ws = wb["OUTER 생산계획"]
    items = []
    for r in range(2, ws.max_row+1):
        b = ws.cell(row=r, column=2).value  # D+1 라인
        d = ws.cell(row=r, column=4).value  # 품번 (접미사 포함)
        e = ws.cell(row=r, column=5).value  # 수량
        if b and str(b).strip() == target_line and d and e:
            items.append({"PROD_NO": str(d).strip(), "QTY": int(e)})
    print(f"[phase1] SD9A01 OUTER D+1: {len(items)}건")
    return items


# ============================================================
# Phase 2: 업로드 엑셀 생성
# ============================================================
def _load_items_from_xlsx(xlsx_path: Path):
    """외부 엑셀 파일에서 (생산일/제품번호/생산량) 추출."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1] and row[2]:
            items.append({"PROD_NO": str(row[1]).strip(), "QTY": int(row[2])})
    return items


def _extract_prod_date(items):
    """items는 PROD_NO/QTY만 — 외부 엑셀 원본에서 생산일 재추출."""
    return None  # caller에서 fallback


def make_upload_xlsx(items, prod_date: datetime, out_path: Path):
    """ERP 양식을 Excel COM(win32com)으로 편집·저장.

    왜 win32com인가:
      openpyxl이 생성·저장한 xlsx는 ERP 서버 파서(아마 Apache POI 구버전)가
      COL2(제품번호)를 빈값으로 파싱하는 이슈 있음 (2026-04-24 세션104 실증).
      OOXML 내부 구조(shared strings, cell type 속성, 시트 XML 네임스페이스)
      차이 때문이며, Microsoft Excel이 직접 저장한 파일만 서버가 정상 파싱.
      → Excel COM으로 실제 Excel 프로세스가 저장해야 호환.

    흐름:
      1. 템플릿(ERP 양식 또는 SSKR 검증본) 복제
      2. Excel COM 열어 R2~last 데이터 ClearContents
      3. 새 데이터 R2부터 작성
      4. Excel이 Save (Visible=False, DisplayAlerts=False로 조용히)
    """
    import shutil
    template = Path(__file__).parent / "template" / "SSKR_D0_template.xlsx"
    if not template.exists():
        raise FileNotFoundError(f"ERP 양식 템플릿 없음: {template}")
    shutil.copy(template, out_path)

    try:
        import win32com.client
    except ImportError:
        raise RuntimeError("win32com 필요 (pywin32 설치 필요). pip install pywin32")

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(out_path))
        try:
            ws = wb.Worksheets("Sheet1")
        except Exception:
            ws = wb.Worksheets(1)
        # 기존 데이터 clear (헤더 R1 유지)
        last = ws.Cells(ws.Rows.Count, 1).End(-4162).Row  # xlUp
        if last > 1:
            ws.Range(f"A2:C{last}").ClearContents()
        # 새 데이터 R2부터 작성
        date_str = prod_date.strftime("%Y-%m-%d")
        for i, it in enumerate(items, start=2):
            ws.Cells(i, 1).Value = date_str
            ws.Cells(i, 2).Value = it["PROD_NO"]
            ws.Cells(i, 3).Value = it["QTY"]
        wb.Save()
        wb.Close()
    finally:
        xl.Quit()
    print(f"[phase2] 업로드 엑셀 생성: {out_path.name} ({len(items)}건) — Excel COM 저장 (서버 파서 호환)")


# ============================================================
# Phase 3: D0 업로드 (selectList + multiList)
# ============================================================
def _wait_d0_popup_frame(page, timeout_sec: float = 25.0):
    """D0 업로드 팝업 iframe 폴링 (frame URL 매칭 + DOM querySelector 보강).

    iframe[name="ifr_user"]는 즉시 생성되지만 src URL이 about:blank → popupPmD0AddnUpload.do
    로 늦게 set되는 경우가 있어, frame URL 매칭만으로는 놓칠 수 있다 (세션107 실증).
    DOM 직접 조회를 보조 신호로 사용해 frame 객체가 반영될 때까지 추가 대기.
    """
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        for fr in page.frames:
            if "popupPmD0AddnUpload" in fr.url:
                return fr
        try:
            has_iframe = page.evaluate(
                "() => !!document.querySelector('iframe[src*=\"popupPmD0AddnUpload\"]')"
            )
            if has_iframe:
                time.sleep(0.5)
                for fr in page.frames:
                    if "popupPmD0AddnUpload" in fr.url:
                        return fr
        except Exception:
            pass
        time.sleep(0.3)
    return None


def d0_upload(page, xlsx_path: Path, parse_only: bool = False):
    """팝업 오픈 → selectList → multiList.

    세션153 A안 2단계: _HTTP_UPLOAD_SESS 세팅 시 d0_upload_via_http로 분기.
    호출자 시그니처 무변경 (page 인자 무시).

    parse_only=True: selectList(엑셀 첨부 + 서버 파싱 + 화면 표시)까지만 수행하고 multiList(DB 저장) 스킵.
    검증 모드 전용 (Phase 4/5 미진행)."""
    if _HTTP_UPLOAD_SESS is not None:
        d0_upload_via_http(_HTTP_UPLOAD_SESS, xlsx_path, parse_only=parse_only)
        # 세션153 A안 2단계 보강: HTTP 업로드 후 브라우저 grid 갱신 (phase4가 grid 데이터 사용).
        # 미호출 시 api_rank_batch가 신규 PROD_NO를 grid에서 못 찾아 매칭 실패.
        if not parse_only:
            try:
                page.evaluate("() => { if (typeof totGridList !== 'undefined' && totGridList.searchListData) totGridList.searchListData(); }")
                time.sleep(3)
                print("[phase3:http] 브라우저 grid 갱신 완료")
            except Exception as e:
                print(f"[phase3:http] grid 갱신 경고 (무시): {e}")
        return

    # 페이지 로드 확인 + 버튼 대기 (세션107 Gemini 지적: 15s → 30s, ERP 초기 로드 지연 대응)
    try:
        page.wait_for_selector("#btnExcelUpload", timeout=30000)
    except Exception:
        print(f"[phase3] #btnExcelUpload 대기 실패 (30s timeout) — 현재 URL: {page.url}")
        raise

    # 팝업 iframe 먼저 검사 (이미 열려있으면 재사용, reload 생략)
    ifr = None
    for fr in page.frames:
        if "popupPmD0AddnUpload" in fr.url:
            ifr = fr; break

    if ifr is not None:
        print(f"[phase3] 기존 팝업 재사용: {ifr.url[:80]}")
    else:
        # 팝업 없음 — overlay 잔재만 있으면 reload 후 재오픈
        has_overlay = page.evaluate("""() => {
            return document.querySelectorAll('.ui-widget-overlay').length > 0
                || Array.from(document.querySelectorAll('.ui-dialog')).some(d => d.offsetParent !== null);
        }""")
        if has_overlay:
            print("[phase3] 팝업 없이 overlay만 잔존 — 페이지 reload")
            try:
                page.reload(timeout=30000, wait_until="domcontentloaded")
                page.wait_for_selector("#btnExcelUpload", timeout=30000)
                time.sleep(3)
            except Exception as e:
                print(f"[phase3] reload err: {e}")
    if ifr is None:
        # 클릭 재시도 (최대 3번) — 폴링은 _wait_d0_popup_frame (25초, DOM 보강)
        for attempt in range(3):
            try:
                page.locator("#btnExcelUpload").first.click()
            except Exception as e:
                print(f"[phase3] 클릭 실패 try={attempt+1}: {e}")
                time.sleep(2)
                continue
            ifr = _wait_d0_popup_frame(page, timeout_sec=25.0)
            if ifr: break
            print(f"[phase3] popup 미등장 try={attempt+1} — 재시도")
            time.sleep(2)
    if ifr is None:
        raise RuntimeError("Excel Upload 팝업 확보 실패 (3회 재시도 실패)")
    try:
        ifr.wait_for_load_state("domcontentloaded", timeout=8000)
        ifr.wait_for_selector("#uploadfrm", timeout=8000)
    except Exception: pass
    print(f"[phase3] popup open: {ifr.url}")

    with open(xlsx_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")

    # selectList
    parse_js = """
    async ({b64, filename}) => {
        const bin = atob(b64);
        const bytes = new Uint8Array(bin.length);
        for (let i=0;i<bin.length;i++) bytes[i]=bin.charCodeAt(i);
        const blob = new Blob([bytes], {type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
        const file = new File([blob], filename, {type: blob.type});
        const form = document.getElementById('uploadfrm');
        const fd = new FormData(form);
        fd.append('files', file);
        const out = await new Promise((resolve) => {
            const to = setTimeout(() => resolve({ok:false, err:'timeout 120s'}), 120000);
            jQuery.ajax({
                type:'POST', url:'/prdtPlanMng/selectListPmD0AddnUpload.do',
                data:fd, processData:false, contentType:false, cache:false,
                success:(res) => { clearTimeout(to); resolve({ok:true, res}); },
                error:(xhr,s,e) => { clearTimeout(to); resolve({ok:false, httpStatus:xhr.status, err:String(e), body:(xhr.responseText||'').slice(0,400)}); }
            });
        });
        if (!out.ok) return out;
        const res = out.res;
        if (res && res.data && res.data.list && typeof popupPmD0AddnUpload !== 'undefined') {
            popupPmD0AddnUpload.excelUploadCallBack(res);
        }
        return {ok:true, listLen: res && res.data && res.data.list ? res.data.list.length : null, statusCode: res.statusCode, statusTxt: res.statusTxt};
    }
    """
    r = ifr.evaluate(parse_js, {"b64": b64, "filename": xlsx_path.name})
    print(f"[phase3 parse] {r}")
    if not r.get("ok") or not r.get("listLen"):
        raise RuntimeError(f"selectList 실패: {r}")

    # 오류 행 검증
    err_check = ifr.evaluate("() => { try { return popupPmD0AddnUpload.$local_grid.getData().filter(r => r.ERROR_FLAG).length; } catch(e){ return -1; } }")
    if err_check > 0:
        raise RuntimeError(f"오류 행 {err_check}건 존재")

    if parse_only:
        print("[phase3 parse-only] selectList 완료 — multiList(DB 저장) 스킵 + 팝업 유지 (사용자 확인용)")
        return

    # multiList (DB 저장)
    save_js = """
    async () => {
        const dataList = popupPmD0AddnUpload.$local_grid.getData();
        const param = {excelList: dataList, ADDN_PRDT_REASON_CD: document.getElementById('ADDN_PRDT_REASON_CD').value};
        return await new Promise((resolve) => {
            const to = setTimeout(() => resolve({ok:false, err:'timeout 30s'}), 30000);
            jQuery.ajax({
                url: '/prdtPlanMng/multiListPmD0AddnUpload.do', type: 'post',
                data: JSON.stringify(param),
                success: (r) => { clearTimeout(to); resolve({ok:true, r}); },
                error: (xhr,s,e) => { clearTimeout(to); resolve({ok:false, httpStatus:xhr.status, body:(xhr.responseText||'').slice(0,400)}); }
            });
        });
    }
    """
    r2 = ifr.evaluate(save_js)
    print(f"[phase3 save] {r2}")
    if not r2.get("ok") or str((r2.get("r") or {}).get("statusCode")) != "200":
        raise RuntimeError(f"multiListPmD0AddnUpload 실패: {r2}")

    # 본 그리드 갱신 + 팝업 닫기
    try:
        ifr.evaluate("""() => {
            try {
                if (parent.totGridList && parent.totGridList.searchListData) parent.totGridList.searchListData();
                if (parent.$ && parent.$('#popupPmD0AddnUpload').dialog) parent.$('#popupPmD0AddnUpload').dialog('close');
            } catch(e){}
        }""")
    except Exception: pass
    time.sleep(3)
    print("[phase3] D0 업로드 완료")


def d0_upload_via_http(sess, xlsx_path: Path, parse_only: bool = False):
    """Phase 3 D0 업로드 — requests 직접 (브라우저·iframe·jQuery 의존 0).

    세션153 A안 2단계 — daily-routine HTTP POST 패턴 ERP 확장.
    실측 캡처 PASS (2026-05-13):
      1) GET popupPmD0AddnUpload.do?callbackFunid=totGridList — XSRF 갱신
      2) POST selectListPmD0AddnUpload.do multipart (files + hidden 3종)
         → {data:{list:[...]}, statusCode:200}
      3) POST multiListPmD0AddnUpload.do JSON ({excelList, ADDN_PRDT_REASON_CD})
         → {statusCode:200}

    핵심 — X-XSRF-TOKEN 헤더는 매 POST 직전 최신 cookie['XSRF-TOKEN']로 갱신해야 함.
    cookie 갱신 시점과 header 시점 불일치 시 ERP 500.

    parse_only=True: selectList만 (multiList=DB 저장 스킵). 검증 전용.
    """
    import requests as _req  # noqa
    POPUP_URL = "http://erp-dev.samsong.com:19100/prdtPlanMng/popupPmD0AddnUpload.do?callbackFunid=totGridList"
    SELECT_URL = "http://erp-dev.samsong.com:19100/prdtPlanMng/selectListPmD0AddnUpload.do"
    MULTI_URL = "http://erp-dev.samsong.com:19100/prdtPlanMng/multiListPmD0AddnUpload.do"

    # 1) popup GET — referer/XSRF 갱신
    sess.get(POPUP_URL, timeout=10)
    sess.headers["Referer"] = POPUP_URL
    sess.headers["X-Requested-With"] = "XMLHttpRequest"

    with open(xlsx_path, "rb") as f:
        file_bytes = f.read()

    # 2) selectList (파일 파싱)
    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    files = {"files": (xlsx_path.name, file_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"callbackFunid": "totGridList", "dataList": "", "ADDN_PRDT_REASON_CD": ""}
    r1 = sess.post(SELECT_URL, files=files, data=data, timeout=120)
    res1 = r1.json() if r1.headers.get("content-type","").startswith("application/json") else {}
    print(f"[phase3:http parse] status={r1.status_code} statusCode={res1.get('statusCode')} listLen={len(res1.get('data',{}).get('list',[]) or [])}")
    if r1.status_code != 200 or str(res1.get("statusCode")) != "200":
        raise RuntimeError(f"selectListPmD0AddnUpload 실패: status={r1.status_code} body={r1.text[:400]}")
    list_data = res1["data"]["list"]
    if not list_data:
        raise RuntimeError(f"selectList 응답 list 빈값: {r1.text[:400]}")
    err_rows = [row for row in list_data if row.get("ERROR_FLAG")]
    if err_rows:
        raise RuntimeError(f"오류 행 {len(err_rows)}건 존재: {err_rows[:3]}")

    if parse_only:
        print(f"[phase3:http parse-only] selectList 완료 ({len(list_data)}건) — multiList 스킵")
        return

    # 3) multiList (DB 저장)
    # ADDN_PRDT_REASON_CD = "002" (신규추가수량) — 2026-05-13 grid row 실측 캡처값.
    # 빈 값으로 보내면 등록사유 칸 비어 있는 채로 ERP DB 등록됨 (사용자 실측 지적).
    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    payload = {"excelList": list_data, "ADDN_PRDT_REASON_CD": "002"}
    r2 = sess.post(MULTI_URL, data=json.dumps(payload),
                   headers={"Content-Type": "application/json; charset=utf-8",
                            "X-XSRF-TOKEN": sess.cookies.get("XSRF-TOKEN", "")},
                   timeout=30)
    res2 = r2.json() if r2.headers.get("content-type","").startswith("application/json") else {}
    print(f"[phase3:http save] status={r2.status_code} statusCode={res2.get('statusCode')} statusTxt={res2.get('statusTxt')}")
    if r2.status_code != 200 or str(res2.get("statusCode")) != "200":
        raise RuntimeError(f"multiListPmD0AddnUpload 실패: status={r2.status_code} body={r2.text[:400]}")
    print(f"[phase3:http] D0 업로드 완료 ({len(list_data)}건)")


# ============================================================
# Phase 4: 야간 서열 배치
# ============================================================
def process_one_row(page, prod_no, target_ext_reg, target_line, save_url, dry_run=False):
    """한 품번 처리: 상단 클릭 → m 선택 → s 로드 → addRow → multiList 임시저장."""
    # 1. rowClick 로직 재현
    click_r = page.evaluate("""(args) => {
        const d = jQuery('#grid_body').pqGrid('option','dataModel').data;
        const idx = d.findIndex(r => r.PROD_NO === args.prodNo && String(r.REG_NO) === String(args.extReg));
        if (idx < 0) return {ok:false, reason:'상단 매칭 실패'};
        const rowData = d[idx];
        if (!rowData.REG_NO) return {ok:false, reason:'REG_NO 없음'};

        window.mSelectRowData = null;
        try { sGridList.$local_grid.options.dataModel.data = []; sGridList.$local_grid.refreshView(); } catch(e){}
        jQuery('#sendMesFlag').val('N');
        window.totSelectRowData = rowData;
        const param = {
            PROD_ID: rowData.PROD_ID, PLAN_DA: rowData.REG_DT, PRDT_QTY: rowData.PRDT_QTY,
            SHIPTDA: rowData.SHIPTDA, EXT_PLAN_REG_NO: rowData.REG_NO, DAY_OPT: jQuery('#dayOpt').val()
        };
        try { mGridList.$local_grid.options.dataModel.data = []; mGridList.$local_grid.refreshView(); } catch(e){}
        mGridList.searchListData(param);
        return {ok:true, idx, prodId: rowData.PROD_ID};
    }""", {"prodNo": prod_no, "extReg": str(target_ext_reg)})
    if not click_r.get("ok"):
        return {"ok":False, "stage":"top_click", "err":click_r.get("reason")}

    # m_grid 로드 대기
    for _ in range(16):
        time.sleep(0.5)
        if page.evaluate("""(e) => { try { return jQuery('#m_grid_body').pqGrid('option','dataModel').data.some(r => String(r.EXT_PLAN_REG_NO) === String(e)); } catch(err){return false;} }""", str(target_ext_reg)):
            break
    else:
        return {"ok":False, "stage":"m_wait"}

    # 2. 라인 선택
    m_r = page.evaluate("""(targetLine) => {
        const m = jQuery('#m_grid_body').pqGrid('option','dataModel').data;
        const idx = m.findIndex(r => r.LINE_CD === targetLine);
        if (idx < 0) return {ok:false, reason:'line 없음', avail:m.map(r=>r.LINE_CD)};
        const rowData = m[idx];
        window.mSelectRowData = rowData;
        try {
            sGridList.searchListData({
                LINE_CD:rowData.LINE_CD, STD_DA:rowData.STD_DA, PLAN_DA:rowData.PLAN_DA,
                LINE_DIV_CD:rowData.LINE_DIV_CD, DAY_OPT:jQuery('#dayOpt').val()
            });
        } catch(e) { return {ok:false, reason:'sGridList err: '+e.message}; }
        return {ok:true, lineDivCd: rowData.LINE_DIV_CD};
    }""", target_line)
    if not m_r.get("ok"):
        return {"ok":False, "stage":"m_select", "err":m_r.get("reason"), "avail":m_r.get("avail")}

    # s_grid 로드 대기 — 고정 3초 (0건이어도 OK, 주간 초기 상황 대응)
    time.sleep(3)
    s_rows = page.evaluate("() => { try { return sGridList.$local_grid.pdata.length; } catch(e){return -1;} }")
    if s_rows < 0:
        return {"ok":False, "stage":"s_wait", "rows":s_rows, "err":"s_grid 미초기화"}

    # 3. addRow 인라인 + 임시저장
    save_js = """
    async (args) => {
        if (!mSelectRowData || mSelectRowData.LINE_CD !== args.targetLine) return {ok:false, reason:'mSelect err'};
        const m = mSelectRowData;
        if (!m.MPRDTN_DIV_CD) return {ok:false, reason:'MPRDTN_DIV_CD 공란'};
        if (!m.PRDT_CATE_CD) return {ok:false, reason:'PRDT_CATE_CD 공란'};

        const rowData = Object.assign({}, m);
        rowData.DAY_OPT          = jQuery('#dayOpt').val();
        rowData.WORK_STATUS_CD   = 'A';
        rowData.WORK_STATUS_NM   = '추가';
        rowData.EXT_PLAN_YN      = 'Y';
        rowData.EXT_PLAN_REG_NO  = m.EXT_PLAN_REG_NO;
        rowData.LINE_NM          = m.LINE_NM; rowData.LINE_CD = m.LINE_CD;
        rowData.PROD_NO          = m.PROD_NO; rowData.PROD_NM = m.PROD_NM; rowData.PROD_REV = m.PROD_REV;
        rowData.PRDT_PLAN_QTY    = m.ADD_PRDT_QTY; rowData.OLD_PRDT_PLAN_QTY = m.ADD_PRDT_QTY;
        rowData.NEXT_PLAN_DA     = m.NEXT_PLAN_DA;

        const addRowIndx = sGridList.$local_grid.pdata.length;
        sGridList.$local_grid.addRow({rowData: rowData, checkEditable:false, rowIndx: addRowIndx});
        sGridList.setBeginTime();

        if (args.dryRun) return {ok:true, dryRun:true, addedRank: rowData.PRDT_RANK, dataListLen: sGridList.$local_grid.getData().length};

        const dataList = sGridList.$local_grid.getData();
        jQuery('#sendMesFlag').val('N');
        const param = JSON.stringify({dataList, PARENT_PROD_ID: totSelectRowData.PROD_ID, sendMesFlag:'N'});
        const out = await new Promise((resolve) => {
            const to = setTimeout(() => resolve({ok:false, err:'timeout 30s'}), 30000);
            jQuery.ajax({
                url: args.saveUrl, type:'post', data: param,
                success: r => { clearTimeout(to); resolve({ok:true, r}); },
                error: (xhr,s,e) => { clearTimeout(to); resolve({ok:false, httpStatus:xhr.status, body:(xhr.responseText||'').slice(0,400)}); }
            });
        });
        return Object.assign({addedRank: rowData.PRDT_RANK, dataListLen: dataList.length}, out);
    }
    """
    r = page.evaluate(save_js, {"targetLine": target_line, "saveUrl": save_url, "dryRun": dry_run})
    return r


def dedupe_night_first_5_via_http(sess, items, check_count=5):
    """야간 첫 N행 dedupe — HTTP totGrid 조회로 대체."""
    if not items:
        return items
    today_str = datetime.now().strftime("%Y-%m-%d")
    grid_all = fetch_tot_grid_via_http(sess, today_str)
    grid = [{"PROD_NO": g["PROD_NO"], "QTY": int(g.get("PRDT_QTY") or g.get("ADD_PRDT_QTY") or 0)}
            for g in grid_all if g.get("REG_DT") == today_str]
    if not grid:
        print(f"[dedupe:http] 상단 grid REG_DT={today_str} 등록분 0건 — 야간 dedupe 스킵")
        return items
    skipped, kept = [], []
    for i, it in enumerate(items):
        if i < check_count:
            match = next((g for g in grid if g["PROD_NO"] == it["PROD_NO"]), None)
            if match:
                skipped.append(it)
                print(f"[dedupe:http] 야간 {i+1}행 PROD_NO={it['PROD_NO']} 야간qty={it['QTY']} 주간qty={match['QTY']} → 제외")
                continue
        kept.append(it)
    print(f"[dedupe:http] 야간 {len(items)}건 → 등록 {len(kept)}건 / 제외 {len(skipped)}건")
    return kept


def dedupe_night_first_5(page, items, check_count=5):
    """세션153 A안 3단계: _HTTP_ONLY_SESS 시 http 버전 분기.
    야간 추출 결과의 첫 N행(기본 5)을 ERP 상단 grid 기등록분과 비교해 중복 제외.

    매칭 기준: REG_DT=오늘 AND **PROD_NO 일치만** (수량 무관, 2026-04-29 사용자 결정 v3.2).
    매칭된 행만 items에서 제외. N+1번째 이후는 그대로.

    2026-04-29 사용자 요청 — SP3M3 야간 1~5행이 주간 등록분과 2중 등록되는 현상 방지.
    초기 v3.1은 PROD_NO+수량 동시 매칭이었으나, 같은 품번이면 수량 다르더라도 작업자 입장에서
    중복 작업 위험 동일하므로 PROD_NO만으로 매칭하도록 정책 단순화.
    """
    if _HTTP_ONLY_SESS is not None:
        return dedupe_night_first_5_via_http(_HTTP_ONLY_SESS, items, check_count)
    if not items:
        return items
    today_str = datetime.now().strftime("%Y-%m-%d")
    grid = page.evaluate("""(today) => {
        try {
            const d = jQuery('#grid_body').pqGrid('option','dataModel').data;
            return d.filter(r => r.REG_DT === today).map(r => ({
                PROD_NO: r.PROD_NO,
                QTY: Number(r.PRDT_QTY || r.ADD_PRDT_QTY || r.PRDT_PLAN_QTY || 0)
            }));
        } catch(e) { return []; }
    }""", today_str)

    if not grid:
        print(f"[dedupe] 상단 grid REG_DT={today_str} 등록분 0건 — 야간 dedupe 스킵 (전량 진행)")
        return items

    skipped, kept = [], []
    for i, it in enumerate(items):
        if i < check_count:
            # 수량 무관, PROD_NO만 매칭 (v3.2)
            match = next((g for g in grid if g["PROD_NO"] == it["PROD_NO"]), None)
            if match:
                skipped.append(it)
                # 로그에는 야간 qty + 주간 등록 qty 둘 다 표시 (참고용)
                print(f"[dedupe] 야간 {i+1}행 PROD_NO={it['PROD_NO']} 야간qty={it['QTY']} 주간qty={match['QTY']} → 품번 일치, 제외")
                continue
        kept.append(it)
    print(f"[dedupe] 야간 추출 {len(items)}건 → 등록 {len(kept)}건 / 제외 {len(skipped)}건 (1~{check_count}행만 검사, 품번 일치 기준)")
    return kept


def fetch_tot_grid_via_http(sess, target_date):
    """totGrid 조회 — GET /prdtPlanMng/selectListDoAddnPrdtPlanInstrMngNew.do.

    실측 캡처 (2026-05-13): GET method, params={searchDate, searchShipDa, searchRegUsrNm=대원테크}.
    POST 보내면 "Request method 'POST' not supported" 거부.
    """
    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    sess.headers["X-Requested-With"] = "XMLHttpRequest"
    sess.headers["Referer"] = "http://erp-dev.samsong.com:19100/prdtPlanMng/viewListDoAddnPrdtPlanInstrMngNew.do"
    sess.headers["ajax"] = "true"
    params = {"searchDate": "", "searchShipDa": target_date, "searchRegUsrNm": "대원테크"}
    r = sess.get("http://erp-dev.samsong.com:19100/prdtPlanMng/selectListDoAddnPrdtPlanInstrMngNew.do",
                 params=params, timeout=15)
    if r.status_code != 200:
        return []
    data = r.json()
    return data.get("data", {}).get("list", []) or []


def dedupe_existing_registrations_via_http(sess, items, prod_date, line_cd):
    """phase1.5 dedupe — requests 직접 (브라우저 의존 0).

    세션153 A안 3단계 — totGrid endpoint 사용.
    """
    if not items:
        return items
    target_date = prod_date.strftime("%Y-%m-%d")
    grid_all = fetch_tot_grid_via_http(sess, target_date)
    grid = [{"PROD_NO": g["PROD_NO"], "REG_NO": g.get("REG_NO"),
             "QTY": int(g.get("PRDT_QTY") or g.get("ADD_PRDT_QTY") or 0),
             "_raw": g}
            for g in grid_all if g.get("REG_DT") == target_date]
    if not grid:
        print(f"[dedupe-day:http] 그리드 REG_DT={target_date} 등록분 0건 — 전량 진행 ({len(items)}건)")
        return items
    existing = {g["PROD_NO"]: g for g in grid}
    skipped, kept = [], []
    for it in items:
        if it["PROD_NO"] in existing:
            g = existing[it["PROD_NO"]]
            print(f"[dedupe-day:http] {it['PROD_NO']} 이미 등록 (REG_NO={g['REG_NO']} qty={g['QTY']}) — 제외")
            skipped.append(it)
        else:
            kept.append(it)
    print(f"[dedupe-day:http] items {len(items)}건 → 등록 {len(kept)}건 / 제외 {len(skipped)}건 (PROD_NO 일치 기준, prod_date={target_date})")
    return kept


def dedupe_existing_registrations(page, items, prod_date, line_cd):
    """세션153 A안 3단계: _HTTP_ONLY_SESS 시 http 버전으로 분기 (호출자 시그니처 무변경).
    같은 prod_date + line_cd ERP 기등록 PROD_NO 제외 (주간/morning 및 --xlsx direct 전용).

    배경: 세션133 5/1 P3 PoC에서 RSP3SC0665 1건 추가 등록 시 이미 morning 20건 등록 상태였음에도
    중복 가드 없이 등록 → MES 잔존 발생. 사용자 요청 "계획 반영 전 이미 등록된 품번 확인 단계 추가".

    검사 기준:
      - 그리드 REG_DT == prod_date + PROD_NO 일치 → 제외
      - 다른 PLAN_DA(주야간 cross)는 제외 안 함 — 같은 PROD_NO 주간/야간 양쪽 등록 허용
      - 야간 1~5행 dedupe(dedupe_night_first_5)와 별개 — 그건 evening 흐름 전용

    Returns: 중복 제외된 items 리스트
    """
    if _HTTP_ONLY_SESS is not None:
        return dedupe_existing_registrations_via_http(_HTTP_ONLY_SESS, items, prod_date, line_cd)
    if not items:
        return items
    target_date = prod_date.strftime("%Y-%m-%d")
    # 그리드 새로고침 (page 진입 직후 자동 검색 안 된 케이스 방어)
    try:
        page.evaluate("() => { try { totGridList.searchListData(); } catch(e){} }")
    except Exception:
        pass
    time.sleep(2)
    grid = page.evaluate("""(args) => {
        try {
            const d = jQuery('#grid_body').pqGrid('option','dataModel').data;
            return d.filter(r => r.REG_DT === args.date)
                    .map(r => ({PROD_NO: r.PROD_NO, REG_NO: r.REG_NO, QTY: Number(r.PRDT_QTY||r.ADD_PRDT_QTY||0)}));
        } catch(e) { return []; }
    }""", {"date": target_date})
    if not grid:
        print(f"[dedupe-day] 그리드 REG_DT={target_date} 등록분 0건 — 전량 진행 ({len(items)}건)")
        return items
    existing = {g["PROD_NO"]: g for g in grid}
    skipped, kept = [], []
    for it in items:
        if it["PROD_NO"] in existing:
            g = existing[it["PROD_NO"]]
            print(f"[dedupe-day] {it['PROD_NO']} 이미 등록 (REG_NO={g['REG_NO']} qty={g['QTY']}) — 제외")
            skipped.append(it)
        else:
            kept.append(it)
    print(f"[dedupe-day] items {len(items)}건 → 등록 {len(kept)}건 / 제외 {len(skipped)}건 (PROD_NO 일치 기준, prod_date={target_date})")
    return kept


def build_requests_session_from_page(page):
    """Playwright page의 cookie + XSRF로 requests Session 구성.

    P2/P3/P4 검증된 헤더 레시피:
      - ajax: true (jQuery prefilter 자동 헤더 — 누락 시 multiList 500 / 8ms 즉시 거부)
      - X-XSRF-TOKEN (매 write 호출 직전 cookie에서 다시 읽어 갱신 필수 — Spring Security 회전)
      - Content-Type: application/json; charset=UTF-8 (P4 단계 2 발견 — form-urlencoded는 % 파싱 에러)
    """
    import requests as _req
    cookies = page.context.cookies()
    sess = _req.Session()
    xsrf = None
    for c in cookies:
        sess.cookies.set(c["name"], c["value"], domain=c["domain"].lstrip("."), path=c.get("path", "/"))
        if c["name"].upper() in ("XSRF-TOKEN", "X-XSRF-TOKEN"):
            xsrf = c["value"]
    sess.headers.update({
        "ajax": "true",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": D0_URL,
        "Origin": "http://erp-dev.samsong.com:19100",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
        "Accept-Language": "ko-KR,ko;q=0.9",
    })
    if xsrf:
        sess.headers["X-XSRF-TOKEN"] = xsrf
    return sess


def refresh_xsrf_from_cookies(sess):
    """매 write 호출 직전 cookie에서 X-XSRF-TOKEN 갱신 (Spring Security 회전 대응)."""
    for c in sess.cookies:
        if c.name.upper() in ("XSRF-TOKEN", "X-XSRF-TOKEN"):
            sess.headers["X-XSRF-TOKEN"] = c.value
            return c.value
    return None


def _sort_idx_map_desc(grid_by_pno):
    """phase4 매뉴얼 4번 룰 — PROD_NO별 REG_NO 내림차순 정렬 (max ext = idx=0).

    같은 PROD_NO가 주야 양쪽 등록된 경우 야간 신규(큰 ext)가 먼저 매핑되어야
    rank가 야간 row에 정확히 박힘. ascending 정렬 시 주간 기존 ext에 덮어 박혀
    SmartMES 누락 사고 발생 (세션152 / 세션155 2회 회귀).

    모든 phase4 매칭 함수는 반드시 이 헬퍼 또는 동등 정렬 사용 필수.
    legacy api_rank_batch(L1311 JS)도 동일 정책(`b-a`) 적용 중.
    """
    for v in grid_by_pno.values():
        v.sort(key=lambda x: int(x.get("REG_NO", 0)), reverse=True)
    return grid_by_pno


def api_rank_batch(page, items, target_line, save_url, sess=None):
    """rank_batch의 옵션 A 하이브리드 변형 — jQuery.ajax POST를 requests 직접 호출로 전환.

    화면 인터랙션(process_one_row dry_run으로 addRow + sGridList 채우기)은 유지,
    multiListMainSubPrdtPlanRankDecideMng.do POST만 requests로.

    sendMesFlag='N' 강제 — final_save(sendMesFlag='Y')는 별도 함수. P3 사고 재발 방지.

    P4 단계 2 검증된 헤더/페이로드 레시피 적용:
      - Content-Type: application/json; charset=UTF-8
      - data = JSON.stringify({dataList, PARENT_PROD_ID, sendMesFlag:'N'}).encode('utf-8')
      - X-XSRF-TOKEN 매 호출 직전 갱신
    """
    import requests as _req
    if sess is None:
        sess = build_requests_session_from_page(page)
    base_url = "http://erp-dev.samsong.com:19100"

    # 상단 grid REG_NO 오름차순 배열 매핑 (세션151: PROD_NO 단일 키 → 다중 등장 매핑)
    # 같은 PROD_NO가 items에 N번 나오면 grid에도 N행 등록되는데, 기존 단일 키는 N건 모두 같은 ext로 매핑되어
    # 1건만 처리되고 나머지(N-1)건 누락. items 등장 순서대로 N번째 REG_NO 매핑.
    today_str = datetime.now().strftime("%Y-%m-%d")
    idx_map = page.evaluate("""(today) => {
        const d = jQuery('#grid_body').pqGrid('option','dataModel').data;
        const map = {};
        d.forEach(r => {
            if (r.REG_DT === today) {
                const key = r.PROD_NO;
                if (!map[key]) map[key] = [];
                map[key].push(parseInt(r.REG_NO));
            }
        });
        // 세션152: 매뉴얼 4번 룰 "EXT_PLAN_REG_NO 최대값 매핑" 준수 — 내림차순.
        // items N번째 등장 시 N번째 큰 값(=가장 최근 phase3 신규 INSERT) 우선.
        // 야간/주간 중복 PROD_NO 케이스: 야간 items 1건 → 최대 ext(=야간 신규) 매핑 → 주간 기존 ext에 덮어쓰기 방지.
        for (const k in map) map[k].sort((a,b) => b - a);
        return map;
    }""", today_str)

    done, failed, missing, fails = 0, 0, 0, []
    pno_seen = {}  # pno -> 이미 처리한 등장 횟수
    for it in items:
        pno = it["PROD_NO"]
        arr = idx_map.get(pno) or []
        n = pno_seen.get(pno, 0)
        if n >= len(arr):
            print(f"[api_phase4] {pno} 상단 grid 매칭 실패 (등장 #{n+1}, grid에 {len(arr)}건만)")
            missing += 1
            pno_seen[pno] = n + 1
            continue
        ext = arr[n]
        pno_seen[pno] = n + 1

        # 1. process_one_row(dry_run=True) — addRow만, jQuery.ajax POST 안 함
        p = process_one_row(page, pno, ext, target_line, save_url, dry_run=True)
        if not p.get("ok") or not p.get("dryRun"):
            print(f"[api_phase4] {pno} process_one_row dry-run 실패: {p}")
            failed += 1
            fails.append({"prod": pno, "stage": "addRow", "reason": p.get("reason") or p.get("err")})
            continue
        time.sleep(0.5)

        # 2. dataList + PARENT_PROD_ID 추출
        dataList = page.evaluate("() => sGridList.$local_grid.getData()")
        parent_prod_id = page.evaluate("() => totSelectRowData ? totSelectRowData.PROD_ID : null")
        if not dataList or not parent_prod_id:
            print(f"[api_phase4] {pno} dataList/PARENT_PROD_ID 추출 실패")
            failed += 1
            fails.append({"prod": pno, "stage": "extract", "reason": "dataList or parent missing"})
            continue

        # 3. requests POST (sendMesFlag='N' 강제)
        refresh_xsrf_from_cookies(sess)
        param = {"dataList": dataList, "PARENT_PROD_ID": parent_prod_id, "sendMesFlag": "N"}
        body = json.dumps(param, ensure_ascii=False).encode("utf-8")
        try:
            resp = sess.post(
                base_url + save_url,
                data=body,
                headers={"Content-Type": "application/json; charset=UTF-8"},
                timeout=30,
            )
        except Exception as e:
            print(f"[api_phase4] {pno} requests POST 예외: {e}")
            failed += 1
            fails.append({"prod": pno, "stage": "post", "reason": str(e)})
            continue

        if resp.status_code != 200:
            print(f"[api_phase4] {pno} POST status={resp.status_code} body={resp.text[:200]}")
            failed += 1
            fails.append({"prod": pno, "stage": "post", "reason": f"http {resp.status_code}"})
            continue

        try:
            r_json = resp.json()
        except Exception:
            print(f"[api_phase4] {pno} JSON parse 실패: {resp.text[:200]}")
            failed += 1
            fails.append({"prod": pno, "stage": "parse", "reason": "json parse"})
            continue

        if str(r_json.get("statusCode")) != "200":
            print(f"[api_phase4] {pno} statusCode={r_json.get('statusCode')} txt={r_json.get('statusTxt')}")
            failed += 1
            fails.append({"prod": pno, "stage": "server", "reason": r_json.get("statusTxt")})
            continue

        # MES 전송 미발생 검증 (sendMesFlag='N' 효과)
        mesMsg = r_json.get("mesMsg", "")
        if mesMsg:
            print(f"[api_phase4] ⚠ {pno} sendMesFlag=N인데 mesMsg 비어있지 않음: {mesMsg[:200]} — 즉시 중단")
            failed += 1
            fails.append({"prod": pno, "stage": "mes_unexpected", "reason": "mesMsg present with sendMesFlag=N"})
            break

        print(f"[api_phase4] {pno} ext={ext} -> OK (api)")
        done += 1

    return {"done": done, "failed": failed, "missing": missing, "fails": fails}


def rank_batch(page, items, target_line, save_url, dry_run=False):
    """엑셀 순서대로 상단 idx 매핑 → process_one_row 반복.

    세션151 보강: 같은 PROD_NO가 items에 N번 나오면(야간 같은 품번 2회 등장) grid에도 N행 등록되는데,
    기존 idx_map은 PROD_NO 단일 키 + REG_NO 최대값만 보관 → N건 모두 같은 ext에 매핑되어
    1건만 라인 배치되고 나머지(N-1)건은 누락. 4/30 evening RSP3SC0395/0396 ext=320225/320227 중복 OK
    + ext=320224/320226 슬롯 빈 케이스가 직접 증거.
    → idx_map을 PROD_NO -> [REG_NO 오름차순 배열]로, items 순회 시 같은 PROD_NO N번째 등장에 N번째 REG_NO 매핑.
    """
    # 상단 grid REG_NO 오름차순 배열 매핑
    idx_map = page.evaluate("""(today) => {
        const d = jQuery('#grid_body').pqGrid('option','dataModel').data;
        const map = {};
        for (let i=0;i<d.length;i++) {
            if (d[i].REG_DT !== today) continue;
            const pno = d[i].PROD_NO;
            if (!map[pno]) map[pno] = [];
            map[pno].push({idx:i, extReg:Number(d[i].REG_NO)});
        }
        // 세션152 evening: 매뉴얼 4번 룰 "EXT_PLAN_REG_NO 최대값 매핑" 준수 — 내림차순.
        // api_rank_batch(L1046)와 동일 정렬 방향 동기화. legacy --legacy-mode fallback 경로에서도
        // 야간/주간 중복 PROD_NO 시 신규 ext(=최대) 우선 매핑하여 주간 기존 행 덮어쓰기 차단.
        // GPT 판정 부분반영 후속 patch (legacy 경로 동기화 누락 지적 반영).
        for (const k in map) map[k].sort((a,b) => b.extReg - a.extReg);
        return map;
    }""", datetime.today().strftime("%Y-%m-%d"))

    done, failed, missing = 0, 0, 0
    fails = []
    pno_seen = {}  # pno -> 이미 처리한 등장 횟수
    for it in items:
        pno = it["PROD_NO"]
        arr = idx_map.get(pno) or []
        n = pno_seen.get(pno, 0)
        if n >= len(arr):
            print(f"[phase4] SKIP missing in grid: {pno} (등장 #{n+1}, grid에 {len(arr)}건만)")
            missing += 1
            pno_seen[pno] = n + 1
            continue
        ext = arr[n]["extReg"]
        pno_seen[pno] = n + 1
        r = process_one_row(page, pno, ext, target_line, save_url, dry_run=dry_run)
        ok = r.get("ok") and (r.get("dryRun") or (r.get("r") and str(r["r"].get("statusCode"))=="200"))
        print(f"[phase4] {pno} ext={ext} -> {'OK' if ok else 'FAIL'}")
        if ok: done += 1
        else:
            failed += 1
            fails.append({"pno":pno, "ext":ext, "result":r})
            if failed >= 3: print("[phase4] 3회 연속 실패 — 중단"); break
        time.sleep(1.2)
    return {"done":done, "failed":failed, "missing":missing, "fails":fails}


# ============================================================
# Phase 5: 최종 저장 (MES 전송)
# ============================================================
def process_one_row_via_http(sess, prod_no, target_ext_reg, target_line, save_url, prod_date, day_opt='1'):
    """phase4 한 row 처리 — requests 직접 (브라우저 page.evaluate 0).

    JS process_one_row 흐름을 그대로 mirror:
      1) totGrid GET → PROD_NO + REG_NO 매칭 row (totSelectRowData)
      2) mGrid GET → target_line LINE_CD 매칭 row (mSelectRowData)
      3) sGrid GET → 현재 누적 dataList
      4) rowData 구성 (m copy + WORK_STATUS_CD=A, EXT_PLAN_YN=Y 등)
      5) POST save_url + {dataList + [rowData], PARENT_PROD_ID, sendMesFlag='N'}

    Returns: {ok, addedRank, parent_prod_id, m_row, ...}
    """
    BASE = "http://erp-dev.samsong.com:19100"
    target_date = prod_date.strftime("%Y-%m-%d")

    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    sess.headers["X-Requested-With"] = "XMLHttpRequest"
    sess.headers["Referer"] = BASE + "/prdtPlanMng/viewListDoAddnPrdtPlanInstrMngNew.do"
    sess.headers["ajax"] = "true"

    # 1) totGrid 조회 + 매칭
    grid = fetch_tot_grid_via_http(sess, target_date)
    tot = next((r for r in grid if r.get("PROD_NO") == prod_no and str(r.get("REG_NO")) == str(target_ext_reg)), None)
    if tot is None:
        return {"ok": False, "stage": "top_click", "err": "totGrid 매칭 실패"}

    # 2) mGrid GET
    m_params = {"PROD_ID": tot["PROD_ID"], "PLAN_DA": tot["REG_DT"], "PRDT_QTY": tot["PRDT_QTY"],
                "SHIPTDA": tot["SHIPTDA"], "EXT_PLAN_REG_NO": tot["REG_NO"], "DAY_OPT": day_opt}
    r = sess.get(BASE + "/prdtPlanMng/selectListDoAddnPrdtConctLineNew.do", params=m_params, timeout=15)
    m_list = r.json().get("data", {}).get("list", []) or []
    m = next((row for row in m_list if row.get("LINE_CD") == target_line), None)
    if m is None:
        return {"ok": False, "stage": "m_select", "err": f"line {target_line} 없음", "avail": [x.get("LINE_CD") for x in m_list]}
    if not m.get("MPRDTN_DIV_CD"):
        return {"ok": False, "stage": "m_validate", "err": "MPRDTN_DIV_CD 공란"}
    if not m.get("PRDT_CATE_CD"):
        return {"ok": False, "stage": "m_validate", "err": "PRDT_CATE_CD 공란"}

    # 3) sGrid GET
    s_params = {"LINE_CD": m["LINE_CD"], "STD_DA": m["STD_DA"], "PLAN_DA": m["PLAN_DA"],
                "LINE_DIV_CD": m["LINE_DIV_CD"], "DAY_OPT": day_opt}
    r2 = sess.get(BASE + "/prdtPlanMng/selectListDoAddnPrdtConctLineDetailNew.do", params=s_params, timeout=15)
    s_data = r2.json().get("data", {}).get("list", []) or []

    # 4) rowData = Object.assign({}, m) + 필드 덮어쓰기 (JS code mirror)
    row_data = dict(m)
    row_data["DAY_OPT"] = day_opt
    row_data["WORK_STATUS_CD"] = "A"
    row_data["WORK_STATUS_NM"] = "추가"
    row_data["EXT_PLAN_YN"] = "Y"
    row_data["EXT_PLAN_REG_NO"] = m["EXT_PLAN_REG_NO"]
    row_data["LINE_NM"] = m["LINE_NM"]
    row_data["LINE_CD"] = m["LINE_CD"]
    row_data["PROD_NO"] = m["PROD_NO"]
    row_data["PROD_NM"] = m["PROD_NM"]
    row_data["PROD_REV"] = m["PROD_REV"]
    row_data["PRDT_PLAN_QTY"] = m["ADD_PRDT_QTY"]
    row_data["OLD_PRDT_PLAN_QTY"] = m["ADD_PRDT_QTY"]
    row_data["NEXT_PLAN_DA"] = m["NEXT_PLAN_DA"]

    # 5) dataList = s_data + [row_data] + POST save_url
    data_list = list(s_data) + [row_data]
    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    payload = {"dataList": data_list, "PARENT_PROD_ID": tot["PROD_ID"], "sendMesFlag": "N"}
    r3 = sess.post(BASE + save_url, data=json.dumps(payload),
                   headers={"Content-Type": "application/json; charset=utf-8",
                            "X-XSRF-TOKEN": sess.cookies.get("XSRF-TOKEN", "")},
                   timeout=30)
    res = r3.json() if r3.headers.get("content-type", "").startswith("application/json") else {}
    if r3.status_code != 200 or str(res.get("statusCode")) != "200":
        return {"ok": False, "stage": "save", "err": f"status={r3.status_code} body={r3.text[:300]}"}
    return {"ok": True, "addedRank": len(data_list), "parent_prod_id": tot["PROD_ID"],
            "m_row": {k: m[k] for k in ("LINE_CD", "STD_DA", "PLAN_DA", "LINE_DIV_CD") if k in m},
            "dataListLen": len(data_list)}


def api_rank_batch_via_http(sess, items, target_line, save_url, prod_date, day_opt='1'):
    """phase4 api_rank_batch HTTP — items 순회하며 process_one_row_via_http 호출.

    같은 PROD_NO N회 등장 시 N번째 REG_NO 매칭 (api_rank_batch와 동일 정책).
    last_m_row / last_parent_prod_id 반환 — final_save_via_http에서 사용.
    """
    target_date = prod_date.strftime("%Y-%m-%d")
    grid = fetch_tot_grid_via_http(sess, target_date)
    grid_by_pno = {}
    for g in grid:
        if g.get("REG_DT") != target_date:
            continue
        grid_by_pno.setdefault(g["PROD_NO"], []).append(g)
    # ⚠️ 매뉴얼 4번 룰 (REG_NO 내림차순 매핑) — 반드시 준수
    # ============================================================
    # 같은 PROD_NO가 주야 양쪽 등록된 경우 max ext(= 야간 신규)부터 매핑.
    # ascending 정렬 시 주간 기존 ext(작은 값)가 idx=0이 되어 야간 rank가
    # 주간 row에 덮어 박힘 → SmartMES 누락 사고.
    # 회귀 이력 — 같은 사고 2회 발생:
    #   세션152 evening: legacy rank_batch idx_map JS sort a-b → b-a 패치
    #   세션155 evening: A안 3단계 신설 시 reverse=True 누락 → 5건 누락 → 보강
    # 새 함수 추가 시 _sort_idx_map_desc(...) 헬퍼 호출 또는 reverse=True 명시 필수.
    _sort_idx_map_desc(grid_by_pno)

    done = failed = missing = 0
    fails = []
    last_parent = None
    last_m_row = None
    pno_idx = {}
    # 세션159: 야간 신규 (pno, ext_reg) 처리 순서 보존 — final_save_via_http가
    # s_data sort에 사용. ERP s_data 응답 정렬 키가 excel 입력 순서와 달라
    # PRDT_RANK가 어긋나는 사고 재발 방지(setBeginTime은 idx+1로 박는다).
    processed_order = []

    for it in items:
        pno = it["PROD_NO"]
        idx = pno_idx.get(pno, 0)
        cands = grid_by_pno.get(pno, [])
        if idx >= len(cands):
            print(f"[api_phase4:http] {pno} 등장 #{idx+1} grid 후보 부족 ({len(cands)}건)")
            missing += 1
            fails.append({"pno": pno, "stage": "match"})
            continue
        target_ext_reg = cands[idx]["REG_NO"]
        pno_idx[pno] = idx + 1

        result = process_one_row_via_http(sess, pno, target_ext_reg, target_line, save_url, prod_date, day_opt)
        if result.get("ok"):
            done += 1
            last_parent = result.get("parent_prod_id")
            last_m_row = result.get("m_row")
            processed_order.append({"PROD_NO": pno, "EXT_PLAN_REG_NO": target_ext_reg})
            print(f"[api_phase4:http] {pno} ext={target_ext_reg} -> OK (rank={result.get('addedRank')})")
        else:
            failed += 1
            fails.append({"pno": pno, "ext": target_ext_reg, "result": result})
            print(f"[api_phase4:http] {pno} ext={target_ext_reg} -> FAIL: {result}")
            if failed >= 3:
                print("[api_phase4:http] 3회 연속 실패 — 중단")
                break
        time.sleep(0.5)

    return {"done": done, "failed": failed, "missing": missing, "fails": fails,
            "last_parent_prod_id": last_parent, "last_m_row": last_m_row,
            "processed_order": processed_order}


def _apply_set_begin_time(s_data):
    """JS sGridList.setBeginTime() Python 재현 — 매 row에 PRDT_RANK + BEGIN_TIME + END_TIME +
    PLAN_DA_S + PLAN_DA_E 부여. 휴식시간 8종 자동 가산.

    JS 알고리즘 정밀 mirror (2026-05-13 캡처):
      idx==0: beginTime = DAY_BEGIN_TIME[0:2]+":00" or "08:00"
      idx>0:  beginTime = 이전 row의 endTime
      tDate = PLAN_DA + beginTime
      upm = UPH_QTY / 60 (분당 생산량)
      tDate += ceil(PRDT_PLAN_QTY / upm) 분
      휴식시간 가산 (start_date ≤ X and end_date ≥ X면 +N분):
        9:50~10:00 +10 / 12:00~13:00 +60 / 14:50~15:00 +10 / 17:00~17:10 +10 /
        21:00~21:10 +10 (PLAN_DA)
        00:00~01:00 +60 / 03:00~03:10 +10 / 05:00~05:10 +10 (NEXT_PLAN_DA)
      PLAN_DA_S/E = beginTime/endTime ≥ "08:00" ? PLAN_DA : NEXT_PLAN_DA
      PRDT_RANK = idx + 1
    """
    import math
    DFLT = "08:00"
    end_time = None
    for idx, obj in enumerate(s_data):
        plan_da = obj.get("PLAN_DA")
        next_plan_da = obj.get("NEXT_PLAN_DA") or plan_da

        if idx == 0:
            day_begin = obj.get("DAY_BEGIN_TIME")
            begin_time = (day_begin[:2] + ":00") if day_begin else "08:00"
        else:
            begin_time = end_time

        t_date = datetime.strptime(f"{plan_da} {begin_time}", "%Y-%m-%d %H:%M")
        start_date = t_date

        prdt_qty = int(obj.get("PRDT_PLAN_QTY") or 0)
        try:
            uph = float(obj.get("UPH_QTY") or 0)
        except (ValueError, TypeError):
            uph = 0
        upm = uph / 60.0 if uph > 0 else 1.0
        add_min = math.ceil(prdt_qty / upm) if upm > 0 else 0
        t_date += timedelta(minutes=add_min)
        end_date = t_date

        breaks = [
            (plan_da, "09:50", "10:00", 10),
            (plan_da, "12:00", "13:00", 60),
            (plan_da, "14:50", "15:00", 10),
            (plan_da, "17:00", "17:10", 10),
            (plan_da, "21:00", "21:10", 10),
            (next_plan_da, "00:00", "01:00", 60),
            (next_plan_da, "03:00", "03:10", 10),
            (next_plan_da, "05:00", "05:10", 10),
        ]
        for d, bs, es, dm in breaks:
            bp = datetime.strptime(f"{d} {bs}", "%Y-%m-%d %H:%M")
            ep = datetime.strptime(f"{d} {es}", "%Y-%m-%d %H:%M")
            if end_date >= bp and start_date <= ep:
                t_date += timedelta(minutes=dm)
                end_date = t_date

        end_time = f"{t_date.hour:02d}:{t_date.minute:02d}"
        plan_da_s = plan_da if begin_time >= DFLT else next_plan_da
        plan_da_e = plan_da if end_time >= DFLT else next_plan_da

        obj["BEGIN_TIME"] = begin_time
        obj["END_TIME"] = end_time
        obj["PRDT_RANK"] = idx + 1
        obj["PLAN_DA_S"] = plan_da_s
        obj["PLAN_DA_E"] = plan_da_e


def _reorder_s_data_by_processed(s_data, processed_order):
    """세션159: ERP s_data 응답을 excel 입력 순서로 재정렬.

    배경: ERP의 selectListDoAddnPrdtConctLineDetailNew.do 응답은 자체 정렬 키로 반환되며
    excel 입력 순서와 다르다. setBeginTime은 idx+1로 PRDT_RANK를 박으므로 sort 누락 시
    SmartMES rank가 어긋난다(세션159 야간 21건 회귀 발생).

    정책:
      - 야간 신규 (pno + ext_reg 매칭)는 processed_order(items 순) 순서로 정렬
      - 그 외(주간 기존)는 원래 s_data 순서 유지 — rank 1~(주간끝)
      - 야간이 뒤로 (rank 주간끝+1~끝)

    s_data row 키 호환: EXT_PLAN_REG_NO 우선, 없으면 REG_NO fallback.
    매칭 실패 시 sort 자체는 무해 — 원래 순서 유지 (=현 버그 그대로, regression 무).
    """
    if not processed_order:
        return s_data
    night_order = {}
    for i, p in enumerate(processed_order):
        key = (p["PROD_NO"], str(p["EXT_PLAN_REG_NO"]))
        night_order[key] = i

    def _row_key(row):
        pno = row.get("PROD_NO")
        ext = row.get("EXT_PLAN_REG_NO") or row.get("REG_NO")
        return (pno, str(ext) if ext is not None else "")

    matched = sum(1 for r in s_data if _row_key(r) in night_order)
    print(f"[phase5:http] s_data 야간 매칭: {matched}/{len(processed_order)} (s_data 총 {len(s_data)}건)")

    indexed = list(enumerate(s_data))
    def _sort_key(pair):
        orig_idx, row = pair
        k = _row_key(row)
        if k in night_order:
            return (1, night_order[k])
        return (0, orig_idx)
    indexed.sort(key=_sort_key)
    return [r for _, r in indexed]


def final_save_via_http(sess, save_url, last_m_row, parent_prod_id, day_opt='1',
                        processed_order=None, send_mes_flag="Y", exclude_prod_nos=None,
                        raise_on_error=True):
    """phase5 final_save — sGrid 재조회 후 sendMesFlag='Y'로 POST.

    last_m_row: api_rank_batch_via_http 결과에서 받은 마지막 처리 m row의 LINE_CD/STD_DA/PLAN_DA/LINE_DIV_CD.
    processed_order: 세션159 추가 — 야간 신규 (pno, ext_reg) items 입력 순서. s_data sort에 사용.
    """
    BASE = "http://erp-dev.samsong.com:19100"
    if last_m_row is None or parent_prod_id is None:
        raise RuntimeError("final_save_via_http: last_m_row/parent_prod_id 누락")

    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    sess.headers["ajax"] = "true"

    # sGrid 재조회
    s_params = {"LINE_CD": last_m_row["LINE_CD"], "STD_DA": last_m_row["STD_DA"],
                "PLAN_DA": last_m_row["PLAN_DA"], "LINE_DIV_CD": last_m_row["LINE_DIV_CD"],
                "DAY_OPT": day_opt}
    r = sess.get(BASE + "/prdtPlanMng/selectListDoAddnPrdtConctLineDetailNew.do",
                 params=s_params, timeout=15)
    s_data = r.json().get("data", {}).get("list", []) or []

    # 세션159: ERP s_data 응답 정렬이 excel 순서와 다르므로 setBeginTime 전 재정렬.
    exclude_set = set(exclude_prod_nos or [])
    if exclude_set:
        before = len(s_data)
        s_data = [row for row in s_data if row.get("PROD_NO") not in exclude_set]
        print(f"[pbom-guard] final_save payload exclude {sorted(exclude_set)}: sGrid {before}->{len(s_data)}건")
        if not s_data:
            raise RuntimeError("final_save_via_http: P-BOM guard 제외 후 전송 대상 0건")
        if processed_order:
            processed_order = [p for p in processed_order if p.get("PROD_NO") not in exclude_set]

    s_data = _reorder_s_data_by_processed(s_data, processed_order)

    # JS sGridList.setBeginTime() 정밀 재현 — 모든 row의 PRDT_RANK + BEGIN_TIME + END_TIME +
    # PLAN_DA_S + PLAN_DA_E 일괄 부여. 휴식시간 8종 자동 가산.
    # 2026-05-13 PoC 실측: 단순 PRDT_RANK 부여만으로는 "Transaction rolled back" -9999 발생 →
    # 시간 계산 비즈니스 로직 재현 필수.
    _apply_set_begin_time(s_data)
    print(f"[phase5:http] setBeginTime 적용 ({len(s_data)}건, ranks 1~{len(s_data)})")

    sess.headers["X-XSRF-TOKEN"] = sess.cookies.get("XSRF-TOKEN", "")
    payload = {"dataList": s_data, "PARENT_PROD_ID": parent_prod_id, "sendMesFlag": send_mes_flag}
    r2 = sess.post(BASE + save_url, data=json.dumps(payload),
                   headers={"Content-Type": "application/json; charset=utf-8",
                            "X-XSRF-TOKEN": sess.cookies.get("XSRF-TOKEN", "")},
                   timeout=120)
    if r2.headers.get("content-type", "").startswith("application/json"):
        res = r2.json()
    else:
        res = {"statusCode": r2.status_code, "mesMsg": r2.text[:1000]}
    print(f"[phase5:http final_save] sendMesFlag={send_mes_flag} status={r2.status_code} statusCode={res.get('statusCode')} mesMsg={(res.get('mesMsg') or '')[:120]}")
    if r2.status_code != 200 or str(res.get("statusCode")) != "200":
        if raise_on_error:
            raise RuntimeError(f"final_save_via_http 실패: status={r2.status_code} body={r2.text[:300]}")
    return res


def final_save_via_http_with_pbom_guard(sess, save_url, last_m_row, parent_prod_id, line_cd,
                                        day_opt='1', processed_order=None):
    """sendMesFlag=N preflight로 P-BOM 미등록을 제외한 뒤 sendMesFlag=Y는 1회만 호출."""
    excluded = set()
    for attempt in range(1, PBOM_GUARD_MAX_ITER + 1):
        res = final_save_via_http(
            sess, save_url, last_m_row, parent_prod_id,
            day_opt=day_opt, processed_order=processed_order,
            send_mes_flag="N", exclude_prod_nos=excluded, raise_on_error=False,
        )
        missing = parse_pbom_missing_from_mes_msg(res, line_cd=line_cd)
        if not missing:
            assert_mes_msg_ok(res, f"pbom-guard preflight line={line_cd}")
            print(f"[pbom-guard] preflight PASS line={line_cd} excluded={sorted(excluded)}")
            break
        new_missing = [p for p in missing if p not in excluded]
        if not new_missing:
            raise RuntimeError(f"[pbom-guard] 반복 P-BOM 미등록 감지: {missing}")
        excluded.update(new_missing)
        print(f"[pbom-guard] preflight {attempt}: P-BOM 미등록 {new_missing} -> 제외 누적 {sorted(excluded)}")
    else:
        raise RuntimeError(f"[pbom-guard] preflight 반복 한도 초과 excluded={sorted(excluded)}")

    res = final_save_via_http(
        sess, save_url, last_m_row, parent_prod_id,
        day_opt=day_opt, processed_order=processed_order,
        send_mes_flag="Y", exclude_prod_nos=excluded,
    )
    assert_mes_msg_ok(res, f"pbom-guard final_save line={line_cd}")
    return res, sorted(excluded)


def final_save(page, save_url, send_mes_flag="Y", exclude_prod_nos=None, raise_on_error=True):
    js = """
    async (args) => {
        jQuery('#sendMesFlag').val(args.sendMesFlag);
        sGridList.setBeginTime();
        const excluded = new Set(args.excludeProdNos || []);
        const dataList = sGridList.$local_grid.getData().filter(r => !excluded.has(r.PROD_NO));
        const param = JSON.stringify({dataList, PARENT_PROD_ID: totSelectRowData.PROD_ID, sendMesFlag:args.sendMesFlag});
        return await new Promise((resolve) => {
            const to = setTimeout(() => resolve({ok:false, err:'timeout 120s'}), 120000);
            jQuery.ajax({
                url: args.saveUrl, type:'post', data: param,
                success: r => { clearTimeout(to); resolve({ok:true, r}); },
                error: (xhr,s,e) => { clearTimeout(to); resolve({ok:false, httpStatus:xhr.status, body:(xhr.responseText||'').slice(0,400)}); }
            });
        });
    }
    """
    r = page.evaluate(js, {"saveUrl": save_url, "sendMesFlag": send_mes_flag, "excludeProdNos": list(exclude_prod_nos or [])})
    print(f"[phase5 final_save] {r}")
    if not r.get("ok") or str((r.get("r") or {}).get("statusCode")) != "200":
        if not raise_on_error:
            res = r.get("r") if isinstance(r.get("r"), dict) else {}
            if not res:
                res = {"statusCode": r.get("httpStatus"), "mesMsg": r.get("body") or str(r)}
            return res
        raise RuntimeError(f"최종 저장 실패: {r}")
    return r["r"]


# ============================================================
# Phase 6: SmartMES 검증
# ============================================================
def final_save_with_pbom_guard(page, save_url, line_cd):
    """Browser path fallback for P-BOM guard. Default operation should prefer --http-only."""
    excluded = set()
    for attempt in range(1, PBOM_GUARD_MAX_ITER + 1):
        res = final_save(page, save_url, send_mes_flag="N", exclude_prod_nos=excluded, raise_on_error=False)
        missing = parse_pbom_missing_from_mes_msg(res, line_cd=line_cd)
        if not missing:
            assert_mes_msg_ok(res, f"pbom-guard preflight browser line={line_cd}")
            print(f"[pbom-guard] browser preflight PASS line={line_cd} excluded={sorted(excluded)}")
            break
        new_missing = [p for p in missing if p not in excluded]
        if not new_missing:
            raise RuntimeError(f"[pbom-guard] browser 반복 P-BOM 미등록 감지: {missing}")
        excluded.update(new_missing)
        print(f"[pbom-guard] browser preflight {attempt}: P-BOM 미등록 {new_missing} -> 제외 누적 {sorted(excluded)}")
    else:
        raise RuntimeError(f"[pbom-guard] browser preflight 반복 한도 초과 excluded={sorted(excluded)}")
    res = final_save(page, save_url, send_mes_flag="Y", exclude_prod_nos=excluded)
    assert_mes_msg_ok(res, f"pbom-guard browser final_save line={line_cd}")
    return res, sorted(excluded)


def verify_smartmes(line_cd: str, prdt_da: datetime, excel_order: list):
    """SmartMES 등록 결과 검증.

    세션151 보강: 기존엔 set 비교라 같은 PROD_NO 중복(엑셀 2건 / server 1건) 누락 검출 불가.
    Counter 기반 카운트·중복 검증 + 건수 일치 검증 추가. 실패 시 False 반환 → main 종료 시 exit 비-0.
    """
    import urllib.request
    from collections import Counter
    headers = {
        "Authorization": f"Bearer {SMARTMES_TOKEN}",
        "tid": SMARTMES_TOKEN, "token": SMARTMES_TOKEN,
        "chnl": "MES", "from": "MESClient", "to": "MES",
        "usrid": "lmes", "Content-Type": "application/json",
    }
    body = json.dumps({"lineCd": line_cd, "prdtDa": prdt_da.strftime("%Y%m%d")}).encode()
    req = urllib.request.Request(f"{SMARTMES_BASE}/v2/prdt/schdl/list.api", data=body, headers=headers)
    try:
        items = json.loads(urllib.request.urlopen(req, timeout=5).read())["rslt"]["items"]
    except Exception as e:
        print(f"[phase6] SmartMES 조회 실패: {e}")
        return False
    r_items = sorted([x for x in items if x["workStatusCd"] == "R"], key=lambda i: i["prdtRank"])
    pnos_server = [x["pno"] for x in r_items]
    pnos_excel = [it["PROD_NO"] for it in excel_order]

    cnt_server = Counter(pnos_server)
    cnt_excel = Counter(pnos_excel)
    excel_set = set(pnos_excel)
    server_set = set(pnos_server)

    issues = []

    # (1) 카운트별 중복 누락 검증 — set 비교 한계 보강
    missing_pnos = []
    for pno, n_excel in cnt_excel.items():
        n_server = cnt_server.get(pno, 0)
        if n_server < n_excel:
            missing_pnos.append({"pno": pno, "excel": n_excel, "server": n_server})
    if missing_pnos:
        issues.append(f"품번별 누락 {len(missing_pnos)}건")

    # (2) 엑셀 PROD_NO 기준 server 등록 합계 vs 엑셀 건수
    server_in_excel_count = sum(cnt_server[p] for p in cnt_server if p in excel_set)
    if server_in_excel_count != len(pnos_excel):
        issues.append(f"건수 불일치 server={server_in_excel_count} excel={len(pnos_excel)}")

    # (3) 순서 검증 — 엑셀에만 있는 / server에만 있는 PROD_NO 제외하고 동일 시퀀스인지
    filtered_server = [p for p in pnos_server if p in excel_set]
    filtered_excel = [p for p in pnos_excel if p in server_set]
    if filtered_server != filtered_excel:
        issues.append("순서 불일치")

    if not issues:
        print(f"[phase6] {line_cd} 서열 순서 엑셀 일치 ✅ ({len(pnos_excel)}건, 카운트·중복·순서 모두 일치)")
        return True

    print(f"[phase6] {line_cd} 검증 실패: {' / '.join(issues)}")
    print(f"  server: {filtered_server[:10]}{'...' if len(filtered_server) > 10 else ''}")
    print(f"  excel : {filtered_excel[:10]}{'...' if len(filtered_excel) > 10 else ''}")
    if missing_pnos:
        print(f"  누락 상세: {missing_pnos}")
    return False


# ============================================================
# 세션 실행
# ============================================================
def run_session_line(page, wb, line, items, prod_date, dry_run, verify_prod_date=None,
                     parse_only=False, no_mes_send=False, api_mode=False, pbom_guard=True):
    """반환값 (2026-05-23 추가): True = Phase 3~6 정상 완료(jobsetup chain 가능),
    False = 실패/부분진행(Phase 4 실패 / Phase 6 SmartMES 불일치 / dry_run / parse_only / no_mes_send).
    기존 호출부는 반환값 무시(파이썬 호환). 신규 --xlsx 호출부는 jobsetup 가드용으로 활용."""
    if not items:
        print(f"[{line}] 추출 0건 — 스킵")
        return False
    cfg = LINE_CONFIG[line]
    xlsx = UPLOAD_DIR / f"d0_{line}_{prod_date.strftime('%Y%m%d')}.xlsx"
    make_upload_xlsx(items, prod_date, xlsx)

    if dry_run:
        print(f"[{line}] dry-run: 엑셀 생성까지만 (추출 {len(items)}건)")
        return False

    if parse_only:
        if _HTTP_ONLY_SESS is not None:
            d0_upload_via_http(_HTTP_ONLY_SESS, xlsx, parse_only=True)
        else:
            d0_upload(page, xlsx, parse_only=True)
        print(f"[{line}] parse-only: 업로드 창에 엑셀 첨부 + 서버 파싱까지만 (추출 {len(items)}건). Phase 4/5 미진행")
        return False

    target_line = "SP3M3" if line == "SP3M3" else "SD9A01"

    # 세션153 A안 3단계 — 완전 브라우저-less path
    if _HTTP_ONLY_SESS is not None:
        sess = _HTTP_ONLY_SESS
        d0_upload_via_http(sess, xlsx, parse_only=False)
        batch = api_rank_batch_via_http(sess, items, target_line, cfg["save_url"], prod_date, day_opt='1')
        print(f"[{line}:http] api_rank_batch: {batch}")
        if batch["failed"] > 0 or batch["missing"] > 0:
            print(f"[{line}:http] 실패/누락 존재 — 최종 저장 보류")
            return False
        if no_mes_send:
            print(f"[{line}:http] --no-mes-send: final_save 차단")
            return False
        final_items = items
        if pbom_guard:
            _, excluded_pnos = final_save_via_http_with_pbom_guard(
                sess, cfg["save_url"], batch["last_m_row"], batch["last_parent_prod_id"],
                target_line, day_opt='1', processed_order=batch.get("processed_order"))
            final_items = filter_items_by_prod_nos(items, excluded_pnos)
        else:
            print(f"[pbom-guard] disabled (--no-pbom-guard): Phase5 preflight skip line={target_line}")
            final_save_via_http(sess, cfg["save_url"], batch["last_m_row"], batch["last_parent_prod_id"],
                                day_opt='1', processed_order=batch.get("processed_order"))
        vp = verify_prod_date or prod_date
        if not verify_smartmes(target_line, vp, final_items):
            PHASE6_FAILED.append((target_line, vp.strftime("%Y-%m-%d")))
            return False
        return True

    d0_upload(page, xlsx)
    if api_mode:
        # 옵션 A 하이브리드 — rank 호출만 requests 직접 POST (sendMesFlag='N' 강제)
        batch = api_rank_batch(page, items, target_line, cfg["save_url"])
        print(f"[{line}] api_rank_batch (Phase 4 hybrid): {batch}")
    else:
        batch = rank_batch(page, items, target_line, cfg["save_url"], dry_run=False)
        print(f"[{line}] rank_batch: {batch}")
    if batch["failed"] > 0:
        print(f"[{line}] 실패 존재 — 최종 저장 보류")
        return False
    if no_mes_send:
        # P3 사고 재발 방지 (세션133 추가) — final_save sendMesFlag='Y' 차단
        # rank 임시저장(sendMesFlag='N')까지만 발생. MES 전송 미발생.
        # 단 ERP DB에는 등록분 + rank 행 잔존 — 사용자 정리 결정 필요
        print(f"[{line}] --no-mes-send: final_save 차단 (sendMesFlag='Y' MES 전송 미실행)")
        print(f"[{line}]   ERP rank 임시저장만 발생. 정리하려면 erp_d0_dedupe.py 또는 화면에서 삭제")
        return False
    final_items = items
    if pbom_guard:
        _, excluded_pnos = final_save_with_pbom_guard(page, cfg["save_url"], target_line)
        final_items = filter_items_by_prod_nos(items, excluded_pnos)
    else:
        print(f"[pbom-guard] disabled (--no-pbom-guard): Phase5 preflight skip line={target_line}")
        final_save(page, cfg["save_url"])
    vp = verify_prod_date or prod_date
    if not verify_smartmes(target_line, vp, final_items):
        PHASE6_FAILED.append((target_line, vp.strftime("%Y-%m-%d")))
        return False
    return True


def determine_session(arg):
    if arg != "auto":
        return arg
    h = datetime.now().hour
    if 6 <= h <= 10: return "morning"
    if 15 <= h <= 22: return "evening"
    raise ValueError(f"시간대({h}시)로 세션 자동 판정 불가 — 명시 지정 필요")


# ============================================================
# 메인
# ============================================================
def _apply_line_locks(args):
    """라인별 자동화 보류 잠금. state/<line>_*.lock 존재 시 해당 라인 차단.

    - --line SD9A01 + sd9a01_outer.lock → exit 0 (작업 미수행)
    - --line ALL    + sd9a01_outer.lock → SP3M3 only로 자동 조정
    """
    if OUTER_LOCK_FILE.exists():
        try:
            reason = OUTER_LOCK_FILE.read_text(encoding="utf-8").strip()
        except Exception:
            reason = "(사유 파일 읽기 실패)"
        print(f"[lock] SD9A01 OUTER 보류 잠금 활성 — {OUTER_LOCK_FILE}")
        print(f"[lock] 해제: 위 파일 삭제 또는 사용자 명시 해제 발화")
        # 사유 첫 줄만 안내 (전체는 파일 직접 확인)
        first_line = reason.split("\n", 1)[0] if reason else ""
        if first_line:
            print(f"[lock] {first_line}")
        if args.line == "SD9A01":
            print("[lock] --line SD9A01 호출 — 잠금으로 작업 미수행, exit 0")
            sys.exit(0)
        if args.line == "ALL":
            print("[lock] --line ALL → SD9A01 자동 skip, SP3M3 only 처리")
            args.line = "SP3M3"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--session", choices=["evening","morning","auto"], required=True)
    ap.add_argument("--line", choices=["SP3M3","SD9A01","ALL"], default="ALL")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--target-date", help="YYYY-MM-DD 파일명 날짜 (기본 자동)")
    ap.add_argument("--prod-date", help="YYYY-MM-DD ERP 등록 생산일자 명시 오버라이드 (공휴일 매핑 등 — evening: target-1 자동 무시 / morning: target 자동 무시)")
    ap.add_argument("--xlsx", help="업로드용 엑셀 파일 경로 직접 지정 (Phase 1 추출 건너뛰기)")
    ap.add_argument("--no-dedupe", action="store_true", help="dedupe 우회 — 같은 PROD_NO가 이미 등록돼 있어도 그대로 추가 등록 (사용자 명시 중복 등록)")
    ap.add_argument("--skip-upload", action="store_true", help="Phase 3 D0 업로드 건너뛰기 (이미 상단에 등록된 경우 Phase 4부터)")
    ap.add_argument("--parse-only", action="store_true", help="검증 모드: 엑셀 업로드창 첨부 + 서버 파싱(selectList)까지만. multiList(DB 저장)/Phase 4-5 모두 스킵")
    ap.add_argument("--no-mes-send", action="store_true", help="Phase 5 final_save(sendMesFlag='Y') 차단 — MES 전송 미발생. P3 PoC/검증용. rank 임시저장(sendMesFlag='N')까지는 발생 — ERP DB 등록분 정리 필요할 수 있음")
    # 세션133 사용자 명시 "기존꺼는 보관만 하고 이제 실제 작업은 하이브리드로 진행" — 기본값 = 하이브리드
    # --legacy-mode 명시 시만 화면 모드(jQuery.ajax) 진입. fallback 보존.
    ap.add_argument("--api-mode", action="store_true", default=True, help="옵션 A 하이브리드 (기본값 활성): rank 호출(Phase 4)을 requests 직접 POST. final_save(Phase 5)는 화면 모드 유지. 세션133 사용자 명시 — 매일 morning 자동 실행이 자연 검증")
    ap.add_argument("--legacy-mode", action="store_true", help="화면 모드 강제 (회귀 fallback). --api-mode를 끄고 jQuery.ajax 흐름 사용. 운영 chain 회귀 시만 사용")
    ap.add_argument("--no-jobsetup", action="store_true", help="morning SP3M3 종료 후 잡셋업 자동 호출 차단 (chain 비활성)")
    ap.add_argument("--day-cut", type=int, default=DAY_CUT_THRESHOLD, help=f"morning SP3M3 주간 누적 컷 임계 (default={DAY_CUT_THRESHOLD}). PoC/특수 케이스만 변경")
    ap.add_argument("--limit", type=int, default=None, help="dedupe 후 등록 대상 N건만 사용 (1건 PoC용, default=전체)")
    ap.add_argument("--exclude", default="", help="phase1 추출 후 제외할 PROD_NO 콤마구분 (E 모드 복구용. 예: --exclude RSP3SC0246). P-BOM 미등록 등 단건 차단으로 전체 fail 시 사용")
    ap.add_argument("--no-pbom-guard", action="store_true", help="P-BOM guard 비활성화. Phase5 sendMesFlag=N preflight 및 자동 제외를 건너뜀")
    ap.add_argument("--http-upload", action="store_true", help="세션153 A안 2단계: phase3 D0 업로드를 requests 직접 (브라우저·iframe·jQuery 0). HTTP OAuth 성공 시만 활성")
    ap.add_argument("--http-only", action="store_true", help="세션153 A안 3단계: 완전 브라우저-less. phase0~6 전부 requests. ensure_chrome_cdp + playwright 0")
    ap.add_argument("--jobsetup-mode", choices=["list-only","dry-run","commit-one","commit-all"], default="commit-all",
                    help="chain에서 호출할 잡셋업 모드 (default=commit-all). 입회 monitoring 시 list-only 권장")
    args = ap.parse_args()

    # --legacy-mode 명시 시 api_mode를 False로 강제 (화면 모드 fallback)
    if args.legacy_mode:
        args.api_mode = False
        print("[mode] --legacy-mode 활성 — 화면 모드(jQuery.ajax) 강제. 하이브리드 미사용")

    # 라인 잠금 체크 (사용자 명시 보류 — 락 파일 존재 시 해당 라인 차단)
    _apply_line_locks(args)

    session = determine_session(args.session)
    print(f"=== /d0-plan session={session} line={args.line} dry_run={args.dry_run} ===")

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if args.target_date:
        target_file_date = datetime.strptime(args.target_date, "%Y-%m-%d")
    else:
        target_file_date = today + timedelta(days=1) if session == "evening" else today
    print(f"파일명 날짜(target_file_date) = {target_file_date.strftime('%Y-%m-%d')}")

    # Phase 0 — HTTP OAuth 우선 (세션153 A안 1단계), 실패 시 브라우저 page.fill fallback
    sess = erp_login_via_http()
    if sess is None:
        print("[phase0] HTTP OAuth 실패 — Chrome page.fill 경로 fallback")
    # 세션153 A안 2단계: --http-upload 활성 + sess 가능 → phase3을 requests로 처리
    if args.http_upload and sess is not None:
        global _HTTP_UPLOAD_SESS
        _HTTP_UPLOAD_SESS = sess
        print("[phase3] --http-upload 활성 — d0_upload requests 경로 사용")

    # 세션153 A안 3단계: --http-only 활성 + sess 가능 → phase0~6 전부 requests
    if args.http_only and sess is not None:
        global _HTTP_ONLY_SESS
        _HTTP_ONLY_SESS = sess
        print("[phase0] --http-only 활성 — 브라우저 launch 스킵, 완전 HTTP path")
        return _main_http_only(args, sess, session, target_file_date)

    ensure_chrome_cdp()
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(CDP_URL)
        page = navigate_to_d0(browser, sess=sess)

        # --xlsx 직접 사용 모드: Phase 1 추출 건너뛰고 그대로 업로드
        if args.xlsx:
            xlsx_path = Path(args.xlsx)
            if not xlsx_path.exists():
                raise FileNotFoundError(f"--xlsx 파일 없음: {xlsx_path}")
            items = _load_items_from_xlsx(xlsx_path)
            prod_date = _extract_prod_date(items) or target_file_date
            line_override = args.line if args.line != "ALL" else "SP3M3"
            print(f"[direct] xlsx={xlsx_path.name} items={len(items)} line={line_override} prod_date={prod_date.strftime('%Y-%m-%d')}")
            # 파일 업로드 전 중복 가드 (세션133 사용자 명시 — 같은 prod_date+PROD_NO 이미 등록된 건 제외)
            items = apply_manual_exclude(items, args.exclude, "direct")
            items = note_pbom_guard_phase1(items, line_override, args.no_pbom_guard)
            items_before = len(items)
            items = dedupe_existing_registrations(page, items, prod_date, line_override)
            if items_before > 0 and len(items) == 0:
                print(f"[direct] 모든 items가 이미 등록됨 — 업로드 스킵, 종료")
                return
            if args.dry_run:
                print("[direct] dry-run: 엑셀 확인까지만")
                return
            # xlsx 그대로 업로드 (Phase 3) + 서열 배치 (Phase 4) + 최종 저장 (Phase 5)
            if args.skip_upload:
                print("[direct] --skip-upload: Phase 3 건너뜀")
            elif args.parse_only:
                d0_upload(page, xlsx_path, parse_only=True)
                print("[direct] parse-only: 엑셀 첨부 + 서버 파싱까지만. Phase 4/5 미진행")
                return
            else:
                # dedupe로 items 줄었으면 xlsx도 재생성 후 업로드 (selectList가 같은 파일 파싱하므로)
                if len(items) < items_before:
                    new_xlsx = UPLOAD_DIR / f"d0_{line_override}_{prod_date.strftime('%Y%m%d')}_dedup.xlsx"
                    make_upload_xlsx(items, prod_date, new_xlsx)
                    print(f"[direct] dedupe 후 재생성: {new_xlsx.name} ({len(items)}건)")
                    d0_upload(page, new_xlsx)
                else:
                    d0_upload(page, xlsx_path)
            cfg = LINE_CONFIG[line_override]
            if args.api_mode:
                batch = api_rank_batch(page, items, line_override, cfg["save_url"])
                print(f"[direct] api_rank_batch (Phase 4 hybrid): {batch}")
            else:
                batch = rank_batch(page, items, line_override, cfg["save_url"], dry_run=False)
                print(f"[direct] rank_batch: {batch}")
            if batch["failed"] > 0:
                print("[direct] 실패 존재 — 최종 저장 보류"); return
            if args.no_mes_send:
                print("[direct] --no-mes-send: final_save 차단 (sendMesFlag='Y' MES 전송 미실행)")
                print("[direct]   ERP rank 임시저장만 발생. 정리하려면 erp_d0_dedupe.py 또는 화면에서 삭제")
                print("=== /d0-plan --xlsx --no-mes-send 완료 ===")
                return
            final_items = items
            if args.no_pbom_guard:
                print(f"[pbom-guard] disabled (--no-pbom-guard): Phase5 preflight skip line={line_override}")
                final_save(page, cfg["save_url"])
            else:
                _, excluded_pnos = final_save_with_pbom_guard(page, cfg["save_url"], line_override)
                final_items = filter_items_by_prod_nos(items, excluded_pnos)
            if not verify_smartmes(line_override, prod_date, final_items):
                PHASE6_FAILED.append((line_override, prod_date.strftime("%Y-%m-%d")))
            # 세션135: --xlsx 분기에도 chain 호출 (line_override=SP3M3 한정)
            if line_override == "SP3M3" and not args.parse_only and not args.no_jobsetup:
                _run_jobsetup_chain(args.jobsetup_mode, prod_date)
            print("=== /d0-plan --xlsx 완료 ===")
            return

        # Phase 1: 파일
        # 사용자 명시 (세션133): morning/evening 자동 실행 시 해당일 파일 없으면 작업 패스 (알림 0, exit 0)
        # — 토요일/공휴일 등 파일 미생성 케이스 자동 skip
        try:
            plan_path = find_plan_file(target_file_date)
        except FileNotFoundError as e:
            print(f"[skip] 해당일 파일 없음 — 작업 패스: {e}")
            print(f"=== /d0-plan {session} skip (no plan file) ===")
            return  # exit 0 정상 종료. recover/verify_run에서 알림 안 띄움
        wb = openpyxl.load_workbook(plan_path, data_only=True, keep_vba=True)

        if session == "evening":
            # SP3M3 야간: ERP 생산일 = 파일명 날짜 - 1 (야간 시작일 = 오늘)
            # SD9A01 OUTER: ERP 생산일 = 파일명 날짜 (내일)
            prod_date_override = datetime.strptime(args.prod_date, "%Y-%m-%d") if args.prod_date else None
            if args.line in ("SP3M3","ALL"):
                items = extract_sp3m3_night(wb)
                # Phase 1.5: 야간 1~5행 dedupe (주간 등록분과 PROD_NO+수량 일치 시 제외)
                items = dedupe_night_first_5(page, items)
                items = apply_manual_exclude(items, args.exclude, "evening SP3M3")
                items = note_pbom_guard_phase1(items, "SP3M3", args.no_pbom_guard)
                prod_date = prod_date_override if prod_date_override else target_file_date - timedelta(days=1)
                if prod_date_override:
                    print(f"[evening] --prod-date 오버라이드: SP3M3 prod_date = {prod_date.strftime('%Y-%m-%d')} (file={target_file_date.strftime('%Y-%m-%d')})")
                run_session_line(page, wb, "SP3M3", items, prod_date, args.dry_run, verify_prod_date=prod_date, parse_only=args.parse_only, no_mes_send=args.no_mes_send, api_mode=args.api_mode, pbom_guard=not args.no_pbom_guard)
            if args.line in ("SD9A01","ALL"):
                items = extract_outer_d1(wb, "SD9M01")
                items = apply_manual_exclude(items, args.exclude, "evening SD9A01")
                items = note_pbom_guard_phase1(items, "SD9A01", args.no_pbom_guard)
                prod_date = prod_date_override if prod_date_override else target_file_date
                if prod_date_override:
                    print(f"[evening] --prod-date 오버라이드: SD9A01 prod_date = {prod_date.strftime('%Y-%m-%d')}")
                run_session_line(page, wb, "SD9A01", items, prod_date, args.dry_run, verify_prod_date=prod_date, parse_only=args.parse_only, no_mes_send=args.no_mes_send, api_mode=args.api_mode, pbom_guard=not args.no_pbom_guard)
        else:  # morning
            # SP3M3 주간: ERP 생산일 = 파일명 날짜 (당일, 어제 저녁 저장된 파일)
            prod_date = datetime.strptime(args.prod_date, "%Y-%m-%d") if args.prod_date else target_file_date
            if args.prod_date:
                print(f"[morning] --prod-date 오버라이드: prod_date = {prod_date.strftime('%Y-%m-%d')} (file={target_file_date.strftime('%Y-%m-%d')})")
            sp3m3_registered = False
            if args.line in ("SP3M3","ALL"):
                items = extract_sp3m3_day(wb, args.day_cut)
                items = apply_manual_exclude(items, args.exclude, "morning SP3M3")
                items = note_pbom_guard_phase1(items, "SP3M3", args.no_pbom_guard)
                # Phase 1.5: 같은 prod_date 이미 등록된 PROD_NO 제외 (세션133 사용자 명시 — 파일 업로드 전 중복 가드)
                items = dedupe_existing_registrations(page, items, prod_date, "SP3M3")
                if args.limit is not None and len(items) > args.limit:
                    print(f"[morning] --limit {args.limit} 적용: {len(items)}건 → {args.limit}건 (PoC)")
                    items = items[:args.limit]
                if not items:
                    print("[morning] dedupe 후 등록 대상 0건 — 업로드 스킵")
                else:
                    # 2026-05-23 수정: run_session_line 반환값으로 sp3m3_registered 결정 — Phase 4 실패 / Phase 6 불일치 시 False
                    sp3m3_registered = run_session_line(page, wb, "SP3M3", items, prod_date, args.dry_run, verify_prod_date=prod_date, parse_only=args.parse_only, no_mes_send=args.no_mes_send, api_mode=args.api_mode, pbom_guard=not args.no_pbom_guard)
            if args.line == "SD9A01":
                print("[morning] SD9A01은 저녁 세션 전용 — 스킵")

            # 세션135: morning SP3M3 등록 성공 시 잡셋업 자동 호출 (chain).
            # 가드: sp3m3_registered=True (run_session_line이 Phase 3~6 정상 완료 반환) + --no-jobsetup 미사용.
            # sp3m3_registered=False는 dry_run/parse_only/no_mes_send/Phase4실패/Phase6불일치 모두 포괄.
            # 실패는 advisory(morning 자체 종료에 영향 없음).
            if sp3m3_registered and not args.no_jobsetup:
                _run_jobsetup_chain(args.jobsetup_mode, prod_date)

        print("=== /d0-plan 완료 ===")

    # 세션151: phase6 검증 실패 누적 시 exit 2로 verify_run 자동복구 진입 신호.
    # final_save 후 SmartMES 카운트·중복·순서 중 하나라도 어긋나면 활성. 누락 의심 시 사용자 알림.
    if PHASE6_FAILED:
        print(f"[phase6] 검증 실패 누적: {PHASE6_FAILED} — exit 2", file=sys.stderr)
        sys.exit(2)


def _main_http_only(args, sess, session, target_file_date):
    """세션153 A안 3단계 — 완전 브라우저-less main 흐름.

    ensure_chrome_cdp + sync_playwright 0. dedupe/d0_upload/api_rank_batch/final_save 모두 http.
    page=None — page 의존 함수는 _HTTP_ONLY_SESS 보고 자동 분기 (래퍼).
    --xlsx 직접 모드 지원 (2026-05-23 추가): 외부 엑셀 5건 등을 SSKR 템플릿으로 변환 후 http-only Phase 3~6.
    """
    # --xlsx 직접 모드 (http-only) — 사용자 첨부 엑셀에서 PROD_NO/QTY 추출 후 등록
    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"--xlsx 파일 없음: {xlsx_path}")
        items = _load_items_from_xlsx(xlsx_path)
        prod_date = _extract_prod_date(items) or target_file_date
        line_override = args.line if args.line != "ALL" else "SP3M3"
        print(f"[direct:http] xlsx={xlsx_path.name} items={len(items)} line={line_override} prod_date={prod_date.strftime('%Y-%m-%d')}")
        page = None
        items = apply_manual_exclude(items, args.exclude, "direct:http")
        items = note_pbom_guard_phase1(items, line_override, args.no_pbom_guard)
        items_before = len(items)
        if args.no_dedupe:
            print(f"[direct:http] --no-dedupe: dedupe 우회 — items {len(items)}건 그대로 진행 (사용자 명시 중복 등록)")
        else:
            items = dedupe_existing_registrations(page, items, prod_date, line_override)
            if items_before > 0 and len(items) == 0:
                print("[direct:http] 모든 items가 이미 등록됨 — 업로드 스킵, 종료")
                print("=== /d0-plan --xlsx --http-only 완료 ===")
                return
        if args.dry_run:
            print(f"[direct:http] dry-run: dedupe 후 등록 대상 {len(items)}건")
            for it in items:
                print(f"  - {it['PROD_NO']} QTY={it['QTY']}")
            print("=== /d0-plan --xlsx --http-only 완료 (dry-run) ===")
            return
        # --skip-upload 가드 (2026-05-23 추가):
        # 기존 브라우저 --xlsx 분기는 skip_upload면 Phase 3 d0_upload 건너뛰고 Phase 4부터 진행했으나,
        # http-only 경로는 run_session_line이 d0_upload_via_http를 무조건 호출 → 같은 xlsx가 중복 업로드되는 위험.
        # 안전 차단: --xlsx --http-only --skip-upload 조합은 명시적 미지원으로 종료. Phase 4부터 진행이 필요하면
        # --xlsx 빼고 --skip-upload만으로 morning/evening 자동 흐름을 사용하거나, 브라우저 모드(--http-only 미사용)로 실행.
        if args.skip_upload:
            print("[direct:http] --xlsx --http-only --skip-upload 조합은 미지원 — d0_upload_via_http가 무조건 호출되어 중복 업로드 위험. 작업 중단", file=sys.stderr)
            sys.exit(2)
        # run_session_line 위임 — http-only 분기(d0_upload_via_http / api_rank_batch_via_http / final_save_via_http / verify_smartmes) 자동 처리
        # 반환값 True = Phase 3~6 전부 정상 (jobsetup 가능). False = Phase 4 실패 / Phase 6 SmartMES 불일치 / parse_only / no_mes_send / dry_run 등 비완료
        run_ok = run_session_line(page, None, line_override, items, prod_date, dry_run=False,
                                  verify_prod_date=prod_date, parse_only=args.parse_only,
                                  no_mes_send=args.no_mes_send, api_mode=args.api_mode,
                                  pbom_guard=not args.no_pbom_guard)
        if line_override == "SP3M3" and run_ok and not args.no_jobsetup:
            _run_jobsetup_chain(args.jobsetup_mode, prod_date)
        elif line_override == "SP3M3" and not run_ok:
            print(f"[direct:http] run_session_line 미완료(Phase 4 실패 / Phase 6 불일치 / parse_only / no_mes_send) — jobsetup chain 차단", file=sys.stderr)
        print("=== /d0-plan --xlsx --http-only 완료 ===")
        if PHASE6_FAILED:
            print(f"[phase6] 검증 실패 누적: {PHASE6_FAILED} — exit 2", file=sys.stderr)
            sys.exit(2)
        return
    try:
        plan_path = find_plan_file(target_file_date)
    except FileNotFoundError as e:
        print(f"[skip:http] 해당일 파일 없음 — 작업 패스: {e}")
        print(f"=== /d0-plan {session} skip (no plan file) http-only ===")
        return
    wb = openpyxl.load_workbook(plan_path, data_only=True, keep_vba=True)
    page = None

    if session == "evening":
        prod_date_override = datetime.strptime(args.prod_date, "%Y-%m-%d") if args.prod_date else None
        if args.line in ("SP3M3", "ALL"):
            items = extract_sp3m3_night(wb)
            items = dedupe_night_first_5(page, items)
            items = apply_manual_exclude(items, args.exclude, "evening:http SP3M3")
            items = note_pbom_guard_phase1(items, "SP3M3", args.no_pbom_guard)
            prod_date = prod_date_override if prod_date_override else target_file_date - timedelta(days=1)
            if prod_date_override:
                print(f"[evening:http] --prod-date 오버라이드: SP3M3 prod_date = {prod_date.strftime('%Y-%m-%d')}")
            if args.limit is not None and len(items) > args.limit:
                print(f"[evening:http] --limit {args.limit} 적용: {len(items)}건 → {args.limit}건")
                items = items[:args.limit]
            run_session_line(page, wb, "SP3M3", items, prod_date, args.dry_run,
                             verify_prod_date=prod_date, parse_only=args.parse_only,
                             no_mes_send=args.no_mes_send, api_mode=args.api_mode,
                             pbom_guard=not args.no_pbom_guard)
        if args.line in ("SD9A01", "ALL"):
            items = extract_outer_d1(wb, "SD9M01")
            items = apply_manual_exclude(items, args.exclude, "evening:http SD9A01")
            items = note_pbom_guard_phase1(items, "SD9A01", args.no_pbom_guard)
            prod_date = prod_date_override if prod_date_override else target_file_date
            if args.limit is not None and len(items) > args.limit:
                items = items[:args.limit]
            run_session_line(page, wb, "SD9A01", items, prod_date, args.dry_run,
                             verify_prod_date=prod_date, parse_only=args.parse_only,
                             no_mes_send=args.no_mes_send, api_mode=args.api_mode,
                             pbom_guard=not args.no_pbom_guard)
    else:  # morning
        prod_date = datetime.strptime(args.prod_date, "%Y-%m-%d") if args.prod_date else target_file_date
        sp3m3_registered = False
        if args.line in ("SP3M3", "ALL"):
            items = extract_sp3m3_day(wb, args.day_cut)
            items = apply_manual_exclude(items, args.exclude, "morning:http SP3M3")
            items = note_pbom_guard_phase1(items, "SP3M3", args.no_pbom_guard)
            items = dedupe_existing_registrations(page, items, prod_date, "SP3M3")
            if args.limit is not None and len(items) > args.limit:
                print(f"[morning:http] --limit {args.limit} 적용: {len(items)}건 → {args.limit}건 (PoC)")
                items = items[:args.limit]
            if not items:
                print("[morning:http] dedupe 후 등록 대상 0건 — 업로드 스킵")
            else:
                # 2026-05-23 수정: run_session_line 반환값으로 결정 — Phase 4 실패 / Phase 6 불일치 시 False → jobsetup 차단
                sp3m3_registered = run_session_line(page, wb, "SP3M3", items, prod_date, args.dry_run,
                                                    verify_prod_date=prod_date, parse_only=args.parse_only,
                                                    no_mes_send=args.no_mes_send, api_mode=args.api_mode,
                                                    pbom_guard=not args.no_pbom_guard)
        if args.line == "SD9A01":
            print("[morning:http] SD9A01은 저녁 세션 전용 — 스킵")

        # sp3m3_registered=False는 dry_run/parse_only/no_mes_send/Phase4실패/Phase6불일치 모두 포괄
        if sp3m3_registered and not args.no_jobsetup:
            _run_jobsetup_chain(args.jobsetup_mode, prod_date)

    print("=== /d0-plan http-only 완료 ===")

    if PHASE6_FAILED:
        print(f"[phase6] 검증 실패 누적: {PHASE6_FAILED} — exit 2", file=sys.stderr)
        sys.exit(2)


def _run_jobsetup_chain(mode: str, prod_date):
    """morning SP3M3 종료 후 잡셋업 v3.x 자동 호출 (chain). advisory — 실패해도 morning 정상 종료."""
    import subprocess
    from pathlib import Path as _P
    script = _P(__file__).resolve().parent.parent / "jobsetup-auto" / "run_jobsetup.py"
    if not script.exists():
        print(f"[jobsetup-chain] script not found: {script} — skip", flush=True)
        return
    prdt_da = prod_date.strftime("%Y%m%d")
    cmd = [sys.executable, str(script), "--mode", mode, "--auto-resolve-pno", "--prdt-da", prdt_da, "--line-cd", "SP3M3"]
    print(f"[jobsetup-chain] {' '.join(cmd)}", flush=True)
    try:
        r = subprocess.run(cmd, timeout=180, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if r.stdout:
            print(r.stdout, flush=True)
        if r.stderr:
            print(r.stderr, flush=True, file=sys.stderr)
        if r.returncode == 0:
            print(f"[jobsetup-chain] OK rc=0", flush=True)
        else:
            print(f"[jobsetup-chain] FAIL rc={r.returncode} (advisory — morning 종료에 영향 없음)", flush=True, file=sys.stderr)
    except subprocess.TimeoutExpired:
        print("[jobsetup-chain] TIMEOUT 180s (advisory — morning 종료에 영향 없음)", flush=True, file=sys.stderr)
    except Exception as e:
        print(f"[jobsetup-chain] ERROR {e!r} (advisory — morning 종료에 영향 없음)", flush=True, file=sys.stderr)


if __name__ == "__main__":
    main()
