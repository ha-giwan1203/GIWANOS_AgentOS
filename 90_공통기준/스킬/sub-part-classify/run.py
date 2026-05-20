# -*- coding: utf-8 -*-
"""sub-part-classify — 라인별 서브품 운영 LIST 자동 채움.

사용법:
    python run.py --line SP3M3
    python run.py --line SD9A01

룰: SKILL.md 참조.
"""
import argparse, copy, datetime, glob, json, os, re, shutil, sys, io
from collections import defaultdict

if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

LIST_PATH = r"C:\Users\User\Desktop\서브품 운영 대상 LIST.xlsx"
SP3M3_REF_DIR = r"C:\Users\User\Desktop\업무리스트\04_생산계획"
SD9_REF = r"\\210.216.217.180\zz-group\15. SP3 메인 CAPA점검\SP3M3\SD9M01 라인 계획표.xlsm"


def is_valid(v):
    if v is None: return False
    s = str(v).strip()
    if not s or s.upper() == 'X' or s.startswith('#'): return False
    return True


def fmt_chassis(c2):
    """차종 옵션2 — 줄바꿈 앞 첫 단어 + 괄호 이후 제거."""
    if not c2: return ""
    code = str(c2).split('\n', 1)[0].strip()
    code = code.split()[0] if code else ''
    code = re.sub(r'\(.*$', '', code).strip()
    return code


def copy_style(src, dst):
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)


def backup(path):
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out = f"{path}.bak_{ts}"
    shutil.copy2(path, out)
    return out


def find_latest_sp3m3_ref():
    """가장 최근 SP3M3 생산지시서 xlsm."""
    files = glob.glob(os.path.join(SP3M3_REF_DIR, "SP3M3_생산지시서_(*).xlsm"))
    files = [f for f in files if not f.endswith('.bak') and '.bak_' not in f]
    if not files:
        raise FileNotFoundError("SP3M3 생산지시서 xlsm not found")
    return max(files, key=os.path.getmtime)


SP3M3_MAPPING = {
    # (LIST 시트 NO, C9 키, 차종 행, 품번 행)
    "SSP3-0023": "NLR 흰색",
    "SSP3-0024": "CLR 노랑",
    "SSP3-0026": "NLR 노랑",
    "SSP3-0053": "P/T CLR흰색",
    "SSP3-0054": "P/T NLR흰색",
    "SSP3-0055": "P/T CLR노랑",
    "SSP3-0056": "P/T NLR노랑",
    "SSP3-0162": "P/T CLR노랑\n(MDS)",
    # SMVO-0032 = SSP3-0162와 동일
}


def run_sp3m3(dry=False):
    """SP3M3 — 비고 키워드 → C9 정확 일치 매핑 + 모듈품번 1:1 입력."""
    ref = find_latest_sp3m3_ref()
    print(f"REF: {ref}")
    wb_ref = openpyxl.load_workbook(ref, data_only=True, keep_vba=False)
    ws_ref = wb_ref["SP3 LINE 기준 정보"]

    rows = []
    for r in range(3, ws_ref.max_row + 1):
        pn = ws_ref.cell(r, 3).value
        mod = ws_ref.cell(r, 5).value
        c2 = ws_ref.cell(r, 2).value
        c9 = ws_ref.cell(r, 9).value
        if pn is None and mod is None:
            continue
        rows.append({"r": r, "part": pn, "mod": mod, "c2": c2, "c9": c9})

    # 1) SSP3 SUB 적용범위 매핑 (NO2~10 사용자 사양 행)
    sub_result = {}
    for sub, c9k in SP3M3_MAPPING.items():
        matched = [x for x in rows if x['c9'] == c9k]
        sub_result[sub] = {
            "chassis": ",".join(sorted(set(fmt_chassis(x['c2']) for x in matched if fmt_chassis(x['c2'])))),
            "parts": ",".join(sorted(set(str(x['part']).strip() for x in matched if is_valid(x['part'])))),
            "count": len(matched),
        }
    sub_result["SMVO-0032"] = sub_result["SSP3-0162"]

    # 2) 모듈품번 1:1 (NO11~)
    bulk = []
    for x in rows:
        if not is_valid(x['mod']):
            continue
        bulk.append({
            "module": str(x['mod']).strip(),
            "part": str(x['part']).strip() if is_valid(x['part']) else "",
            "chassis": fmt_chassis(x['c2']),
        })
    print(f"SUB 매핑: {len(sub_result)}개 / 모듈품번 1:1: {len(bulk)}개")

    if dry:
        for s, v in sub_result.items():
            print(f"  {s}: chassis={v['chassis']} parts({v['count']})")
        print(f"  bulk first: {bulk[:3]}")
        return

    bk = backup(LIST_PATH)
    print(f"BACKUP: {bk}")
    wb = openpyxl.load_workbook(LIST_PATH, data_only=False)
    ws = wb["대원테크_SP3M3"]

    # SUB 사양 행 좌표 (NO2~10): NO2=R13/14, NO3=R18/19, ...
    sub_rows = {
        "SSP3-0023": (13, 14), "SSP3-0024": (18, 19), "SSP3-0026": (23, 24),
        "SSP3-0053": (28, 29), "SSP3-0054": (33, 34), "SSP3-0055": (38, 39),
        "SSP3-0056": (43, 44), "SSP3-0162": (48, 49), "SMVO-0032": (53, 54),
    }
    for sub, (rc, rp) in sub_rows.items():
        v = sub_result[sub]
        ws.cell(rc, 4).value = v["chassis"]
        ws.cell(rp, 4).value = v["parts"]

    # 모듈품번 1:1 (NO11=R80부터)
    NO_START, ROW_START = 11, 80
    TEMPLATE_START = 100  # NO15 템플릿
    template_merges = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= TEMPLATE_START and mr.max_row <= TEMPLATE_START + 4:
            template_merges.append((mr.min_row - TEMPLATE_START, mr.max_row - TEMPLATE_START,
                                    mr.min_col, mr.max_col))
    for i, d in enumerate(bulk):
        no = NO_START + i
        sr = ROW_START + i * 5
        ws.cell(sr, 2).value = no
        ws.cell(sr, 3).value = "품번"
        ws.cell(sr, 4).value = d["module"]
        ws.cell(sr + 1, 3).value = "품명"
        ws.cell(sr + 1, 4).value = "메인SUB ASSY"
        ws.cell(sr + 2, 3).value = "적입수량"
        ws.cell(sr + 2, 4).value = 10
        ws.cell(sr + 3, 3).value = "적용차종"
        ws.cell(sr + 3, 4).value = d["chassis"]
        ws.cell(sr + 4, 3).value = "적용품번"
        ws.cell(sr + 4, 4).value = d["part"]
        if sr > TEMPLATE_START + 4:
            for j in range(5):
                for c in (2, 3, 4):
                    copy_style(ws.cell(TEMPLATE_START + j, c), ws.cell(sr + j, c))
            for (rs, re_, cs, ce) in template_merges:
                ws.merge_cells(start_row=sr + rs, start_column=cs,
                               end_row=sr + re_, end_column=ce)

    wb.save(LIST_PATH)
    print(f"SAVED: NO{NO_START}~NO{NO_START + len(bulk) - 1}")


def run_sd9a01(dry=False):
    """SD9A01 — M(디링 ASSY) + O(앵카 ASSY) unique 추출."""
    print(f"REF: {SD9_REF}")
    wb_ref = openpyxl.load_workbook(SD9_REF, data_only=True, keep_vba=False)
    ws_ref = wb_ref["기준정보"]

    dring = defaultdict(lambda: {"chassis": set(), "parts": set()})
    enka = defaultdict(lambda: {"chassis": set(), "parts": set()})
    for r in range(2, ws_ref.max_row + 1):
        chassis = fmt_chassis(ws_ref.cell(r, 3).value)
        part = ws_ref.cell(r, 2).value
        d_assy = ws_ref.cell(r, 13).value
        e_assy = ws_ref.cell(r, 15).value
        if is_valid(d_assy):
            k = str(d_assy).strip()
            if chassis: dring[k]["chassis"].add(chassis)
            if is_valid(part): dring[k]["parts"].add(str(part).strip())
        if is_valid(e_assy):
            k = str(e_assy).strip()
            if chassis: enka[k]["chassis"].add(chassis)
            if is_valid(part): enka[k]["parts"].add(str(part).strip())
    print(f"디링 unique: {len(dring)} / 엥카 unique: {len(enka)}")

    entries = []
    for k in sorted(dring.keys()):
        v = dring[k]
        entries.append({"name": "디링 ASSY", "qty": 60, "module": k,
                        "chassis": ",".join(sorted(v["chassis"])),
                        "parts": ",".join(sorted(v["parts"]))})
    for k in sorted(enka.keys()):
        v = enka[k]
        entries.append({"name": "엥카 ASSY", "qty": 120, "module": k,
                        "chassis": ",".join(sorted(v["chassis"])),
                        "parts": ",".join(sorted(v["parts"]))})

    if dry:
        for e in entries[:5]:
            print(f"  {e}")
        print(f"  ... 총 {len(entries)}")
        return

    bk = backup(LIST_PATH)
    print(f"BACKUP: {bk}")
    wb = openpyxl.load_workbook(LIST_PATH, data_only=False)
    ws = wb["대원테크_SD9A01"]

    NO_START, ROW_START = 1, 5
    TEMPLATE_START = 75  # NO15 템플릿 (사전에 NO1~15 미리 작성된 시트)
    template_merges = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= TEMPLATE_START and mr.max_row <= TEMPLATE_START + 4:
            template_merges.append((mr.min_row - TEMPLATE_START, mr.max_row - TEMPLATE_START,
                                    mr.min_col, mr.max_col))
    for i, e in enumerate(entries):
        no = NO_START + i
        sr = ROW_START + i * 5
        ws.cell(sr, 2).value = no
        ws.cell(sr, 3).value = "품번"
        ws.cell(sr, 4).value = e["module"]
        ws.cell(sr + 1, 3).value = "품명"
        ws.cell(sr + 1, 4).value = e["name"]
        ws.cell(sr + 2, 3).value = "적입수량"
        ws.cell(sr + 2, 4).value = e["qty"]
        ws.cell(sr + 3, 3).value = "적용차종"
        ws.cell(sr + 3, 4).value = e["chassis"]
        ws.cell(sr + 4, 3).value = "적용품번"
        ws.cell(sr + 4, 4).value = e["parts"]
        if sr > TEMPLATE_START + 4:
            for j in range(5):
                for c in (2, 3, 4):
                    copy_style(ws.cell(TEMPLATE_START + j, c), ws.cell(sr + j, c))
            for (rs, re_, cs, ce) in template_merges:
                ws.merge_cells(start_row=sr + rs, start_column=cs,
                               end_row=sr + re_, end_column=ce)

    wb.save(LIST_PATH)
    print(f"SAVED: NO{NO_START}~NO{NO_START + len(entries) - 1}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--line", required=True, choices=["SP3M3", "SD9A01"])
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--backup-only", action="store_true")
    args = p.parse_args()

    if args.backup_only:
        print(f"BACKUP ONLY: {backup(LIST_PATH)}")
        return

    if args.line == "SP3M3":
        run_sp3m3(dry=args.dry_run)
    elif args.line == "SD9A01":
        run_sd9a01(dry=args.dry_run)


if __name__ == "__main__":
    main()
