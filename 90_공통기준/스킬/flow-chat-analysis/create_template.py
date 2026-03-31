"""
SP3S03 이슈분석 엑셀 생성 — 설비/품질/기타 시트 분리
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = Path(__file__).parent.parent.parent.parent / "06_생산관리" / "통합" / "SP3S03_이슈분석.xlsx"

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _make_header(ws, headers, widths, color="4472C4"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _add_rows(ws, data, wrap_cols=None):
    wrap_cols = wrap_cols or []
    for ri, row in enumerate(data, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            c.alignment = Alignment(vertical="top", wrap_text=(ci in wrap_cols))


def _add_table(ws, name, cols, rows, style="TableStyleMedium2"):
    ref = f"A1:{get_column_letter(cols)}{rows + 1}"
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(t)


def create():
    wb = Workbook()

    # ===== 설비이슈 =====
    ws_eq = wb.active
    ws_eq.title = "설비이슈"
    eq_headers = ["날짜", "시간", "작성자", "중분류", "소분류", "설비/공정명",
                  "내용요약", "조치사항", "비가동(분)", "출처"]
    eq_widths = [12, 8, 18, 10, 12, 15, 40, 30, 10, 8]
    _make_header(ws_eq, eq_headers, eq_widths, "2F5496")

    eq_data = [
        ["2026-03-28", "08:28", "김민건_대원테크", "정지", "라인정지", "", "라인정지 + 선별 진행", "", "", "CHAT"],
        ["2026-03-28", "15:12", "김민건_대원테크", "보전", "예방보전", "유압기", "코킹/보더링/백업지그 유압기 누유방지 + 설비점검", "사전 조치 완료", "", "CHAT"],
        ["2026-03-28", "15:48", "김민건_대원테크", "고장", "비전불량", "비전검사 카메라", "홀더 CLR 비전검사 카메라 교체 + 세팅", "교체 완료", "", "CHAT"],
        ["2026-03-28", "15:51", "김민건_대원테크", "수리", "교체", "유압 펌프", "유압 펌프 오일 탱크 교체 + 오일 13통 주유", "교체 완료", "", "CHAT"],
        ["2026-03-28", "18:11", "", "수리", "교체", "유압실린더", "유압실린더 교체", "교체 완료", "", "CHAT"],
        ["2026-03-30", "08:20", "김대성_대원테크", "정지", "설비정지", "락킹인덱스", "락킹인덱스 #1→#2 공급잡음 → 라인정지", "모니터링", "", "CHAT"],
        ["2026-03-30", "13:01", "김대성_대원테크", "정지", "라인정지", "홀더휠 높이검사", "홀더휠 높이검사공정 라인정지", "", "", "CHAT"],
        ["2026-03-30", "13:33", "김대성_대원테크", "수리", "교체", "LVDT", "LVDT 교체 + 케이블 재연결 → 재가동", "교체 후 재가동", "", "CHAT"],
    ]
    _add_rows(ws_eq, eq_data, wrap_cols=[7, 8])
    _add_table(ws_eq, "설비이슈", len(eq_headers), len(eq_data))

    # ===== 품질이슈 =====
    ws_qt = wb.create_sheet("품질이슈")
    qt_headers = ["날짜", "시간", "작성자", "중분류", "소분류", "품번/라인",
                  "내용요약", "조치사항", "불량수량", "출처"]
    qt_widths = [12, 8, 18, 10, 12, 15, 40, 30, 10, 8]
    _make_header(ws_qt, qt_headers, qt_widths, "C00000")

    qt_data = [
        ["2026-03-28", "08:26", "김민건_대원테크", "검사", "판정불일치", "", "리노텍 검사 미수행 지적", "", "", "CHAT"],
        ["2026-03-28", "08:27", "김민건_대원테크", "조립", "누락", "", "부쉬 누락/이탈/압입 불량 묵인 중단 선언", "", "", "CHAT"],
        ["2026-03-28", "09:56", "김민건_대원테크", "조립", "미삽", "4호기 PLL 8PI", "커버 플레이트 락 미조립 혼입", "선별 지시", "", "CHAT"],
        ["2026-03-28", "08:26", "김민건_대원테크", "검사", "원인미확정", "", "야간부터 사용된 불량품 지적", "", "", "CHAT"],
    ]
    _add_rows(ws_qt, qt_data, wrap_cols=[7, 8])
    _add_table(ws_qt, "품질이슈", len(qt_headers), len(qt_data))

    # ===== 기타이슈 =====
    ws_etc = wb.create_sheet("기타이슈")
    etc_headers = ["날짜", "시간", "작성자", "중분류", "내용요약", "영향범위", "출처"]
    etc_widths = [12, 8, 18, 12, 40, 20, 8]
    _make_header(ws_etc, etc_headers, etc_widths, "548235")

    etc_data = [
        ["2026-03-30", "18:40", "김대성_대원테크", "자재부족", "프레임 2종(SSP3-0032E1, SC30-0001E1) 재고부족 → 생산취소", "금일 생산 취소", "CHAT"],
        ["2026-03-31", "06:45", "김민건_대원테크", "자재부족", "스프링하우징 재고부족 → 06:20 라인마감", "라인 마감", "CHAT"],
        ["2026-03-31", "19:40", "김민건_대원테크", "운영공지", "야간 MES 모듈 품번 스케줄 등록 + 라인 가동", "", "CHAT"],
    ]
    _add_rows(ws_etc, etc_data, wrap_cols=[5, 6])
    _add_table(ws_etc, "기타이슈", len(etc_headers), len(etc_data))

    # ===== 설정 =====
    ws_cfg = wb.create_sheet("설정")
    cfg = {
        "A": ("설비_중분류", ["고장", "수리", "정지", "에러", "보전", "기구"]),
        "B": ("품질_중분류", ["외관", "조립", "검사", "클레임", "원인미확정"]),
        "C": ("기타_중분류", ["자재부족", "결품", "운영공지", "생산계획"]),
        "D": ("출처", ["CHAT", "MES", "EQUIP"]),
    }
    for col, (title, items) in cfg.items():
        ws_cfg[f"{col}1"] = title
        ws_cfg[f"{col}1"].font = Font(bold=True)
        for i, item in enumerate(items, 2):
            ws_cfg[f"{col}{i}"] = item
        ws_cfg.column_dimensions[col].width = 15

    # 저장
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"[DONE] {OUTPUT_PATH}")
    print(f"  설비이슈 {len(eq_data)}건 / 품질이슈 {len(qt_data)}건 / 기타 {len(etc_data)}건")


if __name__ == "__main__":
    create()
