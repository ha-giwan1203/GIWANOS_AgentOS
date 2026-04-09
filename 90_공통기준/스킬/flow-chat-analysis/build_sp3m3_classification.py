from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
SOURCE_CSV = OUTPUT_DIR / "sp3m3_integrated_claim_register.csv"

MASTER_CSV = OUTPUT_DIR / "sp3m3_issue_classified_master.csv"
FACILITY_CSV = OUTPUT_DIR / "sp3m3_facility_issues.csv"
QUALITY_CSV = OUTPUT_DIR / "sp3m3_quality_issues.csv"
MATERIAL_CSV = OUTPUT_DIR / "sp3m3_material_issues.csv"
REVIEW_CSV = OUTPUT_DIR / "sp3m3_review_priority.csv"
REVIEW_TOP_CSV = OUTPUT_DIR / "sp3m3_review_top30.csv"
WORKBOOK_XLSX = OUTPUT_DIR / "sp3m3_issue_management_workbook.xlsx"
OVERRIDES_CSV = BASE_DIR / "classification_overrides.csv"


FACILITY_RULES = [
    ("센서", "감지/배선", ["센서", "감지", "단선"]),
    ("실린더", "실린더/에어", ["실린더", "에어밸브", "에어", "흡착", "진공"]),
    ("비전", "카메라/조명", ["카메라", "렌즈", "조명", "포커스"]),
    ("검사설비", "비전/검사설비", ["비전 ng", "비젼ng", "비전검사", "비젼검사"]),
    ("검사설비", "높이/치수검사", ["높이검사", "치수검사", "작동력 검사"]),
    ("서보/모터", "구동계", ["서보", "모터", "원점 상실", "축 이상"]),
    ("지그/치공구", "지그/고정구", ["지그", "고정", "치공구", "척/언척"]),
    ("공급/로더", "투입/배출", ["로더", "언로더", "투입", "배출", "공급부", "로딩"]),
    ("인덱스/회전", "구동 이상", ["인덱스", "회전", "복귀"]),
    ("배관/윤활", "배관/구리스", ["배관", "구리스", "분배기"]),
    ("압입/하중", "압입 조건", ["압입", "하중"]),
    ("설정조건", "모델/사양 세팅", ["셋팅", "세팅", "모델설정", "사양변경", "조건 불일치"]),
    ("설비파손", "기구부 파손", ["파손", "절단", "충돌"]),
    ("구동계", "슬라이드/드라이버", ["슬라이드", "드라이버", "전진이상"]),
    ("흡착장치", "흡착/진공", ["흡착 불량", "진공"]),
    ("기타설비", "알람/잡음", ["이상 알람", "잡음솔"]),
    ("구동계", "이송/후진", ["후진이상", "이송 불량", "이송불량", "푸셔"]),
]

QUALITY_RULES = [
    ("조립불량", "조립 품질", ["조립불량", "미조립", "조립 지연", "조립 불량"]),
    ("검사NG", "검사 판정", ["검사 ng", "검사ng", "ng 지속", "연속 ng", "판정 ng"]),
    ("누락/이탈", "누락/이탈", ["누락", "이탈"]),
    ("외관/치수", "외관/치수", ["외관", "치수", "휨", "변형"]),
    ("토크/작동력", "기능 검사", ["토크", "작동력"]),
    ("롤러/기어", "부품 조립", ["롤러", "기어", "스풀", "spool"]),
    ("체결불량", "체결 품질", ["체결 불량", "체결불량"]),
    ("조립불량", "각도/위치", ["조립각도", "위치 불량"]),
    ("외관/치수", "버어/휨", ["burr", "휨", "바서포트", "파이프불량"]),
    ("검사NG", "락검사/기능검사", ["락검사 불량", "기능 검사"]),
    ("외관/치수", "미성형/찍힘", ["미성형", "찍힘"]),
]

MATERIAL_RULES = [
    ("자재공급", "공급 지연", ["자재", "도착", "입고지연", "재고부족"]),
    ("품번/사양", "사양 불일치", ["품번", "사양", "서브품번", "sub"]),
]

FACILITY_HINTS = {
    "센서",
    "실린더",
    "에어",
    "에어밸브",
    "카메라",
    "렌즈",
    "조명",
    "비전",
    "비젼",
    "서보",
    "모터",
    "원점",
    "로더",
    "투입",
    "배출",
    "인덱스",
    "회전",
    "복귀",
    "지그",
    "척",
    "배관",
    "구리스",
    "분배기",
    "흡착",
    "파손",
    "절단",
    "단선",
    "드라이버",
    "슬라이드",
    "전진이상",
    "알람",
}

QUALITY_HINTS = {
    "조립불량",
    "검사 ng",
    "검사ng",
    "연속 ng",
    "불량",
    "누락",
    "이탈",
    "휨",
    "변형",
    "토크",
    "작동력",
    "높이검사",
    "체결 불량",
    "조립각도",
    "burr",
    "휨",
    "파이프불량",
}

MATERIAL_HINTS = {
    "자재",
    "도착",
    "입고지연",
    "재고부족",
    "품번",
    "서브품번",
}

MES_CATEGORY_MAP = {
    "설비이상": "설비",
    "품질이상": "품질",
    "자재이상": "자재",
}

ERP_CATEGORY_MAP = {
    "설비이상": "설비",
    "품질이상": "품질",
    "자재이상": "자재",
}


def normalize_text(value: str) -> str:
    text = (value or "").replace("\r", " ").replace("\n", " ").strip().lower()
    return " ".join(text.split())


def text_contains_any(text: str, keywords: set[str] | list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def detect_category_from_text(text: str) -> str | None:
    has_facility = text_contains_any(text, FACILITY_HINTS)
    has_quality = text_contains_any(text, QUALITY_HINTS)
    has_material = text_contains_any(text, MATERIAL_HINTS)

    if has_material and not has_facility:
        return "자재"
    if has_facility and not has_quality:
        return "설비"
    if has_quality and not has_facility:
        return "품질"
    if has_facility and has_quality:
        if "비전검사" in text or "비젼검사" in text or "카메라" in text or "센서" in text:
            return "설비"
        return "품질"
    return None


def detect_confident_text_category(text: str) -> str | None:
    if text_contains_any(text, {"조립불량", "체결 불량", "작동 불량", "토크", "조립각도", "휨 불량", "burr", "검사 ng", "검사ng", "락검사 불량", "미성형", "찍힘"}):
        return "품질"
    if text_contains_any(text, {"재고부족", "계획변경", "입고지연"}):
        return "자재"
    if text_contains_any(text, {"서보 드라이버", "드라이버 이상", "원점 상실", "센서 이상", "흡착 불량", "슬라이드", "후진이상", "이송 불량", "이송불량", "푸셔", "라인정지"}):
        return "설비"
    return None


def classify_category(stop_reason: str, text: str, mes_major: str) -> str:
    erp_category = ERP_CATEGORY_MAP.get((stop_reason or "").strip(), "")
    mes_category = MES_CATEGORY_MAP.get((mes_major or "").strip(), "")
    text_category = detect_category_from_text(text) or ""

    if erp_category == "설비" and text_category in {"품질", "자재"}:
        return text_category
    if erp_category:
        return erp_category
    if mes_category:
        return mes_category
    if text_category:
        return text_category
    return "기타"


def classify_detail(category: str, text: str) -> tuple[str, str]:
    rules = {
        "설비": FACILITY_RULES,
        "품질": QUALITY_RULES,
        "자재": MATERIAL_RULES,
    }.get(category, [])
    for middle, small, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return middle, small
    fallback = {
        "설비": ("기타설비", "기타"),
        "품질": ("기타품질", "기타"),
        "자재": ("기타자재", "기타"),
        "기타": ("기타", "기타"),
    }
    return fallback[category]


def to_number(raw: str) -> float:
    try:
        return float(str(raw).replace(",", "").strip() or 0)
    except ValueError:
        return 0.0


def detail_confidence(category: str, middle: str, text: str) -> bool:
    if category == "기타":
        return False
    if middle.startswith("기타"):
        return False
    if len(text) < 8:
        return False
    return True


def review_level(row: dict[str, str], category: str, middle: str, small: str, text: str) -> tuple[str, str]:
    reasons: list[str] = []
    mes_reg_no = (row.get("mes_reg_no") or "").strip()
    mes_score = int((row.get("mes_match_score") or "0").strip() or "0")
    flow1_score = int((row.get("flow1_score") or "0").strip() or "0")
    cost = to_number(row.get("cost_bill_tot", "0"))
    minutes = to_number(row.get("total_minute", "0"))
    erp_category = ERP_CATEGORY_MAP.get((row.get("stop_reason") or "").strip(), "")
    mes_category = MES_CATEGORY_MAP.get((row.get("mes_issue_major") or "").strip(), "")
    confident = detail_confidence(category, middle, text)
    text_category = detect_confident_text_category(text) or category

    if category == "기타" or middle.startswith("기타"):
        reasons.append("분류애매")
    if erp_category and mes_category and erp_category != mes_category and not (confident and text_category == category):
        reasons.append("ERP-MES분류상충")
    if not mes_reg_no and not confident:
        reasons.append("MES미매칭")
    if mes_reg_no and mes_score <= 0 and not confident:
        reasons.append("MES약매칭")
    if flow1_score <= 0 and not confident and (cost >= 100000 or minutes >= 60):
        reasons.append("Flow근거부족")
    if not mes_reg_no and flow1_score <= 0 and (cost >= 150000 or minutes >= 90):
        reasons.append("고비용/장시간")

    if reasons:
        return "검토필요", ", ".join(dict.fromkeys(reasons))
    return "정상", ""


def load_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_overrides() -> dict[str, dict[str, str]]:
    if not OVERRIDES_CSV.exists():
        return {}
    with OVERRIDES_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    overrides: dict[str, dict[str, str]] = {}
    for row in rows:
        claim_no = (row.get("청구번호") or "").strip()
        if not claim_no:
            continue
        overrides[claim_no] = {
            "대분류": (row.get("대분류") or "").strip(),
            "중분류": (row.get("중분류") or "").strip(),
            "소분류": (row.get("소분류") or "").strip(),
            "판정근거": (row.get("판정근거") or "").strip(),
        }
    return overrides


def transform_rows(rows: list[dict[str, str]], overrides: dict[str, dict[str, str]] | None = None) -> list[dict[str, str]]:
    overrides = overrides or {}
    transformed: list[dict[str, str]] = []
    for row in rows:
        text_pool = " ".join(
            [
                row.get("issue_detail", ""),
                row.get("accept_note", ""),
                row.get("approval_desc", ""),
                row.get("mes_issue_major", ""),
                row.get("mes_issue_minor", ""),
                row.get("flow1_text", ""),
                row.get("flow2_text", ""),
            ]
        )
        normalized = normalize_text(text_pool)
        category = classify_category(row.get("stop_reason", ""), normalized, row.get("mes_issue_major", ""))
        middle, small = classify_detail(category, normalized)
        review_status, review_reason = review_level(row, category, middle, small, normalized)

        item = {
            "발생일자": row.get("event_date", ""),
            "대분류": category,
            "중분류": middle,
            "소분류": small,
            "청구번호": row.get("claim_no", ""),
            "승인상태": row.get("approval_status", ""),
            "라인코드": row.get("line_cd", ""),
            "라인명": row.get("line_nm", ""),
            "공정명": row.get("proc_nm", ""),
            "ERP정지사유": row.get("stop_reason", ""),
            "ERP상세내용": row.get("issue_detail", ""),
            "정지시간(분)": row.get("total_minute", ""),
            "청구금액": row.get("cost_bill_tot", ""),
            "MES등록번호": row.get("mes_reg_no", ""),
            "MES대분류": row.get("mes_issue_major", ""),
            "MES소분류": row.get("mes_issue_minor", ""),
            "MES공정": row.get("mes_proc", ""),
            "MES매칭점수": row.get("mes_match_score", ""),
            "Flow1시간": row.get("flow1_time", ""),
            "Flow1작성자": row.get("flow1_author", ""),
            "Flow1내용": row.get("flow1_text", ""),
            "Flow1점수": row.get("flow1_score", ""),
            "Flow2시간": row.get("flow2_time", ""),
            "Flow2작성자": row.get("flow2_author", ""),
            "Flow2내용": row.get("flow2_text", ""),
            "Flow2점수": row.get("flow2_score", ""),
            "검토상태": review_status,
            "검토사유": review_reason,
        }
        override = overrides.get(item["청구번호"])
        if override:
            item["대분류"] = override["대분류"] or item["대분류"]
            item["중분류"] = override["중분류"] or item["중분류"]
            item["소분류"] = override["소분류"] or item["소분류"]
            item["검토상태"] = "정상"
            item["검토사유"] = override["판정근거"] or "수동확정"

        transformed.append(item)
    return transformed


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        if rows:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)


def autosize(ws) -> None:
    for idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 10), 40)


def write_workbook(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = {
        "원장": rows,
        "설비이슈": [r for r in rows if r["대분류"] == "설비"],
        "품질이슈": [r for r in rows if r["대분류"] == "품질"],
        "자재이슈": [r for r in rows if r["대분류"] == "자재"],
        "검토대상": [r for r in rows if r["검토상태"] == "검토필요"],
    }

    for name, sheet_rows in sheets.items():
        ws = wb.create_sheet(title=name)
        if not sheet_rows:
            ws.append(["데이터 없음"])
            continue
        headers = list(sheet_rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in sheet_rows:
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        autosize(ws)

    code_ws = wb.create_sheet(title="코드관리")
    code_ws.append(["항목", "값", "건수"])
    for cell in code_ws[1]:
        cell.font = Font(bold=True)

    for label in ("대분류", "중분류", "소분류", "검토사유"):
        counts = Counter(r[label] if label != "검토사유" else (r["검토사유"] or "정상") for r in rows)
        for value, count in counts.most_common():
            code_ws.append([label, value, count])
    code_ws.freeze_panes = "A2"
    autosize(code_ws)

    wb.save(WORKBOOK_XLSX)


def sort_review_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def sort_key(row: dict[str, str]) -> tuple[int, float, float]:
        high_impact = 0 if "고비용/장시간" in row["검토사유"] else 1
        return (
            high_impact,
            -to_number(row["청구금액"]),
            -to_number(row["정지시간(분)"]),
        )

    return sorted(rows, key=sort_key)


def main() -> None:
    source_rows = load_rows()
    overrides = load_overrides()
    transformed = transform_rows(source_rows, overrides)

    write_csv(MASTER_CSV, transformed)
    write_csv(FACILITY_CSV, [r for r in transformed if r["대분류"] == "설비"])
    write_csv(QUALITY_CSV, [r for r in transformed if r["대분류"] == "품질"])
    write_csv(MATERIAL_CSV, [r for r in transformed if r["대분류"] == "자재"])
    review_rows = sort_review_rows([r for r in transformed if r["검토상태"] == "검토필요"])
    write_csv(REVIEW_CSV, review_rows)
    write_csv(REVIEW_TOP_CSV, review_rows[:30])
    write_workbook(transformed)

    summary = Counter(r["대분류"] for r in transformed)
    review_count = sum(1 for r in transformed if r["검토상태"] == "검토필요")
    print(
        {
            "master_rows": len(transformed),
            "facility_rows": summary.get("설비", 0),
            "quality_rows": summary.get("품질", 0),
            "material_rows": summary.get("자재", 0),
            "review_rows": review_count,
            "workbook": str(WORKBOOK_XLSX),
        }
    )


if __name__ == "__main__":
    main()
