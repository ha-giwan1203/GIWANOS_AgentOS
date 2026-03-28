#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
생산관리 마스터리스트 자동 생성 스크립트
BI 실적 + 임률단가 + 조립비시스템 API → 생산관리_마스터리스트.xlsx

사용법:
    PYTHONUTF8=1 python build_master.py
"""

import os
import sys
import json
import urllib.request
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ─── 경로 설정 ───────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECT_DIR = os.path.dirname(BASE_DIR)

BI_FILE = os.path.join(BASE_DIR, "BI실적", "대원테크_라인별 생산실적_BI.xlsx")
COST_FILE = os.path.join(PROJECT_DIR, "02_급여단가", "임률단가", "03_대원테크", "임률단가_대원테크_독립계산.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "생산관리_마스터리스트.xlsx")

ASSY_API_URL = "http://ax.samsong.com:33200/api/assembly-cost-operation?"
LINE_API_URL = "http://ax.samsong.com:33200/api/line-operation"

# ─── 서식 상수 ───────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

NUMBER_FMT = '#,##0'
PERCENT_FMT = '0.0%'
DECIMAL_FMT = '#,##0.0'


# ═══════════════════════════════════════════════════════════
#  데이터 로딩
# ═══════════════════════════════════════════════════════════

def load_bi_data():
    """BI 엑셀에서 2026년 데이터만 읽기"""
    print("[1/4] BI 파일 로딩...")
    wb = openpyxl.load_workbook(BI_FILE, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        print("  ERROR: BI 데이터 없음")
        return []

    header = rows[0]
    data = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        rec = dict(zip(header, row))

        # 날짜 필터: 2026년만
        raw_date = rec.get("날짜") or rec.get(header[7])  # H열
        if raw_date is None:
            continue
        if isinstance(raw_date, (datetime, date)):
            dt = raw_date if isinstance(raw_date, date) else raw_date.date()
        elif isinstance(raw_date, (int, float)):
            # 엑셀 시리얼 넘버
            from datetime import timedelta
            dt = (datetime(1899, 12, 30) + timedelta(days=int(raw_date))).date()
        else:
            continue

        if dt.year != 2026:
            continue

        rec["_date"] = dt
        rec["_month"] = f"{dt.year}-{dt.month:02d}"
        data.append(rec)

    print(f"  → 2026년 데이터 {len(data)}행 로딩 완료")
    return data


def load_cost_data():
    """임률단가 엑셀에서 라인별 단가/인건비 정보 읽기"""
    print("[2/4] 임률단가 파일 로딩...")
    wb = openpyxl.load_workbook(COST_FILE, data_only=True)
    ws = wb["업체_단가계산"]

    cost = {}
    for row in ws.iter_rows(min_row=7, max_row=16, values_only=True):
        line_code = row[1]
        if not line_code:
            continue
        cost[str(line_code).strip()] = {
            "라인명": str(row[2] or ""),
            "작업자수": to_num(row[3]),
            "관리자수": to_num(row[4]),
            "실효UPH": to_num(row[5]),
            "작업자임률": to_num(row[6]),
            "관리자임률": to_num(row[7]),
            "총인건비_h": to_num(row[8]),
            "직접노무단가": to_num(row[9]),
            "업체기준최종단가": to_num(row[16]),
        }

    # 라인_입력 시트에서도 보강
    ws2 = wb["라인_입력"]
    for row in ws2.iter_rows(min_row=7, max_row=16, values_only=True):
        lc = str(row[1] or "").strip()
        if lc and lc in cost:
            cost[lc]["입력_작업자수"] = to_num(row[3])
            cost[lc]["입력_관리자수"] = to_num(row[4])
            cost[lc]["입력_실효UPH"] = to_num(row[5])

    wb.close()
    print(f"  → {len(cost)}개 라인 단가 로딩 완료")
    return cost


def load_assy_api_data():
    """조립비현황 API에서 대원테크 2026년 데이터 가져오기"""
    print("[3/4] 조립비시스템 API 호출...")
    try:
        req = urllib.request.Request(ASSY_API_URL)
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = json.loads(resp.read().decode("utf-8"))

        all_data = raw.get("data", [])
        # 대원테크 + 2026년만 필터
        filtered = [
            d for d in all_data
            if d.get("assy_company") == "대원테크"
            and str(d.get("close_date", "")).startswith("2026")
        ]
        print(f"  → 대원테크 2026년 {len(filtered)}건 로딩 완료")
        return filtered
    except Exception as e:
        print(f"  WARNING: API 호출 실패 ({e}) — 정합성 검증 시트 생략")
        return []


def load_line_api_data():
    """라인운영 API에서 대원테크 데이터 가져오기"""
    print("[4/4] 라인운영 API 호출...")
    try:
        req = urllib.request.Request(LINE_API_URL)
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = json.loads(resp.read().decode("utf-8"))

        all_data = raw.get("data", [])
        filtered = [d for d in all_data if d.get("assy_company") == "대원테크"]
        # 라인별로 최신 1건만 (id 기준 역순)
        by_line = {}
        for d in sorted(filtered, key=lambda x: x.get("id", 0), reverse=True):
            lc = d.get("assy_line", "")
            if lc not in by_line:
                by_line[lc] = d
        print(f"  → 대원테크 {len(by_line)}개 라인 로딩 완료")
        return by_line
    except Exception as e:
        print(f"  WARNING: API 호출 실패 ({e}) — 인건비 교차검증 시트 생략")
        return {}


# ═══════════════════════════════════════════════════════════
#  유틸리티
# ═══════════════════════════════════════════════════════════

def to_num(val, default=0):
    """숫자 변환 (None, 빈값 → default)"""
    if val is None or val == "" or val == "-":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_div(a, b, default=0):
    """안전 나눗셈"""
    if not b:
        return default
    return a / b


def get_field(rec, *keys, default=None):
    """레코드에서 여러 키 중 첫 번째 존재하는 값 반환"""
    for k in keys:
        if k in rec and rec[k] is not None:
            return rec[k]
    return default


def write_header(ws, headers, row=1):
    """헤더 행 작성 + 서식"""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_row(ws, row_num, values, formats=None):
    """데이터 행 작성"""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if formats and col <= len(formats) and formats[col - 1]:
            cell.number_format = formats[col - 1]


def auto_width(ws, max_width=25):
    """열 너비 자동 조정"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), max_width)


def apply_conditional_format(ws, col_letter, min_row, max_row, thresholds, reverse=False):
    """조건부서식 적용 (가동효율/인건비비율 등)"""
    cell_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
    if reverse:  # 낮을수록 좋은 경우 (인건비비율)
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="greaterThanOrEqual", formula=[str(thresholds[1])], fill=RED_FILL))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="between", formula=[str(thresholds[0]), str(thresholds[1])], fill=YELLOW_FILL))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="lessThan", formula=[str(thresholds[0])], fill=GREEN_FILL))
    else:  # 높을수록 좋은 경우 (가동효율)
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="greaterThanOrEqual", formula=[str(thresholds[1])], fill=GREEN_FILL))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="between", formula=[str(thresholds[0]), str(thresholds[1])], fill=YELLOW_FILL))
        ws.conditional_formatting.add(cell_range, CellIsRule(
            operator="lessThan", formula=[str(thresholds[0])], fill=RED_FILL))


# ═══════════════════════════════════════════════════════════
#  시트 생성
# ═══════════════════════════════════════════════════════════

def build_daily_sheet(wb, bi_data, cost_data):
    """시트1: 일일실적"""
    print("  시트1: 일일실적...")
    ws = wb.create_sheet("일일실적")
    ws.sheet_properties.tabColor = "4472C4"

    headers = [
        "날짜", "유형", "대표기종", "라인명", "야간구분", "생산인원",
        "근무시간(h)", "비가동(h)", "실가동시간(h)", "생산량(ea)",
        "실적UPH(ea)", "가동효율",
        "업체기준단가(원)", "생산금액(원)", "인건비(원)", "인건비단가(원/ea)",
        "목표달성률(%)", "판정"
    ]
    fmts = [
        None, None, None, None, None, NUMBER_FMT,
        DECIMAL_FMT, DECIMAL_FMT, DECIMAL_FMT, NUMBER_FMT,
        NUMBER_FMT, PERCENT_FMT,
        NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT,
        PERCENT_FMT, None
    ]
    write_header(ws, headers)

    header_keys = list(bi_data[0].keys()) if bi_data else []

    for i, rec in enumerate(sorted(bi_data, key=lambda x: (x["_date"], str(get_field(x, "라인명", default="")))), 1):
        line = str(get_field(rec, "라인명", default="") or "").strip()
        qty = to_num(get_field(rec, "생산량(ea)", "생산량"))
        work_h = to_num(get_field(rec, "근무시간(h)", "근무시간"))
        target_qty = to_num(get_field(rec, "실가동시간 목표수량(ea)", "목표수량(ea)"))
        efficiency = to_num(get_field(rec, "가동효율"))

        # 단가 매칭
        c = cost_data.get(line, {})
        unit_price = c.get("업체기준최종단가", 0)
        labor_h = c.get("총인건비_h", 0)

        prod_amount = qty * unit_price
        labor_cost = labor_h * work_h
        labor_per_ea = safe_div(labor_cost, qty)
        target_rate = safe_div(qty, target_qty)

        # 판정
        if efficiency and isinstance(efficiency, (int, float)):
            eff_val = efficiency if efficiency <= 1 else efficiency / 100
        else:
            eff_val = 0
        if eff_val >= 0.85:
            verdict = "PASS"
        elif eff_val >= 0.80:
            verdict = "WARNING"
        else:
            verdict = "FAIL"

        values = [
            rec["_date"],
            get_field(rec, "유형", default=""),
            get_field(rec, "대표기종", default=""),
            line,
            get_field(rec, "야간구분", default=""),
            to_num(get_field(rec, "생산인원")),
            work_h,
            to_num(get_field(rec, "품질,설비 비가동(h)", "비가동(h)")),
            to_num(get_field(rec, "실가동시간(h)", "실가동시간")),
            qty,
            to_num(get_field(rec, "실적UPH(ea)", "실적UPH")),
            eff_val if eff_val else None,
            unit_price,
            prod_amount,
            labor_cost,
            round(labor_per_ea) if labor_per_ea else 0,
            target_rate if target_rate else None,
            verdict,
        ]
        write_row(ws, i + 1, values, fmts)

    auto_width(ws)

    # 조건부서식: 가동효율(L열)
    if len(bi_data) > 0:
        apply_conditional_format(ws, "L", 2, len(bi_data) + 1, [0.80, 0.85])

    print(f"    → {len(bi_data)}행 작성 완료")
    return ws


def build_monthly_by_line(wb, bi_data, cost_data):
    """시트2: 라인별 월간요약"""
    print("  시트2: 라인별_월간요약...")
    ws = wb.create_sheet("라인별_월간요약")
    ws.sheet_properties.tabColor = "70AD47"

    # 집계
    agg = defaultdict(lambda: {
        "dates": set(), "qty": 0, "work_h": 0, "down_h": 0,
        "active_h": 0, "target_qty": 0, "uph_sum": 0, "upmh_sum": 0,
        "eff_sum": 0, "headcount_sum": 0, "count": 0,
        "prod_amt": 0, "labor_cost": 0, "유형": ""
    })

    for rec in bi_data:
        line = str(get_field(rec, "라인명", default="") or "").strip()
        month = rec["_month"]
        key = (month, line)
        a = agg[key]
        a["dates"].add(rec["_date"])
        qty = to_num(get_field(rec, "생산량(ea)", "생산량"))
        work_h = to_num(get_field(rec, "근무시간(h)", "근무시간"))
        a["qty"] += qty
        a["work_h"] += work_h
        a["down_h"] += to_num(get_field(rec, "품질,설비 비가동(h)", "비가동(h)"))
        a["active_h"] += to_num(get_field(rec, "실가동시간(h)", "실가동시간"))
        a["target_qty"] += to_num(get_field(rec, "실가동시간 목표수량(ea)", "목표수량(ea)"))
        a["uph_sum"] += to_num(get_field(rec, "실적UPH(ea)", "실적UPH"))
        a["upmh_sum"] += to_num(get_field(rec, "UPMH"))
        eff = to_num(get_field(rec, "가동효율"))
        if eff and isinstance(eff, (int, float)):
            a["eff_sum"] += eff if eff <= 1 else eff / 100
        a["headcount_sum"] += to_num(get_field(rec, "생산인원"))
        a["count"] += 1
        a["유형"] = get_field(rec, "유형", default="") or ""

        c = cost_data.get(line, {})
        a["prod_amt"] += qty * c.get("업체기준최종단가", 0)
        a["labor_cost"] += c.get("총인건비_h", 0) * work_h

    headers = [
        "연월", "라인명", "유형", "근무일수", "총생산량(ea)", "일평균생산량(ea)",
        "평균인원", "총근무시간(h)", "총비가동(h)", "비가동비율(%)",
        "평균가동효율", "평균실적UPH(ea)", "평균UPMH",
        "월간생산금액(원)", "월간인건비(원)", "인건비비율(%)",
        "월간목표달성률(%)"
    ]
    fmts = [
        None, None, None, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT,
        DECIMAL_FMT, DECIMAL_FMT, DECIMAL_FMT, PERCENT_FMT,
        PERCENT_FMT, NUMBER_FMT, DECIMAL_FMT,
        NUMBER_FMT, NUMBER_FMT, PERCENT_FMT,
        PERCENT_FMT
    ]
    write_header(ws, headers)

    row_num = 2
    for key in sorted(agg.keys()):
        month, line = key
        a = agg[key]
        days = len(a["dates"])
        cnt = a["count"] or 1
        values = [
            month, line, a["유형"], days,
            a["qty"], round(safe_div(a["qty"], days)),
            round(safe_div(a["headcount_sum"], cnt), 1),
            round(a["work_h"], 1), round(a["down_h"], 1),
            safe_div(a["down_h"], a["work_h"]),
            safe_div(a["eff_sum"], cnt),
            round(safe_div(a["uph_sum"], cnt)),
            round(safe_div(a["upmh_sum"], cnt), 1),
            round(a["prod_amt"]),
            round(a["labor_cost"]),
            safe_div(a["labor_cost"], a["prod_amt"]) if a["prod_amt"] else 0,
            safe_div(a["qty"], a["target_qty"]),
        ]
        write_row(ws, row_num, values, fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")
    return ws, agg


def build_monthly_by_type(wb, monthly_agg):
    """시트3: 유형별 월간요약"""
    print("  시트3: 유형별_월간요약...")
    ws = wb.create_sheet("유형별_월간요약")
    ws.sheet_properties.tabColor = "ED7D31"

    type_agg = defaultdict(lambda: {
        "lines": set(), "qty": 0, "work_h": 0, "down_h": 0,
        "eff_sum": 0, "eff_count": 0, "uph_sum": 0, "uph_count": 0,
        "prod_amt": 0, "labor_cost": 0
    })

    for (month, line), a in monthly_agg.items():
        tp = a["유형"] or "기타"
        key = (month, tp)
        t = type_agg[key]
        t["lines"].add(line)
        t["qty"] += a["qty"]
        t["work_h"] += a["work_h"]
        t["down_h"] += a["down_h"]
        t["prod_amt"] += a["prod_amt"]
        t["labor_cost"] += a["labor_cost"]
        if a["count"]:
            t["eff_sum"] += a["eff_sum"]
            t["eff_count"] += a["count"]
            t["uph_sum"] += a["uph_sum"]
            t["uph_count"] += a["count"]

    headers = [
        "연월", "유형", "라인수", "총생산량(ea)", "총근무시간(h)", "총비가동(h)",
        "비가동비율(%)", "평균가동효율", "평균실적UPH(ea)",
        "월간생산금액(원)", "월간인건비(원)", "인건비비율(%)"
    ]
    fmts = [
        None, None, NUMBER_FMT, NUMBER_FMT, DECIMAL_FMT, DECIMAL_FMT,
        PERCENT_FMT, PERCENT_FMT, NUMBER_FMT,
        NUMBER_FMT, NUMBER_FMT, PERCENT_FMT
    ]
    write_header(ws, headers)

    row_num = 2
    for key in sorted(type_agg.keys()):
        month, tp = key
        t = type_agg[key]
        values = [
            month, tp, len(t["lines"]), t["qty"],
            round(t["work_h"], 1), round(t["down_h"], 1),
            safe_div(t["down_h"], t["work_h"]),
            safe_div(t["eff_sum"], t["eff_count"]),
            round(safe_div(t["uph_sum"], t["uph_count"])),
            round(t["prod_amt"]), round(t["labor_cost"]),
            safe_div(t["labor_cost"], t["prod_amt"]) if t["prod_amt"] else 0,
        ]
        write_row(ws, row_num, values, fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")


def build_cost_trend(wb, monthly_agg):
    """시트4: 금액 월간추이 (크로스탭)"""
    print("  시트4: 금액_월간추이...")
    ws = wb.create_sheet("금액_월간추이")
    ws.sheet_properties.tabColor = "FFC000"

    # 월 목록, 라인 목록
    months = sorted(set(k[0] for k in monthly_agg.keys()))
    lines = sorted(set(k[1] for k in monthly_agg.keys()))

    # 헤더
    headers = ["라인/항목"] + months + ["합계"]
    write_header(ws, headers)

    row_num = 2
    total_prod = [0] * len(months)
    total_labor = [0] * len(months)

    for line in lines:
        # 생산금액 행
        prod_vals = []
        for mi, m in enumerate(months):
            val = round(monthly_agg.get((m, line), {}).get("prod_amt", 0))
            prod_vals.append(val)
            total_prod[mi] += val
        write_row(ws, row_num, [f"{line} 생산금액"] + prod_vals + [sum(prod_vals)],
                  [None] + [NUMBER_FMT] * (len(months) + 1))
        row_num += 1

        # 인건비 행
        labor_vals = []
        for mi, m in enumerate(months):
            val = round(monthly_agg.get((m, line), {}).get("labor_cost", 0))
            labor_vals.append(val)
            total_labor[mi] += val
        write_row(ws, row_num, [f"{line} 인건비"] + labor_vals + [sum(labor_vals)],
                  [None] + [NUMBER_FMT] * (len(months) + 1))
        row_num += 1

        # 인건비비율 행
        ratio_vals = [safe_div(labor_vals[i], prod_vals[i]) for i in range(len(months))]
        write_row(ws, row_num, [f"{line} 인건비비율"] + ratio_vals + [safe_div(sum(labor_vals), sum(prod_vals))],
                  [None] + [PERCENT_FMT] * (len(months) + 1))
        row_num += 1

    # 합계
    write_row(ws, row_num, ["합계 생산금액"] + total_prod + [sum(total_prod)],
              [None] + [NUMBER_FMT] * (len(months) + 1))
    row_num += 1
    write_row(ws, row_num, ["합계 인건비"] + total_labor + [sum(total_labor)],
              [None] + [NUMBER_FMT] * (len(months) + 1))
    row_num += 1
    ratio_total = [safe_div(total_labor[i], total_prod[i]) for i in range(len(months))]
    write_row(ws, row_num, ["합계 인건비비율"] + ratio_total + [safe_div(sum(total_labor), sum(total_prod))],
              [None] + [PERCENT_FMT] * (len(months) + 1))

    auto_width(ws)
    print(f"    → {len(lines)}개 라인 × {len(months)}개월 크로스탭 작성 완료")


def build_labor_analysis(wb, bi_data, cost_data):
    """시트5: 인건비 분석"""
    print("  시트5: 인건비_분석...")
    ws = wb.create_sheet("인건비_분석")
    ws.sheet_properties.tabColor = "9DC3E6"

    # 당월 판단 (최신 월)
    current_month = max(rec["_month"] for rec in bi_data) if bi_data else "2026-03"

    # 당월 데이터 라인별 집계
    line_stats = defaultdict(lambda: {"work_h": 0, "qty": 0})
    for rec in bi_data:
        if rec["_month"] != current_month:
            continue
        line = str(get_field(rec, "라인명", default="") or "").strip()
        line_stats[line]["work_h"] += to_num(get_field(rec, "근무시간(h)", "근무시간"))
        line_stats[line]["qty"] += to_num(get_field(rec, "생산량(ea)", "생산량"))

    headers = [
        "라인코드", "라인명", "작업자수", "관리자수",
        "작업자임률(원/h)", "관리자임률(원/h)", "총인건비/h(원)",
        "직접노무단가(원/ea)", "업체기준최종단가(원/ea)",
        f"당월({current_month}) 총근무시간(h)", f"당월 총생산량(ea)",
        "당월 인건비합계(원)", "당월 생산금액합계(원)", "인건비비율(%)"
    ]
    fmts = [
        None, None, NUMBER_FMT, NUMBER_FMT,
        NUMBER_FMT, NUMBER_FMT, NUMBER_FMT,
        NUMBER_FMT, NUMBER_FMT,
        DECIMAL_FMT, NUMBER_FMT,
        NUMBER_FMT, NUMBER_FMT, PERCENT_FMT
    ]
    write_header(ws, headers)

    row_num = 2
    for line_code in sorted(cost_data.keys()):
        c = cost_data[line_code]
        s = line_stats.get(line_code, {"work_h": 0, "qty": 0})
        labor_total = c["총인건비_h"] * s["work_h"]
        prod_total = c["업체기준최종단가"] * s["qty"]

        values = [
            line_code, c["라인명"], c["작업자수"], c["관리자수"],
            c["작업자임률"], c["관리자임률"], c["총인건비_h"],
            c["직접노무단가"], c["업체기준최종단가"],
            round(s["work_h"], 1), s["qty"],
            round(labor_total), round(prod_total),
            safe_div(labor_total, prod_total),
        ]
        write_row(ws, row_num, values, fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")


def build_line_master(wb, cost_data, bi_data):
    """시트6: 라인 마스터"""
    print("  시트6: 라인_마스터...")
    ws = wb.create_sheet("라인_마스터")
    ws.sheet_properties.tabColor = "A5A5A5"

    # BI에서 라인별 유형/대표기종/야간여부 추출
    line_info = {}
    for rec in bi_data:
        line = str(get_field(rec, "라인명", default="") or "").strip()
        if line not in line_info:
            line_info[line] = {
                "유형": get_field(rec, "유형", default=""),
                "대표기종": get_field(rec, "대표기종", default=""),
                "야간운영": False
            }
        if str(get_field(rec, "야간구분", default="")).strip() in ("야간", "추가"):
            line_info[line]["야간운영"] = True

    headers = [
        "라인코드", "라인명", "유형", "대표기종",
        "작업자수", "관리자수", "실효UPH(ea)", "C/T",
        "업체기준최종단가(원)", "총인건비/h(원)", "야간운영여부"
    ]
    write_header(ws, headers)

    row_num = 2
    for line_code in sorted(cost_data.keys()):
        c = cost_data[line_code]
        info = line_info.get(line_code, {})
        values = [
            line_code, c["라인명"],
            info.get("유형", ""), info.get("대표기종", ""),
            c["작업자수"], c["관리자수"], c["실효UPH"],
            round(safe_div(3600, c["실효UPH"])) if c["실효UPH"] else 0,
            c["업체기준최종단가"], c["총인건비_h"],
            "O" if info.get("야간운영") else "X"
        ]
        write_row(ws, row_num, values, [None, None, None, None, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, DECIMAL_FMT, NUMBER_FMT, NUMBER_FMT, None])
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")


def build_dashboard(wb, bi_data, cost_data, monthly_agg):
    """시트7: 당월 대시보드"""
    print("  시트7: 당월_대시보드...")
    ws = wb.create_sheet("당월_대시보드")
    ws.sheet_properties.tabColor = "FF0000"

    months_sorted = sorted(set(rec["_month"] for rec in bi_data))
    current = months_sorted[-1] if months_sorted else "2026-03"
    prev = months_sorted[-2] if len(months_sorted) >= 2 else None

    # KPI 집계
    cur_qty = sum(a["qty"] for (m, _), a in monthly_agg.items() if m == current)
    cur_amt = sum(a["prod_amt"] for (m, _), a in monthly_agg.items() if m == current)
    cur_labor = sum(a["labor_cost"] for (m, _), a in monthly_agg.items() if m == current)
    prev_qty = sum(a["qty"] for (m, _), a in monthly_agg.items() if m == prev) if prev else 0

    # 상단 KPI
    ws.cell(row=1, column=1, value="당월 대시보드").font = Font(bold=True, size=14)
    ws.cell(row=1, column=3, value=f"기준월: {current}").font = Font(size=11, color="666666")

    kpi_labels = ["총생산량(ea)", "총생산금액(원)", "총인건비(원)", "전월비 증감률", "인건비비율(%)"]
    kpi_values = [
        cur_qty, round(cur_amt), round(cur_labor),
        safe_div(cur_qty - prev_qty, prev_qty) if prev_qty else None,
        safe_div(cur_labor, cur_amt) if cur_amt else 0
    ]
    kpi_fmts = [NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, PERCENT_FMT, PERCENT_FMT]

    for i, (label, val, fmt) in enumerate(zip(kpi_labels, kpi_values, kpi_fmts)):
        col = i * 2 + 1
        cell_label = ws.cell(row=3, column=col, value=label)
        cell_label.font = Font(bold=True, size=10)
        cell_label.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell_val = ws.cell(row=4, column=col, value=val)
        cell_val.font = Font(bold=True, size=12)
        if fmt:
            cell_val.number_format = fmt

    # 하단 라인별 비교표
    row_start = 6
    dash_headers = [
        "라인명", "당월생산량(ea)", "전월생산량(ea)", "증감(%)",
        "당월생산금액(원)", "당월인건비(원)", "인건비비율(%)",
        "당월평균가동효율", "상태"
    ]
    dash_fmts = [
        None, NUMBER_FMT, NUMBER_FMT, PERCENT_FMT,
        NUMBER_FMT, NUMBER_FMT, PERCENT_FMT,
        PERCENT_FMT, None
    ]
    write_header(ws, dash_headers, row=row_start)

    lines = sorted(set(k[1] for k in monthly_agg.keys()))
    row_num = row_start + 1
    for line in lines:
        cur_a = monthly_agg.get((current, line), {})
        prev_a = monthly_agg.get((prev, line), {}) if prev else {}

        cq = cur_a.get("qty", 0)
        pq = prev_a.get("qty", 0)
        ca = cur_a.get("prod_amt", 0)
        cl = cur_a.get("labor_cost", 0)
        c_eff = safe_div(cur_a.get("eff_sum", 0), cur_a.get("count", 1))

        if c_eff >= 0.85:
            status = "PASS"
        elif c_eff >= 0.80:
            status = "WARNING"
        else:
            status = "FAIL"

        values = [
            line, cq, pq,
            safe_div(cq - pq, pq) if pq else None,
            round(ca), round(cl),
            safe_div(cl, ca) if ca else 0,
            c_eff, status
        ]
        write_row(ws, row_num, values, dash_fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {len(lines)}개 라인 대시보드 작성 완료")


def build_validation_sheet(wb, bi_data, cost_data, api_data):
    """시트8: 정합성 검증 (BI vs 조립비시스템 API)"""
    print("  시트8: 정합성_검증...")
    ws = wb.create_sheet("정합성_검증")
    ws.sheet_properties.tabColor = "7030A0"

    if not api_data:
        ws.cell(row=1, column=1, value="API 데이터 미확보 — 정합성 검증 생략")
        print("    → API 데이터 없음, 생략")
        return

    # BI 월간 라인별 집계
    bi_agg = defaultdict(lambda: {"qty": 0, "amt": 0})
    for rec in bi_data:
        line = str(get_field(rec, "라인명", default="") or "").strip()
        month = rec["_month"]
        qty = to_num(get_field(rec, "생산량(ea)", "생산량"))
        c = cost_data.get(line, {})
        bi_agg[(month, line)]["qty"] += qty
        bi_agg[(month, line)]["amt"] += qty * c.get("업체기준최종단가", 0)

    # API 월간 라인별 집계
    api_agg = defaultdict(lambda: {"qty": 0, "amt": 0})
    for d in api_data:
        month = str(d.get("close_date", ""))
        line = str(d.get("assy_line", ""))
        api_agg[(month, line)]["qty"] += to_num(d.get("assy_prod_qty"))
        api_agg[(month, line)]["amt"] += to_num(d.get("assy_amount"))

    # 합집합 키
    all_keys = sorted(set(bi_agg.keys()) | set(api_agg.keys()))

    headers = [
        "연월", "라인명",
        "BI_생산량(ea)", "시스템_조립수량(ea)", "수량차이(ea)",
        "BI_생산금액(원)", "시스템_조립금액(원)", "금액차이(원)",
        "판정"
    ]
    fmts = [None, None, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, None]
    write_header(ws, headers)

    row_num = 2
    for key in all_keys:
        month, line = key
        b = bi_agg.get(key, {"qty": 0, "amt": 0})
        a = api_agg.get(key, {"qty": 0, "amt": 0})
        qty_diff = b["qty"] - a["qty"]
        amt_diff = round(b["amt"] - a["amt"])
        verdict = "일치" if qty_diff == 0 and amt_diff == 0 else "차이있음"

        values = [
            month, line,
            b["qty"], a["qty"], qty_diff,
            round(b["amt"]), round(a["amt"]), amt_diff,
            verdict
        ]
        write_row(ws, row_num, values, fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")


def build_labor_validation(wb, cost_data, line_api_data):
    """시트9: 인건비 교차검증 (엑셀 vs 라인운영 API)"""
    print("  시트9: 인건비_교차검증...")
    ws = wb.create_sheet("인건비_교차검증")
    ws.sheet_properties.tabColor = "00B050"

    if not line_api_data:
        ws.cell(row=1, column=1, value="API 데이터 미확보 — 인건비 교차검증 생략")
        print("    → API 데이터 없음, 생략")
        return

    headers = [
        "라인코드",
        "엑셀_작업자수", "API_작업자수", "작업자수_일치",
        "엑셀_관리자수", "API_관리자수", "관리자수_일치",
        "엑셀_총인건비/h(원)", "API_총인건비/h(원)", "인건비차이(원)",
        "엑셀_최종단가(원)", "판정"
    ]
    fmts = [None, NUMBER_FMT, NUMBER_FMT, None, NUMBER_FMT, NUMBER_FMT, None,
            NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, NUMBER_FMT, None]
    write_header(ws, headers)

    row_num = 2
    for line_code in sorted(cost_data.keys()):
        c = cost_data[line_code]
        api = line_api_data.get(line_code, {})

        api_workers = to_num(api.get("worker_count"))
        api_managers = to_num(api.get("manager_count"))
        api_total_labor = to_num(api.get("total_labor_cost"))

        excel_workers = c["작업자수"]
        excel_managers = c["관리자수"]
        excel_labor_h = c["총인건비_h"]

        worker_match = "O" if excel_workers == api_workers else "X"
        manager_match = "O" if excel_managers == api_managers else "X"
        labor_diff = round(excel_labor_h - api_total_labor) if api_total_labor else None

        all_match = worker_match == "O" and manager_match == "O"
        verdict = "일치" if all_match and (labor_diff is not None and abs(labor_diff) < 100) else "차이있음"
        if not api:
            verdict = "API데이터없음"

        values = [
            line_code,
            excel_workers, api_workers, worker_match,
            excel_managers, api_managers, manager_match,
            excel_labor_h, api_total_labor, labor_diff,
            c["업체기준최종단가"], verdict
        ]
        write_row(ws, row_num, values, fmts)
        row_num += 1

    auto_width(ws)
    print(f"    → {row_num - 2}행 작성 완료")


# ═══════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  생산관리 마스터리스트 생성")
    print(f"  실행시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 파일 존재 확인
    for label, path in [("BI파일", BI_FILE), ("임률단가", COST_FILE)]:
        if not os.path.exists(path):
            print(f"ERROR: {label} 없음 → {path}")
            sys.exit(1)

    # 데이터 로딩
    bi_data = load_bi_data()
    cost_data = load_cost_data()
    api_data = load_assy_api_data()
    line_api = load_line_api_data()

    if not bi_data:
        print("ERROR: 2026년 BI 데이터가 없습니다.")
        sys.exit(1)

    # 워크북 생성
    wb = openpyxl.Workbook()
    # 기본 시트 삭제
    wb.remove(wb.active)

    print("\n시트 생성 시작...")
    build_daily_sheet(wb, bi_data, cost_data)
    _, monthly_agg = build_monthly_by_line(wb, bi_data, cost_data)
    build_monthly_by_type(wb, monthly_agg)
    build_cost_trend(wb, monthly_agg)
    build_labor_analysis(wb, bi_data, cost_data)
    build_line_master(wb, cost_data, bi_data)
    build_dashboard(wb, bi_data, cost_data, monthly_agg)
    build_validation_sheet(wb, bi_data, cost_data, api_data)
    build_labor_validation(wb, cost_data, line_api)

    # 저장
    print(f"\n저장 중: {OUTPUT_FILE}")
    try:
        wb.save(OUTPUT_FILE)
        print(f"✓ 저장 완료: {OUTPUT_FILE}")
    except PermissionError:
        alt = OUTPUT_FILE.replace(".xlsx", f"_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(alt)
        print(f"! 원본 파일 열려있음 → 대체 저장: {alt}")

    # 요약
    print("\n" + "=" * 60)
    print("  생성 완료 요약")
    print("=" * 60)
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row - 1}행")
    print(f"\n  총 {len(wb.worksheets)}개 시트")


if __name__ == "__main__":
    main()
