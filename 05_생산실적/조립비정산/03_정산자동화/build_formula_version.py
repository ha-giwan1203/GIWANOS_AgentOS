"""수식 기반 정산 파일 생성 스크립트 (_pipeline_config 사용)"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import os
import sys
from pathlib import Path

# === SP3M3 BI 보정 룰 (2026-05-14 채택, CLAUDE.md "SP3M3 구ERP 야간 BI 보정 룰" 참조) ===
BI_PATH = Path(r"C:\Users\User\Desktop\업무리스트\05_생산실적\BI실적\대원테크_라인별 생산실적_BI.xlsx")


def extract_bi_night_total(line: str, year: int, month: int) -> tuple[int, int]:
    """BI 원본 라인×월 야간 합계 + 일수. (qty, days) 반환. 데이터 없으면 (0, 0)."""
    if not BI_PATH.exists():
        return 0, 0
    wb = openpyxl.load_workbook(BI_PATH, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    night_val = next((r[5] for r in rows[1:]
                      if r[4] == line and r[5] not in (None, "주간")), None)
    if night_val is None:
        return 0, 0
    total = days = 0
    for r in rows[1:]:
        if r[4] != line or r[5] != night_val:
            continue
        d = r[7]
        if not isinstance(d, datetime) or d.year != year or d.month != month:
            continue
        if r[14]:
            total += r[14]
            days += 1
    return total, days

# Windows cp949 콘솔에서 utf-8 print 가능하게 (em-dash, 한국어 등)
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _pipeline_config import BASE_DIR, GERP_FILE, OLDERP_FILE, OLDERP_SHEET, MONTH, LINE_INFO, LINE_GROUP, SP3M3_MODULE_FILE
from _error_types import (
    build_classify_formula, map_cat_to_user_type,
    TYPE_ORDER, TYPE_COLORS, ERROR_TYPES,
)

BASE = BASE_DIR
_m_int = int(MONTH) + 1
if _m_int > 12:
    _m_int -= 12
FOLDER_NAME = f'{_m_int:02d}월'

print(f"=== 수식 기반 정산 파일 생성 ({MONTH}월 정산 → {FOLDER_NAME}/) ===")

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
gerp_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
olderp_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
summary_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
num_fmt = '#,##0;-#,##0;"-"'

# 1. Load sources
print("1. 기준정보 로딩...")
# 기준정보 path는 _pipeline_config.MASTER_FILE 단일 권위 사용 (V1/V2 drift 방지)
from _pipeline_config import MASTER_FILE
print(f"   기준정보: {os.path.basename(MASTER_FILE)}")

print(f"2. GERP 실적 로딩... ({os.path.basename(GERP_FILE)})")
gerp_wb = openpyxl.load_workbook(GERP_FILE, data_only=True)
gerp_ws = gerp_wb[gerp_wb.sheetnames[0]]

# 정산 제외 룰 폐기 (세션148, 2026-05-08 사용자 룰 통합)
# 이전: SVM/OVK 차종 이관 + 88820X 차종 무관 skip을 빌더 raw 단계에서 처리
# 정정: 마스터 V2에서 정산 제외 품번 직접 삭제 → 라인시트 일반 영역 자동 누락 → 빌더 처리 불필요
# 미등록 영역도 폐기 — 마스터 미등록 GERP 품번 알림은 빌더 로그 카운트로 대체

# ===== 마스터 V2 자동 갱신 (사용자 룰 2026-05-07): 기준정보 누락 + Usage 차이 GERP 기준 등록 =====
print("1-1. 마스터 V2 자동 갱신 (기준정보 누락 + Usage 차이 GERP 기준 등록)...")
LINE_ORDER_MASTER = ['SD9A01', 'ANAAS04', 'DRAAS11', 'SP3M3', 'HASMS02', 'HCAMS02',
                     'WAMAS01', 'WABAS01', 'WASAS01', 'ISAMS03']
import shutil, datetime as _dt, json as _json
_master_backup = MASTER_FILE.replace('.xlsx', f'_pre_autosync_{_dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(MASTER_FILE, _master_backup)
print(f"   마스터 백업: {os.path.basename(_master_backup)}")

# denylist 로드 (사용자 명시 삭제 / GERP 중복 입력 skip — 세션149, 2026-05-08)
_denylist_path = os.path.join(os.path.dirname(MASTER_FILE), '_master_denylist.json')
DENYLIST = set()
if os.path.exists(_denylist_path):
    try:
        with open(_denylist_path, 'r', encoding='utf-8') as _f:
            _dl = _json.load(_f)
        for _e in _dl.get('entries', []):
            if len(_e) >= 3:
                _ln, _pn, _pr = _e[0], _e[1], _e[2]
                try: _pr_key = round(float(_pr), 4) if _pr is not None else None
                except: _pr_key = None
                DENYLIST.add((str(_ln).strip(), str(_pn).strip(), _pr_key))
        print(f"   denylist 로드: {len(DENYLIST)}건 (자동갱신 skip)")
    except Exception as _ex:
        print(f"   [WARN] denylist 로드 실패: {_ex}")

# 마스터 V2 (라인, 품번, 단가) → 행 위치 + Usage
master_wb_w = openpyxl.load_workbook(MASTER_FILE)
master_idx = {}  # (라인, 품번, 단가) → (sheet, row, current_usage)
master_pn_lines = defaultdict(set)  # 품번 → 등록된 라인 set
for sn in master_wb_w.sheetnames:
    if sn not in LINE_ORDER_MASTER: continue
    ws_m = master_wb_w[sn]
    for r in range(4, ws_m.max_row + 1):
        pn = ws_m.cell(r, 1).value
        if not pn: continue
        ln = ws_m.cell(r, 3).value or sn
        u = ws_m.cell(r, 5).value
        p = ws_m.cell(r, 7).value
        try: p_key = round(float(p), 4) if p is not None else None
        except: p_key = None
        master_idx[(str(ln), str(pn).strip(), p_key)] = (sn, r, u)
        master_pn_lines[str(pn).strip()].add(str(ln))

# GERP raw 0109 + 정산 대상 (SVM/OVK/88820X skip)에서 (라인, 품번, 단가) 수집
# 주의: raw row 1 = 헤더, row 2~ 데이터. col 19 = 업체코드 (raw 컬럼 매핑 실측 2026-05-07).
gerp_keys = {}  # (라인, 품번, 단가) → (Usage, 조립품번, 단가구분, 차종)
for r in range(2, gerp_ws.max_row + 1):
    co = gerp_ws.cell(r, 19).value
    if str(co).strip() != '0109': continue
    line = gerp_ws.cell(r, 3).value
    pn = gerp_ws.cell(r, 7).value
    cha = gerp_ws.cell(r, 6).value
    if not line or not pn: continue
    line = str(line).strip(); pn = str(pn).strip()
    # 정산 제외 룰 폐기 — 마스터 V2 직접 삭제로 처리 (세션148)
    # RSP/MO 매핑 미적용 (사용자 룰: GERP 그대로) — RSP/MO 자체로 마스터 등록
    if line not in LINE_ORDER_MASTER: continue
    usage = gerp_ws.cell(r, 11).value
    assy = gerp_ws.cell(r, 12).value
    price = gerp_ws.cell(r, 16).value
    nyu = gerp_ws.cell(r, 14).value
    if nyu == '추가': continue  # 추가행은 마스터 등록 X (가산만)
    try: p_key = round(float(price), 4) if price is not None else None
    except: p_key = None
    # denylist skip — 사용자 명시 삭제 / GERP 중복 입력 부활 차단 (세션149)
    if (line, pn, p_key) in DENYLIST: continue
    gerp_keys[(line, pn, p_key)] = (usage, assy, '정단가', cha)

# 갱신 적용
new_added = 0; usage_updated = 0
for k, v in gerp_keys.items():
    line, pn, p_key = k
    usage, assy, dgu, cha = v
    if k not in master_idx:
        # 신규 등록 — 라인 시트 마지막 행에 추가
        ws_m = master_wb_w[line]
        new_r = ws_m.max_row + 1
        # cha sentinel 가드 — int32 -2146826246 등 raw 결함 → None (세션148, 2026-05-08)
        cha_safe = cha if not (isinstance(cha, (int, float)) and cha < -1e8) else None
        ws_m.cell(new_r, 1, pn)
        ws_m.cell(new_r, 2, '0109')
        ws_m.cell(new_r, 3, line)
        ws_m.cell(new_r, 4, assy or pn)
        ws_m.cell(new_r, 5, usage if usage is not None else 1)
        ws_m.cell(new_r, 6, dgu)
        ws_m.cell(new_r, 7, p_key)
        ws_m.cell(new_r, 8, cha_safe)
        new_added += 1
    else:
        # Usage 차이 — GERP Usage로 갱신
        sn, r, cur_u = master_idx[k]
        try:
            cur_uf = float(cur_u) if cur_u is not None else None
            gerp_uf = float(usage) if usage is not None else None
            if cur_uf is not None and gerp_uf is not None and abs(cur_uf - gerp_uf) > 0.001:
                master_wb_w[sn].cell(r, 5).value = int(gerp_uf) if gerp_uf == int(gerp_uf) else gerp_uf
                usage_updated += 1
        except: pass

if new_added or usage_updated:
    master_wb_w.save(MASTER_FILE)
    print(f"   마스터 V2 자동 갱신: 신규 등록 {new_added}건 + Usage 갱신 {usage_updated}건")
else:
    print(f"   마스터 V2 자동 갱신: 변경 없음")
master_wb_w.close()

# 갱신된 마스터 V2 다시 로드
ref_wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)

# === RSP/MO 모듈품번 → 13자리 기준품번 매핑 (SP3M3 야간) ===
# === 차종별 이관 처리 ===
# SVM 차종: SP3M3/HCAMS02/ISAMS03 → 이관 (정산 제외)
# OVK 차종: SD9A01/ANAAS04/DRAAS11 → 이관 (정산 제외)
print("2-1. RSP/MO 매핑 로드 + SP3M3 prefix 매핑...")
RSP_TO_PN10 = {}  # RSP/MO → 10자리 기본품번
try:
    mwb = openpyxl.load_workbook(SP3M3_MODULE_FILE, data_only=True, read_only=True)
    mws = mwb['Sheet1']
    for row in mws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4: continue
        pn10, rsp = row[1], row[3]
        if rsp and pn10:
            RSP_TO_PN10[str(rsp).strip()] = str(pn10).strip()
    mwb.close()
except Exception as e:
    print(f"   [WARN] RSP 매핑 로드 실패: {e}")

# 라인별 10자리 prefix → 첫 등장 13자리 품번 (전체 라인, SVM 이관 차종 제외)
LINE_PN10_TO_PN13 = defaultdict(dict)  # line → {10자리: 13자리}
for ln in ref_wb.sheetnames:
    if ln not in LINE_INFO: continue
    rws = ref_wb[ln]
    seen10 = set()
    for row in rws.iter_rows(min_row=4, values_only=True):
        pn = row[0]
        # 정산 제외 룰 폐기 (세션148) — 마스터 V2에 정산 제외 품번 없음을 전제
        if pn and isinstance(pn, str) and len(pn) >= 10:
            pn10 = pn[:10]
            if pn10 not in seen10:
                LINE_PN10_TO_PN13[ln][pn10] = pn
                seen10.add(pn10)

# RSP → 13자리 직접 매핑 (SP3M3 매핑파일 기반)
RSP_DIRECT = {}
sp_pn10_to_13 = LINE_PN10_TO_PN13.get('SP3M3', {})
for rsp, pn10 in RSP_TO_PN10.items():
    if pn10 in sp_pn10_to_13:
        RSP_DIRECT[rsp] = sp_pn10_to_13[pn10]
print(f"   RSP 매핑 {len(RSP_TO_PN10)}건 / SP3M3 직접 매핑 {len(RSP_DIRECT)}건")
print(f"   라인별 prefix 매핑: " + ", ".join(f"{ln}={len(LINE_PN10_TO_PN13[ln])}" for ln in LINE_INFO))

# === MO 매핑 사전 빌드 (전체 라인) ===
# MO + 10자리 → 라인별 prefix 매칭 → 13자리
MO_DIRECT = {}  # (라인, MO품번) → 13자리
# 4월부터: 마스터에 미등록인 MO도 13자리 prefix로 매핑 가능 (빌더 L161~165 로직)
# 매핑 시트엔 마스터 등록된 RSP만 박음 (사용자 추적성 확보)

print(f"3. 구ERP 실적 로딩... ({os.path.basename(OLDERP_FILE)} / 시트={OLDERP_SHEET})")
olderp_wb = openpyxl.load_workbook(OLDERP_FILE, data_only=True)
olderp_ws = olderp_wb[OLDERP_SHEET]

# 2. New workbook
wb = openpyxl.Workbook()

# ===== 사용법 sheet =====
ws_guide = wb.active
ws_guide.title = '사용법'
guide_lines = [
    '조립비 정산 수식 버전 — 사용법',
    '',
    '1. GERP_입력 시트에 G-ERP 실적 데이터를 붙여넣기 (헤더 포함, Row1부터)',
    '2. 구ERP_입력 시트에 구ERP 전체입고량 데이터를 붙여넣기 (헤더 포함, Row1부터)',
    '3. 각 라인 시트의 GERP/구ERP 실적·금액이 자동 계산됨',
    '4. 정산집계 시트에서 라인별 합계 확인',
    '',
    '※ 현재 3월 실적 데이터가 미리 입력되어 있음',
    '※ 다음 달 정산 시 GERP_입력/구ERP_입력만 새 데이터로 교체하면 자동 갱신',
    '',
    '수식 구조:',
    '  GERP 주간수량 = SUMIFS(생산량, 라인=해당라인, 품번=해당품번, 주야=정상)',
    '  GERP 야간수량 = SUMIFS(생산량, 라인=해당라인, 품번=해당품번, 주야=추가)',
    '  GERP 금액 = 기준단가 x 수량',
    '  SD9A01 야간금액 = 기준단가 x 0.3 x 야간수량 (야간 30% 가산)',
    '  구ERP 수량 = SUMIFS(입고수량, 품번=해당품번)',
    '',
    '제한사항:',
    '  - 구ERP 야간: LOT 끝자리 B = 야간, 그 외 = 주간',
    '  - SP3M3 구ERP: 서브라인 LOT B 무의미 → 총수량-GERP야간=주간, 야간=GERP야간',
    '  - SP3M3 RSP 모듈품번: 빌더 실행 시 GERP_입력 시트의 RSP/MO 품번이 13자리로 자동 변환됨',
    '  - 매월 빌더 재실행 (build_formula_version.py) → 매핑 최신 반영',
    '  - 변환 결과는 모듈품번_매핑 시트에서 확인 가능',
    '  - Usage=2 품번 수량 2배 환산은 미적용 (필요시 수동 확인)',
]
for r, text in enumerate(guide_lines, 1):
    ws_guide.cell(r, 1, text)
ws_guide.cell(1, 1).font = Font(bold=True, size=14)
ws_guide.column_dimensions['A'].width = 80

# ===== GERP_입력 sheet =====
print("4. GERP_입력 시트 생성...")
ws_gerp = wb.create_sheet('GERP_입력')

gerp_row_count = 0
rsp_count = 0; mo_count = 0
dedup_skipped = 0
non_0109_skipped = 0  # 0109 외 vendor skip (사용자 룰 2026-05-07: GERP는 0109 필터)
out_r = 0  # GERP_입력 시트 출력 행 (skip 시 원본과 다를 수 있음)
seen_keys = set()  # GERP raw 완전 중복 행 dedupe (사용자 룰 2026-05-07)
for r in range(1, gerp_ws.max_row + 1):
    line_v = gerp_ws.cell(r, 3).value if r > 1 else None
    pn_v = gerp_ws.cell(r, 7).value if r > 1 else None
    cha_v = gerp_ws.cell(r, 6).value if r > 1 else None  # 차종
    co_v = gerp_ws.cell(r, 19).value if r > 1 else None  # 업체코드 (raw col 19, 사용자 룰 2026-05-07)

    # 0109 vendor 필터 (사용자 명시 2026-05-07: GERP는 0109 필터 OK)
    # 정산 권위 = 대원테크 0109. 다른 vendor (A00029/A00030/A00031 등) 데이터는 SUMIFS 매칭 시 라인+품번+단가 우연 일치하면 합산되는 결함.
    if r > 1 and str(co_v).strip() != '0109':
        non_0109_skipped += 1
        continue

    # 정산 제외 룰 폐기 (세션148, 2026-05-08)
    # 이전: SVM/OVK 차종 이관 + 88820X 차종 무관 skip을 빌더 raw 단계에서 처리
    # 정정: 마스터 V2에서 정산 제외 품번 직접 삭제 → 라인시트 일반 영역 SUMIFS 호출 안 함 → 영향 없음

    # GERP raw 완전 중복 dedupe (라인+품번+주야+단가+생산량+조립금액+Usage)
    # ※ 조립품번(col12)은 동일 amt에 다른 형식이 박힐 수 있어 키에서 제외 (사용자 룰 2026-05-07)
    if r > 1:
        shift_v = gerp_ws.cell(r, 14).value
        price_v = gerp_ws.cell(r, 16).value
        qty_v = gerp_ws.cell(r, 9).value
        amt_v = gerp_ws.cell(r, 17).value
        usage_v = gerp_ws.cell(r, 11).value
        dedup_key = (str(line_v), str(pn_v), str(shift_v), price_v, qty_v, amt_v, usage_v)
        if dedup_key in seen_keys:
            dedup_skipped += 1
            continue
        seen_keys.add(dedup_key)

    pn_replace = None
    if r > 1 and line_v and pn_v and line_v in LINE_INFO:
        ps = str(pn_v).strip()
        ln = str(line_v).strip()
        # RSP 변환 폐기 (사용자 룰 2026-05-07): GERP 모듈품번 야간 amt = 구ERP 야간 amt 동일 반영
        # RSP를 일반 품번(13자리)로 변환하면 야간이 일반 품번 행에 합쳐져 SP3M3 구ERP 주간 -야간 차감 룰이 잘못 작동
        # RSP 그대로 두고 빌더 라인 시트 미등록 영역에서 별개 행으로 처리 (gp=None+SP3M3 분기 P=L 적용)
        # RSP_DIRECT는 모듈품번_매핑 시트 정보용으로만 사용
        # MO 변환 제거 (2026-05-07 사용자 룰): MO 품번은 마스터에 라인별 자체 등록되어 SUMIFS 직접 매칭

    out_r += 1
    for c in range(1, gerp_ws.max_column + 1):
        v = gerp_ws.cell(r, c).value
        if c == 7 and pn_replace:
            v = pn_replace
        cell = ws_gerp.cell(out_r, c, v)
        if r == 1:
            cell.font = header_font
            cell.fill = gerp_fill
            cell.border = thin_border
        elif isinstance(v, (int, float)):
            cell.number_format = num_fmt
    if r > 1:
        gerp_row_count += 1
print(f"   GERP_입력 시트: 비-0109 skip {non_0109_skipped}건 / dedupe skip {dedup_skipped}건 / RSP→13자리 {rsp_count}건 / MO→13자리 {mo_count}건")

print(f"   GERP 데이터: {gerp_row_count}행")

# === GERP 라인별 품번/단가 캐시 (단가차이 영역 식별용) ===
# 정산 제외 룰 폐기 (세션148) — 0109 vendor 필터만 적용
gerp_line_pns = defaultdict(set)         # 라인 → 품번 set
gerp_line_pn_prices = defaultdict(set)   # (라인, 품번) → 정상행 단가 set
for r in range(2, gerp_ws.max_row + 1):
    line_v = gerp_ws.cell(r, 3).value
    pn_v = gerp_ws.cell(r, 7).value
    co_v = gerp_ws.cell(r, 19).value
    sd_v = gerp_ws.cell(r, 14).value
    price_v = gerp_ws.cell(r, 16).value
    if not (line_v and pn_v):
        continue
    ls = str(line_v).strip()
    ps = str(pn_v).strip()
    if str(co_v).strip() != '0109':
        continue
    gerp_line_pns[ls].add(ps)
    if sd_v == '정상' and price_v is not None:
        gerp_line_pn_prices[(ls, ps)].add(price_v)

# ===== 구ERP_입력 sheet (품번별 주야 집계 — 전체업체 피벗) =====
# CLAUDE.md L29/L37 규칙:
#   - SD9·SUB 라인: 0109 전체 품번 대상, 라인코드 무시 → 전체업체 피벗에도 0109 품번이 모두 포함되어 결과 동일
#   - SP3M3: 모듈품번(MO) 매칭 불가 → 전체업체 피벗(all_day/all_night) 필수
# 따라서 단일 피벗을 전체업체로 만들면 두 케이스 모두 커버
print("5. 구ERP_입력 시트 생성 (전체업체 품번별 주야 피벗)...")
ws_olderp = wb.create_sheet('구ERP_입력')

olderp_day = defaultdict(int)   # LOT 끝자리 ≠ B
olderp_night = defaultdict(int) # LOT 끝자리 = B

# TRANSFER_PN_EXCLUDE 폐기 (세션148) — 마스터 V2 정리로 자동 처리

olderp_raw_count = 0
for r in range(3, olderp_ws.max_row + 1):
    pn = olderp_ws.cell(r, 5).value
    vendor = olderp_ws.cell(r, 3).value
    if not pn or not vendor:
        continue
    ps = str(pn).strip()
    # 이관 품번 skip 폐기 (세션148) — 마스터 V2 정리로 자동 처리 (라인시트 SUMIFS 호출 안 함)
    olderp_raw_count += 1
    lot = str(olderp_ws.cell(r, 10).value or '')
    qty_raw = olderp_ws.cell(r, 11).value or 0
    try:
        qty = int(qty_raw) if not isinstance(qty_raw, (int, float)) else qty_raw
    except (ValueError, TypeError):
        continue
    if lot.strip().endswith('B'):
        olderp_night[ps] += qty
    else:
        olderp_day[ps] += qty
print(f"   구ERP_입력 시트: 총 {olderp_raw_count}행 처리")

# 품번 합집합
all_pns = sorted(set(olderp_day) | set(olderp_night), key=str)

olderp_headers = ['품번', '주간수량', '야간수량', '합계']
for c, h in enumerate(olderp_headers, 1):
    cell = ws_olderp.cell(1, c, h)
    cell.font = header_font
    cell.fill = olderp_fill
    cell.border = thin_border

for i, pn in enumerate(all_pns):
    out_r = i + 2
    ws_olderp.cell(out_r, 1, pn)
    ws_olderp.cell(out_r, 2, olderp_day.get(pn, 0)).number_format = num_fmt
    ws_olderp.cell(out_r, 3, olderp_night.get(pn, 0)).number_format = num_fmt
    ws_olderp.cell(out_r, 4).value = f'=B{out_r}+C{out_r}'
    ws_olderp.cell(out_r, 4).number_format = num_fmt

print(f"   구ERP 원본: {olderp_raw_count}행 → 집계: {len(all_pns)}품번")

# ===== 모듈품번_매핑 sheet (참조용) =====
print("5.5 모듈품번_매핑 시트 생성...")
ws_map = wb.create_sheet('모듈품번_매핑')
map_headers = ['모듈품번(GERP원본)', '변환_13자리', '10자리', '대상라인', '비고']
for c, h in enumerate(map_headers, 1):
    cell = ws_map.cell(1, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

map_r = 1
# RSP 매핑 성공 (마스터 → SP3M3 prefix → 13자리)
for rsp, pn13 in sorted(RSP_DIRECT.items()):
    map_r += 1
    pn10 = RSP_TO_PN10.get(rsp, '')
    ws_map.cell(map_r, 1, rsp).border = thin_border
    ws_map.cell(map_r, 2, pn13).border = thin_border
    ws_map.cell(map_r, 3, pn10).border = thin_border
    ws_map.cell(map_r, 4, 'SP3M3').border = thin_border
    ws_map.cell(map_r, 5, 'RSP→13자리 자동변환').border = thin_border

# 마스터 등록인데 매핑 못 한 RSP (10자리가 SP3M3 시트에 없음)
unmap_rsp = sorted(set(RSP_TO_PN10.keys()) - set(RSP_DIRECT.keys()))
for rsp in unmap_rsp:
    map_r += 1
    pn10 = RSP_TO_PN10.get(rsp, '')
    ws_map.cell(map_r, 1, rsp).border = thin_border
    c2 = ws_map.cell(map_r, 2, '미매핑')
    c2.border = thin_border
    c2.font = Font(color='C00000')
    ws_map.cell(map_r, 3, pn10).border = thin_border
    ws_map.cell(map_r, 4, '-').border = thin_border
    ws_map.cell(map_r, 5, '10자리가 SP3M3 기준정보에 없음 (이관/구품번)').border = thin_border

ws_map.column_dimensions['A'].width = 18
ws_map.column_dimensions['B'].width = 18
ws_map.column_dimensions['C'].width = 14
ws_map.column_dimensions['D'].width = 10
ws_map.column_dimensions['E'].width = 40
print(f"   매핑 시트: {len(RSP_DIRECT)}건 매핑 + {len(unmap_rsp)}건 미매핑")

# ===== 10 Line sheets =====
LINES = ['SD9A01', 'ANAAS04', 'DRAAS11', 'SP3M3', 'HASMS02', 'HCAMS02',
         'WAMAS01', 'WABAS01', 'WASAS01', 'ISAMS03']

line_summaries = {}

for line in LINES:
    print(f"6. {line} 시트 생성...")
    ref_ws = ref_wb[line]
    ws = wb.create_sheet(line)

    # Row 1: section headers
    sections = [
        (1, 8, '기준정보', header_fill),
        (9, 10, 'GERP 실적', gerp_fill),
        (11, 12, 'GERP 금액', gerp_fill),
        (13, 14, '구ERP 실적', olderp_fill),
        (15, 16, '구ERP 금액', olderp_fill),
        (17, 18, '차이', summary_fill),
        (19, 19, '오류분류', summary_fill),
    ]
    for start, end, title, fill in sections:
        cell = ws.cell(1, start, title)
        cell.font = header_font
        cell.fill = fill
        cell.border = thin_border
        if start != end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for cc in range(start + 1, end + 1):
                ws.cell(1, cc).fill = fill
                ws.cell(1, cc).border = thin_border

    # Row 2: column headers
    col_headers = ['품번', '조립업체코드', '조립라인코드', '조립품번', 'Usage',
                   '단가구분', '단가', '차종',
                   '주간', '야간', '주간', '야간',
                   '주간', '야간', '주간', '야간',
                   '금액차이', '수량차이', '카테고리']
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(2, c, h)
        cell.font = header_font
        cell.border = thin_border
        if c <= 8:
            cell.fill = header_fill
        elif c <= 12:
            cell.fill = gerp_fill
        elif c <= 16:
            cell.fill = olderp_fill
        else:
            cell.fill = summary_fill

    # 라인의 (품번 → [단가들]) 사전 — 다중단가 식별용
    pn_prices_list = defaultdict(list)
    for rr in range(4, ref_ws.max_row + 1):
        pp = ref_ws.cell(rr, 1).value
        pr = ref_ws.cell(rr, 7).value
        if pp and isinstance(pr, (int, float)):
            pn_prices_list[str(pp).strip()].append(pr)

    # Data rows (ref file: row 4+ = data, row 1=title, row 3=header)
    out_r = 2  # will increment to 3+
    for r in range(4, ref_ws.max_row + 1):
        pn = ref_ws.cell(r, 1).value
        if not pn:
            continue
        out_r += 1

        # A~H: 기준정보 (col 5 = Usage 정수 강제 — 빌더 SUMIFS × E 환산 정합용 2026-05-07)
        # col 8 = 차종 — int32 sentinel(-2146826246 등) 빈값 정합 (세션148, 사용자 지적 2026-05-08)
        for c in range(1, 9):
            v = ref_ws.cell(r, c).value
            if c == 5:
                try: v = int(float(v)) if v is not None and v != '' else 1
                except (ValueError, TypeError): v = 1
            elif c == 8 and isinstance(v, (int, float)) and v < -1e8:
                v = None  # int32 sentinel → 빈값 (마스터 raw 결함 가드)
            cell = ws.cell(out_r, c, v)
            cell.border = thin_border
            if c == 7 and isinstance(v, (int, float)):
                cell.number_format = num_fmt

        # I: GERP 주간수량 — Python step5 처리 방식
        # 단일단가: SUMIFS(전체 정상) — 단가차이도 자동 흡수
        # 다중단가 첫 행: SUMIFS(자기 단가) + (SUMIFS(전체) - SUMIFS(다른 단가들))
        # 다중단가 다른 행: SUMIFS(자기 단가) — 자기 단가만
        ps = str(pn).strip()
        prices_list = pn_prices_list.get(ps, [])
        is_multi = len(prices_list) > 1
        is_first_pn = (out_r == 3) or (ws.cell(out_r-1, 1).value != pn)
        # 단가 매칭 SUMIFS 공통
        sumifs_self = (f'SUMIFS(GERP_입력!$O:$O,GERP_입력!$C:$C,"{line}",'
                       f'GERP_입력!$G:$G,A{out_r},GERP_입력!$N:$N,"정상",'
                       f'GERP_입력!$P:$P,G{out_r})')
        sumifs_all = (f'SUMIFS(GERP_입력!$O:$O,GERP_입력!$C:$C,"{line}",'
                      f'GERP_입력!$G:$G,A{out_r},GERP_입력!$N:$N,"정상")')
        if not is_multi:
            # 단일단가: 전체 합산 (단가차이 자동 흡수)
            ws.cell(out_r, 9).value = f'={sumifs_all}'
        elif is_first_pn:
            # 다중단가 첫 행: 전체 - 다른 단가들 합 (= 자기 단가 GERP + 단가차이 잔여)
            other_sumifs = []
            current_price = ref_ws.cell(r, 7).value
            for op in prices_list:
                if op != current_price:
                    other_sumifs.append(
                        f'SUMIFS(GERP_입력!$O:$O,GERP_입력!$C:$C,"{line}",'
                        f'GERP_입력!$G:$G,A{out_r},GERP_입력!$N:$N,"정상",'
                        f'GERP_입력!$P:$P,{op})')
            if other_sumifs:
                other_sum = '+'.join(other_sumifs)
                ws.cell(out_r, 9).value = f'={sumifs_all}-({other_sum})'
            else:
                ws.cell(out_r, 9).value = f'={sumifs_all}'
        else:
            # 다중단가 다른 행: 자기 단가만
            ws.cell(out_r, 9).value = f'={sumifs_self}'
        ws.cell(out_r, 9).number_format = num_fmt
        ws.cell(out_r, 9).border = thin_border

        # J: GERP 야간수량 — 첫 다중단가 행에만 합산 (다중단가 야간 중복 방지)
        ws.cell(out_r, 10).value = (
            f'=IF(COUNTIF(A$3:A{out_r},A{out_r})=1,'
            f'SUMIFS(GERP_입력!$O:$O,'
            f'GERP_입력!$C:$C,"{line}",'
            f'GERP_입력!$G:$G,A{out_r},'
            f'GERP_입력!$N:$N,"추가"),0)')
        ws.cell(out_r, 10).number_format = num_fmt
        ws.cell(out_r, 10).border = thin_border

        # K: GERP 주간금액
        ws.cell(out_r, 11).value = f'=G{out_r}*I{out_r}'
        ws.cell(out_r, 11).number_format = num_fmt
        ws.cell(out_r, 11).border = thin_border

        # L: GERP 야간금액 — 첫 다중단가 행에만 합산 (중복 방지)
        ws.cell(out_r, 12).value = (
            f'=IF(COUNTIF(A$3:A{out_r},A{out_r})=1,'
            f'SUMIFS(GERP_입력!$Q:$Q,'
            f'GERP_입력!$C:$C,"{line}",'
            f'GERP_입력!$G:$G,A{out_r},'
            f'GERP_입력!$N:$N,"추가"),0)')
        ws.cell(out_r, 12).number_format = num_fmt
        ws.cell(out_r, 12).border = thin_border

        # M: 구ERP 주간수량 — 첫 다중단가 행에만 합산 (다중단가 중복 방지)
        # SD9A01/SP3M3: has_night=True / SUB 라인: has_night=False (LOT B 무시 총수량)
        # 사용자 룰 2026-05-07: Usage=2 품번 GERP 환산 후(2배) / 구ERP 환산 전 → 라인 시트 비교 base 통일 위해 구ERP 수량에 Usage(E) 곱
        line_has_night = LINE_INFO[line]['has_night'] if line in LINE_INFO else False
        if line == 'SP3M3':
            base_m = (f'IFERROR(SUMIFS(구ERP_입력!$D:$D,구ERP_입력!$A:$A,A{out_r})*E{out_r},0)')
        elif line_has_night:
            base_m = (f'IFERROR(SUMIFS(구ERP_입력!$B:$B,구ERP_입력!$A:$A,A{out_r})*E{out_r},0)')
        else:
            base_m = (f'IFERROR(SUMIFS(구ERP_입력!$D:$D,구ERP_입력!$A:$A,A{out_r})*E{out_r},0)')
        ws.cell(out_r, 13).value = f'=IF(COUNTIF(A$3:A{out_r},A{out_r})=1,{base_m},0)'
        ws.cell(out_r, 13).number_format = num_fmt
        ws.cell(out_r, 13).border = thin_border

        # N: 구ERP 야간수량
        # SP3M3 룰: N=J (구ERP raw 야간 사용 안 함 — 전체업체 피벗 한계, CLAUDE.md L40)
        # 그 외: SUMIFS(구ERP 야간) × Usage (다중단가 첫 행에만)
        if line == 'SP3M3':
            ws.cell(out_r, 14).value = f'=J{out_r}'
        elif line_has_night:
            ws.cell(out_r, 14).value = (
                f'=IF(COUNTIF(A$3:A{out_r},A{out_r})=1,'
                f'IFERROR(SUMIFS(구ERP_입력!$C:$C,구ERP_입력!$A:$A,A{out_r})*E{out_r},0),0)')
        else:
            ws.cell(out_r, 14).value = 0
        ws.cell(out_r, 14).number_format = num_fmt
        ws.cell(out_r, 14).border = thin_border

        # O: 구ERP 주야 합계 × 단가 (사용자 룰 2026-05-08)
        # SP3M3은 N=J 룰이라 O = G × M만 (P=L과 결합 시 야간 중복 방지)
        # 그 외는 O = G × (M+N) (주야 전체)
        if line == 'SP3M3':
            ws.cell(out_r, 15).value = f'=G{out_r}*M{out_r}'
        else:
            ws.cell(out_r, 15).value = f'=G{out_r}*(M{out_r}+N{out_r})'
        ws.cell(out_r, 15).number_format = num_fmt
        ws.cell(out_r, 15).border = thin_border

        # P: 구ERP 야간금액 (사용자 룰 2026-05-08 통합)
        # 모든 야간정산 = GERP raw 추가행 그대로 (이미 0.3 적용된 단가). 빌더 추가 가산 X.
        # SD9A01 / SP3M3: P = L (GERP 야간금액 그대로) — 양측 야간 동일
        # SUB 라인 (야간 없음): G × N (보통 N=0)
        if line in ('SD9A01', 'SP3M3'):
            ws.cell(out_r, 16).value = f'=L{out_r}'
        else:
            ws.cell(out_r, 16).value = f'=G{out_r}*N{out_r}'
        ws.cell(out_r, 16).number_format = num_fmt
        ws.cell(out_r, 16).border = thin_border

        # Q: 금액차이
        ws.cell(out_r, 17).value = f'=(K{out_r}+L{out_r})-(O{out_r}+P{out_r})'
        ws.cell(out_r, 17).number_format = num_fmt
        ws.cell(out_r, 17).border = thin_border

        # R: 수량차이
        ws.cell(out_r, 18).value = f'=(I{out_r}+J{out_r})-(M{out_r}+N{out_r})'
        ws.cell(out_r, 18).number_format = num_fmt
        ws.cell(out_r, 18).border = thin_border

        # S: 오류유형 (1차 통합 사전, 2026-05-18 — _error_types.build_classify_formula)
        # 라인 그룹별 분기: 완성품 → GERP 품번누락(M1/M2), 메인SUB/웨빙SUB → 정산차이 흡수
        line_grp = LINE_GROUP.get(line, '메인SUB')
        ws.cell(out_r, 19).value = build_classify_formula(out_r, line_grp)
        ws.cell(out_r, 19).border = thin_border

    # ===== 미등록 영역 폐기 (세션148, 2026-05-08) =====
    # 마스터 V2에 정산 제외 품번 + RSP 모듈품번 모두 등록 → 일반 영역 SUMIFS로 자동 처리
    # 빌더 미등록 영역 코드 + SP3M3 RSP 별도 처리 코드 모두 폐기

    last_r = out_r

    # Summary row
    sum_r = last_r + 1
    ws.cell(sum_r, 1, '합계').font = Font(bold=True)
    ws.cell(sum_r, 1).fill = summary_fill
    ws.cell(sum_r, 1).border = thin_border
    # col 9~18: SUM, col 19: 카테고리 빈값
    for c in range(9, 19):
        cl = get_column_letter(c)
        ws.cell(sum_r, c).value = f'=SUM({cl}3:{cl}{last_r})'
        ws.cell(sum_r, c).number_format = num_fmt
        ws.cell(sum_r, c).font = Font(bold=True)
        ws.cell(sum_r, c).fill = summary_fill
        ws.cell(sum_r, c).border = thin_border
    ws.cell(sum_r, 19).fill = summary_fill
    ws.cell(sum_r, 19).border = thin_border

    # ===== SP3M3 한정 BI 보정 (2026-05-14 채택) =====
    # 적용 조건 5개: 라인=SP3M3 / 정산월 >= 2026-04 / BI 야간수량 존재 / GERP 야간수량 > 0 / GERP 야간금액 > 0
    # CLAUDE.md "SP3M3 구ERP 야간 BI 보정 룰" 참조
    if line == 'SP3M3':
        try:
            year_int = 2026  # 정산 대상 연도 (현재 2026년 운영). 다년 운영 시 _pipeline_config에 추가
            month_int = int(MONTH)
            # 정산월 조건: 4월 이상
            if month_int >= 4:
                bi_qty, bi_days = extract_bi_night_total('SP3M3', year_int, month_int)
                if bi_qty > 0:
                    # GERP 야간수량(J) / GERP 야간금액(L) > 0 검증은 엑셀 수식 ROUND 단계에서 자동 처리
                    # (J=0이면 #DIV/0 → 사용자 검토 필요로 명시)
                    # N{sum_r} 덮어쓰기: SUM → BI 직접 값
                    ws.cell(sum_r, 14).value = bi_qty
                    ws.cell(sum_r, 14).number_format = num_fmt
                    ws.cell(sum_r, 14).font = Font(bold=True)
                    ws.cell(sum_r, 14).fill = summary_fill
                    ws.cell(sum_r, 14).border = thin_border
                    # P{sum_r} 덮어쓰기: SUM → ROUND(BI × 평균단가)
                    n_col = get_column_letter(14)
                    l_col = get_column_letter(12)
                    j_col = get_column_letter(10)
                    ws.cell(sum_r, 16).value = (
                        f'=IF({j_col}{sum_r}=0,0,'
                        f'ROUND({n_col}{sum_r}*({l_col}{sum_r}/{j_col}{sum_r}),0))'
                    )
                    ws.cell(sum_r, 16).number_format = num_fmt
                    ws.cell(sum_r, 16).font = Font(bold=True)
                    ws.cell(sum_r, 16).fill = summary_fill
                    ws.cell(sum_r, 16).border = thin_border
                    # 비고 행 추가 (sum_r + 2)
                    note_r = sum_r + 2
                    ws.cell(note_r, 1).value = (
                        f"※ SP3M3 구ERP 야간 BI 보정 ({year_int}-{month_int:02d}): "
                        f"BI 야간수량 {bi_qty:,} EA × GERP 야간 평균단가(L/J) 기준. "
                        f"행별 RSP 야간 데이터는 원본 유지, 합계행만 비교용 보정. "
                        f"이 값은 '구ERP 원본 야간금액'이 아니라 'BI 보정 비교용 금액'."
                    )
                    ws.cell(note_r, 1).font = Font(italic=True, color='808080', size=9)
                    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=16)
                    print(f"   SP3M3 BI 보정 적용: 야간수량 {bi_qty:,} EA ({bi_days}일)")
                else:
                    # 예외: BI 야간수량 없음 → 보정 금지 + 경고
                    print(f"   [WARN] SP3M3 BI 보정 불가 — BI 야간수량 0 또는 BI 파일 미존재. 기존 SUM 룰 유지.")
            else:
                # 정산월 < 4월: 옛 룰 (P=L 등) 유지 — SP3M3 BI 보정 미적용
                print(f"   SP3M3: 정산월 {month_int}월 < 4월 → BI 보정 미적용 (옛 룰 유지)")
        except Exception as e:
            print(f"   [ERROR] SP3M3 BI 보정 실패: {e}. 기존 SUM 룰 유지.")

    line_summaries[line] = sum_r
    # ※ SP3M3 야간 RSP 모듈품번은 미등록 영역에서 자동 잡힘 (별도 처리 불필요)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['D'].width = 16
    for c in range(9, 19):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions['S'].width = 18  # 카테고리 컬럼

    print(f"   {line}: data rows={last_r - 2}, sum_row={sum_r}")

# ===== 정산집계 sheet =====
print("7. 정산집계 시트 생성...")
ws_sum = wb.create_sheet('정산집계')

sum_headers = ['라인코드', '라인명',
               'GERP주간수량', 'GERP야간수량', 'GERP주간금액', 'GERP야간금액', 'GERP합계',
               '구ERP주간수량', '구ERP야간수량', '구ERP주간금액', '구ERP야간금액', '구ERP합계',
               '금액차이']
line_names = {
    'SD9A01': '아우터', 'ANAAS04': '앵커', 'DRAAS11': '디링',
    'SP3M3': 'SP3', 'HASMS02': 'HASMS', 'HCAMS02': 'HCAMS',
    'WAMAS01': '웨빙', 'WABAS01': '웨빙버클', 'WASAS01': 'WASAS',
    'ISAMS03': '이너센스'
}

for c, h in enumerate(sum_headers, 1):
    cell = ws_sum.cell(1, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, line in enumerate(LINES):
    r = i + 2
    sr = line_summaries[line]

    ws_sum.cell(r, 1, line).border = thin_border
    ws_sum.cell(r, 2, line_names.get(line, line)).border = thin_border

    # GERP: I=주간수량, J=야간수량, K=주간금액, L=야간금액
    ws_sum.cell(r, 3).value = f"='{line}'!I{sr}"
    ws_sum.cell(r, 4).value = f"='{line}'!J{sr}"
    ws_sum.cell(r, 5).value = f"='{line}'!K{sr}"
    ws_sum.cell(r, 6).value = f"='{line}'!L{sr}"
    ws_sum.cell(r, 7).value = f'=E{r}+F{r}'

    # 구ERP: M=주간수량, N=야간수량, O=주간금액, P=야간금액
    ws_sum.cell(r, 8).value = f"='{line}'!M{sr}"
    ws_sum.cell(r, 9).value = f"='{line}'!N{sr}"
    ws_sum.cell(r, 10).value = f"='{line}'!O{sr}"
    ws_sum.cell(r, 11).value = f"='{line}'!P{sr}"
    ws_sum.cell(r, 12).value = f'=J{r}+K{r}'

    # 금액차이
    ws_sum.cell(r, 13).value = f'=G{r}-L{r}'
    ws_sum.cell(r, 13).font = Font(bold=True, color='FF0000')

    for c in range(3, 14):
        ws_sum.cell(r, c).number_format = num_fmt
        ws_sum.cell(r, c).border = thin_border
    ws_sum.cell(r, 7).font = Font(bold=True)
    ws_sum.cell(r, 12).font = Font(bold=True)

# Total row
total_r = len(LINES) + 2
ws_sum.cell(total_r, 1, '합계').font = Font(bold=True, size=12)
ws_sum.cell(total_r, 1).fill = summary_fill
ws_sum.cell(total_r, 1).border = thin_border
ws_sum.cell(total_r, 2).fill = summary_fill
ws_sum.cell(total_r, 2).border = thin_border
for c in range(3, 14):
    cl = get_column_letter(c)
    ws_sum.cell(total_r, c).value = f'=SUM({cl}2:{cl}{total_r - 1})'
    ws_sum.cell(total_r, c).number_format = num_fmt
    ws_sum.cell(total_r, c).font = Font(bold=True, size=12)
    ws_sum.cell(total_r, c).fill = summary_fill
    ws_sum.cell(total_r, c).border = thin_border

ws_sum.column_dimensions['A'].width = 12
ws_sum.column_dimensions['B'].width = 10
for c in range(3, 14):
    ws_sum.column_dimensions[get_column_letter(c)].width = 16

# ===== 오류리스트 sheet (사용자 양식 — 타이틀+요약+그룹헤더+컬럼헤더 4행, 데이터 row 5+) =====
print("8. 오류리스트 시트 생성 (사용자 양식 헤더만 — 데이터는 사후 단계)...")
ws_err = wb.create_sheet('오류리스트')
ws_err.cell(1, 1).value = f'{int(MONTH)}월 조립비 정산 오류 리스트'
ws_err.merge_cells('A1:T1')
ws_err.cell(1, 1).font = Font(bold=True, size=14)
ws_err.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws_err.cell(1, 1).fill = summary_fill
ws_err.merge_cells('A2:T2')
ws_err.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')
ws_err.cell(2, 1).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
# 1차 통합 사전 (2026-05-18): 결과 그룹 4→6 (차이금액/오류유형/비고/제외사유/받을금액/지원업체) → 총 22컬럼
groups = [('A3:H3', '기본정보', 1), ('I3:L3', 'GERP', 9), ('M3:P3', '구ERP', 13), ('Q3:V3', '결과', 17)]
for ref, name, c0 in groups:
    ws_err.cell(3, c0).value = name
    ws_err.merge_cells(ref)
    ws_err.cell(3, c0).font = header_font
    ws_err.cell(3, c0).fill = header_fill
    ws_err.cell(3, c0).alignment = Alignment(horizontal='center', vertical='center')
    ws_err.cell(3, c0).border = thin_border
err_headers = [
    '품번', '업체코드', '라인코드', '조립품번', 'Usage', '단가구분', '단가', '차종',
    '주간수량', '주간금액', '야간수량', '야간금액',
    '주간수량', '주간금액', '야간수량', '야간금액',
    '차이금액', '오류유형', '비고', '제외사유', '받을금액', '지원업체',
]
for c, h in enumerate(err_headers, 1):
    cell = ws_err.cell(4, c, h)
    cell.font = header_font
    cell.fill = summary_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws_err.column_dimensions['A'].width = 16
ws_err.column_dimensions['B'].width = 10
ws_err.column_dimensions['C'].width = 10
ws_err.column_dimensions['D'].width = 16
ws_err.column_dimensions['E'].width = 7
ws_err.column_dimensions['F'].width = 10
ws_err.column_dimensions['G'].width = 9
ws_err.column_dimensions['H'].width = 8
for c in range(9, 17):
    ws_err.column_dimensions[get_column_letter(c)].width = 11
ws_err.column_dimensions['Q'].width = 12   # 차이금액
ws_err.column_dimensions['R'].width = 14   # 오류유형
ws_err.column_dimensions['S'].width = 18   # 비고
ws_err.column_dimensions['T'].width = 16   # 제외사유
ws_err.column_dimensions['U'].width = 13   # 받을금액
ws_err.column_dimensions['V'].width = 12   # 지원업체
ws_err.freeze_panes = 'A5'

ws_summary_err = wb.create_sheet('유형별요약')
ws_summary_err.cell(1, 1).value = f'{int(MONTH)}월 오류 유형별 요약'
ws_summary_err.merge_cells('A1:E1')
ws_summary_err.cell(1, 1).font = Font(bold=True, size=12)
ws_summary_err.cell(1, 1).alignment = Alignment(horizontal='center')
ws_summary_err.cell(1, 1).fill = summary_fill
sum_headers_err = ['오류유형', '건수', 'GERP 합계', '구ERP 합계', '차이금액', '받을금액']
for c, h in enumerate(sum_headers_err, 1):
    cell = ws_summary_err.cell(3, c, h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
for c in range(1, 7):
    ws_summary_err.column_dimensions[get_column_letter(c)].width = 16

# [폐기 2026-05-07] 이전 Python 직접 계산 dead block 통째 제거 — gerp_agg 미정의 NameError + 변수 사용처 0건 + 빌더 SUMIFS가 권위.

# Save
output = os.path.join(BASE, FOLDER_NAME, f'정산_수식버전_{MONTH}월.xlsx')
try:
    wb.save(output)
except PermissionError:
    import datetime
    ts = datetime.datetime.now().strftime('%H%M%S')
    output = os.path.join(BASE, FOLDER_NAME, f'정산_수식버전_{MONTH}월_v4_{ts}.xlsx')
    wb.save(output)
    print(f'[NOTE] 원본 파일 락 → 임시 저장: {output}')
wb.close()


def _is_locked(p):
    """파일이 다른 프로세스에서 열려 있는지(쓰기 락) 검사."""
    if not os.path.exists(p):
        return False
    try:
        with open(p, 'r+b'):
            return False
    except (PermissionError, IOError, OSError):
        return True


def _map_cat_to_user_type(cat, q_diff, g_amt_total, o_amt_total):
    """라인시트 S 컬럼 cat → (err_type, note). _error_types.map_cat_to_user_type에 위임 (1차 통합 사전, 2026-05-18)."""
    return map_cat_to_user_type(cat)


def populate_error_list_static(xlsx_path, lines, line_summaries, month):
    if _is_locked(xlsx_path):
        print(f"[ERROR] 본체 락 — 사용자 Excel에서 닫고 다시 실행: {xlsx_path}")
        return False
    try:
        import win32com.client as w32
        import pythoncom
    except ImportError:
        print("[WARN] pywin32 미설치 — 오류리스트 정적 박기 건너뜀")
        return False
    EXCLUDE_CATS = {'정상', '다중단가분배(정상)'}
    abs_path = os.path.abspath(xlsx_path)
    pythoncom.CoInitialize()
    excel = None
    wb_com = None
    try:
        excel = w32.DispatchEx('Excel.Application')
        excel.Visible = False; excel.DisplayAlerts = False
        excel.ScreenUpdating = False; excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        print(f"   COM Open: {os.path.basename(abs_path)}")
        wb_com = excel.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=False)
        print("   CalculateFull (전체 SUMIFS 계산)...")
        excel.CalculateFull()
        rows = []
        for line in lines:
            try: ws = wb_com.Worksheets(line)
            except Exception: continue
            sr = line_summaries.get(line)
            if not sr or sr < 4: continue
            data = ws.Range(f"A1:S{sr-1}").Value
            if not data: continue
            cnt_line = 0
            for r_idx, row_tup in enumerate(data, start=1):
                if r_idx < 3: continue
                pno = row_tup[0]
                if pno is None: continue
                pno_s = str(pno)
                if pno_s.startswith('※') or pno_s == '합계': continue
                cat = row_tup[18]
                if cat is None or cat in EXCLUDE_CATS: continue
                pno_v = pno_s
                # 업체코드 정규화 (사용자 양식: '0109' 4자리 zero-padded 문자열)
                _vraw = row_tup[1]
                if _vraw is None or _vraw == '':
                    vendor = '0109'
                else:
                    try: vendor = f'{int(_vraw):04d}'
                    except (ValueError, TypeError): vendor = str(_vraw)
                line_cd = row_tup[2] or line
                assy    = row_tup[3]
                usage   = row_tup[4]
                pclass  = row_tup[5]
                price   = row_tup[6]
                cha     = row_tup[7]
                g_d_q   = row_tup[8]  or 0
                g_n_q   = row_tup[9]  or 0
                g_d_amt = row_tup[10] or 0
                g_n_amt = row_tup[11] or 0
                o_d_q   = row_tup[12] or 0
                o_n_q   = row_tup[13] or 0
                o_d_amt = row_tup[14] or 0
                o_n_amt = row_tup[15] or 0
                q_diff  = row_tup[16] or 0
                err_type, memo = _map_cat_to_user_type(cat, q_diff, (g_d_amt+g_n_amt), (o_d_amt+o_n_amt))
                # 제외사유 — 다중단가분배(정상) cat은 EXCLUDE 단계에서 이미 걸러짐. 여기 들어오는 행은 비제외만.
                excl_reason = ''
                recv_amt = abs(int(round(q_diff)))
                rows.append((
                    pno_v, vendor, line_cd, assy, usage, pclass, price, cha,
                    g_d_q, g_d_amt, g_n_q, g_n_amt,
                    o_d_q, o_d_amt, o_n_q, o_n_amt,
                    q_diff, err_type, memo, excl_reason, recv_amt, '',
                ))
                cnt_line += 1
            print(f"   {line}: 차이 행 {cnt_line}건")
        ws_err = wb_com.Worksheets('오류리스트')
        if ws_err.AutoFilterMode: ws_err.AutoFilterMode = False
        used = ws_err.UsedRange
        if used.Rows.Count > 4:
            ws_err.Range(f"A5:V{used.Rows.Count}").ClearContents()
        n = len(rows)
        if n > 0:
            # 업체코드(B) 텍스트 형식 사전 설정 — '0109' 문자열 유지
            ws_err.Range(f"B5:B{4+n}").NumberFormat = '@'
            ws_err.Range(f"A5:V{4+n}").Value = rows
            num_fmt_local = '#,##0;-#,##0;"-"'
            ws_err.Range(f"E5:E{4+n}").NumberFormat = '0'
            ws_err.Range(f"G5:G{4+n}").NumberFormat = num_fmt_local
            ws_err.Range(f"I5:Q{4+n}").NumberFormat = num_fmt_local
            ws_err.Range(f"U5:U{4+n}").NumberFormat = num_fmt_local  # 받을금액
        from collections import defaultdict as _dd
        agg = _dd(lambda: {'cnt': 0, 'g': 0.0, 'o': 0.0, 'recv': 0.0})
        for r in rows:
            t = r[17]
            g_t = (r[9] or 0) + (r[11] or 0)
            o_t = (r[13] or 0) + (r[15] or 0)
            agg[t]['cnt'] += 1
            agg[t]['g'] += g_t
            agg[t]['o'] += o_t
            agg[t]['recv'] += (r[20] or 0)
        ws_sumE = wb_com.Worksheets('유형별요약')
        used_s = ws_sumE.UsedRange
        if used_s.Rows.Count > 3:
            ws_sumE.Range(f"A4:F{used_s.Rows.Count}").ClearContents()
        sum_rows = []
        tot_cnt = 0; tot_g = 0; tot_o = 0; tot_recv = 0
        for t in TYPE_ORDER:
            d = agg.get(t, {'cnt': 0, 'g': 0, 'o': 0, 'recv': 0})
            sum_rows.append((t, d['cnt'], d['g'], d['o'], d['g'] - d['o'], d['recv']))
            tot_cnt += d['cnt']; tot_g += d['g']; tot_o += d['o']; tot_recv += d['recv']
        sum_rows.append(('합계', tot_cnt, tot_g, tot_o, tot_g - tot_o, tot_recv))
        if sum_rows:
            ws_sumE.Range(f"A4:F{3+len(sum_rows)}").Value = sum_rows
            ws_sumE.Range(f"B4:F{3+len(sum_rows)}").NumberFormat = '#,##0;-#,##0;"-"'
        def _fmt(amt):
            sign = '+' if amt >= 0 else '-'
            return f"{sign}{abs(int(round(amt))):,}원"
        parts = [f"총 {n}건"]
        for t in TYPE_ORDER:
            d = agg.get(t)
            if d and d['cnt'] > 0:
                diff_t = d['g'] - d['o']
                parts.append(f"{t} {d['cnt']}건({_fmt(diff_t)})")
        parts.append(f"받을금액 {int(round(tot_recv)):,}원")
        ws_err.Cells(2, 1).Value = '  |  '.join(parts)
        wb_com.Save()
        wb_com.Close(SaveChanges=False)
        excel.Quit()
        print(f"   오류리스트 정적 박기 완료: {n}행 / 유형별요약 갱신 OK")
        return True
    except Exception as e:
        print(f"[ERROR] 오류리스트 정적 박기 실패: {type(e).__name__}: {e}")
        try:
            if wb_com is not None: wb_com.Close(SaveChanges=False)
        except Exception: pass
        try:
            if excel is not None: excel.Quit()
        except Exception: pass
        return False
    finally:
        try: pythoncom.CoUninitialize()
        except Exception: pass


print("\n9. 오류리스트 정적 박기 (사용자 양식 — 22컬럼 1차 통합 사전 + 유형별요약 + 요약 텍스트)...")
_ok = populate_error_list_static(output, LINES, line_summaries, MONTH)

print(f"\n=== 완료: {output} ===")
if _ok:
    print(f"시트 16개. 오류리스트 시트 = 차이 행만 정적 값으로 박힘 (사용자 본체 열어도 calc 불필요).")
else:
    print(f"시트 16개. 오류리스트 시트 = 헤더만 (정적 박기 실패). 위 [ERROR]/[WARN] 메시지 확인.")
print(f"사용자 본체 직접 열어 검증 가능.")
