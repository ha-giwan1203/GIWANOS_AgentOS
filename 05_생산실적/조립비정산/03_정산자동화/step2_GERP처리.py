#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Step 2 — GERP 처리
0109 필터 → 주야 분리(정상/추가) → 라인별 피벗 생성 → JSON 출력

실행: python step2_gerp처리.py
출력: _cache/step2_gerp.json
"""

import sys, os, json
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from _pipeline_config import *
import pandas as pd
import openpyxl
from datetime import datetime

print("=" * 60)
print("Step 2: GERP 처리")
print("=" * 60)

if not os.path.exists(GERP_FILE):
    print(f"[ERROR] GERP 파일 없음: {GERP_FILE}")
    sys.exit(1)

# ── 로딩 ──────────────────────────────────────────────────────
print(f"\n[1/3] GERP 파일 로딩...")
df_raw = pd.read_excel(GERP_FILE, sheet_name=0, header=None)
data = df_raw.iloc[2:].reset_index(drop=True)

c = GERP_COL
all_gerp = pd.DataFrame({
    'line':       data.iloc[:, c['line']].astype(str).str.strip(),
    'product_no': data.iloc[:, c['product_no']].astype(str).str.strip(),
    'usage':      pd.to_numeric(data.iloc[:, c['usage']], errors='coerce').fillna(1).astype(int),
    'assy_part':  data.iloc[:, c['assy_part']].astype(str).str.strip(),
    'shift':      data.iloc[:, c['shift']].astype(str).str.strip(),
    'qty':        pd.to_numeric(data.iloc[:, c['qty']], errors='coerce').fillna(0).astype(int),
    'unit_price': pd.to_numeric(data.iloc[:, c['unit_price']], errors='coerce').fillna(0),
    'amount':     pd.to_numeric(data.iloc[:, c['amount']], errors='coerce').fillna(0),
    'vendor_cd':  data.iloc[:, c['vendor_cd']].astype(str).str.strip(),
    # 차종 컬럼 (col 5, 0-based) — SVM/OVK 이관품번 제외 분기용 (빌더와 정합)
    'vtype':      data.iloc[:, 5].astype(str).str.strip(),
})

print(f"  GERP 전체: {len(all_gerp):,}행")

# ── 0109 필터 ─────────────────────────────────────────────────
print(f"\n[2/3] 대원테크({VENDOR_CODE}) 필터링 및 주야 분리...")
gerp_dw = all_gerp[all_gerp['vendor_cd'] == VENDOR_CODE].copy()
gerp_dw['shift_type'] = gerp_dw['shift'].map({'정상': '주간', '추가': '야간'}).fillna('주간')
print(f"  대원테크: {len(gerp_dw):,}행")

# ── 이관품번(SVM/OVK) 제외 — 빌더와 동일 로직 ─────────────────
# SVM 차종: SP3M3/HCAMS02/ISAMS03 → 이관 (정산 제외)
# OVK 차종: SD9A01/ANAAS04/DRAAS11 → 이관 (정산 제외)
SVM_TRANSFER_LINES = {'SP3M3', 'HCAMS02', 'ISAMS03'}
OVK_TRANSFER_LINES = {'SD9A01', 'ANAAS04', 'DRAAS11'}
before_n = len(gerp_dw)
svm_mask = gerp_dw['line'].isin(SVM_TRANSFER_LINES) & (gerp_dw['vtype'] == 'SVM')
# OVK: 단 89880X 시작 품번은 OVK여도 이관 아님 (사용자 룰 2026-05-07)
ovk_mask = (
    gerp_dw['line'].isin(OVK_TRANSFER_LINES)
    & (gerp_dw['vtype'] == 'OVK')
    & ~gerp_dw['product_no'].str.startswith('89880X')
)
# 88820X 시작 품번은 완성품 라인 한정 정산 제외 (사용자 룰 2026-06-01 정정 — 구ERP 미등록)
# 이전 룰 "차종/라인 무관 일괄 제외"는 SUB 라인 부당 제외 사고 발생 (5월 39행/1,108,391원)
# 정정: 완성품 3라인(SD9A01/ANAAS04/DRAAS11)만 제외. SUB 라인은 정산 대상.
X88820_TRANSFER_LINES = {'SD9A01', 'ANAAS04', 'DRAAS11'}
x88820_mask = gerp_dw['product_no'].str.startswith('88820X') & gerp_dw['line'].isin(X88820_TRANSFER_LINES)
svm_n = svm_mask.sum(); svm_amt = gerp_dw.loc[svm_mask, 'amount'].sum()
ovk_n = ovk_mask.sum(); ovk_amt = gerp_dw.loc[ovk_mask, 'amount'].sum()
x88820_n = x88820_mask.sum(); x88820_amt = gerp_dw.loc[x88820_mask, 'amount'].sum()
gerp_dw = gerp_dw[~(svm_mask | ovk_mask | x88820_mask)].copy()
print(f"  이관 제외: SVM {svm_n}행({svm_amt:,.0f}원) + OVK {ovk_n}행({ovk_amt:,.0f}원) + 88820X {x88820_n}행({x88820_amt:,.0f}원) = {before_n - len(gerp_dw)}행 / {svm_amt + ovk_amt + x88820_amt:,.0f}원 (정산대상 {len(gerp_dw):,}행)")

# 라인별 요약 출력
for lc in LINE_ORDER:
    sub = gerp_dw[gerp_dw['line'] == lc]
    day_q   = sub[sub['shift_type'] == '주간']['qty'].sum()
    night_q = sub[sub['shift_type'] == '야간']['qty'].sum()
    print(f"  {lc:10s}: 주간 {day_q:>8,}  야간 {night_q:>8,}")

# ── SP3M3 모듈품번(RSP) → 기본품번 매핑 로딩 ──────────────────
rsp_to_base = {}
# 1차: 모듈품번 파일
if os.path.exists(SP3M3_MODULE_FILE):
    wb_mod = openpyxl.load_workbook(SP3M3_MODULE_FILE, data_only=True, read_only=True)
    ws_mod = wb_mod.active
    for row in ws_mod.iter_rows(min_row=2, values_only=True):
        if row[3] and row[1]:  # D=모듈품번, B=기본품번
            rsp = str(row[3]).strip()
            base_pn = str(row[1]).strip()
            rsp_to_base[rsp] = base_pn
    wb_mod.close()
    print(f"\n[RSP매핑] 1차 모듈품번 파일: {len(rsp_to_base)}건")
else:
    print(f"\n[RSP매핑] 1차 파일 없음: {SP3M3_MODULE_FILE}")

# 2차: 라인배정 ENDPART 파일 (1차에 없는 RSP만 보충)
added_from_la = 0
if os.path.exists(LINE_ASSIGN_FILE):
    wb_la = openpyxl.load_workbook(LINE_ASSIGN_FILE, data_only=True, read_only=True)
    ws_la = wb_la[wb_la.sheetnames[0]]
    for row in ws_la.iter_rows(min_row=2, values_only=True):
        if len(row) < 27:
            continue
        rsp = str(row[26]).strip() if row[26] else ''   # col AA (0-based 26)
        pn13 = str(row[8]).strip() if row[8] else ''    # col I (0-based 8)
        if rsp.startswith('RSP') and pn13 and rsp not in rsp_to_base:
            rsp_to_base[rsp] = pn13[:10]  # 컬러코드 제거
            added_from_la += 1
    wb_la.close()
    if added_from_la:
        print(f"[RSP매핑] 2차 라인배정 보충: +{added_from_la}건 → 총 {len(rsp_to_base)}건")
else:
    print(f"[RSP매핑] 2차 파일 없음: {LINE_ASSIGN_FILE}")

# SP3M3 야간행 RSP → 기본품번 변환
rsp_converted = 0
rsp_missing = []
sp3m3_night = gerp_dw[(gerp_dw['line'] == 'SP3M3') & (gerp_dw['shift_type'] == '야간')]
for idx in sp3m3_night.index:
    pn = gerp_dw.at[idx, 'product_no']
    if pn.startswith('RSP'):
        base = rsp_to_base.get(pn)
        if base:
            gerp_dw.at[idx, 'product_no'] = base
            rsp_converted += 1
        else:
            rsp_missing.append(pn)
if rsp_converted:
    print(f"  SP3M3 야간 RSP→기본품번 변환: {rsp_converted}건")
if rsp_missing:
    print(f"  ⚠ RSP 매핑 미존재: {rsp_missing[:10]}")

# ── 피벗 생성 ─────────────────────────────────────────────────
print(f"\n[3/3] 라인별 품번 피벗 생성...")
day_pivot   = {}
night_pivot = {}
day_amt_pivot   = {}
night_amt_pivot = {}
unmatched_lines = []

for lc in LINE_ORDER:
    sub = gerp_dw[gerp_dw['line'] == lc]
    if len(sub) == 0:
        unmatched_lines.append(lc)
        continue

    day_sub   = sub[sub['shift_type'] == '주간']
    night_sub = sub[sub['shift_type'] == '야간']

    if len(day_sub):
        day_pivot[lc] = day_sub.groupby('product_no')['qty'].sum().astype(int).to_dict()
        day_amt_pivot[lc] = day_sub.groupby('product_no')['amount'].sum().to_dict()
    if len(night_sub):
        night_pivot[lc] = night_sub.groupby('product_no')['qty'].sum().astype(int).to_dict()
        night_amt_pivot[lc] = night_sub.groupby('product_no')['amount'].sum().to_dict()

    day_pn   = len(day_pivot.get(lc, {}))
    night_pn = len(night_pivot.get(lc, {}))
    print(f"  {lc:10s}: 주간품번 {day_pn:>4}개  야간품번 {night_pn:>4}개")

if unmatched_lines:
    print(f"\n  ⚠ GERP 데이터 없는 라인: {unmatched_lines}")

# ── GERP 조립품번 lookup: (라인, 품번, 단가) → 조립품번 ──────
gerp_assy_lookup = {}
for _, row in gerp_dw.iterrows():
    key = f"{row['line']}|{row['product_no']}|{row['unit_price']}"
    ap = row.get('assy_part', '')
    if ap and ap != 'nan' and ap != 'None':
        gerp_assy_lookup[key] = ap
print(f"\n  GERP 조립품번 lookup: {len(gerp_assy_lookup):,}건")

# ── JSON 저장 ─────────────────────────────────────────────────
result = {
    "step": 2,
    "timestamp": datetime.now().isoformat(),
    "total_rows": len(gerp_dw),
    "day_pivot": day_pivot,
    "night_pivot": night_pivot,
    "day_amt_pivot": day_amt_pivot,
    "night_amt_pivot": night_amt_pivot,
    "unmatched_lines": unmatched_lines,
    "gerp_assy_lookup": gerp_assy_lookup,
    # 전체 GERP (all vendor) 주야 피벗 — Step4 미매핑 참고용
    "all_gerp_lines": sorted(all_gerp['line'].dropna().unique().tolist()),
}

with open(CACHE_STEP2, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n저장: {CACHE_STEP2}")
print("Step 2 완료")
