#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Step 8 — 오류 리스트 생성
step5_settlement.json에서 GERP vs 구ERP 차이 항목을 추출하여
DB 양식 기반 오류리스트_MM월.xlsx를 생성한다.

실행: python step8_오류리스트.py
출력: {월별폴더}/오류리스트_MM월.xlsx
"""

import sys, os, json
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from _pipeline_config import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
from datetime import datetime

print("=" * 60)
print("Step 8: 오류 리스트 생성")
print("=" * 60)

# ── 데이터 로드 ──
if not os.path.exists(CACHE_STEP5):
    print(f"[ERROR] step5 캐시 없음: {CACHE_STEP5}")
    sys.exit(1)

with open(CACHE_STEP5, 'r', encoding='utf-8') as f:
    s5 = json.load(f)

# 이관 완료 품번 제외 (오류 아님)
EXCLUDED_PATTERNS = ['X9000', 'X9500']

def is_excluded(part_no):
    for pat in EXCLUDED_PATTERNS:
        if pat in str(part_no):
            return True
    return False

print(f"\n[1/3] 오류 항목 수집...")

# ── 오류 항목 수집 ──
errors = []
for lc in LINE_ORDER:
    ld = s5['lines'].get(lc, {})
    for r in ld.get('items', []):
        g_amt = r['gerp_day_amt'] + r['gerp_ngt_amt']
        e_amt = r['erp_day_amt'] + r['erp_ngt_amt']
        diff = g_amt - e_amt
        g_qty = r['gerp_day_qty'] + r['gerp_ngt_qty']
        e_qty = r['erp_day_qty'] + r['erp_ngt_qty']

        if g_amt == e_amt:
            continue

        # 이관 완료 품번은 오류 아님 — 제외
        if is_excluded(r['part_no']):
            continue

        if e_amt == 0 and g_amt > 0:
            err_type = '구실적누락'
        elif g_amt == 0 and e_amt > 0:
            err_type = 'GERP누락'
        elif r['price'] == 0:
            err_type = '기준누락'
        elif g_qty != e_qty:
            err_type = '수량차이'
        else:
            err_type = '정산차이'

        sup_text = ''
        if r.get('support'):
            parts = []
            for s in r['support']:
                v = s['vendor']
                if '[' in v:
                    v = v.split('[')[1].rstrip(']')
                parts.append(f"{v}({s['day_qty']+s['night_qty']})")
            sup_text = ', '.join(parts)

        errors.append({
            'line': lc,
            'part_no': r['part_no'],
            'assy_part': r.get('assy_part', ''),
            'usage': r.get('usage', 1),
            'price_type': r['price_type'],
            'price': r['price'],
            'vtype': r.get('vtype', ''),
            'gerp_day_qty': r['gerp_day_qty'],
            'gerp_day_amt': r['gerp_day_amt'],
            'gerp_ngt_qty': r['gerp_ngt_qty'],
            'gerp_ngt_amt': r['gerp_ngt_amt'],
            'erp_day_qty': r['erp_day_qty'],
            'erp_day_amt': r['erp_day_amt'],
            'erp_ngt_qty': r['erp_ngt_qty'],
            'erp_ngt_amt': r['erp_ngt_amt'],
            'diff': diff,
            'err_type': err_type,
            'sup_text': sup_text,
        })

type_order = {'GERP누락':0, '구실적누락':1, '기준누락':2, '수량차이':3, '정산차이':4}
errors.sort(key=lambda x: (LINE_ORDER.index(x['line']),
                           type_order.get(x['err_type'], 9),
                           -abs(x['diff'])))

print(f"  오류 {len(errors)}건 추출")

# ── 유형별 요약 출력 ──
tcnt = Counter(e['err_type'] for e in errors)
tamt = {}
for e in errors:
    tamt[e['err_type']] = tamt.get(e['err_type'], 0) + e['diff']
for t in ['구실적누락', 'GERP누락', '수량차이', '정산차이', '기준누락']:
    if t in tcnt:
        print(f"    {t}: {tcnt[t]}건 ({tamt[t]:+,}원)")

# ── 스타일 ──
print(f"\n[2/3] 엑셀 생성...")

THIN = Side(style='thin', color='D0D0D0')
BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CA = Alignment(horizontal='center', vertical='center', wrap_text=True)
RA = Alignment(horizontal='right', vertical='center')
LA = Alignment(horizontal='left', vertical='center')
NUM = '#,##0'

DB_FILL = PatternFill('solid', fgColor='1C2B4A')
DB_FONT = Font(name='DejaVu Sans', size=9, bold=True, color='FFFFFF')
G_FILL = PatternFill('solid', fgColor='2B579A')
G_FONT = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
E_FILL = PatternFill('solid', fgColor='2D7D46')
E_FONT = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
RES_FILL = PatternFill('solid', fgColor='4A4A4A')
RES_FONT = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')

DATA = Font(name='맑은 고딕', size=9)
BOLD = Font(name='맑은 고딕', size=9, bold=True)
SUM_FILL = PatternFill('solid', fgColor='E8E8E8')
NEG_FILL = PatternFill('solid', fgColor='FDE8E8')
POS_FILL = PatternFill('solid', fgColor='E8F5E8')

TYPE_FILLS = {
    'GERP누락':   PatternFill('solid', fgColor='FFF2CC'),
    '구실적누락': PatternFill('solid', fgColor='D6E4F7'),
    '기준누락':   PatternFill('solid', fgColor='F4CCCC'),
    '수량차이':   PatternFill('solid', fgColor='FCE5CD'),
    '정산차이':   PatternFill('solid', fgColor='E8D5F5'),
}

# ── 워크북 ──
wb = Workbook()
ws = wb.active
ws.title = '오류리스트'

month_int = int(MONTH)

# Row 1: 타이틀
ws.merge_cells('A1:T1')
ws['A1'] = f'{month_int}월 조립비 정산 오류 리스트'
ws['A1'].font = Font(name='맑은 고딕', bold=True, size=12, color='1B2A4A')
ws['A1'].alignment = LA

# Row 2: 요약
summary = []
for t in ['구실적누락', 'GERP누락', '수량차이', '정산차이', '기준누락']:
    if t in tcnt:
        summary.append(f"{t} {tcnt[t]}건({tamt[t]:+,}원)")
ws.merge_cells('A2:T2')
ws['A2'] = f"총 {len(errors)}건  |  " + '  |  '.join(summary)
ws['A2'].font = Font(name='맑은 고딕', size=9, color='555555')

# Row 3: 그룹 헤더
groups = [
    ('기본정보', 8, DB_FILL, DB_FONT),
    ('GERP', 4, G_FILL, G_FONT),
    ('구ERP', 4, E_FILL, E_FONT),
    ('결과', 4, RES_FILL, RES_FONT),
]
col = 1
for name, span, fill, font in groups:
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+span-1)
    c = ws.cell(3, col, name)
    c.font = font; c.fill = fill; c.alignment = CA; c.border = BDR
    for ci in range(col+1, col+span):
        cc = ws.cell(3, ci)
        cc.fill = fill; cc.border = BDR
    col += span

# Row 4: 세부 헤더
headers = [
    ('품번', 18, DB_FILL, DB_FONT), ('업체코드', 8, DB_FILL, DB_FONT),
    ('라인코드', 10, DB_FILL, DB_FONT), ('조립품번', 16, DB_FILL, DB_FONT),
    ('Usage', 7, DB_FILL, DB_FONT), ('단가구분', 8, DB_FILL, DB_FONT),
    ('단가', 10, DB_FILL, DB_FONT), ('차종', 8, DB_FILL, DB_FONT),
    ('주간수량', 10, G_FILL, G_FONT), ('주간금액', 13, G_FILL, G_FONT),
    ('야간수량', 10, G_FILL, G_FONT), ('야간금액', 13, G_FILL, G_FONT),
    ('주간수량', 10, E_FILL, E_FONT), ('주간금액', 13, E_FILL, E_FONT),
    ('야간수량', 10, E_FILL, E_FONT), ('야간금액', 13, E_FILL, E_FONT),
    ('차이금액', 13, RES_FILL, RES_FONT), ('오류유형', 10, RES_FILL, RES_FONT),
    ('지원업체', 12, RES_FILL, RES_FONT), ('비고', 15, RES_FILL, RES_FONT),
]
for ci, (name, width, fill, font) in enumerate(headers, 1):
    c = ws.cell(4, ci, name)
    c.font = font; c.fill = fill; c.alignment = CA; c.border = BDR
    ws.column_dimensions[get_column_letter(ci)].width = width

# ── 데이터 행 ──
row = 5
for e in errors:
    note = ''
    if e['err_type'] == 'GERP누락':
        note = '구ERP에만 실적'
    elif e['err_type'] == '구실적누락':
        note = 'GERP에만 실적'
    elif e['err_type'] == '기준누락':
        note = '기준단가 없음'

    vals = [
        (e['part_no'], LA, None), (VENDOR_CODE, CA, None),
        (e['line'], CA, None), (e['assy_part'], LA, None),
        (e['usage'], CA, None), (e['price_type'], CA, None),
        (e['price'], RA, NUM), (e['vtype'], CA, None),
        (e['gerp_day_qty'], RA, NUM), (e['gerp_day_amt'], RA, NUM),
        (e['gerp_ngt_qty'], RA, NUM), (e['gerp_ngt_amt'], RA, NUM),
        (e['erp_day_qty'], RA, NUM), (e['erp_day_amt'], RA, NUM),
        (e['erp_ngt_qty'], RA, NUM), (e['erp_ngt_amt'], RA, NUM),
        (e['diff'], RA, NUM), (e['err_type'], CA, None),
        (e['sup_text'], LA, None), (note, LA, None),
    ]
    for ci, (v, align, fmt) in enumerate(vals, 1):
        c = ws.cell(row, ci, v)
        c.font = DATA; c.alignment = align; c.border = BDR
        if fmt:
            c.number_format = fmt

    tf = TYPE_FILLS.get(e['err_type'])
    if tf:
        ws.cell(row, 18).fill = tf

    if e['diff'] < 0:
        ws.cell(row, 17).fill = NEG_FILL
        ws.cell(row, 17).font = Font(name='맑은 고딕', size=9, color='CC0000')
    elif e['diff'] > 0:
        ws.cell(row, 17).fill = POS_FILL
        ws.cell(row, 17).font = Font(name='맑은 고딕', size=9, color='006600')
    row += 1

# ── 합계행 ──
for ci in range(1, 21):
    c = ws.cell(row, ci)
    c.fill = SUM_FILL; c.border = BDR; c.font = BOLD
ws.cell(row, 1, '합계').alignment = CA

for ci, key in [(9,'gerp_day_qty'),(10,'gerp_day_amt'),
                (11,'gerp_ngt_qty'),(12,'gerp_ngt_amt')]:
    c = ws.cell(row, ci, sum(e[key] for e in errors))
    c.number_format = NUM; c.font = BOLD; c.alignment = RA
    c.fill = SUM_FILL; c.border = BDR

for ci, key in [(13,'erp_day_qty'),(14,'erp_day_amt'),
                (15,'erp_ngt_qty'),(16,'erp_ngt_amt')]:
    c = ws.cell(row, ci, sum(e[key] for e in errors))
    c.number_format = NUM; c.font = BOLD; c.alignment = RA
    c.fill = SUM_FILL; c.border = BDR

total_diff = sum(e['diff'] for e in errors)
c = ws.cell(row, 17, total_diff)
c.number_format = NUM; c.font = BOLD; c.alignment = RA
c.fill = NEG_FILL if total_diff < 0 else POS_FILL; c.border = BDR

ws.cell(row, 18, f"{len(errors)}건").font = BOLD
ws.cell(row, 18).alignment = CA; ws.cell(row, 18).fill = SUM_FILL; ws.cell(row, 18).border = BDR

# ── 유형별 요약 시트 ──
ws2 = wb.create_sheet('유형별요약')
ws2['A1'] = '오류 유형별 요약'
ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1B2A4A')

sum_hdrs = ['오류유형', '건수', 'GERP 합계', '구ERP 합계', '차이금액']
for ci, h in enumerate(sum_hdrs, 1):
    c = ws2.cell(3, ci, h)
    c.font = RES_FONT; c.fill = RES_FILL; c.alignment = CA; c.border = BDR
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15

r = 4
for t in ['구실적누락', 'GERP누락', '수량차이', '정산차이', '기준누락']:
    items = [e for e in errors if e['err_type'] == t]
    if not items:
        continue
    g_sum = sum(e['gerp_day_amt']+e['gerp_ngt_amt'] for e in items)
    e_sum = sum(e['erp_day_amt']+e['erp_ngt_amt'] for e in items)
    d = g_sum - e_sum
    vals2 = [t, len(items), g_sum, e_sum, d]
    for ci, v in enumerate(vals2, 1):
        c = ws2.cell(r, ci, v)
        c.font = DATA; c.border = BDR
        c.alignment = CA if ci <= 2 else RA
        if isinstance(v, (int, float)) and ci >= 3:
            c.number_format = NUM
    tf = TYPE_FILLS.get(t)
    if tf:
        ws2.cell(r, 1).fill = tf
    if d < 0:
        ws2.cell(r, 5).fill = NEG_FILL
    elif d > 0:
        ws2.cell(r, 5).fill = POS_FILL
    r += 1

for ci in range(1, 6):
    ws2.cell(r, ci).fill = SUM_FILL; ws2.cell(r, ci).border = BDR; ws2.cell(r, ci).font = BOLD
ws2.cell(r, 1, '합계').alignment = CA
ws2.cell(r, 2, len(errors)).alignment = CA; ws2.cell(r, 2).number_format = NUM
g_all = sum(e['gerp_day_amt']+e['gerp_ngt_amt'] for e in errors)
e_all = sum(e['erp_day_amt']+e['erp_ngt_amt'] for e in errors)
ws2.cell(r, 3, g_all).alignment = RA; ws2.cell(r, 3).number_format = NUM
ws2.cell(r, 4, e_all).alignment = RA; ws2.cell(r, 4).number_format = NUM
ws2.cell(r, 5, total_diff).alignment = RA; ws2.cell(r, 5).number_format = NUM
ws2.cell(r, 5).fill = NEG_FILL if total_diff < 0 else POS_FILL

ws.freeze_panes = 'A5'
ws2.freeze_panes = 'A4'

# ── 저장 ──
print(f"\n[3/3] 저장...")
output_dir = os.path.dirname(OUTPUT_FILE)
error_file = os.path.join(output_dir, f'오류리스트_{MONTH}월.xlsx')
wb.save(error_file)

print(f"\n저장: {error_file}")
print(f"오류 {len(errors)}건 | 차이합계: {total_diff:+,}원")
print("Step 8 완료")
