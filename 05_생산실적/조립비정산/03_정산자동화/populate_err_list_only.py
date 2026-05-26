#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""본체 오류리스트 시트 정적 박기 단독 진입점 (openpyxl, COM-free).

build_formula_version의 populate_error_list_static이 COM 충돌로 실패할 때,
step5 캐시(step5_settlement.json)를 source-of-truth로 본체 오류리스트 시트에
직접 22컬럼 데이터를 박는다. Excel COM 의존성 없음.

사용법:
    python populate_err_list_only.py
"""
import sys, os, json, math
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter

from _pipeline_config import BASE_DIR, MONTH, LINE_ORDER, VENDOR_CODE, CACHE_STEP5, LINE_GROUP
from _error_types import TYPE_ORDER, TYPE_COLORS

MISSING_PRICE_NOTE = '단가 미매핑(기준정보 등록필요)'
GERP_MISSING_SUPPLEMENT_LINES = {'SD9A01'}


def _num(value, default=0):
    try:
        n = float(value)
    except (TypeError, ValueError):
        return default
    if math.isnan(n) or math.isinf(n):
        return default
    return int(n) if n.is_integer() else n


def _is_finished_gerp_missing(row, line_group):
    if line_group != '완성품':
        return False
    if not str(row.get('err_type', '')).startswith('GERP 품번누락'):
        return False
    g_qty = _num(row.get('gerp_day_qty')) + _num(row.get('gerp_ngt_qty'))
    e_qty = _num(row.get('erp_day_qty')) + _num(row.get('erp_ngt_qty'))
    return g_qty == 0 and e_qty > 0


def _is_qty_only_gerp_missing(row, line_group, g_amt=None, e_amt=None):
    if not _is_finished_gerp_missing(row, line_group):
        return False
    if g_amt is None:
        g_amt = _num(row.get('gerp_day_amt')) + _num(row.get('gerp_ngt_amt'))
    if e_amt is None:
        e_amt = _num(row.get('erp_day_amt')) + _num(row.get('erp_ngt_amt'))
    return g_amt == 0 and e_amt == 0

_m_int = int(MONTH) + 1
if _m_int > 12:
    _m_int -= 12
folder = f'{_m_int:02d}월'
output = os.path.join(BASE_DIR, folder, f'정산_수식버전_{MONTH}월.xlsx')

if not os.path.exists(output):
    print(f'[ERROR] 본체 미존재: {output}')
    sys.exit(1)
if not os.path.exists(CACHE_STEP5):
    print(f'[ERROR] step5 캐시 미존재. 먼저 step5_정산계산.py 실행: {CACHE_STEP5}')
    sys.exit(1)

print(f'본체: {output}')
print(f'step5 캐시: {CACHE_STEP5}')

# ── step5 캐시 → 오류 데이터 수집 ──
with open(CACHE_STEP5, encoding='utf-8') as f:
    s5 = json.load(f)

errors = []
excluded = []
for lc in LINE_ORDER:
    ld = s5['lines'].get(lc, {})
    line_group = LINE_GROUP.get(lc, '메인SUB')
    for r in ld.get('items', []):
        if not r.get('is_first_gerp', True):
            continue
        g_amt = _num(r.get('gerp_day_amt')) + _num(r.get('gerp_ngt_amt'))
        e_amt = _num(r.get('erp_day_amt')) + _num(r.get('erp_ngt_amt'))
        diff = g_amt - e_amt
        qty_only_gerp_missing = (
            _is_qty_only_gerp_missing(r, line_group, g_amt, e_amt)
            and r.get('note') == MISSING_PRICE_NOTE
        )
        if g_amt == e_amt and not qty_only_gerp_missing:
            continue
        excl = r.get('excl_reason', '')
        if excl and not qty_only_gerp_missing:
            excluded.append({'line': lc, 'part_no': r['part_no'], 'diff': diff, 'excl': excl})
            continue
        sup_text = ''
        if r.get('support'):
            parts = []
            for s in r['support']:
                v = s['vendor']
                if '[' in v:
                    v = v.split('[')[1].rstrip(']')
                parts.append(f"{v}({s['day_qty']+s['night_qty']})")
            sup_text = ', '.join(parts)
        err_type = 'GERP 품번누락' if qty_only_gerp_missing else r.get('err_type', '정산차이')
        note = MISSING_PRICE_NOTE if qty_only_gerp_missing else r.get('note', '')
        recv_amt = 0 if qty_only_gerp_missing else r.get('recv_amt', abs(diff))
        errors.append({
            'line': lc, 'part_no': r['part_no'], 'assy_part': r.get('assy_part', ''),
            'usage': r.get('usage', 1), 'price_type': r['price_type'],
            'price': r['price'], 'vtype': r.get('vtype', ''),
            'gerp_day_qty': r['gerp_day_qty'], 'gerp_day_amt': r['gerp_day_amt'],
            'gerp_ngt_qty': r['gerp_ngt_qty'], 'gerp_ngt_amt': r['gerp_ngt_amt'],
            'erp_day_qty': r['erp_day_qty'], 'erp_day_amt': r['erp_day_amt'],
            'erp_ngt_qty': r['erp_ngt_qty'], 'erp_ngt_amt': r['erp_ngt_amt'],
            'diff': diff, 'err_type': err_type,
            'note': note, 'recv_amt': recv_amt,
            'sup_text': sup_text,
        })

existing_keys = {(e['line'], e['part_no']) for e in errors}
cache_item_keys = {
    (lc, r.get('part_no'))
    for lc, ld in s5.get('lines', {}).items()
    for r in ld.get('items', [])
}


def _append_line_sheet_gerp_missing():
    """캐시에 빠진 SD9A01 완성품 GERP 품번누락은 라인시트 계산값으로 보강."""
    wb_ro = load_workbook(output, data_only=False, read_only=True)
    added = 0
    try:
        gerp_qty_amt = defaultdict(lambda: {'day_qty': 0, 'night_qty': 0, 'night_amt': 0})
        if 'GERP_입력' in wb_ro.sheetnames:
            ws_gerp = wb_ro['GERP_입력']
            for gr in ws_gerp.iter_rows(min_row=2, values_only=True):
                line = str(gr[2] or '').strip()
                pn = str(gr[6] or '').strip()
                shift = str(gr[13] or '').strip()
                if not line or not pn:
                    continue
                key = (line, pn)
                if shift == '추가':
                    gerp_qty_amt[key]['night_qty'] += _num(gr[14])
                    gerp_qty_amt[key]['night_amt'] += _num(gr[16])
                else:
                    gerp_qty_amt[key]['day_qty'] += _num(gr[14])

        olderp_qty = defaultdict(lambda: {'day_qty': 0, 'night_qty': 0})
        if '구ERP_입력' in wb_ro.sheetnames:
            ws_olderp = wb_ro['구ERP_입력']
            for er in ws_olderp.iter_rows(min_row=2, values_only=True):
                pn = str(er[0] or '').strip()
                if not pn:
                    continue
                olderp_qty[pn]['day_qty'] += _num(er[1])
                olderp_qty[pn]['night_qty'] += _num(er[2])

        for lc in GERP_MISSING_SUPPLEMENT_LINES:
            if LINE_GROUP.get(lc) != '완성품' or lc not in wb_ro.sheetnames:
                continue
            ws_ro = wb_ro[lc]
            for row in ws_ro.iter_rows(min_row=3, values_only=True):
                pn = row[0]
                if not pn or str(pn).startswith('※') or pn == '합계':
                    continue
                if (lc, pn) in existing_keys or (lc, pn) in cache_item_keys:
                    continue
                usage = _num(row[4], 1)
                price = _num(row[6])
                g_src = gerp_qty_amt[(lc, str(pn).strip())]
                e_src = olderp_qty[str(pn).strip()]
                g_day_qty = g_src['day_qty'] * usage
                g_ngt_qty = g_src['night_qty'] * usage
                g_day_amt = price * g_day_qty
                g_ngt_amt = g_src['night_amt']
                e_day_qty = e_src['day_qty'] * usage
                e_ngt_qty = e_src['night_qty'] * usage
                g_qty = g_day_qty + g_ngt_qty
                e_qty = e_day_qty + e_ngt_qty
                if g_qty != 0 or e_qty <= 0:
                    continue
                e_day_amt = price * (e_day_qty + e_ngt_qty)
                e_ngt_amt = g_ngt_amt
                g_amt = g_day_amt + g_ngt_amt
                e_amt = e_day_amt + e_ngt_amt
                diff = g_amt - e_amt
                qty_only_zero_amt = (g_amt == 0 and e_amt == 0)
                errors.append({
                    'line': lc, 'part_no': pn, 'assy_part': row[3] or '',
                    'usage': row[4] or 1, 'price_type': row[5] or '',
                    'price': price, 'vtype': row[7] or '',
                    'gerp_day_qty': g_day_qty, 'gerp_day_amt': g_day_amt,
                    'gerp_ngt_qty': g_ngt_qty, 'gerp_ngt_amt': g_ngt_amt,
                    'erp_day_qty': e_day_qty, 'erp_day_amt': e_day_amt,
                    'erp_ngt_qty': e_ngt_qty, 'erp_ngt_amt': e_ngt_amt,
                    'diff': diff, 'err_type': 'GERP 품번누락',
                    'note': MISSING_PRICE_NOTE if qty_only_zero_amt else 'GERP 신규등록필요',
                    'recv_amt': 0 if qty_only_zero_amt else abs(diff),
                    'sup_text': '',
                })
                existing_keys.add((lc, pn))
                added += 1
    finally:
        wb_ro.close()
    return added


supplemented = _append_line_sheet_gerp_missing()
if supplemented:
    print(f'SD9A01 라인시트 보강: GERP 품번누락 {supplemented}건')

type_order_idx = {t: i for i, t in enumerate(TYPE_ORDER)}
errors.sort(key=lambda x: (LINE_ORDER.index(x['line']),
                           type_order_idx.get(x['err_type'], 99),
                           -abs(x['diff'])))
print(f'오류 {len(errors)}건 / 제외 {len(excluded)}건')

# ── 본체 오류리스트 시트 박기 ──
wb = load_workbook(output, data_only=False, read_only=False)
if '오류리스트' not in wb.sheetnames:
    print('[ERROR] 본체에 오류리스트 시트 없음. build_formula_version.py 먼저 실행.')
    sys.exit(1)

# 가드: 라인시트 S컬럼(오류분류 수식) 보존 검증
# 2026-05-19 회귀 — populate 후 라인시트 S컬럼 0/(전체) 사라짐 사고로 도입
_line_sheets_for_guard = [s for s in LINE_ORDER if s in wb.sheetnames]
_s_before = {}
for s in _line_sheets_for_guard:
    _ws = wb[s]
    _s_before[s] = sum(1 for r in range(3, _ws.max_row+1) if _ws.cell(r, 19).value not in (None, ''))
print(f'[GUARD] 라인시트 S컬럼 load 직후: {_s_before}')

ws = wb['오류리스트']

# row 5 이하 기존 데이터 클리어
for r in range(5, ws.max_row + 1):
    for c in range(1, 23):
        ws.cell(r, c).value = None
        ws.cell(r, c).fill = PatternFill(fill_type=None)

# 스타일
THIN = Side(style='thin', color='D0D0D0')
BDR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CA = Alignment(horizontal='center', vertical='center')
RA = Alignment(horizontal='right', vertical='center')
LA = Alignment(horizontal='left', vertical='center')
NUM = '#,##0'
DATA = Font(name='맑은 고딕', size=9)
NEG_FILL = PatternFill('solid', fgColor='FDE8E8')
POS_FILL = PatternFill('solid', fgColor='E8F5E8')
TYPE_FILLS = {t: PatternFill('solid', fgColor=c) for t, c in TYPE_COLORS.items()}

# 데이터 행
row = 5
for e in errors:
    vals = [
        (e['part_no'], LA, None), (VENDOR_CODE, CA, None),
        (e['line'], CA, None), (e['assy_part'], LA, None),
        (e['usage'], CA, None), (e['price_type'], CA, None),
        (e['price'], RA, NUM), (e['vtype'], CA, None),
        (e['gerp_day_qty'], RA, NUM), (e['gerp_day_amt'], RA, NUM),
        (e['gerp_ngt_qty'], RA, NUM), (e['gerp_ngt_amt'], RA, NUM),
        (e['erp_day_qty'], RA, NUM), (e['erp_day_amt'], RA, NUM),
        (e['erp_ngt_qty'], RA, NUM), (e['erp_ngt_amt'], RA, NUM),
        (e['diff'], RA, NUM), (e['err_type'], CA, None),
        (e['note'], LA, None), ('', CA, None),
        (e['recv_amt'], RA, NUM), (e['sup_text'], LA, None),
    ]
    for ci, (v, align, fmt) in enumerate(vals, 1):
        c = ws.cell(row, ci, v)
        c.font = DATA; c.alignment = align; c.border = BDR
        if fmt:
            c.number_format = fmt
    tf = TYPE_FILLS.get(e['err_type'])
    if tf:
        ws.cell(row, 18).fill = tf
    if e['diff'] < 0:
        ws.cell(row, 17).fill = NEG_FILL
    elif e['diff'] > 0:
        ws.cell(row, 17).fill = POS_FILL
    row += 1

# 요약 텍스트 (row 2)
tcnt = Counter(e['err_type'] for e in errors)
tamt = defaultdict(int)
for e in errors:
    tamt[e['err_type']] += e['diff']
recv_total = sum(e['recv_amt'] for e in errors)
parts = [f"총 {len(errors)}건"]
for t in TYPE_ORDER:
    if t in tcnt:
        sign = '+' if tamt[t] >= 0 else '-'
        parts.append(f"{t} {tcnt[t]}건({sign}{abs(tamt[t]):,}원)")
parts.append(f"받을금액 {recv_total:,}원")
ws.cell(2, 1).value = '  |  '.join(parts)

# ── 유형별요약 시트 박기 ──
if '유형별요약' in wb.sheetnames:
    ws2 = wb['유형별요약']
    # row 4 이하 클리어
    for r in range(4, ws2.max_row + 1):
        for c in range(1, 7):
            ws2.cell(r, c).value = None
            ws2.cell(r, c).fill = PatternFill(fill_type=None)
    BOLD = Font(name='맑은 고딕', size=9, bold=True)
    SUM_FILL = PatternFill('solid', fgColor='E8E8E8')
    r2 = 4
    for t in TYPE_ORDER:
        items = [e for e in errors if e['err_type'] == t]
        if not items:
            continue
        g_sum = sum(e['gerp_day_amt']+e['gerp_ngt_amt'] for e in items)
        e_sum = sum(e['erp_day_amt']+e['erp_ngt_amt'] for e in items)
        d = g_sum - e_sum
        recv = sum(e['recv_amt'] for e in items)
        for ci, v in enumerate([t, len(items), g_sum, e_sum, d, recv], 1):
            c = ws2.cell(r2, ci, v)
            c.font = DATA; c.border = BDR
            c.alignment = CA if ci <= 2 else RA
            if ci >= 3:
                c.number_format = NUM
        tf = TYPE_FILLS.get(t)
        if tf:
            ws2.cell(r2, 1).fill = tf
        if d < 0:
            ws2.cell(r2, 5).fill = NEG_FILL
        elif d > 0:
            ws2.cell(r2, 5).fill = POS_FILL
        if recv > 0:
            ws2.cell(r2, 6).fill = POS_FILL
        r2 += 1
    # 합계
    for ci in range(1, 7):
        ws2.cell(r2, ci).fill = SUM_FILL; ws2.cell(r2, ci).border = BDR; ws2.cell(r2, ci).font = BOLD
    ws2.cell(r2, 1, '합계').alignment = CA
    ws2.cell(r2, 2, len(errors)).alignment = CA
    g_all = sum(e['gerp_day_amt']+e['gerp_ngt_amt'] for e in errors)
    e_all = sum(e['erp_day_amt']+e['erp_ngt_amt'] for e in errors)
    ws2.cell(r2, 3, g_all).alignment = RA; ws2.cell(r2, 3).number_format = NUM
    ws2.cell(r2, 4, e_all).alignment = RA; ws2.cell(r2, 4).number_format = NUM
    total_diff = g_all - e_all
    ws2.cell(r2, 5, total_diff).alignment = RA; ws2.cell(r2, 5).number_format = NUM
    ws2.cell(r2, 5).fill = NEG_FILL if total_diff < 0 else POS_FILL
    ws2.cell(r2, 6, recv_total).alignment = RA; ws2.cell(r2, 6).number_format = NUM
    ws2.cell(r2, 6).fill = POS_FILL

# 가드: save 직전 라인시트 S컬럼 재카운트 — load 직후와 차이 있으면 abort
_s_after = {}
for s in _line_sheets_for_guard:
    _ws = wb[s]
    _s_after[s] = sum(1 for r in range(3, _ws.max_row+1) if _ws.cell(r, 19).value not in (None, ''))
_diffs = {s: (_s_before[s], _s_after[s]) for s in _s_before if _s_before[s] != _s_after[s]}
if _diffs:
    print(f'[GUARD FAIL] 라인시트 S컬럼 손실 감지. abort. diff: {_diffs}')
    sys.exit(2)
print(f'[GUARD OK] 라인시트 S컬럼 보존 확인 ({len(_line_sheets_for_guard)}개 시트)')

wb.save(output)
print(f'\n저장: {output}')
print(f'오류리스트: {len(errors)}행 박힘 / 유형별요약: {len(tcnt)}유형')
print(f'받을금액 합계: {recv_total:,}원')
