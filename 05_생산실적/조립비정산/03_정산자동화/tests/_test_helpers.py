#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
테스트 공통 헬퍼
- 더미 Excel 파일 생성 (GERP, 구ERP, 기준정보)
- _pipeline_config.py 임시 교체/복원 (ConfigPatch 컨텍스트 매니저)
- Step 실행 + JSON 결과 반환
"""

import json
import os
import shutil
import subprocess
import sys
import tempfile
from contextlib import contextmanager

import openpyxl

# ── 경로 상수 ──────────────────────────────────────────────────
TESTS_DIR = os.path.dirname(os.path.abspath(__file__))
PIPELINE_DIR = os.path.abspath(os.path.join(TESTS_DIR, '..'))
CONFIG_FILE = os.path.join(PIPELINE_DIR, '_pipeline_config.py')
PYTHON = r'C:\Users\User\AppData\Local\Programs\Python\Python312\python.exe'

LINE_ORDER = [
    'SD9A01', 'SP3M3', 'WAMAS01', 'WABAS01',
    'ANAAS04', 'DRAAS11', 'WASAS01', 'HCAMS02', 'HASMS02', 'ISAMS03',
]
VENDOR_CODE = '0109'


# ── config 임시 교체 ───────────────────────────────────────────
@contextmanager
def patch_config(tmp_dir: str, month: str = '01'):
    """
    _pipeline_config.py를 백업하고, tmp_dir을 가리키는 임시 버전으로 교체.
    컨텍스트 블록 종료(또는 예외) 시 원본 복원 보장.
    """
    backup = CONFIG_FILE + '.test_bak'
    shutil.copy2(CONFIG_FILE, backup)
    try:
        _write_test_config(tmp_dir, month)
        yield
    finally:
        shutil.copy2(backup, CONFIG_FILE)
        os.remove(backup)


def _write_test_config(tmp_dir: str, month: str):
    """tmp_dir을 참조하는 최소 _pipeline_config.py 작성."""
    master  = os.path.join(tmp_dir, 'master.xlsx').replace('\\', '\\\\')
    gerp    = os.path.join(tmp_dir, 'gerp.xlsx').replace('\\', '\\\\')
    olderp  = os.path.join(tmp_dir, 'olderp.xlsx').replace('\\', '\\\\')
    output  = os.path.join(tmp_dir, f'output_{month}\uc6d4.xlsx').replace('\\', '\\\\')
    cache   = os.path.join(tmp_dir, '_cache').replace('\\', '\\\\')

    content = f"""\
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# AUTO-GENERATED for test — do not edit
import os

BASE_DIR     = r'{tmp_dir}'
CACHE_DIR    = r'{cache}'

MASTER_FILE  = r'{os.path.join(tmp_dir, "master.xlsx")}'
GERP_FILE    = r'{os.path.join(tmp_dir, "gerp.xlsx")}'
OLDERP_FILE  = r'{os.path.join(tmp_dir, "olderp.xlsx")}'
OUTPUT_FILE  = r'{os.path.join(tmp_dir, f"output_{month}월.xlsx")}'

MONTH        = '{month}'
VENDOR_CODE  = '0109'
SP3M3_NIGHT_PRICE = 170

LINE_ORDER = [
    'SD9A01', 'SP3M3', 'WAMAS01', 'WABAS01',
    'ANAAS04', 'DRAAS11', 'WASAS01', 'HCAMS02', 'HASMS02', 'ISAMS03',
]

LINE_INFO = {{
    'SD9A01':  {{'name': '\uc544\uc6b0\ud130',        'type': 'OUTER', 'has_night': True}},
    'SP3M3':   {{'name': '\uba54\uc778',          'type': 'MAIN',  'has_night': True}},
    'WAMAS01': {{'name': '\uc6e8\ube59 ASSY',     'type': 'SUB',   'has_night': False}},
    'WABAS01': {{'name': '\uc6e8\ube59 \uc2a4\ud1a0\ud37c',   'type': 'SUB',   'has_night': False}},
    'ANAAS04': {{'name': '\uc575\ucee4',          'type': 'SUB',   'has_night': False}},
    'DRAAS11': {{'name': '\ub514\ub9c1',          'type': 'SUB',   'has_night': False}},
    'WASAS01': {{'name': '\uc6e8\ube59 \uc2a4\ud1a0\ud37c2',  'type': 'SUB',   'has_night': False}},
    'HCAMS02': {{'name': '\ud640\ub354 CLR ASSY', 'type': 'SUB',   'has_night': False}},
    'HASMS02': {{'name': '\ud640\ub354\uc13c\uc2a4 ASSY', 'type': 'SUB',   'has_night': False}},
    'ISAMS03': {{'name': '\uc774\ub108\uc13c\uc2a4 ASSY', 'type': 'SUB',   'has_night': False}},
}}

OLD_ERP_LINE_MAP = {{
    'TD9':    'SD9A01',
    'D9N6':   'SD9A01',
    'SP3S03': 'SP3M3',
}}

GERP_COL = {{
    'line':       2,
    'product_no': 6,
    'usage':      10,
    'shift':      13,
    'qty':        14,
    'unit_price': 15,
    'amount':     16,
    'vendor_cd':  20,
}}

OLDERP_COL = {{
    'vendor':    2,
    'part_no':   4,
    'qty':       5,
    'line_code': 7,
    'lot_no':    10,
    'unit_cost': 11,
    'amount':    12,
}}

MASTER_COL = {{
    'part_no':    0,
    'vendor_cd':  1,
    'line_code':  2,
    'assy_part':  3,
    'usage':      4,
    'price_type': 5,
    'price':      6,
    'vtype':      7,
}}

CACHE_STEP1 = os.path.join(CACHE_DIR, 'step1_validation.json')
CACHE_STEP2 = os.path.join(CACHE_DIR, 'step2_gerp.json')
CACHE_STEP3 = os.path.join(CACHE_DIR, 'step3_olderp.json')
CACHE_STEP4 = os.path.join(CACHE_DIR, 'step4_matched.json')
CACHE_STEP5 = os.path.join(CACHE_DIR, 'step5_settlement.json')

os.makedirs(CACHE_DIR, exist_ok=True)
"""
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        f.write(content)


# ── 더미 Excel 생성 ────────────────────────────────────────────
def make_gerp(path: str, rows: list = None, ncols: int = 21):
    """
    GERP 더미 파일 생성.
    rows: 각 요소가 딕셔너리 {line, product_no, usage, shift, qty, unit_price, amount, vendor_cd}
    ncols: 총 컬럼 수 (기본 21 = 최소 요구)
    """
    if rows is None:
        rows = [
            dict(line='SD9A01', product_no='TST-001', usage=1,
                 shift='정상', qty=100, unit_price=300, amount=30000, vendor_cd='0109'),
            dict(line='SD9A01', product_no='TST-001', usage=1,
                 shift='추가', qty=20, unit_price=300, amount=6000, vendor_cd='0109'),
            dict(line='SP3M3',  product_no='TST-SP3-001', usage=1,
                 shift='정상', qty=200, unit_price=120, amount=24000, vendor_cd='0109'),
        ]

    wb = openpyxl.Workbook()
    ws = wb.active

    # 헤더 2행
    for _ in range(2):
        ws.append(['HEADER'] * ncols)

    col = {
        'line': 2, 'product_no': 6, 'usage': 10,
        'shift': 13, 'qty': 14, 'unit_price': 15,
        'amount': 16, 'vendor_cd': 20,
    }

    for r in rows:
        row_data = [None] * ncols
        row_data[col['line']]       = r.get('line', '')
        row_data[col['product_no']] = r.get('product_no', '')
        row_data[col['usage']]      = r.get('usage', 1)
        row_data[col['shift']]      = r.get('shift', '정상')
        row_data[col['qty']]        = r.get('qty', 0)
        row_data[col['unit_price']] = r.get('unit_price', 0)
        row_data[col['amount']]     = r.get('amount', 0)
        row_data[col['vendor_cd']]  = r.get('vendor_cd', '0109')
        ws.append(row_data)

    wb.save(path)


def make_olderp(path: str, rows: list = None, ncols: int = 13):
    """
    구ERP 더미 파일 생성 (Sheet1).
    rows: {vendor, part_no, qty, line_code, lot_no, unit_cost, amount}
    """
    if rows is None:
        rows = [
            dict(vendor='0109', part_no='TST-001', qty=95,
                 line_code='TD9', lot_no='2026-03A', unit_cost=300, amount=28500),
            dict(vendor='0109', part_no='TST-001', qty=18,
                 line_code='TD9', lot_no='2026-03B', unit_cost=300, amount=5400),
            dict(vendor='0109', part_no='TST-SP3-001', qty=195,
                 line_code='SP3S03', lot_no='2026-03A', unit_cost=120, amount=23400),
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    for _ in range(2):
        ws.append(['HEADER'] * ncols)

    col = {
        'vendor': 2, 'part_no': 4, 'qty': 5,
        'line_code': 7, 'lot_no': 10, 'unit_cost': 11, 'amount': 12,
    }

    for r in rows:
        row_data = [None] * ncols
        row_data[col['vendor']]    = r.get('vendor', '0109')
        row_data[col['part_no']]   = r.get('part_no', '')
        row_data[col['qty']]       = r.get('qty', 0)
        row_data[col['line_code']] = r.get('line_code', '')
        row_data[col['lot_no']]    = r.get('lot_no', 'xxxA')
        row_data[col['unit_cost']] = r.get('unit_cost', 0)
        row_data[col['amount']]    = r.get('amount', 0)
        ws.append(row_data)

    wb.save(path)


def make_master(path: str, rows_per_line: dict = None, ncols: int = 8):
    """
    기준정보 더미 파일 생성 (라인별 시트 10개).
    rows_per_line: {line_code: [{part_no, vendor_cd, line_code, assy_part, usage, price_type, price, vtype}]}
    """
    DEFAULT_PARTS = {
        'SD9A01': [dict(part_no='TST-001', vendor_cd='0109', line_code='SD9A01',
                        assy_part='', usage=1, price_type='기준', price=300.0, vtype='SD9')],
        'SP3M3':  [dict(part_no='TST-SP3-001', vendor_cd='0109', line_code='SP3M3',
                        assy_part='', usage=1, price_type='기준', price=120.0, vtype='SP3')],
    }
    for lc in LINE_ORDER:
        if lc not in DEFAULT_PARTS:
            DEFAULT_PARTS[lc] = [
                dict(part_no=f'TST-{lc}-001', vendor_cd='0109', line_code=lc,
                     assy_part='', usage=1, price_type='기준', price=500.0, vtype=lc)
            ]

    if rows_per_line is None:
        rows_per_line = DEFAULT_PARTS
    else:
        # 미지정 라인은 기본값으로 채움
        for lc in LINE_ORDER:
            if lc not in rows_per_line:
                rows_per_line[lc] = DEFAULT_PARTS.get(lc, [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    col_order = ['part_no', 'vendor_cd', 'line_code', 'assy_part',
                 'usage', 'price_type', 'price', 'vtype']

    for lc in LINE_ORDER:
        ws = wb.create_sheet(lc)
        # row 0~2: 헤더 (3행)
        for _ in range(3):
            ws.append(['HEADER'] * ncols)
        for r in rows_per_line.get(lc, []):
            row_data = [r.get(c, '') for c in col_order]
            ws.append(row_data)

    wb.save(path)


# ── Step 실행 + 결과 JSON 로드 ─────────────────────────────────
def run_step(step_name: str, cache_path: str) -> dict:
    """
    파이프라인 스크립트를 subprocess로 실행하고, cache_path의 JSON을 반환.
    반환값: {'exit_code': int, 'stdout': str, 'result': dict or None}
    """
    script = os.path.join(PIPELINE_DIR, step_name)
    proc = subprocess.run(
        [PYTHON, script],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding='utf-8',
        errors='replace',
    )

    result = None
    if os.path.exists(cache_path):
        with open(cache_path, encoding='utf-8') as f:
            result = json.load(f)

    return {
        'exit_code': proc.returncode,
        'stdout': proc.stdout,
        'result': result,
    }


# ── 결과 출력 헬퍼 ─────────────────────────────────────────────
def assert_or_fail(condition: bool, message: str):
    if not condition:
        print(f"  [FAIL] {message}")
        sys.exit(1)
    else:
        print(f"  [PASS] {message}")


def section(title: str):
    print(f"\n{'='*55}\n{title}\n{'='*55}")
