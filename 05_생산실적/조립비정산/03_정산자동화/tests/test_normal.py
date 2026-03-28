#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
스모크 테스트 1: 정상 데이터로 전체 파이프라인 통과 확인

기대 결과:
  - Step 1~7 전체 exit code = 0
  - step1_validation.json → status = "OK"
  - step5_settlement.json 존재 + SD9A01 계산값 검증
  - step6_validation.json → overall = "PASS"
  - step7 출력 xlsx 파일 생성 확인
"""

import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(__file__))
from _test_helpers import (
    patch_config, make_gerp, make_olderp, make_master,
    run_step, assert_or_fail, section, PIPELINE_DIR,
)

STEP_SCRIPTS = [
    ('step1_파일검증.py',       'step1_validation.json'),
    ('step2_gerp처리.py',       'step2_gerp.json'),
    ('step3_구erp처리.py',      'step3_olderp.json'),
    ('step4_기준정보매칭.py',   'step4_matched.json'),
    ('step5_정산계산.py',       'step5_settlement.json'),
    ('step6_검증.py',           'step6_validation.json'),
    ('step7_보고서.py',         None),  # 출력은 xlsx
]

print("=" * 55)
print("test_normal: 전체 파이프라인 정상 통과 확인")
print("=" * 55)

with tempfile.TemporaryDirectory() as tmp_dir:
    cache_dir = os.path.join(tmp_dir, '_cache')
    os.makedirs(cache_dir, exist_ok=True)

    # ── 더미 파일 생성 ──────────────────────────────────────────
    section("더미 입력 파일 생성")

    gerp_path   = os.path.join(tmp_dir, 'gerp.xlsx')
    olderp_path = os.path.join(tmp_dir, 'olderp.xlsx')
    master_path = os.path.join(tmp_dir, 'master.xlsx')

    # GERP: SD9A01 주간/야간, SP3M3 주간
    make_gerp(gerp_path, rows=[
        dict(line='SD9A01', product_no='TST-001',     usage=1,
             shift='정상', qty=100, unit_price=300, amount=30000, vendor_cd='0109'),
        dict(line='SD9A01', product_no='TST-001',     usage=1,
             shift='추가', qty=20,  unit_price=300, amount=6000,  vendor_cd='0109'),
        dict(line='SD9A01', product_no='TST-002',     usage=1,
             shift='정상', qty=50,  unit_price=600, amount=30000, vendor_cd='0109'),
        dict(line='SP3M3',  product_no='TST-SP3-001', usage=1,
             shift='정상', qty=200, unit_price=120, amount=24000, vendor_cd='0109'),
        dict(line='SP3M3',  product_no='TST-SP3-001', usage=1,
             shift='추가', qty=30,  unit_price=120, amount=3600,  vendor_cd='0109'),
        # SUB 라인 (WAMAS01)
        dict(line='WAMAS01', product_no='TST-WM-001', usage=1,
             shift='정상', qty=500, unit_price=500, amount=250000, vendor_cd='0109'),
    ])
    print(f"  GERP: {gerp_path}")

    # 구ERP: 대응 데이터
    make_olderp(olderp_path, rows=[
        dict(vendor='0109', part_no='TST-001',     qty=95,  line_code='TD9',
             lot_no='2026-01A', unit_cost=300, amount=28500),
        dict(vendor='0109', part_no='TST-001',     qty=18,  line_code='TD9',
             lot_no='2026-01B', unit_cost=300, amount=5400),
        dict(vendor='0109', part_no='TST-002',     qty=48,  line_code='TD9',
             lot_no='2026-01A', unit_cost=600, amount=28800),
        dict(vendor='0109', part_no='TST-SP3-001', qty=195, line_code='SP3S03',
             lot_no='2026-01A', unit_cost=120, amount=23400),
        dict(vendor='0109', part_no='TST-SP3-001', qty=28,  line_code='SP3S03',
             lot_no='2026-01B', unit_cost=120, amount=3360),
        dict(vendor='OTHER', part_no='TST-WM-001', qty=490, line_code='WMA',
             lot_no='2026-01A', unit_cost=500, amount=245000),
    ])
    print(f"  구ERP: {olderp_path}")

    # 기준정보: 10개 라인 각 1~2 품번
    make_master(master_path, rows_per_line={
        'SD9A01': [
            dict(part_no='TST-001', vendor_cd='0109', line_code='SD9A01',
                 assy_part='', usage=1, price_type='기준', price=300.0, vtype='SD9'),
            dict(part_no='TST-002', vendor_cd='0109', line_code='SD9A01',
                 assy_part='', usage=1, price_type='기준', price=600.0, vtype='SD9'),
        ],
        'SP3M3': [
            dict(part_no='TST-SP3-001', vendor_cd='0109', line_code='SP3M3',
                 assy_part='', usage=1, price_type='기준', price=120.0, vtype='SP3'),
        ],
        'WAMAS01': [
            dict(part_no='TST-WM-001', vendor_cd='0109', line_code='WAMAS01',
                 assy_part='', usage=1, price_type='기준', price=500.0, vtype='WM'),
        ],
    })
    print(f"  기준정보: {master_path}")

    # ── 파이프라인 실행 ─────────────────────────────────────────
    with patch_config(tmp_dir, month='01'):
        section("Step 1~7 순차 실행")

        for script_name, cache_file in STEP_SCRIPTS:
            cache_path = (
                os.path.join(cache_dir, cache_file) if cache_file
                else os.path.join(tmp_dir, 'output_01월.xlsx')
            )

            print(f"\n  → {script_name}")
            ret = run_step(script_name, cache_path)
            print(ret['stdout'][-300:] if len(ret['stdout']) > 300 else ret['stdout'])

            assert_or_fail(
                ret['exit_code'] == 0,
                f"{script_name} exit code = {ret['exit_code']} (기대: 0)",
            )

        # ── 결과 검증 ───────────────────────────────────────────
        section("결과 검증")

        # 1. Step 1 status
        import json
        with open(os.path.join(cache_dir, 'step1_validation.json'), encoding='utf-8') as f:
            s1 = json.load(f)
        assert_or_fail(s1['status'] == 'OK',
                       f"Step1 status={s1['status']} (기대: OK)")

        # 2. Step 5 SD9A01 계산 검증
        with open(os.path.join(cache_dir, 'step5_settlement.json'), encoding='utf-8') as f:
            s5 = json.load(f)

        sd9 = s5['lines'].get('SD9A01', {})
        assert_or_fail('items' in sd9 and len(sd9['items']) > 0,
                       "SD9A01 items 비어있지 않음")

        # TST-001: price=300 (≤500 야간가산), day_qty=100, ngt_qty=20
        # gerp_day_amt = 100*300 = 30000, gerp_ngt_amt = round(20*300*1.3) = 7800
        tst001 = next((r for r in sd9['items'] if r['part_no'] == 'TST-001'), None)
        assert_or_fail(tst001 is not None, "TST-001 품번 존재")
        assert_or_fail(tst001['gerp_day_amt'] == 30000,
                       f"TST-001 gerp_day_amt={tst001['gerp_day_amt']} (기대: 30000)")
        assert_or_fail(tst001['gerp_ngt_amt'] == 7800,
                       f"TST-001 gerp_ngt_amt={tst001['gerp_ngt_amt']} (기대: 7800)")
        assert_or_fail(tst001['price_judgment'] == '야간가산',
                       f"TST-001 price_judgment={tst001['price_judgment']} (기대: 야간가산)")

        # TST-002: price=600 (>500 기본), day_qty=50, ngt_qty=0
        tst002 = next((r for r in sd9['items'] if r['part_no'] == 'TST-002'), None)
        assert_or_fail(tst002 is not None, "TST-002 품번 존재")
        assert_or_fail(tst002['gerp_ngt_qty'] == 0,
                       f"TST-002 gerp_ngt_qty={tst002['gerp_ngt_qty']} (기대: 0)")
        assert_or_fail(tst002['price_judgment'] is None,
                       f"TST-002 price_judgment={tst002['price_judgment']} (기대: null — 야간실적 없음)")

        # 3. SP3M3 야간 고정단가 170원 확인
        sp3 = s5['lines'].get('SP3M3', {})
        sp3_item = next((r for r in sp3['items'] if r['part_no'] == 'TST-SP3-001'), None)
        assert_or_fail(sp3_item is not None, "TST-SP3-001 품번 존재")
        expected_ngt = 30 * 170  # 5100
        assert_or_fail(sp3_item['gerp_ngt_amt'] == expected_ngt,
                       f"SP3M3 야간금액={sp3_item['gerp_ngt_amt']} (기대: {expected_ngt})")

        # 4. Step 6 overall PASS
        with open(os.path.join(cache_dir, 'step6_validation.json'), encoding='utf-8') as f:
            s6 = json.load(f)
        assert_or_fail(s6['overall'] == 'PASS',
                       f"Step6 overall={s6['overall']} (기대: PASS)")
        assert_or_fail(s6['fail'] == 0,
                       f"Step6 fail 건수={s6['fail']} (기대: 0)")

        # 5. Step 7 xlsx 생성
        xlsx_path = os.path.join(tmp_dir, 'output_01월.xlsx')
        assert_or_fail(os.path.exists(xlsx_path),
                       f"Step7 xlsx 생성 확인: {xlsx_path}")

        import openpyxl as _xl
        wb = _xl.load_workbook(xlsx_path, read_only=True)
        assert_or_fail('00_정산집계' in wb.sheetnames,
                       "00_정산집계 시트 존재")
        assert_or_fail('SD9A01' in wb.sheetnames,
                       "SD9A01 시트 존재")
        wb.close()

print("\n" + "=" * 55)
print("test_normal: ALL PASS")
print("=" * 55)
