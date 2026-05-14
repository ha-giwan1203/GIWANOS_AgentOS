#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
기존 기준정보 vs GERP 역설계 기준정보 대조 검증

비교 항목:
  1. 라인별 행 수
  2. 품번 교차 (공통/기존만/신규만)
  3. 공통 품번 단가·조립품번·Usage 일치율
  4. 다중단가 건수 비교
"""

import os
import sys
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _pipeline_config import (
    MASTER_FILE, MASTER_COL, LINE_ORDER, LINE_INFO, BASE_DIR
)

# GERP 역설계 파일 (최신)
GERP_MASTER_DIR = os.path.join(BASE_DIR, '01_기준정보')


def find_latest_gerp_master():
    """GERP 역설계 파일 중 최신 찾기"""
    import glob
    pattern = os.path.join(GERP_MASTER_DIR, '기준정보_GERP역설계_*.xlsx')
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"GERP 역설계 파일 없음: {pattern}")
    return max(files)


def load_master(filepath):
    """기준정보 엑셀 로드 → {line: [rows]}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 라인코드 결정 (시트명 또는 Row1에서)
        line_code = sheet_name
        if sheet_name == '이너센스ASSY':
            line_code = 'ISAMS03'

        # 데이터: Row4부터 (Row1=라인명, Row2=빈행, Row3=헤더)
        rows = []
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r[MASTER_COL['part_no']] is None:
                continue
            rows.append({
                'part_no': str(r[MASTER_COL['part_no']]).strip(),
                'vendor_cd': str(r[MASTER_COL['vendor_cd']]).strip() if r[MASTER_COL['vendor_cd']] else '',
                'line_code': str(r[MASTER_COL['line_code']]).strip() if r[MASTER_COL['line_code']] else line_code,
                'assy_part': str(r[MASTER_COL['assy_part']]).strip() if r[MASTER_COL['assy_part']] else '',
                'usage': r[MASTER_COL['usage']],
                'price_type': str(r[MASTER_COL['price_type']]).strip() if r[MASTER_COL['price_type']] else '',
                'price': r[MASTER_COL['price']] if r[MASTER_COL['price']] else 0,
                'vtype': str(r[MASTER_COL['vtype']]).strip() if r[MASTER_COL['vtype']] else '',
            })

        if rows:
            result[line_code] = rows

    wb.close()
    return result


def compare(old_data, new_data):
    """기존 vs 신규 비교"""
    print("=" * 70)
    print(f"{'라인':<12} {'기존':>6} {'신규':>6} {'공통':>6} {'기존만':>6} {'신규만':>6} {'차이':>6}")
    print("-" * 70)

    total_old = total_new = total_common = total_old_only = total_new_only = 0
    all_diffs = []

    for line in LINE_ORDER:
        old_rows = old_data.get(line, [])
        new_rows = new_data.get(line, [])

        # 품번 기준 비교
        old_parts = set(r['part_no'] for r in old_rows)
        new_parts = set(r['part_no'] for r in new_rows)

        common = old_parts & new_parts
        old_only = old_parts - new_parts
        new_only = new_parts - old_parts

        print(f"{line:<12} {len(old_rows):>6} {len(new_rows):>6} {len(common):>6} {len(old_only):>6} {len(new_only):>6} {len(new_rows)-len(old_rows):>+6}")

        total_old += len(old_rows)
        total_new += len(new_rows)
        total_common += len(common)
        total_old_only += len(old_only)
        total_new_only += len(new_only)

        # 공통 품번 상세 비교
        old_by_part = defaultdict(list)
        new_by_part = defaultdict(list)
        for r in old_rows:
            old_by_part[r['part_no']].append(r)
        for r in new_rows:
            new_by_part[r['part_no']].append(r)

        for part in common:
            old_prices = sorted(set(float(r['price']) for r in old_by_part[part]))
            new_prices = sorted(set(float(r['price']) for r in new_by_part[part]))

            if old_prices != new_prices:
                all_diffs.append({
                    'line': line,
                    'part': part,
                    'type': '단가차이',
                    'old': old_prices,
                    'new': new_prices,
                })

            old_usage = set(r['usage'] for r in old_by_part[part])
            new_usage = set(r['usage'] for r in new_by_part[part])
            if old_usage != new_usage:
                all_diffs.append({
                    'line': line,
                    'part': part,
                    'type': 'Usage차이',
                    'old': old_usage,
                    'new': new_usage,
                })

    print("-" * 70)
    print(f"{'합계':<12} {total_old:>6} {total_new:>6} {total_common:>6} {total_old_only:>6} {total_new_only:>6} {total_new-total_old:>+6}")

    # 다중단가 비교
    print("\n=== 다중단가 비교 ===")
    for label, data in [("기존", old_data), ("신규", new_data)]:
        multi = 0
        for line, rows in data.items():
            parts = defaultdict(set)
            for r in rows:
                parts[r['part_no']].add(float(r['price']))
            multi += sum(1 for prices in parts.values() if len(prices) > 1)
        print(f"  {label}: {multi}건")

    # 차이 상세
    if all_diffs:
        print(f"\n=== 공통 품번 차이 상세 ({len(all_diffs)}건, 최대 20건 표시) ===")
        for d in all_diffs[:20]:
            print(f"  {d['line']}/{d['part']}: {d['type']} 기존={d['old']} → 신규={d['new']}")

    # 기존만 있는 품번 (누락 위험)
    print(f"\n=== 기존에만 있는 품번 (신규 누락, 최대 20건) ===")
    count = 0
    for line in LINE_ORDER:
        old_parts = set(r['part_no'] for r in old_data.get(line, []))
        new_parts = set(r['part_no'] for r in new_data.get(line, []))
        old_only = sorted(old_parts - new_parts)
        for part in old_only[:5]:
            # 해당 품번의 기존 단가
            prices = [r['price'] for r in old_data[line] if r['part_no'] == part]
            print(f"  {line}/{part}: 기존단가={prices}")
            count += 1
            if count >= 20:
                break
        if count >= 20:
            break

    # 신규만 있는 품번 (추가)
    print(f"\n=== 신규에만 있는 품번 (기존 미포함, 최대 20건) ===")
    count = 0
    for line in LINE_ORDER:
        old_parts = set(r['part_no'] for r in old_data.get(line, []))
        new_parts = set(r['part_no'] for r in new_data.get(line, []))
        new_only = sorted(new_parts - old_parts)
        for part in new_only[:5]:
            prices = [r['price'] for r in new_data[line] if r['part_no'] == part]
            print(f"  {line}/{part}: 신규단가={prices}")
            count += 1
            if count >= 20:
                break
        if count >= 20:
            break

    return all_diffs


def main():
    print("=== 기준정보 대조 검증 ===\n")

    gerp_master = find_latest_gerp_master()
    print(f"기존: {MASTER_FILE}")
    print(f"신규: {gerp_master}")
    print()

    old_data = load_master(MASTER_FILE)
    new_data = load_master(gerp_master)

    print(f"기존 시트: {list(old_data.keys())}")
    print(f"신규 시트: {list(new_data.keys())}")
    print()

    diffs = compare(old_data, new_data)

    print(f"\n=== 검증 완료 ===")


if __name__ == "__main__":
    main()
