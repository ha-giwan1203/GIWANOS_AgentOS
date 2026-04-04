#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Step 7 — 보고서 생성
00_정산집계 + 라인별 시트 10개 + 01_차이분석 → 최종 xlsx

실행: python step7_보고서.py
입력: _cache/step5_settlement.json
출력: 05_생산실적/조립비정산/정산결과_{월}월.xlsx  (OUTPUT_FILE)
"""

import sys, os, json
sys.path.insert(0, os.path.dirname(__file__))
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from _pipeline_config import *
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 60)
print("Step 7: 보고서 생성")
print("=" * 60)

if not os.path.exists(CACHE_STEP5):
    print(f"[ERROR] Step5 결과 없음. step5_정산계산.py 먼저 실행하세요.")
    sys.exit(1)

with open(CACHE_STEP5, encoding='utf-8') as f:
    s5 = json.load(f)

lines    = s5['lines']
summary  = s5['summary']
unmapped = []  # 미매핑 품번은 GERP 단가 fallback 적용됨 → 별도 시트 불필요

# step6 검증 결과 로드 (없으면 None)
_step6_path = os.path.join(CACHE_DIR, 'step6_validation.json')
s6 = None
if os.path.exists(_step6_path):
    with open(_step6_path, encoding='utf-8') as f:
        s6 = json.load(f)
    print(f"[검증] step6 결과 로드: overall={s6.get('overall')} "
          f"(CRITICAL_FAIL={s6.get('critical_fail',0)}, WARNING={s6.get('warning',0)})")
else:
    print("[검증] step6_validation.json 없음 — 03_검증결과 시트 생략")

# ── 스타일 정의 ───────────────────────────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1B2A4A')
HDR_FONT  = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=9)
DATA_FONT = Font(name='맑은 고딕', size=9)
BOLD_FONT = Font(name='맑은 고딕', bold=True, size=9)
NEG_FILL  = PatternFill('solid', fgColor='FCE4EC')
POS_FILL  = PatternFill('solid', fgColor='E8F5E9')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
GRAY_FILL = PatternFill('solid', fgColor='F5F5F5')
THIN      = Side(style='thin', color='CCCCCC')
BDR       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT   = '#,##0'

C = Alignment(horizontal='center', vertical='center')
L = Alignment(horizontal='left',   vertical='center')
R = Alignment(horizontal='right',  vertical='center')


def hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = C
        cell.border = BDR


def dc(ws, row, col, value, bold=False, fill=None, fmt=None, align=None):
    cell = ws.cell(row, col, value)
    cell.font = BOLD_FONT if bold else DATA_FONT
    cell.border = BDR
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    if align: cell.alignment = align
    return cell


def auto_w(ws, mn=8, mx=35):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0) + 3
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w, mn), mx)


def _load_gerp_price_lookup():
    """GERP 원본에서 (라인코드, 품번, 단가) → GERP단가 set + (라인, 품번) → [단가들] 반환.
    동일품번 다중단가 지원."""
    if not os.path.exists(GERP_FILE):
        print("[단가] GERP 원본 없음 — GERP단가 공백 처리")
        return {}, {}

    print(f"[단가] GERP 원본에서 단가 추출: {os.path.basename(GERP_FILE)}")
    df_raw = pd.read_excel(GERP_FILE, sheet_name=0, header=None)
    data = df_raw.iloc[2:].reset_index(drop=True)
    c = GERP_COL
    df = pd.DataFrame({
        'line':       data.iloc[:, c['line']].astype(str).str.strip(),
        'product_no': data.iloc[:, c['product_no']].astype(str).str.strip(),
        'unit_price': pd.to_numeric(data.iloc[:, c['unit_price']], errors='coerce').fillna(0),
        'vendor_cd':  data.iloc[:, c['vendor_cd']].astype(str).str.strip(),
    })
    df = df[df['vendor_cd'] == VENDOR_CODE]

    # (라인, 품번) → 단가 set
    from collections import defaultdict
    price_set = defaultdict(set)
    for _, row in df.iterrows():
        if row['unit_price'] > 0:
            price_set[(row['line'], row['product_no'])].add(row['unit_price'])

    print(f"  추출된 (라인,품번) 단가 항목: {len(price_set)}개")
    return dict(price_set)


gerp_price_set = _load_gerp_price_lookup()

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════
# 시트 1: 00_정산집계
# ══════════════════════════════════════════════════════════════
print("\n[1/4] 00_정산집계 시트...")
ws = wb.create_sheet('00_정산집계')

# 스타일
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=13, color='1B2A4A')
SUB_FONT   = Font(name='맑은 고딕', size=9, color='666666')
GRP_FILL   = PatternFill('solid', fgColor='E3E8F0')
GRP_FONT   = Font(name='맑은 고딕', bold=True, size=9, color='1B2A4A')
MATCH_FILL = PatternFill('solid', fgColor='E8F5E9')
DIFF_FILL  = PatternFill('solid', fgColor='FFF2CC')
TOTAL_FILL = PatternFill('solid', fgColor='1B2A4A')
TOTAL_FONT = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
THICK_BOTTOM = Border(left=THIN, right=THIN, top=THIN,
                      bottom=Side(style='medium', color='1B2A4A'))

# 제목 영역
ws.merge_cells('A1:H1')
ws['A1'] = f'{MONTH}월 조립비 정산 집계'
ws['A1'].font = TITLE_FONT
ws['A2'] = '대원테크 (0109)'
ws['A2'].font = SUB_FONT
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 16

# 그룹 헤더 (row 4)
ws.merge_cells('A4:C4')
ws['A4'] = '라인 정보'
ws.merge_cells('D4:E4')
ws['D4'] = '주간'
ws.merge_cells('F4:G4')
ws['F4'] = '야간'
ws.merge_cells('H4:I4')
ws['H4'] = '정산'
ws.merge_cells('J4:K4')
ws['J4'] = '검증'
for c in range(1, 12):
    cell = ws.cell(4, c)
    cell.fill = GRP_FILL
    cell.font = GRP_FONT
    cell.alignment = C
    cell.border = BDR

# 상세 헤더 (row 5)
headers = ['라인코드', '라인명', '유형',
           '주간수량', '주간금액', '야간수량', '야간금액',
           '정산합계', 'GERP원본금액', '차이(정산-원본)', '비고']
for c, h in enumerate(headers, 1):
    ws.cell(5, c, h)
hdr(ws, 5, len(headers))

# GERP 원본 금액 계산 함수
def _gerp_orig_total(ld):
    items = ld.get('items', [])
    return sum(it.get('gerp_orig_day_amt', 0) + it.get('gerp_orig_ngt_amt', 0) for it in items)

def _adjusted_day_night(lc, ld):
    """야간 라인: 정상(주간) 수량/금액에서 야간분을 분리.
    GERP '정상' 행은 전체(주간+야간) 포함, '추가' 행은 야간분.
    → 주간금액 = 정상금액 - 야간기본금액
    → 야간금액 = 야간기본금액 + 야간가산금액(30%)
    """
    has_night = LINE_INFO[lc]['has_night']
    raw_day_amt = ld.get('total_gerp_day_amt', 0)
    raw_ngt_amt = ld.get('total_gerp_ngt_amt', 0)
    raw_day_qty = ld.get('total_gerp_day_qty', 0)
    raw_ngt_qty = ld.get('total_gerp_ngt_qty', 0)

    if not has_night or raw_ngt_qty == 0:
        return raw_day_qty, raw_ngt_qty, raw_day_amt, raw_ngt_amt

    # 야간 기본금액 = 품번별 야간수량 × 기준단가 합산
    ngt_base_amt = 0
    for it in ld.get('items', []):
        ngt_base_amt += round(it.get('gerp_ngt_qty', 0) * it.get('price', 0))

    adj_day_qty = raw_day_qty - raw_ngt_qty
    adj_ngt_qty = raw_ngt_qty
    adj_day_amt = raw_day_amt - ngt_base_amt
    adj_ngt_amt = ngt_base_amt + raw_ngt_amt   # 기본(100%) + 가산(30%)

    return adj_day_qty, adj_ngt_qty, adj_day_amt, adj_ngt_amt

# 데이터 행 (row 6~)
r = 6
for sr in summary:
    lc = sr['line']
    ld = lines.get(lc, {})
    gerp_amt = ld.get('total_gerp_amt', 0)
    orig_amt = _gerp_orig_total(ld)
    price_diff = gerp_amt - orig_amt

    adj_day_qty, adj_ngt_qty, adj_day_amt, adj_ngt_amt = _adjusted_day_night(lc, ld)

    # 비고
    note = ''
    if price_diff > 0:
        note = '단가차이'
    elif price_diff < 0:
        note = '단가차이'

    vals = [
        lc,
        LINE_INFO[lc]['name'],
        LINE_INFO[lc]['type'],
        adj_day_qty,
        adj_day_amt,
        adj_ngt_qty,
        adj_ngt_amt,
        gerp_amt,
        orig_amt,
        price_diff,
        note,
    ]

    row_fill = None
    if price_diff == 0:
        row_fill = MATCH_FILL
    elif abs(price_diff) > 0:
        row_fill = DIFF_FILL

    for c, v in enumerate(vals, 1):
        # 수량/금액 0값 → "-" 변환 (라인코드/라인명/유형/비고 제외)
        if c >= 4 and c <= 10 and isinstance(v, (int, float)) and v == 0:
            v = '-'
        fmt = NUM_FMT if (isinstance(v, (int, float)) and c in (4, 5, 6, 7, 8, 9, 10)) else None
        fill = row_fill if c >= 4 and c <= 10 else None
        al = C if c <= 3 or c == 11 else R
        dc(ws, r, c, v, fmt=fmt, fill=fill, align=al)
    r += 1

# 합계행
NCOLS = 11
for c in range(1, NCOLS + 1):
    cell = ws.cell(r, c)
    cell.fill = TOTAL_FILL
    cell.font = TOTAL_FONT
    cell.border = BDR
    cell.alignment = R

ws.cell(r, 1, '합계').alignment = C
ws.cell(r, 2, '').alignment = C
ws.cell(r, 3, '10개 라인').alignment = C
ws.cell(r, 11, '').alignment = C

grand_day_qty = sum(_adjusted_day_night(lc, lines[lc])[0] for lc in lines)
grand_ngt_qty = sum(_adjusted_day_night(lc, lines[lc])[1] for lc in lines)
grand_day_amt = sum(_adjusted_day_night(lc, lines[lc])[2] for lc in lines)
grand_ngt_amt = sum(_adjusted_day_night(lc, lines[lc])[3] for lc in lines)
grand_gerp = s5['grand_gerp_amt']
grand_orig = sum(_gerp_orig_total(lines[lc]) for lc in lines)
grand_pdiff = grand_gerp - grand_orig

for c, v in [(4, grand_day_qty), (5, grand_day_amt),
             (6, grand_ngt_qty), (7, grand_ngt_amt),
             (8, grand_gerp), (9, grand_orig), (10, grand_pdiff)]:
    ws.cell(r, c, v).number_format = NUM_FMT

# 구분선
r += 2
ws.cell(r, 1, 'GERP vs 구ERP 비교 (참고)').font = GRP_FONT
r += 1

# 구ERP 비교 상세 표 — 주야 분리
# 그룹 헤더
ERP_NCOLS = 14
ws.merge_cells(f'A{r}:B{r}')
ws[f'A{r}'] = '라인 정보'
ws.merge_cells(f'C{r}:G{r}')
ws[f'C{r}'] = 'GERP 정산'
ws.merge_cells(f'H{r}:L{r}')
ws[f'H{r}'] = '구ERP (참고)'
ws.merge_cells(f'M{r}:N{r}')
ws[f'M{r}'] = '비교'
for col in range(1, ERP_NCOLS + 1):
    cell = ws.cell(r, col)
    cell.fill = GRP_FILL
    cell.font = GRP_FONT
    cell.alignment = C
    cell.border = BDR
r += 1

# 상세 헤더
erp_headers = [
    '라인코드', '라인명',
    'GERP\n주간qty', 'GERP\n주간금액', 'GERP\n야간qty', 'GERP\n야간금액', 'GERP\n합계',
    '구ERP\n주간qty', '구ERP\n주간금액', '구ERP\n야간qty', '구ERP\n야간금액', '구ERP\n합계',
    '차이\n(GERP-구ERP)', '비고'
]
for c, h in enumerate(erp_headers, 1):
    ws.cell(r, c, h)
hdr(ws, r, len(erp_headers))
r += 1

for sr in summary:
    lc = sr['line']
    ld = lines.get(lc, {})
    has_night = LINE_INFO[lc]['has_night']
    diff = sr['diff_amt']
    diff_fill = NEG_FILL if diff < 0 else (POS_FILL if diff > 0 else None)

    # GERP 주야 (집계 시트 상단과 동일한 adj 값)
    adj_day_qty, adj_ngt_qty, adj_day_amt, adj_ngt_amt = _adjusted_day_night(lc, ld)
    gerp_total = ld.get('total_gerp_amt', 0)

    # 구ERP 주야
    e_day_qty = ld.get('total_erp_day_qty', 0)
    e_day_amt = ld.get('total_erp_day_amt', 0)
    e_ngt_qty = ld.get('total_erp_ngt_qty', 0) if has_night else 0
    e_ngt_amt = ld.get('total_erp_ngt_amt', 0) if has_night else 0
    erp_total = ld.get('total_erp_amt', 0)

    vals = [
        lc, LINE_INFO[lc]['name'],
        adj_day_qty, adj_day_amt, adj_ngt_qty if has_night else '-', adj_ngt_amt if has_night else '-', gerp_total,
        e_day_qty, e_day_amt, e_ngt_qty if has_night else '-', e_ngt_amt if has_night else '-', erp_total,
        diff, ''
    ]

    for col, v in enumerate(vals, 1):
        if isinstance(v, (int, float)) and v == 0 and col not in (1, 2, 14):
            v = '-'
        fmt = NUM_FMT if isinstance(v, (int, float)) and col not in (1, 2, 14) else None
        fill = diff_fill if col == 13 else None
        al = C if col <= 2 else R
        dc(ws, r, col, v, fmt=fmt, fill=fill, align=al)
    r += 1

# 구ERP 합계행
for c in range(1, ERP_NCOLS + 1):
    ws.cell(r, c).fill = GRAY_FILL
    ws.cell(r, c).font = BOLD_FONT
    ws.cell(r, c).border = BDR

dc(ws, r, 1, '합계', bold=True, fill=GRAY_FILL, align=C)
dc(ws, r, 2, '', fill=GRAY_FILL)
# GERP 합계
g_sum_dq = sum(_adjusted_day_night(lc, lines[lc])[0] for lc in lines)
g_sum_da = sum(_adjusted_day_night(lc, lines[lc])[2] for lc in lines)
g_sum_nq = sum(_adjusted_day_night(lc, lines[lc])[1] for lc in lines)
g_sum_na = sum(_adjusted_day_night(lc, lines[lc])[3] for lc in lines)
dc(ws, r, 3,  g_sum_dq, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 4,  g_sum_da, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 5,  g_sum_nq, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 6,  g_sum_na, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 7,  s5['grand_gerp_amt'], bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
# 구ERP 합계
e_sum_dq = sum(lines[lc].get('total_erp_day_qty', 0) for lc in lines)
e_sum_da = sum(lines[lc].get('total_erp_day_amt', 0) for lc in lines)
e_sum_nq = sum(lines[lc].get('total_erp_ngt_qty', 0) for lc in lines)
e_sum_na = sum(lines[lc].get('total_erp_ngt_amt', 0) for lc in lines)
dc(ws, r, 8,  e_sum_dq, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 9,  e_sum_da, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 10, e_sum_nq, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 11, e_sum_na, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 12, s5['grand_erp_amt'], bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
dc(ws, r, 13, s5['grand_diff_amt'], bold=True, fmt=NUM_FMT, fill=GRAY_FILL)

# 열 너비 수동 설정 (14컬럼 대응)
col_widths = {'A': 12, 'B': 14, 'C': 10, 'D': 14, 'E': 10, 'F': 14, 'G': 16,
              'H': 10, 'I': 14, 'J': 10, 'K': 14, 'L': 16, 'M': 18, 'N': 8}
for col_letter, w in col_widths.items():
    ws.column_dimensions[col_letter].width = w
ws.freeze_panes = 'A6'

# 제목 병합 범위 확장
ws.unmerge_cells('A1:H1')
ws.merge_cells(f'A1:N1')

# ══════════════════════════════════════════════════════════════
# 시트 2~11: 라인별 상세 시트
# ══════════════════════════════════════════════════════════════
print("[2/4] 라인별 상세 시트 10개...")

WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)

for lc in LINE_ORDER:
    ld  = lines.get(lc, {})
    li  = LINE_INFO[lc]
    ws2 = wb.create_sheet(lc)
    has_night = li['has_night']
    items_list = ld.get('items', [])

    # ── 컬럼 정의 ──
    if has_night:
        if lc == 'SD9A01':
            grp_hdrs = [('기본정보', 6), ('GERP 정산', 5), ('구ERP 비교', 4), ('지원분', 3), ('결과', 3)]
            col_hdrs = ['No', '품번', '조립\n품번', 'Usage', '기준\n단가', 'GERP\n단가',
                        '주간\n수량', '주간\n금액', '야간\n수량', '야간\n금액', '합계',
                        '구ERP\n주간수량', '구ERP\n주간금액', '구ERP\n야간수량', '구ERP\n야간금액',
                        '지원\n업체', '지원\n라인', '지원\n수량',
                        '수량차', '금액차', '유형']
        else:  # SP3M3
            grp_hdrs = [('기본정보', 6), ('GERP 정산', 4), ('구ERP 비교', 4), ('결과', 3)]
            col_hdrs = ['No', '품번', '조립\n품번', 'Usage', '기준\n단가', 'GERP\n단가',
                        '주간\n수량', '주간\n금액', '야간\n수량', '야간금액\n(170원)',
                        '구ERP\n주간수량', '구ERP\n주간금액', '구ERP\n야간수량', '구ERP\n야간금액',
                        '수량차', '금액차', '유형']
    else:
        grp_hdrs = [('기본정보', 6), ('GERP 정산', 2), ('구ERP 비교', 2), ('결과', 3)]
        col_hdrs = ['No', '품번', '조립\n품번', 'Usage', '기준\n단가', 'GERP\n단가',
                    'GERP\n수량', 'GERP\n금액',
                    '구ERP\n수량', '구ERP\n금액',
                    '수량차', '금액차', '유형']

    ncols = len(col_hdrs)
    item_count = len(items_list)
    active_count = sum(1 for it in items_list if it['gerp_total_amt'] != 0)

    # ── row 1: 타이틀 (병합) ──
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws2['A1'] = f"{lc} ({li['name']}) — {MONTH}월 정산"
    ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=11, color='1B2A4A')
    ws2.row_dimensions[1].height = 24

    # ── row 2: 메타 정보 (병합) ──
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    inactive_count = item_count - active_count
    meta = f"유형: {li['type']}  |  야간: {'O' if has_night else 'X'}  |  품번: {item_count}건 (실적: {active_count}건, 숨김: {inactive_count}건)  |  합계: {ld.get('total_gerp_amt', 0):,}원"
    ws2['A2'] = meta
    ws2['A2'].font = Font(name='맑은 고딕', size=9, color='666666')

    # ── row 3: 그룹 헤더 (병합) ──
    gc = 1
    for grp_name, span in grp_hdrs:
        ws2.merge_cells(start_row=3, start_column=gc, end_row=3, end_column=gc + span - 1)
        cell = ws2.cell(3, gc, grp_name)
        cell.fill = GRP_FILL
        cell.font = GRP_FONT
        cell.alignment = C
        cell.border = BDR
        for sc in range(gc, gc + span):
            ws2.cell(3, sc).fill = GRP_FILL
            ws2.cell(3, sc).border = BDR
        gc += span

    # ── row 4: 상세 헤더 (줄바꿈) ──
    ws2.row_dimensions[4].height = 32
    for c, h in enumerate(col_hdrs, 1):
        cell = ws2.cell(4, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = WRAP
        cell.border = BDR

    # ── 다중단가 품번 사전 집계 ──
    from collections import Counter
    pn_counter = Counter(it['part_no'] for it in items_list)
    multi_price_pns = {pn for pn, cnt in pn_counter.items() if cnt > 1}

    # ── row 5+: 데이터 ──
    r = 5
    hidden_count = 0
    for i, item in enumerate(items_list, 1):
        g_amt = item['gerp_total_amt']
        e_amt = item['erp_day_amt'] + item['erp_ngt_amt']
        g_qty = item['gerp_day_qty'] + item['gerp_ngt_qty']
        e_qty = item['erp_day_qty'] + item['erp_ngt_qty']
        qty_diff = g_qty - e_qty
        amt_diff = g_amt - e_amt
        diff_fill = NEG_FILL if amt_diff < 0 else (POS_FILL if amt_diff > 0 else None)

        # GERP 단가 조회 — 기준단가와 매칭되는 GERP단가 반환
        price = item['price']
        gp_prices = gerp_price_set.get((lc, item['part_no']), set())
        if price in gp_prices:
            gerp_p = int(price)  # 기준단가와 동일한 GERP단가 존재
        elif len(gp_prices) == 1:
            gerp_p = int(next(iter(gp_prices)))  # 단일 GERP단가
        elif len(gp_prices) > 1:
            gerp_p = int(min(gp_prices))  # 다중단가 중 최소값 (임시)
        else:
            gerp_p = 0
        gp_fill = WARN_FILL if (gerp_p > 0 and price > 0 and price != gerp_p) else None

        # 유형 판정
        price = item['price']
        has_price_diff = (gerp_p > 0 and price > 0 and price != gerp_p)
        has_qty_diff = (qty_diff != 0)
        gerp_missing = (g_amt == 0 and e_amt > 0)   # GERP에 없고 구ERP에 있음
        erp_missing  = (e_amt == 0 and g_amt > 0)   # 구ERP에 없고 GERP에 있음
        ref_missing  = (price == 0 and g_amt > 0)    # 기준정보에 단가 없음
        is_multi_price = (item['part_no'] in multi_price_pns)

        if gerp_missing:
            dtype = 'G실적누락'
        elif ref_missing:
            dtype = '기준누락'
        elif erp_missing:
            dtype = '구실적누락'
        elif is_multi_price:
            dtype = '다중단가'
        elif has_price_diff and has_qty_diff:
            dtype = '단가+수량'
        elif has_price_diff:
            dtype = '단가차이'
        elif has_qty_diff:
            dtype = '수량차이'
        elif amt_diff != 0:
            dtype = '정산차이'
        else:
            dtype = ''

        assy = item.get('assy_part', '')
        if has_night:
            if lc == 'SD9A01':
                # 지원분 정보
                sup_list = item.get('support', [])
                if sup_list:
                    sup_vendor = ', '.join(s['vendor'] for s in sup_list)
                    sup_line   = ', '.join(s['line_code'] for s in sup_list)
                    sup_qty    = sum(s['day_qty'] + s['night_qty'] for s in sup_list)
                else:
                    sup_vendor = '-'
                    sup_line   = '-'
                    sup_qty    = '-'
                vals = [i, item['part_no'], assy, item.get('usage', 1), item['price'], gerp_p,
                        item['gerp_day_qty'], item['gerp_day_amt'],
                        item['gerp_ngt_qty'], item['gerp_ngt_amt'], item['gerp_total_amt'],
                        item['erp_day_qty'], item['erp_day_amt'],
                        item['erp_ngt_qty'], item['erp_ngt_amt'],
                        sup_vendor, sup_line, sup_qty,
                        qty_diff, amt_diff, dtype]
            else:  # SP3M3
                vals = [i, item['part_no'], assy, item.get('usage', 1), item['price'], gerp_p,
                        item['gerp_day_qty'], item['gerp_day_amt'],
                        item['gerp_ngt_qty'], item['gerp_ngt_amt'],
                        item['erp_day_qty'], item['erp_day_amt'],
                        item['erp_ngt_qty'], item['erp_ngt_amt'],
                        qty_diff, amt_diff, dtype]
        else:
            vals = [i, item['part_no'], assy, item.get('usage', 1), item['price'], gerp_p,
                    item['gerp_day_qty'], item['gerp_day_amt'],
                    item['erp_day_qty'], item['erp_day_amt'],
                    qty_diff, amt_diff, dtype]

        # GERP+구ERP 양쪽 실적 없으면 행 숨김
        gerp_total = item['gerp_total_amt']
        erp_total = item['erp_day_amt'] + item.get('erp_ngt_amt', 0)
        if gerp_total == 0 and erp_total == 0:
            ws2.row_dimensions[r].hidden = True
            hidden_count += 1

        for c, v in enumerate(vals, 1):
            # 수량/금액 0값 → "-" 변환 (No, 품번, 조립품번, Usage, 유형 제외)
            if c >= 5 and c < len(vals) and isinstance(v, (int, float)) and v == 0:
                v = '-'
            fmt = NUM_FMT if (isinstance(v, (int, float)) and c >= 5) else None
            if c == 6:  # GERP단가
                cf = gp_fill
            elif c >= len(vals) - 2:  # 수량차, 금액차, 유형 (마지막 3컬럼)
                cf = diff_fill if isinstance(v, (int, float)) and v != 0 else None
            elif c >= 7:
                cf = None
            else:
                cf = None
            dc(ws2, r, c, v, fmt=fmt, fill=cf)
        r += 1

    # ── 합계행 ──
    for c in range(1, ncols + 1):
        ws2.cell(r, c).fill = GRAY_FILL
        ws2.cell(r, c).font = BOLD_FONT
        ws2.cell(r, c).border = BDR
    dc(ws2, r, 1, '합계', bold=True, fill=GRAY_FILL, align=C)

    # 합계 산출
    t_g_qty = ld.get('total_gerp_day_qty', 0) + ld.get('total_gerp_ngt_qty', 0)
    t_e_qty = ld.get('total_erp_day_qty', 0) + ld.get('total_erp_ngt_qty', 0)
    t_g_amt = ld.get('total_gerp_amt', ld.get('total_gerp_day_amt', 0))
    t_e_amt = ld.get('total_erp_day_amt', 0) + ld.get('total_erp_ngt_amt', 0)
    t_qty_diff = t_g_qty - t_e_qty
    t_amt_diff = t_g_amt - t_e_amt

    if has_night:
        if lc == 'SD9A01':
            # GERP: 주간수량(7) 주간금액(8) 야간수량(9) 야간금액(10) 합계(11)
            dc(ws2, r, 7, ld.get('total_gerp_day_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 8, ld.get('total_gerp_day_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 9, ld.get('total_gerp_ngt_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 10, ld.get('total_gerp_ngt_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 11, ld.get('total_gerp_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            # 구ERP: 주간수량(12) 주간금액(13) 야간수량(14) 야간금액(15)
            dc(ws2, r, 12, ld.get('total_erp_day_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 13, ld.get('total_erp_day_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 14, ld.get('total_erp_ngt_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 15, ld.get('total_erp_ngt_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            # 결과: 수량차(16) 금액차(17)
            dc(ws2, r, 16, t_qty_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 17, t_amt_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        else:  # SP3M3
            # GERP: 주간수량(7) 주간금액(8) 야간수량(9) 야간금액(10)
            dc(ws2, r, 7, ld.get('total_gerp_day_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 8, ld.get('total_gerp_day_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 9, ld.get('total_gerp_ngt_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 10, ld.get('total_gerp_ngt_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            # 구ERP: 주간수량(11) 주간금액(12) 야간수량(13) 야간금액(14)
            dc(ws2, r, 11, ld.get('total_erp_day_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 12, ld.get('total_erp_day_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 13, ld.get('total_erp_ngt_qty', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 14, ld.get('total_erp_ngt_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            # 결과: 수량차(15) 금액차(16)
            dc(ws2, r, 15, t_qty_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
            dc(ws2, r, 16, t_amt_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
    else:
        dc(ws2, r, 7, t_g_qty, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws2, r, 8, ld.get('total_gerp_day_amt', 0), bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws2, r, 9, t_e_qty, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws2, r, 10, t_e_amt, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws2, r, 11, t_qty_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws2, r, 12, t_amt_diff, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)

    # ── 열 너비 ──
    detail_widths = {1: 5, 2: 20, 3: 14, 4: 6, 5: 8, 6: 8}
    for ci in range(7, ncols + 1):
        detail_widths[ci] = 14
    detail_widths[ncols] = 10  # 유형 컬럼
    for ci, w in detail_widths.items():
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws2.freeze_panes = 'A5'
    print(f"  {lc}: {item_count}행")

# ══════════════════════════════════════════════════════════════
# 시트: 01_차이분석
# ══════════════════════════════════════════════════════════════
print("[3/4] 01_차이분석 시트...")
ws3 = wb.create_sheet('01_차이분석')

TITLE2_FONT = Font(name='맑은 고딕', bold=True, size=12, color='1B2A4A')
ws3['A1'] = f'{MONTH}월 GERP vs 구ERP 차이 분석'
ws3['A1'].font = TITLE2_FONT

# 그룹 헤더 (row 2)
GRP3_NCOLS = 15
ws3.merge_cells('A2:C2'); ws3['A2'] = '기본정보'
ws3.merge_cells('D2:G2'); ws3['D2'] = 'GERP 정산'
ws3.merge_cells('H2:K2'); ws3['H2'] = '구ERP (참고)'
ws3.merge_cells('L2:M2'); ws3['L2'] = '차이'
ws3['N2'] = '유형'
ws3['O2'] = '비고'
for col in range(1, GRP3_NCOLS + 1):
    cell = ws3.cell(2, col)
    cell.fill = GRP_FILL
    cell.font = GRP_FONT
    cell.alignment = C
    cell.border = BDR

# 상세 헤더 (row 3)
hdrs3 = [
    '라인', '품번', '기준단가',
    'GERP\n주간qty', 'GERP\n주간amt', 'GERP\n야간qty', 'GERP\n합계amt',
    '구ERP\n주간qty', '구ERP\n주간amt', '구ERP\n야간qty', '구ERP\n합계amt',
    '수량차\n(G-E)', '금액차\n(G-E)',
    '유형', '비고'
]
for col, h in enumerate(hdrs3, 1):
    ws3.cell(3, col, h)
hdr(ws3, 3, len(hdrs3))
ws3.row_dimensions[3].height = 30

r = 4
diff_count = 0

for lc in LINE_ORDER:
    ld = lines.get(lc, {})
    items_list = ld.get('items', [])
    has_night = LINE_INFO[lc]['has_night']

    # 다중단가 품번 집계 (라인별)
    from collections import Counter
    pn_counter = Counter(it['part_no'] for it in items_list)
    multi_price_pns_d = {pn for pn, cnt in pn_counter.items() if cnt > 1}

    line_diff_count = 0
    line_gerp_sum = 0
    line_erp_sum  = 0
    line_start_r  = r

    for item in items_list:
        gerp_amt = item['gerp_total_amt']
        erp_amt  = item['erp_day_amt'] + item['erp_ngt_amt']
        diff_amt = gerp_amt - erp_amt

        if diff_amt == 0:
            continue

        g_qty = item['gerp_day_qty'] + item['gerp_ngt_qty']
        e_qty = item['erp_day_qty'] + item['erp_ngt_qty']
        qty_diff = g_qty - e_qty

        # GERP 단가 조회
        ps = gerp_price_set.get((lc, item['part_no']), set())
        price = item['price']
        if price in ps:
            gerp_p = price
        else:
            gerp_p = min(ps) if ps else 0

        has_price_diff = (gerp_p > 0 and price > 0 and price != gerp_p)
        has_qty_diff   = (qty_diff != 0)
        gerp_missing   = (gerp_amt == 0 and erp_amt > 0)
        erp_missing    = (erp_amt == 0 and gerp_amt > 0)
        ref_missing    = (price == 0 and gerp_amt > 0)
        is_multi       = (item['part_no'] in multi_price_pns_d)

        if gerp_missing:
            dtype = 'G실적누락'
        elif ref_missing:
            dtype = '기준누락'
        elif erp_missing:
            dtype = '구실적누락'
        elif is_multi:
            dtype = '다중단가'
        elif has_price_diff and has_qty_diff:
            dtype = '단가+수량'
        elif has_price_diff:
            dtype = '단가차이'
        elif has_qty_diff:
            dtype = '수량차이'
        elif diff_amt != 0:
            dtype = '정산차이'
        else:
            dtype = ''

        note = ''
        if item.get('usage', 1) == 2:
            note = 'Usage=2'
        if lc == 'SP3M3' and item['gerp_ngt_qty'] > 0:
            note += (' ' if note else '') + 'RSP야간'

        fill = NEG_FILL if diff_amt < 0 else POS_FILL

        vals = [
            lc, item['part_no'], price,
            item['gerp_day_qty'], item['gerp_day_amt'],
            item['gerp_ngt_qty'] if has_night else '-', gerp_amt,
            item['erp_day_qty'], item['erp_day_amt'],
            item['erp_ngt_qty'] if has_night else '-', erp_amt,
            qty_diff, diff_amt,
            dtype, note.strip()
        ]

        for col, v in enumerate(vals, 1):
            # 0값 → "-" (라인/품번/단가/유형/비고 제외)
            if col not in (1, 2, 3, 14, 15) and isinstance(v, (int, float)) and v == 0:
                v = '-'
            fmt = NUM_FMT if isinstance(v, (int, float)) and col not in (1, 14, 15) else None
            f = fill if col in (12, 13) else None
            al = C if col in (1, 14, 15) else R
            dc(ws3, r, col, v, fmt=fmt, fill=f, align=al)

        r += 1
        line_diff_count += 1
        line_gerp_sum += gerp_amt
        line_erp_sum  += erp_amt

    # 라인별 소계 행
    if line_diff_count > 0:
        for col in range(1, GRP3_NCOLS + 1):
            ws3.cell(r, col).fill = GRAY_FILL
            ws3.cell(r, col).font = BOLD_FONT
            ws3.cell(r, col).border = BDR
        dc(ws3, r, 1, f'{lc} 소계', bold=True, fill=GRAY_FILL, align=C)
        dc(ws3, r, 7, line_gerp_sum, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        dc(ws3, r, 11, line_erp_sum, bold=True, fmt=NUM_FMT, fill=GRAY_FILL)
        line_total_diff = line_gerp_sum - line_erp_sum
        df = NEG_FILL if line_total_diff < 0 else (POS_FILL if line_total_diff > 0 else GRAY_FILL)
        dc(ws3, r, 13, line_total_diff, bold=True, fmt=NUM_FMT, fill=df)
        r += 1
        diff_count += line_diff_count

# 열 너비
col3_widths = {1:10, 2:22, 3:10, 4:10, 5:14, 6:10, 7:14,
               8:10, 9:14, 10:10, 11:14, 12:12, 13:14, 14:10, 15:10}
for ci, w in col3_widths.items():
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = 'A4'
print(f"  차이항목: {diff_count}건")

# ══════════════════════════════════════════════════════════════
# 시트: 03_검증결과  (step6_validation.json 기반)
# ══════════════════════════════════════════════════════════════
if s6 is not None:
    print("[5/5] 03_검증결과 시트...")
    ws5 = wb.create_sheet('03_검증결과')

    overall  = s6.get('overall', 'N/A')
    crit_n   = s6.get('critical_fail', 0)
    warn_n   = s6.get('warning', 0)
    ts       = s6.get('timestamp', '')[:16].replace('T', ' ')

    # 제목
    ws5.merge_cells('A1:E1')
    ws5['A1'] = f'{MONTH}월 정산 검증 결과 (Step 6)'
    ws5['A1'].font = Font(name='맑은 고딕', bold=True, size=11)
    ws5.row_dimensions[1].height = 22

    # 종합 판정 행
    overall_fill = (
        PatternFill('solid', fgColor='FCE4EC') if overall == 'FAIL' else
        PatternFill('solid', fgColor='FFF2CC') if overall == 'WARNING' else
        PatternFill('solid', fgColor='E8F5E9')
    )
    ws5.merge_cells('A2:E2')
    ws5['A2'] = (f"최종 판정: {overall}  |  CRITICAL FAIL: {crit_n}건  |  "
                 f"WARNING: {warn_n}건  |  실행시각: {ts}")
    ws5['A2'].font = Font(name='맑은 고딕', bold=True, size=10)
    ws5['A2'].fill = overall_fill
    ws5['A2'].alignment = L
    ws5.row_dimensions[2].height = 18

    # 헤더 (row 4)
    for c, h in enumerate(['No', '항목명', '상태', '심각도', '상세'], 1):
        ws5.cell(4, c, h)
    hdr(ws5, 4, 5)

    # ── WARNING / FAIL 항목만 표시 (INFO/PASS 제외)
    WARN_STATUS = {'FAIL', 'WARNING'}
    checks = s6.get('checks', [])
    notable = [c for c in checks if c.get('status') in WARN_STATUS]
    all_checks = checks  # 전체도 함께 표시

    row = 5
    for ck in all_checks:
        st  = ck.get('status', '')
        sev = ck.get('severity', '')
        fill = None
        if st == 'FAIL':
            fill = NEG_FILL
        elif st == 'WARNING':
            fill = WARN_FILL

        dc(ws5, row, 1, ck.get('no'), align=C)
        dc(ws5, row, 2, ck.get('name'), fill=fill)
        dc(ws5, row, 3, st, fill=fill, align=C)
        dc(ws5, row, 4, sev, fill=fill, align=C)
        dc(ws5, row, 5, ck.get('detail', ''), fill=fill)
        row += 1

    # ── WARNING 전용 섹션 (별도 블록)
    if notable:
        row += 1
        ws5.merge_cells(f'A{row}:E{row}')
        ws5.cell(row, 1, '▶ 주의 항목 — WARNING / CRITICAL FAIL')
        ws5.cell(row, 1).font = Font(name='맑은 고딕', bold=True, size=9, color='C62828')
        ws5.cell(row, 1).fill = PatternFill('solid', fgColor='FFEBEE')
        row += 1

        for c, h in enumerate(['No', '항목명', '상태', '심각도', '상세'], 1):
            ws5.cell(row, c, h)
        hdr(ws5, row, 5)
        row += 1

        for ck in notable:
            st  = ck.get('status', '')
            fill = WARN_FILL if st == 'WARNING' else NEG_FILL
            dc(ws5, row, 1, ck.get('no'), align=C)
            dc(ws5, row, 2, ck.get('name'), fill=fill)
            dc(ws5, row, 3, st, fill=fill, align=C)
            dc(ws5, row, 4, ck.get('severity', ''), fill=fill, align=C)
            dc(ws5, row, 5, ck.get('detail', ''), fill=fill)
            row += 1

    # 열 너비
    ws5.column_dimensions['A'].width = 6
    ws5.column_dimensions['B'].width = 38
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 12
    ws5.column_dimensions['E'].width = 55
    print(f"  검증항목: {len(checks)}건 (WARNING/FAIL: {len(notable)}건)")

# ── 저장 ──────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print(f"저장 완료: {OUTPUT_FILE}")
extra = ' | 03_검증결과' if s6 is not None else ''
print(f"시트 구성: 00_정산집계 | {' | '.join(LINE_ORDER)} | 01_차이분석{extra}")
print("Step 7 완료 — 최종 보고서 생성")
