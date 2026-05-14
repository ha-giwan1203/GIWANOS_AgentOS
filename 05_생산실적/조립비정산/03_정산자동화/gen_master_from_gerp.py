#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GERP 실적 데이터에서 기준정보를 역설계하여 기존 양식에 맞게 생성한다.

소스: G-ERP 3월실적.xlsx (대원테크 0109, "정상" 행만)
출력: 01_기준정보/기준정보_GERP역설계_YYYYMMDD.xlsx
양식: 라인별 시트, Row1=라인명, Row2=빈행, Row3=헤더, Row4+=데이터
헤더: 품번 | 업체코드 | 라인 | 조립품번 | Usage | 단가구분 | 단가 | 차종

핵심 규칙:
  - "추가"(야간) 행 제외: 야간수당 30%가 포함된 단가이므로 기준정보 부적합
  - 다중단가 보존: 동일 품번이라도 조립품번/단가가 다르면 별도 행
  - Usage 보존: 1 또는 2 그대로
  - price_type: 단가>0 → 정단가, 단가=0 → 미단가
  - RSP 모듈품번(SP3M3 야간): "정상" 필터로 자동 제외
"""

import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment

# 파이프라인 설정 import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _pipeline_config import (
    GERP_FILE, GERP_COL, MASTER_COL, VENDOR_CODE,
    LINE_ORDER, LINE_INFO, BASE_DIR
)

# ============================================================
# 설정
# ============================================================
OUTPUT_DIR = os.path.join(BASE_DIR, '01_기준정보')
TIMESTAMP = datetime.now().strftime("%Y%m%d")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f'기준정보_GERP역설계_{TIMESTAMP}.xlsx')

HEADER = ['품번', '업체코드', '라인', '조립품번', 'Usage', '단가구분', '단가', '차종']


def load_gerp():
    """GERP 실적 로드 → 대원테크 정상 행만 추출"""
    wb = openpyxl.load_workbook(GERP_FILE, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    filtered = []
    for r in rows:
        vendor = str(r[GERP_COL['vendor_cd']]).strip() if r[GERP_COL['vendor_cd']] else ''
        shift = str(r[GERP_COL['shift']]).strip() if r[GERP_COL['shift']] else ''
        if vendor == VENDOR_CODE and '정상' in shift:
            filtered.append(r)

    print(f"[GERP] 전체 {len(rows)}행 → 대원테크 정상 {len(filtered)}행")
    return filtered


def build_master(gerp_rows):
    """GERP 행을 기준정보 8열 구조로 변환 + 라인별 그룹핑"""
    # 키: (line, part_no, price, assy_part) → 중복 제거 + 다중단가 보존
    seen = {}

    for r in gerp_rows:
        line = r[GERP_COL['line']]
        part_no = r[GERP_COL['product_no']]
        usage = r[GERP_COL['usage']]
        assy_part = r[GERP_COL['assy_part']]
        unit_price = r[GERP_COL['unit_price']]
        vtype = r[5] if r[5] else ''  # col5 = 차종

        # 단가 정규화
        try:
            price = float(unit_price) if unit_price else 0
        except (ValueError, TypeError):
            price = 0

        # 단가구분
        price_type = '정단가' if price > 0 else '미단가'

        # Usage 정규화
        try:
            usage_val = int(usage) if usage else 1
        except (ValueError, TypeError):
            usage_val = 1

        # dedup 키
        key = (line, part_no, price, assy_part)
        if key not in seen:
            seen[key] = {
                'part_no': part_no,
                'vendor_cd': VENDOR_CODE,
                'line_code': line,
                'assy_part': assy_part if assy_part else '',
                'usage': usage_val,
                'price_type': price_type,
                'price': int(price) if price == int(price) else price,
                'vtype': str(vtype) if vtype else '',
            }

    print(f"[변환] 고유 조합: {len(seen)}개")

    # 라인별 그룹핑
    by_line = defaultdict(list)
    for row_data in seen.values():
        by_line[row_data['line_code']].append(row_data)

    # 라인 내 정렬: 품번 → 단가 → 조립품번
    for line in by_line:
        by_line[line].sort(key=lambda x: (str(x['part_no']), float(x['price']), str(x['assy_part'])))

    return by_line


def write_excel(by_line):
    """기존 양식에 맞게 엑셀 출력"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 삭제

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    header_align = Alignment(horizontal='center')

    total_rows = 0

    for line_code in LINE_ORDER:
        if line_code not in by_line:
            print(f"  [경고] {line_code}: GERP 데이터 없음 → 빈 시트 생성")
            rows = []
        else:
            rows = by_line[line_code]

        line_name = LINE_INFO.get(line_code, {}).get('name', line_code)

        # 시트명: ISAMS03 → '이너센스ASSY' 특수 처리 (기존 기준정보 시트명 따름)
        sheet_name = line_code
        if line_code == 'ISAMS03':
            sheet_name = '이너센스ASSY'

        ws = wb.create_sheet(title=sheet_name)

        # Row 1: 라인명
        ws.cell(row=1, column=1, value=f"{line_code} ({line_name})").font = title_font

        # Row 2: 빈행 (그대로 비움)

        # Row 3: 헤더
        for ci, h in enumerate(HEADER, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = header_font
            cell.alignment = header_align

        # Row 4+: 데이터
        for ri, row_data in enumerate(rows, 4):
            ws.cell(row=ri, column=1, value=row_data['part_no'])
            ws.cell(row=ri, column=2, value=row_data['vendor_cd'])
            ws.cell(row=ri, column=3, value=row_data['line_code'])
            ws.cell(row=ri, column=4, value=row_data['assy_part'])
            ws.cell(row=ri, column=5, value=row_data['usage'])
            ws.cell(row=ri, column=6, value=row_data['price_type'])
            ws.cell(row=ri, column=7, value=row_data['price'])
            ws.cell(row=ri, column=8, value=row_data['vtype'])

        total_rows += len(rows)
        print(f"  {line_code} ({sheet_name}): {len(rows)}행")

        # 열 너비 조정
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 8

    wb.save(OUTPUT_FILE)
    print(f"\n[저장] {OUTPUT_FILE}")
    print(f"[합계] {total_rows}행, {len(LINE_ORDER)}시트")
    return total_rows


def print_summary(by_line):
    """다중단가, 미단가 등 주요 통계 출력"""
    print("\n=== 품질 요약 ===")

    # 다중단가 체크
    multi_price_count = 0
    for line, rows in by_line.items():
        parts = defaultdict(list)
        for r in rows:
            parts[r['part_no']].append(r['price'])
        for part, prices in parts.items():
            if len(set(prices)) > 1:
                multi_price_count += 1

    # 미단가 체크
    zero_price = sum(1 for rows in by_line.values() for r in rows if r['price'] == 0)

    # Usage=2 체크
    usage2 = sum(1 for rows in by_line.values() for r in rows if r['usage'] == 2)

    print(f"  다중단가 품번: {multi_price_count}건")
    print(f"  미단가(price=0): {zero_price}건")
    print(f"  Usage=2: {usage2}건")


def main():
    print("=== GERP 기반 기준정보 역설계 ===")
    print(f"소스: {GERP_FILE}")
    print(f"출력: {OUTPUT_FILE}")
    print()

    # 1. GERP 로드
    gerp_rows = load_gerp()

    # 2. 기준정보 변환
    by_line = build_master(gerp_rows)

    # 3. 엑셀 출력
    print("\n[시트별 현황]")
    total = write_excel(by_line)

    # 4. 품질 요약
    print_summary(by_line)

    print("\n=== 완료 ===")
    return total


if __name__ == "__main__":
    main()
