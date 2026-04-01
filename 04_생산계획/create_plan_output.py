"""
PLAN_OUTPUT 동적 조회 시트 생성 스크립트
SP3M3_생산지시서_메크로자동화_v2.xlsm에 PLAN_OUTPUT 시트를 추가한다.

GPT 공동작업 설계안 반영:
- Excel Table(ListObject) 기반 — tblPlanOutput
- SP3 LINE 기준 정보도 Table(tblInfo)로 변환
- MATCH + INDEX 패턴 (INFO_ROW helper 열)
- 직접조회 + 파생계산 분리
- PLAN_RESULT_FIXED 미변경
- 행 추가/삭제 자유 (Table 수식 자동 전파)
"""
import shutil
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

XLSM_PATH = "SP3M3_생산지시서_메크로자동화_v2.xlsm"
BACKUP_PATH = XLSM_PATH + ".before_plan_output.bak"
SHEET_NAME = "PLAN_OUTPUT"
TABLE_NAME = "tblPlanOutput"
REF_TABLE_NAME = "tblInfo"

# 기준 정보 시트
REF_SHEET = "SP3 LINE 기준 정보"
REF_DATA_ROWS = 1300  # 기준 정보 최대 행 (여유 포함)

# --- 컬럼 설계 (GPT 합의안) ---
# (header, type, meta)
# type: input / helper / lookup / derived
# meta: lookup일 때 기준 정보 열 번호(1-based, B=2), 그 외 None
COLUMNS = [
    # 입력 열
    ("NO",            "helper",   None),
    ("KIND",          "input",    None),
    ("SUB_PART",      "input",    None),    # 핵심 키
    ("TIME_TEXT",     "input",    None),
    ("ORDER_QTY",     "input",    None),
    ("CARRY_FLAG",    "input",    None),
    ("PLAN_KEY",      "input",    None),    # 버튼 append용

    # Helper 열 (숨김)
    ("INFO_ROW",      "helper",   None),

    # 직접 조회 열 (MATCH+INDEX)
    ("CAR_TYPE",      "lookup",   2),
    ("PART_NO",       "lookup",   3),
    ("MODULE_PART",   "lookup",   5),
    ("SENSE_SPEC",    "lookup",   6),
    ("SENSE_LOC",     "lookup",   7),
    ("HOUSING",       "lookup",   8),
    ("HOLDER",        "lookup",   9),
    ("TORSION",       "lookup",  10),
    ("TORSION_LOC",   "lookup",  11),
    ("FRAME_LOC",     "lookup",  12),
    ("NOTE",          "lookup",  13),
    ("V_GEAR",        "lookup",  15),
    ("MODEL_1",       "lookup",  16),
    ("MODEL_2",       "lookup",  18),
    ("SUPPLY_LINE",   "lookup",  19),
    ("FRAME_PART",    "lookup",  20),
    ("OLD_MES_PART",  "lookup",  21),
    ("BOX_QTY",       "lookup",  22),
    ("P_TRAY_TYPE",   "lookup",  23),
    ("ANGLE_TYPE",    "lookup",  24),

    # Helper 열 (숨김)
    ("TEXT_ALL",      "helper",   None),

    # 파생 열
    ("MODEL_GROUP",   "derived",  None),
    ("PI_TYPE",       "derived",  None),
    ("PLL_FLAG",      "derived",  None),
    ("TORSION_NUM",   "derived",  None),
    ("CAR_FAMILY",    "derived",  None),
    ("DIR_KEY",       "derived",  None),
    ("PTRAY_FLAG",    "derived",  None),
    ("ANGLE_KEY",     "derived",  None),
]

# 컬럼 인덱스 찾기 (1-based)
COL_IDX = {name: i + 1 for i, (name, _, _) in enumerate(COLUMNS)}


def col_ref(col_name, row, abs_col=True):
    """컬럼명 → 셀 참조 (예: SUB_PART, 2 → $C2)"""
    idx = COL_IDX[col_name]
    letter = get_column_letter(idx)
    if abs_col:
        return f"${letter}{row}"
    return f"{letter}{row}"


def ref_range(col_letter, start_row=2, end_row=None):
    """기준 정보 시트 범위 참조"""
    end = end_row or REF_DATA_ROWS
    return f"'{REF_SHEET}'!${col_letter}${start_row}:${col_letter}${end}"


def ref_col_range(col_num):
    """기준 정보 시트 열 번호 → 범위 참조 (데이터만, 헤더 제외)"""
    letter = get_column_letter(col_num)
    return ref_range(letter)


def make_info_row_formula(row):
    """INFO_ROW: MATCH로 기준 정보에서 SUB_PART 위치 찾기"""
    sub = col_ref("SUB_PART", row)
    ref = ref_range("A")
    return f'=IF({sub}="","",IFERROR(MATCH({sub},{ref},0),""))'


def make_lookup_formula(row, ref_col_num):
    """직접 조회: INDEX(기준정보 열, INFO_ROW)"""
    info_row = col_ref("INFO_ROW", row)
    ref = ref_col_range(ref_col_num)
    return f'=IF({info_row}="","",INDEX({ref},{info_row}))'


def make_no_formula(row):
    """순번: SUB_PART가 있으면 행 번호 기반 순번"""
    sub = col_ref("SUB_PART", row)
    return f'=IF({sub}="","",ROW()-1)'


def make_text_all_formula(row):
    """TEXT_ALL: 파생 계산용 합성 텍스트"""
    cols = ["CAR_TYPE", "PART_NO", "MODULE_PART", "SENSE_SPEC", "NOTE", "TORSION"]
    refs = ",".join(col_ref(c, row) for c in cols)
    return f'=UPPER(TEXTJOIN(" ",TRUE,{refs}))'


def make_derived_formulas(row):
    """파생 열 수식 딕셔너리 반환"""
    car = col_ref("CAR_TYPE", row)
    txt = col_ref("TEXT_ALL", row)
    tor = col_ref("TORSION", row)
    note = col_ref("NOTE", row)
    ptray = col_ref("P_TRAY_TYPE", row)
    pno = col_ref("PART_NO", row)
    mod = col_ref("MODULE_PART", row)
    pflag = col_ref("PTRAY_FLAG", row)
    ang = col_ref("ANGLE_TYPE", row)

    return {
        "MODEL_GROUP": (
            f'=IF({car}="","",LET(x,UPPER(TRIM({car})),'
            f'a,TEXTBEFORE(x&"(","("),'
            f'b,TEXTBEFORE(TRIM(a)&" "," "),'
            f'SUBSTITUTE(b," ","")))'
        ),
        "PI_TYPE": (
            f'=IF({txt}="","",IF(ISNUMBER(SEARCH("8PI",{txt})),"8PI",'
            f'IF(ISNUMBER(SEARCH("10PI",{txt})),"10PI","OTHER")))'
        ),
        "PLL_FLAG": (
            f'=IF({txt}="","",IF(ISNUMBER(SEARCH("PLL",{txt})),"PLL","GEN"))'
        ),
        "TORSION_NUM": (
            f'=IF({tor}="","",IFERROR(--TEXTBEFORE(TRIM({tor})&" "," "),""))'
        ),
        "CAR_FAMILY": (
            f'=IF({car}="","",UPPER(SUBSTITUTE(TEXTBEFORE(TRIM({car})&"(","(")," ","")))'
        ),
        "DIR_KEY": (
            f'=LET(t,UPPER({car}&" "&{note}),'
            f'IF(ISNUMBER(SEARCH("LH",t)),"LH",IF(ISNUMBER(SEARCH("RH",t)),"RH","")))'
        ),
        "PTRAY_FLAG": (
            f'=LET(t,UPPER({ptray}&" "&{car}&" "&{pno}&" "&{mod}&" "&{note}),'
            f'u,SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(t," ",""),"-",""),"_",""),'
            f'IF(OR(ISNUMBER(SEARCH("P/TRAY",u)),ISNUMBER(SEARCH("PTRAY",u))),"PTRAY","GEN"))'
        ),
        "ANGLE_KEY": (
            f'=IF({pflag}<>"PTRAY","",LET(a,SUBSTITUTE(SUBSTITUTE(UPPER(TRIM({ang})),'
            f'"도",""),"°",""),IF(OR(a="0",a="180",a="270",a="90"),a,"")))'
        ),
    }


def main():
    if not os.path.exists(XLSM_PATH):
        raise FileNotFoundError(f"파일 없음: {XLSM_PATH}")

    # 백업
    shutil.copy2(XLSM_PATH, BACKUP_PATH)
    print(f"백업: {BACKUP_PATH}")

    # xlsm 열기 (VBA 매크로 보존)
    wb = load_workbook(XLSM_PATH, keep_vba=True)
    print(f"시트 목록: {wb.sheetnames}")

    # 기준 정보 시트 존재 확인
    if REF_SHEET not in wb.sheetnames:
        raise ValueError(f"기준 정보 시트 '{REF_SHEET}' 없음")

    # --- Step 1: 기준 정보 시트에 tblInfo Table 생성 ---
    ws_info = wb[REF_SHEET]
    info_has_table = False
    for t in ws_info.tables.values():
        if t.displayName == REF_TABLE_NAME:
            info_has_table = True
            break
    if not info_has_table:
        # 실제 데이터 범위 확인
        info_last_row = ws_info.max_row
        info_last_col = min(ws_info.max_column, 24)  # A~X = 24열
        info_ref = f"A1:{get_column_letter(info_last_col)}{info_last_row}"
        info_table = Table(displayName=REF_TABLE_NAME, ref=info_ref)
        info_style = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        info_table.tableStyleInfo = info_style
        ws_info.add_table(info_table)
        print(f"'{REF_TABLE_NAME}' Table 생성: {info_ref}")
    else:
        print(f"'{REF_TABLE_NAME}' Table 이미 존재")

    # --- Step 2: PLAN_OUTPUT 시트 생성 ---
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"기존 '{SHEET_NAME}' 시트 삭제")

    ws = wb.create_sheet(SHEET_NAME)
    print(f"'{SHEET_NAME}' 시트 생성")

    # --- Step 3: 헤더 (1행) ---
    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # --- Step 4: 수식 템플릿 (2~11행, 10행) ---
    TEMPLATE_ROWS = 10
    for row in range(2, 2 + TEMPLATE_ROWS):
        derived = make_derived_formulas(row)

        for col_idx, (header, col_type, meta) in enumerate(COLUMNS, start=1):
            if header == "NO":
                ws.cell(row=row, column=col_idx, value=make_no_formula(row))
            elif header == "INFO_ROW":
                ws.cell(row=row, column=col_idx, value=make_info_row_formula(row))
            elif header == "TEXT_ALL":
                ws.cell(row=row, column=col_idx, value=make_text_all_formula(row))
            elif col_type == "lookup" and meta is not None:
                ws.cell(row=row, column=col_idx, value=make_lookup_formula(row, meta))
            elif col_type == "derived" and header in derived:
                ws.cell(row=row, column=col_idx, value=derived[header])
            # input 열은 빈 셀

    print(f"수식 템플릿: {TEMPLATE_ROWS}행 생성")

    # --- Step 5: Excel Table 생성 ---
    last_col_letter = get_column_letter(len(COLUMNS))
    last_row = 1 + TEMPLATE_ROWS
    table_ref = f"A1:{last_col_letter}{last_row}"

    table = Table(displayName=TABLE_NAME, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    print(f"Table '{TABLE_NAME}' 생성: {table_ref}")

    # --- Step 6: 컬럼 너비 + helper 열 숨김 ---
    widths = {
        "NO": 5, "KIND": 8, "SUB_PART": 12, "TIME_TEXT": 12,
        "ORDER_QTY": 10, "CARRY_FLAG": 10, "PLAN_KEY": 12,
        "INFO_ROW": 8, "CAR_TYPE": 22, "PART_NO": 14, "MODULE_PART": 16,
        "SENSE_SPEC": 16, "SENSE_LOC": 14, "HOUSING": 14, "HOLDER": 12,
        "TORSION": 16, "TORSION_LOC": 14, "FRAME_LOC": 14, "NOTE": 14,
        "V_GEAR": 14, "MODEL_1": 14, "MODEL_2": 14, "SUPPLY_LINE": 12,
        "FRAME_PART": 14, "OLD_MES_PART": 14, "BOX_QTY": 12,
        "P_TRAY_TYPE": 14, "ANGLE_TYPE": 14, "TEXT_ALL": 10,
        "MODEL_GROUP": 14, "PI_TYPE": 10, "PLL_FLAG": 10,
        "TORSION_NUM": 12, "CAR_FAMILY": 14, "DIR_KEY": 8,
        "PTRAY_FLAG": 12, "ANGLE_KEY": 10,
    }
    for col_name, width in widths.items():
        if col_name in COL_IDX:
            letter = get_column_letter(COL_IDX[col_name])
            ws.column_dimensions[letter].width = width

    # Helper 열 숨김
    for hidden_col in ["INFO_ROW", "TEXT_ALL"]:
        letter = get_column_letter(COL_IDX[hidden_col])
        ws.column_dimensions[letter].hidden = True

    print("컬럼 너비 설정 + helper 열 숨김 완료")

    # --- Step 7: 저장 ---
    wb.save(XLSM_PATH)
    wb.close()
    print(f"\n저장 완료: {XLSM_PATH}")

    # --- Step 8: 검증 ---
    wb2 = load_workbook(XLSM_PATH, keep_vba=True, data_only=False)
    ws2 = wb2[SHEET_NAME]
    print(f"\n=== 검증 ===")
    print(f"시트: {SHEET_NAME}")
    print(f"행 수: {ws2.max_row}, 열 수: {ws2.max_column}")

    headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    print(f"헤더({len(headers)}개): {headers}")

    # 수식 샘플 (2행)
    samples = ["INFO_ROW", "CAR_TYPE", "MODEL_GROUP", "PI_TYPE"]
    for s in samples:
        if s in COL_IDX:
            val = ws2.cell(row=2, column=COL_IDX[s]).value
            print(f"  {s}(2행): {val}")

    # Table 확인
    for t_name, t_obj in ws2.tables.items():
        ref = t_obj.ref if hasattr(t_obj, 'ref') else t_obj
        print(f"  Table: {t_name} → {ref}")

    # 기준 정보 Table 확인
    ws_info2 = wb2[REF_SHEET]
    for t_name, t_obj in ws_info2.tables.items():
        ref = t_obj.ref if hasattr(t_obj, 'ref') else t_obj
        print(f"  기준정보 Table: {t_name} → {ref}")

    # 기존 시트 보존
    essential = ["SP3 LINE 기준 정보", "PLAN_RESULT_FIXED", "생산계획"]
    for s in essential:
        status = "OK" if s in wb2.sheetnames else "MISSING"
        print(f"  {s}: {status}")

    wb2.close()
    print("\n=== PLAN_OUTPUT 시트 생성 완료 ===")


if __name__ == "__main__":
    main()
