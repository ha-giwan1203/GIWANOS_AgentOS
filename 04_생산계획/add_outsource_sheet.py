# -*- coding: utf-8 -*-
"""SP3M3 생산지시서에 '외주 발주서' 시트 추가.
- 홀더CLR 10종 (ELR 제외) + 10파이 센서 14종 (분홍 제외)
- SUMIFS 수식으로 '출력용' 시트 D+2 주간계획(R37~R80) 집계
"""
import shutil, sys, io
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).parent
SRC  = ROOT / 'SP3M3_생산지시서_(26.05.19).xlsm'
TS   = datetime.now().strftime('%Y%m%d_%H%M%S')
BAK  = ROOT / f'SP3M3_생산지시서_(26.05.19).xlsm.bak_{TS}'

# 1) 백업
shutil.copy2(SRC, BAK)
print(f'[backup] {BAK.name}')

# 2) 로드 (VBA 보존)
wb = load_workbook(SRC, keep_vba=True)
print(f'[load] sheets={len(wb.sheetnames)}')

# 기존 동명 시트 제거 (재실행 안전)
if '외주 발주서' in wb.sheetnames:
    del wb['외주 발주서']

ws = wb.create_sheet('외주 발주서')

# ---------- 마스터 ----------
HOLDER = [
    'P/T NLR노랑',
    'P/T CLR노랑',
    'P/T CLR흰색',
    'P/T CLR노랑\n(MDS)',
    'P/T NLR흰색',
    'NLR 노랑',
    'CLR 노랑',
    'CLR 노랑\n(MDS)',
]
HOLDER_WHITE = [
    'CLR 흰색',
    'NLR 흰색',
]
SENSOR = [
    '10A (노랑) 10PI\nSSP3-0145',
    '10B (파랑) 10PI\nSSP3-0146',
    '23A10D (하늘) 10PI\nPSP300-0136',
    '8B (연녹색) 10PI\nPSP300-0135',
    '3A (연파랑) 10PI\nSSP3-0189',
    '3B (버건디) 10PI\nSSP3-0190',
    '8A (보라색) 10PI\nPSP300-0134',
    '0도 (녹색) 10PI\nSSP3-0191',
    '23B10D (아이보리)10PI\nPSP300-0137',
    '6.5A (빨강) 10PI\nSSP3-0143',
    '6.5B (주황) 10PI\nSSP3-0144',
    '26A (노랑) 10PI\nSSP3-0177',
    '26B (연두) 10PI\nSSP3-0178',
    '19A12D (흰색) 10PI\nSSP3-0204',
]

# ---------- 스타일 ----------
BOLD   = Font(bold=True)
TITLE  = Font(bold=True, size=16)
SUB    = Font(bold=True, size=11, color='FFFFFF')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center', wrap_text=True)
thin   = Side(style='thin', color='808080')
BOX    = Border(top=thin, bottom=thin, left=thin, right=thin)
FILL_TITLE = PatternFill('solid', fgColor='305496')
FILL_HEAD  = PatternFill('solid', fgColor='D9E1F2')
FILL_SUB   = PatternFill('solid', fgColor='F2F2F2')

# ---------- 보조 셀: 주간 시작 행 (야간 갯수 가변 대응) ----------
# 출력용 B열에서 '◀  D+2  주간계획' 표제 찾고 +2 (표제 다음의 헤더 다음 행)
ws['G1'] = '주간시작행'
ws['G1'].font = Font(size=8, color='808080')
ws['H1'] = '=MATCH("◀  D+2  주간계획",\'출력용\'!B:B,0)+2'
ws['H1'].font = Font(size=8, color='808080')

# ---------- 상단 헤더 ----------
ws.merge_cells('A1:E1')
ws['A1'] = '외주 발주서'
ws['A1'].font = TITLE
ws['A1'].alignment = CENTER

# 라벨(A:B 병합) | 값(C) | 기준라벨(D) | 기준값(E)
for r, label in [(2,'발주일자'), (3,'외주처'), (4,'담당자')]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    a = ws.cell(r, 1, label)
    a.font = BOLD; a.fill = FILL_HEAD; a.alignment = CENTER; a.border = BOX
    ws.cell(r, 3, '').alignment = LEFT
    ws.cell(r, 3).border = BOX

ws['C2'] = "='출력용'!Q1"
ws['C2'].number_format = 'yyyy-mm-dd'
ws['C2'].alignment = LEFT
ws['D2'] = '기준'
ws['D2'].font = BOLD; ws['D2'].fill = FILL_HEAD; ws['D2'].alignment = CENTER; ws['D2'].border = BOX
ws['E2'] = 'D+2 주간계획'
ws['E2'].alignment = LEFT; ws['E2'].border = BOX
# D3,E3,D4,E4 는 빈 박스 노출 방지 (테두리 없음)
for r in (2,3,4):
    ws.row_dimensions[r].height = 22

# ---------- 표 헤더 작성 함수 ----------
def write_section(start_row, title, items, sumifs_match_col):
    """sumifs_match_col: '출력용' 매칭컬럼 (홀더=R, 센서=P)"""
    # 섹션 타이틀
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    c = ws.cell(start_row, 1, f'■ {title}')
    c.font = SUB; c.fill = FILL_TITLE; c.alignment = LEFT

    # 헤더
    hr = start_row + 1
    headers = ['NO', '사양 / 품번', '일 소요량', '단위', '비고']
    for i, h in enumerate(headers, 1):
        cc = ws.cell(hr, i, h)
        cc.font = BOLD; cc.fill = FILL_HEAD; cc.alignment = CENTER; cc.border = BOX

    # 데이터
    qty_first = hr + 1
    for idx, spec in enumerate(items, 1):
        r = hr + idx
        ws.cell(r, 1, idx).alignment = CENTER
        ws.cell(r, 2, spec).alignment = LEFT
        formula = (
            f"=SUMIFS("
            f"INDIRECT(\"'출력용'!K\"&$H$1&\":K1000\"),"
            f"INDIRECT(\"'출력용'!{sumifs_match_col}\"&$H$1&\":{sumifs_match_col}1000\"),"
            f"B{r})"
        )
        ws.cell(r, 3, formula).alignment = RIGHT
        ws.cell(r, 3).number_format = '#,##0'
        ws.cell(r, 4, 'EA').alignment = CENTER
        ws.cell(r, 5, '').alignment = LEFT
        for col in range(1,6):
            ws.cell(r, col).border = BOX

    qty_last = hr + len(items)
    sub_row = qty_last + 1
    # 소계
    ws.cell(sub_row, 1, '소계').alignment = CENTER
    ws.merge_cells(start_row=sub_row, start_column=1, end_row=sub_row, end_column=2)
    ws.cell(sub_row, 3, f'=SUM(C{qty_first}:C{qty_last})').alignment = RIGHT
    ws.cell(sub_row, 3).number_format = '#,##0'
    for col in range(1,6):
        ws.cell(sub_row, col).font = BOLD
        ws.cell(sub_row, col).fill = FILL_SUB
        ws.cell(sub_row, col).border = BOX
    return sub_row  # 소계 row 반환

# ---------- 섹션 1: 홀더CLR (기본) ----------
holder_sub_row = write_section(start_row=6, title='홀더CLR 사양별 일소요량',
                               items=HOLDER, sumifs_match_col='R')

# ---------- 섹션 2: 흰색 (CLR흰색/NLR흰색) — 외주업체 별도 ----------
white_start = holder_sub_row + 2
white_sub_row = write_section(start_row=white_start,
                              title='흰색 사양 (CLR 흰색 / NLR 흰색) — 외주업체 별도',
                              items=HOLDER_WHITE, sumifs_match_col='R')

# ---------- 섹션 3: 센스커버 (MDS 사양과 동일 수량) ----------
cover_title_row = white_sub_row + 2
ws.merge_cells(start_row=cover_title_row, start_column=1, end_row=cover_title_row, end_column=5)
t = ws.cell(cover_title_row, 1, '■ 센스커버 (MDS 사양 동반 발주)')
t.font = SUB; t.fill = FILL_TITLE; t.alignment = LEFT

cover_head_row = cover_title_row + 1
for i, h in enumerate(['NO', '품목', '일 소요량', '단위', '비고'], 1):
    cc = ws.cell(cover_head_row, i, h)
    cc.font = BOLD; cc.fill = FILL_HEAD; cc.alignment = CENTER; cc.border = BOX

cover_row = cover_head_row + 1
ws.cell(cover_row, 1, 1).alignment = CENTER
ws.cell(cover_row, 2, '센스커버 (MDS용)').alignment = LEFT
# 홀더앗세이 컬럼에 '(MDS)' 포함된 모든 사양 합계
ws.cell(cover_row, 3,
        '=SUMIFS('
        'INDIRECT("\'출력용\'!K"&$H$1&":K1000"),'
        'INDIRECT("\'출력용\'!R"&$H$1&":R1000"),'
        '"*(MDS)*")'
       ).alignment = RIGHT
ws.cell(cover_row, 3).number_format = '#,##0'
ws.cell(cover_row, 4, 'EA').alignment = CENTER
ws.cell(cover_row, 5, 'MDS 사양 수량과 동일').alignment = LEFT
for col in range(1,6):
    ws.cell(cover_row, col).border = BOX

cover_sub_row = cover_row + 1
ws.cell(cover_sub_row, 1, '소계').alignment = CENTER
ws.merge_cells(start_row=cover_sub_row, start_column=1, end_row=cover_sub_row, end_column=2)
ws.cell(cover_sub_row, 3, f'=C{cover_row}').alignment = RIGHT
ws.cell(cover_sub_row, 3).number_format = '#,##0'
for col in range(1,6):
    ws.cell(cover_sub_row, col).font = BOLD
    ws.cell(cover_sub_row, col).fill = FILL_SUB
    ws.cell(cover_sub_row, col).border = BOX

# ---------- 섹션 3: 10파이 센서 ----------
gap_row = cover_sub_row + 2
sensor_sub_row = write_section(start_row=gap_row, title='10파이 센서 사양별 일소요량',
                               items=SENSOR, sumifs_match_col='P')

# ---------- 총계 ----------
total_row = sensor_sub_row + 2
ws.cell(total_row, 1, '총 계').alignment = CENTER
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
ws.cell(total_row, 3, f'=C{holder_sub_row}+C{white_sub_row}+C{cover_sub_row}+C{sensor_sub_row}').alignment = RIGHT
ws.cell(total_row, 3).number_format = '#,##0'
for col in range(1,6):
    ws.cell(total_row, col).font = Font(bold=True, size=12)
    ws.cell(total_row, col).fill = FILL_TITLE
    ws.cell(total_row, col).border = BOX
    if col != 3:
        ws.cell(total_row, col).font = Font(bold=True, size=12, color='FFFFFF')

# ---------- 열폭/행높이 ----------
widths = {1: 6, 2: 38, 3: 14, 4: 8, 5: 22}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[1].height = 28
# 사양 행 (개행 포함) 높이 자동 - 32
for r_idx in range(7, total_row+1):
    if ws.cell(r_idx, 2).value and '\n' in str(ws.cell(r_idx, 2).value or ''):
        ws.row_dimensions[r_idx].height = 30

# 시트탭 색
ws.sheet_properties.tabColor = 'C00000'

# 인쇄 영역
ws.print_area = f'A1:E{total_row}'
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True

# 저장
wb.save(SRC)
print(f'[save] {SRC.name}')
print(f'  holder_sub_row=C{holder_sub_row}, sensor_sub_row=C{sensor_sub_row}, total_row=C{total_row}')
print(f'  sheets after = {len(wb.sheetnames)}: {wb.sheetnames}')
