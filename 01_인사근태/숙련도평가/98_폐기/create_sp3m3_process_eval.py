"""
SP3M3 공정별 작업자 숙련도 평가서 생성 스크립트
- 출처: 작업표준서 SP3M3_작업표준서_200730_R1.xlsx
- 배점: 100점 만점 (평가준비20 + 전문강화45 + 기술15 + 생산성10 + 3정5행10)
- 레벨: Lv.4(90+) / Lv.3(80~89) / Lv.2(70~79) / Lv.1(60~69) / 불합격(<60)
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 경로 ───────────────────────────────────────────────────────────────────
BASE_DIR = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
OUTPUT   = os.path.join(BASE_DIR, 'SP3M3_공정별 숙련도 평가서_26년 2분기.xlsx')

# ── 스타일 ─────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

FILL_TITLE   = fill('1F4E79')
FILL_SECTION = fill('BDD7EE')
FILL_EXPERT  = fill('FFF2CC')
FILL_TOTAL   = fill('D6E4F0')

FONT_TITLE   = Font(name='맑은 고딕', bold=True, size=16, color='FFFFFF')
FONT_SECTION = Font(name='맑은 고딕', bold=True, size=10)
FONT_ITEM    = Font(name='맑은 고딕', size=10)
FONT_HEADER  = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
FONT_TOTAL   = Font(name='맑은 고딕', bold=True, size=11)

ALN_C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALN_L  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
BORDER = border()

COL_W = {'A': 12, 'B': 48, 'C': 18, 'D': 7, 'E': 10, 'F': 22}

# ── 공통 평가 항목 정의 ─────────────────────────────────────────────────────
COMMON = {
    '평가준비': {
        'total': 20,
        'method': '육안 확인',
        'items': [
            '작업 전 정리정돈 및 청소 상태',
            '일일 점검표 확인 및 서명',
            '필요 부품 및 자재 사전 준비',
            '바코드/라벨 출력 및 이상 여부 확인',
        ],
    },
    '기술': {
        'total': 15,
        'method': '구술/실연',
        'items': [
            '양품/불량품 판별 능력',
            '부적합품 발생 시 처리 절차 숙지',
            '공정 이상 발생 시 대처 능력',
        ],
    },
    '생산성': {
        'total': 10,
        'method': '실적 확인',
        'items': [
            '작업표준서 활용 상태',
            '시간당 목표 생산량 달성 여부',
        ],
    },
    '3정5행': {
        'total': 10,
        'method': '현장 확인',
        'items': [
            '공구/치공구 정위치 보관',
            '자재 및 완성품 정위치 관리',
        ],
    },
}
EXPERT_TOTAL = 45  # 전문강화 고정 배점

# ── 공정 데이터 (작업표준서 자주검사 항목 기반) ────────────────────────────
PROCESSES = [
    {
        'no': '10',
        'name': '프레임 앗세이 로딩 & 스풀 로딩 & 바코드 부착',
        'level': 'Lv.3',
        'sheet': '공정10',
        'src': '10_20',
        'expert': [
            ('프레임 이종 여부 확인 (차종별/초물표 참조)',       '바코드 자동검사'),
            ('스풀 조립 상태 확인 (이종/누락/정방향)',           '자동검사기(센서)'),
            ('어퍼스테이 이종 여부 확인',                       '자동검사기(센서)'),
            ('사양 확인 및 바코드 부착 상태',                   '육안 확인'),
        ],
    },
    {
        'no': '11',
        'name': 'WEBBING GUIDE 앗세이 조립',
        'level': 'Lv.3',
        'sheet': '공정11',
        'src': '10_20 (웨빙가이드 관련)',
        'expert': [
            ('웨빙가이드 조립 상태 확인 (흔들림/유격 여부)',     '육안 확인'),
            ('웨빙가이드 사양(차종별) 확인 및 이종 여부',        '바코드/조견표'),
            ('웨빙가이드 프레임 안착 방향 확인',                 '육안 확인'),
        ],
    },
    {
        'no': '91',
        'name': 'LOCKING ASSY 조립',
        'level': 'Lv.4',
        'sheet': '공정91',
        'src': '90 + 90-1',
        'expert': [
            # 시트 90 (5개)
            ('샤프트 베어링 부시 공급 상태 (누락 여부)',          '전수 육안'),
            ('락킹 ASS\'Y 공급 상태 (누락 여부)',                '전수 육안'),
            ('락킹 ASS\'Y 조립 상태 (정위치 여부)',              '전수 육안'),
            ('토션 사양 검사 (각인 상태)',                        '전수 육안'),
            ('리턴스프링 조립상태 (유무 여부)',                   '전수 육안'),
            # 시트 90-1 (10개)
            ('토션 공급상태 (누락, 방향성)',                      '전수 육안'),
            ('토션 구경 1차 검사',                               '전수 육안'),
            ('토션, 베이스락 압입상태 확인',                     '높이 게이지'),
            ('베이스락 공급상태 (누락 여부)',                     '전수 육안'),
            ('락플레이트 공급상태 (누락 여부)',                   '전수 육안'),
            ('락플레이트 조립상태 확인',                          '전수 육안'),
            ('커버 플레이트 락 조립상태 확인',                   '전수 육안'),
            ('리턴스프링 조립상태 재확인',                        '전수 육안'),
            ('커버 플레이트 락, 베이스락 압입상태 확인',          '높이 게이지'),
            ('압입청결상태 (이물질 확인)',                        '육안 확인'),
        ],
    },
    {
        'no': '140',
        'name': '볼 가이드 조립 & 파이프 조립 & 플레이트 커버 조립',
        'level': 'Lv.4',
        'sheet': '공정140',
        'src': '130 + 130-1',
        'expert': [
            ('파이프 조립 상태 확인',                            '육안 확인'),
            ('볼 가이드 ASS\'Y 공급 상태 (누락 여부)',           '전수 육안'),
            ('볼 가이드 이너 조립 상태 확인',                    '육안 확인'),
            ('플레이트 커버 공급 상태 (누락 여부)',               '전수 육안'),
            ('플레이트 커버 조립 상태 확인',                     '육안 확인'),
            ('볼가이드 안착면 청결 상태 유지',                   '육안 확인'),
        ],
    },
    {
        'no': '340',
        'name': 'VEHICLE SENSOR & 센서 커버 & 센스 브라켓 조립',
        'level': 'Lv.4',
        'sheet': '공정340',
        'src': '330',
        'expert': [
            ('센스 커버 아웃터, 비클 센서 이너 ASS\'Y 조립 상태', '육안 확인'),
            ('비클 센서, 비클 센서 이너 정사양 및 조립 방향 확인', '육안/조견표'),
            ('센스 커버 아웃터 ASS\'Y 프레임 조립 상태',          '육안 확인'),
            ('비클 센서 ASS\'Y 조립 상태 확인',                   '육안 확인'),
            ('비클 센서 PART 부품 청결 상태',                     '육안 확인'),
            ('센스 커버 아웃터 커버 핀 상태 확인',                '육안 확인'),
        ],
    },
    {
        'no': '430',
        'name': '프레임 앗세이 & 최종 검사 포장',
        'level': 'Lv.4',
        'sheet': '공정430',
        'src': '400',
        'expert': [
            ('스풀 조립 상태 최종 확인',                         '육안 확인'),
            ('부품 조립 외관 상태 확인',                         '육안 확인'),
            ('포장 박스 청결도 상태 확인',                       '육안 확인'),
            ('포장 상태 확인 (규격/수량)',                        '수량 검수'),
        ],
    },
]

# ── 유틸 ───────────────────────────────────────────────────────────────────
def distribute(total, n):
    """total점을 n개에 균등 배분. 나머지는 앞 항목에 +1씩."""
    base, rem = divmod(total, n)
    return [base + (1 if i < rem else 0) for i in range(n)]

def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c

def apply_row_border(ws, row, ncols=6):
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER

def write_section(ws, row, section_name, total, method, items, section_fill):
    """섹션 행 그룹 작성. 반환값: (다음 row, score_cells 리스트)"""
    start = row
    score_cells = []
    per_scores = distribute(total, len(items))

    for score, item in zip(per_scores, items):
        # B: 평가사항
        set_cell(ws, row, 2, item,   FONT_ITEM, align=ALN_L, border=BORDER)
        # C: 확인방법
        set_cell(ws, row, 3, method, FONT_ITEM, align=ALN_C, border=BORDER)
        # D: 배점
        set_cell(ws, row, 4, score,  FONT_ITEM, align=ALN_C, border=BORDER)
        # E: 평가점수 (입력칸)
        ws.cell(row=row, column=5).alignment = ALN_C
        ws.cell(row=row, column=5).border    = BORDER
        # F: 비고
        ws.cell(row=row, column=6).border    = BORDER
        ws.row_dimensions[row].height = 22
        score_cells.append(f'E{row}')
        row += 1

    # A 구분 셀 병합
    if row - 1 > start:
        ws.merge_cells(f'A{start}:A{row-1}')
    label = f'{section_name}\n({total}점)'
    c = ws.cell(row=start, column=1, value=label)
    c.font      = FONT_SECTION
    c.fill      = section_fill
    c.alignment = ALN_C
    c.border    = BORDER

    return row, score_cells

# ── 공정 시트 생성 ──────────────────────────────────────────────────────────
def make_process_sheet(wb, proc):
    ws = wb.create_sheet(title=proc['sheet'])
    for col_letter, width in COL_W.items():
        ws.column_dimensions[col_letter].width = width

    r = 1

    # ─ 타이틀
    ws.merge_cells(f'A{r}:F{r}')
    c = ws.cell(row=r, column=1, value='SP3M3 공정별 작업자 숙련도 평가서')
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = ALN_C
    ws.row_dimensions[r].height = 35
    r += 1

    # ─ 공정 정보 (2행)
    info = [
        [('업체', '대원테크'), ('라인', 'SP3M3'), ('공정번호', proc['no'])],
        [('공정명', proc['name'], True), ('요구레벨', proc['level'])],
    ]
    # 행1: 업체 / 라인 / 공정번호
    for col, (label, val) in enumerate(info[0], 1):
        label_col = col * 2 - 1
        val_col   = col * 2
        set_cell(ws, r, label_col, label, FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
        set_cell(ws, r, val_col,   val,   FONT_ITEM,   None,         ALN_C, BORDER)
    ws.row_dimensions[r].height = 22
    r += 1

    # 행2: 공정명(span) / 요구레벨
    set_cell(ws, r, 1, '공정명', FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
    ws.merge_cells(f'B{r}:D{r}')
    set_cell(ws, r, 2, proc['name'], FONT_ITEM, None, ALN_L, BORDER)
    set_cell(ws, r, 5, '요구레벨',   FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
    set_cell(ws, r, 6, proc['level'],
             Font(name='맑은 고딕', bold=True, size=11), None, ALN_C, BORDER)
    ws.row_dimensions[r].height = 22
    r += 1

    # ─ 작업자 정보
    set_cell(ws, r, 1, '작업자명', FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
    ws.merge_cells(f'B{r}:C{r}')
    ws.cell(row=r, column=2).border = BORDER
    set_cell(ws, r, 4, '사번', FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
    ws.merge_cells(f'E{r}:F{r}')
    ws.cell(row=r, column=5).border = BORDER
    ws.row_dimensions[r].height = 22
    r += 1

    # ─ 평가항목 헤더
    r += 1
    headers = ['구분', '평가사항', '확인방법', '배점', '평가점수', '비고']
    for col, h in enumerate(headers, 1):
        fill_h = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        set_cell(ws, r, col, h, FONT_HEADER, fill_h, ALN_C, BORDER)
    ws.row_dimensions[r].height = 25
    r += 1

    all_score_cells = []

    # ─ 평가준비 (공통)
    sec = COMMON['평가준비']
    r, sc = write_section(ws, r, '평가준비', sec['total'],
                          sec['method'], sec['items'], FILL_SECTION)
    all_score_cells.extend(sc)

    # ─ 전문강화 (공정별)
    expert_items  = [it[0] for it in proc['expert']]
    expert_method = [it[1] for it in proc['expert']]
    n = len(expert_items)
    scores = distribute(EXPERT_TOTAL, n)

    start = r
    for item, method, score in zip(expert_items, expert_method, scores):
        set_cell(ws, r, 2, item,   FONT_ITEM, align=ALN_L, border=BORDER)
        set_cell(ws, r, 3, method, FONT_ITEM, align=ALN_C, border=BORDER)
        set_cell(ws, r, 4, score,  FONT_ITEM, align=ALN_C, border=BORDER)
        ws.cell(row=r, column=5).alignment = ALN_C
        ws.cell(row=r, column=5).border    = BORDER
        ws.cell(row=r, column=6).border    = BORDER
        ws.row_dimensions[r].height = 22
        all_score_cells.append(f'E{r}')
        r += 1

    if r - 1 > start:
        ws.merge_cells(f'A{start}:A{r-1}')
    c = ws.cell(row=start, column=1,
                value=f'전문강화\n({EXPERT_TOTAL}점)\n[{n}항목]')
    c.font = Font(name='맑은 고딕', bold=True, size=10)
    c.fill = FILL_EXPERT; c.alignment = ALN_C; c.border = BORDER

    # ─ 기술 / 생산성 / 3정5행 (공통)
    for name in ('기술', '생산성', '3정5행'):
        sec = COMMON[name]
        r, sc = write_section(ws, r, name, sec['total'],
                              sec['method'], sec['items'], FILL_SECTION)
        all_score_cells.extend(sc)

    # ─ 합계 & 레벨 판정
    r += 1
    sum_formula   = '=SUM(' + ','.join(all_score_cells) + ')'
    level_formula = (
        f'=IF(E{r}>=90,"Lv.4",'
        f'IF(E{r}>=80,"Lv.3",'
        f'IF(E{r}>=70,"Lv.2",'
        f'IF(E{r}>=60,"Lv.1","불합격"))))'
    )
    ws.merge_cells(f'A{r}:C{r}')
    set_cell(ws, r, 1, '합계 (100점)', FONT_TOTAL, FILL_TOTAL, ALN_C, BORDER)
    set_cell(ws, r, 4, 100,            FONT_TOTAL, FILL_TOTAL, ALN_C, BORDER)
    set_cell(ws, r, 5, sum_formula,    FONT_TOTAL, FILL_TOTAL, ALN_C, BORDER)
    set_cell(ws, r, 6, level_formula,
             Font(name='맑은 고딕', bold=True, size=12), FILL_TOTAL, ALN_C, BORDER)
    ws.row_dimensions[r].height = 28
    r += 2

    # ─ 레벨 기준 안내
    ws.merge_cells(f'A{r}:F{r}')
    note = '[ 레벨 기준 ]  Lv.4: 90점 이상   Lv.3: 80~89점   Lv.2: 70~79점   Lv.1: 60~69점   불합격: 60점 미만'
    c = ws.cell(row=r, column=1, value=note)
    c.font = Font(name='맑은 고딕', italic=True, size=9, color='595959')
    c.alignment = ALN_L

# ── 종합 시트 생성 ──────────────────────────────────────────────────────────
def make_summary_sheet(wb):
    ws = wb.create_sheet(title='종합', index=0)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    r = 1
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value='SP3M3 공정별 숙련도 평가서 — 종합 현황 (26년 2분기)')
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = ALN_C
    ws.row_dimensions[r].height = 35
    r += 2

    headers = ['공정번호', '공정명', '요구레벨', '전문강화 항목수', '비고']
    for col, h in enumerate(headers, 1):
        set_cell(ws, r, col, h, FONT_SECTION, FILL_SECTION, ALN_C, BORDER)
    ws.row_dimensions[r].height = 25
    r += 1

    for proc in PROCESSES:
        set_cell(ws, r, 1, proc['no'],             FONT_ITEM, align=ALN_C, border=BORDER)
        set_cell(ws, r, 2, proc['name'],            FONT_ITEM, align=ALN_L, border=BORDER)
        set_cell(ws, r, 3, proc['level'],           FONT_ITEM, align=ALN_C, border=BORDER)
        set_cell(ws, r, 4, len(proc['expert']),     FONT_ITEM, align=ALN_C, border=BORDER)
        set_cell(ws, r, 5, proc['src'],
                 Font(name='맑은 고딕', size=9, color='595959'), align=ALN_C, border=BORDER)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 2
    ws.merge_cells(f'A{r}:E{r}')
    note = '※ 전문강화 배점 = 45점 균등배분 / 공통항목: 평가준비20 + 기술15 + 생산성10 + 3정5행10 = 55점 / 합계 100점'
    c = ws.cell(row=r, column=1, value=note)
    c.font = Font(name='맑은 고딕', italic=True, size=9, color='595959')
    c.alignment = ALN_L

# ── 메인 ───────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    make_summary_sheet(wb)
    for proc in PROCESSES:
        make_process_sheet(wb, proc)
        n = len(proc['expert'])
        scores = distribute(EXPERT_TOTAL, n)
        print(f"  [{proc['sheet']}] 전문강화 {n}항목 | 배점: {scores} | 합={sum(scores)}점")

    wb.save(OUTPUT)
    print(f"\n[완료] 저장: {OUTPUT}")

if __name__ == '__main__':
    main()
