"""
SP3M3 공정별 숙련도 평가서 생성 (샘플 양식 기반)
- 샘플 파일 복사 → 시트 6개 (공정별)
- 변경: 공정명/번호 + 2번 항목(전문강화, rows 13-21, col E) 만 수정
"""
import os
from openpyxl import load_workbook

SAMPLE = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가\공정 평가서표준_SP3M3_샘플.xlsx'
OUTPUT = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가\SP3M3_공정별 숙련도 평가서_26년 2분기_v2.xlsx'

# ── 공정별 전문강화 항목 (작업표준서 기반, 최대 9개) ────────────────────────
PROCESSES = [
    {
        'no': '10',
        'name': '프레임 앗세이 로딩 & 스풀 로딩 & 바코드 부착',
        'sheet': '공정10',
        'expert': [
            '프레임 이종 여부를 확인하고 있는가? (차종별/초물표 참조)',
            '스풀 조립 상태를 확인하고 있는가? (이종/누락/정방향)',
            '어퍼스테이 이종 여부를 확인하고 있는가?',
            '사양 확인 후 바코드를 정확히 부착하고 있는가?',
        ],
    },
    {
        'no': '11',
        'name': 'WEBBING GUIDE 앗세이 조립',
        'sheet': '공정11',
        'expert': [
            '웨빙가이드 조립 상태를 확인하고 있는가? (흔들림/유격 여부)',
            '웨빙가이드 차종별 사양 및 이종 여부를 확인하고 있는가?',
            '웨빙가이드 프레임 안착 방향을 확인하고 있는가?',
        ],
    },
    {
        'no': '91',
        'name': 'LOCKING ASSY 조립',
        'sheet': '공정91',
        'expert': [
            '샤프트 베어링 부시 공급 상태를 확인하고 있는가? (누락 여부)',
            '락킹 ASS\'Y 공급 상태를 확인하고 있는가? (누락 여부)',
            '락킹 ASS\'Y 조립 상태를 확인하고 있는가? (정위치 여부)',
            '토션 사양을 검사하고 있는가? (각인 상태)',
            '리턴스프링 조립상태를 확인하고 있는가? (유무 여부)',
            '토션, 베이스락 압입상태를 확인하고 있는가? (기준: 0±0.05mm)',
            '커버 플레이트 락, 베이스락 압입상태를 확인하고 있는가?',
            '락플레이트 조립상태를 확인하고 있는가?',
            '압입 후 청결상태를 확인하고 있는가? (이물질)',
        ],
    },
    {
        'no': '140',
        'name': '볼 가이드 조립 & 파이프 조립 & 플레이트 커버 조립',
        'sheet': '공정140',
        'expert': [
            '파이프 조립 상태를 확인하고 있는가?',
            '볼 가이드 ASS\'Y 공급 상태를 확인하고 있는가? (누락 여부)',
            '볼 가이드 이너 조립 상태를 확인하고 있는가?',
            '플레이트 커버 공급 상태를 확인하고 있는가? (누락 여부)',
            '플레이트 커버 조립 상태를 확인하고 있는가?',
            '볼가이드 안착면 청결 상태를 유지하고 있는가?',
        ],
    },
    {
        'no': '340',
        'name': 'VEHICLE SENSOR & 센서 커버 & 센스 브라켓 조립',
        'sheet': '공정340',
        'expert': [
            '센스 커버 아웃터, 비클 센서 이너 ASS\'Y 조립 상태를 확인하고 있는가?',
            '비클 센서, 비클 센서 이너 정사양 및 조립 방향을 확인하고 있는가?',
            '센스 커버 아웃터 ASS\'Y 프레임 조립 상태를 확인하고 있는가?',
            '비클 센서 ASS\'Y 조립 상태를 확인하고 있는가?',
            '비클 센서 부품 청결 상태를 유지하고 있는가?',
            '센스 커버 아웃터 커버 핀 상태를 확인하고 있는가?',
        ],
    },
    {
        'no': '430',
        'name': '프레임 앗세이 & 최종 검사 포장',
        'sheet': '공정430',
        'expert': [
            '스풀 조립 상태를 최종 확인하고 있는가?',
            '부품 조립 외관 상태를 확인하고 있는가?',
            '포장 박스 청결도 상태를 확인하고 있는가?',
            '포장 상태(규격/수량)를 확인하고 있는가?',
        ],
    },
]

# ── 셀 위치 (샘플 분석 결과) ──────────────────────────────────────────────
# 첫 번째 공정 열 그룹
COL_PROC1_NO   = 14   # N5  — 공정번호
COL_PROC1_NAME = 19   # S5  — 공정명
COL_PROC1_ITEM = 5    # E   — 전문강화 항목 텍스트 (rows 13-21)

# 두 번째 공정 열 그룹 (지우기 대상)
COL_PROC2_NO   = 45   # AS5
COL_PROC2_NAME = 50   # AX5
COL_PROC2_ITEM = 36   # AJ  (rows 13-21)

# 세 번째 공정 열 그룹 (지우기 대상)
COL_PROC3_NO   = 76   # BX5
COL_PROC3_NAME = 81   # CC5
COL_PROC3_ITEM = 67   # BO  (rows 13-21)

EXPERT_ROW_START = 13
EXPERT_ROW_END   = 21
PROC_INFO_ROW    = 5


def safe_write(ws, row, col, value):
    """머지 셀이면 top-left 셀 찾아서 쓰기"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def fill_process_sheet(ws, proc):
    # 공정번호 / 공정명 업데이트
    safe_write(ws, PROC_INFO_ROW, COL_PROC1_NO,   proc['no'])
    safe_write(ws, PROC_INFO_ROW, COL_PROC1_NAME, proc['name'])

    # 전문강화 항목 채우기 (2번 항목)
    for i in range(EXPERT_ROW_END - EXPERT_ROW_START + 1):
        row = EXPERT_ROW_START + i
        text = proc['expert'][i] if i < len(proc['expert']) else ''
        safe_write(ws, row, COL_PROC1_ITEM, text)

    # 2·3번째 공정 열 지우기
    for col_no, col_name, col_item in [
        (COL_PROC2_NO, COL_PROC2_NAME, COL_PROC2_ITEM),
        (COL_PROC3_NO, COL_PROC3_NAME, COL_PROC3_ITEM),
    ]:
        safe_write(ws, PROC_INFO_ROW, col_no,   '')
        safe_write(ws, PROC_INFO_ROW, col_name, '')
        for i in range(EXPERT_ROW_END - EXPERT_ROW_START + 1):
            safe_write(ws, EXPERT_ROW_START + i, col_item, '')


def main():
    wb = load_workbook(SAMPLE)
    template = wb['양식']

    for proc in PROCESSES:
        ws = wb.copy_worksheet(template)
        ws.title = proc['sheet']
        fill_process_sheet(ws, proc)
        print(f'  [{proc["sheet"]}] {proc["no"]}공정 전문강화 {len(proc["expert"])}항목')

    del wb['양식']
    wb.save(OUTPUT)
    print(f'\n[완료] {OUTPUT}')


if __name__ == '__main__':
    main()
