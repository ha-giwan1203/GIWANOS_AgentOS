"""
SP3M3 공정별 숙련도 평가서 — 공정별 파일 분리 버전 (v3)
- 샘플 파일 복사 → 공정별 개별 .xlsx 파일 6개 생성
- 전문강화 섹션:
    rows 13~17 (5개): 작업표준서 기준 항목
    rows 18~21 (4개): 관리계획서 기준 항목
- 변경 대상: 공정명/번호 + 전문강화 항목 텍스트만 (기타 셀 유지)
"""
import os
import shutil
from openpyxl import load_workbook

BASE    = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
SAMPLE  = os.path.join(BASE, '공정 평가서표준_SP3M3_샘플.xlsx')
OUT_DIR = os.path.join(BASE, 'SP3M3_공정별 평가서')

# ── 공정별 전문강화 항목 (작업표준서 5개 + 관리계획서 4개) ─────────────────
PROCESSES = [
    {
        'no': '10',
        'name': '프레임 앗세이 로딩 & 스풀 로딩 & 바코드 부착',
        'std': [  # 작업표준서 기준 (5개, rows 13~17)
            '프레임 이종 여부를 확인하고 있는가? (차종별/초물표 참조)',
            '스풀 조립 상태를 확인하고 있는가? (이종/누락/정방향)',
            '어퍼스테이 이종 여부를 확인하고 있는가?',
            '웨빙가이드 조립 상태를 확인하고 있는가? (흔들림/유격)',
            '프레임 메인 컨베이어 안착 및 이송을 정확히 수행하는가?',
        ],
        'ctrl': [  # 관리계획서 기준 (4개, rows 18~21)
            '바코드 스티커 사양이 맞는지 확인하고 있는가?',
            '바코드 스티커 방향이 올바른지 확인하고 있는가?',
            '바코드 스티커 부착상태를 제대로 확인하고 있는가?',
            '프레임 이종을 관리기준으로 확인하고 있는가?',
        ],
    },
    {
        'no': '11',
        'name': 'WEBBING GUIDE 앗세이 조립',
        'std': [
            '웨빙가이드 조립 상태를 확인하고 있는가? (흔들림/유격 여부)',
            '웨빙가이드 차종별 사양을 확인하고 있는가?',
            '웨빙가이드 프레임 안착 방향을 확인하고 있는가?',
            '웨빙가이드 이종 여부를 확인하고 있는가?',
            '웨빙가이드 조립 후 검사를 수행하고 있는가?',
        ],
        'ctrl': [
            '웨빙가이드 이종을 관리기준으로 확인하고 있는가?',
            '웨빙가이드 방향을 관리기준으로 확인하고 있는가?',
            '웨빙가이드 조립 상태를 주기적으로 점검하고 있는가?',
            '웨빙가이드 부적합품을 관리계획서 기준으로 처리하고 있는가?',
        ],
    },
    {
        'no': '91',
        'name': 'LOCKING ASSY 조립',
        'std': [
            '샤프트 베어링 부시 공급 상태를 확인하고 있는가? (누락 여부)',
            '락킹 ASS\'Y 조립 상태를 확인하고 있는가? (정위치 여부)',
            '토션 사양(각인 상태)을 검사하고 있는가?',
            '리턴스프링 조립상태를 확인하고 있는가?',
            '에어 압력 상태를 확인하고 있는가? (0.4~0.6 MPa)',
        ],
        'ctrl': [
            '홀더 휠 조립 상태를 관리기준으로 확인하고 있는가?',
            '락킹 ASS\'Y 공급 상태를 주기적으로 점검하고 있는가?',
            '토션/베이스락 압입상태를 관리기준으로 측정하는가? (0±0.05mm)',
            '압입 후 청결 상태(이물질)를 확인하고 있는가?',
        ],
    },
    {
        'no': '140',
        'name': '볼 가이드 조립 & 파이프 조립 & 플레이트 커버 조립',
        'std': [
            '파이프 조립 상태를 확인하고 있는가?',
            '볼 가이드 ASS\'Y 공급 상태를 확인하고 있는가? (누락 여부)',
            '볼 가이드 이너 조립 상태를 확인하고 있는가?',
            '플레이트 커버 공급 상태를 확인하고 있는가? (누락 여부)',
            '플레이트 커버 조립 상태를 확인하고 있는가?',
        ],
        'ctrl': [
            '파이프 조립 상태를 관리기준으로 확인하고 있는가?',
            '볼 가이드 ASS\'Y 조립 상태를 관리기준으로 확인하고 있는가?',
            '볼가이드 이너 조립 상태를 관리기준으로 확인하고 있는가?',
            '플레이트 커버 조립 상태를 관리기준으로 확인하고 있는가?',
        ],
    },
    {
        'no': '340',
        'name': 'VEHICLE SENSOR & 센서 커버 & 센스 브라켓 조립',
        'std': [
            '센스 커버 아웃터, 비클 센서 이너 ASS\'Y 조립 상태를 확인하는가?',
            '비클 센서 정사양 및 조립 방향을 확인하고 있는가?',
            '센스 커버 아웃터 ASS\'Y 프레임 조립 상태를 확인하고 있는가?',
            '비클 센서 ASS\'Y 조립 상태를 확인하고 있는가?',
            '비클 센서 부품 청결 상태를 유지하고 있는가?',
        ],
        'ctrl': [
            '센서 커버 아웃터 핀 압입 상태를 관리기준으로 확인하는가?',
            '비클 센서 볼 조립 상태를 관리기준으로 확인하고 있는가?',
            '센서 볼 마스터(OK/NG)를 시업시 확인하고 있는가?',
            '커버핀 압입 높이를 관리기준으로 확인하고 있는가?',
        ],
    },
    {
        'no': '430',
        'name': '프레임 앗세이 & 최종 검사 포장',
        'std': [
            '스풀 조립 상태를 최종 확인하고 있는가?',
            '부품 조립 외관 상태를 확인하고 있는가?',
            '포장 박스 청결도 상태를 확인하고 있는가?',
            '포장 상태(규격/수량)를 확인하고 있는가?',
            '최종 검사 체크리스트를 준수하고 있는가?',
        ],
        'ctrl': [
            '완제품 외관 검사를 관리기준으로 수행하고 있는가?',
            '포장 수량을 관리계획서 기준으로 확인하고 있는가?',
            '출하 라벨 부착 상태를 확인하고 있는가?',
            '출하 전 최종 검사 기록을 작성하고 있는가?',
        ],
    },
]

# ── 셀 좌표 (샘플 분석 결과) ──────────────────────────────────────────────
COL_PROC1_NO   = 14     # N5   공정번호
COL_PROC1_NAME = 19     # S5   공정명
COL_PROC1_ITEM = 5      # E    전문강화 항목 텍스트

COL_PROC2_NO, COL_PROC2_NAME, COL_PROC2_ITEM = 45, 50, 36     # AS, AX, AJ
COL_PROC3_NO, COL_PROC3_NAME, COL_PROC3_ITEM = 76, 81, 67     # BX, CC, BO

ROW_PROC_INFO = 5

# 전문강화 섹션 행 배치
ROWS_STD  = list(range(13, 18))    # 13~17 (5개) — 작업표준서
ROWS_CTRL = list(range(18, 22))    # 18~21 (4개) — 관리계획서


def safe_write(ws, row, col, value):
    """머지 셀이면 top-left 셀에 쓴다."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def clear_other_processes(ws):
    """2·3번째 공정 열을 모두 지워 단일 공정 레이아웃으로."""
    for col_no, col_name, col_item in [
        (COL_PROC2_NO, COL_PROC2_NAME, COL_PROC2_ITEM),
        (COL_PROC3_NO, COL_PROC3_NAME, COL_PROC3_ITEM),
    ]:
        safe_write(ws, ROW_PROC_INFO, col_no,   '')
        safe_write(ws, ROW_PROC_INFO, col_name, '')
        for r in ROWS_STD + ROWS_CTRL:
            safe_write(ws, r, col_item, '')


def fill_sheet(ws, proc):
    # 공정번호/공정명
    safe_write(ws, ROW_PROC_INFO, COL_PROC1_NO,   proc['no'])
    safe_write(ws, ROW_PROC_INFO, COL_PROC1_NAME, proc['name'])

    # 작업표준서 기준 (5개)
    for i, r in enumerate(ROWS_STD):
        val = proc['std'][i] if i < len(proc['std']) else ''
        safe_write(ws, r, COL_PROC1_ITEM, val)

    # 관리계획서 기준 (4개)
    for i, r in enumerate(ROWS_CTRL):
        val = proc['ctrl'][i] if i < len(proc['ctrl']) else ''
        safe_write(ws, r, COL_PROC1_ITEM, val)

    # 2·3번째 공정 열 지우기
    clear_other_processes(ws)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    for proc in PROCESSES:
        out_path = os.path.join(
            OUT_DIR,
            f"SP3M3_공정{proc['no']} 숙련도 평가서_26년 2분기.xlsx"
        )
        # 샘플 복사 후 수정
        shutil.copy(SAMPLE, out_path)
        wb = load_workbook(out_path)
        ws = wb['양식']
        fill_sheet(ws, proc)
        wb.save(out_path)
        wb.close()
        print(f"  [생성] 공정{proc['no']} → {os.path.basename(out_path)}")

    print(f"\n[완료] 출력 폴더: {OUT_DIR}")


if __name__ == '__main__':
    main()
