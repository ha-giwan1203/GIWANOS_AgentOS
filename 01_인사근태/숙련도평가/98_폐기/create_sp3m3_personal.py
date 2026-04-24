"""
SP3M3 공정별 숙련도 평가서 — 개인별 복사본 생성
- 출처: 기존 SP3M3 주간/야간 숙련도 평가서/ (개인 파일)
- 대상: 공정별 양식 템플릿 (SP3M3_공정별 평가서/)
- 결과: 작업자별 폴더 + 수행공정별 파일 (총 ~67개)
"""
import os
import shutil
from openpyxl import load_workbook

BASE = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
TEMPLATE_DIR = os.path.join(BASE, 'SP3M3_공정별 평가서')
OUT_DIR = os.path.join(BASE, 'SP3M3_개인별 평가서')

# 기존 개인 파일 (주간/야간)
SOURCE_DIRS = [
    (os.path.join(BASE, 'SP3M3 주간 숙련도 평가서'), '주간'),
    (os.path.join(BASE, 'SP3M3 야간 숙련도 평가서'), '야간'),
]

# 기존 파일 셀 위치
PROC_COLS  = [14, 46, 78, 110, 142, 174, 206, 238, 270, 302, 334]
TOTAL_COLS = [10, 42, 74, 106, 138, 170, 202, 234, 266, 298, 330]
LEVEL_COLS = [20, 52, 84, 116, 148, 180, 212, 244, 276, 308, 340]

# 공정번호 → 템플릿 파일명
TEMPLATE_MAP = {
    '10':  'SP3M3_공정10 숙련도 평가서.xlsx',
    '11':  'SP3M3_공정11 숙련도 평가서.xlsx',
    '91':  'SP3M3_공정91 숙련도 평가서.xlsx',
    '140': 'SP3M3_공정140 숙련도 평가서.xlsx',
    '340': 'SP3M3_공정340 숙련도 평가서.xlsx',
    '430': 'SP3M3_공정430 숙련도 평가서.xlsx',
}

# 새 양식 작업자 정보 셀 (이름/사번/직급/평가일)
CELL_NAME    = (2, 22)   # V2
CELL_EMP_ID  = (3, 22)   # V3
CELL_POS     = (3, 26)   # Z3
CELL_DATE    = (5, 28)   # AB5


def safe_write(ws, row, col, value):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def extract_worker(path, shift):
    """기존 개인 파일에서 작업자 정보 + 수행공정 리스트 추출.
       원본 V2 셀이 잘못 입력된 파일 대비, 이름은 파일명 기준."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # 파일명에서 이름 추출 (V2 셀보다 신뢰도 높음)
    name    = os.path.basename(path).split(' 숙련도')[0]
    emp_id  = ws.cell(3, 22).value          # V3
    pos     = ws.cell(3, 26).value          # Z3
    date    = ws.cell(5, 28).value          # AB5
    procs = []
    for pc, tc, lc in zip(PROC_COLS, TOTAL_COLS, LEVEL_COLS):
        pv = ws.cell(5, pc).value
        if pv is not None:
            procs.append({
                'no':    str(pv),
                'total': ws.cell(32, tc).value,
                'level': ws.cell(32, lc).value,
            })
    wb.close()
    return {
        'name':   name,
        'emp_id': emp_id,
        'pos':    pos,
        'date':   date,
        'shift':  shift,
        'procs':  procs,
    }


def create_personal_file(worker, proc):
    """개인 × 공정 파일 1개 생성"""
    template_path = os.path.join(TEMPLATE_DIR, TEMPLATE_MAP[proc['no']])
    if not os.path.exists(template_path):
        print(f'  [경고] 템플릿 없음: 공정{proc["no"]}')
        return None

    folder = os.path.join(OUT_DIR, f'{worker["shift"]}_{worker["name"]}')
    os.makedirs(folder, exist_ok=True)
    out_path = os.path.join(
        folder,
        f'{worker["name"]}_공정{proc["no"]} 숙련도평가서.xlsx'
    )

    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    ws = wb['양식']

    # 작업자 정보 이관
    safe_write(ws, *CELL_NAME,   worker['name'])
    safe_write(ws, *CELL_EMP_ID, worker['emp_id'])
    safe_write(ws, *CELL_POS,    worker['pos'])
    # 평가일 갱신 (기존 날짜 유지)
    if worker['date']:
        safe_write(ws, *CELL_DATE, worker['date'])

    # B33 "평 가 자 의 견" 라벨 원본 복원 + 참고점수 병기 (merged B33:F37)
    merged_text = (
        f'평 가 자 의 견\n\n'
        f'[직전 평가: {proc["total"]}점 / {proc["level"]}]\n'
        f'(기존 공통 양식 기준, 공정별 재평가 필요)'
    )
    safe_write(ws, 33, 2, merged_text)

    wb.save(out_path)
    wb.close()
    return out_path


def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    workers = []
    for src_dir, shift in SOURCE_DIRS:
        for f in sorted(os.listdir(src_dir)):
            if not f.endswith('.xlsx'):
                continue
            w = extract_worker(os.path.join(src_dir, f), shift)
            workers.append(w)

    print(f'[추출] 작업자 {len(workers)}명')

    file_count = 0
    for w in workers:
        procs_list = ', '.join([p['no'] for p in w['procs']])
        print(f'\n  [{w["shift"]}] {w["name"]} ({w["emp_id"]}) 공정: {procs_list}')
        for p in w['procs']:
            path = create_personal_file(w, p)
            if path:
                file_count += 1
                print(f'     - 공정{p["no"]} ({p["total"]}점/{p["level"]}) -> 생성')

    print(f'\n[완료] 총 {file_count}개 파일 생성 / 출력: {OUT_DIR}')


if __name__ == '__main__':
    main()
