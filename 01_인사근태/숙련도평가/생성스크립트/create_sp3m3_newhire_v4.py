# -*- coding: utf-8 -*-
"""
SP3M3 신규 입사자 교육자료 v4 — 공정당 A4 1~2장 (시트 1개)에
신규자가 알아야 할 내용을 한 번에 정리.
"""
from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import win32com.client as win32

BASE = Path(__file__).resolve().parent.parent
STD = BASE / "SP3M3_표준문서"
PROC_FORM = BASE / "SP3M3_공정별 평가서"
PERSONAL = BASE / "SP3M3_개인별 평가서"
OUTDIR = BASE / "SP3M3_신규입사자 교육자료"

xlOpenXMLWorkbook = 51

# 색상
C_WHITE = 2
C_RED = 3
C_BLUE = 5
C_YELLOW = 6
C_GRAY = 15
C_DARK_BLUE = 32
C_LIGHT_GREEN = 35
C_LIGHT_RED = 38
C_LIGHT_YELLOW = 36


# ─────────────────────────────────────────────
# 데이터 로드 + 작업표준서 시트 매핑
# ─────────────────────────────────────────────
def load_processes():
    wb = openpyxl.load_workbook(STD / "SP3M3 라인별공정목록.xlsx", data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        out.append({"no": int(no), "name": (ws.cell(r, 2).value or "").strip(),
                    "level": (ws.cell(r, 3).value or "Lv.3").strip()})
    wb.close()
    return out


def load_eval_forms():
    out = {}
    for fp in PROC_FORM.glob("SP3M3_공정*.xlsx"):
        m = re.search(r"공정(\d+)", fp.name)
        if not m:
            continue
        no = int(m.group(1))
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        items = [ws.cell(r, 5).value for r in range(13, 22)]
        criteria = [(ws.cell(r, 17).value, ws.cell(r, 21).value, ws.cell(r, 25).value)
                    for r in range(13, 22)]
        out[no] = {"items": items, "criteria": criteria}
        wb.close()
    return out


def load_personnel():
    out = defaultdict(list)
    for shift_dir, shift in [("SP3M3 주간", "주간"), ("SP3M3 야간", "야간")]:
        d = PERSONAL / shift_dir
        if not d.exists():
            continue
        for fp in d.glob("*.xlsx"):
            name = fp.stem.replace(" 숙련도평가서", "").strip()
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                n5 = ws["N5"].value
                if not isinstance(n5, (int, float)):
                    try:
                        n5 = int(str(n5))
                    except Exception:
                        continue
                ac_sum = sum(v for r in range(13, 22)
                             if isinstance(v := ws.cell(r, 29).value, (int, float)))
                out[int(n5)].append({"name": name, "shift": shift, "ac_sum": ac_sum})
            wb.close()
    return out


def find_ws_sheet_name(sheetnames, proc_no):
    exact, ranges, near = [], [], []
    for sn in sheetnames:
        nums = [int(n) for n in re.findall(r"\d+", sn)]
        if not nums:
            continue
        if proc_no in nums:
            exact.append(sn)
            continue
        if len(nums) >= 2 and (max(nums) - min(nums)) <= 30:
            if min(nums) <= proc_no <= max(nums):
                ranges.append((max(nums) - min(nums), sn))
                continue
        diff = min(abs(n - proc_no) for n in nums)
        if diff <= 5:
            near.append((diff, sn))
    if exact:
        return exact[0]
    if ranges:
        ranges.sort()
        return ranges[0][1]
    if near:
        near.sort()
        return near[0][1]
    return None


def load_workstd_steps(proc_no, sheetnames_cache):
    """작업표준서에서 해당 공정 시트의 자주검사·작업순서 텍스트만 추출."""
    wb = openpyxl.load_workbook(STD / "SP3M3_작업표준서_200730_R1.xlsx", data_only=True)
    sn = find_ws_sheet_name(wb.sheetnames, proc_no)
    steps, inspect, parts = [], [], set()
    if sn:
        ws = wb[sn]
        in_inspect, in_steps = False, False
        for r in range(1, min(ws.max_row + 1, 60)):
            m = ws.cell(r, 13).value
            if isinstance(m, str):
                if "자   주   검   사" in m or "자주검사" in m.replace(" ", ""):
                    in_inspect, in_steps = True, False
                    continue
                if "공   정   작   업   순   서" in m or "공정작업순서" in m.replace(" ", ""):
                    in_inspect, in_steps = False, True
                    continue
                if "조   건   관   리" in m or "조건관리" in m.replace(" ", ""):
                    in_inspect, in_steps = False, False
                    continue
            if in_inspect:
                item = ws.cell(r, 14).value
                std = ws.cell(r, 17).value
                if isinstance(item, str) and item.strip() and len(item) < 40:
                    inspect.append((item.strip(), (std or "").strip() if isinstance(std, str) else ""))
            if in_steps:
                v = ws.cell(r, 13).value
                if isinstance(v, str) and v.strip():
                    steps.append(v.strip())
    wb.close()
    return {"ref_sheet": sn, "steps": steps, "inspect": inspect}


def extract_parts_from_proc(proc, eval_data, ws_data):
    """공정명·평가항목·작업표준서에서 부품 명사 추출 (한글 명사 휴리스틱)."""
    text = proc["name"] + " "
    if eval_data:
        for it in eval_data["items"]:
            if isinstance(it, str):
                text += it + " "
    for s, _ in ws_data.get("inspect", []):
        text += s + " "
    # 한글 명사 후보 (영어 포함)
    known = ["프레임", "스풀", "웨빙가이드", "어퍼스테이", "바코드", "WEBBING GUIDE",
             "LOCKING", "ASSY", "볼 가이드", "파이프", "플레이트 커버", "VEHICLE SENSOR",
             "센서 커버", "센스 브라켓", "최종 검사", "포장", "릴", "스티커",
             "Locking", "Pawl", "Cam"]
    found = []
    for k in known:
        if k in text and k not in found:
            found.append(k)
    return found


# ─────────────────────────────────────────────
# 1장 빌더 (모든 내용 한 시트에)
# ─────────────────────────────────────────────
def _set_cell(ws, addr, val, *, size=11, bold=False, color=None, bg=None,
              hcenter=False, vcenter=True, wrap=True, border=True):
    c = ws.Range(addr)
    c.Value = val
    c.Font.Name = "맑은 고딕"
    c.Font.Size = size
    c.Font.Bold = bold
    if color is not None:
        c.Font.ColorIndex = color
    if bg is not None:
        c.Interior.ColorIndex = bg
    c.HorizontalAlignment = -4108 if hcenter else -4131
    if vcenter:
        c.VerticalAlignment = -4108
    if wrap:
        c.WrapText = True
    if border:
        c.Borders.LineStyle = 1


def _merge_set(ws, range_addr, val, **kwargs):
    ws.Range(range_addr).Merge()
    _set_cell(ws, range_addr.split(":")[0], val, **kwargs)


def build_one(ws, proc, eval_data, ws_data, parts, mentors):
    """A4 가로 1~2장에 들어가는 신규자 교육자료 1시트."""
    ws.Name = "교육자료"

    # 컬럼 폭 (A~H 8칸 = A4 가로 폭에 맞춤)
    widths = [3, 14, 17, 14, 14, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.Columns(i).ColumnWidth = w

    # === 0) 헤더 ===
    _merge_set(ws, "A1:H1",
               f"공정 {proc['no']}  —  {proc['name']}",
               size=20, bold=True, color=C_WHITE, bg=C_DARK_BLUE, hcenter=True)
    ws.Rows(1).RowHeight = 40

    _merge_set(ws, "A2:D2",
               f"라인: SP3M3 (안전벨트 ASS'Y)    /    요구레벨: {proc['level']}",
               size=12, bold=True, hcenter=True, bg=C_GRAY)
    _merge_set(ws, "E2:H2",
               f"신규자: ____________    사수: ____________    시작일: 26 / __ / __",
               size=11, hcenter=True, bg=C_LIGHT_YELLOW)
    ws.Rows(2).RowHeight = 26

    # === 1) 이 공정이 뭘 하는가 ===
    _merge_set(ws, "A3:H3", "① 이 공정은 무엇을 하는가",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(3).RowHeight = 28

    desc = (f"SP3M3 라인의 {proc['no']}번 공정으로, '{proc['name']}' 작업을 담당합니다.\n"
            f"이 라인은 자동차 안전벨트(Seat Belt Assembly)를 만듭니다. "
            f"단 1개의 이종·누락이 나가면 차량 1대가 리콜 대상이 될 수 있습니다.\n"
            f"→ 빠르게보다 정확하게. 의심되면 사수에게 즉시 확인.")
    _merge_set(ws, "A4:H5", desc, size=12)
    ws.Rows(4).RowHeight = 36
    ws.Rows(5).RowHeight = 36

    # === 2) 사용 부품 ===
    _merge_set(ws, "A6:H6", "② 사용 부품 (작업 전 이름 외우기)",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(6).RowHeight = 28

    parts_str = "  /  ".join(parts) if parts else "(작업표준서 참조)"
    _merge_set(ws, "A7:H7", parts_str, size=13, bold=True, hcenter=True, color=C_DARK_BLUE)
    ws.Rows(7).RowHeight = 36

    # === 3) 작업 순서 ===
    _merge_set(ws, "A8:H8", "③ 작업 순서 (작업표준서 상세 참조 — 현장 비치)",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(8).RowHeight = 28

    steps = ws_data.get("steps") or [
        "(작업표준서 매칭 누락 — 라인반장 보강 필요)"
    ]
    # 최대 8스텝만 핵심으로
    main_steps = [s for s in steps if not s.startswith("◆")][:8]
    if not main_steps:
        main_steps = steps[:8]
    r = 9
    for i, s in enumerate(main_steps, start=1):
        _set_cell(ws, f"A{r}", i, size=12, bold=True, hcenter=True, bg=C_GRAY)
        ws.Range(f"B{r}:H{r}").Merge()
        _set_cell(ws, f"B{r}", s, size=12)
        ws.Rows(r).RowHeight = 24
        r += 1
    if ws_data.get("ref_sheet"):
        ws.Range(f"A{r}:H{r}").Merge()
        _set_cell(ws, f"A{r}",
                  f"※ 자세한 작업 순서는 현장 작업표준서 시트 [{ws_data['ref_sheet']}] 참조 (사진·도면 포함)",
                  size=10, color=C_GRAY, hcenter=True, border=False)
        ws.Rows(r).RowHeight = 18
        r += 1

    # === 4) 합격 vs 불합격 핵심 5개 ===
    _merge_set(ws, f"A{r}:H{r}", "④ 합격(OK) vs 불합격(NG) — 핵심 확인 사항",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(r).RowHeight = 28
    r += 1

    _set_cell(ws, f"A{r}", "NO", size=12, bold=True, hcenter=True, bg=C_GRAY)
    ws.Range(f"B{r}:D{r}").Merge()
    _set_cell(ws, f"B{r}", "확인할 것", size=12, bold=True, hcenter=True, bg=C_GRAY)
    ws.Range(f"E{r}:F{r}").Merge()
    _set_cell(ws, f"E{r}", "✓ OK (이래야 합격)", size=12, bold=True, hcenter=True, bg=C_LIGHT_GREEN)
    ws.Range(f"G{r}:H{r}").Merge()
    _set_cell(ws, f"G{r}", "✗ NG (이러면 불량)", size=12, bold=True, hcenter=True, bg=C_LIGHT_RED)
    ws.Rows(r).RowHeight = 26
    r += 1

    if eval_data:
        items = eval_data["items"][:6]  # 핵심 6개만
        crits = eval_data["criteria"][:6]
        for i, (it, (q, _u, y)) in enumerate(zip(items, crits), start=1):
            _set_cell(ws, f"A{r}", i, size=12, bold=True, hcenter=True)
            ws.Range(f"B{r}:D{r}").Merge()
            _set_cell(ws, f"B{r}", it or "", size=11)
            ws.Range(f"E{r}:F{r}").Merge()
            _set_cell(ws, f"E{r}", q or "", size=11, bold=True, bg=C_LIGHT_GREEN)
            ws.Range(f"G{r}:H{r}").Merge()
            _set_cell(ws, f"G{r}", y or "", size=11, bold=True, bg=C_LIGHT_RED, color=C_RED)
            ws.Rows(r).RowHeight = 36
            r += 1

    # === 5) 꼭 지킬 것 (안전·이종) ===
    _merge_set(ws, f"A{r}:H{r}", "⑤ 꼭 지킬 것 — 안전·이종·일상점검",
               size=14, bold=True, color=C_WHITE, bg=C_RED, hcenter=True)
    ws.Rows(r).RowHeight = 28
    r += 1

    safety = [
        "🛡 시업 시 설비 일상점검표 작성 (에어압 0.4~0.6 MPa)",
        "🛡 차종 변경 시 초물표 확인 후 첫 5개는 사수가 검사",
        "🛡 면장갑·보안경 착용 — 회전부·압입부에 손가락 진입 금지",
        "🛡 이상 발생 시 즉시 작업 중지 후 관리자 보고",
        "🛡 작업 중 의심 시 라인 멈추는 게 NG 흘려보내는 것보다 낫다",
    ]
    for s in safety:
        ws.Range(f"A{r}:H{r}").Merge()
        _set_cell(ws, f"A{r}", s, size=12, bg=C_LIGHT_YELLOW)
        ws.Rows(r).RowHeight = 22
        r += 1

    # === 6) 자주 하는 실수 ===
    _merge_set(ws, f"A{r}:H{r}", "⑥ 신규자가 자주 하는 실수 (사수 작성)",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(r).RowHeight = 28
    r += 1

    # 평가항목 NG 기준에서 실수 사례 자동 채움 (3개)
    if eval_data:
        ng_cases = [(it, y) for it, (_q, _u, y) in zip(eval_data["items"], eval_data["criteria"]) if y][:3]
        for i, (it, y) in enumerate(ng_cases, start=1):
            _set_cell(ws, f"A{r}", i, size=12, bold=True, hcenter=True, bg=C_GRAY)
            ws.Range(f"B{r}:D{r}").Merge()
            _set_cell(ws, f"B{r}", it[:40] if it else "", size=11)
            ws.Range(f"E{r}:H{r}").Merge()
            _set_cell(ws, f"E{r}", f"(자주 발생) {y}", size=11, color=C_RED)
            ws.Rows(r).RowHeight = 30
            r += 1
    # 빈칸 2개 (사수 추가 기입용)
    for i in range(len(ng_cases) + 1 if eval_data else 1, 6):
        _set_cell(ws, f"A{r}", i, size=12, bold=True, hcenter=True, bg=C_GRAY)
        ws.Range(f"B{r}:D{r}").Merge()
        _set_cell(ws, f"B{r}", "", size=11)
        ws.Range(f"E{r}:H{r}").Merge()
        _set_cell(ws, f"E{r}", "(사수 추가 기입)", size=10, color=C_GRAY)
        ws.Rows(r).RowHeight = 28
        r += 1

    # === 7) 사수 + 학습 일정 ===
    _merge_set(ws, f"A{r}:H{r}", "⑦ 사수 배정 + 2주 학습 일정",
               size=14, bold=True, color=C_WHITE, bg=C_BLUE, hcenter=True)
    ws.Rows(r).RowHeight = 28
    r += 1

    # 사수 후보
    _set_cell(ws, f"A{r}", "사수", size=12, bold=True, hcenter=True, bg=C_GRAY)
    ws.Range(f"B{r}:D{r}").Merge()
    day_mentor = next(iter(sorted([m for m in mentors if m["shift"] == "주간"],
                                  key=lambda x: -x["ac_sum"])), None)
    night_mentor = next(iter(sorted([m for m in mentors if m["shift"] == "야간"],
                                    key=lambda x: -x["ac_sum"])), None)
    _set_cell(ws, f"B{r}",
              f"주간 1차 사수: {day_mentor['name'] if day_mentor else '미정'}"
              f"  ({day_mentor['ac_sum'] if day_mentor else '-'} / 45점)",
              size=12, bold=True, color=C_RED)
    ws.Range(f"E{r}:H{r}").Merge()
    _set_cell(ws, f"E{r}",
              f"야간 1차 사수: {night_mentor['name'] if night_mentor else '미정'}"
              f"  ({night_mentor['ac_sum'] if night_mentor else '-'} / 45점)",
              size=12, bold=True, color=C_RED)
    ws.Rows(r).RowHeight = 28
    r += 1

    # 학습 일정
    schedule = [
        ("1일차", "공정 개요·안전 수칙·부품 이름 숙지", "사수 옆 참관"),
        ("3일차", "작업 순서 따라하기 (사수 1:1)", "사수가 1단계씩 시범"),
        ("1주차", "독립 작업 (사수 옆 감독)", "1시간 표준 C/T 내 작업"),
        ("2주차", "자주검사 단독 판정 가능", "사수 1차 일치 확인"),
        ("2주말", f"{proc['level']} 도달 평가 → 단독 작업 인증", "체크리스트 모두 V + 사수 사인"),
    ]
    for stage, std, mth in schedule:
        _set_cell(ws, f"A{r}", stage, size=11, bold=True, hcenter=True, bg=C_GRAY)
        ws.Range(f"B{r}:D{r}").Merge()
        _set_cell(ws, f"B{r}", std, size=11)
        ws.Range(f"E{r}:H{r}").Merge()
        _set_cell(ws, f"E{r}", mth, size=11)
        ws.Rows(r).RowHeight = 22
        r += 1

    # 사인란
    r += 1
    ws.Range(f"A{r}:H{r}").Merge()
    _set_cell(ws, f"A{r}",
              "사수 사인: _____________________      라인장 확인: _____________________      날짜: 26 / __ / __",
              size=11, hcenter=True, border=False)
    ws.Rows(r).RowHeight = 24

    # === 인쇄 설정 (A4 가로 1장에 맞춤) ===
    ws.PageSetup.Orientation = 2  # landscape
    ws.PageSetup.PaperSize = 9  # A4
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False  # 세로는 자동 (1~2장)
    ws.PageSetup.LeftMargin = 14
    ws.PageSetup.RightMargin = 14
    ws.PageSetup.TopMargin = 20
    ws.PageSetup.BottomMargin = 20
    ws.PageSetup.CenterHorizontally = True


def main():
    print("[1/3] 데이터 로드...")
    procs = load_processes()
    ef = load_eval_forms()
    ppl = load_personnel()

    OUTDIR.mkdir(parents=True, exist_ok=True)
    print("[2/3] Excel COM...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    print("[3/3] 6공정 빌드 (공정당 1시트)...")
    src_wb = openpyxl.load_workbook(STD / "SP3M3_작업표준서_200730_R1.xlsx", data_only=True)
    ws_sheetnames = src_wb.sheetnames
    src_wb.close()

    for proc in procs:
        out_path = (OUTDIR / f"SP3M3_공정{proc['no']}_신규입사자 교육자료.xlsx").resolve()
        if out_path.exists():
            out_path.unlink()

        ws_data = load_workstd_steps(proc["no"], ws_sheetnames)
        parts = extract_parts_from_proc(proc, ef.get(proc["no"]), ws_data)

        wb = excel.Workbooks.Add()
        while wb.Sheets.Count > 1:
            wb.Sheets(wb.Sheets.Count).Delete()
        build_one(wb.Sheets(1), proc, ef.get(proc["no"]), ws_data, parts, ppl.get(proc["no"], []))

        wb.SaveAs(str(out_path), FileFormat=xlOpenXMLWorkbook)
        wb.Close(False)
        print(f"  ✓ {out_path.name}  (작업표준 ref={ws_data.get('ref_sheet')}, 부품={len(parts)})")

    excel.ScreenUpdating = True
    excel.Quit()
    print(f"\n완료 — {OUTDIR}")


if __name__ == "__main__":
    main()
