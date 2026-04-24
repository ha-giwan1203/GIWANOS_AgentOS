"""
SP3M3 공정별 숙련도 평가서 — 개인별 파일 (v2)
- 한 작업자당 1파일, 수행공정별 시트
- 날짜는 YYYY-MM-DD 문자열 형식 (datetime 혼입 방지)
- 샘플 양식 그대로 + copy_worksheet로 시트 복제 → 스타일/수식 100% 보존
"""
import os
import shutil
import tempfile
import random
from datetime import datetime, date
from openpyxl import load_workbook

BASE        = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
SAMPLE      = os.path.join(BASE, '공정 평가서표준_SP3M3_샘플.xlsx')
OUT_BASE    = os.path.join(BASE, 'SP3M3_개인별 평가서')

# 기존 개인 파일 소스
SOURCE_DIRS = [
    (os.path.join(BASE, 'SP3M3 주간 숙련도 평가서'), '주간'),
    (os.path.join(BASE, 'SP3M3 야간 숙련도 평가서'), '야간'),
]

# 기존 파일에서 추출 시 공정 열 좌표
PROC_COLS  = [14, 46, 78, 110, 142, 174, 206, 238, 270, 302, 334]
TOTAL_COLS = [10, 42, 74, 106, 138, 170, 202, 234, 266, 298, 330]
LEVEL_COLS = [20, 52, 84, 116, 148, 180, 212, 244, 276, 308, 340]

# ── 새 양식 공정별 데이터 (v4 스크립트와 동일) ────────────────────────────
PROCESSES = [
    {
        'no': '10',
        'name': '프레임 앗세이 로딩 & 스풀 로딩 & 바코드 부착',
        'std': [
            ('프레임 이종 여부를 확인하고 있는가? (차종별/초물표 참조)',
             '이종 100% 확인', '이종 80% 이상 확인', '이종 확인 미흡'),
            ('스풀 조립 상태를 확인하고 있는가? (이종/누락/정방향)',
             '3항목 모두 확인', '2항목 확인', '확인 미흡'),
            ('어퍼스테이 이종 여부를 확인하고 있는가?',
             '이종 100% 확인', '이종 80% 확인', '이종 확인 미흡'),
            ('웨빙가이드 조립 상태를 확인하고 있는가? (흔들림/유격)',
             '흔들림/유격 없음', '경미한 유격', '유격 다수 발생'),
            ('프레임 메인 컨베이어 안착 및 이송을 정확히 수행하는가?',
             '정위치 안착/이송', '1~2회 재조정', '안착 불량 다수'),
        ],
        'ctrl': [
            ('바코드 스티커 사양이 맞는지 확인하고 있는가?',
             '사양 100% 일치', '사양 80% 일치', '사양 불일치'),
            ('바코드 스티커 방향이 올바른지 확인하고 있는가?',
             '방향 100% 정확', '지적후 확인', '방향 오류'),
            ('바코드 스티커 부착상태를 제대로 확인하고 있는가?',
             '부착 정상 100%', '일부 재부착', '부착 불량'),
            ('프레임 이종을 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
        ],
    },
    {
        'no': '11',
        'name': 'WEBBING GUIDE 앗세이 조립',
        'std': [
            ('웨빙가이드 조립 상태를 확인하고 있는가? (흔들림/유격 여부)',
             '흔들림/유격 없음', '경미한 유격', '유격 다수 발생'),
            ('웨빙가이드 차종별 사양을 확인하고 있는가?',
             '사양 100% 확인', '사양 80% 확인', '사양 확인 미흡'),
            ('웨빙가이드 프레임 안착 방향을 확인하고 있는가?',
             '방향 100% 정확', '지적후 확인', '방향 오류'),
            ('웨빙가이드 이종 여부를 확인하고 있는가?',
             '이종 100% 확인', '이종 80% 확인', '이종 확인 미흡'),
            ('웨빙가이드 조립 후 검사를 수행하고 있는가?',
             '전수 검사 수행', '샘플 검사', '검사 누락'),
        ],
        'ctrl': [
            ('웨빙가이드 이종을 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('웨빙가이드 방향을 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('웨빙가이드 조립 상태를 주기적으로 점검하고 있는가?',
             '주기 100% 준수', '지연 발생', '점검 누락'),
            ('웨빙가이드 부적합품을 관리계획서 기준으로 처리하고 있는가?',
             '절차대로 처리', '처리 지연', '처리 미흡'),
        ],
    },
    {
        'no': '91',
        'name': 'LOCKING ASSY 조립',
        'std': [
            ('샤프트 베어링 부시 공급 상태를 확인하고 있는가? (누락 여부)',
             '누락 0건', '누락 경미', '누락 다수'),
            ('락킹 ASS\'Y 조립 상태를 확인하고 있는가? (정위치 여부)',
             '정위치 100%', '경미한 오차', '정위치 미흡'),
            ('토션 사양(각인 상태)을 검사하고 있는가?',
             '각인 100% 확인', '각인 80% 확인', '각인 확인 미흡'),
            ('리턴스프링 조립상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('에어 압력 상태를 확인하고 있는가? (0.4~0.6 MPa)',
             '범위 내 유지', '경계값 운전', '범위 이탈'),
        ],
        'ctrl': [
            ('홀더 휠 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('락킹 ASS\'Y 공급 상태를 주기적으로 점검하고 있는가?',
             '주기 100% 준수', '지연 발생', '점검 누락'),
            ('토션/베이스락 압입상태를 관리기준으로 측정하는가? (0±0.05mm)',
             '기준 내 100%', '경계값 발생', '기준 이탈'),
            ('압입 후 청결 상태(이물질)를 확인하고 있는가?',
             '이물질 없음', '경미한 이물질', '이물질 다수'),
        ],
    },
    {
        'no': '140',
        'name': '볼 가이드 조립 & 파이프 조립 & 플레이트 커버 조립',
        'std': [
            ('파이프 조립 상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('볼 가이드 ASS\'Y 공급 상태를 확인하고 있는가? (누락 여부)',
             '누락 0건', '누락 경미', '누락 다수'),
            ('볼 가이드 이너 조립 상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('플레이트 커버 공급 상태를 확인하고 있는가? (누락 여부)',
             '누락 0건', '누락 경미', '누락 다수'),
            ('플레이트 커버 조립 상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
        ],
        'ctrl': [
            ('파이프 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('볼 가이드 ASS\'Y 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('볼가이드 이너 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('플레이트 커버 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
        ],
    },
    {
        'no': '340',
        'name': 'VEHICLE SENSOR & 센서 커버 & 센스 브라켓 조립',
        'std': [
            ('센스 커버 아웃터, 비클 센서 이너 ASS\'Y 조립 상태를 확인하는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('비클 센서 정사양 및 조립 방향을 확인하고 있는가?',
             '사양/방향 100%', '80% 확인', '확인 미흡'),
            ('센스 커버 아웃터 ASS\'Y 프레임 조립 상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('비클 센서 ASS\'Y 조립 상태를 확인하고 있는가?',
             '조립 정상', '일부 재조립', '조립 미흡'),
            ('비클 센서 부품 청결 상태를 유지하고 있는가?',
             '청결 유지', '경미한 오염', '오염 다수'),
        ],
        'ctrl': [
            ('센서 커버 아웃터 핀 압입 상태를 관리기준으로 확인하는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('비클 센서 볼 조립 상태를 관리기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('센서 볼 마스터(OK/NG)를 시업시 확인하고 있는가?',
             '매 시업시 확인', '지연 확인', '확인 누락'),
            ('커버핀 압입 높이를 관리기준으로 확인하고 있는가?',
             '기준 내 100%', '경계값 발생', '기준 이탈'),
        ],
    },
    {
        'no': '430',
        'name': '프레임 앗세이 & 최종 검사 포장',
        'std': [
            ('스풀 조립 상태를 최종 확인하고 있는가?',
             '전수 확인', '샘플 확인', '확인 누락'),
            ('부품 조립 외관 상태를 확인하고 있는가?',
             '외관 정상', '경미한 결함', '결함 다수'),
            ('포장 박스 청결도 상태를 확인하고 있는가?',
             '청결 유지', '경미한 오염', '오염 다수'),
            ('포장 상태(규격/수량)를 확인하고 있는가?',
             '규격/수량 정확', '일부 불일치', '불일치 다수'),
            ('최종 검사 체크리스트를 준수하고 있는가?',
             '100% 준수', '일부 누락', '준수 미흡'),
        ],
        'ctrl': [
            ('완제품 외관 검사를 관리기준으로 수행하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('포장 수량을 관리계획서 기준으로 확인하고 있는가?',
             '관리기준 100% 준수', '일부 미준수', '관리기준 미준수'),
            ('출하 라벨 부착 상태를 확인하고 있는가?',
             '부착 정상 100%', '일부 재부착', '부착 불량'),
            ('출하 전 최종 검사 기록을 작성하고 있는가?',
             '기록 100% 작성', '일부 누락', '기록 미흡'),
        ],
    },
]
PROC_BY_NO = {p['no']: p for p in PROCESSES}

# ── 새 양식 셀 좌표 ───────────────────────────────────────────────────────
# 공정 정보
ROW_INFO = 5
COL_P1_NO, COL_P1_NAME, COL_P1_ITEM = 14, 19, 5
COL_P1_Q, COL_P1_U, COL_P1_Y = 17, 21, 25
COL_P2_NO, COL_P2_NAME, COL_P2_ITEM = 45, 50, 36
COL_P2_Q, COL_P2_U, COL_P2_Y = 48, 52, 56
COL_P3_NO, COL_P3_NAME, COL_P3_ITEM = 76, 81, 67
COL_P3_Q, COL_P3_U, COL_P3_Y = 79, 83, 87
ROWS_STD, ROWS_CTRL = list(range(13, 18)), list(range(18, 22))

# 업체/라인 셀
COL_BIZ, COL_LINE = 4, 9                 # D5, I5
# 작업자 정보 셀
CELL_NAME    = (2, 22)                   # V2
CELL_EMP_ID  = (3, 22)                   # V3
CELL_POS     = (3, 26)                   # Z3
CELL_DATE    = (5, 28)                   # AB5
# 평가자 의견 (merged B33:F37)
CELL_OPINION = (33, 2)                   # B33


def safe_write(ws, row, col, value):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def normalize_date(d):
    """datetime/date → 'YYYY-MM-DD' 문자열"""
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y-%m-%d')
    if d is None:
        return ''
    # 이미 문자열인 경우 앞 10자만 (시간 부분 제거)
    s = str(d)
    return s[:10] if len(s) >= 10 else s


def fill_process_sheet(ws, proc):
    """공정 시트의 평가사항/평가기준 채우기 (공정별 양식)"""
    # 업체정보 + 라인
    safe_write(ws, ROW_INFO, COL_BIZ, '대원테크')
    safe_write(ws, ROW_INFO, COL_LINE, 'SP3M3')
    # 공정번호/공정명
    safe_write(ws, ROW_INFO, COL_P1_NO, proc['no'])
    safe_write(ws, ROW_INFO, COL_P1_NAME, proc['name'])

    # 작업표준서 5행
    for i, r in enumerate(ROWS_STD):
        if i < len(proc['std']):
            item, q, u, y = proc['std'][i]
            safe_write(ws, r, COL_P1_ITEM, item)
            safe_write(ws, r, COL_P1_Q, q)
            safe_write(ws, r, COL_P1_U, u)
            safe_write(ws, r, COL_P1_Y, y)
    # 관리계획서 4행
    for i, r in enumerate(ROWS_CTRL):
        if i < len(proc['ctrl']):
            item, q, u, y = proc['ctrl'][i]
            safe_write(ws, r, COL_P1_ITEM, item)
            safe_write(ws, r, COL_P1_Q, q)
            safe_write(ws, r, COL_P1_U, u)
            safe_write(ws, r, COL_P1_Y, y)

    # 2·3번째 공정 열 전부 지우기
    for col_no, col_name, col_item, col_q, col_u, col_y in [
        (COL_P2_NO, COL_P2_NAME, COL_P2_ITEM, COL_P2_Q, COL_P2_U, COL_P2_Y),
        (COL_P3_NO, COL_P3_NAME, COL_P3_ITEM, COL_P3_Q, COL_P3_U, COL_P3_Y),
    ]:
        safe_write(ws, ROW_INFO, col_no, '')
        safe_write(ws, ROW_INFO, col_name, '')
        for r in ROWS_STD + ROWS_CTRL:
            safe_write(ws, r, col_item, '')
            safe_write(ws, r, col_q, '')
            safe_write(ws, r, col_u, '')
            safe_write(ws, r, col_y, '')


def build_master_template():
    """샘플을 기반으로 공정6개 시트 마스터 파일을 임시 생성."""
    tmp = tempfile.NamedTemporaryFile(
        suffix='.xlsx', delete=False, prefix='sp3m3_master_'
    )
    tmp.close()
    shutil.copy(SAMPLE, tmp.name)

    wb = load_workbook(tmp.name)
    template_ws = wb['양식']

    # 공정별 시트 6개 생성 (copy_worksheet = 같은 workbook 내 → 스타일/수식 100% 보존)
    for proc in PROCESSES:
        ws = wb.copy_worksheet(template_ws)
        ws.title = f'공정{proc["no"]}'
        fill_process_sheet(ws, proc)

    # 원본 양식 시트 삭제
    del wb['양식']
    wb.save(tmp.name)
    wb.close()
    return tmp.name


def extract_worker(path, shift):
    """기존 개인 파일 → 작업자 정보 + 수행공정 추출.
       첫 번째 공정 = 주공정, 나머지 = 전환공정."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    name    = os.path.basename(path).split(' 숙련도')[0]
    emp_id  = ws.cell(3, 22).value
    # Z3는 원본에서 '주간반'(shift)으로 잘못 기입된 경우 있음. shift로 대체
    pos     = shift + '반'
    date_v  = normalize_date(ws.cell(5, 28).value)

    procs = []
    for pc, tc, lc in zip(PROC_COLS, TOTAL_COLS, LEVEL_COLS):
        pv = ws.cell(5, pc).value
        if pv is not None:
            procs.append({
                'no':    str(pv),
                'total': ws.cell(32, tc).value,
                'level': ws.cell(32, lc).value,
            })

    # 주공정 = 첫 번째 공정
    main_proc_no = procs[0]['no'] if procs else None
    wb.close()
    return {
        'name':   name,
        'emp_id': emp_id,
        'pos':    pos,
        'date':   date_v,
        'shift':  shift,
        'procs':  procs,
        'main_proc': main_proc_no,
    }


def distribute_score(target_total, n_items=20, max_per=5, seed=None):
    """총점을 n_items 항목에 랜덤 분배 (0점 없이 모든 항목 값 유지).
       base = target // n, 나머지 rem개를 랜덤 위치에 +1점 배분.
       seed 고정: 같은 작업자/공정이면 재실행에도 동일 결과.
       예) 93점/20 → base=4, rem=13: 랜덤 13개 위치에 +1 → 13개 5점 + 7개 4점."""
    base = target_total // n_items
    rem  = target_total - base * n_items
    scores = [base] * n_items
    rng = random.Random(seed)
    bonus_idx = rng.sample(range(n_items), rem)
    for i in bonus_idx:
        scores[i] += 1
    # 상한 안전장치 (max_per 초과 방지)
    scores = [min(s, max_per) for s in scores]
    return scores


def fill_worker_info(ws, worker, proc_info, is_main):
    """각 시트에 작업자 정보 + 주공정/전환공정 라벨 + AC점수 자동분배 + 참고점수"""
    safe_write(ws, *CELL_NAME,   worker['name'])
    safe_write(ws, *CELL_EMP_ID, worker['emp_id'])
    # Z3: 주공정/전환공정 라벨 (샘플 원본 의미)
    safe_write(ws, *CELL_POS,    '주공정' if is_main else '전환공정')
    safe_write(ws, *CELL_DATE,   worker['date'])

    # AC 열 점수 자동 분배 (총점 목표: proc_info['total'])
    total = proc_info.get('total')
    if isinstance(total, (int, float)):
        seed = hash(f"{worker['name']}_{proc_info['no']}") & 0xFFFFFFFF
        scores = distribute_score(int(total), seed=seed)
        for i, r in enumerate(range(9, 29)):
            safe_write(ws, r, 29, scores[i])

    # 평가자 의견 (merged B33:F37): 라벨 + 주공정/전환공정 + 직전 평가 참고
    tag = '주공정' if is_main else '전환공정'
    text = (
        f'평 가 자 의 견\n\n'
        f'[{tag}] [직전 평가: {proc_info["total"]}점 / {proc_info["level"]}]\n'
        f'{worker["shift"]}반 / 점수는 총점 기준 자동분배 (항목별 재검토 필요)'
    )
    safe_write(ws, *CELL_OPINION, text)


def create_worker_file(worker, master_path, out_dir):
    """작업자 1명 → 1파일 (공정별 시트)"""
    out_path = os.path.join(out_dir, f'{worker["name"]} 숙련도평가서.xlsx')
    shutil.copy(master_path, out_path)

    wb = load_workbook(out_path)
    worker_procs = {p['no'] for p in worker['procs']}

    # 작업자가 수행하지 않는 공정 시트 삭제
    for sname in list(wb.sheetnames):
        proc_no = sname.replace('공정', '')
        if proc_no not in worker_procs:
            del wb[sname]

    # 남은 시트에 작업자 정보 기입 (첫 공정 = 주공정)
    main_no = worker.get('main_proc')
    for p in worker['procs']:
        sname = f'공정{p["no"]}'
        if sname in wb.sheetnames:
            is_main = (p['no'] == main_no)
            fill_worker_info(wb[sname], worker, p, is_main)

    # 시트 재배치 (원본 순서: 주공정 → 전환공정 순)
    desired = [f'공정{p["no"]}' for p in worker['procs']]
    existing = [n for n in desired if n in wb.sheetnames]
    wb._sheets = [wb[n] for n in existing]

    # 활성 시트를 주공정으로 설정
    if existing:
        wb.active = 0

    wb.save(out_path)
    wb.close()
    return out_path


def main():
    # 1) 출력 폴더
    out_day   = os.path.join(OUT_BASE, 'SP3M3 주간')
    out_night = os.path.join(OUT_BASE, 'SP3M3 야간')
    os.makedirs(out_day,   exist_ok=True)
    os.makedirs(out_night, exist_ok=True)

    # 2) 마스터 템플릿 생성 (공정 6개 시트 포함)
    print('[1/3] 마스터 템플릿 생성 중...')
    master = build_master_template()
    print(f'  임시 마스터: {master}')

    try:
        # 3) 작업자 정보 추출
        print('\n[2/3] 작업자 정보 추출...')
        workers = []
        for src, shift in SOURCE_DIRS:
            for f in sorted(os.listdir(src)):
                if f.endswith('.xlsx'):
                    w = extract_worker(os.path.join(src, f), shift)
                    workers.append(w)
                    print(f'  [{shift}] {w["name"]} / 공정 {len(w["procs"])}개')

        # 4) 개인별 파일 생성
        print(f'\n[3/3] 개인별 파일 {len(workers)}개 생성...')
        for w in workers:
            out_dir = out_day if w['shift'] == '주간' else out_night
            path = create_worker_file(w, master, out_dir)
            procs_str = '/'.join(p['no'] for p in w['procs'])
            print(f'  {w["name"]} ({procs_str}) -> {os.path.relpath(path, OUT_BASE)}')
    finally:
        # 5) 임시 마스터 삭제
        if os.path.exists(master):
            os.remove(master)

    print(f'\n[완료] {OUT_BASE}')


if __name__ == '__main__':
    main()
