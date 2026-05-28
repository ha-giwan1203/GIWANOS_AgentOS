# -*- coding: utf-8 -*-
"""
SP3M3 신규 입사자 공정별 교육자료 자동 생성
- 입력: 라인별공정목록 / 관리계획서 / 작업표준서 / 공정별 평가서 / 개인별 평가서
- 출력: SP3M3_신규입사자 교육자료/SP3M3_공정{N}_신규입사자 교육자료.xlsx (6개) + 투입현황.csv
"""
from __future__ import annotations
import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent  # 01_인사근태/숙련도평가
STD = BASE / "SP3M3_표준문서"
PROC_FORM = BASE / "SP3M3_공정별 평가서"
PERSONAL = BASE / "SP3M3_개인별 평가서"
OUTDIR = BASE / "SP3M3_신규입사자 교육자료"

# ─────────────────────────────────────────────
# 스타일
# ─────────────────────────────────────────────
THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(name="맑은 고딕", size=10, bold=True, color="1F4E79")
BODY_FONT = Font(name="맑은 고딕", size=10)
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
WRAP_LEFT = Alignment(wrap_text=True, vertical="center", horizontal="left")
TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")


# ─────────────────────────────────────────────
# 1. 라인별공정목록 로드
# ─────────────────────────────────────────────
def load_processes():
    """공정 6개 메타: [{'no':10, 'name':'...', 'level':'Lv.3'}, ...]"""
    wb = openpyxl.load_workbook(STD / "SP3M3 라인별공정목록.xlsx", data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        out.append({
            "no": int(no),
            "name": (ws.cell(r, 2).value or "").strip(),
            "level": (ws.cell(r, 3).value or "Lv.3").strip(),
        })
    wb.close()
    return out


# ─────────────────────────────────────────────
# 2. 관리계획서 → 공정번호별 관리항목·관리기준
# ─────────────────────────────────────────────
def load_mgmt_plan():
    """공정번호별 관리항목 리스트.
    A=공정번호, H/I=관리항목(제품/공정), M=관리기준, N=확인방법, O=주기, P=관리방안
    """
    wb = openpyxl.load_workbook(STD / "SP3M3_관리계획서_대원테크.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    result = defaultdict(list)
    current_no = None
    for r in range(14, ws.max_row + 1):
        no_cell = ws.cell(r, 1).value
        if isinstance(no_cell, (int, float)) and no_cell >= 10:
            current_no = int(no_cell)
        if current_no is None:
            continue
        item = ws.cell(r, 8).value or ws.cell(r, 9).value  # H or I
        standard = ws.cell(r, 13).value  # M
        method = ws.cell(r, 14).value  # N
        cycle = ws.cell(r, 15).value  # O
        plan = ws.cell(r, 16).value  # P
        if not any([item, standard, method, cycle, plan]):
            continue
        result[current_no].append({
            "item": (item or "").strip() if isinstance(item, str) else item,
            "standard": (standard or "").strip() if isinstance(standard, str) else standard,
            "method": (method or "").strip() if isinstance(method, str) else method,
            "cycle": (cycle or "").strip() if isinstance(cycle, str) else cycle,
            "plan": (plan or "").strip() if isinstance(plan, str) else plan,
        })
    wb.close()
    return result


# ─────────────────────────────────────────────
# 3. 작업표준서 → 공정번호별 작업스텝/자주검사
# ─────────────────────────────────────────────
def _find_ws_sheet(wb, proc_no):
    """공정번호로 작업표준서 시트 매핑.
    우선순위: (1) 정확 일치 (2) 묶음 범위(10_20 형태) (3) 근접(±5 이내)
    """
    exact, ranges, near = [], [], []
    for sn in wb.sheetnames:
        nums = [int(n) for n in re.findall(r"\d+", sn)]
        if not nums:
            continue
        if proc_no in nums:
            exact.append(sn)
            continue
        # 묶음 시트: 범위 폭 ≤ 30 (10_20, 140_150_160 등)
        if len(nums) >= 2 and (max(nums) - min(nums)) <= 30:
            if min(nums) <= proc_no <= max(nums):
                ranges.append((max(nums) - min(nums), sn))
                continue
        # 근접 매칭 (±5)
        diff = min(abs(n - proc_no) for n in nums)
        if diff <= 5:
            near.append((diff, sn))
    if exact:
        return exact[0]
    if ranges:
        ranges.sort()
        return ranges[0][1]
    if near:
        near.sort()
        return near[0][1]
    return None


def load_workstd():
    """공정번호별 {'steps':[...], 'inspect':[...]} dict."""
    wb = openpyxl.load_workbook(STD / "SP3M3_작업표준서_200730_R1.xlsx", data_only=True)
    procs = load_processes()
    out = {}
    for p in procs:
        no = p["no"]
        sn = _find_ws_sheet(wb, no)
        steps = []
        inspect = []
        ref_sheet = sn
        if sn:
            ws = wb[sn]
            # 자주검사 항목: 행 14~ 사이 NO/항목/기준/주기/검사방법 (M/N/Q/W ~ M=13/N=14/Q=17/W=23/AA=27)
            in_inspect = False
            in_steps = False
            for r in range(1, min(ws.max_row + 1, 60)):
                m = ws.cell(r, 13).value  # col M
                if isinstance(m, str):
                    if "자   주   검   사" in m or "자주검사" in m.replace(" ", ""):
                        in_inspect, in_steps = True, False
                        continue
                    if "공   정   작   업   순   서" in m or "공정작업순서" in m.replace(" ", ""):
                        in_inspect, in_steps = False, True
                        continue
                    if "조   건   관   리" in m or "조건관리" in m.replace(" ", ""):
                        in_inspect, in_steps = False, False
                        continue
                if in_inspect:
                    no_v = ws.cell(r, 13).value
                    item = ws.cell(r, 14).value
                    std = ws.cell(r, 17).value
                    cyc = ws.cell(r, 23).value
                    mth = ws.cell(r, 27).value
                    if item and isinstance(item, str) and item.strip() and "항" not in item[:2]:
                        inspect.append({
                            "no": no_v, "item": item.strip(),
                            "standard": (std or "").strip() if isinstance(std, str) else std,
                            "cycle": (cyc or "").strip() if isinstance(cyc, str) else cyc,
                            "method": (mth or "").strip() if isinstance(mth, str) else mth,
                        })
                if in_steps:
                    v = ws.cell(r, 13).value
                    if isinstance(v, str) and v.strip():
                        steps.append(v.strip())
        out[no] = {"steps": steps, "inspect": inspect, "ref_sheet": ref_sheet}
    wb.close()
    return out


# ─────────────────────────────────────────────
# 4. 공정별 평가서 → 평가항목 9개 + 평가기준
# ─────────────────────────────────────────────
def load_eval_forms():
    """공정번호별 {'items':[(E13..E21)], 'criteria':[(Q,U,Y)*9]}"""
    out = {}
    for fp in PROC_FORM.glob("SP3M3_공정*.xlsx"):
        m = re.search(r"공정(\d+)", fp.name)
        if not m:
            continue
        no = int(m.group(1))
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        items = [ws.cell(r, 5).value for r in range(13, 22)]
        criteria = [
            (ws.cell(r, 17).value, ws.cell(r, 21).value, ws.cell(r, 25).value)
            for r in range(13, 22)
        ]
        out[no] = {"items": items, "criteria": criteria}
        wb.close()
    return out


# ─────────────────────────────────────────────
# 5. 개인별 평가서 → 공정별 투입자·점수
# ─────────────────────────────────────────────
def load_personnel():
    """공정번호별 [{name, shift, ac_sum, role}] (role = 주공정/전환공정)"""
    out = defaultdict(list)
    for shift_dir, shift in [("SP3M3 주간", "주간"), ("SP3M3 야간", "야간")]:
        d = PERSONAL / shift_dir
        if not d.exists():
            continue
        for fp in d.glob("*.xlsx"):
            name = fp.stem.replace(" 숙련도평가서", "").strip()
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                n5 = ws["N5"].value
                z3 = ws["Z3"].value
                if not isinstance(n5, (int, float)):
                    try:
                        n5 = int(str(n5))
                    except Exception:
                        continue
                ac_sum = 0
                for r in range(13, 22):
                    v = ws.cell(r, 29).value
                    if isinstance(v, (int, float)):
                        ac_sum += v
                out[int(n5)].append({
                    "name": name, "shift": shift, "ac_sum": ac_sum,
                    "role": (z3 or "").strip() if isinstance(z3, str) else z3,
                })
            wb.close()
    return out


# ─────────────────────────────────────────────
# 6. 빌드 — 시트 6개짜리 xlsx 생성
# ─────────────────────────────────────────────
def _style_header(cell, text, fill=HEAD_FILL, font=HEAD_FONT):
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = WRAP_CENTER
    cell.border = BORDER


def _style_body(cell, text):
    cell.value = text
    cell.font = BODY_FONT
    cell.alignment = WRAP_LEFT
    cell.border = BORDER


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sheet_overview(ws, proc, personnel_for_proc):
    ws.title = "1.공정개요"
    _set_widths(ws, [4, 18, 24, 28, 20])
    ws["B2"] = f"SP3M3 공정 {proc['no']} 신규 입사자 교육자료"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 28

    rows = [
        ("공정번호", str(proc["no"])),
        ("공정명", proc["name"]),
        ("라인", "SP3M3"),
        ("요구 작업레벨", proc["level"]),
        ("현재 투입인원(주간)", str(sum(1 for p in personnel_for_proc if p["shift"] == "주간"))),
        ("현재 투입인원(야간)", str(sum(1 for p in personnel_for_proc if p["shift"] == "야간"))),
    ]
    r = 4
    for k, v in rows:
        _style_header(ws.cell(r, 2), k, fill=SUB_FILL, font=SUB_FONT)
        _style_body(ws.cell(r, 3), v)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1

    # 사수 후보 (AC 합계 상위 2명, 주간/야간 각각)
    r += 1
    _style_header(ws.cell(r, 2), "사수 후보 (OJT 멘토 — AC합계 상위)", fill=HEAD_FILL, font=HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    headers = ["조", "이름", "역할(주/전환)", "AC합계"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h, fill=SUB_FILL, font=SUB_FONT)
    r += 1
    for shift in ["주간", "야간"]:
        cand = sorted(
            [p for p in personnel_for_proc if p["shift"] == shift],
            key=lambda x: -x["ac_sum"],
        )[:2]
        for c in cand:
            _style_body(ws.cell(r, 2), shift)
            _style_body(ws.cell(r, 3), c["name"])
            _style_body(ws.cell(r, 4), c["role"] or "-")
            _style_body(ws.cell(r, 5), c["ac_sum"])
            r += 1
        if not cand:
            _style_body(ws.cell(r, 2), shift)
            _style_body(ws.cell(r, 3), "(미투입)")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            r += 1

    # 보강 대상 (최저점)
    r += 1
    _style_header(ws.cell(r, 2), "보강 대상 (재교육 우선) — AC합계 하위", fill=HEAD_FILL, font=HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h, fill=SUB_FILL, font=SUB_FONT)
    r += 1
    for shift in ["주간", "야간"]:
        cand = sorted(
            [p for p in personnel_for_proc if p["shift"] == shift],
            key=lambda x: x["ac_sum"],
        )[:2]
        for c in cand:
            _style_body(ws.cell(r, 2), shift)
            _style_body(ws.cell(r, 3), c["name"])
            _style_body(ws.cell(r, 4), c["role"] or "-")
            _style_body(ws.cell(r, 5), c["ac_sum"])
            r += 1


def build_sheet_workstd(ws, proc, ws_data):
    ws.title = "2.작업절차"
    _set_widths(ws, [4, 6, 60, 16, 22])
    ws["B2"] = f"공정 {proc['no']} — 작업 절차 (작업표준서 발췌)"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:E2")

    r = 4
    _style_header(ws.cell(r, 2), "원본 시트")
    _style_body(ws.cell(r, 3), ws_data.get("ref_sheet") or "(매칭 누락 — 작업표준서 직접 확인)")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 2

    _style_header(ws.cell(r, 2), "NO")
    _style_header(ws.cell(r, 3), "작업 순서")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1
    steps = ws_data.get("steps") or [
        "(작업표준서에서 해당 공정 시트가 매칭되지 않음. 아래 GENERIC 항목 사용 + 라인반장 직접 보강 필요)",
        "1. 시업 점검 — 설비 일상점검표 작성 (에어압 0.4~0.6 MPa)",
        "2. 초물 검사 — 차종 변경 시 초물표 확인 후 첫 5개 사수 검사",
        "3. 본 작업 — 공정별 평가서 9개 평가항목 순서대로 수행",
        "4. 자주 검사 — 이종 100% 확인 / 조립 흔들림·유격 확인",
        "5. 종업 정리 — 잔재 정리, 폐기물 분리, 다음 조 인수인계 메모",
    ]
    for i, s in enumerate(steps, start=1):
        _style_body(ws.cell(r, 2), i)
        _style_body(ws.cell(r, 3), s)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    _style_header(ws.cell(r, 2), "자주검사 항목 (작업표준서)")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    headers = ["NO", "항목 / 기준", "주기", "검사방법"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h, fill=SUB_FILL, font=SUB_FONT)
    r += 1
    for ins in ws_data.get("inspect") or []:
        _style_body(ws.cell(r, 2), ins.get("no"))
        _style_body(ws.cell(r, 3), f"{ins.get('item','')}\n[기준] {ins.get('standard','')}")
        _style_body(ws.cell(r, 4), ins.get("cycle"))
        _style_body(ws.cell(r, 5), ins.get("method"))
        ws.row_dimensions[r].height = 36
        r += 1


def build_sheet_mgmt(ws, proc, mgmt_items):
    ws.title = "3.관리기준"
    _set_widths(ws, [4, 24, 28, 16, 14, 14])
    ws["B2"] = f"공정 {proc['no']} — 관리항목·합격기준 (관리계획서)"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:F2")

    r = 4
    headers = ["관리항목", "관리기준", "확인방법", "주기", "관리방안"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h)
    r += 1

    # GENERIC 보강 (관리계획서에 해당 공정 누락 시 SD9A01 패턴)
    if not mgmt_items:
        generic = [
            {"item": "이종 확인", "standard": "차종별 초물표 일치 — 이종 없을 것",
             "method": "초물 작업 시 사수 1차 확인 + 작업자 2차 확인", "cycle": "차종 변경 시", "plan": "초물표"},
            {"item": "안전 점검", "standard": "회전부·압입부 가드 정상 / 비상정지 동작 확인",
             "method": "시업 시 일상점검표 작성", "cycle": "시업시", "plan": "일상점검표"},
            {"item": "설비 점검", "standard": "에어압 0.4~0.6 MPa / 지그 내 이물질 없을 것",
             "method": "압력게이지 + 육안", "cycle": "시업시", "plan": "일상점검표"},
            {"item": "자주검사", "standard": "공정별 평가서 평가항목 9개 — 우수(5점) 기준 부합",
             "method": "공정별 평가서 체크리스트", "cycle": "전수", "plan": "검사 성적서"},
        ]
        for it in generic:
            _style_body(ws.cell(r, 2), it["item"] + " (GENERIC)")
            _style_body(ws.cell(r, 3), it["standard"])
            _style_body(ws.cell(r, 4), it["method"])
            _style_body(ws.cell(r, 5), it["cycle"])
            _style_body(ws.cell(r, 6), it["plan"])
            ws.row_dimensions[r].height = 32
            r += 1
        return
    for it in mgmt_items:
        _style_body(ws.cell(r, 2), it.get("item"))
        _style_body(ws.cell(r, 3), it.get("standard"))
        _style_body(ws.cell(r, 4), it.get("method"))
        _style_body(ws.cell(r, 5), it.get("cycle"))
        _style_body(ws.cell(r, 6), it.get("plan"))
        ws.row_dimensions[r].height = 30
        r += 1


def build_sheet_eval(ws, proc, eval_data):
    ws.title = "4.평가항목"
    _set_widths(ws, [4, 6, 44, 22, 22, 22])
    ws["B2"] = f"공정 {proc['no']} — 평가항목 9개 + 평가기준"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:F2")

    r = 4
    headers = ["NO", "평가 사항", "우수 (5점)", "보통 (3점)", "부족 (1점)"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h)
    r += 1
    if not eval_data:
        _style_body(ws.cell(r, 2), "(공정별 평가서 매칭 실패)")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        return
    items = eval_data["items"]
    criteria = eval_data["criteria"]
    for i, (it, (q, u, y)) in enumerate(zip(items, criteria), start=1):
        _style_body(ws.cell(r, 2), i)
        _style_body(ws.cell(r, 3), it or "")
        _style_body(ws.cell(r, 4), q or "")
        _style_body(ws.cell(r, 5), u or "")
        _style_body(ws.cell(r, 6), y or "")
        ws.row_dimensions[r].height = 36
        r += 1


def build_sheet_checklist(ws, proc):
    ws.title = "5.단계별체크리스트"
    _set_widths(ws, [4, 12, 30, 36, 10])
    ws["B2"] = f"공정 {proc['no']} — 입사 단계별 통과 체크리스트 ({proc['level']} 도달 로드맵)"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:E2")

    r = 4
    headers = ["단계", "통과 기준", "확인 방법", "완료(O/X)"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h)
    r += 1

    target_level = proc["level"]
    stages = [
        ("1일차", "공정 개요·안전 수칙 숙지", "사수 구두 확인 + 안전 서명", ""),
        ("3일차", "작업 순서 이해 / 자주검사 항목 명칭 암기", "사수 질의응답 (10문 중 8문 이상)", ""),
        ("1주차", "독립 작업 가능 (사수 옆 감독)", "1시간 표준 C/T 내 작업 가능", ""),
        ("2주차", "자주검사 5항목 단독 판정 가능", "초물·중물 검사 시 사수 1차 일치 확인", ""),
        ("1개월", f"{target_level} 수준 도달 평가 — 평가항목 9개 ", "공정별 평가서 점검(30/45점 이상)", ""),
        ("2개월", "주공정으로 단독 운영", "1주 무결점 + 사수 사인", ""),
    ]
    for stage, std, mth, mark in stages:
        _style_body(ws.cell(r, 2), stage)
        _style_body(ws.cell(r, 3), std)
        _style_body(ws.cell(r, 4), mth)
        _style_body(ws.cell(r, 5), mark)
        ws.row_dimensions[r].height = 34
        r += 1


def build_sheet_safety(ws, proc, ws_data, mgmt_items):
    ws.title = "6.안전·이종확인"
    _set_widths(ws, [4, 6, 28, 40])
    ws["B2"] = f"공정 {proc['no']} — 안전·이종확인 포인트"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:D2")

    r = 4
    headers = ["NO", "구분", "확인 사항"]
    for i, h in enumerate(headers):
        _style_header(ws.cell(r, 2 + i), h)
    r += 1

    # 작업표준서 inspect 중 '이종' 키워드
    items = []
    for ins in ws_data.get("inspect") or []:
        item = ins.get("item") or ""
        if "이종" in item or "안전" in item or "스위치" in item:
            items.append(("이종/안전", f"{item} — {ins.get('standard','')}"))
    # 관리계획서 중 '이종' 키워드
    for it in mgmt_items or []:
        item = it.get("item") or ""
        std = it.get("standard") or ""
        if "이종" in str(item) + str(std) or "안전" in str(item) + str(std):
            items.append(("관리기준", f"{item} — {std}"))

    # GENERIC 안전 보강 (SD9A01 패턴)
    generic = [
        ("일상점검", "시업 시 설비 일상점검표 작성 (에어압 0.4~0.6MPa 등)"),
        ("정전 시 대응", "정전 발생 시 즉시 작업 중지·관리자 보고·재시작 시 초물 검사"),
        ("개인보호구", "면장갑·보안경 착용 — 회전부·압입부 손가락 진입 금지"),
        ("이종 초물표", "차종 변경 시 초물표 반드시 확인 후 첫 5개 사수 검사"),
    ]
    for g in generic:
        items.append(g)

    for i, (kind, txt) in enumerate(items, start=1):
        _style_body(ws.cell(r, 2), i)
        _style_body(ws.cell(r, 3), kind)
        _style_body(ws.cell(r, 4), txt)
        ws.row_dimensions[r].height = 28
        r += 1


# ─────────────────────────────────────────────
# 7. 메인 빌드
# ─────────────────────────────────────────────
def build_one(proc, mgmt_all, ws_all, eval_all, personnel_all):
    wb = Workbook()
    # 시트 1 (기본)
    s1 = wb.active
    build_sheet_overview(s1, proc, personnel_all.get(proc["no"], []))
    build_sheet_workstd(wb.create_sheet(), proc, ws_all.get(proc["no"], {}))
    build_sheet_mgmt(wb.create_sheet(), proc, mgmt_all.get(proc["no"], []))
    build_sheet_eval(wb.create_sheet(), proc, eval_all.get(proc["no"]))
    build_sheet_checklist(wb.create_sheet(), proc)
    build_sheet_safety(wb.create_sheet(), proc, ws_all.get(proc["no"], {}), mgmt_all.get(proc["no"], []))

    OUTDIR.mkdir(parents=True, exist_ok=True)
    out = OUTDIR / f"SP3M3_공정{proc['no']}_신규입사자 교육자료.xlsx"
    wb.save(out)
    return out


def write_personnel_csv(personnel_all, procs):
    OUTDIR.mkdir(parents=True, exist_ok=True)
    fp = OUTDIR / "투입현황.csv"
    with open(fp, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["공정번호", "공정명", "요구레벨", "주간투입", "야간투입",
                    "평균AC합계", "사수후보(주간)", "사수후보(야간)",
                    "보강대상(주간)", "보강대상(야간)"])
        for p in procs:
            ppl = personnel_all.get(p["no"], [])
            day = [x for x in ppl if x["shift"] == "주간"]
            night = [x for x in ppl if x["shift"] == "야간"]
            avg = round(sum(x["ac_sum"] for x in ppl) / len(ppl), 1) if ppl else 0
            top_d = sorted(day, key=lambda x: -x["ac_sum"])[:1]
            top_n = sorted(night, key=lambda x: -x["ac_sum"])[:1]
            low_d = sorted(day, key=lambda x: x["ac_sum"])[:1]
            low_n = sorted(night, key=lambda x: x["ac_sum"])[:1]
            w.writerow([
                p["no"], p["name"], p["level"],
                len(day), len(night), avg,
                f"{top_d[0]['name']}({top_d[0]['ac_sum']})" if top_d else "-",
                f"{top_n[0]['name']}({top_n[0]['ac_sum']})" if top_n else "-",
                f"{low_d[0]['name']}({low_d[0]['ac_sum']})" if low_d else "-",
                f"{low_n[0]['name']}({low_n[0]['ac_sum']})" if low_n else "-",
            ])
    return fp


def main():
    print("[1/5] 라인별공정목록 로드...")
    procs = load_processes()
    print(f"  → 공정 {len(procs)}개: {[p['no'] for p in procs]}")

    print("[2/5] 관리계획서 로드...")
    mgmt = load_mgmt_plan()
    print(f"  → 매칭 공정: {sorted(mgmt.keys())}")

    print("[3/5] 작업표준서 로드...")
    wsd = load_workstd()
    for p in procs:
        d = wsd.get(p["no"], {})
        print(f"  → 공정{p['no']}: 시트={d.get('ref_sheet')}  steps={len(d.get('steps',[]))}  inspect={len(d.get('inspect',[]))}")

    print("[4/5] 공정별 평가서 + 개인별 평가서 로드...")
    ef = load_eval_forms()
    ppl = load_personnel()
    print(f"  → 평가서 매칭 공정: {sorted(ef.keys())}")
    print(f"  → 개인별 투입공정: {sorted(ppl.keys())}  총 row수={sum(len(v) for v in ppl.values())}")

    print("[5/5] 빌드...")
    OUTDIR.mkdir(parents=True, exist_ok=True)
    outs = []
    for p in procs:
        out = build_one(p, mgmt, wsd, ef, ppl)
        print(f"  ✓ {out.name}")
        outs.append(out)
    csv_fp = write_personnel_csv(ppl, procs)
    print(f"  ✓ {csv_fp.name}")

    print(f"\n완료 — 산출물 {len(outs)+1}개 in {OUTDIR}")


if __name__ == "__main__":
    main()
