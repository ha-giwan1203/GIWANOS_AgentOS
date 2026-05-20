# -*- coding: utf-8 -*-
"""
SP3M3 신규 입사자 교육자료 v5 — 공정당 A4 1~2장 (시트 1개)
디자인 개선: 모던 네이비/파스텔 팔레트, 카드형 레이아웃, 여백 확대.
⑦ 사수 배정 섹션 제거. 학습 일정만 남김.
"""
from __future__ import annotations
import re
from pathlib import Path

import openpyxl
import win32com.client as win32

BASE = Path(__file__).resolve().parent.parent
STD = BASE / "SP3M3_표준문서"
PROC_FORM = BASE / "SP3M3_공정별 평가서"
OUTDIR = BASE / "SP3M3_신규입사자 교육자료"

xlOpenXMLWorkbook = 51


# ─────────────────────────────────────────────
# 모던 색 팔레트 (VBA RGB = r + g*256 + b*65536)
# ─────────────────────────────────────────────
def rgb(r, g, b):
    return r + (g << 8) + (b << 16)


COLOR = {
    "navy": rgb(31, 78, 121),       # 메인 네이비
    "navy_dark": rgb(20, 50, 80),
    "white": rgb(255, 255, 255),
    "gray_100": rgb(247, 248, 250),  # 본문 배경
    "gray_200": rgb(229, 231, 235),  # 카드 테두리
    "gray_500": rgb(107, 114, 128),  # 보조 텍스트
    "gray_700": rgb(55, 65, 81),     # 본문 텍스트
    "blue_light": rgb(219, 234, 254),
    "blue_card": rgb(239, 246, 255),
    "green": rgb(34, 139, 84),
    "green_bg": rgb(220, 252, 231),
    "red": rgb(185, 28, 28),
    "red_bg": rgb(254, 226, 226),
    "yellow_bg": rgb(254, 243, 199),
    "yellow_accent": rgb(245, 158, 11),
}


# ─────────────────────────────────────────────
# 데이터 로드
# ─────────────────────────────────────────────
def load_processes():
    wb = openpyxl.load_workbook(STD / "SP3M3 라인별공정목록.xlsx", data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        out.append({"no": int(no), "name": (ws.cell(r, 2).value or "").strip(),
                    "level": (ws.cell(r, 3).value or "Lv.3").strip()})
    wb.close()
    return out


def load_eval_forms():
    out = {}
    for fp in PROC_FORM.glob("SP3M3_공정*.xlsx"):
        m = re.search(r"공정(\d+)", fp.name)
        if not m:
            continue
        no = int(m.group(1))
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        items = [ws.cell(r, 5).value for r in range(13, 22)]
        criteria = [(ws.cell(r, 17).value, ws.cell(r, 21).value, ws.cell(r, 25).value)
                    for r in range(13, 22)]
        out[no] = {"items": items, "criteria": criteria}
        wb.close()
    return out


def find_ws_sheet_name(sheetnames, proc_no):
    exact, ranges, near = [], [], []
    for sn in sheetnames:
        nums = [int(n) for n in re.findall(r"\d+", sn)]
        if not nums:
            continue
        if proc_no in nums:
            exact.append(sn); continue
        if len(nums) >= 2 and (max(nums) - min(nums)) <= 30:
            if min(nums) <= proc_no <= max(nums):
                ranges.append((max(nums) - min(nums), sn)); continue
        diff = min(abs(n - proc_no) for n in nums)
        if diff <= 5:
            near.append((diff, sn))
    if exact: return exact[0]
    if ranges: ranges.sort(); return ranges[0][1]
    if near: near.sort(); return near[0][1]
    return None


# ─────────────────────────────────────────────
# 공정별 작업순서 OVERRIDE (사용자 확정값 — 자동 추출 결과 대체)
# ─────────────────────────────────────────────
IDENTIFIER_CHECK_STEP = "부품박스 식별표 확인 (차종·품번 일치 — 모든 작업의 시작)"

PROC_OVERRIDES = {
    11: {
        "ref_sheet": "10_20 (사진 1~5)",
        "steps": [
            "프레임에 웨빙가이드를 1차 조립한다",
            "프레임에 웨빙가이드를 2차 조립한다",
            "프레임에 웨빙가이드를 돌려서 어퍼 브라켓에 3차 조립한다",
            "웨빙가이드를 그림과 같이 \"딱\" 소리가 날 때까지 눌러 준다",
            "해당 프레임 사양 확인 한다 (생산계획 품번 참조)",
        ],
    },
    91: {
        "ref_sheet": "90-1",
        "parts": ["토션", "베이스락", "락 플레이트", "리턴 스프링", "카버 플레이트 락"],
        "steps": [
            "토션을 자동 공급한다",
            "토션 사양을(구경) 측정 및 이종 검출을 한다",
            "베이스락을 자동 공급 및 조립을 한다",
            "토션에 베이스락 압입 및 높이 검사를 한다",
            "락 플레이트를 공급 및 조립을 한다",
            "리턴 스프링을 락플레이트 & 베이스락에 조립을 한다",
            "카버 플레이트 락을 베이스락에 공급 & 조립을 한다",
            "카버 플레이트 락 압입 및 높이검사",
            "부적합 발생시 취출 및 이송",
            "양품 취출",
        ],
    },
    140: {
        "ref_sheet": "130",
        "parts": ["파이프", "볼 가이드", "플레이트 커버"],
        "steps": [
            "팔레트 스토퍼가 작동하여 팔레트를 고정한다",
            "파이프와 볼가이드 ASS'Y 의 조립 위치 확인",
            "위치 확인 후 볼가이드 ASS'Y를 조립한다",
            "볼가이드핀 조립후 볼 가이드 이너 누락 검사를 진행한다 (누락 검사기에 삽입하여 누락유무 확인)",
            "누락검사 통과된 제품에 대하여 볼가이드 핀에 맞춰 안착 시킨다",
            "플레이트 커버를 조립한다",
            "후 공정으로 이송한다",
        ],
    },
    340: {
        "ref_sheet": "330",
        "parts": ["센스 커버 아웃터", "비클 센서 이너", "비클 센서 ASS'Y"],
        "steps": [
            "지그에 센스커버 아웃터를 안착한다",
            "지그에 비클 센서를 안착한다",
            "비클 센서에 비클센서 이너를 가 조립 한다",
            "버튼을 눌러 조립한다",
            "조립된 센스커버 아웃터 ASS'Y의 비클 센서를 에어 블로우 처리한다 (상세는 작업표준서 시트 330 참조)",
        ],
    },
    430: {
        "ref_sheet": "400",
        "skip_identifier_check": True,  # 최종검사·포장 — 부품 투입 공정 아님
        "parts_label": "검사·포장 품목 (취급 대상)",
        "parts": ["완제품 ASS'Y", "포장박스"],
        "steps": [
            "팔레트 고정 후 완제품 자동 취출 후 컨베이어에 놓는다",
            "제품을 잡고 외관·소음 검사한다",
            "포장용기 청결 상태를 확인한다",
            "(이후 단계는 작업표준서 시트 400 참조)",
        ],
    },
}


def _next_section_keyword(s):
    """다음 섹션 헤더 감지 — 작업순서 종료 신호.
    헤더는 보통 공백으로 글자 사이가 벌어진 형태('불 량 품   처 리   기 준'). 절차 본문 텍스트에
    '주의사항 참조' 같이 단어가 등장하는 건 종료 X. 헤더 패턴(전체가 헤더 텍스트)만 종료.
    """
    s2 = s.replace(" ", "")
    # 헤더성 키워드 + 셀 전체가 그 헤더에 가까운 경우만 (길이 ≤ 12자)
    if len(s2) > 14:
        return False
    return any(s2 == kw or s2.startswith(kw) for kw in
               ["불량품처리기준", "자체점검", "관리항목", "변경관리", "변경관리이력"])


def _merge_continuation_lines(steps):
    """숫자로 시작하지 않는 줄은 이전 항목에 부착 (줄바꿈된 한 절차)."""
    merged = []
    for s in steps:
        if re.match(r"^\s*\d+[.)]", s) or s.startswith("◆"):
            merged.append(s.strip())
        else:
            if merged:
                merged[-1] = merged[-1].rstrip() + " " + s.strip()
            else:
                merged.append(s.strip())
    return merged


def load_workstd_steps(proc_no):
    """공정별 작업순서 추출. OVERRIDE 우선 → 없으면 작업표준서 자동 추출."""
    # 1) OVERRIDE 있으면 우선 사용
    if proc_no in PROC_OVERRIDES:
        ov = PROC_OVERRIDES[proc_no]
        if ov.get("skip_identifier_check"):
            steps = list(ov["steps"])  # 부품 투입 공정 아님 → 식별표 단계 생략
        else:
            steps = [IDENTIFIER_CHECK_STEP] + list(ov["steps"])
        return {"ref_sheet": ov.get("ref_sheet"), "steps": steps, "inspect": []}

    # 2) 작업표준서 자동 추출
    wb = openpyxl.load_workbook(STD / "SP3M3_작업표준서_200730_R1.xlsx", data_only=True)
    sn = find_ws_sheet_name(wb.sheetnames, proc_no)
    raw_steps, inspect = [], []
    if sn:
        ws = wb[sn]
        in_inspect, in_steps = False, False
        for r in range(1, min(ws.max_row + 1, 60)):
            m = ws.cell(r, 13).value
            if isinstance(m, str):
                if "자   주   검   사" in m or "자주검사" in m.replace(" ", ""):
                    in_inspect, in_steps = True, False; continue
                if "공   정   작   업   순   서" in m or "공정작업순서" in m.replace(" ", ""):
                    in_inspect, in_steps = False, True; continue
                if "조   건   관   리" in m or "조건관리" in m.replace(" ", ""):
                    in_inspect, in_steps = False, False; continue
            if in_inspect:
                item = ws.cell(r, 14).value
                std = ws.cell(r, 17).value
                if isinstance(item, str) and item.strip() and len(item) < 40:
                    inspect.append((item.strip(), (std or "").strip() if isinstance(std, str) else ""))
            if in_steps:
                v = ws.cell(r, 13).value
                if isinstance(v, str) and v.strip():
                    if _next_section_keyword(v):
                        in_steps = False
                        continue
                    raw_steps.append(v.strip())
    wb.close()

    # 3) 줄바꿈된 한 절차 합치기 + ◆ 시업점검 제거
    merged = _merge_continuation_lines(raw_steps)
    main_steps = [s for s in merged if not s.startswith("◆")]
    # 본문 앞의 번호("1. ", "2.") 제거 — 카드의 동그란 배지가 번호 역할
    cleaned = [re.sub(r"^\s*\d+[.)]\s*", "", s) for s in main_steps]
    # 식별표 확인 prepend
    cleaned = [IDENTIFIER_CHECK_STEP] + cleaned
    return {"ref_sheet": sn, "steps": cleaned, "inspect": inspect}


def extract_parts(proc, eval_data, ws_data):
    # OVERRIDE 우선
    ov = PROC_OVERRIDES.get(proc["no"], {})
    if ov.get("parts"):
        return list(ov["parts"])
    text = proc["name"] + " "
    if eval_data:
        for it in eval_data["items"]:
            if isinstance(it, str):
                text += it + " "
    for s, _ in ws_data.get("inspect", []):
        text += s + " "
    known = ["프레임", "스풀", "웨빙가이드", "어퍼스테이", "바코드", "WEBBING GUIDE",
             "LOCKING", "ASSY", "볼 가이드", "파이프", "플레이트 커버", "VEHICLE SENSOR",
             "센서 커버", "센스 브라켓", "최종 검사", "포장", "릴", "스티커"]
    found = []
    for k in known:
        if k in text and k not in found:
            found.append(k)
    return found


# ─────────────────────────────────────────────
# 디자인 헬퍼 (Excel COM)
# ─────────────────────────────────────────────
def style_cell(cell, *, val=None, font_size=11, bold=False, font_color=None,
               bg=None, hcenter=False, vcenter=True, wrap=True,
               border_color=None, border_weight=2):
    if val is not None:
        cell.Value = val
    cell.Font.Name = "맑은 고딕"
    cell.Font.Size = font_size
    cell.Font.Bold = bold
    if font_color is not None:
        cell.Font.Color = font_color
    if bg is not None:
        cell.Interior.Color = bg
    cell.HorizontalAlignment = -4108 if hcenter else -4131
    if vcenter:
        cell.VerticalAlignment = -4108
    if wrap:
        cell.WrapText = True
    if border_color is not None:
        for edge in [7, 8, 9, 10]:  # Left/Top/Bottom/Right
            b = cell.Borders(edge)
            b.LineStyle = 1
            b.Weight = border_weight
            b.Color = border_color


def section_header(ws, row, span_a, span_b, num, title, accent_color):
    """좌측 색 막대 + 진한 네이비 헤더 띠."""
    # 좌측 색 막대 (1셀)
    bar = ws.Cells(row, span_a)
    bar.Value = ""
    bar.Interior.Color = accent_color
    bar.Borders.LineStyle = 0
    # 헤더 띠
    rng = ws.Range(ws.Cells(row, span_a + 1), ws.Cells(row, span_b))
    rng.Merge()
    style_cell(ws.Cells(row, span_a + 1),
               val=f"  {num}  {title}",
               font_size=13, bold=True, font_color=COLOR["white"],
               bg=COLOR["navy"], hcenter=False, vcenter=True)
    ws.Rows(row).RowHeight = 30


def card_box(ws, r1, c1, r2, c2, bg=None):
    """카드형 박스: 얇은 회색 테두리 + 내부 배경색."""
    rng = ws.Range(ws.Cells(r1, c1), ws.Cells(r2, c2))
    if bg is not None:
        rng.Interior.Color = bg
    # 외곽 테두리 (얇은 회색)
    for edge in [7, 8, 9, 10]:
        b = rng.Borders(edge)
        b.LineStyle = 1
        b.Weight = 2
        b.Color = COLOR["gray_200"]


# ─────────────────────────────────────────────
# 빌드 (공정당 1시트)
# ─────────────────────────────────────────────
def build_one(ws, proc, eval_data, ws_data, parts):
    # 시트명은 main()에서 일괄 부여 (단일 파일 다중 시트 — 충돌 방지)
    # 8칸 그리드 (A~H)
    widths = [2, 13, 16, 13, 13, 13, 13, 9]
    for i, w in enumerate(widths, start=1):
        ws.Columns(i).ColumnWidth = w

    # 격자선 제거 (인쇄 미리보기 깔끔하게)
    ws.Application.ActiveWindow.DisplayGridlines = False

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 헤더 (페이지 상단)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 상단 여백
    ws.Rows(1).RowHeight = 8

    # 라인 라벨 (소형)
    ws.Range("B2:H2").Merge()
    style_cell(ws.Cells(2, 2), val="SP3M3  ·  신규 입사자 교육자료",
               font_size=10, font_color=COLOR["gray_500"], hcenter=False)
    ws.Rows(2).RowHeight = 18

    # 메인 타이틀
    ws.Range("B3:H3").Merge()
    style_cell(ws.Cells(2, 2), val=None)
    style_cell(ws.Cells(3, 2),
               val=f"공정 {proc['no']}",
               font_size=11, bold=True, font_color=COLOR["yellow_accent"],
               hcenter=False)
    ws.Rows(3).RowHeight = 22

    ws.Range("B4:H4").Merge()
    style_cell(ws.Cells(4, 2), val=proc["name"],
               font_size=22, bold=True, font_color=COLOR["navy_dark"],
               hcenter=False)
    ws.Rows(4).RowHeight = 36

    # 메타 정보 한 줄 (요구레벨 / 신규자 / 사수 / 시작일)
    ws.Range("B5:C5").Merge()
    style_cell(ws.Cells(5, 2), val=f"요구레벨  {proc['level']}",
               font_size=11, bold=True, font_color=COLOR["navy"],
               bg=COLOR["blue_light"], hcenter=True)
    ws.Range("D5:E5").Merge()
    style_cell(ws.Cells(5, 4), val="신규자  __________",
               font_size=11, font_color=COLOR["gray_700"], hcenter=True)
    ws.Range("F5:G5").Merge()
    style_cell(ws.Cells(5, 6), val="사수  __________",
               font_size=11, font_color=COLOR["gray_700"], hcenter=True)
    ws.Range("H5:H5").Merge()
    style_cell(ws.Cells(5, 8), val="26 / __ / __",
               font_size=10, font_color=COLOR["gray_500"], hcenter=True)
    ws.Rows(5).RowHeight = 24

    # 헤더 하단 구분선 (얇은 네이비 줄)
    ws.Range("B6:H6").Merge()
    ws.Cells(6, 2).Interior.Color = COLOR["navy"]
    ws.Rows(6).RowHeight = 3
    ws.Cells(6, 2).Borders.LineStyle = 0

    # 여백
    ws.Rows(7).RowHeight = 12

    r = 8

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ① 이 공정은 무엇을 하는가
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "①", "이 공정은 무엇을 하는가", COLOR["yellow_accent"])
    r += 1
    desc = (
        f"SP3M3 라인의 {proc['no']}번 공정으로, '{proc['name']}' 작업을 담당합니다.\n"
        f"이 라인은 자동차 안전벨트(Seat Belt Assembly)를 만들며, "
        f"단 1개의 이종·누락이 차량 1대 리콜로 이어집니다.\n"
        f"빠르게보다 정확하게 — 의심되면 사수에게 즉시 확인."
    )
    ws.Range(ws.Cells(r, 2), ws.Cells(r + 2, 8)).Merge()
    style_cell(ws.Cells(r, 2), val=desc,
               font_size=11, font_color=COLOR["gray_700"],
               bg=COLOR["white"], border_color=COLOR["gray_200"])
    for rr in range(r, r + 3):
        ws.Rows(rr).RowHeight = 22
    r += 3
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ② 사용 부품
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ov = PROC_OVERRIDES.get(proc["no"], {})
    parts_label = ov.get("parts_label", "사용 부품 (작업 전 이름 외우기)")
    section_header(ws, r, 2, 8, "②", parts_label, COLOR["yellow_accent"])
    r += 1
    parts_str = "      ·      ".join(parts) if parts else "(작업표준서 참조)"
    ws.Range(ws.Cells(r, 2), ws.Cells(r, 8)).Merge()
    style_cell(ws.Cells(r, 2), val=parts_str,
               font_size=14, bold=True, font_color=COLOR["navy"],
               bg=COLOR["blue_card"], hcenter=True,
               border_color=COLOR["gray_200"])
    ws.Rows(r).RowHeight = 40
    r += 1
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ③ 작업 순서
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "③", "작업 순서  (상세 사진은 작업표준서 참조)", COLOR["yellow_accent"])
    r += 1
    steps = ws_data.get("steps") or ["(작업표준서 매칭 누락 — 라인반장 보강 필요)"]
    main_steps = steps  # 슬라이싱 제한 해제 — 모든 단계 표시
    for i, s in enumerate(main_steps, start=1):
        # 번호 원형 배지
        style_cell(ws.Cells(r, 2),
                   val=str(i), font_size=12, bold=True,
                   font_color=COLOR["white"], bg=COLOR["navy"], hcenter=True,
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 3), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 3), val=s,
                   font_size=11, font_color=COLOR["gray_700"],
                   bg=COLOR["white"], border_color=COLOR["gray_200"])
        ws.Rows(r).RowHeight = 24
        r += 1
    if ws_data.get("ref_sheet"):
        ws.Range(ws.Cells(r, 2), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 2),
                   val=f"※ 상세 작업 순서·사진·도면은 현장 작업표준서 시트 [{ws_data['ref_sheet']}] 참조",
                   font_size=10, font_color=COLOR["gray_500"], hcenter=True)
        ws.Cells(r, 2).Borders.LineStyle = 0
        ws.Rows(r).RowHeight = 20
        r += 1
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ④ 합격(OK) vs 불합격(NG)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "④", "합격(OK) vs 불합격(NG) — 핵심 확인 사항", COLOR["yellow_accent"])
    r += 1
    # 표 헤더
    style_cell(ws.Cells(r, 2), val="NO", font_size=11, bold=True,
               font_color=COLOR["white"], bg=COLOR["navy"], hcenter=True,
               border_color=COLOR["navy"])
    ws.Range(ws.Cells(r, 3), ws.Cells(r, 5)).Merge()
    style_cell(ws.Cells(r, 3), val="확인할 것", font_size=11, bold=True,
               font_color=COLOR["white"], bg=COLOR["navy"], hcenter=True,
               border_color=COLOR["navy"])
    ws.Range(ws.Cells(r, 6), ws.Cells(r, 7)).Merge()
    style_cell(ws.Cells(r, 6), val="✓  OK", font_size=11, bold=True,
               font_color=COLOR["white"], bg=COLOR["green"], hcenter=True,
               border_color=COLOR["green"])
    ws.Cells(r, 8).Value = "✗  NG"
    style_cell(ws.Cells(r, 8), font_size=11, bold=True,
               font_color=COLOR["white"], bg=COLOR["red"], hcenter=True,
               border_color=COLOR["red"])
    ws.Rows(r).RowHeight = 28
    r += 1

    if eval_data:
        items = eval_data["items"][:6]
        crits = eval_data["criteria"][:6]
        for i, (it, (q, _u, y)) in enumerate(zip(items, crits), start=1):
            style_cell(ws.Cells(r, 2), val=i, font_size=11, bold=True,
                       font_color=COLOR["gray_500"], hcenter=True,
                       border_color=COLOR["gray_200"])
            ws.Range(ws.Cells(r, 3), ws.Cells(r, 5)).Merge()
            style_cell(ws.Cells(r, 3), val=it or "", font_size=10,
                       font_color=COLOR["gray_700"], bg=COLOR["white"],
                       border_color=COLOR["gray_200"])
            ws.Range(ws.Cells(r, 6), ws.Cells(r, 7)).Merge()
            style_cell(ws.Cells(r, 6), val=q or "", font_size=10, bold=True,
                       font_color=COLOR["green"], bg=COLOR["green_bg"],
                       border_color=COLOR["gray_200"])
            style_cell(ws.Cells(r, 8), val=y or "", font_size=10, bold=True,
                       font_color=COLOR["red"], bg=COLOR["red_bg"],
                       border_color=COLOR["gray_200"])
            ws.Rows(r).RowHeight = 36
            r += 1
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ⑤ 꼭 지킬 것 (안전·이종)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "⑤", "꼭 지킬 것 — 안전·이종·일상점검", COLOR["red"])
    r += 1
    safety = [
        ("⚙", "시업 시 설비 일상점검표 작성 (에어압 0.4~0.6 MPa)"),
        ("🔄", "차종 변경 시 초물표 확인 후 첫 5개는 사수가 검사"),
        ("🧤", "장갑·귀마개 착용  /  회전부·압입부 손가락 진입 금지"),
        ("🛑", "이상 발생 시 즉시 작업 중지 후 관리자 보고"),
        ("⚠", "의심 시 라인 멈추는 것이 NG 흘려보내는 것보다 낫다"),
    ]
    for icon, text in safety:
        style_cell(ws.Cells(r, 2), val=icon, font_size=13,
                   font_color=COLOR["yellow_accent"], bg=COLOR["yellow_bg"],
                   hcenter=True, border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 3), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 3), val=text, font_size=11,
                   font_color=COLOR["gray_700"], bg=COLOR["yellow_bg"],
                   border_color=COLOR["gray_200"])
        ws.Rows(r).RowHeight = 24
        r += 1
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ⑥ 자주 하는 실수 (사수 작성)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "⑥", "신규자가 자주 하는 실수  (사수 작성)", COLOR["yellow_accent"])
    r += 1
    if eval_data:
        ng_cases = [(it, y) for it, (_q, _u, y) in zip(eval_data["items"], eval_data["criteria"]) if y][:3]
    else:
        ng_cases = []
    for i, (it, y) in enumerate(ng_cases, start=1):
        style_cell(ws.Cells(r, 2), val=i, font_size=11, bold=True,
                   font_color=COLOR["gray_500"], hcenter=True,
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 3), ws.Cells(r, 5)).Merge()
        style_cell(ws.Cells(r, 3), val=(it or "")[:40], font_size=10,
                   font_color=COLOR["gray_700"], bg=COLOR["white"],
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 6), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 6), val=f"자주 발생: {y}",
                   font_size=10, font_color=COLOR["red"], bg=COLOR["red_bg"],
                   border_color=COLOR["gray_200"])
        ws.Rows(r).RowHeight = 28
        r += 1
    # 빈칸 (사수 추가 기입)
    for i in range(len(ng_cases) + 1, 6):
        style_cell(ws.Cells(r, 2), val=i, font_size=11, bold=True,
                   font_color=COLOR["gray_500"], hcenter=True,
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 3), ws.Cells(r, 5)).Merge()
        style_cell(ws.Cells(r, 3), val="", font_size=10,
                   bg=COLOR["white"], border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 6), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 6), val="(사수 추가 기입)",
                   font_size=9, font_color=COLOR["gray_500"],
                   bg=COLOR["gray_100"], border_color=COLOR["gray_200"])
        ws.Rows(r).RowHeight = 26
        r += 1
    ws.Rows(r).RowHeight = 10
    r += 1

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ⑦ 2주 학습 일정
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    section_header(ws, r, 2, 8, "⑦", "2주 학습 일정", COLOR["yellow_accent"])
    r += 1
    schedule = [
        ("1일차", "공정 개요 · 안전 · 부품 이름 숙지", "사수 옆 참관"),
        ("3일차", "작업 순서 따라하기", "사수 1:1 시범"),
        ("1주차", "독립 작업 (사수 감독)", "표준 C/T 내 작업"),
        ("2주차", "자주검사 단독 판정", "사수 1차 일치 확인"),
        ("2주말", f"{proc['level']} 도달 평가 → 단독 작업 인증", "사수 사인 + 라인장 확인"),
    ]
    for stage, std, mth in schedule:
        style_cell(ws.Cells(r, 2), val=stage, font_size=11, bold=True,
                   font_color=COLOR["navy"], bg=COLOR["blue_light"], hcenter=True,
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 3), ws.Cells(r, 5)).Merge()
        style_cell(ws.Cells(r, 3), val=std, font_size=11,
                   font_color=COLOR["gray_700"], bg=COLOR["white"],
                   border_color=COLOR["gray_200"])
        ws.Range(ws.Cells(r, 6), ws.Cells(r, 8)).Merge()
        style_cell(ws.Cells(r, 6), val=mth, font_size=10,
                   font_color=COLOR["gray_500"], bg=COLOR["white"],
                   border_color=COLOR["gray_200"])
        ws.Rows(r).RowHeight = 24
        r += 1

    # 사인란
    ws.Rows(r).RowHeight = 12
    r += 1
    ws.Range(ws.Cells(r, 2), ws.Cells(r, 8)).Merge()
    style_cell(ws.Cells(r, 2),
               val="사수 사인 ___________________      라인장 확인 ___________________      날짜  26 / __ / __",
               font_size=10, font_color=COLOR["gray_500"], hcenter=True)
    ws.Cells(r, 2).Borders.LineStyle = 0
    ws.Rows(r).RowHeight = 22

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # 인쇄 설정
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws.PageSetup.Orientation = 1  # portrait (세로 — A4 1~2장)
    ws.PageSetup.PaperSize = 9  # A4
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False
    ws.PageSetup.LeftMargin = 20
    ws.PageSetup.RightMargin = 20
    ws.PageSetup.TopMargin = 24
    ws.PageSetup.BottomMargin = 24
    ws.PageSetup.CenterHorizontally = True
    ws.PageSetup.PrintGridlines = False


def main():
    print("[1/3] 데이터 로드...")
    procs = load_processes()
    ef = load_eval_forms()

    OUTDIR.mkdir(parents=True, exist_ok=True)
    print("[2/3] Excel COM...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    # 기존 신규자 교육자료 파일 일괄 정리 (분할본·통합본 모두)
    for old in OUTDIR.glob("SP3M3*신규입사자 교육자료.xlsx"):
        old.unlink()

    out_path = (OUTDIR / "SP3M3_신규입사자 교육자료.xlsx").resolve()
    print(f"[3/3] 단일 파일로 6공정 시트 빌드 → {out_path.name}")

    wb = excel.Workbooks.Add()
    while wb.Sheets.Count > 1:
        wb.Sheets(wb.Sheets.Count).Delete()

    first = True
    for proc in procs:
        ws_data = load_workstd_steps(proc["no"])
        parts = extract_parts(proc, ef.get(proc["no"]), ws_data)

        if first:
            sheet = wb.Sheets(1)
            first = False
        else:
            sheet = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))

        build_one(sheet, proc, ef.get(proc["no"]), ws_data, parts)

        # 시트명: 공정번호 + 짧은 공정명 (Excel 31자 제한 + 금지문자 제거)
        short_name = proc["name"].split("&")[0].strip()
        if len(short_name) > 20:
            short_name = short_name[:20]
        sheet_name = f"{proc['no']}.{short_name}"
        sheet_name = re.sub(r'[:\\/?*\[\]]', "", sheet_name)[:31]
        sheet.Name = sheet_name
        print(f"  ✓ 시트 '{sheet_name}' 추가")

    # 첫 시트 활성화
    wb.Sheets(1).Activate()
    wb.SaveAs(str(out_path), FileFormat=xlOpenXMLWorkbook)
    wb.Close(False)

    excel.ScreenUpdating = True
    excel.Quit()
    print(f"\n완료 — {out_path}")


if __name__ == "__main__":
    main()
