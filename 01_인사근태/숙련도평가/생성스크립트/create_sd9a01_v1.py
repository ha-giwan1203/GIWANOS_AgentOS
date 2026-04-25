"""
SD9A01 공정별 숙련도 평가서 v1 — 11개 공정

베이스: create_sp3m3_v4.py 패턴 (샘플 복사 → 공정 특화 9항목 갱신)
- 작업표준서 통합본의 시트별 "자주검사 항목" → std (5개)
- 관리계획서의 공정별 관리항목 → ctrl (4개)
- ERP 표준공정 매핑 (SD9A01_001 ~ SD9A01_011)

실행:
    python create_sd9a01_v1.py
"""
import os
import shutil
import re
from openpyxl import load_workbook

BASE      = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
SAMPLE    = os.path.join(BASE, '공정 평가서표준_SP3M3_샘플.xlsx')
MERGED    = os.path.join(BASE, 'SD9A01_표준문서', 'SD9A01 작업표준서 통합_Rev16.xlsx')
MGRPLAN   = os.path.join(BASE, 'SD9A01_표준문서', 'SD9A01 관리계획서-관리항목및기준목록.xlsx')
OUT_DIR   = os.path.join(BASE, 'SD9A01_공정별 평가서')

# ─── ERP 11공정 매핑 (사용자 화면 캡처 기준) ───────────────────────────────
# (erp_no, erp_name, merged_sheet, std_filter, default_level)
# std_filter: 'all' = 시트 자주검사 전체 / 'package' = 바코드 항목만 / 'main' = 바코드 외
ERP_PROCESSES = [
    ('001', 'PACKAGE 바코드 부착',                                                  '20', 'package', 3),
    ('002', 'RETRACTOR MAIN FRAME 릴 상단 브라켓 압입 & WEBBING GUIDE',           '20', 'main',    3),
    ('003', 'RETRACTOR MAIN FRAME 서브 앗세이 압입 & 리벳 & 릴 하단 브라켓',     '21', 'all',     3),
    ('004', 'WEBBING 웨빙 서브 앗세이 로딩',                                        '30', 'all',     3),
    ('005', '틸트락 작동 검사',                                                     '40', 'all',     3),
    ('006', 'TONGUE 레이저 마킹 조립 & D-RING 서브 앗세이',                         '50', 'all',     3),
    ('007', 'ANCHOR 서브 앗세이 조립 & ANCHOR 앵커 커버',                           '60', 'all',     3),
    ('008', 'WEBBING 실 미싱',                                                      '70', 'all',     3),
    ('009', 'TONGUE STOPPER 텅 스토퍼 조립',                                        '80', 'all',     3),
    ('010', '최종 검사',                                                            '90', 'all',     4),
    ('011', 'PACKAGE 부품식별표 포장',                                              '100','all',     3),
]

# ERP_no ↔ 관리계획서 공정번호 (시트가 다르므로 명시 매핑)
MGRPLAN_PROC_NO_MAP = {
    '001': 10,    # PACKAGE 바코드
    '002': 20,
    '003': None,  # 관리계획서에 21 없음 → 작업표준서 시트 21 자주검사로 대체
    '004': 30,
    '005': 40,
    '006': 50,
    '007': 60,
    '008': 70,
    '009': 80,
    '010': 90,
    '011': 100,
}

# ─── 평가서 셀 좌표 (SP3M3 v4와 동일) ──────────────────────────────────────
COL_PROC1_NO   = 14     # N5
COL_PROC1_NAME = 19     # S5
COL_PROC1_ITEM = 5      # E
COL_PROC1_Q    = 17     # Q
COL_PROC1_U    = 21     # U
COL_PROC1_Y    = 25     # Y

COL_PROC2_NO, COL_PROC2_NAME, COL_PROC2_ITEM = 45, 50, 36
COL_PROC2_Q, COL_PROC2_U, COL_PROC2_Y = 48, 52, 56
COL_PROC3_NO, COL_PROC3_NAME, COL_PROC3_ITEM = 76, 81, 67
COL_PROC3_Q, COL_PROC3_U, COL_PROC3_Y = 79, 83, 87

ROW_PROC_INFO = 5
ROWS_STD  = list(range(13, 18))   # 5행
ROWS_CTRL = list(range(18, 22))   # 4행


# ─── 평가기준 자동 생성 ────────────────────────────────────────────────────
# 자주검사 항목 텍스트에 따라 Q/U/Y 평가기준을 자동 채움
def gen_quy(item_text: str, criterion: str = '') -> tuple:
    """N열 항목 텍스트 + Q열 기준 → SP3M3 패턴의 Q/U/Y 평가기준 3개"""
    s = item_text or ''
    # 키워드 기반 분기
    if any(k in s for k in ('이종', '오조립', '누락')):
        return ('100% 확인 / 이종 없음', '80% 이상 확인', '확인 미흡')
    if any(k in s for k in ('압력', '범위', '에어')):
        return ('범위 내 100%', '경계값 운전', '범위 이탈')
    if any(k in s for k in ('상태', '조립', '부착')):
        return ('정위치 100% 확인', '경미한 조정 1~2회', '조립 불량 다수')
    if any(k in s for k in ('검사', '확인', '점검')):
        return ('전수 검사 수행', '샘플 검사', '검사 누락')
    if any(k in s for k in ('각도', '치수', '길이', '사양')):
        return ('사양 100% 일치', '사양 80% 일치', '사양 불일치')
    # default
    return ('항상 준수', '대체로 준수', '준수 미흡')


def make_eval_question(item_text: str) -> str:
    """N열 항목 텍스트 → '~을 확인하고 있는가?' 형식의 평가사항"""
    s = (item_text or '').strip()
    if not s:
        return ''
    # 이미 의문문 형태면 그대로
    if s.endswith('?') or '있는가' in s:
        return s
    return f'{s} 상태를 확인하고 있는가?'


def make_ctrl_question(category: str, item: str) -> str:
    """관리계획서 분류/관리항목 → '~을 관리기준대로 확인하는가?'"""
    item = (item or '').strip()
    if not item:
        return ''
    cat = (category or '').strip()
    prefix = f'[{cat}] ' if cat else ''
    return f'{prefix}{item}을(를) 관리기준대로 확인하고 있는가?'


# ─── 통합본 시트 자주검사 추출 ──────────────────────────────────────────────
HEADER_PAT = re.compile(r'자\s*주\s*검\s*사\s*항\s*목')


def extract_std_items(ws):
    """통합본 작업표준서 시트에서 자주검사 항목 추출.
    반환: list of (no, item_text, criterion, period)
    """
    # 1) "자주검사 항목" 헤더 찾기
    header_row = None
    for r in range(1, min(ws.max_row+1, 60)):
        for c in range(1, min(ws.max_column+1, 30)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and HEADER_PAT.search(v.replace(' ', '')):
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        return []

    # 2) 헤더 다음 +1 = 컬럼명 행 (NO/항목/기준/주기), +2부터 데이터
    items = []
    for r in range(header_row+2, header_row+30):
        if r > ws.max_row:
            break
        # M열(13)에 NO, N열(14)에 항목, Q열(17)에 기준, W열(23)에 주기
        # 단, 일부 시트는 다를 수 있으니 row 13~14 위치를 검사
        no_v = ws.cell(row=r, column=13).value
        if no_v is None:
            # 빈 행이면 종료 후보
            # 단, 1번 항목 직후 빈 행이 아니라면 계속
            if items and len(items) >= 1:
                # 다음 섹션 시작 가능
                continue
            continue
        # NO가 숫자거나 "특별 / 특성" 같은 특수 텍스트면 항목으로 간주
        no_str = str(no_v).strip()
        # 다음 섹션 헤더 도달 시 중단
        if any(stop in no_str for stop in ('공   정   작   업   순   서', '불 량 품', '공 정')):
            break
        item_text = ws.cell(row=r, column=14).value or ''
        criterion = ws.cell(row=r, column=17).value or ''
        period    = ws.cell(row=r, column=23).value or ''
        item_text = str(item_text).strip()
        criterion = str(criterion).strip()
        if not item_text:
            continue
        items.append((no_str, item_text, criterion, period))
        if len(items) >= 10:
            break

    return items


# ─── 관리계획서 공정별 관리항목 추출 ────────────────────────────────────────
def load_mgrplan_items():
    """관리계획서 → {공정번호: [(분류, 관리항목, 기준, 주기), ...]}"""
    wb = load_workbook(MGRPLAN, data_only=True)
    ws = wb.active
    by_proc = {}
    cur = None
    for r in range(2, ws.max_row+1):
        no = ws.cell(row=r, column=1).value
        nm = ws.cell(row=r, column=2).value
        if no is not None:
            cur = int(no)
            by_proc.setdefault(cur, [])
        if cur is None:
            continue
        cat = ws.cell(row=r, column=4).value or ''   # D열 분류
        item= ws.cell(row=r, column=5).value or ''   # E열 관리항목
        std = ws.cell(row=r, column=8).value or ''   # H열 기준
        prd = ws.cell(row=r, column=10).value or ''  # J열 주기
        item = str(item).strip()
        if not item:
            continue
        by_proc[cur].append((str(cat).strip(), item, str(std).strip(), str(prd).strip()))
    wb.close()
    return by_proc


# ─── std/ctrl 보강 항목 (항목 부족 시 사용) ─────────────────────────────────
GENERIC_STD = [
    ('이종 여부를 확인하고 있는가? (차종별/초물표 참조)',
     '이종 100% 확인', '이종 80% 이상 확인', '이종 확인 미흡'),
    ('설비 시업점검을 수행하고 있는가?',
     '항상 점검', '대체로 점검', '점검 누락'),
    ('작업 공구·치공구 누락 여부를 확인하는가?',
     '누락 0건', '경미한 누락', '누락 다수'),
    ('표준 작업 순서를 준수하고 있는가?',
     '순서 100% 준수', '일부 어긋남', '순서 미준수'),
    ('이상 발생 시 즉시 라인관리자에 보고하는가?',
     '즉시 보고', '지연 보고', '보고 누락'),
]

GENERIC_CTRL = [
    ('관리기준에 따라 정기 점검을 실시하고 있는가?',
     '주기 100% 준수', '지연 발생', '점검 누락'),
    ('부적합품을 관리계획서 절차대로 처리하는가?',
     '절차대로 처리', '처리 지연', '처리 미흡'),
    ('관리한계 이탈 시 즉시 조치를 취하는가?',
     '즉시 조치 100%', '조치 지연', '조치 미흡'),
    ('관리 기록을 정확히 작성·보관하고 있는가?',
     '기록 100% 작성', '일부 누락', '기록 미흡'),
]


# ─── PROCESSES 구성 ────────────────────────────────────────────────────────
def build_std(merged_wb, sheet_no: str, std_filter: str) -> list:
    """5개 std 항목 반환. 부족하면 GENERIC_STD로 보강."""
    if sheet_no not in merged_wb.sheetnames:
        print(f'  [WARN] 통합본에 시트 [{sheet_no}] 없음 → GENERIC_STD 사용')
        return GENERIC_STD[:5]

    ws = merged_wb[sheet_no]
    raw_items = extract_std_items(ws)

    # std_filter에 따라 분리 (시트 20에서 SD9A01_001은 바코드 / SD9A01_002는 그 외)
    package_items = [it for it in raw_items if '바코드' in it[1]]
    main_items    = [it for it in raw_items if '바코드' not in it[1]]

    if std_filter == 'package':
        chosen = package_items
    elif std_filter == 'main':
        chosen = main_items
    else:
        chosen = raw_items

    out = []
    for (no, item, crit, prd) in chosen[:5]:
        eval_q = make_eval_question(item)
        q, u, y = gen_quy(item, crit)
        out.append((eval_q, q, u, y))

    # 부족한 만큼 GENERIC_STD에서 보강
    while len(out) < 5:
        idx = len(out)
        if idx < len(GENERIC_STD):
            out.append(GENERIC_STD[idx])
        else:
            out.append(('-', '-', '-', '-'))
    return out[:5]


def build_ctrl(mgrplan_by_proc: dict, erp_no: str) -> list:
    """관리계획서에서 4개 ctrl 추출. 부족하면 GENERIC_CTRL로 보강."""
    proc_no = MGRPLAN_PROC_NO_MAP.get(erp_no)
    items = mgrplan_by_proc.get(proc_no, []) if proc_no else []

    out = []
    for (cat, item, std, prd) in items[:4]:
        q_text = make_ctrl_question(cat, item)
        if cat in ('특별', '특성', '특별특성'):
            q, u, y = ('100% 확인 (특별특성)', '95% 이상', '95% 미만')
        else:
            q, u, y = ('관리기준 100% 준수', '일부 미준수', '관리기준 미준수')
        out.append((q_text, q, u, y))

    while len(out) < 4:
        idx = len(out)
        if idx < len(GENERIC_CTRL):
            out.append(GENERIC_CTRL[idx])
        else:
            out.append(('-', '-', '-', '-'))
    return out[:4]


def build_processes() -> list:
    merged_wb = load_workbook(MERGED, data_only=False)
    mgrplan_by_proc = load_mgrplan_items()

    procs = []
    for (erp_no, erp_name, sheet_no, std_filter, level) in ERP_PROCESSES:
        std = build_std(merged_wb, sheet_no, std_filter)
        ctrl = build_ctrl(mgrplan_by_proc, erp_no)
        procs.append({
            'no': erp_no,
            'name': erp_name,
            'level': level,
            'std': std,
            'ctrl': ctrl,
            '_src_sheet': sheet_no,
            '_filter': std_filter,
        })

    merged_wb.close()
    return procs


# ─── 셀 갱신 ────────────────────────────────────────────────────────────────
def safe_write(ws, row, col, value):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def write_item_row(ws, row, item_tuple):
    item, q, u, y = item_tuple
    safe_write(ws, row, COL_PROC1_ITEM, item)
    safe_write(ws, row, COL_PROC1_Q, q)
    safe_write(ws, row, COL_PROC1_U, u)
    safe_write(ws, row, COL_PROC1_Y, y)


def clear_row(ws, row):
    for c in (COL_PROC1_ITEM, COL_PROC1_Q, COL_PROC1_U, COL_PROC1_Y):
        safe_write(ws, row, c, '')


def clear_other_processes(ws):
    for col_no, col_name, col_item, col_q, col_u, col_y in [
        (COL_PROC2_NO, COL_PROC2_NAME, COL_PROC2_ITEM,
         COL_PROC2_Q, COL_PROC2_U, COL_PROC2_Y),
        (COL_PROC3_NO, COL_PROC3_NAME, COL_PROC3_ITEM,
         COL_PROC3_Q, COL_PROC3_U, COL_PROC3_Y),
    ]:
        safe_write(ws, ROW_PROC_INFO, col_no, '')
        safe_write(ws, ROW_PROC_INFO, col_name, '')
        for r in ROWS_STD + ROWS_CTRL:
            safe_write(ws, r, col_item, '')
            safe_write(ws, r, col_q, '')
            safe_write(ws, r, col_u, '')
            safe_write(ws, r, col_y, '')


def fill_sheet(ws, proc):
    safe_write(ws, ROW_PROC_INFO, 4, '대원테크')
    safe_write(ws, ROW_PROC_INFO, 9, 'SD9A01')
    safe_write(ws, ROW_PROC_INFO, COL_PROC1_NO, proc['no'])
    safe_write(ws, ROW_PROC_INFO, COL_PROC1_NAME, proc['name'])

    for i, r in enumerate(ROWS_STD):
        if i < len(proc['std']):
            write_item_row(ws, r, proc['std'][i])
        else:
            clear_row(ws, r)

    for i, r in enumerate(ROWS_CTRL):
        if i < len(proc['ctrl']):
            write_item_row(ws, r, proc['ctrl'][i])
        else:
            clear_row(ws, r)

    clear_other_processes(ws)


# ─── 메인 ──────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    procs = build_processes()
    print(f'[빌드] PROCESSES {len(procs)}개 구성 완료')

    for proc in procs:
        out_path = os.path.join(
            OUT_DIR,
            f"SD9A01_공정{proc['no']} 숙련도 평가서.xlsx"
        )
        shutil.copy(SAMPLE, out_path)
        wb = load_workbook(out_path)
        # 시트명 확인 — 샘플은 "양식" 또는 한글 시트명 1개
        ws = wb[wb.sheetnames[0]]
        fill_sheet(ws, proc)
        wb.save(out_path)
        wb.close()
        print(f"  [생성] SD9A01_{proc['no']:>3} {proc['name'][:30]:30s} "
              f"std={len(proc['std'])} ctrl={len(proc['ctrl'])} (시트{proc['_src_sheet']}/{proc['_filter']})")

    print(f"\n[완료] {OUT_DIR}")


if __name__ == '__main__':
    main()
