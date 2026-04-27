"""
SD9A01 산출물 자체검증 스크립트

검증 항목:
  A. 파일 구조 (수, 시트, 행/열)
  B. 공정별 양식 평가사항 무결성
  C. 메타데이터 (공정번호/명/라인/업체)
  D. 개인별 점수·라벨 검증 (총점 일치, 0점 없음, 주공정 라벨)
  E. 일관성 (수행공정 ↔ 시트 일치)
  F. 신규 SD9A01_003 노출 검증
  G. PROCESSES 데이터 의미 검증 (텍스트 키워드)
  H. SP3M3와 좌표 일관성 (회귀 검증)
"""
import os, sys, glob
import openpyxl
from collections import Counter, defaultdict

BASE = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
DIR_FORM   = os.path.join(BASE, 'SD9A01_공정별 평가서')
DIR_PER_D  = os.path.join(BASE, 'SD9A01_개인별 평가서', 'SD9A01 주간')
DIR_PER_N  = os.path.join(BASE, 'SD9A01_개인별 평가서', 'SD9A01 야간')
DIR_SRC_D  = os.path.join(BASE, 'SD9M01 주간 숙련도 평가서')
DIR_SRC_N  = os.path.join(BASE, 'SD9M01 야간 숙련도 평가서')
MERGED     = os.path.join(BASE, 'SD9A01_표준문서', 'SD9A01 작업표준서 통합_Rev16.xlsx')

# 평가서 양식 좌표
ROW_INFO = 5
COL_BIZ, COL_LINE = 4, 9
COL_P1_NO, COL_P1_NAME, COL_P1_ITEM = 14, 19, 5
COL_P1_Q, COL_P1_U, COL_P1_Y = 17, 21, 25
ROWS_STD  = list(range(13, 18))
ROWS_CTRL = list(range(18, 22))

PROC_COLS  = [14, 46, 78, 110, 142, 174, 206, 238, 270, 302, 334]
TOTAL_COLS = [10, 42, 74, 106, 138, 170, 202, 234, 266, 298, 330]

SD9M01_TO_ERP = {'10':'10','20':'20','21':'21','30':'30','40':'40',
                 '50':'50','60':'60','70':'70','80':'80','90':'90','100':'100'}

ERP_NAMES = {
    '10':  'PACKAGE 바코드 부착',
    '20':  'RETRACTOR MAIN FRAME 릴 상단 브라켓 압입 & WEBBING GUIDE',
    '21':  'RETRACTOR MAIN FRAME 서브 앗세이 압입 & 리벳 & 릴 하단 브라켓',
    '30':  'WEBBING 웨빙 서브 앗세이 로딩',
    '40':  '틸트락 작동 검사',
    '50':  'TONGUE 레이저 마킹 조립 & D-RING 서브 앗세이',
    '60':  'ANCHOR 서브 앗세이 조립 & ANCHOR 앵커 커버',
    '70':  'WEBBING 실 미싱',
    '80':  'TONGUE STOPPER 텅 스토퍼 조립',
    '90':  '최종 검사',
    '100': 'PACKAGE 부품식별표 포장',
}
EXPECTED_CODES = ['10','20','21','30','40','50','60','70','80','90','100']

issues = []  # (severity, code, file, msg)

def add_issue(sev, code, f, msg):
    issues.append((sev, code, os.path.basename(f) if f else '', msg))

def get_merged_value(ws, r, c):
    """병합 셀 처리 — 병합 영역의 좌상단 값 반환"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return ws.cell(row=r, column=c).value


# ─── A. 파일 구조 ─────────────────────────────────────────────────────────
def check_a_structure():
    print('\n=== A. 파일 구조 ===')
    forms = sorted(glob.glob(os.path.join(DIR_FORM, '*.xlsx')))
    pers_d = sorted(glob.glob(os.path.join(DIR_PER_D, '*.xlsx')))
    pers_n = sorted(glob.glob(os.path.join(DIR_PER_N, '*.xlsx')))

    if len(forms) != 11:
        add_issue('ERR', 'A1', None, f'공정별 양식 파일 수 11 ≠ {len(forms)}')
    if len(pers_d) != 12:
        add_issue('WARN', 'A2', None, f'개인별 주간 12 ≠ {len(pers_d)}')
    if len(pers_n) != 11:
        add_issue('WARN', 'A3', None, f'개인별 야간 11 ≠ {len(pers_n)}')
    print(f'  공정별 {len(forms)} / 개인별 주간 {len(pers_d)} / 야간 {len(pers_n)}')

    if not os.path.exists(MERGED):
        add_issue('ERR', 'A4', None, f'통합본 미존재: {MERGED}')
    return forms, pers_d, pers_n


# ─── B/C. 공정별 양식 평가사항 + 메타 ──────────────────────────────────────
def check_bc_forms(forms):
    print('\n=== B/C. 공정별 양식 평가사항 + 메타 ===')
    expected_codes = EXPECTED_CODES
    found_codes = []

    for f in forms:
        wb = openpyxl.load_workbook(f, data_only=False)
        if len(wb.sheetnames) != 1:
            add_issue('WARN', 'B1', f, f'시트 수 1 ≠ {len(wb.sheetnames)}')
        ws = wb[wb.sheetnames[0]]
        if ws.max_row != 37:
            add_issue('WARN', 'B2', f, f'max_row 37 ≠ {ws.max_row}')
        if ws.max_column != 95:
            add_issue('WARN', 'B3', f, f'max_column 95 ≠ {ws.max_column}')

        # 메타
        biz = get_merged_value(ws, ROW_INFO, COL_BIZ)
        line = get_merged_value(ws, ROW_INFO, COL_LINE)
        no = get_merged_value(ws, ROW_INFO, COL_P1_NO)
        name = get_merged_value(ws, ROW_INFO, COL_P1_NAME)
        if biz != '대원테크':
            add_issue('WARN', 'C1', f, f'D5 업체 {biz!r} ≠ 대원테크')
        if line != 'SD9A01':
            add_issue('ERR', 'C2', f, f'I5 라인 {line!r} ≠ SD9A01')
        if not no or str(no) not in expected_codes:
            add_issue('ERR', 'C3', f, f'N5 공정번호 {no!r} 미인식')
        else:
            found_codes.append(str(no))
        if name != ERP_NAMES.get(str(no)):
            add_issue('WARN', 'C4', f, f'S5 공정명 {name!r} ≠ ERP_NAMES[{no}]')

        # 평가사항 무결성 (모든 9행 채워짐?)
        empty_std = []
        for r in ROWS_STD:
            v = get_merged_value(ws, r, COL_P1_ITEM)
            if not v or str(v).strip() in ('', '-'):
                empty_std.append(r)
        if empty_std:
            add_issue('WARN', 'B4', f, f'std 빈 행 {empty_std}')

        empty_ctrl = []
        for r in ROWS_CTRL:
            v = get_merged_value(ws, r, COL_P1_ITEM)
            if not v or str(v).strip() in ('', '-'):
                empty_ctrl.append(r)
        if empty_ctrl:
            add_issue('WARN', 'B5', f, f'ctrl 빈 행 {empty_ctrl}')

        # Q/U/Y 모두 채워짐?
        for r in ROWS_STD + ROWS_CTRL:
            for col, label in [(COL_P1_Q,'Q'), (COL_P1_U,'U'), (COL_P1_Y,'Y')]:
                v = get_merged_value(ws, r, col)
                if not v or str(v).strip() in ('', '-'):
                    add_issue('WARN', 'B6', f, f'r{r} {label}열 빈 셀')

        wb.close()

    missing = [c for c in expected_codes if c not in found_codes]
    if missing:
        add_issue('ERR', 'C5', None, f'누락 공정 코드: {missing}')
    print(f'  검사된 공정 코드: {sorted(found_codes)}')


# ─── D/E/F. 개인별 점수·라벨·일관성 ────────────────────────────────────────
def get_original_total(name, shift, sd_no):
    """원본 SD9M01 평가서에서 (이름, 공정번호)의 총점 조회"""
    src_dir = DIR_SRC_D if shift == '주간' else DIR_SRC_N
    path = os.path.join(src_dir, f'{name} 숙련도 평가서 (26.04.01).xlsx')
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    for pc, tc in zip(PROC_COLS, TOTAL_COLS):
        if pc > ws.max_column:
            break
        v = ws.cell(ROW_INFO, pc).value
        if v is None:
            continue
        if str(v).strip() == sd_no:
            t = ws.cell(32, tc).value
            wb.close()
            return t
    wb.close()
    return None


def check_def_personal(pers_d, pers_n):
    print('\n=== D/E/F. 개인별 검증 ===')
    erp_to_sd = {v: k for k, v in SD9M01_TO_ERP.items()}

    has_21 = []  # 공정21 시트 보유자 (박태순만 있어야 함, 신규 공정)

    for path in pers_d + pers_n:
        shift = '주간' if path in pers_d else '야간'
        name = os.path.basename(path).replace(' 숙련도평가서.xlsx', '')
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = wb.sheetnames

        # 첫 시트 = 주공정?
        first_z3 = wb[sheets[0]].cell(3, 26).value
        if first_z3 != '주공정':
            add_issue('ERR', 'D1', path, f'첫 시트 Z3 {first_z3!r} ≠ 주공정')

        # 나머지 = 전환공정?
        for sn in sheets[1:]:
            z3 = wb[sn].cell(3, 26).value
            if z3 != '전환공정':
                add_issue('ERR', 'D2', path, f'{sn} Z3 {z3!r} ≠ 전환공정')

        # 공정21(신규) 노출 검증
        if '공정21' in sheets:
            has_21.append(name)

        # 각 시트 검증
        for sn in sheets:
            ws = wb[sn]
            erp_no = sn.replace('공정', '')
            sd_no = erp_to_sd.get(erp_no)

            # 작업자 정보
            v_name = ws.cell(2, 22).value
            v_emp = ws.cell(3, 22).value
            v_date = ws.cell(5, 28).value
            v_line = ws.cell(5, 9).value
            v_no = ws.cell(5, 14).value
            v_pname = ws.cell(5, 19).value
            if v_name != name:
                add_issue('WARN', 'D3', path, f'{sn} V2 이름 {v_name!r} ≠ {name}')
            if not v_emp:
                add_issue('WARN', 'D4', path, f'{sn} V3 사번 비어있음')
            if v_line != 'SD9A01':
                add_issue('ERR', 'D5', path, f'{sn} I5 라인 {v_line!r} ≠ SD9A01')
            if str(v_no) != erp_no:
                add_issue('WARN', 'D6', path, f'{sn} N5 {v_no!r} ≠ {erp_no}')
            if v_pname != ERP_NAMES.get(erp_no):
                add_issue('WARN', 'D7', path, f'{sn} S5 공정명 불일치')

            # AC열 점수 (col 29, r9~r28 = 20개)
            ac_vals = [ws.cell(r, 29).value for r in range(9, 29)]
            ac_nums = [int(v) for v in ac_vals if isinstance(v, (int, float))]
            ac_sum = sum(ac_nums)
            ac_min = min(ac_nums) if ac_nums else 0
            ac_max = max(ac_nums) if ac_nums else 0
            if ac_min == 0:
                add_issue('WARN', 'D8', path, f'{sn} AC열 0점 발견')
            if ac_max > 5:
                add_issue('WARN', 'D9', path, f'{sn} AC열 5점 초과')
            if len(ac_nums) != 20:
                add_issue('WARN', 'D10', path, f'{sn} AC열 항목 {len(ac_nums)} ≠ 20')

            # 원본 총점 일치
            orig_total = get_original_total(name, shift, sd_no)
            if orig_total is not None and isinstance(orig_total, (int, float)):
                if ac_sum != int(orig_total):
                    add_issue('ERR', 'D11', path, f'{sn} AC합 {ac_sum} ≠ 원본 {int(orig_total)}')

        wb.close()

    # F. 공정21(신규) 노출 — 박태순만 가져야 함
    expected_21_holders = ['박태순']
    if sorted(has_21) != sorted(expected_21_holders):
        add_issue('ERR', 'F1', None,
                  f'공정21(신규) 시트 보유자 {has_21} ≠ {expected_21_holders}')
    print(f'  공정21(신규) 보유자: {has_21}')


# ─── G. PROCESSES 데이터 의미 검증 ────────────────────────────────────────
def check_g_processes(forms):
    print('\n=== G. PROCESSES 데이터 의미 검증 ===')
    f10 = os.path.join(DIR_FORM, 'SD9A01_공정10 숙련도 평가서.xlsx')
    f20 = os.path.join(DIR_FORM, 'SD9A01_공정20 숙련도 평가서.xlsx')
    f21 = os.path.join(DIR_FORM, 'SD9A01_공정21 숙련도 평가서.xlsx')

    # G1 — 공정10 (PACKAGE 바코드, 시트 21 자주검사) "바코드" 키워드
    wb1 = openpyxl.load_workbook(f10)
    ws1 = wb1[wb1.sheetnames[0]]
    std1_text = ' '.join(str(get_merged_value(ws1, r, COL_P1_ITEM) or '')
                         for r in ROWS_STD)
    if '바코드' not in std1_text:
        add_issue('WARN', 'G1', f10, '공정10 std에 "바코드" 키워드 없음 (시트 21 추출 실패 가능성)')
    print(f'  공정10 std "바코드" 포함 {"OK" if "바코드" in std1_text else "MISS"}')
    wb1.close()

    # G2 — 공정20 (RETRACTOR MAIN FRAME, 시트 20 자주검사) — 실드/스테이/WEBBING/어퍼 키워드
    wb2 = openpyxl.load_workbook(f20)
    ws2 = wb2[wb2.sheetnames[0]]
    std2_text = ' '.join(str(get_merged_value(ws2, r, COL_P1_ITEM) or '')
                         for r in ROWS_STD)
    if not any(k in std2_text for k in ('실드', '스테이', 'WEBBING', '어퍼')):
        add_issue('WARN', 'G2', f20, '공정20 std에 시트 20 키워드(실드/스테이/WEBBING) 없음')
    print(f'  공정20 std 시트20 키워드 {"OK" if any(k in std2_text for k in ("실드","스테이","WEBBING","어퍼")) else "MISS"}')
    wb2.close()

    # G3 — 공정21 (신규, OVERRIDE_STD) "리벳/브라켓/RETRACTOR" 키워드
    wb3 = openpyxl.load_workbook(f21)
    ws3 = wb3[wb3.sheetnames[0]]
    std3_text = ' '.join(str(get_merged_value(ws3, r, COL_P1_ITEM) or '')
                         for r in ROWS_STD)
    if not any(k in std3_text for k in ('리벳','브라켓','릴 하단','RETRACTOR')):
        add_issue('WARN', 'G3', f21, '공정21 std에 리벳/브라켓 키워드 없음 (OVERRIDE 적용 실패 가능성)')
    print(f'  공정21 std OVERRIDE 키워드 {"OK" if any(k in std3_text for k in ("리벳","브라켓","릴 하단","RETRACTOR")) else "MISS"}')
    wb3.close()


# ─── H. SP3M3와 좌표 일관성 ───────────────────────────────────────────────
def check_h_coord_compat():
    print('\n=== H. SP3M3 양식과 좌표 일관성 ===')
    sp_form = os.path.join(BASE, 'SP3M3_공정별 평가서', 'SP3M3_공정10 숙련도 평가서.xlsx')
    sd_form = os.path.join(DIR_FORM, 'SD9A01_공정10 숙련도 평가서.xlsx')

    if not os.path.exists(sp_form):
        add_issue('INFO', 'H0', sp_form, 'SP3M3 비교 대상 미존재 (선택)')
        return

    wbs = openpyxl.load_workbook(sp_form)
    wbd = openpyxl.load_workbook(sd_form)
    wss = wbs[wbs.sheetnames[0]]
    wsd = wbd[wbd.sheetnames[0]]

    if (wss.max_row, wss.max_column) != (wsd.max_row, wsd.max_column):
        add_issue('WARN', 'H1', sd_form, f'행/열 SP3M3({wss.max_row},{wss.max_column}) ≠ SD9A01({wsd.max_row},{wsd.max_column})')
    else:
        print(f'  행/열 동일 ({wss.max_row}r × {wss.max_column}c) ✓')

    # 병합 영역 수 동일?
    if len(wss.merged_cells.ranges) != len(wsd.merged_cells.ranges):
        add_issue('WARN', 'H2', sd_form, f'병합 셀 수 SP3M3 {len(wss.merged_cells.ranges)} ≠ SD9A01 {len(wsd.merged_cells.ranges)}')
    else:
        print(f'  병합 셀 수 동일 ({len(wss.merged_cells.ranges)}) ✓')

    wbs.close(); wbd.close()


# ─── 메인 ──────────────────────────────────────────────────────────────────
def main():
    print('=' * 70)
    print('SD9A01 산출물 자체검증')
    print('=' * 70)

    forms, pers_d, pers_n = check_a_structure()
    check_bc_forms(forms)
    check_def_personal(pers_d, pers_n)
    check_g_processes(forms)
    check_h_coord_compat()

    print('\n' + '=' * 70)
    print(f'검증 완료. 이슈 {len(issues)}건')
    print('=' * 70)

    if not issues:
        print('  ✓ 문제 없음')
        return 0

    by_sev = defaultdict(list)
    for sev, code, f, msg in issues:
        by_sev[sev].append((code, f, msg))

    for sev in ['ERR', 'WARN', 'INFO']:
        if sev not in by_sev:
            continue
        print(f'\n[{sev}] {len(by_sev[sev])}건')
        # 같은 code끼리 묶어서 표시 (중복 메시지 대량일 수 있음)
        by_code = defaultdict(list)
        for code, f, msg in by_sev[sev]:
            by_code[(code, msg)].append(f)
        for (code, msg), files in by_code.items():
            if len(files) <= 3:
                files_s = ', '.join(files) if files[0] else '(global)'
            else:
                files_s = f'{files[0]} 외 {len(files)-1}건'
            print(f'  [{code}] {msg}  → {files_s}')

    return 1


if __name__ == '__main__':
    sys.exit(main())
