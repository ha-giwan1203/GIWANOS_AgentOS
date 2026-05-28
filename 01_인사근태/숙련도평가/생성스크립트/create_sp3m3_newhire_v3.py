# -*- coding: utf-8 -*-
"""
SP3M3 신규 입사자 교육자료 v3 — 작업표준서가 못 채우는 부분 위주
작업표준서는 현장 비치 → 복사 X
"""
from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import win32com.client as win32

BASE = Path(__file__).resolve().parent.parent
STD = BASE / "SP3M3_표준문서"
PROC_FORM = BASE / "SP3M3_공정별 평가서"
PERSONAL = BASE / "SP3M3_개인별 평가서"
OUTDIR = BASE / "SP3M3_신규입사자 교육자료"

xlOpenXMLWorkbook = 51

# 색상 (Excel ColorIndex)
C_WHITE = 2
C_RED = 3
C_GREEN = 4
C_BLUE = 5
C_YELLOW = 6
C_GRAY = 15
C_DARK_BLUE = 32
C_LIGHT_GREEN = 35
C_LIGHT_RED = 38
C_LIGHT_YELLOW = 36
C_ORANGE = 45


def load_processes():
    wb = openpyxl.load_workbook(STD / "SP3M3 라인별공정목록.xlsx", data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        out.append({"no": int(no), "name": (ws.cell(r, 2).value or "").strip(),
                    "level": (ws.cell(r, 3).value or "Lv.3").strip()})
    wb.close()
    return out


def load_eval_forms():
    out = {}
    for fp in PROC_FORM.glob("SP3M3_공정*.xlsx"):
        m = re.search(r"공정(\d+)", fp.name)
        if not m:
            continue
        no = int(m.group(1))
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        items = [ws.cell(r, 5).value for r in range(13, 22)]
        criteria = [(ws.cell(r, 17).value, ws.cell(r, 21).value, ws.cell(r, 25).value)
                    for r in range(13, 22)]
        out[no] = {"items": items, "criteria": criteria}
        wb.close()
    return out


def load_personnel():
    out = defaultdict(list)
    for shift_dir, shift in [("SP3M3 주간", "주간"), ("SP3M3 야간", "야간")]:
        d = PERSONAL / shift_dir
        if not d.exists():
            continue
        for fp in d.glob("*.xlsx"):
            name = fp.stem.replace(" 숙련도평가서", "").strip()
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                n5 = ws["N5"].value
                z3 = ws["Z3"].value
                if not isinstance(n5, (int, float)):
                    try:
                        n5 = int(str(n5))
                    except Exception:
                        continue
                ac_sum = sum(v for r in range(13, 22)
                             if isinstance(v := ws.cell(r, 29).value, (int, float)))
                out[int(n5)].append({"name": name, "shift": shift, "ac_sum": ac_sum,
                                     "role": (z3 or "").strip() if isinstance(z3, str) else z3})
            wb.close()
    return out


# ─────────────────────────────────────────────
# 시트 빌더
# ─────────────────────────────────────────────
def _h(cell, val, size=14, bold=True, color=C_WHITE, bg=C_DARK_BLUE, hcenter=True):
    cell.Value = val
    cell.Font.Size = size
    cell.Font.Bold = bold
    cell.Font.ColorIndex = color
    if bg:
        cell.Interior.ColorIndex = bg
    cell.HorizontalAlignment = -4108 if hcenter else -4131
    cell.VerticalAlignment = -4108
    cell.WrapText = True
    cell.Borders.LineStyle = 1


def _b(cell, val, size=12, bg=None, color=None, hcenter=False, bold=False):
    cell.Value = val
    cell.Font.Size = size
    cell.Font.Bold = bold
    if color is not None:
        cell.Font.ColorIndex = color
    if bg is not None:
        cell.Interior.ColorIndex = bg
    cell.HorizontalAlignment = -4108 if hcenter else -4131
    cell.VerticalAlignment = -4108
    cell.WrapText = True
    cell.Borders.LineStyle = 1


def _title(ws, txt, span=6, size=20, color=C_DARK_BLUE):
    ws.Cells(1, 1).Value = txt
    ws.Range(ws.Cells(1, 1), ws.Cells(1, span)).Merge()
    c = ws.Cells(1, 1)
    c.Font.Size = size
    c.Font.Bold = True
    c.Font.ColorIndex = color
    c.HorizontalAlignment = -4108
    ws.Rows(1).RowHeight = 38


def build_sheet1_intro(ws, proc, procs_all):
    """공정 한눈에 — 왜 중요한가 + 전체 흐름 + 이 공정 위치"""
    ws.Name = "1.공정한눈에"
    _title(ws, f"공정 {proc['no']} — {proc['name']}", span=7)

    # 왜 중요한가
    ws.Cells(3, 1).Value = "❗ 이 공정이 왜 중요한가"
    ws.Range(ws.Cells(3, 1), ws.Cells(3, 7)).Merge()
    c = ws.Cells(3, 1)
    c.Font.Size = 15
    c.Font.Bold = True
    c.Interior.ColorIndex = C_LIGHT_RED
    c.HorizontalAlignment = -4108
    ws.Rows(3).RowHeight = 32

    msg = (
        "이 라인은 자동차 안전벨트(Seat Belt Assembly)를 만듭니다.\n"
        "사고 발생 시 단 1mm의 조립 불량이 사람의 목숨을 결정합니다.\n"
        f"공정 {proc['no']}에서 단 1개의 이종·누락이 나가면, 차량 1대 전체가 리콜 대상이 됩니다.\n"
        "→ 빠르게보다 정확하게. 의심되면 무조건 사수에게 확인."
    )
    ws.Cells(4, 1).Value = msg
    ws.Range(ws.Cells(4, 1), ws.Cells(4, 7)).Merge()
    c = ws.Cells(4, 1)
    c.Font.Size = 13
    c.WrapText = True
    c.VerticalAlignment = -4108
    c.Borders.LineStyle = 1
    ws.Rows(4).RowHeight = 100

    # 전체 공정 흐름 (이 공정 강조)
    ws.Cells(6, 1).Value = "📍 SP3M3 라인 전체 공정 흐름"
    ws.Range(ws.Cells(6, 1), ws.Cells(6, 7)).Merge()
    c = ws.Cells(6, 1)
    c.Font.Size = 15
    c.Font.Bold = True
    c.Interior.ColorIndex = C_LIGHT_YELLOW
    c.HorizontalAlignment = -4108
    ws.Rows(6).RowHeight = 28

    r = 7
    for i, p in enumerate(procs_all):
        col = i + 1
        cell = ws.Cells(r, col)
        cell.Value = f"공정 {p['no']}\n{p['name'][:14]}"
        cell.Font.Size = 11
        cell.WrapText = True
        cell.HorizontalAlignment = -4108
        cell.VerticalAlignment = -4108
        cell.Borders.LineStyle = 1
        if p["no"] == proc["no"]:
            cell.Interior.ColorIndex = C_YELLOW
            cell.Font.Bold = True
            cell.Font.ColorIndex = C_RED
        else:
            cell.Interior.ColorIndex = C_GRAY
        ws.Columns(col).ColumnWidth = 16
    ws.Rows(r).RowHeight = 60

    # 요구레벨·기본 정보
    ws.Cells(9, 1).Value = "📋 기본 정보"
    ws.Range(ws.Cells(9, 1), ws.Cells(9, 7)).Merge()
    c = ws.Cells(9, 1)
    c.Font.Size = 15
    c.Font.Bold = True
    c.Interior.ColorIndex = C_LIGHT_GREEN
    c.HorizontalAlignment = -4108
    ws.Rows(9).RowHeight = 28

    info = [
        ("공정번호", str(proc["no"])),
        ("공정명", proc["name"]),
        ("라인", "SP3M3 (안전벨트 ASS'Y)"),
        ("요구 작업레벨", f"{proc['level']}  (신규자는 입사 1개월 내 도달)"),
        ("표준 C/T", "(라인반장 기재)"),
        ("주요 부품", "(작업표준서 참조 — 현장 비치)"),
    ]
    rr = 10
    for k, v in info:
        _b(ws.Cells(rr, 1), k, size=12, bg=C_GRAY, hcenter=True, bold=True)
        ws.Range(ws.Cells(rr, 1), ws.Cells(rr, 2)).Merge()
        _b(ws.Cells(rr, 3), v, size=12)
        ws.Range(ws.Cells(rr, 3), ws.Cells(rr, 7)).Merge()
        ws.Rows(rr).RowHeight = 24
        rr += 1

    ws.PageSetup.Orientation = 2  # landscape
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1


def build_sheet2_ok_ng(ws, proc, eval_data):
    """합격 vs 불합격 9포인트 대비표"""
    ws.Name = "2.합격OK_불합격NG"
    _title(ws, f"공정 {proc['no']} 합격(OK) vs 불합격(NG) 9포인트", span=6)
    ws.Cells(2, 1).Value = "신규자가 1주차에 외울 핵심. 손에 들고 다니며 작업 전 확인."
    ws.Range(ws.Cells(2, 1), ws.Cells(2, 6)).Merge()
    ws.Cells(2, 1).Font.Size = 11
    ws.Cells(2, 1).Font.ColorIndex = C_GRAY
    ws.Cells(2, 1).HorizontalAlignment = -4108

    headers = ["NO", "확인할 것", "✅ 이래야 합격", "❌ 이러면 불량"]
    widths = [5, 40, 28, 28]
    r = 4
    for i, h in enumerate(headers):
        _h(ws.Cells(r, i + 1), h, size=14, bg=C_DARK_BLUE)
    for i, w in enumerate(widths):
        ws.Columns(i + 1).ColumnWidth = w
    ws.Rows(r).RowHeight = 32

    if eval_data:
        for i, (it, (q, _u, y)) in enumerate(zip(eval_data["items"], eval_data["criteria"]), start=1):
            rr = r + i
            _b(ws.Cells(rr, 1), i, size=14, hcenter=True, bold=True)
            _b(ws.Cells(rr, 2), it or "", size=13)
            _b(ws.Cells(rr, 3), q or "", size=13, bg=C_LIGHT_GREEN, bold=True)
            _b(ws.Cells(rr, 4), y or "", size=13, bg=C_LIGHT_RED, bold=True, color=C_RED)
            ws.Rows(rr).RowHeight = 56

    ws.PageSetup.Orientation = 2
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False


def build_sheet3_mistake_knowhow(ws, proc, eval_data):
    """흔한 실수 Top 5 + 사수 노하우 빈양식 (사수가 손글씨 또는 타이핑으로 채움)"""
    ws.Name = "3.실수_노하우"
    _title(ws, f"공정 {proc['no']} 흔한 실수 & 사수 노하우 (사수 작성용)", span=6)
    ws.Cells(2, 1).Value = "▶ 사수가 직접 빈칸을 채워주세요. 신규자가 가장 자주 보게 될 자료입니다."
    ws.Range(ws.Cells(2, 1), ws.Cells(2, 6)).Merge()
    ws.Cells(2, 1).Font.Size = 11
    ws.Cells(2, 1).Font.ColorIndex = C_RED
    ws.Cells(2, 1).Font.Bold = True
    ws.Cells(2, 1).HorizontalAlignment = -4108

    # 흔한 실수 Top 5 — 빈양식
    r = 4
    ws.Cells(r, 1).Value = "🚫 신규자가 자주 하는 실수 TOP 5"
    ws.Range(ws.Cells(r, 1), ws.Cells(r, 6)).Merge()
    c = ws.Cells(r, 1)
    c.Font.Size = 15
    c.Font.Bold = True
    c.Interior.ColorIndex = C_LIGHT_RED
    c.HorizontalAlignment = -4108
    ws.Rows(r).RowHeight = 30
    r += 1

    headers = ["NO", "실수 내용", "왜 그러는가", "어떻게 막을까", "관련 평가항목"]
    widths = [5, 28, 24, 28, 18]
    for i, h in enumerate(headers):
        _h(ws.Cells(r, i + 1), h, size=12, bg=C_DARK_BLUE)
    for i, w in enumerate(widths):
        ws.Columns(i + 1).ColumnWidth = w
    r += 1
    # 평가서 NG 사례(Y열)에서 자동 채워주는 힌트 + 빈칸
    hints = []
    if eval_data:
        for it, (_q, _u, y) in zip(eval_data["items"], eval_data["criteria"]):
            if y:
                hints.append((it, y))
    for i in range(1, 6):
        _b(ws.Cells(r, 1), i, size=13, hcenter=True, bold=True)
        if i <= len(hints):
            it, y = hints[i - 1]
            _b(ws.Cells(r, 2), f"(예시) {y}", size=11, color=C_GRAY)
            _b(ws.Cells(r, 5), it[:25] if it else "", size=10, color=C_GRAY)
        else:
            _b(ws.Cells(r, 2), "", size=12)
            _b(ws.Cells(r, 5), "", size=12)
        _b(ws.Cells(r, 3), "", size=12)
        _b(ws.Cells(r, 4), "", size=12)
        ws.Rows(r).RowHeight = 50
        r += 1

    # 사수 노하우 빈양식
    r += 1
    ws.Cells(r, 1).Value = "💡 사수만 아는 노하우 (이렇게 하면 편하다)"
    ws.Range(ws.Cells(r, 1), ws.Cells(r, 6)).Merge()
    c = ws.Cells(r, 1)
    c.Font.Size = 15
    c.Font.Bold = True
    c.Interior.ColorIndex = C_LIGHT_GREEN
    c.HorizontalAlignment = -4108
    ws.Rows(r).RowHeight = 30
    r += 1

    headers2 = ["NO", "상황 (언제)", "보통 방식", "노하우 방식 (이게 더 편함)"]
    widths2 = [5, 22, 24, 36]
    for i, h in enumerate(headers2):
        _h(ws.Cells(r, i + 1), h, size=12, bg=C_DARK_BLUE)
    for i, w in enumerate(widths2):
        ws.Columns(i + 1).ColumnWidth = w
    r += 1
    for i in range(1, 6):
        _b(ws.Cells(r, 1), i, size=13, hcenter=True, bold=True)
        _b(ws.Cells(r, 2), "", size=12)
        _b(ws.Cells(r, 3), "", size=12)
        _b(ws.Cells(r, 4), "", size=12)
        ws.Rows(r).RowHeight = 44
        r += 1

    ws.PageSetup.Orientation = 2
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False


def build_sheet4_photo_guide(ws, proc, eval_data):
    """NG 사진 첨부 가이드 — 9개 평가항목별 빈칸 (사수가 실제 NG 사진 1장씩 첨부)"""
    ws.Name = "4.NG사진_첨부"
    _title(ws, f"공정 {proc['no']} 실제 NG 사진 첨부 (사수가 채움)", span=4)
    ws.Cells(2, 1).Value = "▶ 사수가 실제 라인에서 발생한 NG 사례를 1장씩 사진으로 찍어 아래 빈칸에 붙여 주세요."
    ws.Range(ws.Cells(2, 1), ws.Cells(2, 4)).Merge()
    ws.Cells(2, 1).Font.Size = 11
    ws.Cells(2, 1).Font.ColorIndex = C_RED
    ws.Cells(2, 1).Font.Bold = True
    ws.Cells(2, 1).HorizontalAlignment = -4108

    headers = ["NO", "확인 항목", "NG 기준 (참고)", "📷 NG 사진 첨부 (빈칸)"]
    widths = [5, 30, 25, 36]
    r = 4
    for i, h in enumerate(headers):
        _h(ws.Cells(r, i + 1), h, size=13, bg=C_DARK_BLUE)
    for i, w in enumerate(widths):
        ws.Columns(i + 1).ColumnWidth = w
    ws.Rows(r).RowHeight = 30

    if eval_data:
        for i, (it, (_q, _u, y)) in enumerate(zip(eval_data["items"], eval_data["criteria"]), start=1):
            rr = r + i
            _b(ws.Cells(rr, 1), i, size=13, hcenter=True, bold=True)
            _b(ws.Cells(rr, 2), it or "", size=12)
            _b(ws.Cells(rr, 3), y or "", size=11, color=C_RED)
            _b(ws.Cells(rr, 4), "", size=11)  # 사진 첨부용 빈칸
            ws.Rows(rr).RowHeight = 100  # 사진 들어갈 공간

    ws.PageSetup.Orientation = 2
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False


def build_sheet5_checklist(ws, proc, eval_data):
    """사수 체크리스트 — 1일/3일/1주/2주 V표"""
    ws.Name = "5.사수체크리스트"
    _title(ws, f"공정 {proc['no']} 사수 체크리스트", span=6)
    ws.Cells(2, 1).Value = "신규자: ___________   사수: ___________   시작일: 26 / __ / __"
    ws.Range(ws.Cells(2, 1), ws.Cells(2, 6)).Merge()
    ws.Cells(2, 1).Font.Size = 13
    ws.Cells(2, 1).HorizontalAlignment = -4108
    ws.Rows(2).RowHeight = 30

    headers = ["NO", "확인 사항", "1일", "3일", "1주", "2주"]
    widths = [5, 50, 8, 8, 8, 8]
    r = 4
    for i, h in enumerate(headers):
        _h(ws.Cells(r, i + 1), h, size=13, bg=C_DARK_BLUE)
    for i, w in enumerate(widths):
        ws.Columns(i + 1).ColumnWidth = w
    ws.Rows(r).RowHeight = 30

    if eval_data:
        for i, it in enumerate(eval_data["items"], start=1):
            rr = r + i
            _b(ws.Cells(rr, 1), i, size=13, hcenter=True, bold=True)
            _b(ws.Cells(rr, 2), it or "", size=12)
            for c in range(3, 7):
                _b(ws.Cells(rr, c), "", size=14, hcenter=True)
            ws.Rows(rr).RowHeight = 38

    last_r = r + 11
    ws.Cells(last_r, 1).Value = "V = 통과 / △ = 보강 필요 / ✗ = 재교육. 2주차 모든 항목 V → 단독 작업 가능 (사수 사인 필요)"
    ws.Range(ws.Cells(last_r, 1), ws.Cells(last_r, 6)).Merge()
    ws.Cells(last_r, 1).Font.Size = 12
    ws.Cells(last_r, 1).Font.ColorIndex = C_RED
    ws.Cells(last_r, 1).Font.Bold = True

    ws.Cells(last_r + 2, 1).Value = "사수 사인: _______________________      날짜: 26 / __ / __      라인장 확인: _______________________"
    ws.Range(ws.Cells(last_r + 2, 1), ws.Cells(last_r + 2, 6)).Merge()
    ws.Cells(last_r + 2, 1).Font.Size = 12

    ws.PageSetup.Orientation = 2
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False


def build_sheet6_mentor(ws, proc, personnel):
    """사수 후보 명단"""
    ws.Name = "6.사수후보"
    _title(ws, f"공정 {proc['no']} OJT 사수 후보 (자동 추출)", span=4)
    ws.Cells(2, 1).Value = "AC합계 = 평가항목 9개 점수 총합 (45점 만점). 상위자를 1차 사수로 지명."
    ws.Range(ws.Cells(2, 1), ws.Cells(2, 4)).Merge()
    ws.Cells(2, 1).Font.Size = 11
    ws.Cells(2, 1).HorizontalAlignment = -4108

    headers = ["조", "이름", "역할", "AC합계 (45점 만점)"]
    widths = [10, 16, 14, 20]
    r = 4
    for i, h in enumerate(headers):
        _h(ws.Cells(r, i + 1), h, size=13, bg=C_DARK_BLUE)
    for i, w in enumerate(widths):
        ws.Columns(i + 1).ColumnWidth = w
    ws.Rows(r).RowHeight = 32

    rr = r + 1
    for shift in ["주간", "야간"]:
        cands = sorted([p for p in personnel if p["shift"] == shift],
                       key=lambda x: -x["ac_sum"])[:3]
        if not cands:
            _b(ws.Cells(rr, 1), shift, size=12, hcenter=True)
            _b(ws.Cells(rr, 2), "(투입 인원 없음)", size=12, color=C_GRAY)
            _b(ws.Cells(rr, 3), "-", size=12, hcenter=True)
            _b(ws.Cells(rr, 4), "-", size=12, hcenter=True)
            rr += 1
            continue
        for idx, c_ in enumerate(cands):
            _b(ws.Cells(rr, 1), shift, size=12, hcenter=True)
            _b(ws.Cells(rr, 2), c_["name"], size=13, hcenter=True,
               bold=(idx == 0), color=C_RED if idx == 0 else None)
            _b(ws.Cells(rr, 3), c_["role"] or "-", size=12, hcenter=True)
            _b(ws.Cells(rr, 4), c_["ac_sum"], size=13, hcenter=True)
            ws.Rows(rr).RowHeight = 28
            rr += 1
        rr += 1

    ws.Cells(rr, 1).Value = "※ 빨강 = 1차 사수 (조별 최고점). 1차 사수가 부재 시 2~3위 순으로 OJT 진행."
    ws.Range(ws.Cells(rr, 1), ws.Cells(rr, 4)).Merge()
    ws.Cells(rr, 1).Font.Size = 11
    ws.Cells(rr, 1).Font.ColorIndex = C_RED


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    print("[1/3] 데이터 로드...")
    procs = load_processes()
    ef = load_eval_forms()
    ppl = load_personnel()

    OUTDIR.mkdir(parents=True, exist_ok=True)

    print("[2/3] Excel COM 시작...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    print("[3/3] 6공정 빌드...")
    for proc in procs:
        out_path = (OUTDIR / f"SP3M3_공정{proc['no']}_신규입사자 교육자료.xlsx").resolve()
        if out_path.exists():
            out_path.unlink()

        wb = excel.Workbooks.Add()
        while wb.Sheets.Count > 1:
            wb.Sheets(wb.Sheets.Count).Delete()

        s1 = wb.Sheets(1)
        build_sheet1_intro(s1, proc, procs)
        s2 = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        build_sheet2_ok_ng(s2, proc, ef.get(proc["no"]))
        s3 = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        build_sheet3_mistake_knowhow(s3, proc, ef.get(proc["no"]))
        s4 = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        build_sheet4_photo_guide(s4, proc, ef.get(proc["no"]))
        s5 = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        build_sheet5_checklist(s5, proc, ef.get(proc["no"]))
        s6 = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        build_sheet6_mentor(s6, proc, ppl.get(proc["no"], []))

        wb.SaveAs(str(out_path), FileFormat=xlOpenXMLWorkbook)
        wb.Close(False)
        print(f"  ✓ {out_path.name}")

    excel.ScreenUpdating = True
    excel.Quit()
    print(f"\n완료 — {OUTDIR}")


if __name__ == "__main__":
    main()
