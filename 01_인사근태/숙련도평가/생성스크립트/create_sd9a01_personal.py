"""
SD9A01 개인별 숙련도 평가서 — 23명 (주간 12 + 야간 11)

베이스: create_sp3m3_personal_v2.py
- 입력: SD9M01 주간/야간 23명 기존 평가서
- 마스터 템플릿: 공정 11개 시트 (create_sd9a01_v1.py PROCESSES 사용)
- 작업자별 처리:
  - 수행공정만 시트 유지, 나머지 삭제
  - 주공정 시트 맨 앞 + Z3 = "주공정"/"전환공정"
  - SD9M01 공정번호(10/20/.../100) → ERP 코드(001/002/.../011) 변환
  - AC열 점수 자동 분배 (총점 보존, 0점 없음)

실행:
    python create_sd9a01_personal.py
"""
import os
import sys
import shutil
import tempfile
import random
from datetime import datetime, date
from openpyxl import load_workbook

# create_sd9a01_v1 모듈에서 PROCESSES + 매핑 가져오기
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from create_sd9a01_v1 import build_processes, ERP_PROCESSES  # noqa


BASE        = r'C:\Users\User\Desktop\업무리스트\01_인사근태\숙련도평가'
SAMPLE      = os.path.join(BASE, '공정 평가서표준_SP3M3_샘플.xlsx')
OUT_BASE    = os.path.join(BASE, 'SD9A01_개인별 평가서')

# 입력: SD9M01 주간 / 야간 (라인 코드는 SD9M01이지만 ERP 라인은 SD9A01)
SOURCE_DIRS = [
    (os.path.join(BASE, 'SD9M01 주간 숙련도 평가서'), '주간'),
    (os.path.join(BASE, 'SD9M01 야간 숙련도 평가서'), '야간'),
]

# 좌표 (SP3M3 동일)
PROC_COLS  = [14, 46, 78, 110, 142, 174, 206, 238, 270, 302, 334]
TOTAL_COLS = [10, 42, 74, 106, 138, 170, 202, 234, 266, 298, 330]
LEVEL_COLS = [20, 52, 84, 116, 148, 180, 212, 244, 276, 308, 340]

# 새 양식 셀 좌표 (공정별 양식 첫 번째 공정 컬럼만 사용)
ROW_INFO = 5
COL_P1_NO, COL_P1_NAME, COL_P1_ITEM = 14, 19, 5
COL_P1_Q, COL_P1_U, COL_P1_Y = 17, 21, 25
COL_P2_NO, COL_P2_NAME, COL_P2_ITEM = 45, 50, 36
COL_P2_Q, COL_P2_U, COL_P2_Y = 48, 52, 56
COL_P3_NO, COL_P3_NAME, COL_P3_ITEM = 76, 81, 67
COL_P3_Q, COL_P3_U, COL_P3_Y = 79, 83, 87
ROWS_STD, ROWS_CTRL = list(range(13, 18)), list(range(18, 22))

COL_BIZ, COL_LINE = 4, 9
CELL_NAME    = (2, 22)
CELL_EMP_ID  = (3, 22)
CELL_POS     = (3, 26)
CELL_DATE    = (5, 28)
CELL_OPINION = (33, 2)

# SD9M01 공정번호(작업표준서 시트번호 기준) → ERP 코드 매핑
SD9M01_TO_ERP = {
    '10':  '001',
    '20':  '002',
    '21':  '003',
    '30':  '004',
    '40':  '005',
    '50':  '006',
    '60':  '007',
    '70':  '008',
    '80':  '009',
    '90':  '010',
    '100': '011',
}


def safe_write(ws, row, col, value):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value


def normalize_date(d):
    if isinstance(d, (datetime, date)):
        return d.strftime('%Y-%m-%d')
    if d is None:
        return ''
    s = str(d)
    return s[:10] if len(s) >= 10 else s


def fill_process_sheet(ws, proc):
    """공정별 양식 시트 채우기 (한 시트 = 한 공정, std 5 + ctrl 4)"""
    safe_write(ws, ROW_INFO, COL_BIZ, '대원테크')
    safe_write(ws, ROW_INFO, COL_LINE, 'SD9A01')
    safe_write(ws, ROW_INFO, COL_P1_NO, proc['no'])
    safe_write(ws, ROW_INFO, COL_P1_NAME, proc['name'])

    for i, r in enumerate(ROWS_STD):
        if i < len(proc['std']):
            item, q, u, y = proc['std'][i]
            safe_write(ws, r, COL_P1_ITEM, item)
            safe_write(ws, r, COL_P1_Q, q)
            safe_write(ws, r, COL_P1_U, u)
            safe_write(ws, r, COL_P1_Y, y)
    for i, r in enumerate(ROWS_CTRL):
        if i < len(proc['ctrl']):
            item, q, u, y = proc['ctrl'][i]
            safe_write(ws, r, COL_P1_ITEM, item)
            safe_write(ws, r, COL_P1_Q, q)
            safe_write(ws, r, COL_P1_U, u)
            safe_write(ws, r, COL_P1_Y, y)

    # 2·3번 공정 컬럼 비우기
    for col_no, col_name, col_item, col_q, col_u, col_y in [
        (COL_P2_NO, COL_P2_NAME, COL_P2_ITEM, COL_P2_Q, COL_P2_U, COL_P2_Y),
        (COL_P3_NO, COL_P3_NAME, COL_P3_ITEM, COL_P3_Q, COL_P3_U, COL_P3_Y),
    ]:
        safe_write(ws, ROW_INFO, col_no, '')
        safe_write(ws, ROW_INFO, col_name, '')
        for r in ROWS_STD + ROWS_CTRL:
            safe_write(ws, r, col_item, '')
            safe_write(ws, r, col_q, '')
            safe_write(ws, r, col_u, '')
            safe_write(ws, r, col_y, '')


def build_master_template(processes):
    """공정 11개 시트 마스터 임시 파일 생성."""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, prefix='sd9a01_master_')
    tmp.close()
    shutil.copy(SAMPLE, tmp.name)

    wb = load_workbook(tmp.name)
    template_ws = wb[wb.sheetnames[0]]
    orig_name = template_ws.title

    for proc in processes:
        ws = wb.copy_worksheet(template_ws)
        ws.title = f"공정{proc['no']}"
        fill_process_sheet(ws, proc)

    # 원본 양식 시트 삭제
    del wb[orig_name]
    wb.save(tmp.name)
    wb.close()
    return tmp.name


def extract_worker(path, shift):
    """기존 SD9M01 평가서 → 작업자 정보 + 수행공정 (ERP 코드로 변환)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    name    = os.path.basename(path).split(' 숙련도')[0]
    emp_id  = ws.cell(*CELL_EMP_ID).value
    date_v  = normalize_date(ws.cell(*CELL_DATE).value)

    procs = []
    for pc, tc, lc in zip(PROC_COLS, TOTAL_COLS, LEVEL_COLS):
        if pc > ws.max_column:
            break
        pv = ws.cell(ROW_INFO, pc).value
        if pv is None:
            continue
        sd_no = str(pv).strip()
        erp_no = SD9M01_TO_ERP.get(sd_no)
        if erp_no is None:
            print(f'  [WARN] {name}: 미매핑 공정번호 [{sd_no}] (skip)')
            continue
        procs.append({
            'no':       erp_no,        # ERP 코드 사용
            'sd9m01_no': sd_no,
            'total':    ws.cell(32, tc).value,
            'level':    ws.cell(32, lc).value,
        })

    main_proc_no = procs[0]['no'] if procs else None
    wb.close()
    return {
        'name':   name,
        'emp_id': emp_id,
        'date':   date_v,
        'shift':  shift,
        'procs':  procs,
        'main_proc': main_proc_no,
    }


def distribute_score(target_total, n_items=20, max_per=5, seed=None):
    base = target_total // n_items
    rem  = target_total - base * n_items
    scores = [base] * n_items
    rng = random.Random(seed)
    bonus_idx = rng.sample(range(n_items), rem) if rem else []
    for i in bonus_idx:
        scores[i] += 1
    scores = [min(s, max_per) for s in scores]
    return scores


def fill_worker_info(ws, worker, proc_info, is_main):
    safe_write(ws, *CELL_NAME,   worker['name'])
    safe_write(ws, *CELL_EMP_ID, worker['emp_id'])
    safe_write(ws, *CELL_POS,    '주공정' if is_main else '전환공정')
    safe_write(ws, *CELL_DATE,   worker['date'])

    total = proc_info.get('total')
    if isinstance(total, (int, float)):
        seed = hash(f"{worker['name']}_{proc_info['no']}") & 0xFFFFFFFF
        scores = distribute_score(int(total), seed=seed)
        for i, r in enumerate(range(9, 29)):
            safe_write(ws, r, 29, scores[i])

    tag = '주공정' if is_main else '전환공정'
    text = (
        f'평 가 자 의 견\n\n'
        f'[{tag}] [직전 평가: {proc_info.get("total")}점 / {proc_info.get("level")}]\n'
        f'{worker["shift"]}반 / 점수는 총점 기준 자동분배 (항목별 재검토 필요)'
    )
    safe_write(ws, *CELL_OPINION, text)


def create_worker_file(worker, master_path, out_dir):
    out_path = os.path.join(out_dir, f'{worker["name"]} 숙련도평가서.xlsx')
    shutil.copy(master_path, out_path)

    wb = load_workbook(out_path)
    worker_procs = {p['no'] for p in worker['procs']}

    # 미수행 공정 시트 삭제
    for sname in list(wb.sheetnames):
        proc_no = sname.replace('공정', '')
        if proc_no not in worker_procs:
            del wb[sname]

    main_no = worker.get('main_proc')
    for p in worker['procs']:
        sname = f"공정{p['no']}"
        if sname in wb.sheetnames:
            is_main = (p['no'] == main_no)
            fill_worker_info(wb[sname], worker, p, is_main)

    # 시트 재배치 (수행공정 순)
    desired = [f"공정{p['no']}" for p in worker['procs']]
    existing = [n for n in desired if n in wb.sheetnames]
    if existing:
        wb._sheets = [wb[n] for n in existing]
        wb.active = 0

    wb.save(out_path)
    wb.close()
    return out_path


def main():
    out_day   = os.path.join(OUT_BASE, 'SD9A01 주간')
    out_night = os.path.join(OUT_BASE, 'SD9A01 야간')
    os.makedirs(out_day,   exist_ok=True)
    os.makedirs(out_night, exist_ok=True)

    print('[1/3] PROCESSES 11개 빌드 (통합본+관리계획서)...')
    processes = build_processes()
    print(f'  {len(processes)}개 공정 빌드 완료')

    print('\n[2/3] 마스터 템플릿 생성 (공정 11개 시트)...')
    master = build_master_template(processes)
    print(f'  임시 마스터: {master}')

    try:
        print('\n[3/3] 작업자 정보 추출 + 개인별 파일 생성...')
        workers = []
        for src, shift in SOURCE_DIRS:
            for f in sorted(os.listdir(src)):
                if f.endswith('.xlsx'):
                    w = extract_worker(os.path.join(src, f), shift)
                    workers.append(w)

        for w in workers:
            out_dir = out_day if w['shift'] == '주간' else out_night
            path = create_worker_file(w, master, out_dir)
            procs_str = '/'.join(p['no'] for p in w['procs'])
            print(f"  [{w['shift']}] {w['name']:8s} ({len(w['procs']):2d}공정: {procs_str})")
    finally:
        if os.path.exists(master):
            os.remove(master)

    print(f'\n[완료] {OUT_BASE}')
    total_files = sum(len(os.listdir(d)) for d in [out_day, out_night] if os.path.isdir(d))
    print(f'  총 파일 수: {total_files}')


if __name__ == '__main__':
    main()
