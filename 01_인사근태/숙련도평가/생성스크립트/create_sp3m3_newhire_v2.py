# -*- coding: utf-8 -*-
"""
SP3M3 신규 입사자 교육자료 v2 — 프린트해서 실제 교육 가능한 형태
- 1번 시트: 작업표준서 해당 공정 시트 통째 복사 (이미지·도면·서식 보존) = 교육 본문
- 2번 시트: 합격/불합격 큰 글자 요약 (NG vs OK)
- 3번 시트: 사수 체크 1장 (V표 양식)
- 4번 시트: 사수 후보 / OJT 매칭

Excel COM 사용 (Sheet.Copy로 도형·이미지 보존).
"""
from __future__ import annotations
import csv
import re
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
import win32com.client as win32

BASE = Path(__file__).resolve().parent.parent
STD = BASE / "SP3M3_표준문서"
PROC_FORM = BASE / "SP3M3_공정별 평가서"
PERSONAL = BASE / "SP3M3_개인별 평가서"
OUTDIR = BASE / "SP3M3_신규입사자 교육자료"

# Excel 상수
xlOpenXMLWorkbook = 51


# ─────────────────────────────────────────────
# 데이터 로드 (v1과 동일)
# ─────────────────────────────────────────────
def load_processes():
    wb = openpyxl.load_workbook(STD / "SP3M3 라인별공정목록.xlsx", data_only=True)
    ws = wb.active
    out = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        out.append({"no": int(no), "name": (ws.cell(r, 2).value or "").strip(),
                    "level": (ws.cell(r, 3).value or "Lv.3").strip()})
    wb.close()
    return out


def load_eval_forms():
    out = {}
    for fp in PROC_FORM.glob("SP3M3_공정*.xlsx"):
        m = re.search(r"공정(\d+)", fp.name)
        if not m:
            continue
        no = int(m.group(1))
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        items = [ws.cell(r, 5).value for r in range(13, 22)]
        criteria = [(ws.cell(r, 17).value, ws.cell(r, 21).value, ws.cell(r, 25).value)
                    for r in range(13, 22)]
        out[no] = {"items": items, "criteria": criteria}
        wb.close()
    return out


def load_personnel():
    out = defaultdict(list)
    for shift_dir, shift in [("SP3M3 주간", "주간"), ("SP3M3 야간", "야간")]:
        d = PERSONAL / shift_dir
        if not d.exists():
            continue
        for fp in d.glob("*.xlsx"):
            name = fp.stem.replace(" 숙련도평가서", "").strip()
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                n5 = ws["N5"].value
                z3 = ws["Z3"].value
                if not isinstance(n5, (int, float)):
                    try:
                        n5 = int(str(n5))
                    except Exception:
                        continue
                ac_sum = sum(v for r in range(13, 22)
                             if isinstance(v := ws.cell(r, 29).value, (int, float)))
                out[int(n5)].append({"name": name, "shift": shift, "ac_sum": ac_sum,
                                     "role": (z3 or "").strip() if isinstance(z3, str) else z3})
            wb.close()
    return out


# ─────────────────────────────────────────────
# 작업표준서 시트 매핑
# ─────────────────────────────────────────────
def find_ws_sheet_name(ws_book_sheetnames, proc_no):
    exact, ranges, near = [], [], []
    for sn in ws_book_sheetnames:
        nums = [int(n) for n in re.findall(r"\d+", sn)]
        if not nums:
            continue
        if proc_no in nums:
            exact.append(sn)
            continue
        if len(nums) >= 2 and (max(nums) - min(nums)) <= 30:
            if min(nums) <= proc_no <= max(nums):
                ranges.append((max(nums) - min(nums), sn))
                continue
        diff = min(abs(n - proc_no) for n in nums)
        if diff <= 5:
            near.append((diff, sn))
    if exact:
        return exact[0]
    if ranges:
        ranges.sort()
        return ranges[0][1]
    if near:
        near.sort()
        return near[0][1]
    return None


# ─────────────────────────────────────────────
# 큰 글자 요약·체크리스트 시트 생성 (COM)
# ─────────────────────────────────────────────
def build_summary_sheet(ws_obj, proc, eval_data):
    """합격/불합격 큰 글자 요약."""
    ws_obj.Name = "1.합격OK_불합격NG"
    ws_obj.Cells(1, 1).Value = f"공정 {proc['no']} — {proc['name']}"
    ws_obj.Range("A1:F1").Merge()
    c1 = ws_obj.Cells(1, 1)
    c1.Font.Size = 22
    c1.Font.Bold = True
    c1.Font.ColorIndex = 5  # blue
    c1.HorizontalAlignment = -4108  # center
    ws_obj.Rows(1).RowHeight = 40

    ws_obj.Cells(2, 1).Value = f"요구 레벨: {proc['level']}  /  신규자가 1개월 안에 도달해야 하는 수준"
    ws_obj.Range("A2:F2").Merge()
    ws_obj.Cells(2, 1).Font.Size = 12
    ws_obj.Cells(2, 1).Font.ColorIndex = 16

    # 헤더
    headers = ["NO", "확인 사항 (이것을 보세요)", "OK (이래야 합격)", "NG (이러면 불량)"]
    widths = [6, 50, 32, 32]
    r = 4
    for i, h in enumerate(headers):
        cell = ws_obj.Cells(r, i + 1)
        cell.Value = h
        cell.Font.Size = 14
        cell.Font.Bold = True
        cell.Font.ColorIndex = 2  # white
        cell.Interior.ColorIndex = 32  # dark blue
        cell.HorizontalAlignment = -4108
        cell.Borders.LineStyle = 1
    for i, w in enumerate(widths):
        ws_obj.Columns(i + 1).ColumnWidth = w
    ws_obj.Rows(r).RowHeight = 36

    if not eval_data:
        ws_obj.Cells(r + 1, 1).Value = "(공정평가서 매칭 실패)"
        return

    items = eval_data["items"]
    criteria = eval_data["criteria"]
    for i, (it, (q, _u, y)) in enumerate(zip(items, criteria), start=1):
        rr = r + i
        ws_obj.Cells(rr, 1).Value = i
        ws_obj.Cells(rr, 2).Value = it or ""
        ws_obj.Cells(rr, 3).Value = q or ""
        ws_obj.Cells(rr, 4).Value = y or ""
        for c in range(1, 5):
            cell = ws_obj.Cells(rr, c)
            cell.Font.Size = 13
            cell.WrapText = True
            cell.VerticalAlignment = -4108  # center
            cell.Borders.LineStyle = 1
        ws_obj.Cells(rr, 3).Interior.ColorIndex = 35  # light green
        ws_obj.Cells(rr, 4).Interior.ColorIndex = 38  # light red
        ws_obj.Rows(rr).RowHeight = 56

    # 인쇄 설정
    ws_obj.PageSetup.Orientation = 2  # landscape
    ws_obj.PageSetup.FitToPagesWide = 1
    ws_obj.PageSetup.FitToPagesTall = False


def build_checklist_sheet(ws_obj, proc, eval_data):
    """사수가 V표 칠 수 있는 체크리스트 1장."""
    ws_obj.Name = "2.사수체크리스트"
    ws_obj.Cells(1, 1).Value = f"공정 {proc['no']} 사수 체크리스트 (신규자: ____________  사수: ____________  날짜: ____/____)"
    ws_obj.Range("A1:F1").Merge()
    c1 = ws_obj.Cells(1, 1)
    c1.Font.Size = 16
    c1.Font.Bold = True
    c1.HorizontalAlignment = -4108
    ws_obj.Rows(1).RowHeight = 36

    headers = ["NO", "확인 사항", "1일", "3일", "1주", "2주"]
    widths = [6, 55, 8, 8, 8, 8]
    r = 3
    for i, h in enumerate(headers):
        cell = ws_obj.Cells(r, i + 1)
        cell.Value = h
        cell.Font.Size = 13
        cell.Font.Bold = True
        cell.Interior.ColorIndex = 15
        cell.HorizontalAlignment = -4108
        cell.Borders.LineStyle = 1
    for i, w in enumerate(widths):
        ws_obj.Columns(i + 1).ColumnWidth = w
    ws_obj.Rows(r).RowHeight = 32

    if eval_data:
        for i, it in enumerate(eval_data["items"], start=1):
            rr = r + i
            ws_obj.Cells(rr, 1).Value = i
            ws_obj.Cells(rr, 2).Value = it or ""
            for c in range(1, 7):
                cell = ws_obj.Cells(rr, c)
                cell.Font.Size = 12
                cell.WrapText = True
                cell.VerticalAlignment = -4108
                cell.Borders.LineStyle = 1
            ws_obj.Cells(rr, 2).HorizontalAlignment = -4131  # left
            ws_obj.Rows(rr).RowHeight = 40

    # 안내 문구
    last_r = r + 10
    ws_obj.Cells(last_r, 1).Value = "※ V = 통과 / △ = 보강 필요 / X = 재교육. 2주차 모든 항목 V 시 단독 작업 가능."
    ws_obj.Range(ws_obj.Cells(last_r, 1), ws_obj.Cells(last_r, 6)).Merge()
    ws_obj.Cells(last_r, 1).Font.Size = 11
    ws_obj.Cells(last_r, 1).Font.ColorIndex = 3

    ws_obj.PageSetup.Orientation = 2
    ws_obj.PageSetup.FitToPagesWide = 1
    ws_obj.PageSetup.FitToPagesTall = False


def build_mentor_sheet(ws_obj, proc, personnel):
    """사수 후보 / OJT 매칭."""
    ws_obj.Name = "3.사수후보"
    ws_obj.Cells(1, 1).Value = f"공정 {proc['no']} OJT 사수 후보"
    ws_obj.Range("A1:D1").Merge()
    c1 = ws_obj.Cells(1, 1)
    c1.Font.Size = 16
    c1.Font.Bold = True
    c1.HorizontalAlignment = -4108
    ws_obj.Rows(1).RowHeight = 32

    headers = ["조", "이름", "역할", "AC합계"]
    widths = [10, 18, 14, 12]
    r = 3
    for i, h in enumerate(headers):
        cell = ws_obj.Cells(r, i + 1)
        cell.Value = h
        cell.Font.Bold = True
        cell.Font.Size = 12
        cell.Interior.ColorIndex = 15
        cell.HorizontalAlignment = -4108
        cell.Borders.LineStyle = 1
    for i, w in enumerate(widths):
        ws_obj.Columns(i + 1).ColumnWidth = w

    rr = r + 1
    for shift in ["주간", "야간"]:
        cands = sorted([p for p in personnel if p["shift"] == shift],
                       key=lambda x: -x["ac_sum"])[:3]
        if not cands:
            ws_obj.Cells(rr, 1).Value = shift
            ws_obj.Cells(rr, 2).Value = "(투입 인원 없음)"
            for c in range(1, 5):
                ws_obj.Cells(rr, c).Borders.LineStyle = 1
                ws_obj.Cells(rr, c).Font.Size = 11
            rr += 1
            continue
        for c_ in cands:
            ws_obj.Cells(rr, 1).Value = shift
            ws_obj.Cells(rr, 2).Value = c_["name"]
            ws_obj.Cells(rr, 3).Value = c_["role"] or "-"
            ws_obj.Cells(rr, 4).Value = c_["ac_sum"]
            for c in range(1, 5):
                cell = ws_obj.Cells(rr, c)
                cell.Font.Size = 12
                cell.Borders.LineStyle = 1
                cell.HorizontalAlignment = -4108
            rr += 1
        rr += 1  # 빈 줄

    ws_obj.Cells(rr, 1).Value = "※ AC합계 = 평가항목 9개 점수 총합 (45점 만점). 상위 1명을 1차 사수로 지정."
    ws_obj.Range(ws_obj.Cells(rr, 1), ws_obj.Cells(rr, 4)).Merge()
    ws_obj.Cells(rr, 1).Font.Size = 11
    ws_obj.Cells(rr, 1).Font.ColorIndex = 3


def build_fallback_workinst_sheet(ws_obj, proc):
    """작업표준서 시트 매칭 실패 시 보강 안내 카드."""
    ws_obj.Name = "0.작업표준_보강필요"
    ws_obj.Cells(1, 1).Value = f"⚠ 공정 {proc['no']} 작업표준서 매칭 누락"
    ws_obj.Range("A1:E1").Merge()
    c1 = ws_obj.Cells(1, 1)
    c1.Font.Size = 18
    c1.Font.Bold = True
    c1.Font.ColorIndex = 3
    c1.HorizontalAlignment = -4108
    ws_obj.Rows(1).RowHeight = 40

    msg = (f"\n공정 {proc['no']} ({proc['name']}) 는 작업표준서 200730_R1 파일에 별도 시트가 없습니다."
           f"\n\n[조치]\n"
           f"1. 라인반장이 동급 공정의 작업표준서 시트를 인쇄해서 첨부.\n"
           f"2. 또는 별도 SOP 사진을 1장 찍어 이 파일에 삽입.\n"
           f"3. 합격/불합격 기준 + 체크리스트 + 사수후보 시트는 동일하게 사용 가능.\n"
           f"\n현재 시트 구성:\n"
           f"  · 1.합격OK_불합격NG — 평가항목 9개 기준\n"
           f"  · 2.사수체크리스트 — 1일/3일/1주/2주 V표\n"
           f"  · 3.사수후보 — OJT 멘토 추천\n")
    ws_obj.Cells(3, 1).Value = msg
    ws_obj.Range("A3:E20").Merge()
    cell = ws_obj.Cells(3, 1)
    cell.Font.Size = 13
    cell.WrapText = True
    cell.VerticalAlignment = -4160  # top


# ─────────────────────────────────────────────
# 메인 빌드
# ─────────────────────────────────────────────
def main():
    print("[1/4] 데이터 로드...")
    procs = load_processes()
    ef = load_eval_forms()
    ppl = load_personnel()
    print(f"  공정 {len(procs)}개, 평가서 {len(ef)}개, 개인 row {sum(len(v) for v in ppl.values())}")

    OUTDIR.mkdir(parents=True, exist_ok=True)

    print("[2/4] Excel COM 실행...")
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    print("[3/4] 작업표준서 열기 (이미지·도형 보존용)...")
    src_path = (STD / "SP3M3_작업표준서_200730_R1.xlsx").resolve()
    src_wb = excel.Workbooks.Open(str(src_path))
    src_sheetnames = [src_wb.Sheets(i + 1).Name for i in range(src_wb.Sheets.Count)]

    print("[4/4] 공정 6개 빌드...")
    for proc in procs:
        out_path = (OUTDIR / f"SP3M3_공정{proc['no']}_신규입사자 교육자료.xlsx").resolve()
        if out_path.exists():
            out_path.unlink()

        # 신규 wb 생성
        dst_wb = excel.Workbooks.Add()
        # 기본 시트 1개 남기고 정리
        while dst_wb.Sheets.Count > 1:
            dst_wb.Sheets(dst_wb.Sheets.Count).Delete()

        sn = find_ws_sheet_name(src_sheetnames, proc["no"])
        if sn:
            # 작업표준서 해당 시트를 dst_wb로 복사 (이미지·도형 그대로)
            src_wb.Sheets(sn).Copy(Before=dst_wb.Sheets(1))
            copied = dst_wb.Sheets(1)
            copied.Name = f"0.작업표준_공정{proc['no']}"
            # 빈 기본 시트 제거
            dst_wb.Sheets(dst_wb.Sheets.Count).Delete()
            print(f"  공정{proc['no']}: 작업표준서 '{sn}' 복사 ✓")
        else:
            # fallback 카드
            build_fallback_workinst_sheet(dst_wb.Sheets(1), proc)
            print(f"  공정{proc['no']}: 작업표준서 매칭 실패 → 보강 카드 ⚠")

        # 2~4번 시트 추가
        s2 = dst_wb.Sheets.Add(After=dst_wb.Sheets(dst_wb.Sheets.Count))
        build_summary_sheet(s2, proc, ef.get(proc["no"]))

        s3 = dst_wb.Sheets.Add(After=dst_wb.Sheets(dst_wb.Sheets.Count))
        build_checklist_sheet(s3, proc, ef.get(proc["no"]))

        s4 = dst_wb.Sheets.Add(After=dst_wb.Sheets(dst_wb.Sheets.Count))
        build_mentor_sheet(s4, proc, ppl.get(proc["no"], []))

        dst_wb.SaveAs(str(out_path), FileFormat=xlOpenXMLWorkbook)
        dst_wb.Close(False)
        print(f"    → 저장: {out_path.name}")

    src_wb.Close(False)
    excel.ScreenUpdating = True
    excel.Quit()
    print(f"\n완료 — {OUTDIR}")


if __name__ == "__main__":
    main()
