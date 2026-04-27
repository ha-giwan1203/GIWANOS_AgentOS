"""
센스커버 조립공정 부적합 가능성 — 단순 정리본 (사용자 제시 4건)
- 1시트 구성: 문제점 + 영상 썸네일 + 하이퍼링크
"""
import sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

BASE = Path(r"C:\Users\User\Desktop\업무리스트\06_생산관리\품질")
VIDEO = BASE / "센스커버 조립공정.mp4"
THUMB = BASE / "_frames_analysis" / "frame_05.jpg"
OUT = BASE / "센스커버_조립공정_부적합가능성분석_20260427_v2.xlsx"

THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
BODY = Font(name="맑은 고딕", size=11)
BODY_B = Font(name="맑은 고딕", size=11, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "부적합 가능성"
ws.sheet_view.showGridLines = False

# 제목
ws.merge_cells("B2:E2")
c = ws["B2"]
c.value = "센스커버 조립공정 — 오조립 발생 가능성"
c.font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
c.fill = HDR_FILL
c.alignment = CENTER
ws.row_dimensions[2].height = 32

# 메타
ws.cell(row=3, column=2, value="작성일").font = BODY_B
ws.cell(row=3, column=2).border = BORDER
ws.cell(row=3, column=2).alignment = CENTER
ws.merge_cells("C3:E3")
ws.cell(row=3, column=3, value="2026-04-27   /   대상 품번: 89870CU100").font = BODY
ws.cell(row=3, column=3).alignment = LEFT
ws.cell(row=3, column=3).border = BORDER

# 표 헤더
hdr = ["NO", "구분", "문제점 (오조립 가능성)"]
for i, h in enumerate(hdr, start=2):
    cell = ws.cell(row=5, column=i, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.merge_cells("D5:E5")
ws.row_dimensions[5].height = 26

# 4개 항목
items = [
    ("1", "스펙·색상 동일",
     "센스CP 사양/스펙 사이즈, 색상이 동일하면 오조립 가능성"),
    ("2", "비전검사 고장",
     "비전검사 고장 시 수작업 진행으로 인한 오조립 가능성"),
    ("3", "색상 판단",
     "작업자가 품번으로 사양 확인하여야 하나 색상으로 판단 시 오조립 가능성 (유사 색상)"),
    ("4", "재작업",
     "수작업 재작업 진행 시 품번 미확인으로 오사양 조립 가능성"),
]
r = 6
for no, gubun, content in items:
    ws.cell(row=r, column=2, value=no).font = BODY_B
    ws.cell(row=r, column=3, value=gubun).font = BODY_B
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    ws.cell(row=r, column=4, value=content).font = BODY
    for col in range(2, 6):
        cc = ws.cell(row=r, column=col)
        cc.border = BORDER
        cc.alignment = CENTER if col in (2, 3) else LEFT
    ws.row_dimensions[r].height = 36
    r += 1

# 동영상 영역 헤더
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
hc = ws.cell(row=r, column=2, value="원본 동영상")
hc.font = HDR_FONT
hc.fill = HDR_FILL
hc.alignment = CENTER
hc.border = BORDER
ws.row_dimensions[r].height = 24

# 하이퍼링크 셀 (썸네일 위)
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
link_cell = ws.cell(row=r, column=2,
    value="▶ 클릭하여 동영상 재생  —  센스커버 조립공정.mp4")
link_cell.font = Font(name="맑은 고딕", size=12, bold=True,
                     color="0563C1", underline="single")
link_cell.fill = PatternFill("solid", fgColor="FFF2CC")
link_cell.alignment = CENTER
link_cell.border = BORDER
link_cell.hyperlink = "센스커버 조립공정.mp4"
ws.row_dimensions[r].height = 28

# 썸네일 (다음 행에 앵커)
r += 1
thumb_anchor = f"B{r}"
img = XLImage(str(THUMB))
img.width = 220
img.height = 390
img.anchor = thumb_anchor
ws.add_image(img)
# 썸네일 영역 행 높이 확보
for rr in range(r, r + 18):
    ws.row_dimensions[rr].height = 22

# 컬럼 너비
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 32

wb.save(OUT)
print(f"saved: {OUT}")
